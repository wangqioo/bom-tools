# -*- coding: utf-8 -*-
"""BOM Tools Web 公共模块 — 跨工具共享的工具函数与常量"""

import os, sys, json, io, zipfile, threading, uuid, shutil, time, csv, math, re
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
from typing import Dict, List, Optional, Tuple

# ── 确保可以导入父目录的 openpyxl ──
_parent = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _parent not in sys.path:
    sys.path.insert(0, _parent)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string

try:
    import requests
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "-q"])
    import requests

from flask import Flask, render_template, request, send_file, jsonify, session, redirect, url_for

# ── 目录 ──
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")


# ════════════════════════════════════════════════════════════════
#  工具函数
# ════════════════════════════════════════════════════════════════

def _unique_path(path):
    """若 path 已存在或被占用，则自动叠加 (1)(2)… 直到找到可写路径。"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f"{base}({n}){ext}"
        if not os.path.exists(candidate):
            try:
                with open(candidate, "ab"):
                    pass
                os.remove(candidate)
                return candidate
            except PermissionError:
                pass
        n += 1


def _cell_str(val):
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, list):
        return " ".join(
            item.get("text", "") or item.get("link", "")
            for item in val if isinstance(item, dict)
        ).strip()
    return str(val).strip()


def _cleanup_old_files(directory, minutes=30):
    now = time.time()
    for f in os.listdir(directory):
        fp = os.path.join(directory, f)
        try:
            if os.path.isfile(fp) and now - os.path.getmtime(fp) > minutes * 60:
                os.remove(fp)
        except Exception:
            pass


def _natural_sort_key(s):
    parts = re.split(r'(\d+)', str(s or '').upper())
    return [int(p) if p.isdigit() else p for p in parts]


def _col_int(s):
    """列标识符转列索引：'A'→1, '1'→1, 'AB'→28"""
    try:
        return int(s) if s.isdigit() else column_index_from_string(s.upper())
    except Exception:
        return None


# ── 飞书预设表格（被首页和多个工具的 GET 复用）──
FEISHU_PRESET_TABLES = [
    {"name": "MLCC", "token": "shthq7d9W17DSo7cwuFhtIg7KPf", "category": "优选库"},
    {"name": "电阻", "token": "shthqdJvubPLY8mrO8qkOMXmGiw", "category": "优选库"},
    {"name": "电解电容", "token": "shthqO56eTG9DyJX60nDaaMGp0e", "category": "优选库"},
    {"name": "网络变压器/电感器", "token": "shthquBGGQB8twAgmJWSBwFY5he", "category": "优选库"},
    {"name": "磁珠", "token": "shthq1EJlpgHfBqBGNediRZdt8c", "category": "优选库"},
    {"name": "晶体晶振", "token": "shthqpy6hg6hVD78VNPElmcSF6d", "category": "优选库"},
    {"name": "保险丝", "token": "shthqpsZCFkwGjn62CjRYCpKrHg", "category": "优选库"},
    {"name": "纽扣电池", "token": "shthqngCNYdufotGIcL6Vn6gWWh", "category": "优选库"},
    {"name": "滤波器/共模扼流圈", "token": "shthqaZbFrblnh3V0A3ahhOj4Og", "category": "优选库"},
    {"name": "Power IC优选库", "token": "shthqz7lKPJt9UGF4FOIU1uyTUh", "category": "优选库"},
    {"name": "功能IC优选库", "token": "shthq4b1PTCh1HqyalUal6aTYte", "category": "优选库"},
    {"name": "DBG分立器件优选库", "token": "shthqEZrwmemvVULwrhmyAmlfzd", "category": "优选库"},
    {"name": "连接器", "token": "shthqE9sVI2DkIBYkSkLcUdNxvn", "category": "优选库"},
    {"name": "Cable", "token": "shthqpYELkJAH7b0uPn1HRcEyLg", "category": "优选库"},
    {"name": "客户物料型号与HQ料号对应关系", "token": "shthq1R9G7zSp5hvTISGNDOWjme", "category": "对应关系库"},
]
