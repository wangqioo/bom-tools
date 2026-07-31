"""Shared legacy Excel reader for standalone BOM Tools scripts."""

from __future__ import annotations

import hashlib
import os
import subprocess
import tempfile
from pathlib import Path

import openpyxl
from openpyxl import Workbook


def _converted_path(source: Path) -> Path:
    cache_dir = Path(tempfile.gettempdir()) / "bom-tools-xls"
    cache_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(str(source.resolve()).encode("utf-8")).hexdigest()[:16]
    return cache_dir / f"{source.stem}_{digest}.xlsx"


def _convert_with_xlrd(source: Path, destination: Path) -> None:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Missing xlrd. Install it with: pip install xlrd") from exc

    workbook = xlrd.open_workbook(str(source))
    target = Workbook()
    try:
        for index in range(workbook.nsheets):
            source_sheet = workbook.sheet_by_index(index)
            sheet = target.active if index == 0 else target.create_sheet()
            sheet.title = source_sheet.name or f"Sheet{index + 1}"
            for row in range(source_sheet.nrows):
                for column in range(source_sheet.ncols):
                    cell = source_sheet.cell(row, column)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate_as_datetime(value, workbook.datemode)
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        value = bool(value)
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        value = xlrd.error_text_from_code.get(int(value), "#ERROR!")
                    sheet.cell(row=row + 1, column=column + 1, value=value)
        target.save(destination)
    finally:
        workbook.release_resources()
        target.close()


def _ps_quote(value: Path) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _convert_with_excel(source: Path, destination: Path) -> None:
    if os.name != "nt":
        raise RuntimeError("This protected XLS file requires Windows Excel to convert")
    script = (
        "$ErrorActionPreference='Stop';"
        f"$src={_ps_quote(source)};$dst={_ps_quote(destination)};"
        "$excel=New-Object -ComObject Excel.Application;"
        "$excel.Visible=$false;$excel.DisplayAlerts=$false;$excel.AutomationSecurity=3;"
        "try{$wb=$excel.Workbooks.Open($src);$wb.SaveAs($dst,51);$wb.Close($false)}"
        "finally{$excel.Quit();[Runtime.InteropServices.Marshal]::ReleaseComObject($excel)|Out-Null}"
    )
    subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        check=True,
        capture_output=True,
        timeout=90,
    )


def open_workbook_compat(path: str | os.PathLike[str], **kwargs):
    """Open XLSX directly or convert a legacy XLS file before opening it."""
    source = Path(path)
    if source.suffix.lower() != ".xls":
        return openpyxl.load_workbook(source, **kwargs)

    destination = _converted_path(source)
    if not destination.exists() or destination.stat().st_mtime < source.stat().st_mtime:
        try:
            _convert_with_xlrd(source, destination)
        except Exception:
            try:
                _convert_with_excel(source, destination)
            except Exception as exc:
                raise ValueError(
                    "Unable to read this .xls file. Install xlrd or open it in authorized Excel and save as .xlsx."
                ) from exc
    return openpyxl.load_workbook(destination, **kwargs)
