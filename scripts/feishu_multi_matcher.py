# -*- coding: utf-8 -*-
"""
飞书多表格匹配工具 v3.0（华勤内部网关版）

功能：
  - 预置 14 个常用飞书表格，可自行增删改名称/Token
  - 每个表格支持选择多个 Sheet（合并读取）
  - 每个表格独立配置匹配键（最多 3 对，AND 逻辑）和提取列
  - 一次运行可匹配所有已启用表格，按顺序命中第一个匹配的表格
  - 输出增加「来源表格」列，配置持久化保存到 feishu_match_config.json

认证：华勤内部 API 网关（origin + 工号，无需 App Secret）
依赖：pip install openpyxl requests（首次运行自动安装）
"""

import sys, subprocess
for _pkg in ["openpyxl", "requests"]:
    try:
        __import__(_pkg)
    except ImportError:
        print(f"安装 {_pkg} ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", _pkg])

import os, json, threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from excel_compat import open_workbook_compat
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

# ─────────────────────── 预置表格 ───────────────────────
# category 字段区分两类库：优选库 / 对应关系库
PRESET_TABLES = [
    {"name": "MLCC",            "token": "shthq7d9W17DSo7cwuFhtIg7KPf",  "category": "优选库"},
    {"name": "电阻",             "token": "shthqdJvubPLY8mrO8qkOMXmGiw",  "category": "优选库"},
    {"name": "电解电容",          "token": "shthqO56eTG9DyJX60nDaaMGp0e", "category": "优选库"},
    {"name": "网络变压器/电感器",  "token": "shthquBGGQB8twAgmJWSBwFY5he", "category": "优选库"},
    {"name": "磁珠",             "token": "shthq1EJlpgHfBqBGNediRZdt8c",  "category": "优选库"},
    {"name": "晶体晶振",          "token": "shthqpy6hg6hVD78VNPElmcSF6d", "category": "优选库"},
    {"name": "保险丝",            "token": "shthqpsZCFkwGjn62CjRYCpKrHg", "category": "优选库"},
    {"name": "纽扣电池",          "token": "shthqngCNYdufotGIcL6Vn6gWWh", "category": "优选库"},
    {"name": "滤波器/共模扼流圈",  "token": "shthqaZbFrblnh3V0A3ahhOj4Og","category": "优选库"},
    {"name": "Power IC优选库",    "token": "shthqz7lKPJt9UGF4FOIU1uyTUh", "category": "优选库"},
    {"name": "功能IC优选库",      "token": "shthq4b1PTCh1HqyalUal6aTYte", "category": "优选库"},
    {"name": "DBG分立器件优选库",  "token": "shthqEZrwmemvVULwrhmyAmlfzd","category": "优选库"},
    {"name": "连接器",            "token": "shthqE9sVI2DkIBYkSkLcUdNxvn", "category": "优选库"},
    {"name": "Cable",            "token": "shthqpYELkJAH7b0uPn1HRcEyLg",  "category": "优选库"},
    # ── 对应关系库 ──
    {"name": "客户物料型号与HQ料号对应关系",
                                 "token": "shthq1R9G7zSp5hvTISGNDOWjme",  "category": "对应关系库"},
]

CATEGORIES = ["优选库", "对应关系库"]   # 分类顺序（新增时可选）

CONFIG_FILE = "feishu_match_config.json"

# ─────────────────────── 工具函数 ───────────────────────
def _unique_path(path):
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        c = f"{base}({n}){ext}"
        if not os.path.exists(c):
            try:
                with open(c, "ab"): pass
                os.remove(c)
                return c
            except PermissionError:
                pass
        n += 1

def _cell_str(val):
    if val is None: return ""
    if isinstance(val, (int, float)): return str(val)
    if isinstance(val, str): return val.strip()
    if isinstance(val, list):
        return " ".join(item.get("text") or item.get("link") or ""
                        for item in val if isinstance(item, dict)).strip()
    return str(val).strip()

# ─────────────────────── API 函数 ───────────────────────
def hq_get_sheets(base_url, origin, user_id, token):
    url = f"{base_url}/fs/sheet/v1/spreadsheetsMetainfo"
    r = requests.get(url, params={"origin": origin, "userId": user_id,
                                  "spreadsheetToken": token}, timeout=15)
    r.raise_for_status()
    d = r.json()
    if d.get("code") not in (0, 200):
        raise RuntimeError(f"获取 Sheet 列表失败：{d.get('msg')}（code={d.get('code')}）")
    return [s for s in d["data"]["sheets"] if s.get("title")]

def hq_read_sheet(base_url, origin, user_id, token,
                  sheet_id, row_count=200000, col_count=100,
                  batch_size=3000, progress_cb=None):
    end_col = get_column_letter(max(col_count, 26))
    all_rows, start = [], 1
    while start <= max(row_count, 1):
        end = min(start + batch_size - 1, row_count)
        params = {"origin": origin, "userId": user_id,
                  "spreadsheetToken": token,
                  "range": f"{sheet_id}!A{start}:{end_col}{end}"}
        r = requests.get(f"{base_url}/fs/sheet/v1/getSheetsValue",
                         params=params, timeout=60)
        r.raise_for_status()
        d = r.json()
        if d.get("code") not in (0, 200):
            raise RuntimeError(f"读取失败：{d.get('msg')}")
        batch = d["data"]["valueRange"].get("values") or []
        if all_rows and batch:
            batch = batch[1:]
        if not batch:
            break
        all_rows.extend(batch)
        if progress_cb:
            progress_cb(len(all_rows))
        expected = end - start + 1
        skip = 1 if start > 1 else 0
        if len(batch) < expected - skip:
            break
        start = end + 1
    while all_rows and not any(_cell_str(v) for v in all_rows[-1]):
        all_rows.pop()
    return all_rows

def hq_read_table(base_url, origin, user_id, token,
                  sheets_meta, active_ids, progress_cb=None):
    """读取并合并多个 Sheet 的数据。
    progress_cb(sheet_title, n_rows_this_sheet) 每批次回调。
    """
    combined, headers_set = [], False
    for s in sheets_meta:
        if s["sheetId"] not in active_ids:
            continue
        title = s.get("title", s["sheetId"])
        # 包装 progress_cb，附加当前 sheet 标题
        def _pcb(n, _t=title):
            if progress_cb:
                progress_cb(_t, n)
        rows = hq_read_sheet(base_url, origin, user_id, token,
                             s["sheetId"],
                             row_count=s.get("rowCount", 200000),
                             col_count=s.get("columnCount", 100),
                             batch_size=3000, progress_cb=_pcb)
        if not rows:
            continue
        if not headers_set:
            combined = rows
            headers_set = True
        else:
            combined.extend(rows[1:])
    return combined

# ─────────────────────── 匹配核心 ───────────────────────
def do_match_multi(local_ws, local_header_row,
                   prepared_tables, all_fetch_cols,
                   out_file, log_cb):
    """
    多表格匹配：每行本地数据依次尝试各表格，首个命中为准。
    prepared_tables: [{name, local_key_cols, lookup, fetch_col_names}, ...]
    all_fetch_cols:  所有表格提取列名的有序并集
    """
    max_local_col = local_ws.max_column
    local_header = [local_ws.cell(row=local_header_row, column=ci).value
                    for ci in range(1, max_local_col + 1)]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "匹配结果"
    thin = Side(style="thin")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")
    hq_fill  = PatternFill("solid", fgColor="FFFF00")
    src_fill = PatternFill("solid", fgColor="BDD7EE")

    out_hdrs = list(local_header) + all_fetch_cols + ["来源表格"]
    for ci, h in enumerate(out_hdrs, 1):
        c = ws_out.cell(row=1, column=ci, value=h or "")
        c.font = Font(bold=True)
        c.fill = (PatternFill("solid", fgColor="FFC000")
                  if ci > max_local_col else hdr_fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    ws_out.row_dimensions[1].height = 22
    for ci in range(1, max_local_col + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 18
    for ci in range(max_local_col + 1, len(out_hdrs) + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 24

    dr = 2
    matched = unmatched = 0

    for ri in range(local_header_row + 1, local_ws.max_row + 1):
        row_vals = [local_ws.cell(row=ri, column=ci).value
                    for ci in range(1, max_local_col + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue

        found = False
        for pt in prepared_tables:
            key = tuple(_cell_str(row_vals[lc - 1]) for lc in pt["local_key_cols"])
            matches = pt["lookup"].get(key, [])
            if not matches:
                continue
            first = True
            for mdict in matches:
                for ci, val in enumerate(row_vals, 1):
                    c = ws_out.cell(row=dr, column=ci, value=val if first else None)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = bdr
                for j, col_name in enumerate(all_fetch_cols):
                    c = ws_out.cell(row=dr, column=max_local_col + j + 1,
                                    value=mdict.get(col_name, ""))
                    c.fill = hq_fill
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = bdr
                src_col = max_local_col + len(all_fetch_cols) + 1
                c = ws_out.cell(row=dr, column=src_col,
                                value=pt["name"] if first else "")
                c.fill = src_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = bdr
                first = False
                dr += 1
            found = True
            matched += 1
            break

        if not found:
            for ci, val in enumerate(row_vals, 1):
                c = ws_out.cell(row=dr, column=ci, value=val)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = bdr
            for j in range(len(all_fetch_cols)):
                ws_out.cell(row=dr, column=max_local_col + j + 1).border = bdr
            c = ws_out.cell(row=dr, column=max_local_col + len(all_fetch_cols) + 1,
                            value="未匹配")
            c.border = bdr
            unmatched += 1
            dr += 1

    wb_out.save(out_file)
    total = dr - 2
    log_cb(f"写入 {total} 行：{matched} 个匹配成功，{unmatched} 个未匹配")
    return total, matched, unmatched


# ═══════════════════════ 主程序 ════════════════════════
class FeishuMatchApp(tk.Tk):
    # ── 初始化 ────────────────────────────────────────

    def __init__(self):
        super().__init__()
        self.title("飞书多表格匹配工具 v3.0（华勤网关版）")
        self.geometry("980x720")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # 本地文件
        self.wb = None
        self.ws = None
        self.local_headers = []

        # 网关配置（全局）
        self.v_base_url = tk.StringVar(value="https://mcenter.huaqin.com")
        self.v_origin   = tk.StringVar(value="cli_a96ac38049f8d0e5")
        self.v_user_id  = tk.StringVar(value="100448405")
        self.v_in_path  = tk.StringVar()
        self.v_sheet    = tk.StringVar()
        self.v_hdr_row  = tk.IntVar(value=1)
        self.v_out_path = tk.StringVar(value="飞书匹配结果.xlsx")

        # 表格库
        self.tables: list = []
        self._tab2_cur_idx  = None   # 当前在 tab2 右侧显示的表格
        self._tab2_sv: dict = {}     # tab2 sheet checkboxes {sheetId: BooleanVar}

        # 列映射
        self._tab3_cur_idx = None
        self._tab3_key_vars = [(tk.StringVar(), tk.StringVar()) for _ in range(3)]
        self._tab3_col_vars: dict = {}
        self._tab3_chk_canvas = None
        self._tab3_chk_inner  = None
        self._tab3_chk_win    = None

        self._spinning = False

        self._load_config()
        self._build_ui()

    def _make_table_entry(self, preset=None, category="优选库"):
        return {
            "name":             preset["name"]     if preset else "新表格",
            "token":            preset["token"]    if preset else "",
            "category":         preset.get("category", category) if preset else category,
            "enabled":          False,
            "active_sheet_ids": [],
            "local_key_names":  ["", "", ""],
            "feishu_key_names": ["", "", ""],
            "fetch_col_names":  [],
            # 运行时（不持久化）
            "_sheets":    [],
            "_rows":      [],
            "_headers":   [],
            "_connected": False,
            "_loaded":    False,
        }

    # ── 配置持久化 ────────────────────────────────────

    def _cfg_path(self):
        try:
            base = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            base = os.getcwd()
        return os.path.join(base, CONFIG_FILE)

    def _load_config(self):
        p = self._cfg_path()
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                self.v_base_url.set(cfg.get("base_url", self.v_base_url.get()))
                self.v_origin.set(cfg.get("origin",   self.v_origin.get()))
                self.v_user_id.set(cfg.get("user_id", self.v_user_id.get()))
                self.tables = []
                for st in cfg.get("tables", []):
                    t = self._make_table_entry()
                    t.update({k: v for k, v in st.items() if not k.startswith("_")})
                    self.tables.append(t)
                return
            except Exception:
                pass
        self.tables = [self._make_table_entry(p) for p in PRESET_TABLES]

    def _save_config(self):
        cfg = {
            "base_url": self.v_base_url.get(),
            "origin":   self.v_origin.get(),
            "user_id":  self.v_user_id.get(),
            "tables": [{k: v for k, v in t.items() if not k.startswith("_")}
                       for t in self.tables],
        }
        try:
            with open(self._cfg_path(), "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self._log(f"[配置保存失败] {e}")

    def _on_close(self):
        self._save_tab3_to_table()
        self._save_config()
        self.destroy()

    # ── 界面构建 ──────────────────────────────────────

    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=8, pady=6)
        self.nb = nb
        self.tab1 = ttk.Frame(nb); nb.add(self.tab1, text="  本地文件  ")
        self.tab2 = ttk.Frame(nb); nb.add(self.tab2, text="  表格库  ")
        self.tab3 = ttk.Frame(nb); nb.add(self.tab3, text="  列映射  ")
        self.tab4 = ttk.Frame(nb); nb.add(self.tab4, text="  运行  ")
        self._build_tab1()
        self._build_tab2()
        self._build_tab3()
        self._build_tab4()

    # ═══ Tab1：本地文件 ════════════════════════════════

    def _build_tab1(self):
        p = self.tab1

        f1 = ttk.LabelFrame(p, text="本地 Excel 文件", padding=8)
        f1.pack(fill="x", padx=10, pady=6)
        tk.Label(f1, text="路径：").grid(row=0, column=0, sticky="w")
        ttk.Entry(f1, textvariable=self.v_in_path, width=56).grid(row=0, column=1, padx=6)
        ttk.Button(f1, text="浏览...", command=self._browse_local).grid(row=0, column=2)

        f2 = ttk.LabelFrame(p, text="Sheet / 表头行", padding=8)
        f2.pack(fill="x", padx=10, pady=4)
        tk.Label(f2, text="Sheet：").grid(row=0, column=0, sticky="w")
        self.cb_sheet = ttk.Combobox(f2, textvariable=self.v_sheet, width=24, state="readonly")
        self.cb_sheet.grid(row=0, column=1, padx=4, sticky="w")
        self.cb_sheet.bind("<<ComboboxSelected>>", lambda e: self._load_local_sheet())
        tk.Label(f2, text="  表头行：").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(f2, from_=1, to=20, textvariable=self.v_hdr_row, width=5).grid(row=0, column=3)
        ttk.Button(f2, text="刷新", command=self._load_local_sheet).grid(row=0, column=4, padx=8)

        f3 = ttk.LabelFrame(p, text="表头预览（前5行）", padding=6)
        f3.pack(fill="x", padx=10, pady=4)
        self.preview_tree = ttk.Treeview(f3, height=10, show="headings")
        sx = ttk.Scrollbar(f3, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(xscrollcommand=sx.set)
        self.preview_tree.pack(fill="x")
        sx.pack(fill="x")

        ttk.Button(p, text="下一步：配置表格库 →",
                   command=lambda: self.nb.select(1)).pack(anchor="e", padx=12, pady=8)

    # ═══ Tab2：表格库 ══════════════════════════════════

    def _build_tab2(self):
        p = self.tab2

        # 网关设置（折叠式 LabelFrame）
        gw = ttk.LabelFrame(p, text="网关设置", padding=6)
        gw.pack(fill="x", padx=10, pady=(6, 2))
        gw_fields = [("网关地址：", self.v_base_url, 36),
                     ("Origin：",  self.v_origin,   28),
                     ("工号：",    self.v_user_id,  14)]
        for i, (lbl, var, w) in enumerate(gw_fields):
            tk.Label(gw, text=lbl).grid(row=0, column=i*2, sticky="w", padx=(8 if i else 0, 0))
            ttk.Entry(gw, textvariable=var, width=w).grid(row=0, column=i*2+1, padx=4)
        ttk.Button(gw, text="保存网关配置",
                   command=self._save_config).grid(row=0, column=7, padx=6)

        # 全局操作栏：一键更新全部表格
        act_bar = tk.Frame(p)
        act_bar.pack(fill="x", padx=10, pady=(2, 4))
        self.btn_update_all = ttk.Button(
            act_bar, text="⟳ 一键更新全部表格数据",
            command=self._do_batch_update_all)
        self.btn_update_all.pack(side="left")
        self._lbl_update_all = tk.Label(act_bar, text="", fg="#555",
                                        font=("Arial", 9))
        self._lbl_update_all.pack(side="left", padx=8)

        # 主体：左右分栏
        paned = tk.PanedWindow(p, orient="horizontal", sashwidth=5, sashrelief="raised")
        paned.pack(fill="both", expand=True, padx=10, pady=6)

        # 左侧：表格列表
        left = tk.Frame(paned)
        paned.add(left, minsize=200)

        # 新增/删除按钮在列表上方，等宽填满
        btn_bar = tk.Frame(left)
        btn_bar.pack(fill="x", pady=(0, 2))
        ttk.Button(btn_bar, text="＋ 新增", command=self._add_table).pack(
            side="left", fill="x", expand=True)
        ttk.Button(btn_bar, text="－ 删除", command=self._del_table).pack(
            side="left", fill="x", expand=True)

        # show="tree headings" 启用层级显示（分类行 + 表格行）
        tv_frame = tk.Frame(left)
        tv_frame.pack(fill="both", expand=True)
        self.tv_tables = ttk.Treeview(tv_frame, columns=("loaded", "match"),
                                      show="tree headings", height=18)
        self.tv_tables.heading("#0",     text="表格名称")
        self.tv_tables.heading("loaded", text="已加载")
        self.tv_tables.heading("match",  text="参与匹配")
        self.tv_tables.column("#0",     width=128)
        self.tv_tables.column("loaded", width=48, anchor="center", stretch=False)
        self.tv_tables.column("match",  width=68, anchor="center", stretch=False)
        sb = ttk.Scrollbar(tv_frame, command=self.tv_tables.yview)
        self.tv_tables.configure(yscrollcommand=sb.set)
        self.tv_tables.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.tv_tables.bind("<<TreeviewSelect>>", self._on_table_select)

        # 右侧：详情面板
        right = tk.Frame(paned)
        paned.add(right, minsize=360)
        self._tab2_right = right
        self._tab2_detail_frame = tk.Frame(right)
        self._tab2_detail_frame.pack(fill="both", expand=True, padx=6, pady=4)
        tk.Label(self._tab2_detail_frame, text="← 点击左侧表格查看详情",
                 fg="#888", font=("Arial", 10)).pack(pady=30)

        self._refresh_table_list()

    def _refresh_table_list(self):
        sel = self.tv_tables.selection()
        self.tv_tables.delete(*self.tv_tables.get_children())
        cat_colors = {"优选库": "#1a5fa8", "对应关系库": "#7a3a9a"}

        # 按 category 分组，保持 CATEGORIES 顺序，未知分类排末尾
        from collections import defaultdict
        grouped = defaultdict(list)
        for i, t in enumerate(self.tables):
            grouped[t.get("category", "优选库")].append((i, t))
        all_cats = list(CATEGORIES) + [c for c in grouped if c not in CATEGORIES]

        for cat in all_cats:
            if cat not in grouped:
                continue
            cat_iid = f"__cat__{cat}"
            self.tv_tables.insert("", "end", iid=cat_iid, text=f"▸ {cat}",
                                  values=("", ""), open=True,
                                  tags=(f"cat_{cat}",))
            self.tv_tables.tag_configure(f"cat_{cat}",
                                         foreground=cat_colors.get(cat, "#333"),
                                         font=("Arial", 9, "bold"))
            for i, t in grouped[cat]:
                # 已加载列：有数据→绿✓；已连接→蓝◎；否则灰—
                if t["_loaded"]:
                    loaded_icon = "✓";  loaded_tag = "ld_yes"
                elif t["_connected"]:
                    loaded_icon = "◎";  loaded_tag = "ld_conn"
                else:
                    loaded_icon = "—";  loaded_tag = "ld_no"

                # 参与匹配列：启用且已加载→绿●；启用未加载→橙！；未启用→灰○
                if t["enabled"] and t["_loaded"]:
                    match_icon = "●";   match_tag = "mt_yes"
                elif t["enabled"] and not t["_loaded"]:
                    match_icon = "！";  match_tag = "mt_warn"
                else:
                    match_icon = "○";   match_tag = "mt_no"

                row_tag = f"row_{i}"
                self.tv_tables.insert(cat_iid, "end", iid=str(i),
                                      text=t["name"],
                                      values=(loaded_icon, match_icon),
                                      tags=(row_tag,))
                # 行整体颜色：取两列中优先级高的颜色
                if t["enabled"] and t["_loaded"]:
                    row_fg = "#2a8a2a"   # 绿：就绪
                elif t["enabled"]:
                    row_fg = "#e07000"   # 橙：启用未加载
                elif t["_loaded"]:
                    row_fg = "#2d6cdf"   # 蓝：已加载未启用
                else:
                    row_fg = "#888"      # 灰：空闲
                self.tv_tables.tag_configure(row_tag, foreground=row_fg)

        # 恢复选中
        for s in sel:
            if self.tv_tables.exists(s):
                self.tv_tables.selection_set(s)

    def _on_table_select(self, _event=None):
        sel = self.tv_tables.selection()
        if not sel:
            return
        iid = sel[0]
        # 忽略分类行点击
        if iid.startswith("__cat__"):
            return
        idx = int(iid)
        self._tab2_cur_idx = idx
        self._build_table_detail(idx)

    def _build_table_detail(self, idx):
        t = self.tables[idx]
        frame = self._tab2_detail_frame
        for w in frame.winfo_children():
            w.destroy()

        # 分类
        r1b = tk.Frame(frame); r1b.pack(fill="x", pady=2)
        tk.Label(r1b, text="分类：", width=7, anchor="w").pack(side="left")
        v_cat = tk.StringVar(value=t.get("category", CATEGORIES[0]))
        cb_cat = ttk.Combobox(r1b, textvariable=v_cat, values=CATEGORIES,
                              state="readonly", width=16)
        cb_cat.pack(side="left", padx=4)

        # 名称 + 启用
        r1 = tk.Frame(frame); r1.pack(fill="x", pady=3)
        tk.Label(r1, text="名称：", width=7, anchor="w").pack(side="left")
        v_name = tk.StringVar(value=t["name"])
        ttk.Entry(r1, textvariable=v_name, width=18).pack(side="left", padx=4)
        v_en = tk.BooleanVar(value=t["enabled"])
        tk.Checkbutton(r1, text="启用（参与匹配）", variable=v_en,
                       command=lambda: self._toggle_enable(idx, v_en.get())
                       ).pack(side="left", padx=10)

        # Token
        r2 = tk.Frame(frame); r2.pack(fill="x", pady=3)
        tk.Label(r2, text="Token：", width=7, anchor="w").pack(side="left")
        v_token = tk.StringVar(value=t["token"])
        ttk.Entry(r2, textvariable=v_token, width=42).pack(side="left", padx=4)

        ttk.Button(frame, text="保存名称 / Token",
                   command=lambda: self._save_table_name_token(
                       idx, v_name.get(), v_token.get(), v_cat.get())).pack(anchor="w", pady=2)

        ttk.Separator(frame).pack(fill="x", pady=6)

        # 连接
        r3 = tk.Frame(frame); r3.pack(fill="x")
        btn_conn = ttk.Button(r3, text="连接并获取 Sheet 列表")
        btn_conn.pack(side="left")
        lbl_conn = tk.Label(r3, text="", fg="#555"); lbl_conn.pack(side="left", padx=8)
        btn_conn.configure(command=lambda: self._do_connect_table(idx, btn_conn, lbl_conn))

        # Sheet 选择区
        sf = ttk.LabelFrame(frame, text="选择 Sheet（可多选）", padding=6)
        sf.pack(fill="x", pady=6)
        self._tab2_sheet_frame = sf
        self._tab2_sv = {}

        if t["_connected"] and t["_sheets"]:
            lbl_conn.configure(text="已连接", fg="#2a8a2a")
            self._build_sheet_checkboxes(idx)
        else:
            tk.Label(sf, text="请先点击「连接并获取 Sheet 列表」", fg="#888").pack()

        ttk.Separator(frame).pack(fill="x", pady=4)

        # 读取数据
        r4 = tk.Frame(frame); r4.pack(fill="x")
        btn_load = ttk.Button(r4, text="读取选中 Sheet 数据")
        btn_load.pack(side="left")
        lbl_load = tk.Label(r4, text="", fg="#555"); lbl_load.pack(side="left", padx=8)
        btn_load.configure(command=lambda: self._do_load_table(idx, btn_load, lbl_load))
        if t["_loaded"]:
            lbl_load.configure(text=f"✓ 已加载 {max(len(t['_rows'])-1,0)} 行",
                               fg="#2a8a2a")

        ttk.Separator(frame).pack(fill="x", pady=6)

        # 快捷更新（连接 + 用已保存的 Sheet 选择直接读取）
        r5 = tk.Frame(frame); r5.pack(fill="x")
        btn_upd = ttk.Button(r5, text="⟳ 更新本地数据")
        btn_upd.pack(side="left")
        lbl_upd = tk.Label(r5, text="（自动连接并重新读取已选 Sheet）",
                           fg="#888", font=("Arial", 8))
        lbl_upd.pack(side="left", padx=6)
        btn_upd.configure(command=lambda: self._do_update_table(
            idx, btn_upd, lbl_upd, btn_conn, lbl_conn, btn_load, lbl_load))

        self._tab2_lbl_conn = lbl_conn
        self._tab2_lbl_load = lbl_load
        self._tab2_btn_conn = btn_conn
        self._tab2_btn_load = btn_load

    def _build_sheet_checkboxes(self, idx):
        t = self.tables[idx]
        sf = self._tab2_sheet_frame
        for w in sf.winfo_children():
            w.destroy()
        self._tab2_sv = {}

        # 可滚动区域，Sheet 多时不会溢出
        canvas = tk.Canvas(sf, height=130, highlightthickness=0)
        vsb = ttk.Scrollbar(sf, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        inner = tk.Frame(canvas)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(win_id, width=e.width))
        # 鼠标滚轮（仅当鼠标在该 canvas 上时响应）
        canvas.bind("<Enter>",
                    lambda e, c=canvas: c.bind_all(
                        "<MouseWheel>",
                        lambda ev, c=c: c.yview_scroll(-1*(ev.delta//120), "units")))
        canvas.bind("<Leave>",
                    lambda e, c=canvas: c.unbind_all("<MouseWheel>"))

        for s in t["_sheets"]:
            var = tk.BooleanVar(value=s["sheetId"] in t["active_sheet_ids"])
            self._tab2_sv[s["sheetId"]] = var
            tk.Checkbutton(inner, text=s["title"], variable=var,
                           anchor="w").pack(anchor="w", fill="x")

    def _toggle_enable(self, idx, val):
        self.tables[idx]["enabled"] = val
        self._refresh_table_list()
        self._save_config()

    def _save_table_name_token(self, idx, name, token, category=None):
        self.tables[idx]["name"]  = name.strip() or self.tables[idx]["name"]
        self.tables[idx]["token"] = token.strip()
        if category:
            self.tables[idx]["category"] = category
        self._refresh_table_list()
        self._save_config()
        messagebox.showinfo("已保存", f"表格「{self.tables[idx]['name']}」名称和 Token 已更新")

    def _add_table(self):
        # 弹窗让用户选择分类
        dlg = tk.Toplevel(self)
        dlg.title("新增表格 – 选择分类")
        dlg.resizable(False, False)
        dlg.grab_set()
        tk.Label(dlg, text="请选择表格所属分类：", padx=16, pady=12).pack()
        v_cat = tk.StringVar(value=CATEGORIES[0])
        for cat in CATEGORIES:
            tk.Radiobutton(dlg, text=cat, variable=v_cat, value=cat,
                           anchor="w").pack(anchor="w", padx=24)
        confirmed = [False]
        def _ok():
            confirmed[0] = True
            dlg.destroy()
        def _cancel():
            dlg.destroy()
        bf = tk.Frame(dlg); bf.pack(pady=10)
        ttk.Button(bf, text="确定", command=_ok).pack(side="left", padx=6)
        ttk.Button(bf, text="取消", command=_cancel).pack(side="left", padx=6)
        self.wait_window(dlg)
        if not confirmed[0]:
            return
        self.tables.append(self._make_table_entry(category=v_cat.get()))
        self._refresh_table_list()
        new_iid = str(len(self.tables) - 1)
        self.tv_tables.selection_set(new_iid)
        self.tv_tables.see(new_iid)
        self._build_table_detail(len(self.tables) - 1)
        self._save_config()

    def _del_table(self):
        sel = self.tv_tables.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选中要删除的表格")
            return
        iid = sel[0]
        if iid.startswith("__cat__"):
            messagebox.showwarning("提示", "请选中具体的表格行，不能删除分类节点")
            return
        idx = int(iid)
        name = self.tables[idx]["name"]
        if not messagebox.askyesno("确认", f"确认删除表格「{name}」？"):
            return
        self.tables.pop(idx)
        self._tab2_cur_idx = None
        for w in self._tab2_detail_frame.winfo_children():
            w.destroy()
        tk.Label(self._tab2_detail_frame, text="← 点击左侧表格查看详情",
                 fg="#888", font=("Arial", 10)).pack(pady=30)
        self._refresh_table_list()
        self._save_config()

    def _do_connect_table(self, idx, btn, lbl):
        t = self.tables[idx]
        token = t["token"].strip()
        if not token:
            messagebox.showwarning("提示", "请先填写并保存 Token")
            return
        btn.configure(state="disabled")
        lbl.configure(text="连接中...", fg="#2d6cdf")
        threading.Thread(target=self._connect_bg, args=(idx, btn, lbl),
                         daemon=True).start()

    def _connect_bg(self, idx, btn, lbl):
        t = self.tables[idx]
        try:
            sheets = hq_get_sheets(self.v_base_url.get().rstrip("/"),
                                   self.v_origin.get(), self.v_user_id.get(),
                                   t["token"])
            t["_sheets"]    = sheets
            t["_connected"] = True
            self.after(0, lambda: self._on_connect_ok(idx, btn, lbl))
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: self._on_connect_fail(btn, lbl, msg))

    def _on_connect_ok(self, idx, btn, lbl):
        btn.configure(state="normal")
        lbl.configure(text="已连接", fg="#2a8a2a")
        self._build_sheet_checkboxes(idx)
        self._refresh_table_list()
        self._log(f"[{self.tables[idx]['name']}] 连接成功，"
                  f"{len(self.tables[idx]['_sheets'])} 个 Sheet")

    def _on_connect_fail(self, btn, lbl, msg):
        btn.configure(state="normal")
        lbl.configure(text="连接失败", fg="red")
        messagebox.showerror("连接失败", msg)

    def _do_load_table(self, idx, btn, lbl):
        t = self.tables[idx]
        if not t["_connected"]:
            messagebox.showwarning("提示", "请先连接")
            return
        active_ids = [sid for sid, var in self._tab2_sv.items() if var.get()]
        if not active_ids:
            messagebox.showwarning("提示", "请至少选择一个 Sheet")
            return
        t["active_sheet_ids"] = active_ids
        self._save_config()
        btn.configure(state="disabled")
        lbl.configure(text="读取中...", fg="#2d6cdf")

        def progress(sheet_title, n):
            self.after(0, lambda t=sheet_title, x=n:
                       lbl.configure(text=f"「{t}」已读 {x} 行...", fg="#2d6cdf"))

        threading.Thread(target=self._load_bg, args=(idx, btn, lbl, progress),
                         daemon=True).start()

    def _load_bg(self, idx, btn, lbl, progress_cb):
        t = self.tables[idx]
        try:
            rows = hq_read_table(self.v_base_url.get().rstrip("/"),
                                 self.v_origin.get(), self.v_user_id.get(),
                                 t["token"], t["_sheets"],
                                 t["active_sheet_ids"], progress_cb)
            t["_rows"]    = rows
            t["_headers"] = [_cell_str(v) for v in rows[0]] if rows else []
            t["_loaded"]  = True
            cnt = max(len(rows) - 1, 0)
            self.after(0, lambda: self._on_load_ok(idx, btn, lbl, cnt))
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: self._on_load_fail(btn, lbl, msg))

    def _on_load_ok(self, idx, btn, lbl, cnt):
        btn.configure(state="normal")
        lbl.configure(text=f"✓ 已加载 {cnt} 行", fg="#2a8a2a")
        self._refresh_table_list()
        # 如果 Tab3 当前正显示此表格，刷新列
        if self._tab3_cur_idx == idx:
            self._refresh_tab3_cols()
        self._log(f"[{self.tables[idx]['name']}] 数据加载完成：{cnt} 行")

    def _on_load_fail(self, btn, lbl, msg):
        btn.configure(state="normal")
        lbl.configure(text="读取失败", fg="red")
        messagebox.showerror("读取失败", msg)

    # ── 快捷更新：连接 + 用已保存 Sheet 选择直接读取 ──────────

    def _do_update_table(self, idx, btn_upd, lbl_upd,
                         btn_conn, lbl_conn, btn_load, lbl_load):
        t = self.tables[idx]
        if not t["token"].strip():
            messagebox.showwarning("提示", "请先填写并保存 Token")
            return
        for b in (btn_upd, btn_conn, btn_load):
            b.configure(state="disabled")
        lbl_upd.configure(text="连接中...", fg="#2d6cdf")
        threading.Thread(target=self._update_table_bg,
                         args=(idx, btn_upd, lbl_upd, btn_conn, lbl_conn,
                               btn_load, lbl_load),
                         daemon=True).start()

    def _update_table_bg(self, idx, btn_upd, lbl_upd,
                         btn_conn, lbl_conn, btn_load, lbl_load):
        t = self.tables[idx]
        base   = self.v_base_url.get().rstrip("/")
        origin = self.v_origin.get()
        uid    = self.v_user_id.get()

        # 1. 连接
        try:
            sheets = hq_get_sheets(base, origin, uid, t["token"])
            t["_sheets"]    = sheets
            t["_connected"] = True
            self.after(0, lambda: lbl_conn.configure(text="已连接", fg="#2a8a2a"))
        except Exception as e:
            msg = str(e)
            def _fail_conn():
                for b in (btn_upd, btn_conn, btn_load):
                    b.configure(state="normal")
                lbl_upd.configure(text="连接失败", fg="red")
                messagebox.showerror("连接失败", msg)
            self.after(0, _fail_conn)
            return

        # 2. 校验 active_sheet_ids 与线上 Sheet 的一致性
        online_ids   = {s["sheetId"] for s in sheets}
        online_titles = {s["sheetId"]: s.get("title", s["sheetId"]) for s in sheets}
        saved_ids    = t["active_sheet_ids"]

        if not saved_ids:
            # 从未配置过 → 自动选全部
            active_ids = list(online_ids)
            t["active_sheet_ids"] = active_ids
            self.after(0, lambda: self._log(
                f"[{t['name']}] 首次加载，自动选择全部 {len(active_ids)} 个 Sheet"))
        else:
            stale_ids = [sid for sid in saved_ids if sid not in online_ids]
            active_ids = [sid for sid in saved_ids if sid in online_ids]

            if stale_ids:
                # 有失效 ID，自动剔除并提示
                t["active_sheet_ids"] = active_ids
                stale_warn = "、".join(
                    online_titles.get(sid, sid) for sid in stale_ids)
                def _warn_stale(w=stale_warn, n=t["name"], left=len(active_ids)):
                    self._log(f"[{n}] ⚠ 以下 Sheet 在线上已不存在，已自动剔除：{w}")
                    if left == 0:
                        self._log(f"[{n}] ✗ 无有效 Sheet 可读取，请在表格库重新选择 Sheet")
                self.after(0, _warn_stale)

            if not active_ids:
                # 所有保存的 Sheet 都失效
                def _abort():
                    for b in (btn_upd, btn_conn, btn_load):
                        b.configure(state="normal")
                    lbl_upd.configure(text="所有 Sheet 已失效，请重新选择", fg="red")
                    messagebox.showwarning(
                        "Sheet 失效",
                        f"表格「{t['name']}」中保存的 Sheet 在线上均已不存在，\n"
                        "请点击「连接并获取 Sheet 列表」重新勾选 Sheet。")
                self.after(0, _abort)
                return

        self.after(0, lambda: lbl_upd.configure(text="读取中...", fg="#2d6cdf"))

        # 3. 读取
        try:
            def progress(sheet_title, n):
                self.after(0, lambda tt=sheet_title, x=n:
                           lbl_upd.configure(text=f"「{tt}」已读 {x} 行...",
                                             fg="#2d6cdf"))
            rows = hq_read_table(base, origin, uid,
                                 t["token"], t["_sheets"],
                                 active_ids, progress)
            t["_rows"]    = rows
            t["_headers"] = [_cell_str(v) for v in rows[0]] if rows else []
            t["_loaded"]  = True
            cnt = max(len(rows) - 1, 0)
            self._save_config()

            def _ok():
                for b in (btn_upd, btn_conn, btn_load):
                    b.configure(state="normal")
                lbl_upd.configure(text=f"✓ 已更新，{cnt} 行", fg="#2a8a2a")
                lbl_load.configure(text=f"✓ 已加载 {cnt} 行", fg="#2a8a2a")
                self._refresh_table_list()
                self._build_sheet_checkboxes(idx)
                if self._tab3_cur_idx == idx:
                    self._refresh_tab3_cols()
                self._log(f"[{t['name']}] 更新完成：{cnt} 行")
            self.after(0, _ok)
        except Exception as e:
            msg = str(e)
            def _fail_load():
                for b in (btn_upd, btn_conn, btn_load):
                    b.configure(state="normal")
                lbl_upd.configure(text="读取失败", fg="red")
                messagebox.showerror("读取失败", msg)
            self.after(0, _fail_load)

    # ═══ Tab3：列映射 ══════════════════════════════════

    def _build_tab3(self):
        p = self.tab3

        # 顶部：选择表格
        top = tk.Frame(p); top.pack(fill="x", padx=10, pady=6)
        tk.Label(top, text="配置表格：", font=("Arial", 10, "bold")).pack(side="left")
        self.cb_t3_table = ttk.Combobox(top, width=24, state="readonly")
        self.cb_t3_table.pack(side="left", padx=6)
        self.cb_t3_table.bind("<<ComboboxSelected>>", self._on_tab3_table_change)
        self.btn_t3_refresh = ttk.Button(top, text="刷新表格列表",
                   command=self._do_refresh_tab3_table_list)
        self.btn_t3_refresh.pack(side="left", padx=4)
        self.lbl_t3_hint = tk.Label(top, text="（请先在「表格库」连接并读取数据）",
                                    fg="#888")
        self.lbl_t3_hint.pack(side="left", padx=8)

        ttk.Separator(p).pack(fill="x", padx=10, pady=2)

        # 主体滚动区
        container = tk.Frame(p)
        container.pack(fill="both", expand=True, padx=10, pady=4)

        # 匹配键区域
        kf = ttk.LabelFrame(container, text="匹配键（AND 逻辑，第1对必填，第2/3对可选）",
                            padding=8)
        kf.pack(fill="x", pady=4)
        tk.Label(kf, text="", width=10).grid(row=0, column=0)
        tk.Label(kf, text="本地列", width=22, anchor="center",
                 font=("Arial", 9, "bold"), fg="#333").grid(row=0, column=1, padx=4)
        tk.Label(kf, text="飞书列", width=22, anchor="center",
                 font=("Arial", 9, "bold"), fg="#333").grid(row=0, column=2, padx=4)

        self._t3_key_cbs_local  = []
        self._t3_key_cbs_feishu = []
        labels = ["第1对（必填）：", "第2对（可选）：", "第3对（可选）："]
        colors = ["#c00000", "#555", "#555"]
        for i in range(3):
            tk.Label(kf, text=labels[i], fg=colors[i], width=10,
                     anchor="w").grid(row=i+1, column=0, sticky="w", pady=4)
            cb_l = ttk.Combobox(kf, textvariable=self._tab3_key_vars[i][0],
                                width=22, state="readonly")
            cb_l.grid(row=i+1, column=1, padx=4, sticky="w")
            cb_f = ttk.Combobox(kf, textvariable=self._tab3_key_vars[i][1],
                                width=22, state="readonly")
            cb_f.grid(row=i+1, column=2, padx=4, sticky="w")
            if i > 0:
                ttk.Button(kf, text="清除",
                           command=lambda a=i: (
                               self._tab3_key_vars[a][0].set(""),
                               self._tab3_key_vars[a][1].set("")
                           )).grid(row=i+1, column=3, padx=4)
            self._t3_key_cbs_local.append(cb_l)
            self._t3_key_cbs_feishu.append(cb_f)

        # 提取列区域
        cf = ttk.LabelFrame(container, text="要从飞书提取的列", padding=8)
        cf.pack(fill="x", pady=4)
        tb = tk.Frame(cf); tb.pack(anchor="w", pady=(0, 4))
        ttk.Button(tb, text="全选",
                   command=lambda: [v.set(True) for v in self._tab3_col_vars.values()]
                   ).pack(side="left", padx=(0, 4))
        ttk.Button(tb, text="全不选",
                   command=lambda: [v.set(False) for v in self._tab3_col_vars.values()]
                   ).pack(side="left")
        # 可滚动 checkbox 区
        outer = tk.Frame(cf, bd=1, relief="sunken"); outer.pack(fill="x")
        self._tab3_chk_canvas = tk.Canvas(outer, height=120, highlightthickness=0)
        csb = ttk.Scrollbar(outer, orient="vertical",
                            command=self._tab3_chk_canvas.yview)
        self._tab3_chk_canvas.configure(yscrollcommand=csb.set)
        self._tab3_chk_canvas.pack(side="left", fill="both", expand=True)
        csb.pack(side="right", fill="y")
        self._tab3_chk_inner = tk.Frame(self._tab3_chk_canvas)
        self._tab3_chk_win = self._tab3_chk_canvas.create_window(
            (0, 0), window=self._tab3_chk_inner, anchor="nw")
        self._tab3_chk_inner.bind(
            "<Configure>",
            lambda e: self._tab3_chk_canvas.configure(
                scrollregion=self._tab3_chk_canvas.bbox("all")))
        self._tab3_chk_canvas.bind(
            "<Configure>",
            lambda e: self._tab3_chk_canvas.itemconfig(
                self._tab3_chk_win, width=e.width))
        self._tab3_chk_canvas.bind_all(
            "<MouseWheel>",
            lambda e: self._tab3_chk_canvas.yview_scroll(
                -1*(e.delta//120), "units"))
        tk.Label(self._tab3_chk_inner,
                 text="请先在「表格库」读取此表格数据，再回此处配置",
                 fg="#888").grid(row=0, column=0, columnspan=4, pady=8)

        # 保存按钮
        ttk.Button(container, text="保存此表格列映射配置",
                   command=self._save_tab3_to_table_and_file).pack(anchor="w", pady=6)
        self.lbl_t3_save = tk.Label(container, text="", fg="#555")
        self.lbl_t3_save.pack(anchor="w")
        # 列失效警告（有失效列时显示）
        self.lbl_t3_col_warn = tk.Label(container, text="", fg="#c00000",
                                        font=("Arial", 9), wraplength=420,
                                        justify="left")
        self.lbl_t3_col_warn.pack(anchor="w")

        self._refresh_tab3_table_list()

    def _do_refresh_tab3_table_list(self):
        """点击「刷新表格列表」按钮时调用：立刻显示提示，再执行刷新。"""
        self.btn_t3_refresh.configure(state="disabled")
        self.lbl_t3_hint.configure(text="刷新中...", fg="#2d6cdf")
        self.after(20, self._refresh_tab3_table_list_and_restore)

    def _refresh_tab3_table_list_and_restore(self):
        self._refresh_tab3_table_list()
        self.btn_t3_refresh.configure(state="normal")

    def _refresh_tab3_table_list(self):
        loaded = [t["name"] for t in self.tables if t["_loaded"]]
        self.cb_t3_table["values"] = loaded
        if loaded:
            self.lbl_t3_hint.configure(
                text=f"找到 {len(loaded)} 个已加载表格", fg="#2a8a2a")
            self.after(2000, lambda: self.lbl_t3_hint.configure(text=""))
        else:
            self.lbl_t3_hint.configure(text="（请先在「表格库」读取数据）", fg="#c00000")

    def _on_tab3_table_change(self, _event=None):
        self._save_tab3_to_table()
        name = self.cb_t3_table.get()
        idx = next((i for i, t in enumerate(self.tables) if t["name"] == name), None)
        if idx is None:
            return
        self._tab3_cur_idx = idx
        # 先立刻显示"加载中"提示，避免 UI 假死
        self.lbl_t3_hint.configure(text=f"加载「{name}」列信息...", fg="#2d6cdf")
        self.cb_t3_table.configure(state="disabled")
        self.after(15, lambda: self._load_tab3_table(idx))

    def _load_tab3_table(self, idx):
        """实际构建 Tab3 列配置区域（延迟执行，避免主线程卡顿）。"""
        t = self.tables[idx]
        fs_headers = t["_headers"]

        # 检测飞书匹配键是否有失效项，将失效值加回下拉列表（带 ⚠ 前缀）
        stale_keys = []
        for i in range(3):
            fk = t["feishu_key_names"][i]
            if fk and fk not in fs_headers:
                stale_keys.append(fk)
        for cb in self._t3_key_cbs_feishu:
            extra = [f"⚠ {k}（已失效）" for k in stale_keys]
            cb["values"] = fs_headers + extra

        for cb in self._t3_key_cbs_local:
            cb["values"] = self.local_headers
        # 恢复已保存的 key 值（失效的值仍然保留显示，提醒用户修改）
        for i in range(3):
            self._tab3_key_vars[i][0].set(t["local_key_names"][i])
            self._tab3_key_vars[i][1].set(t["feishu_key_names"][i])

        # 刷新列 checkboxes
        self._refresh_tab3_cols()
        self.lbl_t3_save.configure(text="")

        # 更新警告标签
        stale_fetch = [c for c in t["fetch_col_names"] if c not in fs_headers]
        warn_parts = []
        if stale_keys:
            warn_parts.append(f"匹配键已失效：{' / '.join(stale_keys)}")
        if stale_fetch:
            warn_parts.append(f"提取列已失效：{' / '.join(stale_fetch)}")
        if warn_parts:
            self.lbl_t3_col_warn.configure(
                text="⚠ 飞书列发生变化，请重新配置以下列：\n" + "\n".join(warn_parts))
        else:
            self.lbl_t3_col_warn.configure(text="")

        # 恢复 UI 状态
        self.cb_t3_table.configure(state="readonly")
        self.lbl_t3_hint.configure(text="")

    def _refresh_tab3_cols(self):
        if self._tab3_cur_idx is None:
            return
        t = self.tables[self._tab3_cur_idx]
        inner = self._tab3_chk_inner
        for w in inner.winfo_children():
            w.destroy()
        self._tab3_col_vars = {}
        if not t["_headers"]:
            tk.Label(inner, text="暂无列数据", fg="#888").grid(row=0, column=0, pady=8)
            return
        saved = set(t["fetch_col_names"])
        current = [h for h in t["_headers"] if h]
        current_set = set(current)
        for i, h in enumerate(current):
            var = tk.BooleanVar(value=(h in saved))
            self._tab3_col_vars[h] = var
            tk.Checkbutton(inner, text=h, variable=var, anchor="w",
                           font=("Consolas", 9), padx=6, pady=1
                           ).grid(row=i // 3, column=i % 3, sticky="w")
        # 已失效的提取列（保存在配置中但线上已不存在）
        stale_fetch = [c for c in t["fetch_col_names"] if c not in current_set]
        if stale_fetch:
            offset = len(current)
            sep_row = offset // 3 + (1 if offset % 3 else 0)
            tk.Label(inner, text="以下列在飞书中已不存在，请重新选择：",
                     fg="#c00000", font=("Arial", 8)
                     ).grid(row=sep_row, column=0, columnspan=3, sticky="w", pady=(4, 0))
            for j, c in enumerate(stale_fetch):
                var = tk.BooleanVar(value=True)   # 保持勾选，让用户看到并手动取消
                self._tab3_col_vars[c] = var
                tk.Checkbutton(inner, text=f"⚠ {c}", variable=var, anchor="w",
                               fg="#c00000", font=("Consolas", 9), padx=6, pady=1
                               ).grid(row=sep_row + 1 + j // 3,
                                      column=j % 3, sticky="w")

    def _save_tab3_to_table(self):
        idx = self._tab3_cur_idx
        if idx is None or idx >= len(self.tables):
            return
        t = self.tables[idx]
        for i in range(3):
            t["local_key_names"][i]  = self._tab3_key_vars[i][0].get()
            t["feishu_key_names"][i] = self._tab3_key_vars[i][1].get()
        t["fetch_col_names"] = [name for name, var in self._tab3_col_vars.items()
                                if var.get()]

    def _save_tab3_to_table_and_file(self):
        self._save_tab3_to_table()
        self._save_config()
        self.lbl_t3_save.configure(text="✓ 已保存", fg="#2a8a2a")

    # ═══ Tab4：运行 ════════════════════════════════════

    def _build_tab4(self):
        p = self.tab4

        f1 = ttk.LabelFrame(p, text="输出文件", padding=8)
        f1.pack(fill="x", padx=10, pady=6)
        r = tk.Frame(f1); r.pack(fill="x")
        ttk.Entry(r, textvariable=self.v_out_path, width=52).pack(side="left")
        ttk.Button(r, text="另存为...", command=self._browse_output).pack(side="left", padx=6)

        f2 = ttk.LabelFrame(p, text="已启用表格状态", padding=8)
        f2.pack(fill="x", padx=10, pady=4)
        self._run_status_frame = f2
        self._refresh_run_status()

        # 开始匹配按钮（自动加载未就绪的已启用表格）
        self.btn_run = tk.Button(p, text="开始匹配",
                                 font=("Arial", 13, "bold"),
                                 bg="#2d6cdf", fg="white", relief="flat",
                                 padx=24, pady=8, command=self._do_run)
        self.btn_run.pack(pady=(8, 2))
        self.lbl_run_status = tk.Label(p, text="", font=("Arial", 10))
        self.lbl_run_status.pack()

        self.log_box = scrolledtext.ScrolledText(p, font=("Consolas", 9),
                                                 state="disabled",
                                                 bg="#1e1e1e", fg="#d4d4d4",
                                                 height=10)
        self.log_box.pack(fill="both", expand=True, padx=10, pady=6)
        ttk.Button(p, text="清空日志",
                   command=lambda: (self.log_box.configure(state="normal"),
                                    self.log_box.delete("1.0", "end"),
                                    self.log_box.configure(state="disabled"))
                   ).pack(anchor="e", padx=10, pady=2)

    def _refresh_run_status(self):
        f = self._run_status_frame
        for w in f.winfo_children():
            w.destroy()
        enabled = [t for t in self.tables if t["enabled"]]
        if not enabled:
            tk.Label(f, text="尚未启用任何表格（请在「表格库」中勾选启用）",
                     fg="#c00000").pack(anchor="w")
            return
        for t in enabled:
            if t["_loaded"]:
                fc = len(t["fetch_col_names"])
                kp = sum(1 for lk, fk in zip(t["local_key_names"], t["feishu_key_names"])
                         if lk and fk)
                cfg_ok = fc > 0 and kp > 0
                txt = (f"✓  {t['name']}  —  已加载 {max(len(t['_rows'])-1,0)} 行"
                       f"，{kp} 对匹配键，{fc} 列提取"
                       + ("" if cfg_ok else "  ⚠ 请在「列映射」完成配置"))
                clr = "#2a8a2a" if cfg_ok else "#e07000"
            else:
                txt = f"⚠  {t['name']}  —  已启用，未读取数据"
                clr = "#c00000"
            tk.Label(f, text=txt, fg=clr, anchor="w").pack(anchor="w", pady=1)

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")])
        if path:
            self.v_out_path.set(path)

    # ── 一键加载全部已启用表格 ─────────────────────────────

    def _do_batch_update_all(self):
        """连接并更新所有配置了 Token 的表格（不限是否启用）。"""
        targets = [(i, t) for i, t in enumerate(self.tables)
                   if t["token"].strip()]
        if not targets:
            messagebox.showwarning("提示", "没有配置 Token 的表格")
            return
        self._set_batch_btns("disabled")
        self._lbl_update_all.configure(text="更新中...", fg="#2d6cdf")
        self._log(f"\n── 一键更新全部（{len(targets)} 张表）：开始 ──")
        threading.Thread(target=self._batch_load_bg, args=(targets,),
                         daemon=True).start()

    def _set_batch_btns(self, state):
        """统一设置所有批量操作按钮的状态。"""
        for btn in (getattr(self, "btn_update_all", None),
                    getattr(self, "btn_run",        None)):
            if btn:
                try:
                    btn.configure(state=state)
                except Exception:
                    pass

    def _batch_load_bg(self, targets):
        base   = self.v_base_url.get().rstrip("/")
        origin = self.v_origin.get()
        uid    = self.v_user_id.get()
        ok_cnt = 0
        total  = len(targets)
        for seq, (i, t) in enumerate(targets, start=1):
            name = t["name"]
            # 进度标签：更新中 (X/Y)：表名
            def _set_progress(n=name, s=seq, tot=total):
                if hasattr(self, "_lbl_update_all"):
                    self._lbl_update_all.configure(
                        text=f"更新中 ({s}/{tot})：{n}", fg="#2d6cdf")
            self.after(0, _set_progress)
            # 1. 连接（获取 sheet 元信息）
            self.after(0, lambda n=name: self._log(f"[{n}] 连接中..."))
            try:
                sheets = hq_get_sheets(base, origin, uid, t["token"])
                t["_sheets"]    = sheets
                t["_connected"] = True
                self.after(0, lambda n=name, s=sheets:
                           self._log(f"[{n}] 连接成功，{len(s)} 个 Sheet"))
            except Exception as e:
                self.after(0, lambda n=name, e=e:
                           self._log(f"[{n}] 连接失败：{e}"))
                self.after(0, self._refresh_table_list)
                continue

            # 2. 校验 active_sheet_ids 与线上 Sheet 一致性
            online_ids    = {s["sheetId"] for s in t["_sheets"]}
            online_titles = {s["sheetId"]: s.get("title", s["sheetId"])
                             for s in t["_sheets"]}
            saved_ids = t["active_sheet_ids"]

            if not saved_ids:
                active_ids = list(online_ids)
                t["active_sheet_ids"] = active_ids
                self.after(0, lambda n=name, c=len(active_ids):
                           self._log(f"[{n}] 首次加载，自动选择全部 {c} 个 Sheet"))
            else:
                stale = [sid for sid in saved_ids if sid not in online_ids]
                active_ids = [sid for sid in saved_ids if sid in online_ids]
                if stale:
                    stale_names = "、".join(
                        online_titles.get(sid, sid) for sid in stale)
                    t["active_sheet_ids"] = active_ids
                    self.after(0, lambda n=name, w=stale_names:
                               self._log(f"[{n}] ⚠ Sheet 已失效并自动剔除：{w}"))
                if not active_ids:
                    self.after(0, lambda n=name:
                               self._log(f"[{n}] ✗ 所有 Sheet 均已失效，跳过，"
                                         "请在表格库重新选择 Sheet"))
                    self.after(0, self._refresh_table_list)
                    continue

            # 3. 读取数据
            self.after(0, lambda n=name: self._log(f"[{n}] 读取数据中..."))
            try:
                def _progress(sheet_title, n_rows, n=name):
                    self.after(0, lambda t=sheet_title, x=n_rows, nm=n:
                               self._log(f"[{nm}] 「{t}」已读 {x} 行..."))
                rows = hq_read_table(base, origin, uid,
                                     t["token"], t["_sheets"],
                                     active_ids, _progress)
                t["_rows"]    = rows
                t["_headers"] = [_cell_str(v) for v in rows[0]] if rows else []
                t["_loaded"]  = True
                cnt = max(len(rows) - 1, 0)
                ok_cnt += 1
                self.after(0, lambda n=name, c=cnt:
                           self._log(f"[{n}] ✓ 加载完成，{c} 行"))
                self.after(0, self._refresh_table_list)
            except Exception as e:
                self.after(0, lambda n=name, e=e:
                           self._log(f"[{n}] 读取失败：{e}"))
                self.after(0, self._refresh_table_list)

        self._save_config()
        self.after(0, self._on_batch_load_done, ok_cnt, len(targets))

    def _on_batch_load_done(self, ok_cnt, total):
        self._set_batch_btns("normal")
        if hasattr(self, "_lbl_update_all"):
            self._lbl_update_all.configure(
                text=f"✓ 完成 {ok_cnt}/{total}", fg="#2a8a2a")
        self._refresh_run_status()
        self._refresh_table_list()
        self._log(f"── 批量更新完成：{ok_cnt}/{total} 个表格成功 ──")

    # ── 运行匹配 ──────────────────────────────────────────

    def _do_run(self):
        self._save_tab3_to_table()
        self._refresh_run_status()

        if not self.ws:
            messagebox.showerror("错误", "请先在「本地文件」选择 BOM 文件")
            return

        enabled = [(i, t) for i, t in enumerate(self.tables)
                   if t["enabled"] and t["token"].strip()]
        if not enabled:
            messagebox.showerror("错误", "没有已启用的表格，请先在「表格库」中启用并配置 Token")
            return

        out_file = _unique_path(self.v_out_path.get().strip() or "飞书匹配结果.xlsx")
        self._set_batch_btns("disabled")
        self._start_spinner()
        self.nb.select(3)
        self._log(f"\n── 开始匹配（含自动加载未就绪表格）──")
        threading.Thread(
            target=self._run_full_bg,
            args=(enabled, out_file),
            daemon=True,
        ).start()

    def _run_full_bg(self, enabled, out_file):
        """后台：1. 加载所有未就绪的已启用表 → 2. 执行匹配。"""
        base   = self.v_base_url.get().rstrip("/")
        origin = self.v_origin.get()
        uid    = self.v_user_id.get()

        # ── 阶段 1：加载未就绪的表 ──────────────────────────
        need_load = [(i, t) for i, t in enabled if not t["_loaded"]]
        if need_load:
            self.after(0, lambda n=len(need_load):
                       self._log(f"[准备] 有 {n} 张表尚未加载，自动加载中..."))
        for i, t in need_load:
            name = t["name"]
            # 连接
            self.after(0, lambda n=name: self._log(f"[{n}] 连接中..."))
            try:
                sheets = hq_get_sheets(base, origin, uid, t["token"])
                t["_sheets"]    = sheets
                t["_connected"] = True
            except Exception as e:
                self.after(0, lambda n=name, e=e:
                           self._log(f"[{n}] 连接失败：{e}，已跳过"))
                continue

            # 校验 active_sheet_ids
            online_ids    = {s["sheetId"] for s in t["_sheets"]}
            saved_ids     = t["active_sheet_ids"]
            if not saved_ids:
                active_ids = list(online_ids)
                t["active_sheet_ids"] = active_ids
            else:
                stale = [sid for sid in saved_ids if sid not in online_ids]
                active_ids = [sid for sid in saved_ids if sid in online_ids]
                if stale:
                    stale_names = "、".join(
                        s.get("title", s["sheetId"])
                        for s in t["_sheets"] if s["sheetId"] in stale)
                    t["active_sheet_ids"] = active_ids
                    self.after(0, lambda n=name, w=stale_names:
                               self._log(f"[{n}] ⚠ 失效 Sheet 已剔除：{w}"))
                if not active_ids:
                    self.after(0, lambda n=name:
                               self._log(f"[{n}] ✗ 所有 Sheet 均已失效，已跳过"))
                    continue

            # 读取
            self.after(0, lambda n=name: self._log(f"[{n}] 读取数据中..."))
            try:
                def _prog(sheet_title, n_rows, n=name):
                    self.after(0, lambda tt=sheet_title, x=n_rows, nm=n:
                               self._log(f"[{nm}] 「{tt}」已读 {x} 行..."))
                rows = hq_read_table(base, origin, uid,
                                     t["token"], t["_sheets"],
                                     active_ids, _prog)
                t["_rows"]    = rows
                t["_headers"] = [_cell_str(v) for v in rows[0]] if rows else []
                t["_loaded"]  = True
                cnt = max(len(rows) - 1, 0)
                self.after(0, lambda n=name, c=cnt:
                           self._log(f"[{n}] ✓ 加载完成，{c} 行"))
            except Exception as e:
                self.after(0, lambda n=name, e=e:
                           self._log(f"[{n}] 读取失败：{e}，已跳过"))

        self._save_config()
        self.after(0, self._refresh_table_list)
        self.after(0, self._refresh_run_status)

        # ── 阶段 2：构建 prepared_tables ────────────────────
        hr = self.v_hdr_row.get()
        local_headers = [self.ws.cell(row=hr, column=ci).value
                         for ci in range(1, self.ws.max_column + 1)]

        prepared_tables = []
        for _, t in enabled:
            if not t["_loaded"]:
                self.after(0, lambda n=t["name"]:
                           self._log(f"[{n}] 加载未成功，跳过匹配"))
                continue
            tname = t["name"]
            fs_header_set = set(t["_headers"])
            local_key_cols, feishu_key_cols = [], []
            for lk, fk in zip(t["local_key_names"], t["feishu_key_names"]):
                if not lk or not fk:
                    continue
                # 检测匹配键是否还在线上列中
                if fk not in fs_header_set:
                    self.after(0, lambda n=tname, f=fk:
                               self._log(f"[{n}] ⚠ 飞书匹配键「{f}」已不存在，此键对跳过"))
                    continue
                try:
                    lc = next(ci+1 for ci, h in enumerate(local_headers)
                              if _cell_str(h) == lk)
                except StopIteration:
                    self.after(0, lambda n=tname, l=lk:
                               self._log(f"[{n}] ⚠ 本地匹配键「{l}」在表头中未找到，此键对跳过"))
                    continue
                fc = t["_headers"].index(fk)
                local_key_cols.append(lc)
                feishu_key_cols.append(fc)
            if not local_key_cols:
                self.after(0, lambda n=tname:
                           self._log(f"[{n}] ✗ 无有效匹配键（请到「列映射」重新配置），跳过"))
                continue
            if not t["fetch_col_names"]:
                self.after(0, lambda n=tname:
                           self._log(f"[{n}] ✗ 未选择提取列，跳过"))
                continue

            fetch_idxs = []
            stale_fetch = []
            for col_name in t["fetch_col_names"]:
                if col_name in fs_header_set:
                    fetch_idxs.append(t["_headers"].index(col_name))
                else:
                    fetch_idxs.append(-1)
                    stale_fetch.append(col_name)
            if stale_fetch:
                self.after(0, lambda n=tname, s=stale_fetch:
                           self._log(f"[{n}] ⚠ 提取列已失效（输出空白）：{' / '.join(s)}"))

            lookup = {}
            for row in t["_rows"][1:]:
                key = tuple(_cell_str(row[fc]) if fc < len(row) else ""
                            for fc in feishu_key_cols)
                if not any(key):
                    continue
                vals = {col_name: (_cell_str(row[idx]) if 0 <= idx < len(row) else "")
                        for col_name, idx in zip(t["fetch_col_names"], fetch_idxs)}
                lookup.setdefault(key, []).append(vals)

            prepared_tables.append({
                "name":            t["name"],
                "local_key_cols":  local_key_cols,
                "lookup":          lookup,
                "fetch_col_names": t["fetch_col_names"],
            })

        if not prepared_tables:
            def _no_tables():
                self._stop_spinner()
                self._set_batch_btns("normal")
                messagebox.showerror("错误", "没有可用的已启用表格（请检查列映射配置）")
            self.after(0, _no_tables)
            return

        seen, all_fetch_cols = set(), []
        for pt in prepared_tables:
            for col_name in pt["fetch_col_names"]:
                if col_name not in seen:
                    seen.add(col_name)
                    all_fetch_cols.append(col_name)

        self.after(0, lambda: self._log(f"\n── 开始匹配 ──"))
        self.after(0, lambda: self._log(
            f"启用表格：{', '.join(pt['name'] for pt in prepared_tables)}"))
        self.after(0, lambda: self._log(
            f"输出列：{', '.join(all_fetch_cols)} + 来源表格"))
        self.after(0, lambda: self._log(
            f"输出文件：{os.path.abspath(out_file)}"))

        # ── 阶段 3：执行匹配 ────────────────────────────────
        self._run_bg(prepared_tables, all_fetch_cols, out_file)

    def _run_bg(self, prepared_tables, all_fetch_cols, out_file):
        try:
            total, matched, unmatched = do_match_multi(
                self.ws, self.v_hdr_row.get(),
                prepared_tables, all_fetch_cols,
                out_file, self._log,
            )
            abs_path = os.path.abspath(out_file)
            self._log(f"\n✅ 完成！输出：{abs_path}")
            try:
                folder = os.path.dirname(abs_path)
                if sys.platform == "win32":   os.startfile(folder)
                elif sys.platform == "darwin": subprocess.Popen(["open", folder])
                else:                          subprocess.Popen(["xdg-open", folder])
            except Exception:
                pass
            self.after(0, self._stop_spinner)
            self.after(0, lambda: self.lbl_run_status.configure(
                text=f"✅ 完成！共 {total} 行，命中 {matched}，未匹配 {unmatched}",
                fg="#2a8a2a"))
            self.after(0, lambda: messagebox.showinfo(
                "完成",
                f"匹配完成！\n\n共 {total} 行\n命中：{matched}\n未匹配：{unmatched}\n\n{abs_path}"))
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            self._log(f"\n❌ 错误：{e}\n{tb}")
            self.after(0, self._stop_spinner)
            self.after(0, lambda: self.lbl_run_status.configure(
                text="出错，请查看日志", fg="red"))
            self.after(0, lambda: messagebox.showerror("错误", str(e)))
        finally:
            self.after(0, lambda: self._set_batch_btns("normal"))

    # ═══ 本地文件 ══════════════════════════════════════

    def _browse_local(self):
        path = filedialog.askopenfilename(
            title="选择本地 BOM 文件",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("所有文件", "*.*")])
        if not path:
            return
        self.v_in_path.set(path)
        default_out = os.path.join(os.path.dirname(path), "飞书匹配结果.xlsx")
        self.v_out_path.set(default_out)
        self._log(f"本地文件：{path}")
        threading.Thread(target=self._load_wb_bg, args=(path,), daemon=True).start()

    def _load_wb_bg(self, path):
        try:
            wb = open_workbook_compat(path, data_only=True)
            self.wb = wb
            self.after(0, self._on_wb_loaded)
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: messagebox.showerror("错误", f"无法打开文件：\n{msg}"))

    def _on_wb_loaded(self):
        self.cb_sheet["values"] = self.wb.sheetnames
        self.v_sheet.set(self.wb.sheetnames[0])
        self._load_local_sheet()

    def _load_local_sheet(self):
        if not self.wb:
            return
        name = self.v_sheet.get()
        if name not in self.wb.sheetnames:
            return
        self.ws = self.wb[name]
        self._log(f"Sheet：{name}（{self.ws.max_row} 行 × {self.ws.max_column} 列）")
        self._update_preview()
        self._refresh_local_headers()

    def _update_preview(self):
        tree = self.preview_tree
        tree.delete(*tree.get_children())
        if not self.ws:
            return
        mc = min(self.ws.max_column, 10)
        cols = [get_column_letter(i) for i in range(1, mc + 1)]
        tree["columns"] = cols
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=100, anchor="w")
        for ri in range(1, min(11, self.ws.max_row + 1)):
            vals = [str(self.ws.cell(row=ri, column=ci).value or "")[:24]
                    for ci in range(1, mc + 1)]
            tree.insert("", "end", values=vals)

    def _refresh_local_headers(self):
        hr = self.v_hdr_row.get()
        headers = []
        for ci in range(1, self.ws.max_column + 1):
            h = self.ws.cell(row=hr, column=ci).value
            if h:
                headers.append(str(h).strip())
        self.local_headers = headers
        # 同步更新 Tab3 本地列下拉
        for cb in self._t3_key_cbs_local:
            cb["values"] = headers
        self._log(f"本地表头（{len(headers)} 列）：{', '.join(headers[:6])}"
                  f"{'…' if len(headers) > 6 else ''}")

    # ═══ 日志 & 转圈 ════════════════════════════════════

    def _log(self, msg):
        def _w():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _w)

    def _start_spinner(self):
        self._spinning = True
        self._spin_step = 0
        self._spin()

    def _spin(self):
        if not self._spinning:
            return
        frames = ["◐ 匹配中...", "◓ 匹配中...", "◑ 匹配中...", "◒ 匹配中..."]
        self.lbl_run_status.configure(text=frames[self._spin_step % 4], fg="#2d6cdf")
        self._spin_step += 1
        self._spin_job = self.after(200, self._spin)

    def _stop_spinner(self):
        self._spinning = False
        if hasattr(self, "_spin_job"):
            self.after_cancel(self._spin_job)


if __name__ == "__main__":
    app = FeishuMatchApp()
    app.mainloop()
