# -*- coding: utf-8 -*-
"""
PLM 上传工具 v1.6
整机BOM配置表 → PLM 系统导入格式

【源文件（整机BOM配置表）关键列】
  序号   —— 相同序号 = 同一组主二供
  HQ PN  —— HQ 料号，写入 PLM「料号」列
  主二供 —— 主供 / 二供 / 三供 / 四供…
  用量   —— 单耗

【转换规则】
  1. 用量为空           → 不导入
  2. 用量为 0           → 导入，PLM「单耗」列留空
  3. 主供 + 用量 > 0    → 导入，PLM「单耗」填实际用量
  4. 二供/三供/…        → 导入，「单耗」留空，「主辅BOM标记」填「二供」「三供」…

【PLM 输出格式】
  从 A 列开始：序号(A) 料号(B) 型号(C) 物料描述(D) 单耗(E) …
  第 3 行为列表头，第 4 行起为数据

依赖：pip install openpyxl
运行：python plm_upload.py
"""

import sys, subprocess

try:
    import openpyxl
except ImportError:
    print("未检测到 openpyxl，正在自动安装，请稍候...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    print("安装完成！正在启动程序...")
    import openpyxl

from excel_compat import open_workbook_compat

import os, threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# ── PLM 输出列定义（从 A 列 = 第 1 列开始）──────────────────────
PLM_COL_START  = 1    # A 列
PLM_HEADER_ROW = 3    # 列表头在第 3 行
PLM_DATA_ROW   = 4    # 数据从第 4 行开始

PLM_HEADERS = [
    "序号",
    "料号",
    "型号",
    "物料描述",
    "单耗",
    "替代关系\n(A:完全替代/N:独供/X:不完全替代)",
    "位号",
    "生产厂家",
    "是否环保",
    "温敏属性",
    "备注",
    "主辅BOM标记\n(仅允许填写二供/三供/四供/五供/六供/七供/八供)",
    "MBG优选属性",
    "CBG优选属性",
    "DBG优选属性",
    "首制程",
    "次制程",
    "次制程单耗",
    "是否可量产下单",
    "次制程位号",
    "ABG优选属性",
    "IFM_PART",
    "PCD_PART",
    "是否受EAR管控",
    "ECCN",
]

# 关键列在 PLM_HEADERS 列表中的偏移（0-based）
PLM_IDX_SEQ  = 0   # A：序号
PLM_IDX_HQPN = 1   # B：料号 ← HQ PN
PLM_IDX_QTY  = 4   # E：单耗（只填主供，替代料留空让 PLM 自动归组）


# ── 工具函数 ────────────────────────────────────────────────────

def _unique_path(path):
    """若文件已存在，自动叠加 (1)(2)… 后缀。"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while os.path.exists(f"{base}({i}){ext}"):
        i += 1
    return f"{base}({i}){ext}"


def _safe_qty(v):
    """解析用量；空值/None 返回 None，否则返回 float。"""
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _col_to_int(s):
    """列字母转数字，失败返回 None。"""
    s = str(s).strip().upper()
    if not s:
        return None
    try:
        return column_index_from_string(s)
    except Exception:
        return None


def _int_to_col(n):
    return get_column_letter(n) if n else ""


def _detect_columns(ws, header_row):
    """
    扫描表头行，返回 ({字段名: 列号}, 已读列名列表)。
    字段名：seq / hq_pn / supply_type / qty
    宽松匹配：忽略换行、空格、括号内附注。
    """
    result = {}
    found_headers = []
    # 向后多扫几列，防止 max_column 偏小
    scan_cols = max((ws.max_column or 0) + 5, 30)
    for ci in range(1, scan_cols + 1):
        raw = ws.cell(row=header_row, column=ci).value
        if raw is None:
            continue
        # 清洗：去掉换行、多余空格，保留原始文本供调试
        h  = str(raw).replace("\n", "").replace("\r", "").strip()
        hl = h.lower().replace(" ", "")
        if h:
            found_headers.append(f"{_int_to_col(ci)}:{h}")
        if "序号" in h:
            result.setdefault("seq", ci)
        if "hq" in hl and "pn" in hl:
            result.setdefault("hq_pn", ci)
        if "主二供" in h or "主供" in h:
            result.setdefault("supply_type", ci)
        if "用量" in h or "单耗" in h:
            result.setdefault("qty", ci)
    return result, found_headers


# ── 转换核心 ─────────────────────────────────────────────────────

def do_convert(in_file, sheet_name, header_row,
               col_hqpn, col_stype, col_qty,
               project_name, out_file, log_cb):
    """
    读取源 BOM，写入 PLM 格式 Excel。
    返回 (total_written, skipped_empty_qty)。
    """
    wb_in = open_workbook_compat(in_file, data_only=True)
    ws_in = wb_in[sheet_name]
    max_col = ws_in.max_column

    # 收集非空数据行
    data_rows = []
    for ri in range(header_row + 1, ws_in.max_row + 1):
        rv = {ci: ws_in.cell(row=ri, column=ci).value
              for ci in range(1, max_col + 1)}
        if not any(v is not None and str(v).strip() for v in rv.values()):
            continue
        data_rows.append(rv)

    # ── 创建输出工作簿 ───────────────────────────────────────────
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "PLM导入"

    bdr = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    meta_font = Font(bold=True, size=10)

    # ── 行 1-2：PLM 模板头部信息 ─────────────────────────────────
    # 行1: 料号: [  ]  描述: [  ]  项目配置名: [project]  工程师: [  ]
    ws_out.cell(row=1, column=1, value="料号:").font = meta_font
    ws_out.cell(row=1, column=3, value="描述:").font = meta_font
    ws_out.cell(row=1, column=5, value="项目配置名:").font = meta_font
    ws_out.cell(row=1, column=6, value=project_name or "").font = Font(size=10)
    ws_out.cell(row=1, column=7, value="工程师:").font = meta_font

    # 行2: 版本: [  ]  替代项 [  ]  BOM名称: [  ]  归档部门: [  ]
    ws_out.cell(row=2, column=1, value="版本:").font = meta_font
    ws_out.cell(row=2, column=3, value="替代项").font = meta_font
    ws_out.cell(row=2, column=5, value="BOM名称:").font = meta_font
    ws_out.cell(row=2, column=7, value="归档部门:").font = meta_font

    # ── 行 3：列表头 ──────────────────────────────────────────────
    for offset, hdr_txt in enumerate(PLM_HEADERS):
        col_num = PLM_COL_START + offset
        c = ws_out.cell(row=PLM_HEADER_ROW, column=col_num, value=hdr_txt)
        c.font = Font(bold=True, color="FF0000", size=9)   # 红色加粗，与模板一致
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = bdr
        ws_out.column_dimensions[get_column_letter(col_num)].width = 14

    # 料号列加宽
    ws_out.column_dimensions[
        get_column_letter(PLM_COL_START + PLM_IDX_HQPN)].width = 22
    ws_out.row_dimensions[PLM_HEADER_ROW].height = 60

    # ── 写数据行（从第 4 行开始）────────────────────────────────
    dr = PLM_DATA_ROW
    total = 0
    skipped = 0
    seq = 0   # PLM 序号计数器

    for rv in data_rows:
        hqpn    = rv.get(col_hqpn)
        stype   = rv.get(col_stype)
        qty_raw = rv.get(col_qty)

        # HQ PN 为空 → 跳过
        if not hqpn or str(hqpn).strip() == "":
            continue

        # 用量为空 → 不导入
        qty = _safe_qty(qty_raw)
        if qty is None:
            skipped += 1
            log_cb(f"  跳过（用量为空）: {str(hqpn).strip()}")
            continue

        # 判断主供 / 替代料
        stype_str  = str(stype or "").strip()
        is_primary = (stype_str == "主供" or stype_str == "")

        # 主供行才递增序号
        if is_primary:
            seq += 1

        def wc(idx, val, row=dr):
            cc = ws_out.cell(row=row, column=PLM_COL_START + idx, value=val)
            cc.alignment = Alignment(horizontal="left", vertical="center")
            cc.border = bdr

        # 序号（同一组相同序号）
        wc(PLM_IDX_SEQ, seq)

        # 料号 ← HQ PN（每行必填）
        wc(PLM_IDX_HQPN, str(hqpn).strip())

        # 单耗：主供且用量 > 0 才填；替代料留空，PLM系统自动归组
        if is_primary and qty > 0:
            wc(PLM_IDX_QTY, qty)

        dr += 1
        total += 1

    wb_out.save(out_file)
    return total, skipped


# ── GUI ──────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PLM 上传工具 v1.6")
        self.resizable(False, False)

        self._in_file   = tk.StringVar()
        self._sheet     = tk.StringVar()
        self._hdr_row   = tk.StringVar(value="4")
        self._col_hqpn  = tk.StringVar()
        self._col_stype = tk.StringVar()
        self._col_qty   = tk.StringVar()
        self._proj_name = tk.StringVar()
        self._hdr_debounce = None   # 防抖定时器

        self._build_ui()

        # 表头行变更时自动触发识别（防抖 800ms）
        self._hdr_row.trace_add("write", self._on_hdr_change)


    # ── 界面构建 ──────────────────────────────────────────────
    def _build_ui(self):
        pad = dict(padx=8, pady=4)

        # ── 第一步：选择文件 ──────────────────────────────────
        f1 = ttk.LabelFrame(self, text="第一步：选择整机BOM配置表")
        f1.pack(fill="x", padx=10, pady=(10, 4))

        r = ttk.Frame(f1); r.pack(fill="x", **pad)
        ttk.Label(r, text="文件路径：").pack(side="left")
        ttk.Entry(r, textvariable=self._in_file, width=52).pack(side="left", padx=4)
        ttk.Button(r, text="浏览...", command=self._browse).pack(side="left")

        r2 = ttk.Frame(f1); r2.pack(fill="x", **pad)
        ttk.Label(r2, text="Sheet：").pack(side="left")
        self._sheet_cb = ttk.Combobox(r2, textvariable=self._sheet,
                                      state="readonly", width=22)
        self._sheet_cb.pack(side="left", padx=4)
        self._sheet_cb.bind("<<ComboboxSelected>>", lambda e: self._auto_detect())
        ttk.Label(r2, text="  表头行：").pack(side="left")
        ttk.Entry(r2, textvariable=self._hdr_row, width=4).pack(side="left")
        ttk.Button(r2, text="加载 / 刷新", command=self._load).pack(side="left", padx=8)

        # ── 第二步：列映射 ────────────────────────────────────
        f2 = ttk.LabelFrame(self, text="第二步：列映射  （自动识别，可手动填列字母覆盖，如 E）")
        f2.pack(fill="x", padx=10, pady=4)

        self._detect_log = tk.Text(f2, height=2, state="disabled",
                                   bg="#f5f5f5", font=("Courier", 9))
        self._detect_log.pack(fill="x", **pad)

        cf = ttk.Frame(f2); cf.pack(fill="x", **pad)
        for i, (lbl, var) in enumerate([
            ("HQ PN 列",  self._col_hqpn),
            ("主二供列",  self._col_stype),
            ("用量列",    self._col_qty),
        ]):
            ttk.Label(cf, text=lbl + "：").grid(row=0, column=i * 2,     sticky="e", padx=4)
            ttk.Entry(cf, textvariable=var, width=6).grid(row=0, column=i * 2 + 1, padx=2)

        # ── 第三步：输出设置 ──────────────────────────────────
        f3 = ttk.LabelFrame(self, text="第三步：输出设置")
        f3.pack(fill="x", padx=10, pady=4)

        r3 = ttk.Frame(f3); r3.pack(fill="x", **pad)
        ttk.Label(r3, text="项目名称：").pack(side="left")
        ttk.Entry(r3, textvariable=self._proj_name, width=32).pack(side="left", padx=4)
        ttk.Label(r3, text="（写入PLM文件头部，可留空）",
                  foreground="gray").pack(side="left")

        # ── 转换按钮 ──────────────────────────────────────────
        bf = ttk.Frame(self); bf.pack(pady=6)
        self._btn = ttk.Button(bf, text="开始转换", command=self._start)
        self._btn.pack()

        # ── 日志 ──────────────────────────────────────────────
        f4 = ttk.LabelFrame(self, text="日志")
        f4.pack(fill="both", expand=True, padx=10, pady=(4, 10))

        self._log = tk.Text(f4, height=12, state="disabled",
                            font=("Courier", 9))
        sb = ttk.Scrollbar(f4, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True, padx=4, pady=4)

    # ── 文件浏览 ──────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="选择整机BOM配置表",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xls"),
                       ("所有文件",   "*.*")],
        )
        if path:
            self._in_file.set(path)
            self._load()

    # ── 加载文件 & 自动识别列 ────────────────────────────────
    def _load(self):
        path = self._in_file.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "请先选择有效的 Excel 文件")
            return
        try:
            wb = open_workbook_compat(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            messagebox.showerror("加载失败", str(e))
            return

        self._sheet_cb["values"] = sheets
        if sheets and self._sheet.get() not in sheets:
            self._sheet.set(sheets[0])
        self._auto_detect()

    def _on_hdr_change(self, *_):
        """表头行输入变化时防抖触发识别（延迟 800ms，避免逐字触发）。"""
        if self._hdr_debounce:
            self.after_cancel(self._hdr_debounce)
        self._hdr_debounce = self.after(800, self._auto_detect)

    def _auto_detect(self):
        path  = self._in_file.get().strip()
        sheet = self._sheet.get()
        try:
            hdr_row = int(self._hdr_row.get())
        except ValueError:
            hdr_row = 4
        try:
            wb   = open_workbook_compat(path, read_only=True, data_only=True)
            ws   = wb[sheet]
            found, raw_headers = _detect_columns(ws, hdr_row)
            wb.close()
        except Exception as e:
            self._dlog(f"识别失败: {e}")
            return

        label_map = {"hq_pn": "HQ PN", "supply_type": "主二供", "qty": "用量"}
        var_map   = {"hq_pn": self._col_hqpn,
                     "supply_type": self._col_stype,
                     "qty": self._col_qty}
        parts = []
        for field, ci in found.items():
            if field not in var_map:
                continue
            col_l = _int_to_col(ci)
            var_map[field].set(col_l)
            parts.append(f"{label_map[field]} → {col_l}({ci})")

        if parts:
            self._dlog("自动识别：" + "  |  ".join(parts))
        else:
            # 显示实际读到的列名，帮助排查
            preview = "  ".join(raw_headers[:15]) if raw_headers else "（该行为空）"
            self._dlog(f"未识别到目标列。第{hdr_row}行实际内容：\n{preview}")

    def _dlog(self, text):
        self._detect_log.configure(state="normal")
        self._detect_log.delete("1.0", "end")
        self._detect_log.insert("end", text)
        self._detect_log.configure(state="disabled")

    # ── 日志写入（线程安全）──────────────────────────────────
    def _wlog(self, text):
        def _do():
            self._log.configure(state="normal")
            self._log.insert("end", text + "\n")
            self._log.see("end")
            self._log.configure(state="disabled")
        self.after(0, _do)

    # ── 开始转换 ──────────────────────────────────────────────
    def _start(self):
        path = self._in_file.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "请先选择源 BOM 文件")
            return
        sheet = self._sheet.get()
        if not sheet:
            messagebox.showwarning("提示", "请选择 Sheet")
            return
        try:
            hdr_row = int(self._hdr_row.get())
        except ValueError:
            messagebox.showwarning("提示", "表头行请填数字")
            return

        col_hqpn  = _col_to_int(self._col_hqpn.get())
        col_stype = _col_to_int(self._col_stype.get())
        col_qty   = _col_to_int(self._col_qty.get())

        missing = []
        if not col_hqpn:  missing.append("HQ PN 列")
        if not col_stype: missing.append("主二供列")
        if not col_qty:   missing.append("用量列")
        if missing:
            messagebox.showwarning("提示",
                f"以下列未识别，请手动填写列字母：\n{'、'.join(missing)}")
            return

        out_file = _unique_path(
            os.path.join(os.path.dirname(os.path.abspath(path)),
                         "PLM导入BOM.xlsx")
        )

        self._btn.config(state="disabled")
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")
        self._wlog(f"输入文件：{path}")
        self._wlog(f"输出文件：{out_file}")
        self._wlog("转换中，请稍候...")

        def _run():
            try:
                total, skipped = do_convert(
                    path, sheet, hdr_row,
                    col_hqpn, col_stype, col_qty,
                    self._proj_name.get().strip(),
                    out_file,
                    self._wlog,
                )
                self._wlog(
                    f"\n完成！共写入 {total} 行，"
                    f"跳过 {skipped} 行（用量为空不导入）。"
                )
                self._wlog(f"输出文件：{out_file}")
                self.after(0, lambda: messagebox.showinfo(
                    "转换完成",
                    f"完成！共写入 {total} 行，跳过 {skipped} 行（用量为空不导入）。\n\n{out_file}"
                ))
                # 打开输出文件所在文件夹
                def _open_dir():
                    d = os.path.dirname(out_file)
                    if sys.platform == "win32":
                        os.startfile(d)
                    elif sys.platform == "darwin":
                        subprocess.Popen(["open", d])
                    else:
                        subprocess.Popen(["xdg-open", d])
                self.after(0, _open_dir)
            except Exception as e:
                import traceback
                self._wlog(f"\n错误：{e}")
                self._wlog(traceback.format_exc())
                self.after(0, lambda: messagebox.showerror("错误", str(e)))
            finally:
                self.after(0, lambda: self._btn.config(state="normal"))

        threading.Thread(target=_run, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
