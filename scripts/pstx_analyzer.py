# -*- coding: utf-8 -*-
"""
PSTX 原理图分析工具 v1.3
解析 Cadence Packager-XL 导出的 pstxprt.dat / pstxnet.dat

功能：BOM 管理 / 网络拓扑 / DRC / 电容降额 / 电阻检查 / 元件查询 / Excel 导出 / 层次化页码映射

依赖：pip install openpyxl
运行：python pstx_analyzer.py
"""

import sys
import subprocess

try:
    import openpyxl
except ImportError:
    print("未检测到 openpyxl，正在自动安装...")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

import os
import re
import threading
import tkinter as tk
from collections import Counter, defaultdict
from pathlib import Path
from tkinter import ttk, filedialog, messagebox, scrolledtext
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════
# 零、页码解析辅助（内联自 pstx_page_logic.py）
# ══════════════════════════════════════════════════════════

_PAGE_TOKEN_RE = re.compile(
    r'(?<![A-Z0-9])PAGE(?:[_\-/ ]*)(\d+)([A-Z]?)(?![A-Z0-9])',
    re.IGNORECASE,
)
_PATH_SEGMENT_RE = re.compile(
    r'^(?P<head>.+?)\((?P<view>[^)]+)\)\s*:\s*(?P<tail>.+)$',
    re.IGNORECASE,
)
_SECTION_PATH_RE = re.compile(
    r'(?ims)^\s*SECTION_NUMBER\s+(?P<num>\d+)\s*\n\s*\'(?P<path>[^\']+)\'\s*:',
)
_PAGE_NUMBER_LINE_RE = re.compile(
    r"""^\s*["']?PAGE_NUMBER["']?\s*(?:=|:)\s*["']?(?P<value>[A-Z0-9_./ -]+?)["']?\s*[;,]?\s*$""",
    re.IGNORECASE,
)


_MODULE_ORDER_LINE_RE = re.compile(
    r'^\s*(?P<path>@\S+)\s+(?P<unk1>\d+)\s+(?P<unk2>\d+)\s+(?P<start>\d+)\s+(?P<count>\d+)\s+(?P<flag>\d+)\s*$',
)


def _natural_sort_key(value: str):
    parts = re.split(r'(\d+)', str(value or '').upper())
    return [int(p) if p.isdigit() else p for p in parts]


def _normalize_page_token(match: re.Match) -> str:
    num = str(int(match.group(1)))
    suffix = match.group(2).upper()
    return f'PAGE{num}{suffix}'


def _normalize_page_label(page_label: str) -> str:
    value = str(page_label or '').strip().upper()
    if not value:
        return ''
    matches = list(_PAGE_TOKEN_RE.finditer(value))
    if not matches:
        return value
    normalized = [_normalize_page_token(m) for m in matches]
    return normalized[0] if len(normalized) == 1 else ' / '.join(normalized)


def _coerce_page_number(value: str) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    if not text.upper().startswith('PAGE'):
        text = f'PAGE{text}'
    return _normalize_page_label(text)


def _clean_page_csv_value(value: str) -> str:
    text = str(value or '').strip().rstrip(';,').strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {'"', "'"}:
        text = text[1:-1].strip()
    return text


def _iter_text_with_fallback_encodings(file_path) -> List[str]:
    """支持 utf-16 等多种编码读取文件，依次尝试，返回去重后的文本列表"""
    try:
        raw_bytes = Path(file_path).read_bytes()
    except OSError:
        return []
    texts: List[str] = []
    seen: set = set()
    for enc in ['utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be', 'utf-8', 'gb18030', 'cp936']:
        try:
            text = raw_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
        if text and text not in seen:
            seen.add(text)
            texts.append(text)
    fallback = raw_bytes.decode('utf-8', errors='replace')
    if fallback and fallback not in seen:
        texts.append(fallback)
    return texts


def _extract_page_number_from_text(text: str) -> str:
    if not text:
        return ''
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = _PAGE_NUMBER_LINE_RE.match(line)
        if match:
            pn = _coerce_page_number(_clean_page_csv_value(match.group('value')))
            if pn:
                return pn
    rows = []
    for raw_line in text.splitlines():
        parts = [_clean_page_csv_value(p) for p in raw_line.split(',')]
        rows.append(parts)
        for idx, part in enumerate(parts):
            if part.upper() != 'PAGE_NUMBER':
                continue
            for follower in parts[idx + 1:]:
                pn = _coerce_page_number(_clean_page_csv_value(follower))
                if pn:
                    return pn
    for row_idx, parts in enumerate(rows):
        header_idxs = [i for i, p in enumerate(parts) if p.upper() == 'PAGE_NUMBER']
        for col_idx in header_idxs:
            for data_row in rows[row_idx + 1:]:
                if col_idx >= len(data_row):
                    continue
                pn = _coerce_page_number(_clean_page_csv_value(data_row[col_idx]))
                if pn:
                    return pn
    for regex in [
        re.compile(r'(?im)["\']?PAGE_NUMBER["\']?\s*[,=:\t;]\s*["\']?([A-Z0-9_./ -]+?)["\']?\s*[;,]?(?:$|\r|\n)'),
        re.compile(r'(?im)^["\']?PAGE_NUMBER["\']?\s*[,;\t]\s*["\']?([A-Z0-9_./ -]+?)["\']?\s*[;,]?(?:$|\r|\n)'),
    ]:
        m = regex.search(text)
        if m:
            pn = _coerce_page_number(_clean_page_csv_value(m.group(1)))
            if pn:
                return pn
    return ''


def _read_page_number_from_csv(csv_path) -> str:
    for text in _iter_text_with_fallback_encodings(csv_path):
        pn = _extract_page_number_from_text(text)
        if pn:
            return pn
    return ''


def _iter_page_csv_paths(project_root: Path) -> List[Path]:
    candidates: Dict[str, Path] = {}
    direct_sch = project_root / 'sch_1'
    if direct_sch.is_dir():
        for csv_path in direct_sch.iterdir():
            if (csv_path.is_file() and csv_path.suffix.lower() == '.csv'
                    and csv_path.stem.lower().startswith('page')):
                candidates[str(csv_path.resolve())] = csv_path
    for csv_path in project_root.rglob('page*.csv'):
        if csv_path.is_file() and csv_path.parent.name.lower() == 'sch_1':
            candidates[str(csv_path.resolve())] = csv_path
    return sorted(candidates.values(), key=lambda p: _natural_sort_key(str(p)))


def _build_page_csv_index(project_root: str) -> Dict:
    root = Path(project_root).expanduser()
    index = {
        'root': str(root), 'by_logical_page': defaultdict(list),
        'warnings': [], 'count': 0, 'scanned': 0, 'skipped_paths': [],
    }
    if not project_root or not root.exists():
        if project_root:
            index['warnings'].append(f'项目根路径不存在：{root}')
        return index
    csv_paths = _iter_page_csv_paths(root)
    index['scanned'] = len(csv_paths)
    for csv_path in csv_paths:
        real_page = _coerce_page_number(csv_path.stem)
        if not real_page:
            index['skipped_paths'].append(str(csv_path))
            continue
        logical_page = _read_page_number_from_csv(csv_path)
        if not logical_page:
            index['skipped_paths'].append(str(csv_path))
            continue
        index['by_logical_page'][logical_page].append({
            'path': str(csv_path), 'resolved_page': real_page,
            'is_root_sch1': csv_path.parent == (root / 'sch_1'),
        })
        index['count'] += 1
    if index['scanned'] == 0:
        index['warnings'].append(f'未在项目根路径下找到任何 sch_1/page*.csv：{root}')
    elif index['count'] == 0:
        samples = '；'.join(index['skipped_paths'][:3])
        index['warnings'].append(f'已扫描 {index["scanned"]} 个 page*.csv，但没有读出任何 PAGE_NUMBER' +
                                  (f'；例如：{samples}' if samples else ''))
    return index


def _parse_page_map_line(raw_line: str) -> Optional[Dict]:
    parts = re.split(r'\s+', str(raw_line or '').strip(), maxsplit=2)
    if len(parts) < 3:
        return None
    lp = _coerce_page_number(parts[0])
    rp = _coerce_page_number(parts[1])
    if not lp or not rp:
        return None
    return {'logical_page': lp, 'real_page': rp, 'page_name': parts[2].strip()}


def _build_page_map_index(project_root: str) -> Dict:
    root = Path(project_root).expanduser()
    index = {'root': str(root), 'by_logical_page': defaultdict(list), 'warnings': [], 'count': 0}
    if not project_root or not root.exists():
        return index
    file_paths = []
    direct = root / 'sch_1' / 'page.map'
    if direct.is_file():
        file_paths.append(direct)
    for path in root.rglob('page.map'):
        if path.is_file() and path not in file_paths:
            file_paths.append(path)
    for path in file_paths:
        matched = False
        for text in _iter_text_with_fallback_encodings(path):
            for raw_line in text.splitlines():
                parsed = _parse_page_map_line(raw_line)
                if not parsed:
                    continue
                lp = parsed['logical_page']
                rp = parsed['real_page']
                index['by_logical_page'][lp].append({
                    'path': str(path), 'logical_page': lp, 'resolved_page': rp,
                    'page_name': parsed['page_name'],
                    'is_root_sch1': path.parent == (root / 'sch_1'),
                })
                index['count'] += 1
                matched = True
            if matched:
                break
    return index


def _resolve_unique_real_page(index: Optional[Dict], logical_page: str) -> Tuple[str, str]:
    if not index or not logical_page:
        return '', 'none'
    entries = index.get('by_logical_page', {}).get(logical_page, [])
    if not entries:
        return '', 'none'
    real_pages = sorted({e.get('resolved_page', '') for e in entries if e.get('resolved_page')},
                        key=_natural_sort_key)
    if len(real_pages) != 1:
        return '', 'ambiguous'
    return real_pages[0], 'unique'


def _extract_path_segments(path_text: str) -> List[Dict]:
    raw = str(path_text or '').strip()
    if not raw:
        return []
    segments = []
    for chunk in [s.strip() for s in raw.split('@') if s.strip()]:
        match = _PATH_SEGMENT_RE.match(chunk)
        if not match:
            continue
        head = match.group('head').strip()
        view = match.group('view').strip()
        tail = match.group('tail').strip()
        pm = _PAGE_TOKEN_RE.search(tail)
        if not pm:
            continue
        lib, _, cell = head.rpartition('.')
        segments.append({
            'raw': chunk, 'head': head, 'lib': lib.strip(),
            'cell': (cell or head).strip(), 'view': view,
            'raw_page': _normalize_page_token(pm), 'tail': tail,
        })
    return segments


def _extract_section_paths(block_text: str) -> List[Dict]:
    entries = []
    for m in _SECTION_PATH_RE.finditer(str(block_text or '')):
        entries.append({'section_number': m.group('num'), 'path': m.group('path').strip()})
    return entries


def _select_component_page_source(block_text: str, attrs: Dict) -> Tuple[str, str]:
    section_paths = _extract_section_paths(block_text)
    if section_paths:
        preferred = next((e for e in section_paths if e.get('section_number') == '1'), section_paths[0])
        path_text = preferred.get('path', '').strip()
        if path_text:
            return path_text, 'section_path'
    c_path = str(attrs.get('C_PATH', '')).strip()
    if c_path:
        return c_path, 'c_path'
    drawing = str(attrs.get('DRAWING', '')).strip()
    if drawing:
        return drawing, 'drawing'
    return '', 'none'


def _extract_top_level_logical_page(path_text: str) -> str:
    segments = _extract_path_segments(path_text)
    for seg in segments:
        if seg.get('view', '').upper() == 'SCH_1':
            return seg.get('raw_page', '')
    if segments:
        return segments[0].get('raw_page', '')
    return _normalize_page_label(path_text).split(' / ')[0] if path_text else ''


def _extract_submodule_page(path_text: str) -> str:
    sch_segs = [s for s in _extract_path_segments(path_text) if s.get('view', '').upper() == 'SCH_1']
    return sch_segs[1].get('raw_page', '') if len(sch_segs) == 2 else ''


def _pick_top_schematic_segment(path_text: str, page_map_index: Optional[Dict],
                                 page_csv_index: Optional[Dict]) -> Dict:
    sch_segs = [s for s in _extract_path_segments(path_text) if s.get('view', '').upper() == 'SCH_1']
    if not sch_segs:
        return {}
    # 优先找与项目根目录同名的模块
    root_name = ''
    for idx in [page_map_index, page_csv_index]:
        if idx and idx.get('root'):
            try:
                root_name = Path(str(idx['root'])).name.upper()
                break
            except Exception:
                pass
    if root_name:
        exact = [s for s in sch_segs if s.get('cell', '').upper() == root_name]
        if exact:
            return exact[0]
    # 优先找在 root sch_1 中有页码的
    root_pages = set()
    for idx in [page_map_index, page_csv_index]:
        if idx:
            for lp, entries in idx.get('by_logical_page', {}).items():
                if any(e.get('is_root_sch1') for e in entries):
                    root_pages.add(lp)
    if root_pages:
        root_matches = [s for s in sch_segs if s.get('raw_page', '') in root_pages]
        if root_matches:
            return root_matches[0]
    return sch_segs[0]


def _resolve_component_page(comp: Dict, page_map_index: Optional[Dict],
                             page_csv_index: Optional[Dict]) -> str:
    logical_path = str(comp.get('page_path_raw', '') or comp.get('drawing', ''))
    top_seg = _pick_top_schematic_segment(logical_path, page_map_index, page_csv_index)
    top_logical = top_seg.get('raw_page', '') or _extract_top_level_logical_page(logical_path)
    if not top_logical:
        return ''
    pm_real, _ = _resolve_unique_real_page(page_map_index, top_logical)
    csv_real, _ = _resolve_unique_real_page(page_csv_index, top_logical)
    return pm_real or csv_real or top_logical


def _iter_named_files(project_root: Path, filename: str) -> List[Path]:
    """递归查找指定文件名，优先 sch_1 目录下的"""
    candidates: Dict[str, Path] = {}
    direct = project_root / 'sch_1' / filename
    if direct.is_file():
        candidates[str(direct.resolve())] = direct
    for found in project_root.rglob(filename):
        if found.is_file():
            candidates[str(found.resolve())] = found
    return sorted(candidates.values(), key=lambda p: _natural_sort_key(str(p)))


def _build_module_order_index(project_root: str) -> Dict:
    """解析 module_order.dat / module_order 文件，建立子模块→主模块页映射"""
    root = Path(project_root).expanduser()
    index = {
        'root': str(root), 'by_key': defaultdict(list),
        'warnings': [], 'count': 0, 'files': [],
    }
    if not project_root or not root.exists():
        return index
    file_paths = _iter_named_files(root, 'module_order.dat')
    if not file_paths:
        file_paths = _iter_named_files(root, 'module_order')
    index['files'] = [str(p) for p in file_paths]
    seen_entries = set()
    for path in file_paths:
        matched = False
        for text in _iter_text_with_fallback_encodings(path):
            in_section = False
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                upper_line = line.upper()
                if upper_line == 'START_MODULEORDER':
                    in_section = True
                    continue
                if upper_line == 'END_MODULEORDER':
                    in_section = False
                    continue
                if not in_section or not line.startswith('@'):
                    continue
                m = _MODULE_ORDER_LINE_RE.match(line)
                if not m:
                    continue
                key = str(m.group('path') or '').strip().upper()
                entry = {
                    'path': m.group('path'),
                    'path_key': key,
                    'start_real_page': _coerce_page_number(m.group('start')),
                    'page_count': int(m.group('count')),
                    'flag': int(m.group('flag')),
                }
                sig = (key, entry['start_real_page'], entry['page_count'], entry['flag'])
                if sig in seen_entries:
                    matched = True
                    continue
                seen_entries.add(sig)
                index['by_key'][key].append(entry)
                index['count'] += 1
                matched = True
            if matched:
                break
    if file_paths and index['count'] == 0:
        index['warnings'].append(f'已扫描 {len(file_paths)} 个 module_order，但没有读出有效映射')
    return index


def resolve_component_pages(components: Dict, project_root: str = '') -> List[str]:
    """用 page.map / page*.csv / module_order 把逻辑页转换为真实页，返回警告列表"""
    if not project_root:
        for comp in components.values():
            if not comp.get('page'):
                lp = _extract_top_level_logical_page(
                    str(comp.get('page_path_raw', '') or comp.get('drawing', '')))
                comp['page'] = lp
                comp['page_logical'] = lp
        return []
    pm_index = _build_page_map_index(project_root)
    csv_index = _build_page_csv_index(project_root)
    mo_index = _build_module_order_index(project_root)
    warnings = list(pm_index.get('warnings', [])) + list(csv_index.get('warnings', [])) + list(mo_index.get('warnings', []))
    for comp in components.values():
        if comp.get('page_real'):
            continue
        logical_path = str(comp.get('page_path_raw', '') or comp.get('drawing', ''))
        top_seg = _pick_top_schematic_segment(logical_path, pm_index, csv_index)
        top_logical = top_seg.get('raw_page', '') or _extract_top_level_logical_page(logical_path)
        pm_real, _ = _resolve_unique_real_page(pm_index, top_logical)
        csv_real, _ = _resolve_unique_real_page(csv_index, top_logical)
        real_page = pm_real or csv_real or ''
        comp['page'] = real_page or top_logical
        comp['page_logical'] = top_logical
        comp['page_real'] = real_page
        # module_order 子模块页映射（仅对层次化设计中的子模块有效）
        if mo_index and mo_index.get('by_key'):
            submodule_page = _extract_submodule_page(logical_path)
            if submodule_page:
                sch_segs = [s for s in _extract_path_segments(logical_path) if s.get('view', '').upper() == 'SCH_1']
                if len(sch_segs) >= 2:
                    # 构造 module_order key（取子模块路径的 @ 链）
                    for child_idx in range(len(sch_segs) - 1, 0, -1):
                        parent_chain = '@'.join(s['raw'] for s in sch_segs[:child_idx])
                        child = sch_segs[child_idx]
                        raw_key = f'@{parent_chain}@{child["head"]}({child["view"]})'
                        key = raw_key.strip().upper()
                        entries = mo_index.get('by_key', {}).get(key, [])
                        if len(entries) == 1:
                            entry = entries[0]
                            try:
                                start_match = re.search(r'(\d+)', str(entry.get('start_real_page', '')))
                                local_match = re.search(r'(\d+)', submodule_page)
                                if start_match and local_match:
                                    start = int(start_match.group(1))
                                    local = int(local_match.group(1))
                                    mapped = _coerce_page_number(str(start + local - 1))
                                    if mapped:
                                        comp['page_submodule_mapped'] = mapped
                                        # 子模块映射结果覆盖 page（更精确）
                                        comp['page'] = mapped
                            except Exception:
                                pass
                            break
    return warnings




# ══════════════════════════════════════════════════════════
# 一、PST 文件解析
# ══════════════════════════════════════════════════════════

def _join_continuations(text: str) -> str:
    normalized = str(text or '').replace('\r\n', '\n').replace('\r', '\n')
    lines = normalized.split('\n')
    result, buf = [], ''
    for line in lines:
        stripped = line.rstrip()
        if stripped.endswith('~'):
            buf += stripped[:-1]
        else:
            buf += line
            result.append(buf)
            buf = ''
    if buf:
        result.append(buf)
    return '\n'.join(result)


def _split_named_blocks(text: str, marker: str) -> List[str]:
    """用编译后的正则按 marker 分块，比 re.split 更健壮"""
    return re.split(rf'(?:^|\n){re.escape(marker)}\n', text)[1:]


def _extract_attrs(text: str) -> Dict[str, str]:
    attrs = {}
    for m in re.finditer(r"\b([A-Z][A-Z0-9_]*)\s*=\s*'([^']*)'", text):
        key, val = m.group(1), m.group(2)
        if key not in attrs:
            attrs[key] = val
    return attrs


def _get_comp_type(refdes: str, part_name: str) -> str:
    pn = part_name.lower()
    type_rules = [
        (['cap_pol'],                           'CAP_POL'),
        (['cap_hdl', 'cap_'],                   'CAP'),
        (['res_hdl', 'res_'],                   'RES'),
        (['ind_hdl', 'ind_', 'ferrite', 'fer_hdl', 'fb_hdl'], 'IND'),
        (['osc_', 'crystal', 'xtal'],           'XTAL'),
        (['conn_', 'connector'],                'CONN'),
        (['led_'],                              'LED'),
        (['diode', '_d_hdl'],                   'DIODE'),
        (['mosfet', 'mos_', 'nmos', 'pmos', 'nfet', 'pfet'], 'FET'),
        (['bjt', 'transistor', 'npn', 'pnp'],  'BJT'),
        (['fuse'],                              'FUSE'),
        (['sw_hdl', 'switch'],                  'SWITCH'),
        (['testpoint', 'test_point', 'tp_hdl'], 'TESTPOINT'),
        (['transformer', 'xfmr'],              'TRANSFORMER'),
    ]
    for keywords, ctype in type_rules:
        if any(k in pn for k in keywords):
            return ctype
    prefix = (re.match(r'[A-Za-z]+', refdes) or re.match(r'', '')).group(0).upper()
    prefix_map = {
        'C': 'CAP', 'PC': 'CAP', 'R': 'RES', 'L': 'IND', 'FB': 'IND',
        'U': 'IC', 'J': 'CONN', 'P': 'CONN', 'CN': 'CONN', 'Q': 'FET',
        'D': 'DIODE', 'LED': 'LED', 'Y': 'XTAL', 'F': 'FUSE',
        'SW': 'SWITCH', 'TP': 'TESTPOINT', 'T': 'TRANSFORMER',
    }
    return prefix_map.get(prefix, 'IC')


def _infer_project_root_from_data_paths(*paths: str) -> str:
    """从 pstxprt/pstxnet 文件路径自动推断项目根目录"""
    raw_paths = [str(p or '').strip() for p in paths if str(p or '').strip()]
    if not raw_paths:
        return ''
    candidates = []
    for path_text in raw_paths:
        try:
            p = Path(path_text).expanduser().resolve()
            candidates.append(p)
        except OSError:
            continue
    if not candidates:
        return ''
    for p in candidates:
        if p.parent.name.lower() == 'packaged':
            return str(p.parent.parent)
    try:
        common = Path(os.path.commonpath([str(p.parent) for p in candidates]))
    except ValueError:
        common = candidates[0].parent
    if common.name.lower() == 'packaged':
        return str(common.parent)
    return str(common)


def parse_pstxprt(content: str) -> Dict[str, dict]:
    text = _join_continuations(content)
    components = {}
    for block in _split_named_blocks(text, 'PART_NAME'):
        m = re.match(r"(\S+)\s+'([^']+)'", block.split('\n')[0].strip())
        if not m:
            continue
        refdes, part_name = m.group(1), m.group(2)
        attrs = _extract_attrs(block)
        page_path_raw, page_path_source = _select_component_page_source(block, attrs)
        logical_page = _extract_top_level_logical_page(page_path_raw or attrs.get('DRAWING', ''))
        # PHYS_PAGE 是工程师印刷原理图上看到的实际页码。
        # 但层次化设计中，深度≥2的子模块内元件 PHYS_PAGE 是子模块内页码，不是主图页码。
        # 只有直接放置在顶层（路径中仅1个 SCH_1 层级）时，PHYS_PAGE 才是主图物理页码。
        phys_raw = attrs.get('PHYS_PAGE', '').strip()
        path_for_depth = page_path_raw or attrs.get('DRAWING', '')
        sch1_depth = len(re.findall(r'\(sch_1\)', path_for_depth, re.IGNORECASE))
        phys_page = f'PAGE{phys_raw}' if (phys_raw.isdigit() and sch1_depth <= 1) else ''
        components[refdes] = {
            'refdes':           refdes,
            'part_name':        part_name,
            'hq_code':          attrs.get('HQ_CODE', ''),
            'value':            attrs.get('VALUE', ''),
            'package':          attrs.get('PACKAGE', ''),
            'material':         attrs.get('MATERIAL', ''),
            'tolerance':        attrs.get('TOLERANCE', ''),
            'voltage':          attrs.get('VOLTAGE', ''),
            'current':          attrs.get('CURRENT', ''),
            'power':            attrs.get('POWER', ''),
            'bom_option':       attrs.get('BOM_OPTION', ''),
            'bom_cost':         attrs.get('BOM_COST', ''),
            'room':             attrs.get('ROOM', ''),
            'drawing':          attrs.get('DRAWING', ''),
            'page_path_raw':    page_path_raw,
            'page_path_source': page_path_source,
            'page':             phys_page or logical_page,
            'page_logical':     logical_page,
            'page_real':        phys_page,
            'comp_type':        _get_comp_type(refdes, part_name),
        }
    return components


def parse_pstxnet(content: str) -> Dict[str, List[dict]]:
    text = _join_continuations(content)
    nets = {}
    node_re     = re.compile(r'NODE_NAME\s+(\S+)\s+(\S+)')
    pin_name_re = re.compile(r"'([^']+)'\s*:")
    for block in _split_named_blocks(text, 'NET_NAME'):
        m = re.search(r"'([^']+)'", block)
        if not m:
            continue
        net_name = m.group(1)
        nodes = []
        matches = list(node_re.finditer(block))
        for idx, nm in enumerate(matches):
            next_start = matches[idx + 1].start() if idx + 1 < len(matches) else len(block)
            after = block[nm.end():next_start]
            pn_match = pin_name_re.search(after)
            nodes.append({
                'refdes':   nm.group(1),
                'pin':      nm.group(2),
                'pin_name': pn_match.group(1) if pn_match else nm.group(2),
            })
        if nodes:
            nets[net_name] = nodes
    return nets


def parse_all(prt_content: str, net_content: str):
    components = parse_pstxprt(prt_content)
    nets       = parse_pstxnet(net_content)
    comp_nets: Dict[str, Dict[str, str]] = {}
    for net_name, nodes in nets.items():
        for node in nodes:
            rd = node['refdes']
            if rd not in comp_nets:
                comp_nets[rd] = {}
            comp_nets[rd][node['pin']] = net_name
    for refdes, comp in components.items():
        comp['nets'] = comp_nets.get(refdes, {})
    return components, nets, comp_nets


def _is_depop_option(bom_option: str) -> bool:
    return str(bom_option or '').strip().upper() in {'DEPOP', 'DNP'}


def _display_bom_option(bom_option: str) -> str:
    v = str(bom_option or '').strip().upper()
    return v or '默认'


# ══════════════════════════════════════════════════════════
# 二、BOM 分析
# ══════════════════════════════════════════════════════════

COMP_TYPE_CN = {
    'CAP': '电容', 'CAP_POL': '电解/钽电容', 'RES': '电阻',
    'IND': '电感/磁珠', 'IC': 'IC 芯片', 'CONN': '连接器',
    'DIODE': '二极管', 'LED': 'LED', 'FET': 'MOS/FET',
    'BJT': '三极管', 'XTAL': '晶振', 'FUSE': '保险丝',
    'SWITCH': '开关', 'TESTPOINT': '测试点', 'TRANSFORMER': '变压器',
}
_TYPE_ORDER = list(COMP_TYPE_CN.keys())


def build_bom(components: Dict):
    detail_normal, detail_depop = [], []
    for comp in components.values():
        ctype = comp.get('comp_type', '')
        row = {
            '位号':          comp['refdes'],
            '料号':          comp.get('hq_code', ''),
            '描述':          comp.get('part_name', ''),
            '值':            comp.get('value', ''),
            '封装':          comp.get('package', ''),
            '耐压/额定电压': comp.get('voltage', ''),
            '额定功率':      comp.get('power', ''),
            '精度':          comp.get('tolerance', ''),
            '材质':          comp.get('material', ''),
            '类型':          COMP_TYPE_CN.get(ctype, ctype),
            '_ctype':        ctype,
            '页面':          comp.get('page_submodule_mapped', '') or comp.get('page', ''),
            'ROOM':          comp.get('room', ''),
        }
        (detail_depop if _is_depop_option(comp.get('bom_option', '')) else detail_normal).append(row)

    def _merge(detail):
        if not detail:
            return []
        groups = {}
        for row in detail:
            key = row['料号'] or row['描述']
            if key not in groups:
                groups[key] = {
                    '料号': row['料号'], '位号列表': [], '数量': 0,
                    '描述': row['描述'], '值': row['值'], '封装': row['封装'],
                    '耐压': row['耐压/额定电压'], '额定功率': row['额定功率'],
                    '精度': row['精度'], '材质': row['材质'],
                    '类型': row['类型'], '_ctype': row['_ctype'],
                }
            groups[key]['位号列表'].append(row['位号'])
            groups[key]['数量'] += 1
        merged = list(groups.values())
        merged.sort(key=lambda r: (
            _TYPE_ORDER.index(r['_ctype']) if r['_ctype'] in _TYPE_ORDER else 99, r['料号']))
        for i, r in enumerate(merged, 1):
            r['序号'] = i
            r['位号列表'] = ', '.join(sorted(r['位号列表'], key=_natural_sort_key))
            del r['_ctype']
        return merged

    def _clean(rows):
        return [{k: v for k, v in r.items() if k != '_ctype'} for r in rows]

    return _clean(detail_normal), _clean(detail_depop), _merge(detail_normal), _merge(detail_depop)


# ══════════════════════════════════════════════════════════
# 三、网络分析
# ══════════════════════════════════════════════════════════

_DIFF_SUFFIX_PAIRS = [
    ('_P', '_N'), ('_DP', '_DN'), ('.P', '.N'),
    ('_TXPLUS', '_TXMINUS'), ('_RXPLUS', '_RXMINUS'),
]


def _get_diff_net_info(net_name: str, upper_name_map: Dict[str, str]) -> Optional[Dict[str, str]]:
    """获取单个网络的差分对信息（极性、配对网络）"""
    upper_name = (net_name or '').upper()
    for pos_suffix, neg_suffix in _DIFF_SUFFIX_PAIRS:
        pos_upper = pos_suffix.upper()
        neg_upper = neg_suffix.upper()
        if upper_name.endswith(pos_upper):
            partner = upper_name_map.get(upper_name[:-len(pos_upper)] + neg_upper)
            if partner:
                return {'base': net_name[:-len(pos_suffix)], 'polarity': 'P', 'partner': partner}
        elif upper_name.endswith(neg_upper):
            partner = upper_name_map.get(upper_name[:-len(neg_upper)] + pos_upper)
            if partner:
                return {'base': net_name[:-len(neg_suffix)], 'polarity': 'N', 'partner': partner}
    return None


def _collect_diff_pairs(nets: Dict) -> Dict[str, dict]:
    diff_pairs: Dict[str, dict] = {}
    upper_map = {name.upper(): name for name in nets}
    for net_name in nets:
        upper = net_name.upper()
        for pos_sfx, neg_sfx in _DIFF_SUFFIX_PAIRS:
            pu, nu = pos_sfx.upper(), neg_sfx.upper()
            if upper.endswith(pu):
                partner = upper_map.get(upper[:-len(pu)] + nu)
                if partner:
                    diff_pairs[net_name[:-len(pos_sfx)]] = {'P': net_name, 'N': partner}
                    break
            elif upper.endswith(nu):
                partner = upper_map.get(upper[:-len(nu)] + pu)
                base = net_name[:-len(neg_sfx)]
                if partner and base not in diff_pairs:
                    diff_pairs[base] = {'P': partner, 'N': net_name}
                    break
    return diff_pairs


def analyze_networks(nets: Dict, components: Dict) -> dict:
    single_node = {k: v for k, v in nets.items() if len(v) == 1}
    gnd_nets    = {k: v for k, v in nets.items() if _net_is_gnd(k)}
    power_nets  = {k: v for k, v in nets.items() if _net_is_power(k) and k not in gnd_nets}
    diff_pairs  = _collect_diff_pairs(nets)
    page_counter: Counter = Counter()
    for comp in components.values():
        page_counter[comp.get('page', '') or 'UNKNOWN'] += 1
    return {
        'total': len(nets), 'single_node': single_node,
        'gnd_nets': gnd_nets, 'power_nets': power_nets,
        'diff_pairs': diff_pairs, 'page_counter': page_counter,
    }


# ══════════════════════════════════════════════════════════
# 四、DRC 设计检查
# ══════════════════════════════════════════════════════════

_VALID_BOM_OPTIONS = {'', 'DEPOP', 'OPTION', 'MAIN_PLD', 'MAIN', 'ALT', 'DNP'}
_FUZZY_KEYWORDS    = sorted(opt for opt in _VALID_BOM_OPTIONS if opt)


def _edit_distance(a: str, b: str) -> int:
    a, b = a.upper(), b.upper()
    if a == b: return 0
    if not a:  return len(b)
    if not b:  return len(a)
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        prev = dp[:]
        dp[0] = i + 1
        for j, cb in enumerate(b):
            dp[j+1] = min(prev[j] + (0 if ca == cb else 1), dp[j]+1, prev[j+1]+1)
    return dp[len(b)]


def check_drc(components: Dict, nets: Dict) -> dict:
    missing_hq, missing_val, missing_pkg, tbd_attrs, single_pin, unnamed = [], [], [], [], [], []
    bom_option_components = []
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype == 'TESTPOINT':
            continue
        base = {'位号': refdes, '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': comp.get('page', '')}
        if not comp.get('hq_code'):  missing_hq.append(base.copy())
        if not comp.get('value'):    missing_val.append(base.copy())
        if not comp.get('package'):  missing_pkg.append(base.copy())
        for attr in ('voltage', 'current', 'power'):
            val = comp.get(attr, '')
            if val and 'TBD' in val.upper():
                tbd_attrs.append({'位号': refdes, '属性': attr.upper(), '当前值': val,
                                  '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': comp.get('page', '')})
        # BOM_OPTION 元件清单
        bom_option = str(comp.get('bom_option', '') or '').strip().upper()
        if bom_option:
            bom_option_components.append({
                '位号': refdes,
                '类型': COMP_TYPE_CN.get(ctype, ctype),
                'BOM_OPTION值': bom_option,
                '是否DEPOP': '是' if _is_depop_option(bom_option) else '否',
                '页面': comp.get('page', ''),
            })

    for net_name, nodes in nets.items():
        if len(nodes) == 1:
            n = nodes[0]
            comp = components.get(n['refdes'], {})
            if comp.get('comp_type') != 'TESTPOINT' and not re.search(r'^UNNAMED_', net_name, re.I):
                single_pin.append({'网络名': net_name, '连接元件': n['refdes'],
                                    '引脚': n['pin_name'],
                                    '页面': comp.get('page', '')})
        if re.search(r'^UNNAMED_', net_name, re.I):
            unnamed.append({'网络名': net_name, '节点数': len(nodes)})

    # 计算每个 BOM_OPTION 值的拼写风险
    risk_per_value: Dict[str, str] = {}
    for val in set(str(comp.get('bom_option', '') or '').strip().upper() for comp in components.values()):
        if not val:
            continue
        if val in _VALID_BOM_OPTIONS:
            risk_per_value[val] = '✅ 合法'
        else:
            min_d = min(_edit_distance(val, kw) for kw in _FUZZY_KEYWORDS)
            risk_per_value[val] = '❌ 疑似拼错' if min_d <= 2 else '⚠ 未知值'

    # 将风险信息写入每个元件行
    for item in bom_option_components:
        item['拼写风险'] = risk_per_value.get(item['BOM_OPTION值'], '')

    return {
        'missing_hq_code': missing_hq, 'missing_value': missing_val,
        'missing_package': missing_pkg, 'tbd_attrs': tbd_attrs,
        'single_pin_nets': single_pin, 'unnamed_nets': unnamed,
        'bom_option_components': sorted(bom_option_components, key=lambda r: _natural_sort_key(r['位号'])),
    }



# ══════════════════════════════════════════════════════════
# 五、电容降额分析
# ══════════════════════════════════════════════════════════

# PG/OD/OC 信号网络模式：这类网络电压由外部上拉决定，不推断
_OD_SKIP_PATTERNS = re.compile(
    r'\bPG\b|PGOOD|_PG_|_PGD\b|PG_N|PWRGD|POWER_GOOD'
    r'|\bFAULT\b|_FAULT|VR_FAULT'
    r'|\bALERT\b|_ALERT|SMBALERT'
    r'|\bSDA\b|\bSCL\b'
    r'|\bOC_N\b|_OC\b'
    r'|\bPRSNT\b|\bPRESENT\b'
    r'|\bINT_N\b|\bIRQ_N\b',
    re.IGNORECASE,
)


def _split_net_tokens(net_name: str) -> List[str]:
    return [tok for tok in re.split(r'[_./-]+', (net_name or '').upper()) if tok]


def _first_net_token(net_name: str) -> str:
    tokens = _split_net_tokens(net_name)
    return tokens[0] if tokens else (net_name or '').upper()


_POWER_TOKEN_RE = re.compile(
    r'(?:VCC|VDD|VBAT|VCORE|VCCIO|PVDD|PVCC|AVDD|DVDD|VBUS)[A-Z0-9]*',
    re.IGNORECASE,
)
_GROUND_TOKEN_RE = re.compile(
    r'(?:[A-Z0-9]*GND[A-Z0-9]*|[A-Z0-9]*VSS[A-Z0-9]*|0V|0)',
    re.IGNORECASE,
)


def _token_is_power(token: str) -> bool:
    m = re.fullmatch(r'P?(\d+)V(\d*)', token.upper())
    if m:
        return True
    return bool(_POWER_TOKEN_RE.fullmatch(token))


def _token_is_ground(token: str) -> bool:
    return bool(_GROUND_TOKEN_RE.fullmatch(token))


def _net_is_power(net: str) -> bool:
    return _token_is_power(_first_net_token(net))


def _net_is_gnd(net: str) -> bool:
    return _token_is_ground(_first_net_token(net))


def _parse_voltage_from_token(token: str) -> Optional[float]:
    m = re.fullmatch(r'P?(\d+)V(\d*)', token.upper())
    if not m:
        return None
    int_part, frac_part = m.groups()
    return float(f'{int_part}.{frac_part}') if frac_part else float(int_part)


def _infer_voltage(net_name: str) -> Optional[float]:
    """从网络名首 token 推断电压（新版：基于 token 解析，不误判 PG/OD 信号）"""
    if _OD_SKIP_PATTERNS.search(net_name):
        return None
    token = _first_net_token(net_name)
    if _token_is_ground(token):
        return 0.0
    return _parse_voltage_from_token(token)


def _is_od_net(net_name: str) -> bool:
    return bool(_OD_SKIP_PATTERNS.search(net_name))


def _matches_prefix_with_boundary(name: str, prefix: str) -> bool:
    if not prefix:
        return False
    name = (name or '').upper()
    prefix = prefix.upper()
    if not name.startswith(prefix):
        return False
    return len(name) == len(prefix) or name[len(prefix)] in '_./-'


def _match_custom_voltage(net_name: str, custom_volt_map: Optional[Dict]) -> Optional[float]:
    if not custom_volt_map:
        return None
    best: Optional[Tuple[int, float]] = None
    for key, volt in custom_volt_map.items():
        prefix = str(key).strip().upper()
        if prefix and _matches_prefix_with_boundary(net_name, prefix):
            if best is None or len(prefix) > best[0]:
                best = (len(prefix), float(volt))
    return best[1] if best else None


def _collect_component_nets(nets: Dict) -> Dict[str, List[str]]:
    comp_nets: Dict[str, List[str]] = defaultdict(list)
    for net_name, nodes in nets.items():
        for node in nodes:
            comp_nets[node['refdes']].append(net_name)
    return comp_nets


def _unique_component_nets(comp_nets: Dict, refdes: str) -> List[str]:
    return list(dict.fromkeys(comp_nets.get(refdes, [])))


def _find_ac_coupling_candidates(components: Dict,
                                  comp_nets: Dict[str, List[str]],
                                  nets: Dict) -> Dict[str, dict]:
    """查找 AC 耦合电容候选：两端都接差分对同极性 net，且有镜像电容"""
    upper_map = {name.upper(): name for name in nets}

    def _get_diff_info(net_name):
        upper = net_name.upper()
        for pos_sfx, neg_sfx in _DIFF_SUFFIX_PAIRS:
            pu, nu = pos_sfx.upper(), neg_sfx.upper()
            if upper.endswith(pu):
                partner = upper_map.get(upper[:-len(pu)] + nu)
                if partner:
                    return {'polarity': 'P', 'partner': partner}
            elif upper.endswith(nu):
                partner = upper_map.get(upper[:-len(nu)] + pu)
                if partner:
                    return {'polarity': 'N', 'partner': partner}
        return None

    cap_pairs: Dict[str, Tuple[str, str]] = {}
    caps_by_pair: Dict[frozenset, List[str]] = defaultdict(list)
    for refdes, comp in components.items():
        if comp.get('comp_type') not in ('CAP', 'CAP_POL'):
            continue
        unique_nets = _unique_component_nets(comp_nets, refdes)
        if len(unique_nets) != 2:
            continue
        na, nb = unique_nets
        if _net_is_power(na) or _net_is_power(nb) or _net_is_gnd(na) or _net_is_gnd(nb):
            continue
        cap_pairs[refdes] = (na, nb)
        caps_by_pair[frozenset((na, nb))].append(refdes)

    candidates: Dict[str, dict] = {}
    for refdes, (na, nb) in cap_pairs.items():
        ia = _get_diff_info(na)
        ib = _get_diff_info(nb)
        if not ia or not ib or ia['polarity'] != ib['polarity']:
            continue
        partner_pair = frozenset((ia['partner'], ib['partner']))
        mirror_caps = sorted([c for c in caps_by_pair.get(partner_pair, []) if c != refdes],
                             key=_natural_sort_key)
        if not mirror_caps:
            continue
        candidates[refdes] = {
            'nets': (na, nb), 'mirror_nets': sorted(partner_pair, key=_natural_sort_key),
            'mirror_caps': mirror_caps, 'polarity': ia['polarity'],
        }
    return candidates


def _calc_board_max_voltage(nets: Dict, custom_volt_map: Optional[Dict]) -> float:
    """扫描全板所有网络名，推断板级最高工作电压（用于快速 pass 高额定电容）"""
    max_v = 0.0
    for net_name in nets:
        v = _match_custom_voltage(net_name, custom_volt_map)
        if v is None:
            v = _infer_voltage(net_name)
        if v is not None and v > max_v:
            max_v = v
    return max_v


def analyze_derating(components: Dict, nets: Dict,
                     pct: float = 70.0,
                     custom_volt_map: Optional[Dict[str, float]] = None,
                     include_depop: bool = False) -> List[dict]:
    """pct: 工作电压上限占额定电压的百分比（默认 70%）"""
    comp_nets = _collect_component_nets(nets)
    ac_coupling_caps = _find_ac_coupling_candidates(components, comp_nets, nets)
    board_max_v = _calc_board_max_voltage(nets, custom_volt_map)

    rows = []
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype not in ('CAP', 'CAP_POL'):
            continue
        if not include_depop and _is_depop_option(comp.get('bom_option', '')):
            continue
        connected_nets = _unique_component_nets(comp_nets, refdes)
        rated_str = comp.get('voltage', '')
        source_type = ''
        max_v, from_net, derating = None, '', None

        if not rated_str:
            status = '⚪ 无额定电压'
        else:
            m = re.match(r'([\d.]+)\s*V', rated_str.strip(), re.I)
            rated_v = float(m.group(1)) if m else None
            if rated_v is None:
                status = '⚪ 无法解析额定电压'
            elif board_max_v > 0 and rated_v * (pct / 100) >= board_max_v:
                # 额定电压 × 降额比 ≥ 板级最高电压 → 无论接哪个网络都安全，直接 pass
                threshold_v = rated_v * (pct / 100)
                status = (f'✅ 板级直通 (额定{rated_v:.0f}V×{pct:.0f}%={threshold_v:.1f}V'
                          f' ≥ 板级最高{board_max_v:.1f}V)')
                source_type = '板级直通'
            elif refdes in ac_coupling_caps:
                status = '⚪ 疑似 AC 耦合电容，不推断电压'
                source_type = 'AC 耦合候选'
            else:
                known_nets = []
                ground_present = False
                for net_name in connected_nets:
                    if _net_is_gnd(net_name):
                        ground_present = True
                    # PG/OD 信号：标记为特殊，跳过电压推断
                    if _is_od_net(net_name) and not _net_is_gnd(net_name):
                        continue
                    v = _match_custom_voltage(net_name, custom_volt_map)
                    src = 'custom_map' if v is not None else ''
                    if v is None:
                        v = _infer_voltage(net_name)
                        if v is not None:
                            src = 'net_token'
                    if v is None:
                        continue
                    if v == 0:
                        ground_present = True
                    known_nets.append((net_name, float(v), src))

                positives: Dict[float, Tuple[str, str]] = {}
                for net_name, v, src in known_nets:
                    if v > 0:
                        positives.setdefault(round(v, 6), (net_name, src))

                od_nets = [n for n in connected_nets if _is_od_net(n) and not _net_is_gnd(n)]

                if not ground_present:
                    if od_nets:
                        status = f'⚪ PG/OD信号（{od_nets[0]}），工作电压由上拉决定，请手动确认'
                    else:
                        status = '⚪ 无法判断（未连接地）'
                elif not positives:
                    if od_nets:
                        status = f'⚪ PG/OD信号（{od_nets[0]}），工作电压由上拉决定，请手动确认'
                    else:
                        status = '⚪ 无法推断工作电压'
                elif len(positives) > 1:
                    status = '⚪ 无法判断（连接多个不同电位）'
                else:
                    rounded_v, (from_net, src) = next(iter(positives.items()))
                    max_v = rounded_v
                    source_type = '自定义映射' if src == 'custom_map' else '网络首 token'
                    usage_pct = max_v / rated_v * 100
                    derating = rated_v / max_v
                    if usage_pct <= pct:
                        status = f'✅ 合格 ({usage_pct:.0f}% ≤ {pct:.0f}%)'
                    else:
                        status = f'❌ 不合格 ({usage_pct:.0f}% > {pct:.0f}%)'

        rows.append({
            '位号':            refdes,
            '值':              comp.get('value', ''),
            '封装':            comp.get('package', ''),
            '类型':            COMP_TYPE_CN.get(ctype, ctype),
            '额定电压':        rated_str,
            '推断工作电压(V)': str(max_v) if max_v is not None else '',
            '推断来源网络':    from_net,
            '推断来源类型':    source_type,
            '所有连接网络':    ', '.join(connected_nets),
            '降额比':          f'{derating:.2f}' if derating is not None else '',
            '状态':            status,
            '页面':            comp.get('page', ''),
            'DEPOP':           'Y' if _is_depop_option(comp.get('bom_option', '')) else '',
        })
    rows.sort(key=lambda r: (
        0 if r['状态'].startswith('❌') else 1 if r['状态'].startswith('✅') else 2,
        _natural_sort_key(r.get('位号', '')),
    ))
    return rows



# ══════════════════════════════════════════════════════════
# 六、电阻检查（上拉 / 下拉 / 串阻 / OD/OC / 芯片Pin总览）
# ══════════════════════════════════════════════════════════

def _parse_ohms(value_str: str) -> Optional[float]:
    if not value_str:
        return None
    s = re.sub(r'\s', '', value_str.upper())
    s = s.replace('Ω', 'R').replace('OHM', 'R').replace('OHMS', 'R')
    m = re.match(r'^([\d.]+)([KMGR]?)$', s)
    if m:
        val = float(m.group(1))
        return val * {'K': 1e3, 'M': 1e6, 'G': 1e9, 'R': 1, '': 1}.get(m.group(2), 1)
    # 支持 4K7 → 4.7k 写法
    embedded = re.match(r'^(\d+)([KMGR])(\d+)$', s)
    if embedded:
        val = float(f'{embedded.group(1)}.{embedded.group(3)}')
        return val * {'K': 1e3, 'M': 1e6, 'G': 1e9, 'R': 1}.get(embedded.group(2), 1)
    return None


_CHIP_REFDES_RE = re.compile(r'^(?:XU|PU|U)[A-Z0-9]+$', re.IGNORECASE)


def _is_chip_component(refdes: str, comp: Dict) -> bool:
    return comp.get('comp_type') == 'IC' and bool(_CHIP_REFDES_RE.match(refdes or ''))


# OD/OC 信号名关键词（用于多证据判定）
_OD_STRONG_TOKENS = {'SDA', 'SCL', 'SMBALERT', 'SMBDAT', 'SMBDATA', 'SMBCLK', 'OD', 'OC'}
_OD_WEAK_TOKENS = {'ALERT', 'FAULT', 'IRQ', 'INT', 'PGOOD', 'PWROK', 'PWRGD', 'PRSNT', 'PRESENT'}


def _od_oc_evidence_from_name(value: str, source_label: str) -> List[Tuple[str, str]]:
    tokens = set(re.findall(r'[A-Z0-9]+', (value or '').upper()))
    evidence = []
    for tok in _OD_STRONG_TOKENS:
        if tok in tokens:
            evidence.append(('strong', f'{source_label} 含 {tok}'))
    for tok in _OD_WEAK_TOKENS:
        if tok in tokens:
            evidence.append(('weak', f'{source_label} 含 {tok}'))
    return evidence


def _classify_od_oc_evidence(net_name: str, nodes: List[dict],
                              components: Dict) -> Optional[Dict]:
    evidence = []
    chip_nodes = []
    for node in nodes:
        refdes = node.get('refdes', '')
        comp = components.get(refdes, {})
        if not _is_chip_component(refdes, comp):
            continue
        chip_nodes.append(node)
        evidence.extend(_od_oc_evidence_from_name(
            node.get('pin_name', node.get('pin', '')), f'{refdes}.{node.get("pin", "")}'))
    if not chip_nodes:
        return None
    evidence.extend(_od_oc_evidence_from_name(net_name, '网络名'))
    strong = [t for lvl, t in evidence if lvl == 'strong']
    weak = [t for lvl, t in evidence if lvl == 'weak']
    if not strong and len(weak) < 2:
        return None
    unique_evidence = list(dict.fromkeys(strong + weak))
    chip_pins = ', '.join(dict.fromkeys(
        f'{n["refdes"]}.{n["pin"]}({n.get("pin_name", n["pin"])})' for n in chip_nodes))
    return {
        '芯片引脚': chip_pins,
        '判定依据': '; '.join(unique_evidence[:6]),
        'confidence': 'medium' if strong else 'low',
    }


def _classify_series_bias_ratio(series_ohms, bias_ohms):
    if series_ohms is None or bias_ohms is None or bias_ohms <= 0:
        return None, '⚪ 阻值缺失，无法计算'
    ratio = series_ohms / bias_ohms
    if bias_ohms < 1000 and ratio > 0.1:
        return ratio, '❌ 高风险'
    if ratio >= 0.33:
        return ratio, '❌ 高风险'
    if ratio > 0.1:
        return ratio, '⚠️ 关注'
    return ratio, '✅ 正常'


def _format_entry_list(entries: List[dict], key: str) -> str:
    return ', '.join(dict.fromkeys(str(e.get(key, '')) for e in entries if e.get(key, '') != ''))


def analyze_resistors(components: Dict, nets: Dict, exclude_depop: bool = True) -> dict:
    """检测上拉/下拉/串阻相关设计问题，含双向扫描和芯片Pin总览"""
    pullups:   Dict[str, list] = defaultdict(list)
    pulldowns: Dict[str, list] = defaultdict(list)
    series_list: list = []
    series_by_net: Dict[str, list] = defaultdict(list)
    indirect_pullups: Dict[str, list] = defaultdict(list)
    indirect_pulldowns: Dict[str, list] = defaultdict(list)
    node_lookup: Dict[Tuple[str, str], str] = {}

    for net_name, nodes in nets.items():
        for node in nodes:
            node_lookup[(node['refdes'], node['pin'])] = node.get('pin_name', node['pin'])

    for refdes, comp in components.items():
        if comp.get('comp_type') != 'RES':
            continue
        if exclude_depop and _is_depop_option(comp.get('bom_option', '')):
            continue
        pin_nets = list(dict.fromkeys(comp.get('nets', {}).values()))
        if len(pin_nets) != 2:
            continue
        net_a, net_b = pin_nets[0], pin_nets[1]
        ohms = _parse_ohms(comp.get('value', ''))
        val_str = comp.get('value', '')
        page = comp.get('page', '')
        bom_option = comp.get('bom_option', '')

        a_pwr, b_pwr = _net_is_power(net_a), _net_is_power(net_b)
        a_gnd, b_gnd = _net_is_gnd(net_a),   _net_is_gnd(net_b)

        entry_base = {'refdes': refdes, 'ohms': ohms, 'value': val_str, 'page': page, 'bom_option': bom_option}
        if a_pwr and not b_pwr and not b_gnd:
            pullups[net_b].append({**entry_base, 'power_net': net_a})
        elif b_pwr and not a_pwr and not a_gnd:
            pullups[net_a].append({**entry_base, 'power_net': net_b})
        elif a_gnd and not b_gnd and not b_pwr:
            pulldowns[net_b].append(entry_base.copy())
        elif b_gnd and not a_gnd and not a_pwr:
            pulldowns[net_a].append(entry_base.copy())
        elif not a_pwr and not b_pwr and not a_gnd and not b_gnd:
            series_list.append({**entry_base, 'net_a': net_a, 'net_b': net_b})
            series_by_net[net_a].append({**entry_base, 'other_net': net_b})
            series_by_net[net_b].append({**entry_base, 'other_net': net_a})

    # ── 检查1：重复上拉 ─────────────────────────────────
    dup_pullups = []
    for sig_net, pu_list in sorted(pullups.items()):
        if len(pu_list) < 2:
            continue
        group = sorted(pu_list, key=lambda r: _natural_sort_key(r.get('refdes', '')))
        dup_pullups.append({
            '信号网络':  sig_net,
            '上拉数量':  len(group),
            '位号':      ', '.join(r['refdes'] for r in group),
            '阻值':      ', '.join(r['value']  for r in group),
            '上拉电源':  ', '.join(dict.fromkeys(r['power_net'] for r in group)),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(r['bom_option']) for r in group)),
            '页面':      ', '.join(dict.fromkeys(r['page'] for r in group)),
        })

    # ── 检查2：重复下拉 ─────────────────────────────────
    dup_pulldowns = []
    for sig_net, pd_list in sorted(pulldowns.items()):
        if len(pd_list) < 2:
            continue
        group = sorted(pd_list, key=lambda r: _natural_sort_key(r.get('refdes', '')))
        dup_pulldowns.append({
            '信号网络': sig_net,
            '下拉数量': len(group),
            '位号':     ', '.join(r['refdes'] for r in group),
            '阻值':     ', '.join(r['value']  for r in group),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(r['bom_option']) for r in group)),
            '页面':     ', '.join(dict.fromkeys(r['page'] for r in group)),
        })

    # ── 检查3：串阻 + 偏置电阻分压风险（双向扫描）─────
    divider_risks = []
    seen_pairs: set = set()
    seen_indirect: set = set()
    for sr in sorted(series_list, key=lambda r: _natural_sort_key(r.get('refdes', ''))):
        for bias_net, affected_net in ((sr['net_a'], sr['net_b']), (sr['net_b'], sr['net_a'])):
            for bias_kind, bias_map, indirect_map in [
                ('上拉', pullups, indirect_pullups),
                ('下拉', pulldowns, indirect_pulldowns),
            ]:
                for bias in bias_map.get(bias_net, []):
                    # 记录间接偏置
                    ik = (affected_net, bias_kind, bias['refdes'], sr['refdes'])
                    if ik not in seen_indirect:
                        seen_indirect.add(ik)
                        indirect_map[affected_net].append({
                            **bias, 'via_refdes': sr['refdes'],
                            'via_value': sr['value'], 'via_ohms': sr['ohms'],
                            'source_net': bias_net, 'other_net': affected_net,
                        })
                    pair_key = (sr['refdes'], bias['refdes'], bias_kind, bias_net, affected_net)
                    if pair_key in seen_pairs:
                        continue
                    seen_pairs.add(pair_key)
                    ratio, status = _classify_series_bias_ratio(sr['ohms'], bias.get('ohms'))
                    ref_net = bias.get('power_net', '') if bias_kind == '上拉' else 'GND'
                    pages = ', '.join(dict.fromkeys(v for v in [sr.get('page', ''), bias.get('page', '')] if v))
                    divider_risks.append({
                        '串阻位号':    sr['refdes'],
                        '串阻值':      sr['value'],
                        '串阻网络A':   sr['net_a'],
                        '串阻网络B':   sr['net_b'],
                        '偏置类型':    bias_kind,
                        '偏置位号':    bias['refdes'],
                        '偏置值':      bias['value'],
                        '偏置所在网络': bias_net,
                        '偏置参考网络': ref_net,
                        '受影响网络':  affected_net,
                        '串/偏置比':   f'{ratio:.3f}' if ratio is not None else '',
                        '偏置 < 1k':  '是' if (bias.get('ohms') or 0) < 1000 else '否',
                        '说明':        f'{bias_kind}位于 {bias_net} 侧，通过 {sr["refdes"]} 影响 {affected_net}',
                        '状态':        status,
                        '页面':        pages,
                    })
    divider_risks.sort(key=lambda r: (
        0 if r['状态'].startswith('❌') else 1 if r['状态'].startswith('⚠') else 2,
        _natural_sort_key(r.get('串阻位号', '')),
    ))

    # ── 检查4：OD/OC 信号缺上拉（多证据判定）──────────
    od_missing = []
    for net_name in sorted(nets.keys()):
        if _net_is_power(net_name) or _net_is_gnd(net_name):
            continue
        nodes = nets[net_name]
        evidence = _classify_od_oc_evidence(net_name, nodes, components)
        if not evidence:
            continue
        if pullups.get(net_name) or indirect_pullups.get(net_name):
            continue
        od_missing.append({
            '网络名':   net_name,
            '节点数':   len(nodes),
            '连接元件': ', '.join(dict.fromkeys(n['refdes'] for n in nodes[:6])),
            '芯片引脚': evidence['芯片引脚'],
            '判定依据': evidence['判定依据'],
            '上拉状态': '未找到直接上拉/隔串阻上拉',
            '说明':     '疑似 OD/OC 信号，未找到上拉电阻',
        })

    # ── 芯片 Pin 电阻状态总览 ────────────────────────────
    chip_pin_rows = []
    for refdes, comp in sorted(components.items(), key=lambda item: _natural_sort_key(item[0])):
        if not _is_chip_component(refdes, comp):
            continue
        for pin, net_name in sorted(comp.get('nets', {}).items(), key=lambda item: _natural_sort_key(item[0])):
            pin_name = node_lookup.get((refdes, pin), pin)
            s_entries = series_by_net.get(net_name, [])
            pu_entries = pullups.get(net_name, [])
            pd_entries = pulldowns.get(net_name, [])
            ipu_entries = indirect_pullups.get(net_name, [])
            ipd_entries = indirect_pulldowns.get(net_name, [])
            chip_pin_rows.append({
                '芯片位号': refdes,
                '引脚':     pin,
                '引脚名':   pin_name,
                '网络名':   net_name,
                '有串阻':   '是' if s_entries else '否',
                '串阻数量': len(s_entries),
                '串阻位号': _format_entry_list(s_entries, 'refdes'),
                '串阻另一端': _format_entry_list(s_entries, 'other_net'),
                '有上拉':   '是' if pu_entries else '否',
                '上拉数量': len(pu_entries),
                '上拉位号': _format_entry_list(pu_entries, 'refdes'),
                '上拉电源': _format_entry_list(pu_entries, 'power_net'),
                '隔串阻上拉数量': len(ipu_entries),
                '隔串阻上拉位号': _format_entry_list(ipu_entries, 'refdes'),
                '有下拉':   '是' if pd_entries else '否',
                '下拉数量': len(pd_entries),
                '下拉位号': _format_entry_list(pd_entries, 'refdes'),
                '隔串阻下拉数量': len(ipd_entries),
                '页面':     comp.get('page', ''),
            })

    return {
        'dup_pullups':         dup_pullups,
        'dup_pulldowns':       dup_pulldowns,
        'divider_risks':       divider_risks,
        'od_missing':          od_missing,
        'chip_pin_rows':       chip_pin_rows,
        'pullups':             dict(pullups),
        'pulldowns':           dict(pulldowns),
        'indirect_pullups':    dict(indirect_pullups),
        'indirect_pulldowns':  dict(indirect_pulldowns),
        'series_by_net':       dict(series_by_net),
    }



# ══════════════════════════════════════════════════════════
# 八、Excel 导出
# ══════════════════════════════════════════════════════════

_BL = PatternFill("solid", fgColor="1F4E79")
_OR = PatternFill("solid", fgColor="C55A11")
_GR = PatternFill("solid", fgColor="375623")
_GY = PatternFill("solid", fgColor="595959")
_RF = PatternFill("solid", fgColor="FFCCCC")
_WF = Font(color="FFFFFF", bold=True, size=10)
_BF = Font(bold=True, size=10)
_NF = Font(size=10)
_CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LA = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_TH = Side(style='thin')
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


def _xl_hdr(ws, row_idx, fill):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill; cell.font = _WF; cell.alignment = _CA; cell.border = _BD


def _xl_autowidth(ws, mx=50):
    for col in ws.columns:
        vals = [str(c.value or '') for c in col]
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(v) for v in vals), default=8) + 2, mx)


def _xl_write_rows(ws, rows: List[dict], fill, hl_col=None, freeze=True):
    if not rows:
        ws.append(['（无数据）']); return
    hdrs = list(rows[0].keys())
    ws.append(hdrs); _xl_hdr(ws, ws.max_row, fill)
    hl_idx = hdrs.index(hl_col) if hl_col in hdrs else None
    for row in rows:
        ws.append(list(row.values()))
        ri  = ws.max_row
        red = hl_idx is not None and '❌' in str(ws.cell(ri, hl_idx+1).value or '')
        for cell in ws[ri]:
            cell.border = _BD; cell.alignment = _LA; cell.font = _NF
            if red: cell.fill = _RF
    _xl_autowidth(ws)
    if freeze: ws.freeze_panes = 'A2'


def _xl_section(ws, title, fill):
    ws.append([title])
    for cell in ws[ws.max_row]:
        cell.fill = fill; cell.font = _WF; cell.border = _BD
    ws.append([])


def export_to_excel(data: dict, out_path: str) -> str:
    base, ext = os.path.splitext(out_path)
    n, path = 1, out_path
    while os.path.exists(path):
        path = f'{base}({n}){ext}'; n += 1

    wb = Workbook(); wb.remove(wb.active)
    project = data.get('project_name', '')
    na  = data.get('net_analysis', {})
    drc = data.get('drc', {})
    drt = data.get('derating', [])
    mn  = data.get('bom_normal_merged', [])
    md  = data.get('bom_depop_merged', [])
    res = data.get('resistor_analysis', {})

    # 概览
    ws = wb.create_sheet('概览')
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 16
    drc_total = sum(len(v) for v in drc.values() if isinstance(v, list))
    fail = sum(1 for r in drt if r.get('状态', '').startswith('❌'))
    for label, val in [
        ('项目名称', project),
        ('贴装元件种类数', len(mn)),
        ('贴装元件总数',  sum(r.get('数量', 0) for r in mn)),
        ('DEPOP 元件种类数', len(md)),
        ('DEPOP 元件总数',  sum(r.get('数量', 0) for r in md)),
        ('网络总数', na.get('total', '')),
        ('单端网络数（疑似漏连）', len(na.get('single_node', {}))),
        ('电源网络数', len(na.get('power_nets', {}))),
        ('差分对数', len(na.get('diff_pairs', {}))),
        ('DRC 问题总数', drc_total),
        ('电容降额不合格数', fail),
    ]:
        ws.append([label, val])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _BD
            cell.font = _BF if cell.column == 1 else _NF
            cell.alignment = _LA

    # BOM
    ws = wb.create_sheet('BOM_贴装'); _xl_write_rows(ws, mn, _BL)
    ws = wb.create_sheet('BOM_DEPOP'); _xl_write_rows(ws, md, _OR)
    ws = wb.create_sheet('BOM_明细')
    all_d = [{'DEPOP': '', **r} for r in data.get('bom_normal_detail', [])] + \
            [{'DEPOP': 'Y', **r} for r in data.get('bom_depop_detail', [])]
    _xl_write_rows(ws, all_d, _GY)

    # 网络分析
    ws = wb.create_sheet('网络分析'); ws.freeze_panes = None
    _xl_section(ws, '电源网络', _BL)
    _xl_write_rows(ws, [{'网络名': k, '节点数': len(v)}
                        for k, v in sorted(na.get('power_nets', {}).items(), key=lambda x: -len(x[1]))],
                   _BL, freeze=False)
    ws.append([])
    _xl_section(ws, 'GND 网络', _GR)
    _xl_write_rows(ws, [{'网络名': k, '节点数': len(v)}
                        for k, v in sorted(na.get('gnd_nets', {}).items(), key=lambda x: -len(x[1]))],
                   _GR, freeze=False)
    ws.append([])
    _xl_section(ws, '差分对', _OR)
    _xl_write_rows(ws, [{'基础名': b, 'P端网络': pr['P'], 'N端网络': pr['N']}
                        for b, pr in sorted(na.get('diff_pairs', {}).items())],
                   _OR, freeze=False)
    ws.append([])
    _xl_section(ws, '单端网络（疑似漏连）', _GY)
    _xl_write_rows(ws, [{'网络名': k, '连接元件': v[0]['refdes'], '引脚': v[0]['pin_name']}
                        for k, v in sorted(na.get('single_node', {}).items())],
                   _GY, freeze=False)
    ws.append([])
    _xl_section(ws, '各页面元件数', _BL)
    _xl_write_rows(ws, [{'页面': p, '元件数': c}
                        for p, c in sorted(na.get('page_counter', {}).items())],
                   _BL, freeze=False)
    _xl_autowidth(ws)

    # 设计检查
    ws = wb.create_sheet('设计检查'); ws.freeze_panes = None
    for title, key, fill in [
        ('TBD 待确认属性', 'tbd_attrs',       _OR),
        ('缺少料号',       'missing_hq_code',  _RF),
        ('缺少 VALUE',     'missing_value',     _RF),
        ('缺少封装',       'missing_package',   _RF),
        ('单端网络',       'single_pin_nets',   _GY),
        ('未命名网络',     'unnamed_nets',      _GY),
        ('BOM_OPTION 元件清单（含拼写风险）', 'bom_option_components', _BL),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, drc.get(key, []), fill, freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    # 降额
    ws = wb.create_sheet('降额分析')
    _xl_write_rows(ws, drt, _BL, hl_col='状态')

    # 电阻检查
    ws = wb.create_sheet('电阻检查'); ws.freeze_panes = None
    for title, key, hl, fill in [
        ('串阻分压风险', 'divider_risks', '状态', _OR),
        ('重复上拉', 'dup_pullups', None, _BL),
        ('重复下拉', 'dup_pulldowns', None, _GY),
        ('OD/OC 缺上拉', 'od_missing', None, _GR),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, res.get(key, []), fill, hl_col=hl, freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    # 芯片 Pin 总览
    chip_rows = res.get('chip_pin_rows', [])
    if chip_rows:
        ws = wb.create_sheet('芯片Pin总览')
        _xl_write_rows(ws, chip_rows, _BL)

    wb.save(path)
    return path



# ══════════════════════════════════════════════════════════
# 九、GUI 辅助函数
# ══════════════════════════════════════════════════════════

def _make_tree(parent, columns, height=12):
    outer = tk.Frame(parent)
    tree  = ttk.Treeview(outer, columns=columns, show='headings', height=height)
    vsb   = ttk.Scrollbar(outer, orient='vertical',   command=tree.yview)
    hsb   = ttk.Scrollbar(outer, orient='horizontal',  command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    tree.grid(row=0, column=0, sticky='nsew')
    vsb.grid(row=0, column=1, sticky='ns')
    hsb.grid(row=1, column=0, sticky='ew')
    outer.grid_rowconfigure(0, weight=1)
    outer.grid_columnconfigure(0, weight=1)
    return outer, tree


def _sort_tree(tree, col, reverse: bool):
    items = [(tree.set(iid, col), iid) for iid in tree.get_children('')]
    try:
        items.sort(key=lambda t: (float(t[0]) if t[0] else float('-inf')), reverse=reverse)
    except ValueError:
        items.sort(key=lambda t: _natural_sort_key(t[0]), reverse=reverse)
    for idx, (_, iid) in enumerate(items):
        tree.move(iid, '', idx)
    arrow = ' ▲' if not reverse else ' ▼'
    for c in tree['columns']:
        base = tree.heading(c, 'text').rstrip(' ▲▼')
        if c == col:
            tree.heading(c, text=base + arrow,
                         command=lambda _c=c: _sort_tree(tree, _c, not reverse))
        else:
            tree.heading(c, text=base,
                         command=lambda _c=c: _sort_tree(tree, _c, False))


def _fill_tree(tree, rows: list, columns: list = None):
    tree.delete(*tree.get_children())
    if not rows:
        return
    cols = columns or list(rows[0].keys())
    tree['columns'] = cols
    for c in cols:
        tree.heading(c, text=c, anchor='w',
                     command=lambda _c=c: _sort_tree(tree, _c, False))
        tree.column(c, width=min(max(len(c)*9, 80), 220), anchor='w', stretch=True)
    for row in rows:
        tree.insert('', 'end', values=[str(row.get(c, '')) for c in cols])


# ══════════════════════════════════════════════════════════
# 十、主 GUI 类
# ══════════════════════════════════════════════════════════

class PstxApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title('PSTX 原理图分析工具 v1.3')
        self.geometry('1060x740')
        self.minsize(900, 600)
        self.resizable(True, True)

        self._components = {}; self._nets = {}
        self._dn = []; self._dd = []; self._mn = []; self._md = []
        self._na = {}; self._drc = {}; self._drt = []; self._res = {}

        self.prt_path    = tk.StringVar()
        self.net_path    = tk.StringVar()
        self.proj_root   = tk.StringVar()  # 项目根目录（用于页码解析）
        self.project_var = tk.StringVar()
        self.bom_filter  = tk.StringVar(value='贴装')
        self.bom_search  = tk.StringVar()
        self.query_mode  = tk.StringVar(value='位号')
        self.query_text  = tk.StringVar()
        self.ratio_var   = tk.DoubleVar(value=70.0)
        self.include_depop_var = tk.BooleanVar(value=False)  # DEPOP 排除开关

        self.bom_filter.trace_add('write', lambda *_: self._refresh_bom())
        self.bom_search.trace_add('write', lambda *_: self._refresh_bom())

        self._build_ui()

    # ──────── UI 搭建 ─────────────────────────────────────

    def _section(self, parent, title):
        f = ttk.LabelFrame(parent, text=title, padding=8)
        f.pack(fill='x', padx=10, pady=4)
        return f

    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=6)
        self.nb = nb
        for text, builder in [
            ('  文件加载  ', self._build_load),
            ('  BOM 管理  ', self._build_bom),
            ('  网络分析  ', self._build_net),
            ('  设计检查  ', self._build_drc),
            ('  电阻检查  ', self._build_res),
            ('  电容降额  ', self._build_derating),
            ('  元件查询  ', self._build_query),
            ('  日志      ', self._build_log),
        ]:
            f = ttk.Frame(nb); nb.add(f, text=text); builder(f)

    # ── Tab：文件加载 ──────────────────────────────────────

    def _build_load(self, p):
        fa = tk.Frame(p, bg='#e8f0fe', relief='groove', bd=1)
        fa.pack(fill='x', padx=10, pady=(10, 4))
        inner = tk.Frame(fa, bg='#e8f0fe')
        inner.pack(padx=14, pady=10)
        tk.Label(inner, text='快速加载', font=('Arial', 11, 'bold'),
                 bg='#e8f0fe', fg='#1a3a8f').pack(side='left', padx=(0, 10))
        tk.Button(inner, text='  选择文件夹…  ', font=('Arial', 10, 'bold'),
                  bg='#2d6cdf', fg='white', relief='flat', padx=8, pady=4,
                  command=self._auto_detect).pack(side='left')
        self.auto_detect_lbl = tk.Label(
            inner,
            text='选择 worklib 中该项目的文件夹，自动识别并填入下方路径',
            bg='#e8f0fe', fg='#444')
        self.auto_detect_lbl.pack(side='left', padx=12)

        sep_row = tk.Frame(p)
        sep_row.pack(fill='x', padx=10, pady=(6, 0))
        ttk.Separator(sep_row, orient='horizontal').pack(side='left', fill='x', expand=True, pady=6)
        tk.Label(sep_row, text='  或手动选择各文件  ', fg='#888').pack(side='left')
        ttk.Separator(sep_row, orient='horizontal').pack(side='left', fill='x', expand=True, pady=6)

        for label, var, is_folder in [
            ('pstxprt.dat  【必须】元件属性 — 位号、料号、封装、电气参数等', self.prt_path, False),
            ('pstxnet.dat  【必须】网络连接 — 引脚与网络的映射关系',          self.net_path, False),
            ('项目根目录   【可选】用于 page.map/page*.csv 页码解析',          self.proj_root, True),
        ]:
            f = self._section(p, label)
            tk.Label(f, text='路径：').grid(row=0, column=0, sticky='w')
            ttk.Entry(f, textvariable=var, width=58).grid(row=0, column=1, padx=6)
            if is_folder:
                ttk.Button(f, text='浏览…',
                           command=lambda v=var: self._browse_folder(v)).grid(row=0, column=2)
            else:
                ttk.Button(f, text='浏览…',
                           command=lambda v=var: self._browse_dat(v)).grid(row=0, column=2)

        fp = self._section(p, '项目名称（导出报告用）')
        tk.Label(fp, text='项目名称：').grid(row=0, column=0, sticky='w')
        ttk.Entry(fp, textvariable=self.project_var, width=40).grid(
            row=0, column=1, padx=6, sticky='w')

        # DEPOP 排查开关
        fd = self._section(p, '分析选项')
        ttk.Checkbutton(fd, text='DEPOP 元件参与分析（勾选后 BOM_OPTION=DEPOP/DNP 的元件也纳入降额/电阻检查）',
                        variable=self.include_depop_var).pack(side='left', padx=4)

        br = tk.Frame(p); br.pack(pady=14)
        self.parse_btn = tk.Button(
            br, text='开始解析', font=('Arial', 13, 'bold'),
            bg='#2d6cdf', fg='white', relief='flat',
            padx=24, pady=10, command=self._run_parse)
        self.parse_btn.pack(side='left', padx=8)
        self.load_status = tk.Label(br, text='', font=('Arial', 11))
        self.load_status.pack(side='left', padx=8)

        fo = self._section(p, '解析概览')
        self.overview_text = tk.Text(fo, height=7, font=('Consolas', 9),
                                      state='disabled', bg='#f5f5f5', relief='flat')
        self.overview_text.pack(fill='x')

    # ── Tab：BOM 管理 ──────────────────────────────────────

    def _build_bom(self, p):
        ctrl = self._section(p, '筛选 / 搜索')
        for val, txt in [('贴装', '贴装元件'), ('DEPOP', 'DEPOP'), ('全部', '全部')]:
            ttk.Radiobutton(ctrl, text=txt, variable=self.bom_filter,
                            value=val).pack(side='left', padx=10)
        ttk.Separator(ctrl, orient='vertical').pack(side='left', fill='y', padx=6)
        tk.Label(ctrl, text='搜索：').pack(side='left')
        ttk.Entry(ctrl, textvariable=self.bom_search, width=26).pack(side='left', padx=4)

        outer, self.bom_tree = _make_tree(p, ['位号', '料号', '值', '封装', '类型', '页面'], height=16)
        outer.pack(fill='both', expand=True, padx=10, pady=4)

        bot = tk.Frame(p); bot.pack(fill='x', padx=10, pady=4)
        self.bom_count_lbl = tk.Label(bot, text='', fg='#444')
        self.bom_count_lbl.pack(side='left')
        ttk.Button(bot, text='导出 Excel', command=self._export_excel).pack(side='right')

    # ── Tab：网络分析 ──────────────────────────────────────

    def _build_net(self, p):
        fs = self._section(p, '汇总')
        self.net_stat = tk.Text(fs, height=2, font=('Consolas', 9),
                                 state='disabled', bg='#f5f5f5', relief='flat')
        self.net_stat.pack(fill='x')

        sub = ttk.Notebook(p); sub.pack(fill='both', expand=True, padx=10, pady=4)
        for title, cols, attr in [
            ('电源网络',     ['网络名', '节点数'],           '_tree_power'),
            ('GND 网络',     ['网络名', '节点数'],           '_tree_gnd'),
            ('差分对',       ['基础名', 'P端网络', 'N端网络'],'_tree_diff'),
            ('单端网络',     ['网络名', '连接元件', '引脚'],  '_tree_single'),
            ('各页面元件数', ['页面', '元件数'],             '_tree_pages'),
        ]:
            f = ttk.Frame(sub); sub.add(f, text=f'  {title}  ')
            outer, tree = _make_tree(f, cols, height=14)
            outer.pack(fill='both', expand=True)
            setattr(self, attr, tree)

    # ── Tab：设计检查 ──────────────────────────────────────

    def _build_drc(self, p):
        sub = ttk.Notebook(p); sub.pack(fill='both', expand=True, padx=10, pady=8)
        for title, cols, attr in [
            ('缺料号',     ['位号', '类型', '页面'],                                       '_tree_drc_hq'),
            ('缺 VALUE',   ['位号', '类型', '页面'],                                       '_tree_drc_val'),
            ('缺封装',     ['位号', '类型', '页面'],                                       '_tree_drc_pkg'),
            ('TBD 属性',   ['位号', '属性', '当前值', '类型', '页面'],                     '_tree_drc_tbd'),
            ('单端网络',   ['网络名', '连接元件', '引脚', '页面'],                         '_tree_drc_single'),
            ('BOM_OPTION', ['位号', '类型', 'BOM_OPTION值', '是否DEPOP', '拼写风险', '页面'], '_tree_drc_bom_opt'),
            ('未命名网络',     ['网络名', '节点数'],                                         '_tree_drc_unnamed'),
        ]:
            f = ttk.Frame(sub); sub.add(f, text=f'  {title}  ')
            outer, tree = _make_tree(f, cols, height=15)
            outer.pack(fill='both', expand=True)
            setattr(self, attr, tree)

    # ── Tab：电阻检查 ──────────────────────────────────────

    def _build_res(self, p):
        sub = ttk.Notebook(p); sub.pack(fill='both', expand=True, padx=10, pady=8)
        for title, cols, attr in [
            ('串阻分压风险',
             ['串阻位号', '串阻值', '串阻网络A', '串阻网络B', '偏置类型', '偏置位号',
              '偏置值', '偏置所在网络', '受影响网络', '串/偏置比', '偏置 < 1k', '状态', '页面'],
             '_tree_res_div'),
            ('重复上拉',
             ['信号网络', '上拉数量', '位号', '阻值', '上拉电源', 'BOM_OPTION', '页面'],
             '_tree_res_dup_pu'),
            ('重复下拉',
             ['信号网络', '下拉数量', '位号', '阻值', 'BOM_OPTION', '页面'],
             '_tree_res_dup_pd'),
            ('OD/OC 缺上拉',
             ['网络名', '节点数', '连接元件', '芯片引脚', '判定依据', '上拉状态', '说明'],
             '_tree_res_od'),
            ('芯片Pin总览',
             ['芯片位号', '引脚', '引脚名', '网络名', '有串阻', '串阻位号', '串阻另一端',
              '有上拉', '上拉位号', '上拉电源', '有下拉', '下拉位号', '页面'],
             '_tree_res_chip'),
        ]:
            f = ttk.Frame(sub); sub.add(f, text=f'  {title}  ')
            outer, tree = _make_tree(f, cols, height=14)
            outer.pack(fill='both', expand=True)
            setattr(self, attr, tree)

    # ── Tab：电容降额 ──────────────────────────────────────

    def _build_derating(self, p):
        fc = self._section(p, '参数设置')
        tk.Label(fc, text='工作电压上限（额定电压的 %）：').grid(row=0, column=0, sticky='w')
        ttk.Scale(fc, from_=50, to=100, orient='horizontal',
                  variable=self.ratio_var, length=200).grid(row=0, column=1, padx=8)
        self.ratio_lbl = tk.Label(fc, text='70%', width=6, font=('Arial', 11, 'bold'))
        self.ratio_lbl.grid(row=0, column=2)
        self.ratio_var.trace_add('write', lambda *_: self.ratio_lbl.configure(
            text=f'{self.ratio_var.get():.0f}%'))

        tk.Label(fc, text='自定义电压映射\n（每行 NET前缀=电压V）：',
                 justify='left').grid(row=1, column=0, sticky='nw', pady=6)
        self.volt_entry = tk.Text(fc, height=3, width=38, font=('Consolas', 9))
        self.volt_entry.grid(row=1, column=1, columnspan=2, padx=8, sticky='w')
        self.volt_entry.insert('1.0', '# 示例：VBUS=5.0\n# P12V_AUX=12.0')

        btn_row = tk.Frame(fc); btn_row.grid(row=2, column=1, sticky='w', pady=4)
        ttk.Button(btn_row, text='重新计算',
                   command=self._recalc_derating).pack(side='left')
        self._rules_visible = False
        self._rules_btn = ttk.Button(btn_row, text='查看内置电压匹配规则 ▾',
                                     command=self._toggle_rules)
        self._rules_btn.pack(side='left', padx=12)

        self._rules_frame = tk.Frame(fc, relief='sunken', bd=1, bg='#f8f8f8')
        self._rules_frame.grid(row=3, column=0, columnspan=3, sticky='ew', padx=0, pady=(0, 4))
        self._rules_frame.grid_remove()

        rules_txt = scrolledtext.ScrolledText(
            self._rules_frame, font=('Consolas', 9), height=10,
            bg='#f8f8f8', relief='flat', state='normal')
        rules_txt.pack(fill='both', expand=True, padx=6, pady=4)
        algo_text = (
            "【工作电压推断算法（v1.2 升级版）】\n"
            "  1. 读取该电容连接的所有网络名\n"
            "  2. 首先检测 AC 耦合电容（两端均接差分对同极性网络），跳过推断\n"
            "  3. 如果连接了 PG/OD 信号网络（如 PGOOD/FAULT/ALERT），标注特殊状态\n"
            "  4. 对其余网络，取首 token 用正则匹配电压（如 P3V3_AON → 3.3V）\n"
            "  5. 要求同时有接地网络才认为电压有效（单端电容跳过）\n"
            "  6. 用户可填写自定义映射（NET前缀=电压），优先级高于内置规则\n"
            "  7. 合格条件：工作电压 ≤ 额定电压 × 上限百分比（默认 70%）\n\n"
            "【PG/OD 信号识别（不推断电压）】\n"
            "  包含：PGOOD / _PG_ / PWRGD / FAULT / ALERT / SMBALERT\n"
            "        SDA / SCL / OC_N / PRSNT / INT_N / IRQ_N 等\n\n"
            "【AC 耦合电容识别（不推断电压）】\n"
            "  两端均连接差分对同极性网络，且存在镜像电容（差分另一极）\n"
        )
        rules_txt.insert('1.0', algo_text)
        rules_txt.configure(state='disabled')

        cols = ['位号', '值', '封装', '类型', '额定电压', '推断工作电压(V)',
                '推断来源网络', '推断来源类型', '降额比', '状态', '页面', 'DEPOP']
        outer, self.drt_tree = _make_tree(p, cols, height=14)
        outer.pack(fill='both', expand=True, padx=10, pady=4)
        self.drt_stat = tk.Label(p, text='', fg='#555')
        self.drt_stat.pack(anchor='w', padx=10, pady=2)

    # ── Tab：元件查询 ──────────────────────────────────────

    def _build_query(self, p):
        fc = self._section(p, '查询')
        for val, txt in [('位号', '按位号 (refdes)'), ('网络名', '按网络名')]:
            ttk.Radiobutton(fc, text=txt, variable=self.query_mode,
                            value=val).pack(side='left', padx=10)
        ttk.Separator(fc, orient='vertical').pack(side='left', fill='y', padx=6)
        ttk.Entry(fc, textvariable=self.query_text, width=30).pack(side='left', padx=4)
        ttk.Button(fc, text='查询', command=self._do_query).pack(side='left', padx=4)

        fr = self._section(p, '查询结果')
        self.query_result = scrolledtext.ScrolledText(
            fr, font=('Consolas', 10), state='disabled',
            bg='#1e1e1e', fg='#d4d4d4', relief='flat', height=22)
        self.query_result.pack(fill='both', expand=True)

    # ── Tab：日志 ──────────────────────────────────────────

    def _build_log(self, p):
        self.log = scrolledtext.ScrolledText(
            p, font=('Consolas', 9), state='disabled',
            bg='#1e1e1e', fg='#d4d4d4', relief='flat')
        self.log.pack(fill='both', expand=True, padx=8, pady=8)
        ttk.Button(p, text='清空日志',
                   command=self._clear_log).pack(anchor='e', padx=8, pady=4)


    # ──────── 事件处理 ────────────────────────────────────

    def _browse_dat(self, var):
        path = filedialog.askopenfilename(
            title='选择 .dat 文件',
            filetypes=[('DAT 文件', '*.dat'), ('所有文件', '*.*')])
        if path:
            var.set(path); self._log(f'选择文件：{path}')

    def _browse_folder(self, var):
        folder = filedialog.askdirectory(title='选择项目根目录')
        if folder:
            var.set(folder); self._log(f'选择目录：{folder}')

    def _auto_detect(self):
        folder = filedialog.askdirectory(title='选择包含 PST 文件的文件夹')
        if not folder:
            return
        candidates = [folder] + [
            os.path.join(folder, d) for d in os.listdir(folder)
            if os.path.isdir(os.path.join(folder, d))
        ]
        targets = {'pstxprt.dat': self.prt_path, 'pstxnet.dat': self.net_path}
        found, missing = {}, []
        for name, var in targets.items():
            hit = None
            for d in candidates:
                p = os.path.join(d, name)
                if os.path.isfile(p):
                    hit = p; break
            if hit:
                var.set(hit); found[name] = hit
            else:
                missing.append(name)
        # 尝试推断项目根目录（pstxprt.dat 所在目录的上级可能是项目根）
        if 'pstxprt.dat' in found and not self.proj_root.get():
            prt_dir = os.path.dirname(found['pstxprt.dat'])
            if os.path.basename(prt_dir).lower() == 'packaged':
                self.proj_root.set(os.path.dirname(prt_dir))
        if found:
            msg = f'找到 {len(found)} 个：{", ".join(found.keys())}'
            if missing:
                msg += f'    未找到：{", ".join(missing)}'
            color = '#2a8a2a' if 'pstxprt.dat' in found and 'pstxnet.dat' in found else '#b06000'
        else:
            msg = '未在该目录下找到任何 PST 文件'
            color = 'red'
        self.auto_detect_lbl.configure(text=msg, fg=color)
        self._log(f'\n自动识别文件夹：{folder}')
        for name, path in found.items():
            self._log(f'  ✅ {name} → {path}')
        for name in missing:
            self._log(f'  ⚪ {name} 未找到')

    # ──────── 解析流程 ────────────────────────────────────

    def _run_parse(self):
        if not self.prt_path.get().strip() or not self.net_path.get().strip():
            messagebox.showerror('错误', '请先选择 pstxprt.dat 和 pstxnet.dat')
            return
        self.parse_btn.configure(state='disabled')
        self._start_spinner('解析中')
        threading.Thread(target=self._do_parse, daemon=True).start()

    def _do_parse(self):
        try:
            self._log('\n── 开始解析 ──────────────────')
            # 使用多编码回退读取文件
            prt_texts = _iter_text_with_fallback_encodings(self.prt_path.get())
            net_texts = _iter_text_with_fallback_encodings(self.net_path.get())
            if not prt_texts or not net_texts:
                raise FileNotFoundError('无法读取文件，请检查路径')
            prt = prt_texts[0]
            net = net_texts[0]
            self._log(f'  pstxprt：{len(prt):,} 字节    pstxnet：{len(net):,} 字节')

            comps, nets, _ = parse_all(prt, net)
            self._log(f'  元件：{len(comps)} 个    网络：{len(nets)} 个')

            # 页码解析（优先使用项目根目录，否则自动推断）
            proj_root = self.proj_root.get().strip()
            if not proj_root:
                proj_root = _infer_project_root_from_data_paths(
                    self.prt_path.get(), self.net_path.get())
                if proj_root:
                    self._log(f'  🔍 自动推断项目根目录：{proj_root}')
                    self.after(0, lambda: self.proj_root.set(proj_root))
            page_warnings = resolve_component_pages(comps, proj_root)
            if proj_root:
                self._log(f'  项目根目录：{proj_root}')
            if page_warnings:
                for w in page_warnings:
                    self._log(f'  ⚠ 页码：{w}')

            include_depop = self.include_depop_var.get()
            self._log(f'  DEPOP 元件参与分析：{"是" if include_depop else "否"}')

            dn, dd, mn, md = build_bom(comps)
            na  = analyze_networks(nets, comps)
            drc = check_drc(comps, nets)
            drt = analyze_derating(comps, nets, self.ratio_var.get(), self._volt_map(), include_depop)
            res = analyze_resistors(comps, nets, exclude_depop=not include_depop)

            self._components = comps; self._nets = nets
            self._dn = dn; self._dd = dd; self._mn = mn; self._md = md
            self._na = na; self._drc = drc; self._drt = drt; self._res = res

            drc_total = sum(len(v) for v in drc.values() if isinstance(v, list))
            self._log(f'  贴装 {len(mn)} 种 / {sum(r.get("数量",0) for r in mn)} 个')
            self._log(f'  DRC 问题：{drc_total} 项    '
                      f'降额不合格：{sum(1 for r in drt if r["状态"].startswith("❌"))} 个')
            self.after(0, self._on_parse_done)
        except Exception as e:
            import traceback
            self._log(f'❌ {e}\n{traceback.format_exc()}')
            self.after(0, lambda: self._stop_spinner('❌ 解析失败'))
            self.after(0, lambda: messagebox.showerror('错误', str(e)))
        finally:
            self.after(0, lambda: self.parse_btn.configure(state='normal'))

    def _on_parse_done(self):
        self._stop_spinner('✅ 解析完成')
        self._update_overview()
        self._refresh_bom()
        self._refresh_net()
        self._refresh_drc()
        self._refresh_res()
        self._refresh_derating()
        self.nb.select(1)

    def _update_overview(self):
        na = self._na; drc = self._drc; mn = self._mn; md = self._md; drt = self._drt
        drc_total = sum(len(v) for v in drc.values() if isinstance(v, list))
        fail = sum(1 for r in drt if r.get('状态', '').startswith('❌'))
        lines = [
            f'贴装元件：{len(mn)} 种 / {sum(r.get("数量",0) for r in mn)} 个',
            f'DEPOP：   {len(md)} 种 / {sum(r.get("数量",0) for r in md)} 个',
            f'网络：{na.get("total",0)} 个   电源：{len(na.get("power_nets",{}))}   '
            f'GND：{len(na.get("gnd_nets",{}))}   差分对：{len(na.get("diff_pairs",{}))}',
            f'单端网络（疑似漏连）：{len(na.get("single_node",{}))}',
            f'DRC 问题：{drc_total}   电容降额不合格：{fail}',
            f'DEPOP 参与分析：{"是" if self.include_depop_var.get() else "否"}',
        ]
        self.overview_text.configure(state='normal')
        self.overview_text.delete('1.0', 'end')
        self.overview_text.insert('end', '\n'.join(lines))
        self.overview_text.configure(state='disabled')

    # ──────── BOM ─────────────────────────────────────────

    def _refresh_bom(self):
        mode = self.bom_filter.get()
        kw   = self.bom_search.get().strip().lower()
        src  = (self._dn if mode == '贴装' else
                self._dd if mode == 'DEPOP' else
                self._dn + self._dd)
        if kw:
            src = [r for r in src if any(kw in str(v).lower() for v in r.values())]
        _fill_tree(self.bom_tree, src, ['位号', '料号', '值', '封装', '类型', '页面'])
        self.bom_count_lbl.configure(text=f'共 {len(src)} 行')

    # ──────── 网络 ─────────────────────────────────────────

    def _refresh_net(self):
        na = self._na
        stat = (f'网络总数：{na.get("total",0)}    '
                f'电源：{len(na.get("power_nets",{}))}    '
                f'GND：{len(na.get("gnd_nets",{}))}    '
                f'差分对：{len(na.get("diff_pairs",{}))}    '
                f'单端（疑似漏连）：{len(na.get("single_node",{}))}')
        self.net_stat.configure(state='normal')
        self.net_stat.delete('1.0', 'end')
        self.net_stat.insert('end', stat)
        self.net_stat.configure(state='disabled')

        _fill_tree(self._tree_power,
                   [{'网络名': k, '节点数': len(v)}
                    for k, v in sorted(na.get('power_nets', {}).items(), key=lambda x: -len(x[1]))],
                   ['网络名', '节点数'])
        _fill_tree(self._tree_gnd,
                   [{'网络名': k, '节点数': len(v)}
                    for k, v in sorted(na.get('gnd_nets', {}).items(), key=lambda x: -len(x[1]))],
                   ['网络名', '节点数'])
        _fill_tree(self._tree_diff,
                   [{'基础名': b, 'P端网络': pr['P'], 'N端网络': pr['N']}
                    for b, pr in sorted(na.get('diff_pairs', {}).items())],
                   ['基础名', 'P端网络', 'N端网络'])
        _fill_tree(self._tree_single,
                   [{'网络名': k, '连接元件': v[0]['refdes'], '引脚': v[0]['pin_name']}
                    for k, v in sorted(na.get('single_node', {}).items())],
                   ['网络名', '连接元件', '引脚'])
        _fill_tree(self._tree_pages,
                   [{'页面': pg, '元件数': cnt}
                    for pg, cnt in sorted(na.get('page_counter', {}).items(),
                                          key=lambda x: _natural_sort_key(x[0]))],
                   ['页面', '元件数'])

    # ──────── DRC ─────────────────────────────────────────

    def _refresh_drc(self):
        drc = self._drc
        _fill_tree(self._tree_drc_hq,     drc.get('missing_hq_code', []),  ['位号', '类型', '页面'])
        _fill_tree(self._tree_drc_val,    drc.get('missing_value', []),    ['位号', '类型', '页面'])
        _fill_tree(self._tree_drc_pkg,    drc.get('missing_package', []),  ['位号', '类型', '页面'])
        _fill_tree(self._tree_drc_tbd,    drc.get('tbd_attrs', []),
                   ['位号', '属性', '当前值', '类型', '页面'])
        _fill_tree(self._tree_drc_single, drc.get('single_pin_nets', []),
                   ['网络名', '连接元件', '引脚', '页面'])
        _fill_tree(self._tree_drc_bom_opt, drc.get('bom_option_components', []),
                   ['位号', '类型', 'BOM_OPTION值', '是否DEPOP', '拼写风险', '页面'])
        _fill_tree(self._tree_drc_unnamed, drc.get('unnamed_nets', []),
                   ['网络名', '节点数'])

    # ──────── 电阻检查 ─────────────────────────────────────

    def _refresh_res(self):
        ra = self._res
        _fill_tree(self._tree_res_div, ra.get('divider_risks', []),
                   ['串阻位号', '串阻值', '串阻网络A', '串阻网络B', '偏置类型', '偏置位号',
                    '偏置值', '偏置所在网络', '受影响网络', '串/偏置比', '偏置 < 1k', '状态', '页面'])
        _fill_tree(self._tree_res_dup_pu, ra.get('dup_pullups', []),
                   ['信号网络', '上拉数量', '位号', '阻值', '上拉电源', 'BOM_OPTION', '页面'])
        _fill_tree(self._tree_res_dup_pd, ra.get('dup_pulldowns', []),
                   ['信号网络', '下拉数量', '位号', '阻值', 'BOM_OPTION', '页面'])
        _fill_tree(self._tree_res_od, ra.get('od_missing', []),
                   ['网络名', '节点数', '连接元件', '芯片引脚', '判定依据', '上拉状态', '说明'])
        _fill_tree(self._tree_res_chip, ra.get('chip_pin_rows', []),
                   ['芯片位号', '引脚', '引脚名', '网络名', '有串阻', '串阻位号', '串阻另一端',
                    '有上拉', '上拉位号', '上拉电源', '有下拉', '下拉位号', '页面'])

    # ──────── 降额 ─────────────────────────────────────────

    def _refresh_derating(self):
        cols = ['位号', '值', '封装', '类型', '额定电压', '推断工作电压(V)',
                '推断来源网络', '推断来源类型', '降额比', '状态', '页面', 'DEPOP']
        _fill_tree(self.drt_tree, self._drt, cols)
        total = len(self._drt)
        fail  = sum(1 for r in self._drt if r.get('状态', '').startswith('❌'))
        ok    = sum(1 for r in self._drt if r.get('状态', '').startswith('✅'))
        self.drt_stat.configure(
            text=f'共 {total} 个电容  |  ✅ 合格 {ok}  |  ❌ 不合格 {fail}  |  ⚪ 无法判断 {total-fail-ok}')

    def _recalc_derating(self):
        if not self._components:
            messagebox.showwarning('提示', '请先解析文件'); return
        self._drt = analyze_derating(
            self._components, self._nets, self.ratio_var.get(), self._volt_map(),
            self.include_depop_var.get())
        self._refresh_derating()
        self._log(f'降额重新计算完成（上限={self.ratio_var.get():.0f}%）')

    def _toggle_rules(self):
        self._rules_visible = not self._rules_visible
        if self._rules_visible:
            self._rules_frame.grid()
            self._rules_btn.configure(text='收起内置电压匹配规则 ▴')
        else:
            self._rules_frame.grid_remove()
            self._rules_btn.configure(text='查看内置电压匹配规则 ▾')

    def _volt_map(self):
        result = {}
        for line in self.volt_entry.get('1.0', 'end').splitlines():
            line = line.strip()
            if not line or line.startswith('#'): continue
            if '=' in line:
                k, _, v = line.partition('=')
                try: result[k.strip()] = float(v.strip())
                except ValueError: pass
        return result or None

    # ──────── 查询 ─────────────────────────────────────────

    def _do_query(self):
        kw   = self.query_text.get().strip()
        mode = self.query_mode.get()
        if not kw: return
        if not self._components:
            messagebox.showwarning('提示', '请先解析文件'); return

        lines = []
        if mode == '位号':
            comp = self._components.get(kw) or next(
                (v for k, v in self._components.items() if k.upper() == kw.upper()), None)
            if comp:
                lines.append(f'═══ 元件：{kw} ═══')
                for k, v in comp.items():
                    if k == 'nets': continue
                    lines.append(f'  {k:<20} {v}')
                lines += ['', '  引脚 → 网络：']
                for pin, net in sorted(comp.get('nets', {}).items(), key=lambda x: _natural_sort_key(x[0])):
                    lines.append(f'    pin {pin:<6} → {net}')
            else:
                matched = sorted(r for r in self._components if kw.upper() in r.upper())
                lines.append('未找到精确匹配，模糊结果：' if matched else f'未找到位号：{kw}')
                lines.extend(f'  {r}' for r in matched[:50])
        else:
            nodes = self._nets.get(kw) or self._nets.get(
                next((k for k in self._nets if k.upper() == kw.upper()), ''))
            if nodes:
                net_key = kw if kw in self._nets else next(k for k in self._nets if k.upper() == kw.upper())
                lines.append(f'═══ 网络：{net_key}（{len(nodes)} 个节点）═══')
                for n in nodes:
                    comp = self._components.get(n['refdes'], {})
                    desc = comp.get('value', '') or comp.get('part_name', '')
                    lines.append(f'  {n["refdes"]:<10} pin {n["pin"]:<6} ({n["pin_name"]:<12}) {desc}')
            else:
                matched = sorted(k for k in self._nets if kw.upper() in k.upper())
                lines.append('未找到精确匹配，模糊结果：' if matched else f'未找到网络：{kw}')
                lines.extend(f'  {nm}  ({len(self._nets[nm])} nodes)' for nm in matched[:50])

        self.query_result.configure(state='normal')
        self.query_result.delete('1.0', 'end')
        self.query_result.insert('end', '\n'.join(lines))
        self.query_result.configure(state='disabled')

    # ──────── 导出 ─────────────────────────────────────────

    def _export_excel(self):
        if not self._components:
            messagebox.showwarning('提示', '请先解析文件'); return
        out = filedialog.asksaveasfilename(
            title='保存分析报告', initialfile='pstx_分析报告.xlsx',
            defaultextension='.xlsx', filetypes=[('Excel 文件', '*.xlsx')])
        if not out: return
        self._log('\n导出 Excel 中…')
        threading.Thread(target=self._do_export, args=(out,), daemon=True).start()

    def _do_export(self, path):
        try:
            actual = export_to_excel({
                'project_name':      self.project_var.get().strip() or '未命名项目',
                'bom_normal_detail': self._dn, 'bom_depop_detail': self._dd,
                'bom_normal_merged': self._mn, 'bom_depop_merged': self._md,
                'net_analysis': self._na, 'drc': self._drc, 'derating': self._drt,
                'resistor_analysis': self._res,
            }, path)
            self._log(f'✅ 导出完成：{actual}')
            self.after(0, lambda: messagebox.showinfo('完成', f'导出成功！\n{actual}'))
            folder = os.path.dirname(os.path.abspath(actual))
            try:
                if sys.platform == 'win32':    os.startfile(folder)
                elif sys.platform == 'darwin': subprocess.Popen(['open', folder])
                else:                          subprocess.Popen(['xdg-open', folder])
            except Exception:
                pass
        except Exception as e:
            import traceback
            self._log(f'❌ 导出失败：{e}\n{traceback.format_exc()}')
            self.after(0, lambda: messagebox.showerror('错误', str(e)))

    # ──────── Spinner + 日志 ──────────────────────────────

    def _start_spinner(self, label='处理中'):
        self._spinning = True; self._spin_step = 0; self._spin_label = label; self._spin()

    def _spin(self):
        if not self._spinning: return
        f = ['◐', '◓', '◑', '◒'][self._spin_step % 4]
        self.load_status.configure(text=f'{f} {self._spin_label}，请稍候…', fg='#2d6cdf')
        self._spin_step += 1
        self._spin_job = self.after(200, self._spin)

    def _stop_spinner(self, msg=''):
        self._spinning = False
        if hasattr(self, '_spin_job'): self.after_cancel(self._spin_job)
        color = '#2a8a2a' if msg.startswith('✅') else ('red' if '❌' in msg else '#333')
        self.load_status.configure(text=msg, fg=color)

    def _log(self, msg):
        def _w():
            self.log.configure(state='normal')
            self.log.insert('end', msg + '\n'); self.log.see('end')
            self.log.configure(state='disabled')
        self.after(0, _w)

    def _clear_log(self):
        self.log.configure(state='normal'); self.log.delete('1.0', 'end')
        self.log.configure(state='disabled')


if __name__ == '__main__':
    app = PstxApp()
    app.mainloop()
