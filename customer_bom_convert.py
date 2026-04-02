# -*- coding: utf-8 -*-
"""
客户BOM → 内部整机BOM评审格式转换脚本
使用方法：python customer_bom_convert.py
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== 配置区 ==========
INPUT_FILE  = "客户BOM.xlsx"       # 客户BOM文件名，改成实际文件名
INPUT_SHEET = 0                    # Sheet索引，0=第一个Sheet；也可填Sheet名如 "Sheet1"
HEADER_ROW  = 1                    # 客户BOM表头行号（第几行是列名）

# 客户BOM列名关键字（脚本会自动匹配，不区分大小写）
COL_NAME_KEYWORD    = "名"          # 物料名称列（含此关键字的列）
COL_QTY_KEYWORD     = "用量"        # 用量列
COL_BRAND_KEYWORD   = "品牌型号"    # 品牌型号列（含 厂商:型号||厂商:型号 格式）

OUTPUT_FILE = "内部评审BOM.xlsx"    # 输出文件名
PROJECT_NAME = "项目名称"           # 项目名称（可修改）
# ============================

SUPPLIER_LABELS = ["主供", "二供", "三供", "四供", "五供", "六供", "七供", "八供", "九供", "十供"]


def find_col(ws, header_row, keyword):
    """在表头行查找含有keyword的列，返回列号（1-based）"""
    for cell in ws[header_row]:
        if cell.value and keyword in str(cell.value):
            return cell.column
    return None


def parse_brand_model(brand_model_str):
    """
    解析品牌型号字符串，返回 [(厂商, 型号), ...] 列表
    支持格式：厂商:型号||厂商:型号  或  厂商：型号  （全角冒号）
    """
    if not brand_model_str or str(brand_model_str).strip() == "":
        return []

    raw = str(brand_model_str).strip()
    # 统一分隔符
    raw = raw.replace("：", ":")   # 全角冒号 → 半角

    entries = [e.strip() for e in raw.split("||") if e.strip()]
    result = []
    for entry in entries:
        if ":" in entry:
            parts = entry.split(":", 1)
            brand = parts[0].strip()
            model = parts[1].strip()
        else:
            brand = ""
            model = entry.strip()
        result.append((brand, model))
    return result


def write_review_bom(rows, output_file, project_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "SW节点整机BOM配置"

    # ---------- 颜色定义 ----------
    GREEN   = "92D050"
    YELLOW  = "FFFF00"
    ORANGE  = "FFC000"
    DARK_BG = "1F4E79"

    def cell_style(cell, bold=False, bg=None, font_color="000000",
                   h_align="center", v_align="center", wrap=False, font_size=11):
        cell.font = Font(bold=bold, color=font_color, size=font_size)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)

    # ---------- 第1行：项目名称 ----------
    ws.merge_cells("A1:A2")
    ws["A1"] = "项目名称"
    cell_style(ws["A1"], bold=True, bg=GREEN, font_color="000000", font_size=14)

    ws.merge_cells("B1:B2")
    ws["B1"] = project_name
    cell_style(ws["B1"], bold=True, bg=GREEN, font_color="000000", font_size=14)

    ws.merge_cells("E1:I2")
    ws["E1"] = "整机BOM配置表"
    cell_style(ws["E1"], bold=True, bg=GREEN, font_color="000000", font_size=16)

    ws["J1"] = "配置说明"
    cell_style(ws["J1"], bold=True, bg=GREEN)
    ws["K1"] = "TBD"
    cell_style(ws["K1"], bg="BDD7EE")

    # ---------- 第3行：SW节点 ----------
    ws.merge_cells("A3:I3")
    ws["A3"] = "SW节点HQ SN"
    cell_style(ws["A3"], bold=True, bg=YELLOW, font_color="FF0000", font_size=12)
    ws["K3"] = ""
    cell_style(ws["K3"], bg=ORANGE, font_color="FF0000")

    # ---------- 第4行：表头 ----------
    headers = ["序号", "组件子类", "虚拟层/物料", "物料类型", "HQ PN",
               "物料名称", "厂商型号", "厂商", "主二供", "", "用量"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell_style(cell, bold=True, bg="D9D9D9", font_color="000000")
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[4].height = 22

    # ---------- 数据行（从第5行开始）----------
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    data_row = 5
    for item in rows:
        seq        = item["seq"]
        suppliers  = item["suppliers"]   # [(厂商, 型号, 用量), ...]
        name       = item["name"]

        for s_idx, (brand, model, qty) in enumerate(suppliers):
            supplier_label = SUPPLIER_LABELS[s_idx] if s_idx < len(SUPPLIER_LABELS) else f"{s_idx+1}供"

            values = [seq, "", "", "", "", name, model, brand, supplier_label, "", qty]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            data_row += 1

    # ---------- 列宽 ----------
    col_widths = [6, 10, 12, 10, 18, 30, 30, 20, 8, 6, 8]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 20

    wb.save(output_file)
    print(f"输出完成：{output_file}，共 {data_row - 5} 行数据（含替代料展开）")


def convert(input_file, input_sheet, output_file, project_name):
    print(f"读取客户BOM：{input_file}")
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"[错误] 无法打开文件：{e}")
        return

    if isinstance(input_sheet, int):
        ws = wb.worksheets[input_sheet]
    else:
        ws = wb[input_sheet]

    print(f"使用Sheet：{ws.title}")

    # 自动查找列
    col_name  = find_col(ws, HEADER_ROW, COL_NAME_KEYWORD)
    col_qty   = find_col(ws, HEADER_ROW, COL_QTY_KEYWORD)
    col_brand = find_col(ws, HEADER_ROW, COL_BRAND_KEYWORD)

    if not all([col_name, col_qty, col_brand]):
        print(f"[错误] 列识别失败 → 名={col_name}, 用量={col_qty}, 品牌型号={col_brand}")
        print("请检查 HEADER_ROW 是否设置正确，或手动在配置区指定列号")
        return

    print(f"列位置 → 名={col_name}, 用量={col_qty}, 品牌型号={col_brand}")

    rows = []
    seq = 0

    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        name_val  = ws.cell(row=row_idx, column=col_name).value
        qty_val   = ws.cell(row=row_idx, column=col_qty).value
        brand_val = ws.cell(row=row_idx, column=col_brand).value

        # 跳过空行
        if not name_val and not brand_val:
            continue

        suppliers_raw = parse_brand_model(brand_val)
        if not suppliers_raw:
            # 没有品牌型号信息，仍保留该行（厂商和型号留空）
            suppliers_raw = [("", "")]

        # 处理用量
        try:
            main_qty = float(qty_val) if qty_val not in (None, "") else 0
            main_qty = int(main_qty) if main_qty == int(main_qty) else main_qty
        except (ValueError, TypeError):
            main_qty = qty_val

        suppliers_with_qty = []
        for s_idx, (brand, model) in enumerate(suppliers_raw):
            qty = main_qty if s_idx == 0 else 0
            suppliers_with_qty.append((brand, model, qty))

        seq += 1
        rows.append({
            "seq": seq,
            "name": str(name_val).strip() if name_val else "",
            "suppliers": suppliers_with_qty,
        })

    print(f"解析到 {len(rows)} 个物料，展开后共 {sum(len(r['suppliers']) for r in rows)} 行")
    write_review_bom(rows, output_file, project_name)


if __name__ == "__main__":
    convert(INPUT_FILE, INPUT_SHEET, OUTPUT_FILE, PROJECT_NAME)
