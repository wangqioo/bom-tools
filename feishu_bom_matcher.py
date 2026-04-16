# -*- coding: utf-8 -*-
"""
飞书优选库匹配工具 v3.0
- 支持多张优选库表格，每张有多个 Sheet
- 本地 SQLite 缓存，同步一次即可离线匹配
- 每次拿到新表格 token，添加进来配置好列映射即可纳入全局匹配

认证方式：企业内部 API 网关（origin + 工号，无需 App Secret）
依赖：pip install openpyxl requests（首次运行自动安装）
运行：python feishu_bom_matcher.py
"""

import sys, subprocess
for _pkg in ["openpyxl", "requests"]:
    try:
        __import__(_pkg)
    except ImportError:
        print(f"正在安装 {_pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", _pkg, "-q"])

import os, json, sqlite3, threading, time
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

VERSION      = "v3.0"
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE  = os.path.join(BASE_DIR, "feishu_libraries.json")
CACHE_FILE   = os.path.join(BASE_DIR, "feishu_cache.db")
DEFAULT_BASE_URL = "https://your-gateway.example.com"  # 替换为实际网关地址
DEFAULT_ORIGIN   = ""   # 替换为你的 App ID
DEFAULT_USER_ID  = ""   # 替换为你的工号

# ─────────────────────────────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────────────────────────────
def _cell_str(val):
    if val is None: return ""
    if isinstance(val, (int, float)): return str(val)
    if isinstance(val, str): return val.strip()
    if isinstance(val, list):
        parts = [item.get("text") or item.get("link") or ""
                 for item in val if isinstance(item, dict)]
        return " ".join(parts).strip()
    return str(val).strip()

def _unique_path(path):
    if not os.path.exists(path): return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        c = f"{base}({n}){ext}"
        if not os.path.exists(c):
            try:
                with open(c, "ab"): pass
                os.remove(c)
                return c
            except PermissionError: pass
        n += 1

# ─────────────────────────────────────────────────────────────────────
# API 层
# ─────────────────────────────────────────────────────────────────────
class FeishuAPI:
    def __init__(self, base_url, origin, user_id):
        self.base_url = base_url.rstrip("/")
        self.origin   = origin
        self.user_id  = user_id

    def get_sheets(self, token):
        url = f"{self.base_url}/fs/sheet/v1/spreadsheetsMetainfo"
        r = requests.get(url, params={"origin": self.origin, "userId": self.user_id,
                                      "spreadsheetToken": token}, timeout=15)
        r.raise_for_status()
        data = r.json()
        if data.get("code") not in (0, 200):
            raise RuntimeError(f"获取表格元信息失败: {data.get('msg')} (code={data.get('code')})")
        return [s for s in data["data"]["sheets"] if s.get("title")]

    def read_sheet(self, token, sheet_id, row_count=5000):
        safe = min(max(row_count, 50), 10000)
        url  = f"{self.base_url}/fs/sheet/v1/getSheetsValue"
        r = requests.get(url, params={"origin": self.origin, "userId": self.user_id,
                                      "spreadsheetToken": token,
                                      "range": f"{sheet_id}!A1:Z{safe}"}, timeout=60)
        r.raise_for_status()
        data = r.json()
        if data.get("code") not in (0, 200):
            raise RuntimeError(f"读取数据失败: {data.get('msg')} (code={data.get('code')})")
        values = data["data"]["valueRange"].get("values") or []
        while values and not any(_cell_str(v) for v in values[-1]):
            values.pop()
        return values

# ─────────────────────────────────────────────────────────────────────
# 配置管理
# ─────────────────────────────────────────────────────────────────────
class ConfigManager:
    def __init__(self):
        self.data = self._load()

    def _load(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {"base_url": DEFAULT_BASE_URL, "origin": DEFAULT_ORIGIN,
                "user_id": DEFAULT_USER_ID, "libraries": []}

    def save(self):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    @property
    def base_url(self): return self.data.get("base_url", DEFAULT_BASE_URL)
    @property
    def origin(self):   return self.data.get("origin", DEFAULT_ORIGIN)
    @property
    def user_id(self):  return self.data.get("user_id", DEFAULT_USER_ID)
    @property
    def libraries(self): return self.data.get("libraries", [])

    def add_library(self, name, token, sheets_cfg):
        lib = {"id": f"lib_{int(time.time()*1000)}", "name": name,
               "token": token, "sheets": sheets_cfg, "last_sync": None}
        self.data["libraries"].append(lib)
        self.save()
        return lib

    def update_library(self, lib_id, **kwargs):
        for lib in self.data["libraries"]:
            if lib["id"] == lib_id:
                lib.update(kwargs); break
        self.save()

    def delete_library(self, lib_id):
        self.data["libraries"] = [l for l in self.data["libraries"] if l["id"] != lib_id]
        self.save()

    def get_library(self, lib_id):
        return next((l for l in self.data["libraries"] if l["id"] == lib_id), None)

# ─────────────────────────────────────────────────────────────────────
# 缓存管理（SQLite）
# ─────────────────────────────────────────────────────────────────────
class CacheManager:
    def __init__(self):
        self.conn = sqlite3.connect(CACHE_FILE, check_same_thread=False)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lib_id TEXT, lib_name TEXT, sheet_name TEXT,
                key_value TEXT COLLATE NOCASE,
                hq_no TEXT, brand TEXT, spec TEXT, description TEXT,
                raw_data TEXT, synced_at TEXT)""")
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_key ON materials(key_value COLLATE NOCASE)")
        self.conn.commit()

    def clear_library(self, lib_id):
        self.conn.execute("DELETE FROM materials WHERE lib_id=?", (lib_id,))
        self.conn.commit()

    def insert_rows(self, rows):
        self.conn.executemany(
            "INSERT INTO materials(lib_id,lib_name,sheet_name,key_value,"
            "hq_no,brand,spec,description,raw_data,synced_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
            rows)
        self.conn.commit()

    def search(self, key_value):
        cur = self.conn.execute(
            "SELECT lib_name,sheet_name,key_value,hq_no,brand,spec,description,raw_data "
            "FROM materials WHERE key_value=? COLLATE NOCASE", (key_value,))
        return cur.fetchall()

    def stats(self):
        return self.conn.execute(
            "SELECT lib_id,lib_name,COUNT(*) FROM materials GROUP BY lib_id,lib_name"
        ).fetchall()

    def count(self):
        return self.conn.execute("SELECT COUNT(*) FROM materials").fetchone()[0]

    def close(self):
        self.conn.close()

# ─────────────────────────────────────────────────────────────────────
# 同步引擎
# ─────────────────────────────────────────────────────────────────────
def sync_library(api, cache, lib, log_cb):
    lib_id, lib_name, token = lib["id"], lib["name"], lib["token"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cache.clear_library(lib_id)
    total = 0
    enabled = [s for s in lib.get("sheets", []) if s.get("enabled", True)]
    log_cb(f"[{lib_name}] 开始同步 {len(enabled)} 个 Sheet...")

    for sheet_cfg in enabled:
        sid, title = sheet_cfg["sheet_id"], sheet_cfg["title"]
        k_col = sheet_cfg.get("key_col", "")
        if not k_col:
            log_cb(f"  [{title}] 跳过（未配置关键列）"); continue
        try:
            values = api.read_sheet(token, sid, sheet_cfg.get("row_count", 5000))
        except Exception as e:
            log_cb(f"  [{title}] 读取失败: {e}"); continue
        if not values:
            log_cb(f"  [{title}] 无数据，跳过"); continue

        headers = [_cell_str(h) for h in values[0]]
        def idx(col):
            if not col: return -1
            try: return headers.index(col)
            except ValueError: return -1

        ki = idx(k_col)
        if ki == -1:
            log_cb(f"  [{title}] 找不到关键列 '{k_col}'，跳过"); continue

        hi = idx(sheet_cfg.get("hq_no_col", ""))
        bi = idx(sheet_cfg.get("brand_col", ""))
        si = idx(sheet_cfg.get("spec_col", ""))
        di = idx(sheet_cfg.get("desc_col", ""))

        rows = []
        for row in values[1:]:
            def g(i): return _cell_str(row[i]) if 0 <= i < len(row) else ""
            kv = g(ki)
            if not kv: continue
            raw = json.dumps({headers[j]: _cell_str(row[j])
                              for j in range(min(len(headers), len(row)))},
                             ensure_ascii=False)
            rows.append((lib_id, lib_name, title, kv, g(hi), g(bi), g(si), g(di), raw, now))
        cache.insert_rows(rows)
        total += len(rows)
        log_cb(f"  [{title}] ✓ {len(rows)} 行")

    log_cb(f"[{lib_name}] 同步完成，共 {total} 行")
    return total

# ─────────────────────────────────────────────────────────────────────
# BOM 匹配 & 输出
# ─────────────────────────────────────────────────────────────────────
def match_bom(ws, header_row, key_col_name, cache, output_cols, show_source, out_path, log_cb):
    max_col = ws.max_column
    headers = [_cell_str(ws.cell(row=header_row, column=ci).value)
               for ci in range(1, max_col + 1)]
    try:
        key_idx = headers.index(key_col_name)
    except ValueError:
        raise ValueError(f"BOM中找不到列 '{key_col_name}'，现有列: {headers}")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "匹配结果"
    thin = Side(style="thin")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill     = PatternFill("solid", fgColor="D9D9D9")
    hq_hdr_fill  = PatternFill("solid", fgColor="FFC000")
    hq_data_fill = PatternFill("solid", fgColor="FFFF00")

    extra_names = [c[0] for c in output_cols]
    if show_source:
        extra_names += ["来源库", "来源Sheet"]
    out_headers = headers + extra_names

    for ci, h in enumerate(out_headers, 1):
        c = ws_out.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True)
        c.fill      = hq_hdr_fill if ci > max_col else hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr
    ws_out.row_dimensions[1].height = 22
    for ci in range(1, max_col + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 16
    for ci in range(max_col + 1, max_col + len(extra_names) + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 24

    field_map = {"hq_no": 3, "brand": 4, "spec": 5, "description": 6}
    dr = 2
    matched = 0
    unmatched = 0

    for ri in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=ri, column=ci).value for ci in range(1, max_col + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        key     = _cell_str(row_vals[key_idx])
        matches = cache.search(key) if key else []
        rows_to_write = matches if matches else [None]
        first = True

        for m in rows_to_write:
            for ci, val in enumerate(row_vals, 1):
                c = ws_out.cell(row=dr, column=ci, value=val if first else None)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border    = bdr
            extra_vals = [m[field_map[f]] if m else "" for _, f in output_cols]
            if show_source:
                extra_vals += [m[0] if m else "", m[1] if m else ""]
            for j, val in enumerate(extra_vals):
                c = ws_out.cell(row=dr, column=max_col + j + 1, value=val)
                c.fill      = hq_data_fill
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border    = bdr
            first = False
            dr   += 1

        if matches: matched   += 1
        else:       unmatched += 1

    out_path = _unique_path(out_path)
    wb_out.save(out_path)
    log_cb(f"匹配完成: {matched} 个匹配成功, {unmatched} 个未匹配")
    log_cb(f"输出: {os.path.abspath(out_path)}")
    return dr - 2, matched, unmatched, out_path

# ─────────────────────────────────────────────────────────────────────
# 列映射配置对话框
# ─────────────────────────────────────────────────────────────────────
class MappingDialog(tk.Toplevel):
    def __init__(self, parent, sheet_cfg, headers, lib_name):
        super().__init__(parent)
        self.title(f"配置列映射 — {lib_name} / {sheet_cfg['title']}")
        self.grab_set()
        self.resizable(False, False)
        self.result = None
        self._cfg  = dict(sheet_cfg)
        choices    = ["（不使用）"] + headers
        fields = [
            ("key_col",   "匹配关键列 *",  "BOM中的厂家型号与此列对应（必填）"),
            ("hq_no_col", "HQ料号列 *",    "输出为 HQ料号（必填）"),
            ("brand_col", "制造商列",       "输出为 HQ制造商（可选）"),
            ("spec_col",  "规格型号列",     "输出为 HQ规格型号（可选）"),
            ("desc_col",  "描述列",         "输出为 HQ描述（可选）"),
        ]
        f = ttk.LabelFrame(self, text=f"  {sheet_cfg['title']}  ", padding=12)
        f.pack(padx=14, pady=10, fill="x")
        self._vars = {}
        for i, (key, lbl, tip) in enumerate(fields):
            tk.Label(f, text=lbl, width=14, anchor="w").grid(row=i, column=0, sticky="w", pady=4)
            cur = sheet_cfg.get(key, "")
            var = tk.StringVar(value=cur if cur else "（不使用）")
            ttk.Combobox(f, textvariable=var, values=choices, width=28,
                         state="readonly").grid(row=i, column=1, padx=6, sticky="w")
            tk.Label(f, text=tip, fg="#777").grid(row=i, column=2, sticky="w", padx=4)
            self._vars[key] = var
        row_btn = tk.Frame(self); row_btn.pack(pady=8)
        ttk.Button(row_btn, text="确定", command=self._ok, width=10).pack(side="left", padx=6)
        ttk.Button(row_btn, text="取消", command=self.destroy, width=10).pack(side="left")

    def _ok(self):
        cfg = dict(self._cfg)
        for key, var in self._vars.items():
            v = var.get()
            cfg[key] = "" if v == "（不使用）" else v
        if not cfg.get("key_col") or not cfg.get("hq_no_col"):
            messagebox.showwarning("提示", "匹配关键列和HQ料号列必须选择", parent=self)
            return
        self.result = cfg
        self.destroy()

# ─────────────────────────────────────────────────────────────────────
# 添加/编辑优选库对话框
# ─────────────────────────────────────────────────────────────────────
class AddLibraryDialog(tk.Toplevel):
    def __init__(self, parent, api, log_cb, existing=None):
        super().__init__(parent)
        self.title("添加优选库" if not existing else "编辑优选库")
        self.geometry("800x600")
        self.grab_set()
        self.result    = None
        self._api      = api
        self._log_cb   = log_cb
        self._cfgs     = []
        self._existing = existing
        self._build()
        if existing:
            self.v_name.set(existing["name"])
            self.v_token.set(existing["token"])

    def _build(self):
        top = ttk.LabelFrame(self, text="基础信息", padding=10)
        top.pack(fill="x", padx=12, pady=8)
        tk.Label(top, text="库名称：",    width=12, anchor="w").grid(row=0, column=0, sticky="w", pady=3)
        self.v_name = tk.StringVar()
        ttk.Entry(top, textvariable=self.v_name, width=36).grid(row=0, column=1, sticky="w", padx=4)
        tk.Label(top, text="表格Token：", width=12, anchor="w").grid(row=1, column=0, sticky="w", pady=3)
        self.v_token = tk.StringVar()
        ttk.Entry(top, textvariable=self.v_token, width=46).grid(row=1, column=1, sticky="w", padx=4)
        self.btn_fetch = ttk.Button(top, text="获取 Sheet 列表", command=self._do_fetch)
        self.btn_fetch.grid(row=1, column=2, padx=8)
        self.lbl_fetch = tk.Label(top, text="", fg="#555")
        self.lbl_fetch.grid(row=1, column=3)

        mid = ttk.LabelFrame(self, text="Sheet 列表（双击行可配置该 Sheet 的列映射）", padding=8)
        mid.pack(fill="both", expand=True, padx=12, pady=4)
        cols = ("title","key_col","hq_no_col","brand_col","enabled")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", height=8)
        for c, w, lbl in [("title",130,"Sheet名称"),("key_col",140,"关键列"),
                           ("hq_no_col",120,"HQ料号列"),("brand_col",110,"制造商列"),
                           ("enabled",60,"启用")]:
            self.tree.heading(c, text=lbl)
            self.tree.column(c, width=w, anchor="w")
        sb = ttk.Scrollbar(mid, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")
        self.tree.bind("<Double-1>", self._on_dbl)
        tk.Label(self,
                 text="提示：双击 Sheet 配置列映射。未配置关键列的 Sheet 同步时跳过。",
                 fg="#888").pack(anchor="w", padx=14)
        btn_row = tk.Frame(self); btn_row.pack(pady=10)
        self.btn_ok = ttk.Button(btn_row, text="确定", command=self._ok,
                                 state="disabled", width=12)
        self.btn_ok.pack(side="left", padx=6)
        ttk.Button(btn_row, text="取消", command=self.destroy, width=10).pack(side="left")

    def _do_fetch(self):
        token = self.v_token.get().strip()
        if not token:
            messagebox.showwarning("提示", "请先填写表格Token", parent=self); return
        self.btn_fetch.configure(state="disabled")
        self.lbl_fetch.configure(text="获取中...", fg="#2d6cdf")
        threading.Thread(target=self._fetch_bg, args=(token,), daemon=True).start()

    def _fetch_bg(self, token):
        try:
            sheets = self._api.get_sheets(token)
            self.after(0, lambda: self._on_fetch_ok(sheets))
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: self._on_fetch_fail(msg))

    def _on_fetch_ok(self, sheets):
        existing_map = {c["sheet_id"]: c for c in (self._existing or {}).get("sheets", [])}
        self._cfgs = []
        for s in sheets:
            sid = s["sheetId"]
            cfg = dict(existing_map.get(sid, {
                "sheet_id": sid, "enabled": True,
                "key_col": "", "hq_no_col": "", "brand_col": "",
                "spec_col": "", "desc_col": ""}))
            cfg["title"]     = s["title"]
            cfg["sheet_id"]  = sid
            cfg["row_count"] = s.get("rowCount", cfg.get("row_count", 5000))
            self._cfgs.append(cfg)
        self.btn_fetch.configure(state="normal")
        self.lbl_fetch.configure(text=f"✓ {len(sheets)} 个Sheet", fg="#2a8a2a")
        self._refresh_tree()
        self.btn_ok.configure(state="normal")

    def _on_fetch_fail(self, msg):
        self.btn_fetch.configure(state="normal")
        self.lbl_fetch.configure(text="失败", fg="red")
        messagebox.showerror("错误", msg, parent=self)

    def _refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for cfg in self._cfgs:
            self.tree.insert("", "end", iid=cfg["sheet_id"], values=(
                cfg["title"],
                cfg.get("key_col") or "（未配置）",
                cfg.get("hq_no_col") or "（未配置）",
                cfg.get("brand_col") or "",
                "✓" if cfg.get("enabled", True) else "✗"))

    def _on_dbl(self, _event):
        item = self.tree.focus()
        if not item: return
        cfg = next((c for c in self._cfgs if c["sheet_id"] == item), None)
        if not cfg: return
        threading.Thread(target=self._open_mapping, args=(cfg,), daemon=True).start()

    def _open_mapping(self, cfg):
        token = self.v_token.get().strip()
        try:
            values  = self._api.read_sheet(token, cfg["sheet_id"], 3)
            headers = [_cell_str(v) for v in (values[0] if values else [])]
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: messagebox.showerror("错误", f"读取表头失败: {msg}", parent=self))
            return
        lib_name = self.v_name.get() or "（未命名）"
        self.after(0, lambda: self._show_mapping(cfg, headers, lib_name))

    def _show_mapping(self, cfg, headers, lib_name):
        dlg = MappingDialog(self, cfg, headers, lib_name)
        self.wait_window(dlg)
        if dlg.result:
            for i, c in enumerate(self._cfgs):
                if c["sheet_id"] == cfg["sheet_id"]:
                    self._cfgs[i] = dlg.result; break
            self._refresh_tree()

    def _ok(self):
        name  = self.v_name.get().strip()
        token = self.v_token.get().strip()
        if not name:
            messagebox.showwarning("提示", "请填写库名称", parent=self); return
        if not token:
            messagebox.showwarning("提示", "请填写表格Token", parent=self); return
        self.result = {"name": name, "token": token, "sheets": self._cfgs}
        self.destroy()

# ─────────────────────────────────────────────────────────────────────
# 主界面
# ─────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"飞书优选库匹配工具 {VERSION}")
        self.geometry("940x720")
        self.resizable(True, True)
        self.cfg   = ConfigManager()
        self.cache = CacheManager()
        self.api   = FeishuAPI(self.cfg.base_url, self.cfg.origin, self.cfg.user_id)
        self.bom_wb = None
        self.bom_ws = None
        self._build_ui()
        self._refresh_lib_tree()
        self._update_cache_label()

    def _build_ui(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=10, pady=8)
        for name, attr in [("库管理","_tab_lib"),("同步数据","_tab_sync"),
                            ("BOM匹配","_tab_bom"),("日志","_tab_log")]:
            f = ttk.Frame(self.nb)
            setattr(self, attr, f)
            self.nb.add(f, text=f"  {name}  ")
        self._build_lib()
        self._build_sync()
        self._build_bom()
        self._build_log()

    # ── Tab: 库管理 ──────────────────────────────────────────────────
    def _build_lib(self):
        p = self._tab_lib
        bar = tk.Frame(p); bar.pack(fill="x", padx=10, pady=6)
        ttk.Button(bar, text="+ 添加优选库", command=self._add_lib).pack(side="left", padx=3)
        ttk.Button(bar, text="编辑选中",     command=self._edit_lib).pack(side="left", padx=3)
        ttk.Button(bar, text="删除选中",     command=self._del_lib).pack(side="left", padx=3)
        ttk.Button(bar, text="刷新",         command=self._refresh_lib_tree).pack(side="right", padx=3)
        cols = ("name","token","sheets","last_sync","records")
        self.lib_tree = ttk.Treeview(p, columns=cols, show="headings", height=14)
        for c, w, lbl in [("name",200,"库名称"),("token",230,"Token（部分）"),
                           ("sheets",80,"Sheet数"),("last_sync",165,"最后同步"),
                           ("records",100,"缓存行数")]:
            self.lib_tree.heading(c, text=lbl)
            self.lib_tree.column(c, width=w, anchor="w")
        sb = ttk.Scrollbar(p, command=self.lib_tree.yview)
        self.lib_tree.configure(yscrollcommand=sb.set)
        self.lib_tree.pack(side="left", fill="both", expand=True, padx=(10,0), pady=4)
        sb.pack(side="left", fill="y", pady=4, padx=(0,10))

    # ── Tab: 同步数据 ────────────────────────────────────────────────
    def _build_sync(self):
        p = self._tab_sync
        top = tk.Frame(p); top.pack(fill="x", padx=12, pady=8)
        self.lbl_cache = tk.Label(top, text="", font=("Arial", 11))
        self.lbl_cache.pack(side="left")
        btn_row = tk.Frame(p); btn_row.pack(fill="x", padx=12, pady=4)
        ttk.Button(btn_row, text="全部重新同步", command=self._sync_all).pack(side="left", padx=4)
        ttk.Button(btn_row, text="同步选中库",   command=self._sync_sel).pack(side="left", padx=4)
        self.prog = ttk.Progressbar(p, mode="determinate")
        self.prog.pack(fill="x", padx=12, pady=6)
        self.lbl_prog = tk.Label(p, text="就绪", fg="#555")
        self.lbl_prog.pack(anchor="w", padx=14)
        f = ttk.LabelFrame(p, text="各库同步状态", padding=8)
        f.pack(fill="both", expand=True, padx=12, pady=4)
        cols = ("name","sheets","records","last_sync","status")
        self.sync_tree = ttk.Treeview(f, columns=cols, show="headings", height=10)
        for c, w, lbl in [("name",210,"库名称"),("sheets",80,"Sheet数"),
                           ("records",90,"缓存行数"),("last_sync",165,"最后同步"),
                           ("status",100,"状态")]:
            self.sync_tree.heading(c, text=lbl)
            self.sync_tree.column(c, width=w, anchor="w")
        sb2 = ttk.Scrollbar(f, command=self.sync_tree.yview)
        self.sync_tree.configure(yscrollcommand=sb2.set)
        self.sync_tree.pack(side="left", fill="both", expand=True)
        sb2.pack(side="left", fill="y")

    # ── Tab: BOM匹配 ─────────────────────────────────────────────────
    def _build_bom(self):
        p = self._tab_bom
        f1 = ttk.LabelFrame(p, text="本地 BOM 文件", padding=10)
        f1.pack(fill="x", padx=12, pady=6)
        tk.Label(f1, text="文件：").grid(row=0, column=0, sticky="w")
        self.v_bom_path = tk.StringVar()
        ttk.Entry(f1, textvariable=self.v_bom_path, width=54).grid(row=0, column=1, padx=6)
        ttk.Button(f1, text="浏览...", command=self._browse_bom).grid(row=0, column=2)
        tk.Label(f1, text="Sheet：").grid(row=1, column=0, sticky="w", pady=4)
        self.v_bom_sheet = tk.StringVar()
        self.cb_bom_sheet = ttk.Combobox(f1, textvariable=self.v_bom_sheet,
                                         width=24, state="readonly")
        self.cb_bom_sheet.grid(row=1, column=1, sticky="w", padx=6)
        self.cb_bom_sheet.bind("<<ComboboxSelected>>", lambda _: self._load_bom_sheet())
        tk.Label(f1, text="  表头行：").grid(row=1, column=2, sticky="w")
        self.v_bom_hdr = tk.IntVar(value=1)
        ttk.Spinbox(f1, from_=1, to=20, textvariable=self.v_bom_hdr, width=5).grid(row=1, column=3)
        ttk.Button(f1, text="刷新", command=self._load_bom_sheet).grid(row=1, column=4, padx=4)

        f2 = ttk.LabelFrame(p, text="匹配设置", padding=10)
        f2.pack(fill="x", padx=12, pady=4)
        tk.Label(f2, text="BOM关键列：", width=12, anchor="w").grid(row=0, column=0, sticky="w")
        self.v_bom_key = tk.StringVar()
        self.cb_bom_key = ttk.Combobox(f2, textvariable=self.v_bom_key,
                                       width=28, state="readonly")
        self.cb_bom_key.grid(row=0, column=1, padx=6, sticky="w")
        tk.Label(f2, text="（BOM中用于全库匹配的列，如「厂家型号」）",
                 fg="#666").grid(row=0, column=2, sticky="w")
        tk.Label(f2, text="输出内容：", width=12, anchor="w").grid(row=1, column=0, sticky="w", pady=4)
        chk = tk.Frame(f2); chk.grid(row=1, column=1, columnspan=2, sticky="w")
        self.v_out_hq    = tk.BooleanVar(value=True)
        self.v_out_brand = tk.BooleanVar(value=True)
        self.v_out_spec  = tk.BooleanVar(value=True)
        self.v_out_desc  = tk.BooleanVar(value=True)
        self.v_out_src   = tk.BooleanVar(value=True)
        for text, var in [("HQ料号",self.v_out_hq),("HQ制造商",self.v_out_brand),
                          ("HQ规格型号",self.v_out_spec),("HQ描述",self.v_out_desc),
                          ("来源库/Sheet",self.v_out_src)]:
            ttk.Checkbutton(chk, text=text, variable=var).pack(side="left", padx=4)

        f3 = ttk.LabelFrame(p, text="输出文件", padding=10)
        f3.pack(fill="x", padx=12, pady=4)
        self.v_out_path = tk.StringVar(value="BOM匹配结果.xlsx")
        ttk.Entry(f3, textvariable=self.v_out_path, width=54).pack(side="left")
        ttk.Button(f3, text="另存为...", command=self._browse_out).pack(side="left", padx=6)

        self.btn_match = tk.Button(p, text="开始全库匹配", font=("Arial", 13, "bold"),
                                   bg="#2d6cdf", fg="white", relief="flat",
                                   padx=20, pady=10, command=self._do_match)
        self.btn_match.pack(pady=10)
        self.lbl_match = tk.Label(p, text="", font=("Arial", 10))
        self.lbl_match.pack()
        self.lbl_warn = tk.Label(p, text="", fg="orange", font=("Arial", 9))
        self.lbl_warn.pack()

    # ── Tab: 日志 ────────────────────────────────────────────────────
    def _build_log(self):
        p = self._tab_log
        self.log_box = scrolledtext.ScrolledText(
            p, font=("Consolas", 9), state="disabled",
            bg="#1e1e1e", fg="#d4d4d4", relief="flat")
        self.log_box.pack(fill="both", expand=True, padx=8, pady=8)
        ttk.Button(p, text="清空日志",
                   command=lambda: (self.log_box.configure(state="normal"),
                                    self.log_box.delete("1.0", "end"),
                                    self.log_box.configure(state="disabled"))
                   ).pack(anchor="e", padx=8, pady=4)

    # ── 库管理 ────────────────────────────────────────────────────────
    def _refresh_lib_tree(self):
        self.lib_tree.delete(*self.lib_tree.get_children())
        stats = {r[0]: r[2] for r in self.cache.stats()}
        for lib in self.cfg.libraries:
            n_en  = sum(1 for s in lib.get("sheets", []) if s.get("enabled", True))
            n_all = len(lib.get("sheets", []))
            self.lib_tree.insert("", "end", iid=lib["id"], values=(
                lib["name"], lib["token"][:28] + "...",
                f"{n_en}/{n_all}",
                lib.get("last_sync") or "从未同步",
                stats.get(lib["id"], 0)))
        self._refresh_sync_tree()

    def _refresh_sync_tree(self):
        self.sync_tree.delete(*self.sync_tree.get_children())
        stats = {r[0]: r[2] for r in self.cache.stats()}
        for lib in self.cfg.libraries:
            n_en  = sum(1 for s in lib.get("sheets", []) if s.get("enabled", True))
            n_all = len(lib.get("sheets", []))
            self.sync_tree.insert("", "end", iid=lib["id"], values=(
                lib["name"], f"{n_en}/{n_all}",
                stats.get(lib["id"], 0),
                lib.get("last_sync") or "从未同步",
                "✓ 已同步" if lib.get("last_sync") else "○ 未同步"))

    def _add_lib(self):
        dlg = AddLibraryDialog(self, self.api, self._log)
        self.wait_window(dlg)
        if dlg.result:
            lib = self.cfg.add_library(dlg.result["name"], dlg.result["token"],
                                       dlg.result["sheets"])
            self._log(f"已添加库：{lib['name']}")
            self._refresh_lib_tree()
            messagebox.showinfo("成功",
                f"库「{lib['name']}」已添加！\n请切换到「同步数据」页面进行首次同步。")

    def _edit_lib(self):
        sel = self.lib_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个库"); return
        lib = self.cfg.get_library(sel[0])
        if not lib: return
        dlg = AddLibraryDialog(self, self.api, self._log, existing=lib)
        self.wait_window(dlg)
        if dlg.result:
            self.cfg.update_library(lib["id"], name=dlg.result["name"],
                                    token=dlg.result["token"],
                                    sheets=dlg.result["sheets"])
            self._log(f"已更新库：{dlg.result['name']}")
            self._refresh_lib_tree()

    def _del_lib(self):
        sel = self.lib_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个库"); return
        lib = self.cfg.get_library(sel[0])
        if not lib: return
        if not messagebox.askyesno("确认", f"删除库「{lib['name']}」？\n本地缓存数据也会删除。"):
            return
        self.cache.clear_library(lib["id"])
        self.cfg.delete_library(lib["id"])
        self._log(f"已删除库：{lib['name']}")
        self._refresh_lib_tree()
        self._update_cache_label()

    # ── 同步 ─────────────────────────────────────────────────────────
    def _update_cache_label(self):
        n = self.cache.count()
        self.lbl_cache.configure(text=f"本地缓存：共 {n} 条物料记录")
        if hasattr(self, "lbl_warn"):
            self.lbl_warn.configure(
                text="⚠ 缓存为空，请先同步优选库" if n == 0
                else f"已缓存 {n} 条记录，可直接匹配")

    def _sync_all(self):
        if not self.cfg.libraries:
            messagebox.showwarning("提示", "还没有添加任何优选库"); return
        threading.Thread(target=self._sync_bg, args=(self.cfg.libraries,), daemon=True).start()

    def _sync_sel(self):
        sel = self.sync_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择要同步的库"); return
        libs = [self.cfg.get_library(s) for s in sel]
        libs = [l for l in libs if l]
        threading.Thread(target=self._sync_bg, args=(libs,), daemon=True).start()

    def _sync_bg(self, libs):
        self.after(0, lambda: self.prog.configure(maximum=len(libs), value=0))
        for i, lib in enumerate(libs):
            name = lib["name"]
            self.after(0, lambda n=name: self.lbl_prog.configure(
                text=f"同步中：{n}...", fg="#2d6cdf"))
            try:
                sync_library(self.api, self.cache, lib, self._log)
                self.cfg.update_library(lib["id"],
                    last_sync=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            except Exception as e:
                self._log(f"[{name}] 同步失败: {e}")
            self.after(0, lambda v=i+1: self.prog.configure(value=v))
        self.after(0, lambda: self.lbl_prog.configure(text="✓ 同步完成", fg="#2a8a2a"))
        self.after(0, self._refresh_lib_tree)
        self.after(0, self._update_cache_label)
        self._log("全部同步完成")

    # ── BOM 匹配 ─────────────────────────────────────────────────────
    def _browse_bom(self):
        path = filedialog.askopenfilename(
            title="选择BOM文件",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("所有文件", "*.*")])
        if not path: return
        self.v_bom_path.set(path)
        self.v_out_path.set(os.path.join(os.path.dirname(path), "BOM匹配结果.xlsx"))
        threading.Thread(target=self._load_bom_bg, args=(path,), daemon=True).start()

    def _load_bom_bg(self, path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            self.bom_wb = wb
            self.after(0, self._on_bom_loaded)
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: messagebox.showerror("错误", f"无法打开文件: {msg}"))

    def _on_bom_loaded(self):
        self.cb_bom_sheet["values"] = self.bom_wb.sheetnames
        self.v_bom_sheet.set(self.bom_wb.sheetnames[0])
        self._load_bom_sheet()

    def _load_bom_sheet(self):
        if not self.bom_wb: return
        name = self.v_bom_sheet.get()
        if name not in self.bom_wb.sheetnames: return
        self.bom_ws = self.bom_wb[name]
        hr = self.v_bom_hdr.get()
        headers = [_cell_str(self.bom_ws.cell(row=hr, column=ci).value)
                   for ci in range(1, self.bom_ws.max_column + 1)
                   if self.bom_ws.cell(row=hr, column=ci).value]
        self.cb_bom_key["values"] = headers
        default = next((h for h in headers if "型号" in h or "part" in h.lower()),
                       headers[0] if headers else "")
        if not self.v_bom_key.get() and default:
            self.v_bom_key.set(default)
        self._log(f"BOM表头({len(headers)}列): {', '.join(headers[:8])}")

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if path: self.v_out_path.set(path)

    def _do_match(self):
        if not self.bom_ws:
            messagebox.showerror("错误", "请先选择BOM文件"); return
        if self.cache.count() == 0:
            messagebox.showerror("错误", "缓存为空，请先同步优选库"); return
        key_col = self.v_bom_key.get()
        if not key_col:
            messagebox.showerror("错误", "请选择BOM关键列"); return
        output_cols = []
        if self.v_out_hq.get():    output_cols.append(("HQ料号",    "hq_no"))
        if self.v_out_brand.get(): output_cols.append(("HQ制造商",  "brand"))
        if self.v_out_spec.get():  output_cols.append(("HQ规格型号","spec"))
        if self.v_out_desc.get():  output_cols.append(("HQ描述",    "description"))
        out_path = self.v_out_path.get().strip() or "BOM匹配结果.xlsx"
        self.btn_match.configure(state="disabled")
        self.lbl_match.configure(text="全库匹配中...", fg="#2d6cdf")
        self.nb.select(3)
        threading.Thread(
            target=self._match_bg,
            args=(self.v_bom_hdr.get(), key_col, output_cols,
                  self.v_out_src.get(), out_path),
            daemon=True).start()

    def _match_bg(self, hr, key_col, output_cols, show_src, out_path):
        try:
            total, matched, unmatched, out = match_bom(
                self.bom_ws, hr, key_col, self.cache,
                output_cols, show_src, out_path, self._log)
            self.after(0, lambda: self.lbl_match.configure(
                text=f"✓ 完成！{matched} 个匹配成功，{unmatched} 个未匹配",
                fg="#2a8a2a"))
            try:
                folder = os.path.dirname(os.path.abspath(out))
                if sys.platform == "win32": os.startfile(folder)
                elif sys.platform == "darwin": subprocess.Popen(["open", folder])
            except Exception:
                pass
            abs_out = os.path.abspath(out)
            self.after(0, lambda: messagebox.showinfo("完成",
                f"匹配完成！\n\n共 {total} 行\n匹配成功：{matched} 个物料\n"
                f"未匹配：{unmatched} 个物料\n\n{abs_out}"))
        except Exception as e:
            import traceback
            msg, tb = str(e), traceback.format_exc()
            self._log(f"错误: {msg}\n{tb}")
            self.after(0, lambda: self.lbl_match.configure(text="出错，请查看日志", fg="red"))
            self.after(0, lambda: messagebox.showerror("错误", msg))
        finally:
            self.after(0, lambda: self.btn_match.configure(state="normal"))

    # ── 日志 ─────────────────────────────────────────────────────────
    def _log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        def _w():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", f"[{ts}] {msg}\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _w)

    def on_close(self):
        self.cache.close()
        self.destroy()


if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()
