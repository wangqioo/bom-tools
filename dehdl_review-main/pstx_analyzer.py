# -*- coding: utf-8 -*-
"""
PSTX 原理图分析工具 v1.1
解析 Cadence Packager-XL 导出的 pstxprt.dat / pstxnet.dat

功能：BOM 管理 / 网络拓扑 / DRC / 电容降额 / 元件查询 / Excel 导出

依赖：pip install openpyxl
运行：python pstx_analyzer.py
"""

import sys
import subprocess

try:
    import openpyxl
except ImportError:
    print("未检测到 openpyxl，正在自动安装...")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

import os
import ntpath
import re
from collections import Counter, defaultdict, deque
from pathlib import Path, PureWindowsPath
from typing import Dict, List, Optional, Tuple

import pstx_page_logic as page_logic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════
# 一、PST 文件解析
# ══════════════════════════════════════════════════════════

def _join_continuations(text: str) -> str:
    normalized = str(text or '').replace('\r\n', '\n').replace('\r', '\n')
    lines = normalized.split('\n')
    result, buf = [], ''
    for line in lines:
        stripped = line.rstrip()
        if stripped.endswith('~'):
            buf += stripped[:-1]
        else:
            buf += line
            result.append(buf)
            buf = ''
    if buf:
        result.append(buf)
    return '\n'.join(result)


def _extract_attrs(text: str) -> Dict[str, str]:
    attrs = {}
    for m in re.finditer(r"\b([A-Z][A-Z0-9_]*)\s*=\s*'([^']*)'", text):
        key, val = m.group(1), m.group(2)
        if key not in attrs:
            attrs[key] = val
    return attrs


# ══════════════════════════════════════════════════════════
# 一、页码解析入口
#   具体 page.map / page*.csv / module_order 逻辑集中在 pstx_page_logic.py，
#   这里仅保留历史测试和业务代码依赖的轻量兼容包装。
# ══════════════════════════════════════════════════════════


def _normalize_page_label(page_label: str) -> str:
    return page_logic.normalize_page_label(page_label)


def _extract_top_level_logical_page(path_text: str) -> str:
    return page_logic.extract_top_level_page(path_text)


def _extract_section_paths(block_text: str) -> List[Dict[str, str]]:
    return page_logic.extract_section_paths(block_text)


def _select_component_page_source(block_text: str, attrs: Dict[str, str]) -> Tuple[str, str]:
    sources = page_logic.select_component_page_sources(block_text, attrs)
    return sources.get('logical_path_raw', ''), sources.get('logical_path_source', 'none')


def _extract_page_number_from_text(text: str) -> str:
    return page_logic._extract_page_number_from_text(text)


def _read_page_number_from_csv(csv_path: Path) -> str:
    return page_logic._read_page_number_from_csv(Path(csv_path))


def _build_page_csv_index(project_root: str) -> Dict[str, object]:
    return page_logic.build_page_csv_index(project_root)


def analyze_page_mappings(page_index: Optional[Dict[str, object]]) -> Dict[str, object]:
    return page_logic.build_page_mapping_rows(None, page_index)


def _prepare_page_resolution(project_root: str) -> Dict[str, object]:
    page_csv_index = page_logic.build_page_csv_index(project_root) if project_root else None
    page_map_index = page_logic.build_page_map_index(project_root) if project_root else None
    module_order_index = page_logic.build_module_order_index(project_root) if project_root else None
    page_mapping = page_logic.build_page_mapping_rows(page_map_index, page_csv_index)

    warnings: List[str] = []
    for index in [page_csv_index, page_map_index, module_order_index]:
        if index:
            warnings.extend(index.get('warnings', []))
    warnings.extend(page_mapping.get('warnings', []))
    return {
        'page_csv_index': page_csv_index,
        'page_map_index': page_map_index,
        'module_order_index': module_order_index,
        'page_mapping': page_mapping,
        'warnings': warnings,
    }


def _apply_component_pages(components: Dict[str, dict],
                           page_context: Optional[Dict[str, object]]) -> None:
    context = page_context or {}
    page_map_index = context.get('page_map_index')
    page_csv_index = context.get('page_csv_index')
    module_order_index = context.get('module_order_index')
    for comp in components.values():
        page_info = page_logic.resolve_component_page_info(
            comp,
            page_map_index=page_map_index,
            page_csv_index=page_csv_index,
            module_order_index=module_order_index,
        )
        display_page = page_info.get('page_real', '')
        comp['page'] = display_page
        comp['page_logical'] = page_info.get('page_logical', '')
        comp['page_raw'] = page_info.get('page_logical', '')
        comp['page_real'] = display_page
        comp['page_submodule_real'] = page_info.get('page_submodule_real', '')
        comp['page_submodule_mapped'] = page_info.get('page_submodule_mapped', '')
        comp['page_context'] = page_info.get('page_context', '')
        comp['page_context_real'] = page_info.get('page_context_real', '')
        comp['page_source'] = page_info.get('page_logical_source', '')
        comp['page_real_source'] = page_info.get('page_real_source', '')
        comp['page_validation_status'] = page_info.get('page_validation_status', '')
        comp['page_mapping_ok'] = page_info.get('page_mapping_ok', '')
        comp['page_mapping_note'] = page_info.get('page_validation_note', '')
        comp['page_validation_note'] = page_info.get('page_validation_note', '')
        comp['page_map_real'] = page_info.get('page_map_real', '')
        comp['page_map_state'] = page_info.get('page_map_state', '')
        comp['page_csv_real'] = page_info.get('page_csv_real', '')
        comp['page_csv_state'] = page_info.get('page_csv_state', '')
        comp['module_order_key'] = page_info.get('module_order_key', '')
        comp['module_order_state'] = page_info.get('module_order_state', '')
        comp['module_order_local_page'] = page_info.get('module_order_local_page', '')
        comp['module_order_start_page'] = page_info.get('module_order_start_page', '')
        comp['module_order_page_count'] = page_info.get('module_order_page_count', '')
        comp['page_submodule_mapping_note'] = page_info.get('page_submodule_mapping_note', '')


def resolve_component_pages(components: Dict[str, dict], project_root: str = '') -> List[str]:
    page_context = _prepare_page_resolution(project_root)
    _apply_component_pages(components, page_context)
    return list(page_context.get('warnings', []))


def _component_logical_page(comp: Dict) -> str:
    return _normalize_page_label(
        comp.get('page_logical', '')
        or comp.get('page_raw', '')
        or comp.get('page', '')
    )


def _component_display_page(comp: Dict) -> str:
    return _normalize_page_label(comp.get('page_real', '') or comp.get('page', ''))


def _component_submodule_mapped_page(comp: Dict) -> str:
    return _normalize_page_label(comp.get('page_submodule_mapped', ''))


def _component_page_fields(comp: Dict) -> Dict[str, str]:
    return {
        '页面': _component_display_page(comp),
        '子模块映射主模块真实页': _component_submodule_mapped_page(comp),
    }


def _looks_like_windows_path(path_text: str) -> bool:
    return bool(re.match(r'^[A-Za-z]:[\\/]', path_text)) or '\\' in path_text


def _infer_project_root_from_data_paths(*paths: str) -> str:
    raw_paths = [str(raw_path or '').strip().strip('"') for raw_path in paths if str(raw_path or '').strip()]
    if not raw_paths:
        return ''

    windows_mode = any(_looks_like_windows_path(path_text) for path_text in raw_paths)
    candidates = []
    for path_text in raw_paths:
        try:
            if windows_mode:
                candidates.append(PureWindowsPath(path_text))
            else:
                candidates.append(Path(path_text).expanduser().resolve())
        except OSError:
            continue
    if not candidates:
        return ''

    for path in candidates:
        parent = path.parent
        if parent.name.lower() == 'packaged':
            return str(parent.parent)

    parent_strings = [str(path.parent) for path in candidates]
    try:
        if windows_mode:
            common_parent = PureWindowsPath(ntpath.commonpath(parent_strings))
        else:
            common_parent = Path(os.path.commonpath(parent_strings))
    except ValueError:
        common_parent = candidates[0].parent
    if common_parent.name.lower() == 'packaged':
        return str(common_parent.parent)
    return str(common_parent)


def _get_comp_type(refdes: str, part_name: str) -> str:
    pn = part_name.lower()
    type_rules = [
        (['cap_pol'],                           'CAP_POL'),
        (['cap_hdl', 'cap_'],                   'CAP'),
        (['res_hdl', 'res_'],                   'RES'),
        (['ind_hdl', 'ind_', 'ferrite', 'fer_hdl', 'fb_hdl'], 'IND'),
        (['osc_', 'crystal', 'xtal'],           'XTAL'),
        (['conn_', 'connector'],                'CONN'),
        (['led_'],                              'LED'),
        (['diode', '_d_hdl'],                   'DIODE'),
        (['mosfet', 'mos_', 'nmos', 'pmos', 'nfet', 'pfet'], 'FET'),
        (['bjt', 'transistor', 'npn', 'pnp'],  'BJT'),
        (['fuse'],                              'FUSE'),
        (['sw_hdl', 'switch'],                  'SWITCH'),
        (['testpoint', 'test_point', 'tp_hdl'], 'TESTPOINT'),
        (['transformer', 'xfmr'],              'TRANSFORMER'),
    ]
    for keywords, ctype in type_rules:
        if any(k in pn for k in keywords):
            return ctype
    prefix = (re.match(r'[A-Za-z]+', refdes) or re.match(r'', '')).group(0).upper()
    prefix_map = {
        'C': 'CAP', 'PC': 'CAP', 'R': 'RES', 'L': 'IND', 'FB': 'IND',
        'U': 'IC', 'J': 'CONN', 'P': 'CONN', 'CN': 'CONN', 'Q': 'FET',
        'D': 'DIODE', 'LED': 'LED', 'Y': 'XTAL', 'F': 'FUSE',
        'SW': 'SWITCH', 'TP': 'TESTPOINT', 'T': 'TRANSFORMER',
    }
    return prefix_map.get(prefix, 'IC')


def _split_named_blocks(text: str, marker: str) -> List[str]:
    return re.split(rf'(?:^|\n){re.escape(marker)}\n', text)[1:]


def _split_net_tokens(net_name: str) -> List[str]:
    return [tok for tok in re.split(r'[_./-]+', (net_name or '').upper()) if tok]


def _first_net_token(net_name: str) -> str:
    tokens = _split_net_tokens(net_name)
    return tokens[0] if tokens else (net_name or '').upper()


def _matches_prefix_with_boundary(name: str, prefix: str) -> bool:
    if not prefix:
        return False
    name = (name or '').upper()
    prefix = prefix.upper()
    if not name.startswith(prefix):
        return False
    return len(name) == len(prefix) or name[len(prefix)] in '_./-'


def _parse_voltage_token(token: str) -> Optional[float]:
    m = re.fullmatch(r'P?(\d+)V(\d*)', token.upper())
    if not m:
        return None
    int_part, frac_part = m.groups()
    return float(f'{int_part}.{frac_part}') if frac_part else float(int_part)


_POWER_TOKEN_RE = re.compile(
    r'(?:VCC|VDD|VBAT|VCORE|VCCIO|PVDD|PVCC|AVDD|DVDD|VBUS)[A-Z0-9]*',
    re.IGNORECASE,
)
_GROUND_TOKEN_RE = re.compile(
    r'(?:GND|AGND|SGND|PGND|DGND|VSS[A-Z0-9]*)',
    re.IGNORECASE,
)


def _token_is_power(token: str) -> bool:
    return _parse_voltage_token(token) is not None or bool(_POWER_TOKEN_RE.fullmatch(token))


def _token_is_ground(token: str) -> bool:
    return bool(_GROUND_TOKEN_RE.fullmatch(token))


_DIFF_SUFFIX_PAIRS = [
    ('_P', '_N'),
    ('_DP', '_DN'),
    ('.P', '.N'),
    ('_TXPLUS', '_TXMINUS'),
    ('_RXPLUS', '_RXMINUS'),
]


def _get_diff_net_info(net_name: str, upper_name_map: Dict[str, str]) -> Optional[Dict[str, str]]:
    upper_name = (net_name or '').upper()
    for pos_suffix, neg_suffix in _DIFF_SUFFIX_PAIRS:
        pos_upper = pos_suffix.upper()
        neg_upper = neg_suffix.upper()
        if upper_name.endswith(pos_upper):
            partner = upper_name_map.get(upper_name[:-len(pos_upper)] + neg_upper)
            if partner:
                return {
                    'base': net_name[:-len(pos_suffix)],
                    'polarity': 'P',
                    'partner': partner,
                }
        elif upper_name.endswith(neg_upper):
            partner = upper_name_map.get(upper_name[:-len(neg_upper)] + pos_upper)
            if partner:
                return {
                    'base': net_name[:-len(neg_suffix)],
                    'polarity': 'N',
                    'partner': partner,
                }
    return None


def _collect_diff_pairs(nets: Dict) -> Dict[str, dict]:
    diff_pairs: Dict[str, dict] = {}
    upper_name_map = {name.upper(): name for name in nets}
    for net_name in nets:
        info = _get_diff_net_info(net_name, upper_name_map)
        if not info:
            continue
        base = info['base']
        if info['polarity'] == 'P':
            diff_pairs[base] = {'P': net_name, 'N': info['partner']}
        elif base not in diff_pairs:
            diff_pairs[base] = {'P': info['partner'], 'N': net_name}
    return diff_pairs


def _collect_component_nets(nets: Dict) -> Dict[str, List[str]]:
    comp_nets: Dict[str, List[str]] = defaultdict(list)
    for net_name, nodes in nets.items():
        for node in nodes:
            comp_nets[node['refdes']].append(net_name)
    return comp_nets


def _unique_component_nets(comp_nets: Dict[str, List[str]], refdes: str) -> List[str]:
    return list(dict.fromkeys(comp_nets.get(refdes, [])))


def _find_ac_coupling_candidates(components: Dict,
                                 comp_nets: Dict[str, List[str]],
                                 nets: Dict) -> Dict[str, dict]:
    upper_name_map = {name.upper(): name for name in nets}
    diff_info_map = {
        net_name: info
        for net_name in nets
        if (info := _get_diff_net_info(net_name, upper_name_map))
    }
    cap_pairs: Dict[str, Tuple[str, str]] = {}
    caps_by_pair: Dict[frozenset, List[str]] = defaultdict(list)

    for refdes, comp in components.items():
        if comp.get('comp_type') not in ('CAP', 'CAP_POL'):
            continue
        unique_nets = _unique_component_nets(comp_nets, refdes)
        if len(unique_nets) != 2:
            continue
        net_a, net_b = unique_nets
        if _net_is_power(net_a) or _net_is_power(net_b) or _net_is_gnd(net_a) or _net_is_gnd(net_b):
            continue
        cap_pairs[refdes] = (net_a, net_b)
        caps_by_pair[frozenset((net_a, net_b))].append(refdes)

    candidates: Dict[str, dict] = {}
    for refdes, (net_a, net_b) in cap_pairs.items():
        info_a = diff_info_map.get(net_a)
        info_b = diff_info_map.get(net_b)
        if not info_a or not info_b:
            continue
        if info_a['polarity'] != info_b['polarity']:
            continue
        partner_pair = frozenset((info_a['partner'], info_b['partner']))
        mirror_caps = sorted(
            (cap for cap in caps_by_pair.get(partner_pair, []) if cap != refdes),
            key=_natural_sort_key,
        )
        if not mirror_caps:
            continue
        candidates[refdes] = {
            'nets': (net_a, net_b),
            'mirror_nets': sorted(partner_pair, key=_natural_sort_key),
            'mirror_caps': mirror_caps,
            'polarity': info_a['polarity'],
        }
    return candidates


def _natural_sort_key(value: str):
    parts = re.split(r'(\d+)', str(value or '').upper())
    return [int(p) if p.isdigit() else p for p in parts]


_RESULT_KIND_LABELS = {
    'confirmed': '确定结论',
    'candidate': '候选判断',
    'indeterminate': '无法判断',
}
_SEVERITY_LABELS = {
    'high': '高',
    'medium': '中',
    'low': '低',
}
_CONFIDENCE_LABELS = {
    'high': '高',
    'medium': '中',
    'low': '低',
}


def _meta_fields(result_kind: str, severity: str, confidence: str, reason_code: str) -> Dict[str, str]:
    return {
        '结论类型': _RESULT_KIND_LABELS[result_kind],
        '严重级别': _SEVERITY_LABELS[severity],
        '置信度': _CONFIDENCE_LABELS[confidence],
        '原因代码': reason_code,
    }


def _with_meta(row: Dict[str, str], result_kind: str, severity: str,
               confidence: str, reason_code: str) -> Dict[str, str]:
    merged = dict(row)
    merged.update(_meta_fields(result_kind, severity, confidence, reason_code))
    return merged


def _count_result_kinds(rows: List[dict]) -> Counter:
    counter: Counter = Counter()
    for row in rows:
        kind = row.get('结论类型')
        if kind:
            counter[kind] += 1
    return counter


def _iter_list_rows(mapping: dict, keys: List[str]) -> List[dict]:
    rows: List[dict] = []
    for key in keys:
        value = mapping.get(key, [])
        if isinstance(value, list):
            rows.extend([row for row in value if isinstance(row, dict)])
    return rows




def parse_pstxnet(content: str) -> Dict[str, List[dict]]:
    text = _join_continuations(content)
    nets = {}
    node_re     = re.compile(r'NODE_NAME\s+(\S+)\s+(\S+)')
    pin_name_re = re.compile(r"'([^']+)'\s*:")
    for block in _split_named_blocks(text, 'NET_NAME'):
        m = re.search(r"'([^']+)'", block)
        if not m:
            continue
        net_name = m.group(1)
        nodes = []
        matches = list(node_re.finditer(block))
        for idx, nm in enumerate(matches):
            next_start = matches[idx + 1].start() if idx + 1 < len(matches) else len(block)
            after = block[nm.end():next_start]
            pn_match = pin_name_re.search(after)
            nodes.append({
                'refdes':   nm.group(1),
                'pin':      nm.group(2),
                'pin_name': pn_match.group(1) if pn_match else nm.group(2),
            })
        if nodes:
            nets[net_name] = nodes
    return nets


def parse_all(prt_content: str, net_content: str):
    components = parse_pstxprt(prt_content)
    nets       = parse_pstxnet(net_content)
    comp_nets: Dict[str, Dict[str, str]] = {}
    for net_name, nodes in nets.items():
        for node in nodes:
            rd = node['refdes']
            if rd not in comp_nets:
                comp_nets[rd] = {}
            comp_nets[rd][node['pin']] = net_name
    for refdes, comp in components.items():
        comp['nets'] = comp_nets.get(refdes, {})
    return components, nets, comp_nets


def parse_pstxprt(content: str) -> Dict[str, dict]:
    text = _join_continuations(content)
    components = {}
    for block in _split_named_blocks(text, 'PART_NAME'):
        m = re.match(r"(\S+)\s+'([^']+)'", block.split('\n')[0].strip())
        if not m:
            continue
        refdes, part_name = m.group(1), m.group(2)
        attrs = _extract_attrs(block)
        page_sources = page_logic.select_component_page_sources(block, attrs)
        logical_path_raw = page_sources.get('logical_path_raw', '')
        logical_path_source = page_sources.get('logical_path_source', 'none')
        real_path_raw = page_sources.get('real_path_raw', '')
        real_path_source = page_sources.get('real_path_source', 'none')
        components[refdes] = {
            'refdes': refdes,
            'part_name': part_name,
            'hq_code': attrs.get('HQ_CODE', ''),
            'value': attrs.get('VALUE', ''),
            'package': attrs.get('PACKAGE', ''),
            'material': attrs.get('MATERIAL', ''),
            'tolerance': attrs.get('TOLERANCE', ''),
            'voltage': attrs.get('VOLTAGE', ''),
            'current': attrs.get('CURRENT', ''),
            'power': attrs.get('POWER', ''),
            'bom_option': attrs.get('BOM_OPTION', ''),
            'bom_cost': attrs.get('BOM_COST', ''),
            'room': attrs.get('ROOM', ''),
            'drawing': attrs.get('DRAWING', ''),
            'page_path_raw': logical_path_raw,
            'page_path_source': logical_path_source,
            'page_path_logical_raw': logical_path_raw,
            'page_path_logical_source': logical_path_source,
            'page_path_real_raw': real_path_raw,
            'page_path_real_source': real_path_source,
            'page': '',
            'page_logical': page_logic.extract_top_level_page(logical_path_raw or attrs.get('DRAWING', '')),
            'page_real': '',
            'page_submodule_real': page_logic.extract_submodule_page(real_path_raw),
            'page_submodule_mapped': '',
            'comp_type': _get_comp_type(refdes, part_name),
        }
    return components


def _normalize_bom_option(value: str) -> str:
    return (value or '').strip().upper()


def _display_bom_option(value: str) -> str:
    normalized = _normalize_bom_option(value)
    return normalized or '默认'


def _is_depop_option(bom_option: str) -> bool:
    return _normalize_bom_option(bom_option) in {'DEPOP', 'DNP'}


def _clone_components(components: Dict[str, dict], allowed_nets: Optional[set] = None) -> Dict[str, dict]:
    cloned: Dict[str, dict] = {}
    for refdes, comp in components.items():
        cloned_comp = dict(comp)
        nets_map = dict(comp.get('nets', {}))
        if allowed_nets is not None:
            nets_map = {pin: net for pin, net in nets_map.items() if net in allowed_nets}
        cloned_comp['nets'] = nets_map
        cloned[refdes] = cloned_comp
    return cloned


def _collect_depop_refdes(components: Dict[str, dict]) -> List[str]:
    return sorted(
        [refdes for refdes, comp in components.items() if _is_depop_option(comp.get('bom_option', ''))],
        key=_natural_sort_key,
    )


def _build_analysis_scope(components: Dict[str, dict],
                          nets: Dict[str, List[dict]],
                          *,
                          include_depop: bool) -> Tuple[Dict[str, dict], Dict[str, List[dict]], List[str], List[str]]:
    depop_refdes = _collect_depop_refdes(components)
    if include_depop or not depop_refdes:
        active_nets = {
            net_name: [dict(node) for node in nodes]
            for net_name, nodes in nets.items()
        }
        active_components = _clone_components(components, set(active_nets.keys()))
        return active_components, active_nets, depop_refdes, []

    excluded = set(depop_refdes)
    active_nets: Dict[str, List[dict]] = {}
    for net_name, nodes in nets.items():
        filtered_nodes = [dict(node) for node in nodes if node.get('refdes') not in excluded]
        if filtered_nodes:
            active_nets[net_name] = filtered_nodes

    active_components = _clone_components(
        {refdes: comp for refdes, comp in components.items() if refdes not in excluded},
        set(active_nets.keys()),
    )
    return active_components, active_nets, depop_refdes, depop_refdes[:]


# ══════════════════════════════════════════════════════════
# 二、BOM 分析
# ══════════════════════════════════════════════════════════

COMP_TYPE_CN = {
    'CAP': '电容', 'CAP_POL': '电解/钽电容', 'RES': '电阻',
    'IND': '电感/磁珠', 'IC': 'IC 芯片', 'CONN': '连接器',
    'DIODE': '二极管', 'LED': 'LED', 'FET': 'MOS/FET',
    'BJT': '三极管', 'XTAL': '晶振', 'FUSE': '保险丝',
    'SWITCH': '开关', 'TESTPOINT': '测试点', 'TRANSFORMER': '变压器',
}
_TYPE_ORDER = list(COMP_TYPE_CN.keys())


def build_bom(components: Dict):
    detail_normal, detail_depop = [], []
    for comp in components.values():
        ctype = comp.get('comp_type', '')
        row = {
            '位号':          comp['refdes'],
            '料号':          comp.get('hq_code', ''),
            '描述':          comp.get('part_name', ''),
            '值':            comp.get('value', ''),
            '封装':          comp.get('package', ''),
            '耐压/额定电压': comp.get('voltage', ''),
            '额定功率':      comp.get('power', ''),
            '精度':          comp.get('tolerance', ''),
            '材质':          comp.get('material', ''),
            '类型':          COMP_TYPE_CN.get(ctype, ctype),
            '_ctype':        ctype,
            'ROOM':          comp.get('room', ''),
        }
        row.update(_component_page_fields(comp))
        (detail_depop if _is_depop_option(comp.get('bom_option', '')) else detail_normal).append(row)

    def _merge(detail):
        if not detail:
            return []
        groups = {}
        for row in detail:
            key = (
                ('pn', row['料号']) if row['料号'] else
                ('desc', row['描述'], row['值'], row['封装'], row['耐压/额定电压'],
                 row['额定功率'], row['精度'], row['材质'], row['类型'])
            )
            if key not in groups:
                groups[key] = {
                    '料号': row['料号'], '位号列表': [], '数量': 0,
                    '描述': row['描述'], '值': row['值'], '封装': row['封装'],
                    '耐压': row['耐压/额定电压'], '额定功率': row['额定功率'],
                    '精度': row['精度'], '材质': row['材质'],
                    '类型': row['类型'], '_ctype': row['_ctype'],
                }
            groups[key]['位号列表'].append(row['位号'])
            groups[key]['数量'] += 1
        merged = list(groups.values())
        merged.sort(key=lambda r: (
            _TYPE_ORDER.index(r['_ctype']) if r['_ctype'] in _TYPE_ORDER else 99,
            r['料号'],
            r['描述'],
            r['值'],
            r['封装'],
        ))
        for i, r in enumerate(merged, 1):
            r['序号'] = i
            r['位号列表'] = ', '.join(sorted(r['位号列表'], key=_natural_sort_key))
            del r['_ctype']
        return merged

    def _clean(rows):
        return [{k: v for k, v in r.items() if k != '_ctype'} for r in rows]

    return _clean(detail_normal), _clean(detail_depop), _merge(detail_normal), _merge(detail_depop)


# ══════════════════════════════════════════════════════════
# 三、网络分析
# ══════════════════════════════════════════════════════════

def analyze_networks(nets: Dict, components: Dict) -> dict:
    single_node = {k: v for k, v in nets.items() if len(v) == 1}
    gnd_nets    = {k: v for k, v in nets.items()
                   if _net_is_gnd(k)}
    power_nets  = {k: v for k, v in nets.items()
                   if _net_is_power(k)
                   and k not in gnd_nets}
    diff_pairs = _collect_diff_pairs(nets)
    page_counter: Counter = Counter()
    for comp in components.values():
        page_label = _component_display_page(comp) or 'UNKNOWN'
        page_counter[page_label] += 1
    page_rows = [
        {'页面': page, '元件数': count}
        for page, count in sorted(
            page_counter.items(),
            key=lambda item: (item[0] == 'UNKNOWN', _natural_sort_key(item[0])),
        )
    ]
    power_rows = [
        _with_meta({'网络名': k, '节点数': len(v)}, 'candidate', 'low', 'medium', 'net_name_power_token')
        for k, v in sorted(power_nets.items(), key=lambda x: (-len(x[1]), x[0].upper()))
    ]
    gnd_rows = [
        _with_meta({'网络名': k, '节点数': len(v)}, 'candidate', 'low', 'high', 'net_name_ground_token')
        for k, v in sorted(gnd_nets.items(), key=lambda x: (-len(x[1]), x[0].upper()))
    ]
    diff_rows = [
        _with_meta({'基础名': b, 'P端网络': pr['P'], 'N端网络': pr['N']},
                   'candidate', 'low', 'medium', 'paired_name_suffix')
        for b, pr in sorted(diff_pairs.items(), key=lambda x: x[0].upper())
    ]
    single_rows = [
        _with_meta({'网络名': k, '连接元件': v[0]['refdes'], '引脚': v[0]['pin_name']},
                   'candidate', 'medium', 'low', 'single_node_net')
        for k, v in sorted(single_node.items(), key=lambda x: x[0].upper())
    ]
    return {
        'total': len(nets), 'single_node': single_node,
        'gnd_nets': gnd_nets, 'power_nets': power_nets,
        'diff_pairs': diff_pairs, 'page_counter': page_counter,
        'page_rows': page_rows,
        'power_net_rows': power_rows,
        'gnd_net_rows': gnd_rows,
        'diff_pair_rows': diff_rows,
        'single_node_rows': single_rows,
    }


# ══════════════════════════════════════════════════════════
# 四、DRC 设计检查
# ══════════════════════════════════════════════════════════

_VALID_BOM_OPTIONS = {'', 'DEPOP', 'OPTION', 'MAIN_PLD', 'MAIN', 'ALT', 'DNP'}
_FUZZY_KEYWORDS    = sorted(opt for opt in _VALID_BOM_OPTIONS if opt)


def _edit_distance(a: str, b: str) -> int:
    a, b = a.upper(), b.upper()
    if a == b: return 0
    if not a:  return len(b)
    if not b:  return len(a)
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        prev = dp[:]
        dp[0] = i + 1
        for j, cb in enumerate(b):
            dp[j+1] = min(prev[j] + (0 if ca == cb else 1), dp[j]+1, prev[j+1]+1)
    return dp[len(b)]


def check_drc(components: Dict, nets: Dict, *, option_components_source: Optional[Dict] = None) -> dict:
    missing_hq, missing_val, missing_pkg, tbd_attrs, single_pin, unnamed = [], [], [], [], [], []
    bom_option_components = []
    option_source = option_components_source if option_components_source is not None else components
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype == 'TESTPOINT':
            continue
        page_label = _component_display_page(comp)
        page_fields = _component_page_fields(comp)
        base = {'位号': refdes, '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': page_label}
        base.update(page_fields)
        if not comp.get('hq_code'):
            missing_hq.append(_with_meta(base.copy(), 'confirmed', 'high', 'high', 'missing_hq_code'))
        if not comp.get('value'):
            missing_val.append(_with_meta(base.copy(), 'confirmed', 'high', 'high', 'missing_value'))
        if not comp.get('package'):
            missing_pkg.append(_with_meta(base.copy(), 'confirmed', 'high', 'high', 'missing_package'))
        for attr in ('voltage', 'current', 'power'):
            val = comp.get(attr, '')
            if val and 'TBD' in val.upper():
                tbd_attrs.append(_with_meta({
                    '位号': refdes, '属性': attr.upper(), '当前值': val,
                    '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': page_label
                }, 'confirmed', 'medium', 'high', 'tbd_attribute'))
                tbd_attrs[-1].update(page_fields)
    for net_name, nodes in nets.items():
        if len(nodes) == 1:
            n = nodes[0]
            comp = components.get(n['refdes'], {})
            if comp.get('comp_type') != 'TESTPOINT' and not re.search(r'^UNNAMED_', net_name, re.I):
                single_pin.append(_with_meta({
                    '网络名': net_name, '连接元件': n['refdes'],
                    '引脚': n['pin_name'], '页面': _component_display_page(comp)
                }, 'candidate', 'medium', 'low', 'single_pin_net'))
                single_pin[-1].update(_component_page_fields(comp))
        if re.search(r'^UNNAMED_', net_name, re.I):
            unnamed.append(_with_meta({'网络名': net_name, '节点数': len(nodes)},
                                      'candidate', 'medium', 'high', 'unnamed_net'))
    option_map: Dict[str, List[str]] = defaultdict(list)
    for refdes, comp in option_source.items():
        ctype = comp.get('comp_type', '')
        if ctype == 'TESTPOINT':
            continue
        bom_option = _normalize_bom_option(comp.get('bom_option', ''))
        if bom_option:
            bom_option_components.append({
                '位号': refdes,
                '类型': COMP_TYPE_CN.get(ctype, ctype),
                'BOM_OPTION值': bom_option,
                '是否DEPOP': '是' if _is_depop_option(bom_option) else '否',
                '页面': _component_display_page(comp),
            })
            bom_option_components[-1].update(_component_page_fields(comp))
        option_map[_normalize_bom_option(comp.get('bom_option'))].append(refdes)
    typos = []
    for val, refs in sorted(option_map.items()):
        if val in _VALID_BOM_OPTIONS:
            continue
        min_d   = min(_edit_distance(val, kw) for kw in _FUZZY_KEYWORDS)
        nearest = min(_FUZZY_KEYWORDS, key=lambda kw: _edit_distance(val, kw))
        typos.append(_with_meta({
            '实际填写值': val, '疑似应为': nearest if min_d <= 2 else '未知',
            '编辑距离': min_d, '使用该值的位号': ', '.join(sorted(refs, key=_natural_sort_key)),
            '数量': len(refs), '风险': '❌ 疑似拼错' if min_d <= 2 else '⚠ 未知值'
        }, 'candidate', 'medium', 'medium' if min_d <= 2 else 'low', 'bom_option_typo'))
    return {
        'missing_hq_code': missing_hq, 'missing_value': missing_val,
        'missing_package': missing_pkg, 'tbd_attrs': tbd_attrs,
        'single_pin_nets': single_pin, 'unnamed_nets': unnamed,
        'bom_option_typos': typos,
        'bom_option_components': sorted(
            bom_option_components,
            key=lambda r: (_natural_sort_key(r['位号']), _natural_sort_key(r['页面'])),
        ),
    }


_DRC_ISSUE_KEYS = [
    'missing_hq_code', 'missing_value', 'missing_package',
    'tbd_attrs', 'single_pin_nets', 'unnamed_nets', 'bom_option_typos',
]
_DRC_REPORT_KEYS = ['bom_option_components']


# ══════════════════════════════════════════════════════════
# 五、电容降额分析
# ══════════════════════════════════════════════════════════

_SAFE_VOLTAGE_EXAMPLES: List[Tuple[str, float]] = [
    ('P12V_AUX', 12.0),
    ('12V_MAIN', 12.0),
    ('P5V_STBY', 5.0),
    ('P3V3_AON', 3.3),
    ('P1V8_S0', 1.8),
    ('P1V05_RTC', 1.05),
    ('GND', 0.0),
]


def _match_custom_voltage(net_name: str,
                          custom_volt_map: Optional[Dict[str, float]]) -> Optional[float]:
    if not custom_volt_map:
        return None
    best: Optional[Tuple[int, float]] = None
    for key, volt in custom_volt_map.items():
        prefix = str(key).strip().upper()
        if prefix and _matches_prefix_with_boundary(net_name, prefix):
            if best is None or len(prefix) > best[0]:
                best = (len(prefix), float(volt))
    return best[1] if best else None


def _infer_voltage(net_name: str) -> Optional[float]:
    token = _first_net_token(net_name)
    if _token_is_ground(token):
        return 0.0
    return _parse_voltage_token(token)


def _collect_global_max_voltage(nets: Dict,
                                custom_volt_map: Optional[Dict[str, float]] = None
                                ) -> Tuple[Optional[float], str, str]:
    max_voltage: Optional[float] = None
    max_net = ''
    max_source = ''
    for net_name in nets:
        v = _match_custom_voltage(net_name, custom_volt_map)
        source = 'custom_map' if v is not None else ''
        if v is None:
            v = _infer_voltage(net_name)
            if v is not None:
                source = 'net_token'
        if v is None or v <= 0:
            continue
        if max_voltage is None or float(v) > max_voltage:
            max_voltage = float(v)
            max_net = net_name
            max_source = source
    return max_voltage, max_net, max_source


def analyze_derating(components: Dict, nets: Dict,
                     pct: float = 70.0,
                     custom_volt_map: Optional[Dict[str, float]] = None) -> List[dict]:
    """pct: 工作电压上限占额定电压的百分比，默认 70（即工作电压 ≤ 额定 × 70% 视为合格）"""
    comp_nets = _collect_component_nets(nets)
    ac_coupling_caps = _find_ac_coupling_candidates(components, comp_nets, nets)
    global_max_v, global_max_net, global_max_source = _collect_global_max_voltage(nets, custom_volt_map)

    rows = []
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype not in ('CAP', 'CAP_POL'):
            continue
        connected_nets = _unique_component_nets(comp_nets, refdes)
        rated_str = comp.get('voltage', '')
        source_type = ''
        if not rated_str:
            status, derating, max_v, from_net = '⚪ 无额定电压', None, None, ''
            meta = _meta_fields('indeterminate', 'medium', 'high', 'missing_rated_voltage')
        else:
            m = re.match(r'([\d.]+)\s*V', rated_str.strip(), re.I)
            rated_v = float(m.group(1)) if m else None
            if rated_v is None:
                status, derating, max_v, from_net = '⚪ 无法解析额定电压', None, None, ''
                meta = _meta_fields('indeterminate', 'medium', 'high', 'unparsed_rated_voltage')
            elif global_max_v is not None and global_max_v <= 12.0 and rated_v >= 50.0:
                max_v = global_max_v
                from_net = global_max_net
                derating = rated_v / max_v if max_v else None
                source_type = '全局最大电压(自定义映射)' if global_max_source == 'custom_map' else '全局最大电压(网络名 token)'
                status = f'✅ 合格 (全局最大电压 {max_v:.1f}V ≤ 12V，50V 高耐压器件直接通过)'
                meta = _meta_fields(
                    'confirmed' if global_max_source == 'custom_map' else 'candidate',
                    'low',
                    'high' if global_max_source == 'custom_map' else 'medium',
                    'global_max_voltage_under_12v_high_rated_cap',
                )
            elif refdes in ac_coupling_caps:
                status, derating, max_v, from_net = '⚪ 无法判断（疑似 AC 耦合）', None, None, ''
                source_type = 'AC 耦合候选'
                meta = _meta_fields('indeterminate', 'low', 'medium', 'ac_coupling_candidate')
            else:
                max_v, from_net = None, ''
                known_nets: List[Tuple[str, float, str]] = []
                ground_present = False
                for net_name in connected_nets:
                    if _net_is_gnd(net_name):
                        ground_present = True
                    v = _match_custom_voltage(net_name, custom_volt_map)
                    source = 'custom_map' if v is not None else ''
                    if v is None:
                        v = _infer_voltage(net_name)
                        if v is not None:
                            source = 'net_token'
                    if v is None:
                        continue
                    if v == 0:
                        ground_present = True
                    known_nets.append((net_name, float(v), source))

                positives: Dict[float, Tuple[str, str]] = {}
                for net_name, v, source in known_nets:
                    if v > 0:
                        positives.setdefault(round(v, 6), (net_name, source))

                if not ground_present:
                    status, derating = '⚪ 无法判断（未连接地）', None
                    meta = _meta_fields('indeterminate', 'low', 'high', 'no_ground_reference')
                elif not positives:
                    status, derating = '⚪ 无法推断工作电压', None
                    meta = _meta_fields('indeterminate', 'low', 'high', 'no_positive_voltage_evidence')
                elif len(positives) > 1:
                    status, derating = '⚪ 无法判断（连接多个不同电位）', None
                    meta = _meta_fields('indeterminate', 'medium', 'high', 'multiple_positive_rails')
                else:
                    rounded_v, (from_net, source) = next(iter(positives.items()))
                    max_v = rounded_v
                    source_type = '自定义映射' if source == 'custom_map' else '网络首 token'
                    usage_pct = max_v / rated_v * 100        # 工作电压占额定的 %
                    derating  = rated_v / max_v              # 仍保留降额比供参考
                    if usage_pct <= pct:
                        status = f'✅ 合格 ({usage_pct:.0f}% ≤ {pct:.0f}%)'
                    else:
                        status = f'❌ 不合格 ({usage_pct:.0f}% > {pct:.0f}%)'
                    if source == 'custom_map':
                        meta = _meta_fields(
                            'confirmed',
                            'high' if status.startswith('❌') else 'low',
                            'high',
                            'custom_voltage_map',
                        )
                    else:
                        meta = _meta_fields(
                            'candidate',
                            'medium' if status.startswith('❌') else 'low',
                            'medium',
                            'single_positive_rail_token',
                        )
        rows.append({
            '位号':            refdes,
            '值':              comp.get('value', ''),
            '封装':            comp.get('package', ''),
            '类型':            COMP_TYPE_CN.get(ctype, ctype),
            '额定电压':        rated_str,
            '推断工作电压(V)': str(max_v) if max_v is not None else '',
            '推断来源网络':    from_net,
            '推断来源类型':    source_type,
            '所有连接网络':    ', '.join(connected_nets),
            '降额比':          f'{derating:.2f}' if derating is not None else '',
            '状态':            status,
            **meta,
            '页面':            _component_display_page(comp),
            'DEPOP':           'Y' if _is_depop_option(comp.get('bom_option', '')) else '',
        })
        rows[-1].update(_component_page_fields(comp))
    rows.sort(key=lambda r: (
        0 if r['状态'].startswith('❌') else 1 if r.get('结论类型') == '候选判断' else 2,
        {'高': 0, '中': 1, '低': 2}.get(r.get('严重级别', ''), 9),
        _natural_sort_key(r.get('位号', '')),
    ))
    return rows


# ══════════════════════════════════════════════════════════
# 六、电阻检查（上拉 / 下拉 / 串阻）
# ══════════════════════════════════════════════════════════

def _parse_ohms(value_str: str) -> Optional[float]:
    """解析电阻值字符串为欧姆数，支持 k/M/R/Ω 后缀，如 10k→10000, 4.7k→4700, 100R→100"""
    if not value_str:
        return None
    s = re.sub(r'\s', '', value_str.upper())
    s = s.replace('Ω', 'R').replace('Ω', 'R').replace('欧', 'R')
    s = re.sub(r'OHMS?$', 'R', s)
    s = re.sub(r'([KMG])R$', r'\1', s)
    m = re.match(r'^([\d.]+)([KMGR]?)$', s)
    if not m:
        embedded = re.match(r'^(\d+)([KMGR])(\d+)$', s)
        if not embedded:
            return None
        val = float(f'{embedded.group(1)}.{embedded.group(3)}')
        return val * {'K': 1e3, 'M': 1e6, 'G': 1e9, 'R': 1}.get(embedded.group(2), 1)
    val = float(m.group(1))
    return val * {'K': 1e3, 'M': 1e6, 'G': 1e9, 'R': 1, '': 1}.get(m.group(2), 1)


def _net_is_power(net: str) -> bool:
    return _token_is_power(_first_net_token(net))


def _net_is_gnd(net: str) -> bool:
    return _token_is_ground(_first_net_token(net))


_CHIP_REFDES_RE = re.compile(r'^(?:XU|PU|U)[A-Z0-9]+$', re.IGNORECASE)
_REFDES_SUFFIX_RE = re.compile(r'(?<=\d)([A-Z]+\d+)$', re.IGNORECASE)


def _is_chip_component(refdes: str, comp: Dict) -> bool:
    return comp.get('comp_type') == 'IC' and bool(_CHIP_REFDES_RE.match(refdes or ''))


def _extract_refdes_suffix_group(refdes: str) -> str:
    match = _REFDES_SUFFIX_RE.search((refdes or '').strip().upper())
    return match.group(1) if match else ''


def _extract_pin_hierarchy_labels(pin_name: str) -> List[str]:
    raw = str(pin_name or '').strip()
    if '@' not in raw:
        return []

    labels: List[str] = []
    for segment in raw.split('@'):
        segment = segment.strip()
        if not segment:
            continue
        head = segment.split(':', 1)[0].strip()
        if '(' in head:
            head = head.split('(', 1)[0].strip()
        label = head.rsplit('.', 1)[-1].strip()
        if label:
            labels.append(label)
    return labels


def _extract_pin_submodule_info(pin_name: str) -> Tuple[str, str]:
    labels = _extract_pin_hierarchy_labels(pin_name)
    if len(labels) >= 2:
        return labels[-2], ' / '.join(labels[:-1])
    if labels:
        return labels[0], labels[0]
    return '', ''


def _format_entry_list(entries: List[dict], key: str) -> str:
    return ', '.join(dict.fromkeys(str(entry.get(key, '')) for entry in entries if entry.get(key, '') != ''))


def _merge_display_values(*values: str) -> str:
    merged: List[str] = []
    for value in values:
        for part in str(value or '').split(','):
            part = part.strip()
            if part and part not in merged:
                merged.append(part)
    return ', '.join(merged)


def _series_edge_sort_key(entry: Dict[str, object]):
    return (_natural_sort_key(entry.get('refdes', '')), _natural_sort_key(entry.get('other_net', '')))


def _series_bias_sort_key(entry: Dict[str, object]):
    return (
        int(entry.get('via_hop_count', 0) or 0),
        _natural_sort_key(entry.get('refdes', '')),
        _natural_sort_key(entry.get('source_net', '')),
    )


def _series_chain_field(chain: List[dict], key: str) -> str:
    return ' -> '.join(str(edge.get(key, '')) for edge in chain if edge.get(key, ''))


def _series_chain_pages(chain: List[dict], key: str) -> str:
    return ', '.join(
        dict.fromkeys(str(edge.get(key, '')) for edge in chain if edge.get(key, ''))
    )


def _series_chain_total_ohms(chain: List[dict]) -> Optional[float]:
    total = 0.0
    for edge in chain:
        ohms = edge.get('ohms')
        if ohms is None:
            return None
        total += float(ohms)
    return total


MAX_SERIES_WALK_HOPS = 8
MAX_SERIES_WALK_RESULTS = 512


def _walk_series_paths(start_net: str, series_by_net: Dict[str, list]) -> List[dict]:
    if not start_net:
        return []

    queue = deque([{'net': start_net, 'chain': [], 'net_chain': [start_net]}])
    seen_paths = set()
    results: List[dict] = []

    while queue:
        state = queue.popleft()
        current_net = state['net']
        if len(state['chain']) >= MAX_SERIES_WALK_HOPS:
            continue
        for edge in sorted(series_by_net.get(current_net, []), key=_series_edge_sort_key):
            next_net = edge.get('other_net', '')
            if not next_net or next_net in state['net_chain']:
                continue
            if any(prev_edge.get('refdes', '') == edge.get('refdes', '') for prev_edge in state['chain']):
                continue
            next_chain = state['chain'] + [edge]
            next_net_chain = state['net_chain'] + [next_net]
            path_key = (
                next_net,
                tuple(
                    (chain_edge.get('refdes', ''), chain_edge.get('other_net', ''))
                    for chain_edge in next_chain
                ),
            )
            if path_key in seen_paths:
                continue
            seen_paths.add(path_key)
            results.append({
                'source_net': start_net,
                'target_net': next_net,
                'chain': next_chain,
                'net_chain': next_net_chain,
                'hop_count': len(next_chain),
                'via_refdes_chain': _series_chain_field(next_chain, 'refdes'),
                'via_value_chain': _series_chain_field(next_chain, 'value'),
                'via_net_chain': ' -> '.join(next_net_chain),
                'via_pages': _series_chain_pages(next_chain, 'page'),
                'via_mapped_pages': _series_chain_pages(next_chain, 'mapped_page'),
                'via_total_ohms': _series_chain_total_ohms(next_chain),
            })
            if len(results) >= MAX_SERIES_WALK_RESULTS:
                return results
            queue.append({
                'net': next_net,
                'chain': next_chain,
                'net_chain': next_net_chain,
            })

    return results


def _build_indirect_bias_maps(pullups: Dict[str, list],
                              pulldowns: Dict[str, list],
                              series_by_net: Dict[str, list]) -> Tuple[Dict[str, list], Dict[str, list]]:
    indirect_pullups: Dict[str, list] = defaultdict(list)
    indirect_pulldowns: Dict[str, list] = defaultdict(list)

    for start_net in sorted(series_by_net.keys(), key=_natural_sort_key):
        seen_keys = {'pullup': set(), 'pulldown': set()}
        for path in _walk_series_paths(start_net, series_by_net):
            remote_net = path['target_net']
            for bias_kind, direct_map, indirect_map in (
                ('pullup', pullups, indirect_pullups),
                ('pulldown', pulldowns, indirect_pulldowns),
            ):
                for bias in sorted(direct_map.get(remote_net, []), key=lambda row: _natural_sort_key(row.get('refdes', ''))):
                    dedupe_key = (remote_net, bias.get('refdes', ''), path['via_refdes_chain'])
                    if dedupe_key in seen_keys[bias_kind]:
                        continue
                    seen_keys[bias_kind].add(dedupe_key)
                    indirect_map[start_net].append({
                        **bias,
                        'source_net': remote_net,
                        'other_net': start_net,
                        'via_refdes': path['via_refdes_chain'],
                        'via_value': path['via_value_chain'],
                        'via_ohms': path['via_total_ohms'],
                        'via_refdes_chain': path['via_refdes_chain'],
                        'via_value_chain': path['via_value_chain'],
                        'via_net_chain': path['via_net_chain'],
                        'via_hop_count': path['hop_count'],
                        'via_pages': path['via_pages'],
                        'via_mapped_pages': path['via_mapped_pages'],
                    })

    return dict(indirect_pullups), dict(indirect_pulldowns)


def _name_tokens(value: str) -> List[str]:
    return re.findall(r'[A-Z0-9]+', (value or '').upper())


def _od_oc_evidence_from_name(value: str, source_label: str) -> List[Tuple[str, str]]:
    tokens = set(_name_tokens(value))
    evidence: List[Tuple[str, str]] = []
    normalized = (value or '').upper()

    if 'SDA' in tokens:
        evidence.append(('strong', f'{source_label}包含 SDA'))
    if 'SCL' in tokens:
        evidence.append(('strong', f'{source_label}包含 SCL'))
    if 'SMBALERT' in tokens or ({'SMB', 'ALERT'} <= tokens):
        evidence.append(('strong', f'{source_label}包含 SMBALERT'))
    if 'SMBDAT' in tokens or 'SMBDATA' in tokens:
        evidence.append(('strong', f'{source_label}包含 SMBDAT(A)'))
    if 'SMBCLK' in tokens or ({'SMB', 'CLK'} <= tokens):
        evidence.append(('strong', f'{source_label}包含 SMBCLK'))
    if {'OPEN', 'DRAIN'} <= tokens or 'OPENDRAIN' in tokens:
        evidence.append(('strong', f'{source_label}显式标注 OPEN_DRAIN'))
    if {'OPEN', 'COLLECTOR'} <= tokens or 'OPENCOLLECTOR' in tokens:
        evidence.append(('strong', f'{source_label}显式标注 OPEN_COLLECTOR'))
    if ' OD ' in f' {normalized} ' or tokens.intersection({'OD', 'OC'}):
        evidence.append(('strong', f'{source_label}显式包含 OD/OC'))

    weak_token_pairs = [
        ({'ALERT'}, 'ALERT'),
        ({'FAULT'}, 'FAULT'),
        ({'IRQ'}, 'IRQ'),
        ({'INT'}, 'INT'),
        ({'PGOOD'}, 'PGOOD'),
        ({'PWROK'}, 'PWROK'),
        ({'PWRGD'}, 'PWRGD'),
        ({'PRSNT'}, 'PRSNT'),
        ({'PRESENT'}, 'PRESENT'),
    ]
    for required_tokens, label in weak_token_pairs:
        if required_tokens <= tokens:
            evidence.append(('weak', f'{source_label}包含 {label}'))

    return evidence


def _classify_od_oc_evidence(net_name: str,
                             nodes: List[dict],
                             components: Dict) -> Optional[Dict[str, str]]:
    chip_nodes = []
    evidence: List[Tuple[str, str]] = []

    for node in nodes:
        refdes = node.get('refdes', '')
        comp = components.get(refdes, {})
        if not _is_chip_component(refdes, comp):
            continue
        chip_nodes.append(node)
        evidence.extend(_od_oc_evidence_from_name(node.get('pin_name', node.get('pin', '')),
                                                  f'{refdes}.{node.get("pin", "")}'))

    if not chip_nodes:
        return None

    evidence.extend(_od_oc_evidence_from_name(net_name, '网络名'))
    strong = [text for level, text in evidence if level == 'strong']
    weak = [text for level, text in evidence if level == 'weak']
    if not strong and len(weak) < 2:
        return None

    unique_evidence = list(dict.fromkeys(strong + weak))
    chip_pins = ', '.join(
        dict.fromkeys(
            f'{node["refdes"]}.{node["pin"]}({node.get("pin_name", node["pin"])})'
            for node in chip_nodes
        )
    )
    return {
        '芯片引脚': chip_pins,
        '判定依据': '; '.join(unique_evidence[:6]),
        'confidence': 'medium' if strong else 'low',
        'reason_code': 'od_oc_strong_name_without_pullup' if strong else 'od_oc_multi_hint_without_pullup',
    }


def _classify_series_bias_ratio(series_ohms: Optional[float],
                                bias_ohms: Optional[float]) -> Tuple[Optional[float], str, Dict[str, str]]:
    if series_ohms is None or bias_ohms is None or bias_ohms <= 0:
        return None, '⚪ 阻值缺失，无法计算', _meta_fields('indeterminate', 'low', 'high', 'missing_resistance_value')

    ratio = series_ohms / bias_ohms
    if bias_ohms < 1000 and ratio > 0.1:
        return ratio, '❌ 高风险', _meta_fields('candidate', 'high', 'medium', 'series_bias_ratio_high')
    if ratio >= 0.33:
        return ratio, '❌ 高风险', _meta_fields('candidate', 'high', 'medium', 'series_bias_ratio_high')
    if ratio > 0.1:
        return ratio, '⚠️ 关注', _meta_fields('candidate', 'medium', 'medium', 'series_bias_ratio_warn')
    return ratio, '✅ 正常', _meta_fields('candidate', 'low', 'medium', 'series_bias_ratio_ok')


def _analyze_resistors_multi_series(components: Dict, nets: Dict, *, exclude_depop: bool = True) -> dict:
    pullups: Dict[str, list] = defaultdict(list)
    pulldowns: Dict[str, list] = defaultdict(list)
    series_by_net: Dict[str, list] = defaultdict(list)
    node_lookup: Dict[Tuple[str, str], str] = {}

    for net_name, nodes in nets.items():
        for node in nodes:
            node_lookup[(node['refdes'], node['pin'])] = node.get('pin_name', node['pin'])

    for refdes, comp in components.items():
        if comp.get('comp_type') != 'RES':
            continue
        pin_nets = list(dict.fromkeys(comp.get('nets', {}).values()))
        if len(pin_nets) != 2:
            continue

        net_a, net_b = pin_nets[0], pin_nets[1]
        ohms = _parse_ohms(comp.get('value', ''))
        value = comp.get('value', '')
        page = _component_display_page(comp)
        mapped_page = _component_submodule_mapped_page(comp)
        bom_option = comp.get('bom_option', '')
        if exclude_depop and _is_depop_option(bom_option):
            continue

        a_pwr, b_pwr = _net_is_power(net_a), _net_is_power(net_b)
        a_gnd, b_gnd = _net_is_gnd(net_a), _net_is_gnd(net_b)

        if a_pwr and not b_pwr and not b_gnd:
            pullups[net_b].append({
                'refdes': refdes,
                'ohms': ohms,
                'value': value,
                'power_net': net_a,
                'page': page,
                'mapped_page': mapped_page,
                'bom_option': bom_option,
            })
        elif b_pwr and not a_pwr and not a_gnd:
            pullups[net_a].append({
                'refdes': refdes,
                'ohms': ohms,
                'value': value,
                'power_net': net_b,
                'page': page,
                'mapped_page': mapped_page,
                'bom_option': bom_option,
            })
        elif a_gnd and not b_gnd and not b_pwr:
            pulldowns[net_b].append({
                'refdes': refdes,
                'ohms': ohms,
                'value': value,
                'page': page,
                'mapped_page': mapped_page,
                'bom_option': bom_option,
            })
        elif b_gnd and not a_gnd and not a_pwr:
            pulldowns[net_a].append({
                'refdes': refdes,
                'ohms': ohms,
                'value': value,
                'page': page,
                'mapped_page': mapped_page,
                'bom_option': bom_option,
            })
        elif not a_pwr and not b_pwr and not a_gnd and not b_gnd:
            edge_a = {
                'refdes': refdes,
                'ohms': ohms,
                'value': value,
                'page': page,
                'mapped_page': mapped_page,
                'bom_option': bom_option,
                'other_net': net_b,
            }
            edge_b = dict(edge_a, other_net=net_a)
            series_by_net[net_a].append(edge_a)
            series_by_net[net_b].append(edge_b)

    indirect_pullups, indirect_pulldowns = _build_indirect_bias_maps(pullups, pulldowns, series_by_net)

    dup_pullups = []
    for sig_net, pu_list in sorted(pullups.items()):
        if len(pu_list) < 2:
            continue
        group = sorted(pu_list, key=lambda row: _natural_sort_key(row.get('refdes', '')))
        row = _with_meta({
            '信号网络': sig_net,
            '上拉数量': len(group),
            '位号': ', '.join(item['refdes'] for item in group),
            '阻值': ', '.join(item['value'] for item in group),
            '上拉电源': ', '.join(dict.fromkeys(item['power_net'] for item in group)),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(item['bom_option']) for item in group)),
            '页面': ', '.join(dict.fromkeys(item['page'] for item in group)),
        }, 'candidate', 'medium', 'medium', 'multiple_pullup_paths')
        row['子模块映射主模块真实页'] = ', '.join(
            dict.fromkeys(item.get('mapped_page', '') for item in group if item.get('mapped_page', ''))
        )
        dup_pullups.append(row)

    dup_pulldowns = []
    for sig_net, pd_list in sorted(pulldowns.items()):
        if len(pd_list) < 2:
            continue
        group = sorted(pd_list, key=lambda row: _natural_sort_key(row.get('refdes', '')))
        row = _with_meta({
            '信号网络': sig_net,
            '下拉数量': len(group),
            '位号': ', '.join(item['refdes'] for item in group),
            '阻值': ', '.join(item['value'] for item in group),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(item['bom_option']) for item in group)),
            '页面': ', '.join(dict.fromkeys(item['page'] for item in group)),
        }, 'candidate', 'medium', 'medium', 'multiple_pulldown_paths')
        row['子模块映射主模块真实页'] = ', '.join(
            dict.fromkeys(item.get('mapped_page', '') for item in group if item.get('mapped_page', ''))
        )
        dup_pulldowns.append(row)

    divider_risks = []
    for bias_kind, indirect_map in (('上拉', indirect_pullups), ('下拉', indirect_pulldowns)):
        for affected_net, entries in sorted(indirect_map.items(), key=lambda item: _natural_sort_key(item[0])):
            for bias in sorted(entries, key=_series_bias_sort_key):
                ratio, status, meta = _classify_series_bias_ratio(bias.get('via_ohms'), bias.get('ohms'))
                ref_net = bias.get('power_net', '') if bias_kind == '上拉' else 'GND'
                row = {
                    '串阻位号': bias.get('via_refdes_chain', ''),
                    '串阻值': bias.get('via_value_chain', ''),
                    '串阻网络A': affected_net,
                    '串阻网络B': bias.get('source_net', ''),
                    '串阻经过网络': bias.get('via_net_chain', ''),
                    '串阻跳数': bias.get('via_hop_count', 0),
                    '偏置类型': bias_kind,
                    '偏置位号': bias['refdes'],
                    '偏置值': bias['value'],
                    '偏置所在网络': bias.get('source_net', ''),
                    '偏置参考网络': ref_net,
                    '受影响网络': affected_net,
                    '串/偏置比': f'{ratio:.3f}' if ratio is not None else '',
                    '偏置 < 1k': '是' if (bias.get('ohms') or 0) < 1000 else '否',
                    '说明': (
                        f'{bias_kind}位于 {bias.get("source_net", "")} 侧，'
                        f'通过 {bias.get("via_refdes_chain", "")} 影响 {affected_net}'
                    ),
                    '状态': status,
                    **meta,
                    '页面': _merge_display_values(bias.get('via_pages', ''), bias.get('page', '')),
                }
                row['子模块映射主模块真实页'] = _merge_display_values(
                    bias.get('via_mapped_pages', ''),
                    bias.get('mapped_page', ''),
                )
                divider_risks.append(row)
    divider_risks.sort(key=lambda row: (
        0 if row['状态'].startswith('❌') else 1 if row['状态'].startswith('⚠') else 2,
        _natural_sort_key(row.get('串阻位号', '')),
        _natural_sort_key(row.get('偏置位号', '')),
    ))

    od_missing = []
    for net_name in sorted(nets.keys()):
        if _net_is_power(net_name) or _net_is_gnd(net_name):
            continue
        nodes = nets[net_name]
        evidence = _classify_od_oc_evidence(net_name, nodes, components)
        if not evidence:
            continue
        if pullups.get(net_name) or indirect_pullups.get(net_name):
            continue
        row = _with_meta({
            '网络名': net_name,
            '节点数': len(nodes),
            '连接元件': ', '.join(dict.fromkeys(node['refdes'] for node in nodes[:6])),
            '芯片引脚': evidence['芯片引脚'],
            '判定依据': evidence['判定依据'],
            '上拉状态': '未找到直接上拉/隔串阻上拉',
            '说明': '疑似 OD/OC 信号，但当前可见网络和隔一个或多串阻的可达网络上都未找到上拉电阻',
        }, 'candidate', 'medium', evidence['confidence'], evidence['reason_code'])
        row['子模块映射主模块真实页'] = ', '.join(
            dict.fromkeys(
                _component_submodule_mapped_page(components.get(node['refdes'], {}))
                for node in nodes
                if _component_submodule_mapped_page(components.get(node['refdes'], {}))
            )
        )
        od_missing.append(row)

    chip_pin_rows = []
    for refdes, comp in sorted(components.items(), key=lambda item: _natural_sort_key(item[0])):
        if not _is_chip_component(refdes, comp):
            continue
        for pin, net_name in sorted(comp.get('nets', {}).items(), key=lambda item: _natural_sort_key(item[0])):
            pin_name = node_lookup.get((refdes, pin), pin)
            submodule, submodule_path = _extract_pin_submodule_info(pin_name)
            series_entries = sorted(series_by_net.get(net_name, []), key=_series_edge_sort_key)
            pullup_entries = sorted(pullups.get(net_name, []), key=lambda row: _natural_sort_key(row.get('refdes', '')))
            pulldown_entries = sorted(pulldowns.get(net_name, []), key=lambda row: _natural_sort_key(row.get('refdes', '')))
            indirect_pullup_entries = sorted(indirect_pullups.get(net_name, []), key=_series_bias_sort_key)
            indirect_pulldown_entries = sorted(indirect_pulldowns.get(net_name, []), key=_series_bias_sort_key)
            row = {
                '芯片位号': refdes,
                '引脚': pin,
                '引脚名': pin_name,
                '后缀组': _extract_refdes_suffix_group(refdes),
                '子模块': submodule,
                '子模块路径': submodule_path,
                '网络名': net_name,
                '有串阻': '是' if series_entries else '否',
                '串阻数量': len(series_entries),
                '串阻位号': _format_entry_list(series_entries, 'refdes'),
                '串阻另一端网络': _format_entry_list(series_entries, 'other_net'),
                '串阻BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(entry.get('bom_option', '')) for entry in series_entries)),
                '有上拉': '是' if pullup_entries else '否',
                '上拉数量': len(pullup_entries),
                '上拉位号': _format_entry_list(pullup_entries, 'refdes'),
                '上拉电源': _format_entry_list(pullup_entries, 'power_net'),
                '上拉BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(entry.get('bom_option', '')) for entry in pullup_entries)),
                '隔串阻上拉数量': len(indirect_pullup_entries),
                '隔串阻上拉位号': _format_entry_list(indirect_pullup_entries, 'refdes'),
                '隔串阻上拉来源网络': _format_entry_list(indirect_pullup_entries, 'source_net'),
                '隔串阻上拉电源': _format_entry_list(indirect_pullup_entries, 'power_net'),
                '隔串阻上拉串阻链': _format_entry_list(indirect_pullup_entries, 'via_refdes_chain'),
                '有下拉': '是' if pulldown_entries else '否',
                '下拉数量': len(pulldown_entries),
                '下拉位号': _format_entry_list(pulldown_entries, 'refdes'),
                '下拉BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(entry.get('bom_option', '')) for entry in pulldown_entries)),
                '隔串阻下拉数量': len(indirect_pulldown_entries),
                '隔串阻下拉位号': _format_entry_list(indirect_pulldown_entries, 'refdes'),
                '隔串阻下拉来源网络': _format_entry_list(indirect_pulldown_entries, 'source_net'),
                '隔串阻下拉串阻链': _format_entry_list(indirect_pulldown_entries, 'via_refdes_chain'),
                '页面': _component_display_page(comp),
                '页码一一对应': comp.get('page_mapping_ok', ''),
            }
            row.update(_component_page_fields(comp))
            chip_pin_rows.append(row)

    return {
        'dup_pullups': dup_pullups,
        'dup_pulldowns': dup_pulldowns,
        'divider_risks': divider_risks,
        'od_missing': od_missing,
        'chip_pin_rows': chip_pin_rows,
        'pullups': dict(pullups),
        'pulldowns': dict(pulldowns),
        'indirect_pullups': dict(indirect_pullups),
        'indirect_pulldowns': dict(indirect_pulldowns),
        'series_by_net': dict(series_by_net),
    }


def analyze_resistors(components: Dict, nets: Dict, *, exclude_depop: bool = True) -> dict:
    """检测上拉/下拉/串阻相关设计问题"""
    return _analyze_resistors_multi_series(components, nets, exclude_depop=exclude_depop)

    pullups:   Dict[str, list] = defaultdict(list)
    pulldowns: Dict[str, list] = defaultdict(list)
    series_list: list = []
    series_by_net: Dict[str, list] = defaultdict(list)
    node_lookup: Dict[Tuple[str, str], str] = {}

    for net_name, nodes in nets.items():
        for node in nodes:
            node_lookup[(node['refdes'], node['pin'])] = node.get('pin_name', node['pin'])

    for refdes, comp in components.items():
        if comp.get('comp_type') != 'RES':
            continue
        pin_nets = list(dict.fromkeys(comp.get('nets', {}).values()))
        if len(pin_nets) != 2:
            continue
        net_a, net_b = pin_nets[0], pin_nets[1]
        ohms    = _parse_ohms(comp.get('value', ''))
        val_str = comp.get('value', '')
        page    = _component_display_page(comp)
        mapped_page = _component_submodule_mapped_page(comp)
        bom_option = comp.get('bom_option', '')
        if exclude_depop and _is_depop_option(bom_option):
            continue

        a_pwr, b_pwr = _net_is_power(net_a), _net_is_power(net_b)
        a_gnd, b_gnd = _net_is_gnd(net_a),   _net_is_gnd(net_b)

        if a_pwr and not b_pwr and not b_gnd:
            pullups[net_b].append({'refdes': refdes, 'ohms': ohms, 'value': val_str,
                                   'power_net': net_a, 'page': page, 'mapped_page': mapped_page,
                                   'bom_option': bom_option})
        elif b_pwr and not a_pwr and not a_gnd:
            pullups[net_a].append({'refdes': refdes, 'ohms': ohms, 'value': val_str,
                                   'power_net': net_b, 'page': page, 'mapped_page': mapped_page,
                                   'bom_option': bom_option})
        elif a_gnd and not b_gnd and not b_pwr:
            pulldowns[net_b].append({'refdes': refdes, 'ohms': ohms,
                                     'value': val_str, 'page': page, 'mapped_page': mapped_page,
                                     'bom_option': bom_option})
        elif b_gnd and not a_gnd and not a_pwr:
            pulldowns[net_a].append({'refdes': refdes, 'ohms': ohms,
                                     'value': val_str, 'page': page, 'mapped_page': mapped_page,
                                     'bom_option': bom_option})
        elif not a_pwr and not b_pwr and not a_gnd and not b_gnd:
            series_list.append({'refdes': refdes, 'net_a': net_a, 'net_b': net_b,
                                 'ohms': ohms, 'value': val_str, 'page': page,
                                 'mapped_page': mapped_page, 'bom_option': bom_option})
            series_by_net[net_a].append({
                'refdes': refdes, 'ohms': ohms, 'value': val_str, 'page': page,
                'mapped_page': mapped_page, 'bom_option': bom_option, 'other_net': net_b,
            })
            series_by_net[net_b].append({
                'refdes': refdes, 'ohms': ohms, 'value': val_str, 'page': page,
                'mapped_page': mapped_page, 'bom_option': bom_option, 'other_net': net_a,
            })

    # ── 检查1：重复上拉 ───────────────────────────────────
    indirect_pullups, indirect_pulldowns = _build_indirect_bias_maps(pullups, pulldowns, series_by_net)

    dup_pullups = []
    for sig_net, pu_list in sorted(pullups.items()):
        if len(pu_list) < 2:
            continue
        group = sorted(pu_list, key=lambda row: _natural_sort_key(row.get('refdes', '')))
        dup_pullups.append(_with_meta({
            '信号网络':  sig_net,
            '上拉数量':  len(group),
            '位号':      ', '.join(r['refdes'] for r in group),
            '阻值':      ', '.join(r['value']  for r in group),
            '上拉电源':  ', '.join(dict.fromkeys(r['power_net'] for r in group)),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(r['bom_option']) for r in group)),
            '页面':      ', '.join(dict.fromkeys(r['page'] for r in group)),
        }, 'candidate', 'medium', 'medium', 'multiple_pullup_paths'))
        dup_pullups[-1]['子模块映射主模块真实页'] = ', '.join(
            dict.fromkeys(r.get('mapped_page', '') for r in group if r.get('mapped_page', ''))
        )

    # ── 检查2：重复下拉 ───────────────────────────────────
    dup_pulldowns = []
    for sig_net, pd_list in sorted(pulldowns.items()):
        if len(pd_list) < 2:
            continue
        group = sorted(pd_list, key=lambda row: _natural_sort_key(row.get('refdes', '')))
        dup_pulldowns.append(_with_meta({
            '信号网络': sig_net,
            '下拉数量': len(group),
            '位号':     ', '.join(r['refdes'] for r in group),
            '阻值':     ', '.join(r['value']  for r in group),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(r['bom_option']) for r in group)),
            '页面':     ', '.join(dict.fromkeys(r['page'] for r in group)),
        }, 'candidate', 'medium', 'medium', 'multiple_pulldown_paths'))
        dup_pulldowns[-1]['子模块映射主模块真实页'] = ', '.join(
            dict.fromkeys(r.get('mapped_page', '') for r in group if r.get('mapped_page', ''))
        )

    # ── 检查3：串阻 + 偏置电阻分压风险 ────────────────────
    divider_risks = []
    seen_pairs: set = set()
    seen_indirect: set = set()
    for sr in sorted(series_list, key=lambda row: _natural_sort_key(row.get('refdes', ''))):
        for bias_net, affected_net in ((sr['net_a'], sr['net_b']), (sr['net_b'], sr['net_a'])):
            for bias_kind, bias_entries, indirect_map in [
                ('上拉', pullups.get(bias_net, []), indirect_pullups),
                ('下拉', pulldowns.get(bias_net, []), indirect_pulldowns),
            ]:
                for bias in bias_entries:
                    indirect_key = (affected_net, bias_kind, bias['refdes'], sr['refdes'])
                    if indirect_key not in seen_indirect:
                        seen_indirect.add(indirect_key)
                        indirect_map[affected_net].append({
                            **bias,
                            'via_refdes': sr['refdes'],
                            'via_value': sr['value'],
                            'via_ohms': sr['ohms'],
                            'source_net': bias_net,
                            'other_net': affected_net,
                        })

                    pair_key = (sr['refdes'], bias['refdes'], bias_kind, bias_net, affected_net)
                    if pair_key in seen_pairs:
                        continue
                    seen_pairs.add(pair_key)

                    ratio, status, meta = _classify_series_bias_ratio(sr['ohms'], bias['ohms'])
                    ref_net = bias.get('power_net', '') if bias_kind == '上拉' else 'GND'
                    pages = ', '.join(dict.fromkeys(v for v in [sr.get('page', ''), bias.get('page', '')] if v))
                    divider_risks.append({
                        '串阻位号': sr['refdes'],
                        '串阻值': sr['value'],
                        '串阻网络A': sr['net_a'],
                        '串阻网络B': sr['net_b'],
                        '偏置类型': bias_kind,
                        '偏置位号': bias['refdes'],
                        '偏置值': bias['value'],
                        '偏置所在网络': bias_net,
                        '偏置参考网络': ref_net,
                        '受影响网络': affected_net,
                        '串/偏置比': f'{ratio:.3f}' if ratio is not None else '',
                        '偏置 < 1k': '是' if (bias.get('ohms') or 0) < 1000 else '否',
                        '说明': f'{bias_kind}位于 {bias_net} 侧，通过 {sr["refdes"]} 影响 {affected_net}',
                        '状态': status,
                        **meta,
                        '页面': pages,
                    })
                    divider_risks[-1]['子模块映射主模块真实页'] = ', '.join(
                        dict.fromkeys(
                            v for v in [sr.get('mapped_page', ''), bias.get('mapped_page', '')] if v
                        )
                    )
    divider_risks.sort(key=lambda r: (
        0 if r['状态'].startswith('❌') else 1 if r['状态'].startswith('⚠') else 2,
        _natural_sort_key(r.get('串阻位号', '')),
        _natural_sort_key(r.get('偏置位号', '')),
    ))

    # ── 检查4：OD/OC 信号缺上拉 ──────────────────────────
    od_missing = []
    for net_name in sorted(nets.keys()):
        if _net_is_power(net_name) or _net_is_gnd(net_name):
            continue
        nodes = nets[net_name]
        evidence = _classify_od_oc_evidence(net_name, nodes, components)
        if not evidence:
            continue
        if pullups.get(net_name) or indirect_pullups.get(net_name):
            continue
        od_missing.append(_with_meta({
            '网络名': net_name,
            '节点数': len(nodes),
            '连接元件': ', '.join(dict.fromkeys(node['refdes'] for node in nodes[:6])),
            '芯片引脚': evidence['芯片引脚'],
            '判定依据': evidence['判定依据'],
            '上拉状态': '未找到直接上拉/隔串阻上拉',
            '说明': '疑似 OD/OC 信号，但当前可见网络和隔一只串阻的相邻网络上都未找到上拉电阻',
        }, 'candidate', 'medium', evidence['confidence'], evidence['reason_code']))
        od_missing[-1]['子模块映射主模块真实页'] = ', '.join(
            dict.fromkeys(
                _component_submodule_mapped_page(components.get(node['refdes'], {}))
                for node in nodes
                if _component_submodule_mapped_page(components.get(node['refdes'], {}))
            )
        )

    chip_pin_rows = []
    for refdes, comp in sorted(components.items(), key=lambda item: _natural_sort_key(item[0])):
        if not _is_chip_component(refdes, comp):
            continue
        for pin, net_name in sorted(comp.get('nets', {}).items(), key=lambda item: _natural_sort_key(item[0])):
            pin_name = node_lookup.get((refdes, pin), pin)
            submodule, submodule_path = _extract_pin_submodule_info(pin_name)
            series_entries = series_by_net.get(net_name, [])
            pullup_entries = pullups.get(net_name, [])
            pulldown_entries = pulldowns.get(net_name, [])
            indirect_pullup_entries = indirect_pullups.get(net_name, [])
            indirect_pulldown_entries = indirect_pulldowns.get(net_name, [])
            chip_pin_rows.append({
                '芯片位号': refdes,
                '引脚': pin,
                '引脚名': pin_name,
                '后缀组': _extract_refdes_suffix_group(refdes),
                '子模块': submodule,
                '子模块路径': submodule_path,
                '网络名': net_name,
                '有串阻': '是' if series_entries else '否',
                '串阻数量': len(series_entries),
                '串阻位号': _format_entry_list(series_entries, 'refdes'),
                '串阻另一端网络': _format_entry_list(series_entries, 'other_net'),
                '串阻BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(e.get('bom_option', '')) for e in series_entries)),
                '有上拉': '是' if pullup_entries else '否',
                '上拉数量': len(pullup_entries),
                '上拉位号': _format_entry_list(pullup_entries, 'refdes'),
                '上拉电源': _format_entry_list(pullup_entries, 'power_net'),
                '上拉BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(e.get('bom_option', '')) for e in pullup_entries)),
                '隔串阻上拉数量': len(indirect_pullup_entries),
                '隔串阻上拉位号': _format_entry_list(indirect_pullup_entries, 'refdes'),
                '隔串阻上拉来源网络': _format_entry_list(indirect_pullup_entries, 'source_net'),
                '有下拉': '是' if pulldown_entries else '否',
                '下拉数量': len(pulldown_entries),
                '下拉位号': _format_entry_list(pulldown_entries, 'refdes'),
                '下拉BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(e.get('bom_option', '')) for e in pulldown_entries)),
                '隔串阻下拉数量': len(indirect_pulldown_entries),
                '隔串阻下拉位号': _format_entry_list(indirect_pulldown_entries, 'refdes'),
                '隔串阻下拉来源网络': _format_entry_list(indirect_pulldown_entries, 'source_net'),
                '页面': _component_display_page(comp),
                '页码一一对应': comp.get('page_mapping_ok', ''),
            })
            chip_pin_rows[-1].update(_component_page_fields(comp))

    return {
        'dup_pullups':    dup_pullups,
        'dup_pulldowns':  dup_pulldowns,
        'divider_risks':  divider_risks,
        'od_missing':     od_missing,
        'chip_pin_rows':  chip_pin_rows,
        'pullups':        dict(pullups),
        'pulldowns':      dict(pulldowns),
        'indirect_pullups': dict(indirect_pullups),
        'indirect_pulldowns': dict(indirect_pulldowns),
        'series_by_net':  dict(series_by_net),
    }


def analyze_project_contents(prt_content: str,
                             net_content: str,
                             *,
                             project_name: str = '',
                             project_root: str = '',
                             ratio_limit: float = 70.0,
                             custom_volt_map: Optional[Dict[str, float]] = None,
                             include_depop: bool = False) -> dict:
    components, nets, comp_nets = parse_all(prt_content, net_content)
    page_context = _prepare_page_resolution(project_root)
    page_mapping = page_context.get('page_mapping', {})
    page_warnings = list(page_context.get('warnings', []))
    _apply_component_pages(components, page_context)
    bom_normal_detail, bom_depop_detail, bom_normal_merged, bom_depop_merged = build_bom(components)
    analysis_components, analysis_nets, depop_refdes, excluded_depop_refdes = _build_analysis_scope(
        components,
        nets,
        include_depop=include_depop,
    )
    analysis_comp_nets = {refdes: dict(comp.get('nets', {})) for refdes, comp in analysis_components.items()}
    net_analysis = analyze_networks(analysis_nets, analysis_components)
    drc = check_drc(analysis_components, analysis_nets, option_components_source=components)
    derating = analyze_derating(analysis_components, analysis_nets, ratio_limit, custom_volt_map)
    resistor_analysis = analyze_resistors(analysis_components, analysis_nets, exclude_depop=not include_depop)
    if depop_refdes:
        if include_depop:
            page_warnings.append(
                f'DEPOP 排查开关已开启：共有 {len(depop_refdes)} 个 BOM_OPTION=DEPOP/DNP 元件继续参与后续分析。'
            )
        else:
            preview = ', '.join(depop_refdes[:8])
            suffix = ' ...' if len(depop_refdes) > 8 else ''
            page_warnings.append(
                f'DEPOP 排查开关默认关闭：已在后续分析中忽略 {len(depop_refdes)} 个 BOM_OPTION=DEPOP/DNP 元件'
                f'（{preview}{suffix}）。'
            )
    return {
        'project_name': project_name,
        'project_root': project_root,
        'components': analysis_components,
        'nets': analysis_nets,
        'comp_nets': analysis_comp_nets,
        'all_components': components,
        'all_nets': nets,
        'all_comp_nets': comp_nets,
        'bom_normal_detail': bom_normal_detail,
        'bom_depop_detail': bom_depop_detail,
        'bom_normal_merged': bom_normal_merged,
        'bom_depop_merged': bom_depop_merged,
        'net_analysis': net_analysis,
        'page_mapping_rows': page_mapping.get('rows', []),
        'drc': drc,
        'derating': derating,
        'resistor_analysis': resistor_analysis,
        'ratio_limit': ratio_limit,
        'custom_volt_map': custom_volt_map or None,
        'include_depop': include_depop,
        'depop_refdes': depop_refdes,
        'excluded_depop_refdes': excluded_depop_refdes,
        'page_warnings': page_warnings,
    }


def query_project_data(components: Dict,
                       nets: Dict,
                       mode: str,
                       keyword: str) -> dict:
    kw = (keyword or '').strip()
    if not kw:
        return {
            'title': '空查询',
            'lines': ['请输入位号或网络名。'],
            'mode': mode,
            'view': 'empty',
            'entity_type': '',
            'match_type': 'empty',
            'summary': {},
            'cards': [],
            'items': [],
        }

    lines: List[str] = []
    if mode == '位号':
        comp = components.get(kw)
        if comp is None:
            comp = next((value for refdes, value in components.items() if refdes.upper() == kw.upper()), None)
        if comp:
            value_text = comp.get('value', '') or comp.get('part_name', '')
            display_page = _component_display_page(comp)
            mapped_page = _component_submodule_mapped_page(comp)
            prop_items = [
                {'label': '位号', 'value': str(comp.get('refdes', kw))},
                {'label': '类型', 'value': str(COMP_TYPE_CN.get(comp.get('comp_type', ''), comp.get('comp_type', '')))},
                {'label': '料号', 'value': str(comp.get('hq_code', ''))},
                {'label': '值', 'value': str(comp.get('value', ''))},
                {'label': '封装', 'value': str(comp.get('package', ''))},
                {'label': 'BOM_OPTION', 'value': _display_bom_option(comp.get('bom_option', ''))},
                {'label': '页面', 'value': display_page},
                {'label': '页码一一对应', 'value': str(comp.get('page_mapping_ok', ''))},
                {'label': '页码来源', 'value': str(comp.get('page_source', ''))},
                {'label': 'ROOM', 'value': str(comp.get('room', ''))},
                {'label': 'DRAWING', 'value': str(comp.get('drawing', ''))},
            ]
            prop_items = [item for item in prop_items if item['value']]
            if mapped_page:
                prop_items.insert(7, {'label': '子模块映射主模块真实页', 'value': mapped_page})
            lines.append(f'◆ 元件：{comp.get("refdes", kw)}')
            for item in prop_items:
                lines.append(f'  {item["label"]:<16} {item["value"]}')
            lines += ['', '  引脚 -> 网络：']
            pin_rows = []
            for pin, net_name in sorted(comp.get('nets', {}).items(), key=lambda item: _natural_sort_key(item[0])):
                lines.append(f'    pin {pin:<6} -> {net_name}')
                pin_rows.append({'pin': pin, 'net': net_name})
            return {
                'title': comp.get('refdes', kw),
                'lines': lines,
                'mode': mode,
                'view': 'component',
                'entity_type': 'component',
                'match_type': 'exact',
                'summary': {
                    'title': comp.get('refdes', kw),
                    'subtitle': value_text,
                    'meta': [
                        {'label': '封装', 'value': str(comp.get('package', ''))},
                        {'label': '页面', 'value': display_page},
                        {'label': '子模块映射主模块真实页', 'value': mapped_page},
                        {'label': '页码一一对应', 'value': str(comp.get('page_mapping_ok', ''))},
                        {'label': '页码来源', 'value': str(comp.get('page_source', ''))},
                        {'label': '料号', 'value': str(comp.get('hq_code', ''))},
                    ],
                },
                'cards': [
                    {'title': '元件属性', 'kind': 'properties', 'items': prop_items},
                    {'title': '引脚连接', 'kind': 'pins', 'items': pin_rows},
                ],
                'items': [],
            }

        matched = sorted(refdes for refdes in components if kw.upper() in refdes.upper())
        lines.append('未找到精确匹配，模糊结果：' if matched else f'未找到位号：{kw}')
        lines.extend(f'  {refdes}' for refdes in matched[:50])
        items = []
        for refdes in matched[:50]:
            comp = components.get(refdes, {})
            items.append({
                'title': refdes,
                'subtitle': str(comp.get('value', '') or comp.get('part_name', '')),
                'meta': [
                    {'label': '封装', 'value': str(comp.get('package', ''))},
                    {'label': '页面', 'value': _component_display_page(comp)},
                    {'label': '子模块映射主模块真实页', 'value': _component_submodule_mapped_page(comp)},
                    {'label': '页码一一对应', 'value': str(comp.get('page_mapping_ok', ''))},
                ],
                'keyword': refdes,
            })
        return {
            'title': kw,
            'lines': lines,
            'mode': mode,
            'view': 'list',
            'entity_type': 'component',
            'match_type': 'fuzzy' if matched else 'missing',
            'summary': {
                'title': kw,
                'subtitle': '模糊匹配结果' if matched else '未找到位号',
                'meta': [{'label': '结果数', 'value': str(len(items))}],
            },
            'cards': [],
            'items': items,
        }

    nodes = nets.get(kw)
    exact_name = kw
    if nodes is None:
        exact_name = next((name for name in nets if name.upper() == kw.upper()), kw)
        nodes = nets.get(exact_name)
    if nodes:
        lines.append(f'◆ 网络：{exact_name}（{len(nodes)} 个节点）')
        node_rows = []
        for node in nodes:
            comp = components.get(node['refdes'], {})
            desc = comp.get('value', '') or comp.get('part_name', '')
            lines.append(f'  {node["refdes"]:<10} pin {node["pin"]:<6} ({node["pin_name"]:<12}) {desc}')
            node_rows.append({
                'refdes': node['refdes'],
                'pin': node['pin'],
                'pin_name': node['pin_name'],
                'desc': desc,
                '页面': _component_display_page(comp),
                '页码一一对应': str(comp.get('page_mapping_ok', '')),
            })
            node_rows[-1].update({
                '子模块映射主模块真实页': _component_submodule_mapped_page(comp),
            })
        return {
            'title': exact_name,
            'lines': lines,
            'mode': mode,
            'view': 'network',
            'entity_type': 'network',
            'match_type': 'exact',
            'summary': {
                'title': exact_name,
                'subtitle': f'{len(nodes)} 个连接节点',
                'meta': [
                    {'label': '节点数', 'value': str(len(nodes))},
                    {'label': '页面覆盖', 'value': str(len({row["页面"] for row in node_rows if row["页面"]}))},
                ],
            },
            'cards': [
                {'title': '网络节点', 'kind': 'nodes', 'items': node_rows},
            ],
            'items': [],
        }

    matched = sorted(name for name in nets if kw.upper() in name.upper())
    lines.append('未找到精确匹配，模糊结果：' if matched else f'未找到网络：{kw}')
    lines.extend(f'  {name}  ({len(nets[name])} nodes)' for name in matched[:50])
    items = [{
        'title': name,
        'subtitle': f'{len(nets[name])} 个节点',
        'meta': [],
        'keyword': name,
    } for name in matched[:50]]
    return {
        'title': kw,
        'lines': lines,
        'mode': mode,
        'view': 'list',
        'entity_type': 'network',
        'match_type': 'fuzzy' if matched else 'missing',
        'summary': {
            'title': kw,
            'subtitle': '模糊匹配结果' if matched else '未找到网络',
            'meta': [{'label': '结果数', 'value': str(len(items))}],
        },
        'cards': [],
        'items': items,
    }


# ══════════════════════════════════════════════════════════
# 八、Excel 导出
# ══════════════════════════════════════════════════════════

_BL = PatternFill("solid", fgColor="1F4E79")
_OR = PatternFill("solid", fgColor="C55A11")
_GR = PatternFill("solid", fgColor="375623")
_GY = PatternFill("solid", fgColor="595959")
_RF = PatternFill("solid", fgColor="FFCCCC")
_WF = Font(color="FFFFFF", bold=True, size=10)
_BF = Font(bold=True, size=10)
_NF = Font(size=10)
_CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LA = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_TH = Side(style='thin')
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


def _xl_hdr(ws, row_idx, fill):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill; cell.font = _WF; cell.alignment = _CA; cell.border = _BD


def _xl_autowidth(ws, mx=50):
    for col in ws.columns:
        vals = [str(c.value or '') for c in col]
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(v) for v in vals), default=8) + 2, mx)


def _xl_write_rows(ws, rows: List[dict], fill, hl_col=None, freeze=True):
    if not rows:
        ws.append(['（无数据）']); return
    hdrs = list(rows[0].keys())
    ws.append(hdrs); _xl_hdr(ws, ws.max_row, fill)
    hl_idx = hdrs.index(hl_col) if hl_col in hdrs else None
    for row in rows:
        ws.append([row.get(h, '') for h in hdrs])
        ri  = ws.max_row
        red = hl_idx is not None and '❌' in str(ws.cell(ri, hl_idx+1).value or '')
        for cell in ws[ri]:
            cell.border = _BD; cell.alignment = _LA; cell.font = _NF
            if red: cell.fill = _RF
    _xl_autowidth(ws)
    if freeze: ws.freeze_panes = 'A2'


def _xl_section(ws, title, fill):
    ws.append([title])
    for cell in ws[ws.max_row]:
        cell.fill = fill; cell.font = _WF; cell.border = _BD
    ws.append([])


def export_to_excel(data: dict, out_path: str) -> str:
    base, ext = os.path.splitext(out_path)
    n, path = 1, out_path
    while os.path.exists(path):
        path = f'{base}({n}){ext}'; n += 1

    wb = Workbook(); wb.remove(wb.active)
    project = data.get('project_name', '')
    na  = data.get('net_analysis', {})
    drc = data.get('drc', {})
    drt = data.get('derating', [])
    res = data.get('resistor_analysis', {})
    mn  = data.get('bom_normal_merged', [])
    md  = data.get('bom_depop_merged', [])
    network_rows = _iter_list_rows(na, ['power_net_rows', 'gnd_net_rows', 'diff_pair_rows', 'single_node_rows'])
    drc_rows = _iter_list_rows(drc, _DRC_ISSUE_KEYS)
    resistor_rows = _iter_list_rows(res, ['divider_risks', 'dup_pullups', 'dup_pulldowns', 'od_missing'])
    net_kind_counts = _count_result_kinds(network_rows)
    drc_kind_counts = _count_result_kinds(drc_rows)
    drt_kind_counts = _count_result_kinds(drt)
    resistor_kind_counts = _count_result_kinds(resistor_rows)

    # 概览
    ws = wb.create_sheet('概览')
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 16
    drc_total = sum(len(drc.get(key, [])) for key in _DRC_ISSUE_KEYS)
    fail = sum(1 for r in drt if r.get('状态', '').startswith('❌'))
    for label, val in [
        ('项目名称', project),
        ('贴装元件种类数', len(mn)),
        ('贴装元件总数',  sum(r.get('数量', 0) for r in mn)),
        ('DEPOP 元件种类数', len(md)),
        ('DEPOP 元件总数',  sum(r.get('数量', 0) for r in md)),
        ('网络总数', na.get('total', '')),
        ('候选单节点网络数', len(na.get('single_node', {}))),
        ('候选电源网络数', len(na.get('power_nets', {}))),
        ('候选差分对数', len(na.get('diff_pairs', {}))),
        ('DRC 问题总数', drc_total),
        ('电容降额不合格数', fail),
        ('网络候选判断数', net_kind_counts.get('候选判断', 0)),
        ('DRC 确定结论数', drc_kind_counts.get('确定结论', 0)),
        ('DRC 候选判断数', drc_kind_counts.get('候选判断', 0)),
        ('降额确定结论数', drt_kind_counts.get('确定结论', 0)),
        ('降额候选判断数', drt_kind_counts.get('候选判断', 0)),
        ('降额无法判断数', drt_kind_counts.get('无法判断', 0)),
        ('电阻候选判断数', resistor_kind_counts.get('候选判断', 0)),
        ('电阻无法判断数', resistor_kind_counts.get('无法判断', 0)),
    ]:
        ws.append([label, val])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _BD
            cell.font = _BF if cell.column == 1 else _NF
            cell.alignment = _LA

    # BOM
    ws = wb.create_sheet('BOM_贴装'); _xl_write_rows(ws, mn, _BL)
    ws = wb.create_sheet('BOM_DEPOP'); _xl_write_rows(ws, md, _OR)
    ws = wb.create_sheet('BOM_明细')
    all_d = [{'DEPOP': '', **r} for r in data.get('bom_normal_detail', [])] + \
            [{'DEPOP': 'Y', **r} for r in data.get('bom_depop_detail', [])]
    _xl_write_rows(ws, all_d, _GY)

    # 网络分析
    ws = wb.create_sheet('网络分析'); ws.freeze_panes = None
    _xl_section(ws, '候选电源网络', _BL)
    _xl_write_rows(ws, na.get('power_net_rows', []), _BL, freeze=False)
    ws.append([])
    _xl_section(ws, '候选 GND 网络', _GR)
    _xl_write_rows(ws, na.get('gnd_net_rows', []), _GR, freeze=False)
    ws.append([])
    _xl_section(ws, '候选差分对', _OR)
    _xl_write_rows(ws, na.get('diff_pair_rows', []), _OR, freeze=False)
    ws.append([])
    _xl_section(ws, '单节点候选网络', _GY)
    _xl_write_rows(ws, na.get('single_node_rows', []), _GY, freeze=False)
    ws.append([])
    _xl_section(ws, '各页面元件数', _BL)
    _xl_write_rows(ws, na.get('page_rows', []), _BL, freeze=False)
    ws.append([])
    _xl_section(ws, '逻辑页/真实页映射检查', _GY)
    _xl_write_rows(ws, data.get('page_mapping_rows', []), _GY, freeze=False)
    _xl_autowidth(ws)

    # 设计检查
    ws = wb.create_sheet('设计检查'); ws.freeze_panes = None
    for title, key, fill in [
        ('TBD 待确认属性', 'tbd_attrs',       _OR),
        ('缺少料号',       'missing_hq_code',  _RF),
        ('缺少 VALUE',     'missing_value',     _RF),
        ('缺少封装',       'missing_package',   _RF),
        ('单端候选网络',   'single_pin_nets',   _GY),
        ('未命名网络',     'unnamed_nets',      _GY),
        ('BOM_OPTION 候选拼写', 'bom_option_typos', _OR),
        ('BOM_OPTION 元件', 'bom_option_components', _BL),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, drc.get(key, []), fill, freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    # 电阻检查
    ws = wb.create_sheet('电阻检查'); ws.freeze_panes = None
    for title, key, fill in [
        ('串阻分压候选风险', 'divider_risks',  _OR),
        ('重复上拉候选',     'dup_pullups',    _RF),
        ('重复下拉候选',     'dup_pulldowns',  _RF),
        ('OD/OC 候选缺上拉', 'od_missing',     _GY),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, res.get(key, []), fill, hl_col='状态', freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    ws = wb.create_sheet('芯片引脚电阻')
    _xl_write_rows(ws, res.get('chip_pin_rows', []), _BL)

    # 降额
    ws = wb.create_sheet('降额分析')
    _xl_write_rows(ws, drt, _BL, hl_col='状态')

    wb.save(path)
    return path


def main() -> int:
    # Backward-compatible launcher: keep python pstx_analyzer.py useful while
    # the actual desktop shell lives in pstx_local_ui.py.
    import sys as _sys
    if 'pstx_analyzer' not in _sys.modules:
        _sys.modules['pstx_analyzer'] = _sys.modules[__name__]
    from pstx_local_ui import main as _local_ui_main
    return _local_ui_main()


if __name__ == '__main__':
    raise SystemExit(main())
