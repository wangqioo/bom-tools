# -*- coding: utf-8 -*-
"""
Excel 导出模块
纯 openpyxl，无 pandas 依赖
"""

import os
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────── 样式常量 ───────────────────────────────────────
_BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
_ORANGE_FILL = PatternFill("solid", fgColor="C55A11")
_GREEN_FILL  = PatternFill("solid", fgColor="375623")
_GRAY_FILL   = PatternFill("solid", fgColor="595959")
_RED_FILL    = PatternFill("solid", fgColor="FFCCCC")

_WHITE_FONT  = Font(color="FFFFFF", bold=True, size=10)
_BLACK_BOLD  = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)

_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

_THIN  = Side(style='thin')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ─────── 辅助 ───────────────────────────────────────────

def _header_row(ws, row_idx: int, fill: PatternFill):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill      = fill
            cell.font      = _WHITE_FONT
            cell.alignment = _CENTER
            cell.border    = _BORDER


def _auto_width(ws, max_w: int = 50):
    for col in ws.columns:
        vals = [str(c.value or '') for c in col]
        best = max((len(v) for v in vals), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(best + 2, max_w)


def _write_rows(ws, rows: List[dict], fill: PatternFill,
                highlight_col: Optional[str] = None, freeze: bool = True):
    """写表头 + 数据行"""
    if not rows:
        ws.append(['（无数据）'])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    _header_row(ws, ws.max_row, fill)

    hl_idx = (headers.index(highlight_col) if highlight_col in headers else None)

    for row in rows:
        ws.append(list(row.values()))
        r_idx = ws.max_row
        needs_red = False
        if hl_idx is not None:
            cell_val = str(ws.cell(r_idx, hl_idx + 1).value or '')
            if '❌' in cell_val:
                needs_red = True
        for cell in ws[r_idx]:
            cell.border    = _BORDER
            cell.alignment = _LEFT
            cell.font      = _NORMAL_FONT
            if needs_red:
                cell.fill = _RED_FILL

    _auto_width(ws)
    if freeze:
        ws.freeze_panes = 'A2'


def _section_title(ws, title: str, fill: PatternFill):
    ws.append([title])
    for cell in ws[ws.max_row]:
        cell.fill      = fill
        cell.font      = _WHITE_FONT
        cell.border    = _BORDER
    ws.append([])


# ══════════════════════════════════════════════════════════
# 主导出函数
# ══════════════════════════════════════════════════════════

def export_to_excel(data: dict, out_path: str) -> str:
    """
    data 字段:
      project_name        : str
      bom_normal_detail   : list[dict]
      bom_depop_detail    : list[dict]
      bom_normal_merged   : list[dict]
      bom_depop_merged    : list[dict]
      net_analysis        : dict
      drc                 : dict
      derating            : list[dict]
      components          : dict
    返回实际写入路径
    """
    out_path = _unique_path(out_path)
    wb = Workbook()
    wb.remove(wb.active)

    project = data.get('project_name', '')

    # 概览
    ws = wb.create_sheet('概览')
    _write_overview(ws, data, project)

    # BOM 贴装（汇总）
    ws = wb.create_sheet('BOM_贴装')
    _write_rows(ws, data.get('bom_normal_merged', []), _BLUE_FILL)

    # BOM DEPOP（汇总）
    ws = wb.create_sheet('BOM_DEPOP')
    _write_rows(ws, data.get('bom_depop_merged', []), _ORANGE_FILL)

    # BOM 明细
    ws = wb.create_sheet('BOM_明细')
    detail_all = []
    for r in data.get('bom_normal_detail', []):
        detail_all.append({'DEPOP': '', **r})
    for r in data.get('bom_depop_detail', []):
        detail_all.append({'DEPOP': 'Y', **r})
    _write_rows(ws, detail_all, _GRAY_FILL)

    # 网络分析
    ws = wb.create_sheet('网络分析')
    _write_net_analysis(ws, data.get('net_analysis', {}))

    # 设计检查
    ws = wb.create_sheet('设计检查')
    _write_drc(ws, data.get('drc', {}))

    # 降额分析
    ws = wb.create_sheet('降额分析')
    _write_rows(ws, data.get('derating', []), _BLUE_FILL, highlight_col='状态')

    wb.save(out_path)
    return out_path


# ─────── 各 Sheet 内容 ──────────────────────────────────

def _write_overview(ws, data: dict, project: str):
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16

    na  = data.get('net_analysis', {})
    drc = data.get('drc', {})
    derating = data.get('derating', [])

    normal_merged = data.get('bom_normal_merged', [])
    depop_merged  = data.get('bom_depop_merged', [])
    drc_total     = sum(len(v) for v in drc.values() if isinstance(v, list))
    derating_fail = sum(1 for r in derating if r.get('状态', '').startswith('❌'))

    rows = [
        ('项目名称',              project),
        ('贴装元件种类数',         len(normal_merged)),
        ('贴装元件总数',           sum(r.get('数量', 0) for r in normal_merged)),
        ('DEPOP 元件种类数',       len(depop_merged)),
        ('DEPOP 元件总数',         sum(r.get('数量', 0) for r in depop_merged)),
        ('网络总数',               na.get('total', '')),
        ('单端网络数（疑似漏连）', len(na.get('single_node', {}))),
        ('电源网络数',             len(na.get('power_nets', {}))),
        ('差分对数',               len(na.get('diff_pairs', {}))),
        ('DRC 问题总数',           drc_total),
        ('电容降额不合格数',       derating_fail),
    ]
    for label, val in rows:
        ws.append([label, val])
    for row in ws.iter_rows():
        for cell in row:
            cell.border    = _BORDER
            cell.font      = _BLACK_BOLD if cell.column == 1 else _NORMAL_FONT
            cell.alignment = _LEFT


def _write_net_analysis(ws, na: dict):
    ws.freeze_panes = None

    _section_title(ws, '电源网络', _BLUE_FILL)
    power = [{'网络名': k, '节点数': len(v)}
             for k, v in sorted(na.get('power_nets', {}).items(), key=lambda x: -len(x[1]))]
    _write_rows(ws, power, _BLUE_FILL, freeze=False)
    ws.append([])

    _section_title(ws, 'GND 网络', _GREEN_FILL)
    gnd = [{'网络名': k, '节点数': len(v)}
           for k, v in sorted(na.get('gnd_nets', {}).items(), key=lambda x: -len(x[1]))]
    _write_rows(ws, gnd, _GREEN_FILL, freeze=False)
    ws.append([])

    _section_title(ws, '差分对', _ORANGE_FILL)
    diff = [{'基础名': base, 'P端网络': pair['P'], 'N端网络': pair['N']}
            for base, pair in sorted(na.get('diff_pairs', {}).items())]
    _write_rows(ws, diff, _ORANGE_FILL, freeze=False)
    ws.append([])

    _section_title(ws, '单端网络（疑似漏连）', _GRAY_FILL)
    single = [{'网络名': k, '连接元件': v[0]['refdes'], '引脚': v[0]['pin_name']}
              for k, v in sorted(na.get('single_node', {}).items())]
    _write_rows(ws, single, _GRAY_FILL, freeze=False)
    ws.append([])

    _section_title(ws, '各页面元件数统计', _BLUE_FILL)
    pages = [{'页面': p, '元件数': c}
             for p, c in sorted(na.get('page_counter', {}).items())]
    _write_rows(ws, pages, _BLUE_FILL, freeze=False)

    _auto_width(ws)


def _write_drc(ws, drc: dict):
    ws.freeze_panes = None
    sections = [
        ('TBD 待确认属性', 'tbd_attrs',       _ORANGE_FILL),
        ('缺少料号',       'missing_hq_code',  _RED_FILL),
        ('缺少 VALUE',     'missing_value',     _RED_FILL),
        ('缺少封装',       'missing_package',   _RED_FILL),
        ('单端网络',       'single_pin_nets',   _GRAY_FILL),
        ('未命名网络',     'unnamed_nets',      _GRAY_FILL),
        ('BOM_OPTION 拼写', 'bom_option_typos', _ORANGE_FILL),
    ]
    for title, key, fill in sections:
        _section_title(ws, title, fill)
        rows = drc.get(key, [])
        _write_rows(ws, rows, fill, freeze=False)
        ws.append([])
    _auto_width(ws)


# ─────── 工具 ───────────────────────────────────────────

def _unique_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f'{base}({n}){ext}'
        if not os.path.exists(candidate):
            try:
                with open(candidate, 'ab'):
                    pass
                os.remove(candidate)
                return candidate
            except PermissionError:
                pass
        n += 1
