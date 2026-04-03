# -*- coding: utf-8 -*-
"""
BOM 转换工具 v5.2
格式A：品牌型号合并列（|| 或多空格分隔，如 MURATA:GRM188||SAMSUNG:CL10）
格式B：厂家/型号分开列，分号分隔（如 YAGEO;KOA / RC0805;RK73）
格式C：制造商/型号分开列，冒号分隔，制造商含编号（如 1630-大毅科技[全称]:0362-RALEC[全称]）

输出模式：
  HQ格式    → 转换为HQ内部评审BOM（整机BOM配置表）
  原格式展开 → 保留客户BOM所有列，仅将供应商拆成多行，厂商/型号各一列

依赖：pip install openpyxl
运行：python bom_gui.py
"""

import sys, subprocess, importlib, importlib.util

def _ensure(pkg, import_name=None):
    name = import_name or pkg
    if importlib.util.find_spec(name) is None:
        import tkinter as _tk; import tkinter.messagebox as _mb
        _r = _tk.Tk(); _r.withdraw()
        if not _mb.askyesno("缺少依赖",
                f"未检测到 {pkg}，是否自动安装？\n（需要联网，约几秒钟）"):
            sys.exit(0)
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

_ensure("openpyxl")

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import os, re, threading

def _unique_path(path):
    """若 path 已存在或被占用，则自动叠加 (1)(2)… 直到找到可写路径。"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f"{base}({n}){ext}"
        if not os.path.exists(candidate):
            # 再确认能写入（防止文件存在但被占用）
            try:
                with open(candidate, "ab"):
                    pass
                os.remove(candidate)
                return candidate
            except PermissionError:
                pass
        n += 1

# ───────────────────── 解析逻辑 ─────────────────────────────

SUPPLIER_LABELS = ["主供","二供","三供","四供","五供","六供","七供","八供","九供","十供"]

def parse_combined(raw):
    if not raw or str(raw).strip() == "": return []
    s = str(raw).strip().replace("：",":").replace("∥","||").replace("‖","||")
    if "||" in s:
        entries = [e.strip() for e in re.split(r"\|\|", s) if e.strip()]
    elif re.search(r'[^\s]+:[^\s]+\s{2,}[^\s]+:[^\s]+', s):
        entries = [e.strip() for e in re.split(r'\s{2,}', s) if e.strip()]
    else:
        entries = [s.strip()]
    result = []
    for entry in entries:
        if ":" in entry:
            b, m = entry.split(":", 1); result.append((b.strip(), m.strip()))
        elif "/" in entry and len(entry.split("/")) == 2:
            b, m = entry.split("/", 1); result.append((b.strip(), m.strip()))
        elif entry:
            result.append(("", entry.strip()))
    return result

def parse_split(brand_raw, model_raw):
    brands = [b.strip() for b in str(brand_raw or "").split(";") if b.strip()] if brand_raw else []
    models = [m.strip() for m in str(model_raw or "").split(";") if m.strip()] if model_raw else []
    result = []
    for i in range(max(len(brands), len(models), 1)):
        b = brands[i] if i < len(brands) else ""
        m = models[i] if i < len(models) else ""
        if b or m: result.append((b, m))
    return result

def parse_format_c(brand_raw, model_raw):
    brand_names = []
    if brand_raw:
        s = str(brand_raw).strip()
        matches = re.findall(r'\d{4}-([^\[:\]]+)\[', s)
        if matches:
            brand_names = [m.strip() for m in matches]
        else:
            brand_names = [b.strip() for b in s.split(":") if b.strip()]
    models = [m.strip() for m in str(model_raw or "").split(":") if m.strip()] if model_raw else []
    result = []
    for i in range(max(len(brand_names), len(models), 1)):
        b = brand_names[i] if i < len(brand_names) else ""
        m = models[i] if i < len(models) else ""
        if b or m: result.append((b, m))
    return result

def parse_suppliers(bv, mv, fmt):
    if fmt == "C": return parse_format_c(bv, mv)
    if fmt == "B": return parse_split(bv, mv)
    return parse_combined(bv)

def safe_qty(qv):
    try:
        q = float(qv)
        return int(q) if q == int(q) else q
    except:
        return qv if qv not in (None, "") else ""

# ───────────────────── 列检测 ─────────────────────────────

def detect_columns(ws, header_row):
    data_rows = list(range(header_row + 1, min(header_row + 11, ws.max_row + 1)))
    all_cols = {}
    for ci in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_row, column=ci).value
        hs = str(hv).strip() if hv else ""
        letter = get_column_letter(ci)
        samples = [ws.cell(row=r, column=ci).value for r in data_rows]
        strs = [str(v).strip() for v in samples if v is not None]
        role = "other"; score = 0

        b_code = sum(1 for v in strs if re.search(r'\d{4}-[^\[]+\[', v))
        if b_code >= 2 or (any(k in hs for k in ["制造商","Manufacturer"]) and "型号" not in hs and b_code >= 1):
            role = "brand_code"; score = b_code * 25 + (50 if "制造商" in hs else 0)
        m_code = sum(1 for v in strs if ":" in v and not re.search(r'\d{4}-[^\[]+\[', v) and "||" not in v)
        if "制造商型号" in hs or "Manufacturer P/N" in hs:
            if role == "other": role = "model_code"; score = 85
        elif m_code >= 3 and role == "other": role = "model_code"; score = m_code * 12

        b_comb = sum(1 for v in strs if "||" in v or re.search(r"[A-Za-z0-9]+:[A-Za-z0-9]", v))
        if role == "other" and (b_comb >= 2 or "品牌型号" in hs):
            role = "brand_combined"; score = b_comb * 20 + (40 if "品牌型号" in hs else 0)
        b_split = sum(1 for v in strs if ";" in v and not re.search(r"[A-Za-z0-9]+:[A-Za-z0-9]", v))
        if any(k in hs for k in ["厂家","厂商","Manufacturer","Brand"]) and role == "other":
            role = "brand_split"; score = 80
        elif b_split >= 3 and role == "other": role = "brand_split"; score = b_split * 15
        m_split = sum(1 for v in strs if ";" in v)
        if "型号" in hs and "品牌" not in hs and "制造商" not in hs and role == "other":
            role = "model_split"; score = 80
        elif m_split >= 3 and role == "other": role = "model_split"; score = m_split * 12

        numeric = sum(1 for v in samples if v is not None and str(v).replace(".", "").isdigit())
        if any(k in hs for k in ["用量","数量","qty","quantity","Quantity"]): role = "qty"; score = 85
        elif numeric >= len(data_rows) * 0.6 and role == "other": role = "qty"; score = numeric * 10
        avg_len = sum(len(v) for v in strs) / max(len(strs), 1)
        if any(k in hs for k in ["名称","品名","物料名","描述","项目描述","description","Description"]):
            if role == "other": role = "name"; score = 75
        elif avg_len > 8 and role == "other": role = "name"; score = int(avg_len * 2)
        all_cols[ci] = {"letter": letter, "header": hs, "role": role, "score": score, "sample": strs[:3]}
    best = {}
    for ci, info in all_cols.items():
        r = info["role"]
        if r != "other" and (r not in best or info["score"] > best[r]["score"]):
            best[r] = {"ci": ci, **info}
    return all_cols, best

# ───────────────────── 输出：HQ格式 ─────────────────────────

def write_review_bom(rows, output_file, project_name):
    wb = Workbook(); ws = wb.active; ws.title = "SW节点整机BOM配置"
    GREEN = "92D050"; YELLOW = "FFFF00"; ORANGE = "FFC000"
    thin = Side(style="thin"); bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    def S(cell, bold=False, bg=None, color="000000", h="center", v="center", size=11):
        cell.font = Font(bold=bold, color=color, size=size)
        if bg: cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=h, vertical=v)
    ws.merge_cells("A1:A2"); ws["A1"] = "项目名称"; S(ws["A1"], bold=True, bg=GREEN, size=14)
    ws.merge_cells("B1:B2"); ws["B1"] = project_name; S(ws["B1"], bold=True, bg=GREEN, size=14)
    ws.merge_cells("E1:I2"); ws["E1"] = "整机BOM配置表"; S(ws["E1"], bold=True, bg=GREEN, size=16)
    ws["J1"] = "配置说明"; S(ws["J1"], bold=True, bg=GREEN)
    ws["K1"] = "TBD"; S(ws["K1"], bg="BDD7EE")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A3:I3"); ws["A3"] = "SW节点HQ SN"
    S(ws["A3"], bold=True, bg=YELLOW, color="FF0000", size=12)
    ws["K3"] = ""; S(ws["K3"], bg=ORANGE, color="FF0000"); ws.row_dimensions[3].height = 20
    headers = ["序号","组件子类","虚拟层/物料","物料类型","HQ PN","物料名称","厂商型号","厂商","主二供","","用量"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h); S(c, bold=True, bg="D9D9D9"); c.border = bdr
    ws.row_dimensions[4].height = 22
    dr = 5
    for item in rows:
        for si, (brand, model, qty) in enumerate(item["suppliers"]):
            label = SUPPLIER_LABELS[si] if si < len(SUPPLIER_LABELS) else f"{si+1}供"
            for ci, val in enumerate([item["seq"],"","","","",item["name"],model,brand,label,"",qty], 1):
                c = ws.cell(row=dr, column=ci, value=val); c.border = bdr
                c.alignment = Alignment(horizontal="center", vertical="center")
            dr += 1
    for i, w in enumerate([6,10,12,10,18,35,30,20,8,6,8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(output_file); return dr - 5

# ───────────────────── 输出：原格式展开 ──────────────────────

def write_expanded_bom(ws_in, header_row, col_brand, col_model, col_qty, fmt, out_file):
    """
    保留客户BOM所有列，将供应商信息拆成多行。
    格式A：品牌型号合并列拆成 厂商 + 型号 两列（在原位置展开）。
    格式B/C：厂家列和型号列已分开，各行写入对应供应商的厂商和型号。
    主供保留原用量，替代料用量写0。
    """
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="D9D9D9")
    max_col = ws_in.max_column

    # 构建输出列映射：list of (type, src_ci, header)
    # type: "seq"=序号, "orig"=原样复制, "brand"=写厂商, "model"=写型号
    out_map = [("seq", None, "序号")]   # 第一列固定为序号
    for ci in range(1, max_col + 1):
        h = ws_in.cell(row=header_row, column=ci).value or ""
        if fmt == "A":
            if ci == col_brand:
                out_map.append(("brand", ci, "厂商"))
                out_map.append(("model", None, "型号"))   # 插入新列
            else:
                out_map.append(("orig", ci, str(h)))
        else:  # B 或 C
            if ci == col_brand:
                out_map.append(("brand", ci, "厂商"))
            elif col_model and ci == col_model:
                out_map.append(("model", ci, "型号"))
            else:
                out_map.append(("orig", ci, str(h)))

    # 写表头
    for out_ci, (typ, _, h) in enumerate(out_map, 1):
        c = ws_out.cell(row=1, column=out_ci, value=h)
        c.font = Font(bold=True)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
        ws_out.column_dimensions[get_column_letter(out_ci)].width = 6 if typ == "seq" else 18

    # 写数据行
    dr = 2; total = 0; skipped = 0; seq = 0
    for ri in range(header_row + 1, ws_in.max_row + 1):
        row_vals = {ci: ws_in.cell(row=ri, column=ci).value for ci in range(1, max_col + 1)}
        if not any(v is not None and str(v).strip() for v in row_vals.values()):
            skipped += 1; continue

        bv = row_vals.get(col_brand)
        mv = row_vals.get(col_model) if col_model else None
        qv = row_vals.get(col_qty)
        suppliers = parse_suppliers(bv, mv, fmt)
        if not suppliers: suppliers = [("", "")]
        mq = safe_qty(qv)
        seq += 1  # 同一组替代料共享同一序号

        for si, (brand, model) in enumerate(suppliers):
            for out_ci, (typ, src_ci, _) in enumerate(out_map, 1):
                if typ == "seq":
                    val = seq
                elif si == 0:
                    # 主供：填所有列
                    if typ == "brand":   val = brand
                    elif typ == "model": val = model
                    else: val = mq if src_ci == col_qty else row_vals.get(src_ci)
                else:
                    # 替代料：只填厂商、型号，其余留空（用量也留空）
                    if typ == "brand":   val = brand
                    elif typ == "model": val = model
                    else:                val = None
                c = ws_out.cell(row=dr, column=out_ci, value=val)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = bdr
            dr += 1; total += 1

    wb_out.save(out_file)
    return total, skipped

# ───────────────────── GUI ───────────────────────────────────

class BomApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BOM 转换工具 v5.2")
        self.geometry("820x700")
        self.resizable(True, True)

        self.wb = None; self.ws = None
        self.col_name_var    = tk.StringVar()
        self.col_qty_var     = tk.StringVar()
        self.col_brand_var   = tk.StringVar()
        self.col_model_var   = tk.StringVar()
        self.header_row_var  = tk.IntVar(value=1)
        self.project_var     = tk.StringVar()
        self.output_var      = tk.StringVar(value="展开多行BOM.xlsx")
        self.input_path      = tk.StringVar()
        self.sheet_var       = tk.StringVar()
        self.fmt_var         = tk.StringVar(value="auto")
        self.output_mode_var = tk.StringVar(value="expand")  # "expand" | "hq"

        self.output_mode_var.trace_add("write", self._on_mode_change)
        self.fmt_var.trace_add("write", self._on_fmt_change)
        self._build_ui()

    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=8)
        self.nb = nb
        self.tab1 = ttk.Frame(nb); nb.add(self.tab1, text="  第一步：选择文件  ")
        self.tab2 = ttk.Frame(nb); nb.add(self.tab2, text="  第二步：列映射  ")
        self.tab3 = ttk.Frame(nb); nb.add(self.tab3, text="  第三步：输出设置  ")
        self.tab4 = ttk.Frame(nb); nb.add(self.tab4, text="  日志  ")
        self._build_tab1(); self._build_tab2(); self._build_tab3(); self._build_tab4()

    def _section(self, parent, title):
        f = ttk.LabelFrame(parent, text=title, padding=10)
        f.pack(fill="x", padx=12, pady=5)
        return f

    # ── Tab1 ──────────────────────────────────────────────────

    def _build_tab1(self):
        p = self.tab1
        f1 = self._section(p, "客户 BOM 文件")
        tk.Label(f1, text="文件路径：").grid(row=0, column=0, sticky="w")
        ttk.Entry(f1, textvariable=self.input_path, width=52).grid(row=0, column=1, padx=6)
        ttk.Button(f1, text="浏览...", command=self._browse_file).grid(row=0, column=2)

        f2 = self._section(p, "Sheet / 表头行")
        tk.Label(f2, text="Sheet：").grid(row=0, column=0, sticky="w")
        self.sheet_cb = ttk.Combobox(f2, textvariable=self.sheet_var, width=28, state="readonly")
        self.sheet_cb.grid(row=0, column=1, padx=6, sticky="w")
        self.sheet_cb.bind("<<ComboboxSelected>>", lambda e: self._load_sheet())
        tk.Label(f2, text="  表头行：").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(f2, from_=1, to=10, textvariable=self.header_row_var, width=5).grid(row=0, column=3)
        ttk.Button(f2, text="重新扫描", command=self._scan_columns).grid(row=0, column=4, padx=8)

        f3 = self._section(p, "文件预览（前5行）")
        self.preview_tree = ttk.Treeview(f3, height=5, show="headings")
        sx = ttk.Scrollbar(f3, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(xscrollcommand=sx.set)
        self.preview_tree.pack(fill="x"); sx.pack(fill="x")

        ttk.Button(p, text="下一步：确认列映射 →",
                   command=lambda: self.nb.select(1)).pack(anchor="e", padx=12, pady=8)

    # ── Tab2 ──────────────────────────────────────────────────

    def _build_tab2(self):
        p = self.tab2
        ff = self._section(p, "品牌型号格式")
        row = tk.Frame(ff); row.pack(anchor="w")
        for val, txt in [("auto","自动识别"),
                          ("A","格式A：合并列（MURATA:GRM188||SAMSUNG:CL10 或多空格）"),
                          ("B","格式B：分开列，分号分隔（YAGEO;KOA / RC0805;RK73）"),
                          ("C","格式C：分开列，冒号分隔（1630-大毅[全称]:0362-RALEC[全称]）")]:
            ttk.Radiobutton(row, text=txt, variable=self.fmt_var,
                            value=val).pack(side="left", padx=8)

        fm = self._section(p, "列位置（填列字母，如 A / D / G）")
        rows_def = [
            ("物料名称列",  self.col_name_var,  "物料品名/描述所在列"),
            ("用量列",      self.col_qty_var,   "用量/数量所在列"),
            ("品牌/厂家列", self.col_brand_var, "格式A=品牌型号合并；格式B=厂家列（分号）；格式C=制造商列（含编号）"),
            ("型号列",      self.col_model_var, "格式B/C需填写；格式A无需填写（自动拆分）"),
        ]
        for i, (lbl, var, hint) in enumerate(rows_def):
            tk.Label(fm, text=lbl+"：", anchor="w", width=14).grid(row=i, column=0, sticky="w", pady=3)
            e = ttk.Entry(fm, textvariable=var, width=8)
            e.grid(row=i, column=1, padx=6)
            if lbl == "型号列":
                self.col_model_entry = e
                self.col_model_hint  = tk.Label(fm, text=hint, fg="#666", wraplength=500, justify="left")
                self.col_model_hint.grid(row=i, column=2, sticky="w", padx=6)
            else:
                tk.Label(fm, text=hint, fg="#666", wraplength=500,
                         justify="left").grid(row=i, column=2, sticky="w", padx=6)

        f2 = self._section(p, "自动扫描结果")
        self.detect_text = tk.Text(f2, height=10, font=("Consolas", 9),
                                    state="disabled", bg="#fafafa", relief="flat")
        self.detect_text.pack(fill="x")

        ttk.Button(p, text="下一步：设置输出 →",
                   command=lambda: self.nb.select(2)).pack(anchor="e", padx=12, pady=8)

    # ── Tab3 ──────────────────────────────────────────────────

    def _build_tab3(self):
        p = self.tab3

        fm = self._section(p, "输出模式")
        for val, txt, desc in [
            ("expand", "原格式展开", "保留客户BOM所有列，将供应商拆成多行，厂商/型号各一列"),
            ("hq",     "转为HQ格式",  "输出为 整机BOM配置表（需填写项目名称）"),
        ]:
            r = tk.Frame(fm); r.pack(anchor="w", pady=2)
            ttk.Radiobutton(r, text=txt, variable=self.output_mode_var,
                            value=val).pack(side="left")
            tk.Label(r, text="  "+desc, fg="#555").pack(side="left")

        self.hq_frame = self._section(p, "项目信息（仅HQ格式需要）")
        tk.Label(self.hq_frame, text="项目名称：").grid(row=0, column=0, sticky="w")
        ttk.Entry(self.hq_frame, textvariable=self.project_var, width=42).grid(
            row=0, column=1, padx=6, sticky="w")

        f2 = self._section(p, "输出文件")
        tk.Label(f2, text="输出文件名：").grid(row=0, column=0, sticky="w")
        ttk.Entry(f2, textvariable=self.output_var, width=42).grid(row=0, column=1, padx=6, sticky="w")
        ttk.Button(f2, text="另存为...", command=self._browse_output).grid(row=0, column=2, padx=6)

        self.run_btn = tk.Button(p, text="开始转换", font=("Arial", 13, "bold"),
                                  bg="#2d6cdf", fg="white", relief="flat",
                                  padx=20, pady=10, command=self._run_convert)
        self.run_btn.pack(pady=16)
        self.status_label = tk.Label(p, text="", font=("Arial", 11))
        self.status_label.pack()

        self._on_mode_change()

    # ── Tab4 ──────────────────────────────────────────────────

    def _build_tab4(self):
        p = self.tab4
        self.log = scrolledtext.ScrolledText(p, font=("Consolas", 9), state="disabled",
                                              bg="#1e1e1e", fg="#d4d4d4", relief="flat")
        self.log.pack(fill="both", expand=True, padx=8, pady=8)
        ttk.Button(p, text="清空日志", command=self._clear_log).pack(anchor="e", padx=8, pady=4)

    # ── 事件 ─────────────────────────────────────────────────

    def _on_fmt_change(self, *_):
        """格式A：型号列不需要，自动清空并禁用；B/C/auto：启用。"""
        if not hasattr(self, "col_model_entry"): return
        fmt = self.fmt_var.get()
        if fmt == "A":
            self.col_model_var.set("")
            self.col_model_entry.configure(state="disabled")
            self.col_model_hint.configure(text="格式A合并列，无需填写", fg="#aaa")
        else:
            self.col_model_entry.configure(state="normal")
            self.col_model_hint.configure(
                text="格式B/C需填写；格式A无需填写（自动拆分）", fg="#666")

    def _default_out_path(self, mode=None):
        """根据当前模式和输入文件路径，生成默认输出路径。"""
        if mode is None:
            mode = self.output_mode_var.get()
        name = "内部评审BOM.xlsx" if mode == "hq" else "展开多行BOM.xlsx"
        in_path = self.input_path.get().strip()
        if in_path:
            return os.path.join(os.path.dirname(in_path), name)
        return name

    def _on_mode_change(self, *_):
        if not hasattr(self, "hq_frame"): return
        mode = self.output_mode_var.get()
        if mode == "hq":
            self.hq_frame.configure(style="")
            for w in self.hq_frame.winfo_children():
                w.configure(state="normal")
        else:
            for w in self.hq_frame.winfo_children():
                try: w.configure(state="disabled")
                except: pass
        self.output_var.set(self._default_out_path(mode))

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="选择客户BOM",
            filetypes=[("Excel文件","*.xlsx *.xlsm *.xls"),("所有文件","*.*")])
        if not path: return
        self.input_path.set(path)
        self.output_var.set(self._default_out_path())   # 自动对应输入文件目录
        self._log(f"文件：{path}")
        try:
            self.wb = openpyxl.load_workbook(path, data_only=True)
            self.sheet_cb["values"] = self.wb.sheetnames
            self.sheet_var.set(self.wb.sheetnames[0])
            self._load_sheet()
        except Exception as e:
            messagebox.showerror("错误", f"无法打开文件：\n{e}")

    def _load_sheet(self):
        if not self.wb: return
        self.ws = self.wb[self.sheet_var.get()]
        self._log(f"Sheet：{self.ws.title}（{self.ws.max_row}行 × {self.ws.max_column}列）")
        self._update_preview(); self._scan_columns()

    def _update_preview(self):
        tree = self.preview_tree; tree.delete(*tree.get_children())
        if not self.ws: return
        mc = min(self.ws.max_column, 9)
        cols = [get_column_letter(i) for i in range(1, mc + 1)]
        tree["columns"] = cols
        for c in cols: tree.heading(c, text=c); tree.column(c, width=110, anchor="w")
        for ri in range(1, min(6, self.ws.max_row + 1)):
            vals = [str(self.ws.cell(row=ri, column=ci).value or "")[:26] for ci in range(1, mc+1)]
            tree.insert("", "end", values=vals)

    def _scan_columns(self):
        if not self.ws: messagebox.showwarning("提示","请先选择文件"); return
        hr = self.header_row_var.get()
        all_cols, best = detect_columns(self.ws, hr)
        role_label = {
            "brand_combined": "✅ 品牌型号(合并A)",
            "brand_split":    "✅ 厂家(分开B)",
            "model_split":    "✅ 型号(分开B)",
            "brand_code":     "✅ 制造商(编号C)",
            "model_code":     "✅ 制造商型号(C)",
            "qty":            "✅ 用量",
            "name":           "✅ 物料名称",
            "other":          "   -",
        }
        self.detect_text.configure(state="normal")
        self.detect_text.delete("1.0", "end")
        lines = [f"{'列':<4} {'表头':<22} {'识别用途':<24} 样本\n" + "─"*70]
        for ci in sorted(all_cols.keys()):
            info = all_cols[ci]
            rz = role_label.get(info["role"], "   -")
            sample = " | ".join(info["sample"][:2])[:32]
            lines.append(f" {info['letter']:<4} {info['header']:<22} {rz:<24} {sample}")
        self.detect_text.insert("end", "\n".join(lines))
        self.detect_text.configure(state="disabled")

        if "name"  in best: self.col_name_var.set(best["name"]["letter"])
        if "qty"   in best: self.col_qty_var.set(best["qty"]["letter"])
        if "brand_code" in best:
            self.col_brand_var.set(best["brand_code"]["letter"])
            if "model_code" in best: self.col_model_var.set(best["model_code"]["letter"])
            self.fmt_var.set("C")
        elif "brand_combined" in best:
            self.col_brand_var.set(best["brand_combined"]["letter"])
            self.col_model_var.set(""); self.fmt_var.set("A")
        elif "brand_split" in best:
            self.col_brand_var.set(best["brand_split"]["letter"])
            if "model_split" in best: self.col_model_var.set(best["model_split"]["letter"])
            self.fmt_var.set("B")

        self._log(f"扫描完成 → 名称={self.col_name_var.get()} 用量={self.col_qty_var.get()} "
                  f"品牌/厂家={self.col_brand_var.get()} 型号={self.col_model_var.get() or '-'} "
                  f"格式={self.fmt_var.get()}")

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel文件","*.xlsx")])
        if path: self.output_var.set(path)

    def _run_convert(self):
        if not self.ws:
            messagebox.showerror("错误","请先选择输入文件"); return
        if not self.col_brand_var.get():
            messagebox.showerror("错误","请确认列映射（第二步）"); return
        if self.output_mode_var.get() == "hq" and not self.project_var.get().strip():
            messagebox.showerror("错误","HQ格式需要填写项目名称（第三步）"); return
        self.run_btn.configure(state="disabled")
        self.status_label.configure(text="转换中...", fg="#2d6cdf")
        self.nb.select(3)
        threading.Thread(target=self._do_convert, daemon=True).start()

    def _do_convert(self):
        try:
            hr        = self.header_row_var.get()
            col_brand = column_index_from_string(self.col_brand_var.get().upper())
            col_model = column_index_from_string(self.col_model_var.get().upper()) \
                        if self.col_model_var.get().strip() else None
            col_qty   = column_index_from_string(self.col_qty_var.get().upper()) \
                        if self.col_qty_var.get().strip() else None
            col_name  = column_index_from_string(self.col_name_var.get().upper()) \
                        if self.col_name_var.get().strip() else None
            fmt       = self.fmt_var.get()
            out_file  = self.output_var.get().strip()
            mode      = self.output_mode_var.get()

            # 自动推断格式
            if fmt == "auto":
                if col_model:
                    sample = str(self.ws.cell(row=hr+1, column=col_brand).value or "")
                    fmt = "C" if re.search(r'\d{4}-[^\[]+\[', sample) else "B"
                else:
                    fmt = "A"

            # 若同名文件已存在或被占用，自动在文件名后叠加序号
            out_file = _unique_path(out_file)

            self._log(f"\n开始转换（格式{fmt}，模式={'原格式展开' if mode=='expand' else 'HQ格式'}）")

            if mode == "expand":
                if not col_qty:
                    messagebox.showerror("错误","原格式展开需要指定用量列"); return
                total, skipped = write_expanded_bom(
                    self.ws, hr, col_brand, col_model, col_qty, fmt, out_file)
                self._log(f"跳过空行：{skipped}")
                self._log(f"共写入 {total} 行")
            else:
                # HQ格式：需要 col_name
                project = self.project_var.get().strip()
                rows = []; seq = 0; skipped = 0
                for ri in range(hr + 1, self.ws.max_row + 1):
                    nv = self.ws.cell(row=ri, column=col_name).value if col_name else ""
                    qv = self.ws.cell(row=ri, column=col_qty).value if col_qty else ""
                    bv = self.ws.cell(row=ri, column=col_brand).value
                    mv = self.ws.cell(row=ri, column=col_model).value if col_model else None
                    if not nv and not bv: skipped += 1; continue
                    sr = parse_suppliers(bv, mv, fmt)
                    if not sr: sr = [("", "")]
                    mq = safe_qty(qv)
                    suppliers = [(b, m, mq if i == 0 else 0) for i, (b, m) in enumerate(sr)]
                    seq += 1
                    rows.append({"seq": seq, "name": str(nv or "").strip(), "suppliers": suppliers})
                self._log(f"解析：{len(rows)} 个物料（跳过空行 {skipped}）")
                total = write_review_bom(rows, out_file, project)
                self._log(f"共写入 {total} 行")

            abs_path = os.path.abspath(out_file)
            self._log(f"输出：{abs_path}\n✅ 转换成功！")
            self.after(0, lambda: self.status_label.configure(
                text=f"✅ 完成！共 {total} 行", fg="#2a8a2a"))
            self.after(0, lambda: messagebox.showinfo("完成", f"转换成功！\n{abs_path}"))

        except Exception as e:
            import traceback
            self._log(f"\n❌ 错误：{e}\n{traceback.format_exc()}")
            self.after(0, lambda: self.status_label.configure(text="❌ 转换失败，请查看日志", fg="red"))
            self.after(0, lambda: messagebox.showerror("错误", str(e)))
        finally:
            self.after(0, lambda: self.run_btn.configure(state="normal"))

    def _log(self, msg):
        def _w():
            self.log.configure(state="normal")
            self.log.insert("end", msg + "\n"); self.log.see("end")
            self.log.configure(state="disabled")
        self.after(0, _w)

    def _clear_log(self):
        self.log.configure(state="normal"); self.log.delete("1.0", "end")
        self.log.configure(state="disabled")


if __name__ == "__main__":
    app = BomApp()
    app.mainloop()
