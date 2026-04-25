import os
import tempfile
import unittest
from pathlib import Path, PureWindowsPath

from openpyxl import load_workbook
from openpyxl import Workbook

import pstx_analyzer
import pstx_csa_geometry
import pstx_page_logic


def make_cap(refdes='C1', rated='6.3V', bom_option='', page='PAGE1', page_real=''):
    display_page = page_real or ''
    return {
        'refdes': refdes,
        'part_name': 'CAP_0402',
        'hq_code': '',
        'value': '0.1uF',
        'package': '0402',
        'material': '',
        'tolerance': '',
        'voltage': rated,
        'current': '',
        'power': '',
        'bom_option': bom_option,
        'bom_cost': '',
        'room': '',
        'drawing': 'SCH_PAGE1',
        'page': display_page,
        'page_logical': page,
        'page_real': display_page,
        'comp_type': 'CAP',
        'nets': {},
    }


def make_ic(refdes='U1', nets=None, bom_option='', page='PAGE1', page_real=''):
    display_page = page_real or ''
    return {
        'refdes': refdes,
        'part_name': 'IC_CPU',
        'hq_code': 'PN_IC',
        'value': 'CPU',
        'package': 'BGA',
        'material': '',
        'tolerance': '',
        'voltage': '',
        'current': '',
        'power': '',
        'bom_option': bom_option,
        'bom_cost': '',
        'room': '',
        'drawing': 'SCH_PAGE1',
        'page': display_page,
        'page_logical': page,
        'page_real': display_page,
        'comp_type': 'IC',
        'nets': nets or {},
    }


def make_res(refdes, net_a, net_b, value='10k', bom_option='', page='PAGE1', page_real=''):
    display_page = page_real or ''
    return {
        'refdes': refdes,
        'part_name': 'RES_0402',
        'hq_code': '',
        'value': value,
        'package': '0402',
        'material': '',
        'tolerance': '',
        'voltage': '',
        'current': '',
        'power': '',
        'bom_option': bom_option,
        'bom_cost': '',
        'room': '',
        'drawing': 'SCH_PAGE1',
        'page': display_page,
        'page_logical': page,
        'page_real': display_page,
        'comp_type': 'RES',
        'nets': {'1': net_a, '2': net_b},
    }


CSA_PAGE_T_WITH_DOT = (
    "FILE_TYPE = MACRO_DRAWING;\n"
    "SET PAGE_NUMBER P1;\n"
    "WIRE 16 -1 (0 0)(100 0);\n"
    "FORCEPROP 2 LAST SIG_NAME T_H\n"
    "WIRE 16 -1 (50 0)(50 100);\n"
    "FORCEPROP 2 LAST SIG_NAME T_V\n"
    "DOT 1 (50 0);\n"
    "CIRCLE 16 -1 (1000 1000)(1100 1000);\n"
)

CSA_PAGE_DOTLESS_CROSS = (
    "FILE_TYPE = MACRO_DRAWING;\n"
    "SET PAGE_NUMBER P2;\n"
    "WIRE 16 -1 (200 0)(300 0);\n"
    "FORCEPROP 2 LAST SIG_NAME CROSS_NO_DOT_H\n"
    "WIRE 16 -1 (250 -50)(250 50);\n"
    "FORCEPROP 2 LAST SIG_NAME CROSS_NO_DOT_V\n"
    "CIRCLE 16 -1 (2000 2000) 150;\n"
)

CSA_PAGE_DOT_CROSS = (
    "FILE_TYPE = MACRO_DRAWING;\n"
    "SET PAGE_NUMBER P3;\n"
    "WIRE 16 -1 (400 0)(500 0);\n"
    "FORCEPROP 2 LAST SIG_NAME CROSS_DOT_H\n"
    "WIRE 16 -1 (450 -50)(450 50);\n"
    "FORCEPROP 2 LAST SIG_NAME CROSS_DOT_V\n"
    "DOT 1 (450 0);\n"
)

CSA_PAGE_SPLIT_CROSS_WITH_ARC = (
    "FILE_TYPE = MACRO_DRAWING;\n"
    "SET PAGE_NUMBER P4;\n"
    "WIRE 16 -1 (600 0)(650 0);\n"
    "FORCEPROP 2 LAST SIG_NAME SPLIT_CROSS_L\n"
    "WIRE 16 -1 (650 0)(700 0);\n"
    "FORCEPROP 2 LAST SIG_NAME SPLIT_CROSS_R\n"
    "WIRE 16 -1 (650 -50)(650 0);\n"
    "FORCEPROP 2 LAST SIG_NAME SPLIT_CROSS_D\n"
    "WIRE 16 -1 (650 0)(650 50);\n"
    "FORCEPROP 2 LAST SIG_NAME SPLIT_CROSS_U\n"
    "DOT 1 (650 0);\n"
    "ARC 16 -1 (3000 3000)(3100 3000)(3050 3050);\n"
)


class CsaGeometryTests(unittest.TestCase):
    def test_csa_geometry_matches_reference_demo_rules(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir()
            (sch_dir / 'page1.csa').write_text(CSA_PAGE_T_WITH_DOT, encoding='utf-8')
            (sch_dir / 'page2.csa').write_text(CSA_PAGE_DOTLESS_CROSS, encoding='utf-8')
            (sch_dir / 'page3.csa').write_text(CSA_PAGE_DOT_CROSS, encoding='utf-8')
            (sch_dir / 'page4.csa').write_text(CSA_PAGE_SPLIT_CROSS_WITH_ARC, encoding='utf-8')

            result = pstx_csa_geometry.analyze_csa_geometry(root)

        self.assertTrue(result['enabled'])
        self.assertEqual(4, result['page_count'])
        self.assertEqual(2, result['cross_count'])
        self.assertEqual(3, result['circle_count'])
        self.assertEqual(['PAGE3', 'PAGE4'], [row['页面'] for row in result['dot_cross_rows']])
        self.assertEqual(['(450,0)', '(650,0)'], [row['坐标'] for row in result['dot_cross_rows']])
        self.assertEqual(0, next(row for row in result['summary_rows'] if row['页面'] == 'PAGE1')['DOT四向十字数'])
        self.assertEqual(0, next(row for row in result['summary_rows'] if row['页面'] == 'PAGE2')['DOT四向十字数'])

    def test_analyze_project_contents_includes_csa_geometry_when_sch1_has_csa(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / 'sch_1').mkdir()
            (root / 'sch_1' / 'page3.csa').write_text(CSA_PAGE_DOT_CROSS, encoding='utf-8')
            bundle = pstx_analyzer.analyze_project_contents(
                "PART_NAME\nU1 'IC_CPU'\nHQ_CODE='PN'\nVALUE='CPU'\nPACKAGE='BGA'\n",
                "NET_NAME\n'N1'\nNODE_NAME U1 1\n'PIN1':\n",
                project_name='demo',
                project_root=str(root),
            )

        self.assertEqual(1, bundle['csa_geometry']['page_count'])
        self.assertEqual(1, bundle['csa_geometry']['cross_count'])
        self.assertEqual('PAGE3', bundle['csa_geometry']['dot_cross_rows'][0]['页面'])


class ParseTests(unittest.TestCase):
    def test_parse_pstxprt_handles_marker_at_file_start(self):
        content = (
            "PART_NAME\n"
            "C1 'CAP_0402'\n"
            "VALUE='1uF'\n"
            "PACKAGE='0402'\n"
            "DRAWING='SCH_PAGE1'\n"
        )
        components = pstx_analyzer.parse_pstxprt(content)
        self.assertIn('C1', components)
        self.assertEqual('1uF', components['C1']['value'])

    def test_parse_pstxprt_handles_crlf_newlines(self):
        content = (
            "PART_NAME\r\n"
            "C1 'CAP_0402'\r\n"
            "VALUE='1uF'\r\n"
            "PACKAGE='0402'\r\n"
            "DRAWING='SCH_PAGE1'\r\n"
        )
        components = pstx_analyzer.parse_pstxprt(content)
        self.assertIn('C1', components)

    def test_parse_pstxprt_normalizes_page_tokens_with_separator_and_suffix(self):
        content = (
            "PART_NAME\n"
            "U1 'IC_CPU'\n"
            "DRAWING='ROOT/PAGE_02A'\n"
        )
        components = pstx_analyzer.parse_pstxprt(content)
        self.assertEqual('', components['U1']['page'])
        self.assertEqual('PAGE2A', components['U1']['page_logical'])

    def test_parse_pstxprt_preserves_hierarchical_page_chain(self):
        content = (
            "PART_NAME\n"
            "U1 'IC_CPU'\n"
            "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1"
            "@GPU_2SW_BOARD_LIB.HQPWR_EFUSE_TPS259260_12VIN_4A(SCH_1):PAGE1_I14'\n"
        )
        components = pstx_analyzer.parse_pstxprt(content)
        self.assertEqual('', components['U1']['page'])
        self.assertEqual('PAGE242', components['U1']['page_logical'])

    def test_parse_pstxprt_prefers_section_path_over_submodule_drawing_for_page_source(self):
        content = (
            "PART_NAME\n"
            "C1A104 'CAP_HDL-HQ17101005HS0,100NF,10%,0402,X7R,50V':\n"
            "SECTION_NUMBER 1\n"
            " '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE144_I70"
            "@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE1_I17"
            "@HQ_CAP.CAP_HDL(CHIPS)':\n"
            " C_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70"
            "@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page1_i17"
            "@hq_cap.cap_hdl(chips)',\n"
            " PATH='I17',\n"
            " DRAWING='@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE1',\n"
            " PHYS_PAGE='1',\n"
            " BOM_OPTION='DEPOP',\n"
            " PACKAGE='0402',\n"
            " HQ_CODE='HQ17101005HS0',\n"
            " VALUE='100NF'\n"
        )
        components = pstx_analyzer.parse_pstxprt(content)
        self.assertEqual('section_path', components['C1A104']['page_path_source'])
        self.assertEqual(
            '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE144_I70'
            '@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE1_I17'
            '@HQ_CAP.CAP_HDL(CHIPS)',
            components['C1A104']['page_path_raw'],
        )
        self.assertEqual('PAGE144', components['C1A104']['page_logical'])

    def test_resolve_component_pages_prefers_outer_schematic_segment(self):
        components = {
            'U1': {
                'refdes': 'U1',
                'drawing': (
                    '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'
                    '@GPU_2SW_BOARD_LIB.HQPWR_EFUSE_TPS259260_12VIN_4A(SCH_1):PAGE1_I14'
                    '@HQ_IC.TPS259261DRCR_11P_HDL(CHIPS)'
                ),
                'page': 'PAGE242 / PAGE1',
            },
        }
        warnings = pstx_analyzer.resolve_component_pages(components)
        self.assertEqual([], warnings)
        self.assertEqual('', components['U1']['page'])
        self.assertEqual('PAGE242', components['U1']['page_logical'])
        self.assertEqual('PAGE242', components['U1']['page_raw'])
        self.assertEqual('', components['U1']['page_real'])
        self.assertEqual('GPU_2SW_BOARD:PAGE242', components['U1']['page_context'])
        self.assertEqual('', components['U1']['page_context_real'])
        self.assertEqual('drawing', components['U1']['page_source'])
        self.assertEqual('none', components['U1']['page_real_source'])
        self.assertEqual('', components['U1']['page_mapping_ok'])

    def test_resolve_component_pages_uses_page_csv_mapping_for_outer_schematic_segment(self):
        components = {
            'U1': {
                'refdes': 'U1',
                'drawing': (
                    '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'
                    '@GPU_2SW_BOARD_LIB.HQPWR_EFUSE_TPS259260_12VIN_4A(SCH_1):PAGE1_I14'
                    '@HQ_IC.TPS259261DRCR_11P_HDL(CHIPS)'
                ),
                'page': 'PAGE242 / PAGE1',
            },
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            top_sch = root / 'GPU_2SW_BOARD' / 'sch_1'
            top_sch.mkdir(parents=True)
            (top_sch / 'page518.csv').write_text('NAME,PAGE_NUMBER\nTOP,242\n', encoding='utf-8')

            child_sch = root / 'HQPWR_EFUSE_TPS259260_12VIN_4A' / 'sch_1'
            child_sch.mkdir(parents=True)
            (child_sch / 'page12.csv').write_text('NAME,PAGE_NUMBER\nCHILD,1\n', encoding='utf-8')

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        self.assertEqual('PAGE518', components['U1']['page'])
        self.assertEqual('PAGE242', components['U1']['page_logical'])
        self.assertEqual('PAGE518', components['U1']['page_real'])
        self.assertEqual('drawing', components['U1']['page_source'])
        self.assertEqual('page_csv', components['U1']['page_real_source'])
        self.assertEqual('是', components['U1']['page_mapping_ok'])
        self.assertEqual('GPU_2SW_BOARD:PAGE242', components['U1']['page_context'])
        self.assertEqual('GPU_2SW_BOARD:PAGE518', components['U1']['page_context_real'])

    def test_resolve_component_pages_ignores_child_page_when_child_segment_appears_first(self):
        components = {
            'U1': {
                'refdes': 'U1',
                'drawing': (
                    '@GPU_2SW_BOARD_LIB.HQPWR_EFUSE_TPS259260_12VIN_4A(SCH_1):PAGE1_I14'
                    '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'
                    '@HQ_IC.TPS259261DRCR_11P_HDL(CHIPS)'
                ),
                'page': 'PAGE1 / PAGE242',
            },
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / 'GPU_2SW_BOARD'
            root.mkdir(parents=True)
            top_sch = root / 'sch_1'
            top_sch.mkdir(parents=True)
            (top_sch / 'page518.csv').write_text('"PAGE_NUMBER" = 242;', encoding='utf-8')
            (top_sch / 'page12.csv').write_text('"PAGE_NUMBER" = 1;', encoding='utf-8')

            child_sch = root / 'HQPWR_EFUSE_TPS259260_12VIN_4A' / 'sch_1'
            child_sch.mkdir(parents=True)
            (child_sch / 'page12.csv').write_text('"PAGE_NUMBER" = 1;', encoding='utf-8')

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        self.assertEqual('PAGE242', components['U1']['page_logical'])
        self.assertEqual('PAGE518', components['U1']['page'])
        self.assertEqual('PAGE518', components['U1']['page_real'])
        self.assertEqual('page_csv', components['U1']['page_real_source'])

    def test_resolve_component_pages_reads_direct_project_root_sch1_pages(self):
        components = {
            'U1': {
                'refdes': 'U1',
                'drawing': '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1',
                'page': 'PAGE242',
            },
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            direct_sch = root / 'sch_1'
            direct_sch.mkdir(parents=True)
            (direct_sch / 'page518.csv').write_text('NAME,PAGE_NUMBER\nTOP,242\n', encoding='utf-8')

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        self.assertEqual('PAGE518', components['U1']['page'])
        self.assertEqual('PAGE518', components['U1']['page_real'])
        self.assertEqual('page_csv', components['U1']['page_real_source'])

    def test_infer_project_root_from_packaged_data_paths(self):
        prt_path = r'E:\demo\GPU_2SW_BOARD\packaged\pstxprt.dat'
        net_path = r'E:\demo\GPU_2SW_BOARD\packaged\pstxnet.dat'
        project_root = pstx_analyzer._infer_project_root_from_data_paths(prt_path, net_path)
        self.assertEqual(str(PureWindowsPath(r'E:\demo\GPU_2SW_BOARD')), project_root)

    def test_resolve_component_pages_uses_global_page_number_mapping_without_cell_filter(self):
        components = {
            'U1': {
                'refdes': 'U1',
                'drawing': '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1',
                'page': 'PAGE242',
            },
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            unrelated_sch = root / 'UNRELATED_BLOCK' / 'sch_1'
            unrelated_sch.mkdir(parents=True)
            (unrelated_sch / 'page900.csv').write_text('NAME,PAGE_NUMBER\nX,242\n', encoding='utf-8')

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        self.assertEqual('PAGE900', components['U1']['page'])
        self.assertEqual('PAGE900', components['U1']['page_real'])
        self.assertEqual('drawing', components['U1']['page_source'])
        self.assertEqual('page_csv', components['U1']['page_real_source'])
        self.assertEqual('是', components['U1']['page_mapping_ok'])

    def test_resolve_component_pages_leaves_real_page_blank_when_logical_page_hits_multiple_real_pages(self):
        components = {
            'U1': {
                'refdes': 'U1',
                'drawing': '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1',
                'page': 'PAGE242',
            },
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_sch = root / 'BLOCK_A' / 'sch_1'
            first_sch.mkdir(parents=True)
            (first_sch / 'page518.csv').write_text('NAME,PAGE_NUMBER\nA,242\n', encoding='utf-8')

            second_sch = root / 'BLOCK_B' / 'sch_1'
            second_sch.mkdir(parents=True)
            (second_sch / 'page900.csv').write_text('NAME,PAGE_NUMBER\nB,242\n', encoding='utf-8')

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertTrue(any('PAGE242 同时命中多个真实页' in warning for warning in warnings))
        self.assertEqual('', components['U1']['page'])
        self.assertEqual('', components['U1']['page_real'])
        self.assertEqual('page_csv_ambiguous', components['U1']['page_real_source'])
        self.assertEqual('否', components['U1']['page_mapping_ok'])

    def test_build_page_csv_index_reports_scanned_but_unparsed_csv_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            direct_sch = root / 'sch_1'
            direct_sch.mkdir(parents=True)
            (direct_sch / 'page518.csv').write_text('NAME,VALUE\nTOP,242\n', encoding='utf-8')

            index = pstx_analyzer._build_page_csv_index(str(root))

        self.assertEqual(1, index['scanned'])
        self.assertEqual(1, index['matched_root_sch1'])
        self.assertEqual(0, index['count'])
        self.assertTrue(any('没有读出任何 PAGE_NUMBER' in warning for warning in index['warnings']))

    def test_read_page_number_from_csv_prefers_exact_assignment_format_with_semicolon(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / 'page518.csv'
            csv_path.write_text('"PAGE_NUMBER" = 242;\nNAME = TOP;\n', encoding='utf-8')

            page_number = pstx_analyzer._read_page_number_from_csv(csv_path)

        self.assertEqual('PAGE242', page_number)

    def test_read_page_number_from_utf16_csv_with_assignment_format(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / 'page518.csv'
            csv_path.write_text('"PAGE_NUMBER" = "242";\n"NAME" = "TOP";\n', encoding='utf-16')

            page_number = pstx_analyzer._read_page_number_from_csv(csv_path)

        self.assertEqual('PAGE242', page_number)

    def test_build_page_csv_index_reads_utf16_assignment_page_csv(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            direct_sch = root / 'sch_1'
            direct_sch.mkdir(parents=True)
            (direct_sch / 'page518.csv').write_text(
                '"PAGE_NUMBER" = 242;\n"NAME" = "TOP";\n',
                encoding='utf-16',
            )

            index = pstx_analyzer._build_page_csv_index(str(root))

        self.assertEqual(1, index['count'])
        self.assertIn('PAGE242', index['by_logical_page'])
        self.assertEqual('PAGE518', index['by_logical_page']['PAGE242'][0]['resolved_page'])

    def test_analyze_page_mappings_marks_reverse_conflict_as_not_one_to_one(self):
        page_index = {
            'by_logical_page': {
                'PAGE242': [{'resolved_page': 'PAGE518', 'cell': 'TOP', 'path': 'a/page518.csv'}],
                'PAGE300': [{'resolved_page': 'PAGE518', 'cell': 'TOP', 'path': 'b/page518.csv'}],
            }
        }
        result = pstx_analyzer.analyze_page_mappings(page_index)
        row_map = {row['逻辑页']: row for row in result['rows']}
        self.assertEqual('否', row_map['PAGE242']['是否一一对应'])
        self.assertEqual('真实页对应多个逻辑页', row_map['PAGE242']['状态'])
        self.assertTrue(any('PAGE518 同时被多个逻辑页复用' in warning for warning in result['warnings']))

    def test_parse_pstxnet_handles_marker_at_file_start(self):
        content = (
            "NET_NAME\n"
            "'P1V8'\n"
            "NODE_NAME C1 1\n"
            "'POS':\n"
        )
        nets = pstx_analyzer.parse_pstxnet(content)
        self.assertIn('P1V8', nets)
        self.assertEqual('POS', nets['P1V8'][0]['pin_name'])

    def test_parse_pstxnet_finds_pin_name_beyond_fixed_window(self):
        content = (
            "\nNET_NAME\n"
            "'NET1'\n"
            "NODE_NAME U1 1\n"
            f"{'X' * 220}'GPIO1':\n"
        )
        nets = pstx_analyzer.parse_pstxnet(content)
        self.assertEqual('GPIO1', nets['NET1'][0]['pin_name'])


class PageModelV3Tests(unittest.TestCase):
    def _sample_part_block(self):
        return (
            "PART_NAME\n"
            "C1A104 'CAP_HDL-HQ17101005HS0,100NF,10%,0402,X7R,50V':\n"
            "SECTION_NUMBER 1\n"
            " '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE144_I70"
            "@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE1_I17"
            "@HQ_CAP.CAP_HDL(CHIPS)':\n"
            " C_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70"
            "@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page1_i17"
            "@hq_cap.cap_hdl(chips)',\n"
            " P_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page114_i70"
            "@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page1_i17"
            "@hq_cap.cap_hdl(chips)',\n"
            " PATH='I17',\n"
            " DRAWING='@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE1',\n"
            " PHYS_PAGE='1',\n"
            " BOM_OPTION='DEPOP',\n"
            " PACKAGE='0402',\n"
            " HQ_CODE='HQ17101005HS0',\n"
            " VALUE='100NF'\n"
        )

    def _deep_hierarchy_part_block(self):
        return (
            "PART_NAME\n"
            "C9A001 'CAP_HDL-HQ99999999,10NF,10%,0402,X7R,16V':\n"
            "SECTION_NUMBER 1\n"
            " '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE144_I70"
            "@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE3_I17"
            "@GPU_2SW_BOARD_LIB.GRAND_CHILD_BLOCK(SCH_1):PAGE2_I5"
            "@HQ_CAP.CAP_HDL(CHIPS)':\n"
            " C_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70"
            "@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page3_i17"
            "@gpu_2sw_board_lib.grand_child_block(sch_1):page2_i5"
            "@hq_cap.cap_hdl(chips)',\n"
            " P_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page114_i70"
            "@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page3_i17"
            "@gpu_2sw_board_lib.grand_child_block(sch_1):page2_i5"
            "@hq_cap.cap_hdl(chips)',\n"
            " DRAWING='@GPU_2SW_BOARD_LIB.GRAND_CHILD_BLOCK(SCH_1):PAGE2',\n"
            " HQ_CODE='HQ99999999',\n"
            " VALUE='10NF'\n"
        )

    def _pex90144_part_block(self):
        return (
            "PART_NAME\n"
            "C1A101 'CAP_HDL-HQ171010060D0,220NF,10%,0201,X6S,6.3V':\n"
            "REUSE_INSTANCE='PEX90144_CBB_V1A101',\n"
            "REUSE_NAME='PEX90144_CBB_V1',\n"
            "REUSE_PID='906';\n"
            "SECTION_NUMBER 1\n"
            " '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE112_I167"
            "@GPU_2SW_BOARD_LIB.PEX90144_CBB_V1(SCH_1):PAGE1_I155"
            "@HQ_CAP.CAP_HDL(CHIPS)':\n"
            " C_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page112_i167"
            "@gpu_2sw_board_lib.pex90144_cbb_v1(sch_1):page1_i155"
            "@hq_cap.cap_hdl(chips)',\n"
            " P_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page24_i167"
            "@gpu_2sw_board_lib.pex90144_cbb_v1(sch_1):page1_i155"
            "@hq_cap.cap_hdl(chips)',\n"
            " PATH='I155',\n"
            " DRAWING='@GPU_2SW_BOARD_LIB.PEX90144_CBB_V1(SCH_1):PAGE1',\n"
            " PHYS_PAGE='1',\n"
            " CDS_LIB='hq_cap',\n"
            " CDS_PART_NAME='CAP_HDL-HQ171010060D0,220NF,10%,0201,X6S,6.3V',\n"
            " TOLERANCE='10%',\n"
            " PACKAGE='0201',\n"
            " MATERIAL='X6S',\n"
            " HQ_CODE='HQ171010060D0',\n"
            " VOLTAGE='6.3V',\n"
            " VALUE='220NF',\n"
            " REUSE_PID='906',\n"
            " SUBDESIGN_SUFFIX='101',\n"
            " SUBDESIGN_NAME='PEX90144_CBB_V1',\n"
            " REUSE_INSTANCE='PEX90144_CBB_V1A101',\n"
            " REUSE_NAME='PEX90144_CBB_V1';\n"
        )

    def test_parse_pstxprt_keeps_c_path_for_logical_and_p_path_for_real(self):
        components = pstx_analyzer.parse_pstxprt(self._sample_part_block())
        comp = components['C1A104']
        self.assertEqual('section_path', comp['page_path_logical_source'])
        self.assertEqual('p_path', comp['page_path_real_source'])
        self.assertEqual('PAGE144', comp['page_logical'])
        self.assertEqual('PAGE1', comp['page_submodule_real'])
        self.assertTrue(comp['page_path_real_raw'].startswith('@gpu_2sw_board_lib.gpu_2sw_board'))

    def test_parse_page_map_line_reads_logical_then_real_then_name(self):
        parsed = pstx_page_logic._parse_page_map_line('144 114 TOP')
        self.assertEqual(
            {
                'logical_page': 'PAGE144',
                'real_page': 'PAGE114',
                'page_name': 'TOP',
            },
            parsed,
        )

    def test_parse_page_map_line_keeps_full_name_segment_after_real_page(self):
        parsed = pstx_page_logic._parse_page_map_line('144   114   TOP MAIN BLOCK')
        self.assertEqual('PAGE144', parsed['logical_page'])
        self.assertEqual('PAGE114', parsed['real_page'])
        self.assertEqual('TOP MAIN BLOCK', parsed['page_name'])

    def test_build_page_map_index_reads_name_segment_with_spaces(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page.map').write_text('144 114 TOP MAIN BLOCK\n', encoding='utf-8')

            index = pstx_page_logic.build_page_map_index(str(root))

        entries = index['by_logical_page']['PAGE144']
        self.assertEqual(1, len(entries))
        self.assertEqual('PAGE114', entries[0]['resolved_page'])
        self.assertEqual('TOP MAIN BLOCK', entries[0]['page_name'])

    def test_resolve_component_pages_prefers_p_path_and_computes_mapped_submodule_page(self):
        components = pstx_analyzer.parse_pstxprt(self._sample_part_block())
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page114.csv').write_text('"PAGE_NUMBER" = 144;\n', encoding='utf-8')
            (sch_dir / 'page.map').write_text('144 114 TOP\n', encoding='utf-8')
            (root / 'module_order').write_text(
                'Version 15.0\n'
                'START_MODULEORDER\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70'
                '@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1) 0 1 177 34 0\n'
                'END_MODULEORDER\n',
                encoding='utf-8',
            )

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        comp = components['C1A104']
        self.assertEqual('PAGE144', comp['page_logical'])
        self.assertEqual('PAGE114', comp['page_real'])
        self.assertEqual('PAGE114', comp['page'])
        self.assertEqual('PAGE1', comp['page_submodule_real'])
        self.assertEqual('PAGE177', comp['page_submodule_mapped'])
        self.assertEqual('p_path', comp['page_real_source'])
        self.assertEqual('是', comp['page_mapping_ok'])

    def test_check_drc_bom_option_components_show_real_and_mapped_pages(self):
        components = pstx_analyzer.parse_pstxprt(self._sample_part_block())
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page114.csv').write_text('"PAGE_NUMBER" = 144;\n', encoding='utf-8')
            (sch_dir / 'page.map').write_text('144 114 TOP\n', encoding='utf-8')
            (root / 'module_order').write_text(
                'Version 15.0\n'
                'START_MODULEORDER\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70'
                '@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1) 0 1 177 34 0\n'
                'END_MODULEORDER\n',
                encoding='utf-8',
            )
            pstx_analyzer.resolve_component_pages(components, str(root))

        result = pstx_analyzer.check_drc(components, {}, option_components_source=components)
        row = result['bom_option_components'][0]
        self.assertEqual('PAGE114', row['页面'])
        self.assertEqual('PAGE177', row['子模块映射主模块真实页'])

    def test_page_map_cross_check_prefers_root_sch1_over_child_sch1(self):
        components = pstx_analyzer.parse_pstxprt(
            "PART_NAME\n"
            "U1 'IC_CPU':\n"
            " C_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70@hq_ic.cpu(chips)',\n"
            " DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE144'\n"
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / 'sch_1').mkdir(parents=True)
            (root / 'sch_1' / 'page.map').write_text('144 114 ROOT_TOP\n', encoding='utf-8')
            (root / 'sch_1' / 'page114.csv').write_text('"PAGE_NUMBER" = 144;\n', encoding='utf-8')
            child_sch = root / 'reuse_block' / 'sch_1'
            child_sch.mkdir(parents=True)
            (child_sch / 'page.map').write_text('144 999 CHILD_SHOULD_NOT_WIN\n', encoding='utf-8')

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        comp = components['U1']
        self.assertEqual('PAGE144', comp['page_logical'])
        self.assertEqual('PAGE114', comp['page_real'])
        self.assertEqual('page_map', comp['page_real_source'])
        self.assertEqual('unique', comp['page_map_state'])
        self.assertEqual('是', comp['page_mapping_ok'])

    def test_module_order_prefers_logical_path_key_when_p_path_exists(self):
        components = pstx_analyzer.parse_pstxprt(self._sample_part_block())
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page.map').write_text('144 114 TOP\n', encoding='utf-8')
            (sch_dir / 'page114.csv').write_text('"PAGE_NUMBER" = 144;\n', encoding='utf-8')
            (root / 'module_order').write_text(
                'Version 15.0\n'
                'START_MODULEORDER\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70'
                '@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1) 0 1 177 34 0\n'
                'END_MODULEORDER\n',
                encoding='utf-8',
            )

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        comp = components['C1A104']
        self.assertEqual('PAGE114', comp['page_real'])
        self.assertEqual('PAGE177', comp['page_submodule_mapped'])
        self.assertEqual('unique', comp['module_order_state'])
        self.assertIn('@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE144_I70', comp['module_order_key'])

    def test_module_order_mapping_rejects_submodule_page_out_of_range(self):
        part_block = self._sample_part_block().replace('page1_i17', 'page35_i17').replace('PAGE1_I17', 'PAGE35_I17')
        components = pstx_analyzer.parse_pstxprt(part_block)
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page.map').write_text('144 114 TOP\n', encoding='utf-8')
            (sch_dir / 'page114.csv').write_text('"PAGE_NUMBER" = 144;\n', encoding='utf-8')
            (root / 'module_order').write_text(
                'Version 15.0\n'
                'START_MODULEORDER\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70'
                '@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1) 0 1 177 34 0\n'
                'END_MODULEORDER\n',
                encoding='utf-8',
            )

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        comp = components['C1A104']
        self.assertEqual('PAGE35', comp['page_submodule_real'])
        self.assertEqual('', comp['page_submodule_mapped'])
        self.assertEqual('local_page_out_of_range', comp['module_order_state'])
        self.assertIn('超出 module_order 页数 34', comp['page_submodule_mapping_note'])


    def test_resolve_component_pages_maps_deepest_module_order_for_nested_reuse(self):
        components = pstx_analyzer.parse_pstxprt(self._deep_hierarchy_part_block())
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page114.csv').write_text('"PAGE_NUMBER" = 144;\n', encoding='utf-8')
            (sch_dir / 'page.map').write_text('144 114 TOP\n', encoding='utf-8')
            (root / 'module_order').write_text(
                'Version 15.0\n'
                'START_MODULEORDER\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70'
                '@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1) 0 1 177 34 0\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70'
                '@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page3_i17'
                '@gpu_2sw_board_lib.grand_child_block(sch_1) 0 1 250 10 0\n'
                'END_MODULEORDER\n',
                encoding='utf-8',
            )

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        comp = components['C9A001']
        self.assertEqual('PAGE114', comp['page_real'])
        self.assertEqual('PAGE2', comp['page_submodule_real'])
        self.assertEqual('PAGE251', comp['page_submodule_mapped'])
        self.assertIn('@GPU_2SW_BOARD_LIB.GRAND_CHILD_BLOCK(SCH_1)', comp['module_order_key'])
        self.assertEqual('PAGE2', comp['module_order_local_page'])

    def test_resolve_component_pages_reads_module_order_dat_and_maps_pex90144_sample(self):
        components = pstx_analyzer.parse_pstxprt(self._pex90144_part_block())
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page24.csv').write_text('"PAGE_NUMBER" = 112;\n', encoding='utf-8')
            (sch_dir / 'page.map').write_text('112 24 TOP\n', encoding='utf-8')
            (root / 'module_order.dat').write_text(
                'Version 15.0\n'
                'START_MODULEORDER\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1) 0 1 1 176 0\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page112_i167'
                '@gpu_2sw_board_lib.pex90144_cbb_v1(sch_1) 0 1 177 34 0\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page4_i1'
                '@gpu_2sw_board_lib.pex90144_cbb_v1(sch_1) 0 1 211 34 1\n'
                '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page239_i64'
                '@gpu_2sw_board_lib.i2c_sw_tpt29548_000_cbb_v3(sch_1) 0 1 245 1 0\n'
                'END_MODULEORDER\n',
                encoding='utf-8',
            )

            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        comp = components['C1A101']
        self.assertEqual('PAGE112', comp['page_logical'])
        self.assertEqual('PAGE24', comp['page_real'])
        self.assertEqual('PAGE1', comp['page_submodule_real'])
        self.assertEqual('PAGE177', comp['page_submodule_mapped'])
        self.assertEqual('p_path', comp['page_real_source'])
        self.assertEqual('unique', comp['module_order_state'])
        self.assertIn('@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE112_I167', comp['module_order_key'])

    def test_module_order_deduplicates_identical_dat_and_extensionless_files(self):
        components = pstx_analyzer.parse_pstxprt(self._pex90144_part_block())
        module_order_text = (
            'Version 15.0\n'
            'START_MODULEORDER\n'
            '@gpu_2sw_board_lib.gpu_2sw_board(sch_1) 0 1 1 176 0\n'
            '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page112_i167'
            '@gpu_2sw_board_lib.pex90144_cbb_v1(sch_1) 0 1 177 34 0\n'
            'END_MODULEORDER\n'
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sch_dir = root / 'sch_1'
            sch_dir.mkdir(parents=True)
            (sch_dir / 'page24.csv').write_text('"PAGE_NUMBER" = 112;\n', encoding='utf-8')
            (sch_dir / 'page.map').write_text('112 24 TOP\n', encoding='utf-8')
            (root / 'module_order.dat').write_text(module_order_text, encoding='utf-8')
            (root / 'module_order').write_text(module_order_text, encoding='utf-8')

            index = pstx_page_logic.build_module_order_index(str(root))
            warnings = pstx_analyzer.resolve_component_pages(components, str(root))

        self.assertEqual([], warnings)
        self.assertEqual(2, index['count'])
        self.assertEqual(2, index['duplicate_count'])
        comp = components['C1A101']
        self.assertEqual('PAGE177', comp['page_submodule_mapped'])
        self.assertEqual('unique', comp['module_order_state'])


class BomTests(unittest.TestCase):
    def test_build_bom_does_not_merge_distinct_items_without_part_number(self):
        components = {
            'R1': {
                'refdes': 'R1', 'hq_code': '', 'part_name': 'RES_0402', 'value': '10k',
                'package': '0402', 'voltage': '', 'power': '', 'tolerance': '1%',
                'material': 'thick', 'comp_type': 'RES', 'page': 'PAGE1', 'room': '',
                'bom_option': '',
            },
            'R2': {
                'refdes': 'R2', 'hq_code': '', 'part_name': 'RES_0402', 'value': '4.7k',
                'package': '0402', 'voltage': '', 'power': '', 'tolerance': '1%',
                'material': 'thick', 'comp_type': 'RES', 'page': 'PAGE1', 'room': '',
                'bom_option': '',
            },
        }
        _, _, merged, _ = pstx_analyzer.build_bom(components)
        self.assertEqual(2, len(merged))
        self.assertEqual({'10k', '4.7k'}, {row['值'] for row in merged})

    def test_build_bom_treats_dnp_as_depop(self):
        components = {
            'C1': {
                'refdes': 'C1', 'hq_code': '', 'part_name': 'CAP_0402', 'value': '1uF',
                'package': '0402', 'voltage': '6.3V', 'power': '', 'tolerance': '',
                'material': '', 'comp_type': 'CAP', 'page': 'PAGE1', 'room': '',
                'bom_option': 'DNP',
            },
        }
        detail_normal, detail_depop, _, merged_depop = pstx_analyzer.build_bom(components)
        self.assertEqual([], detail_normal)
        self.assertEqual(1, len(detail_depop))
        self.assertEqual(1, len(merged_depop))

    def test_build_bom_prefers_real_page_display(self):
        components = {
            'R1': make_res('R1', 'P3V3', 'GPIO1', page='PAGE242', page_real='PAGE518'),
        }
        detail_normal, _, _, _ = pstx_analyzer.build_bom(components)
        self.assertEqual('PAGE518', detail_normal[0]['页面'])


class RuleTests(unittest.TestCase):
    def test_build_analysis_scope_excludes_depop_components_by_default(self):
        components = {
            'U1': make_ic('U1', {'1': 'GPIO1'}),
            'R1': make_res('R1', 'P3V3', 'GPIO1', bom_option='DEPOP'),
            'R2': make_res('R2', 'P3V3', 'GPIO1'),
        }
        nets = {
            'GPIO1': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'GPIO1'},
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [
                {'refdes': 'R1', 'pin': '1', 'pin_name': '1'},
                {'refdes': 'R2', 'pin': '1', 'pin_name': '1'},
            ],
        }
        active_components, active_nets, depop_refdes, excluded_refdes = pstx_analyzer._build_analysis_scope(
            components,
            nets,
            include_depop=False,
        )
        self.assertEqual(['R1'], depop_refdes)
        self.assertEqual(['R1'], excluded_refdes)
        self.assertNotIn('R1', active_components)
        self.assertEqual({'U1', 'R2'}, {node['refdes'] for node in active_nets['GPIO1']})

    def test_check_drc_keeps_bom_option_component_list_from_raw_source(self):
        analysis_components = {'U1': make_ic('U1')}
        raw_components = {
            'U1': make_ic('U1'),
            'R1': make_res('R1', 'P3V3', 'GPIO1', bom_option='DEPOP', page_real='PAGE518'),
        }
        result = pstx_analyzer.check_drc(analysis_components, {}, option_components_source=raw_components)
        row_map = {row['位号']: row for row in result['bom_option_components']}
        self.assertIn('R1', row_map)
        self.assertEqual('是', row_map['R1']['是否DEPOP'])
        self.assertEqual('PAGE518', row_map['R1']['页面'])

    def test_check_drc_bom_option_components_use_top_level_logical_page_mapping(self):
        components = {
            'R1': {
                'refdes': 'R1',
                'part_name': 'RES_0402',
                'hq_code': '',
                'value': '10k',
                'package': '0402',
                'material': '',
                'tolerance': '',
                'voltage': '',
                'current': '',
                'power': '',
                'bom_option': 'DEPOP',
                'bom_cost': '',
                'room': '',
                'drawing': (
                    '@GPU_2SW_BOARD_LIB.HQPWR_EFUSE_TPS259260_12VIN_4A(SCH_1):PAGE1_I14'
                    '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'
                ),
                'page': 'PAGE242 / PAGE1',
                'comp_type': 'RES',
                'nets': {'1': 'P3V3', '2': 'GPIO1'},
            },
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / 'GPU_2SW_BOARD'
            root.mkdir(parents=True)
            top_sch = root / 'sch_1'
            top_sch.mkdir(parents=True)
            (top_sch / 'page518.csv').write_text('"PAGE_NUMBER" = 242;', encoding='utf-8')
            (top_sch / 'page12.csv').write_text('"PAGE_NUMBER" = 1;', encoding='utf-8')
            child_sch = root / 'HQPWR_EFUSE_TPS259260_12VIN_4A' / 'sch_1'
            child_sch.mkdir(parents=True)
            (child_sch / 'page12.csv').write_text('"PAGE_NUMBER" = 1;', encoding='utf-8')
            pstx_analyzer.resolve_component_pages(components, str(root))
        rows = pstx_analyzer.check_drc({}, {}, option_components_source=components)['bom_option_components']
        self.assertEqual('PAGE518', rows[0]['页面'])

    def test_extract_refdes_suffix_group_prefers_trailing_letter_digit_cluster(self):
        self.assertEqual('A1', pstx_analyzer._extract_refdes_suffix_group('PU1A1'))
        self.assertEqual('A1', pstx_analyzer._extract_refdes_suffix_group('R1A1'))
        self.assertEqual('', pstx_analyzer._extract_refdes_suffix_group('U1'))

    def test_extract_pin_submodule_info_uses_parent_hierarchy_before_leaf_symbol(self):
        pin_name = (
            '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'
            '@GPU_2SW_BOARD_LIB.HQPWR_EFUSE_TPS259260_12VIN_4A(SCH_1):PAGE1_I14'
            '@HQ_IC.TPS259261DRCR_11P_HDL(CHIPS)'
        )
        submodule, submodule_path = pstx_analyzer._extract_pin_submodule_info(pin_name)
        self.assertEqual('HQPWR_EFUSE_TPS259260_12VIN_4A', submodule)
        self.assertEqual('GPU_2SW_BOARD / HQPWR_EFUSE_TPS259260_12VIN_4A', submodule_path)

    def test_parse_ohms_supports_embedded_notation(self):
        self.assertEqual(4.7, pstx_analyzer._parse_ohms('4R7'))
        self.assertEqual(1500.0, pstx_analyzer._parse_ohms('1K5'))

    def test_parse_ohms_supports_ohm_word_suffixes(self):
        self.assertEqual(10, pstx_analyzer._parse_ohms('10OHM'))
        self.assertEqual(10, pstx_analyzer._parse_ohms('10OHMS'))
        self.assertEqual(10000, pstx_analyzer._parse_ohms('10KOHM'))
        self.assertEqual(4700, pstx_analyzer._parse_ohms('4.7KΩ'))

    def test_derating_does_not_infer_signal_like_pg_p1v8(self):
        components = {'C1': make_cap()}
        nets = {
            'PG_P1V8': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        rows = pstx_analyzer.analyze_derating(components, nets)
        self.assertIn('无法推断', rows[0]['状态'])
        self.assertEqual('', rows[0]['推断工作电压(V)'])

    def test_derating_requires_ground_and_single_known_positive_rail(self):
        components = {'C1': make_cap(rated='16V')}
        nets = {
            'P5V': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'P3V3': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        rows = pstx_analyzer.analyze_derating(components, nets)
        self.assertIn('未连接地', rows[0]['状态'])

    def test_custom_voltage_map_matches_prefix_boundary_only(self):
        components = {'C1': make_cap()}
        nets = {
            'PG_P1V8': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        rows = pstx_analyzer.analyze_derating(components, nets, custom_volt_map={'P1V8': 1.8})
        self.assertIn('无法推断', rows[0]['状态'])

    def test_exact_custom_voltage_map_can_override_signal_net_when_user_declares_it(self):
        components = {'C1': make_cap()}
        nets = {
            'PG_P1V8': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        rows = pstx_analyzer.analyze_derating(components, nets, custom_volt_map={'PG_P1V8': 1.8})
        self.assertEqual('1.8', rows[0]['推断工作电压(V)'])

    def test_derating_token_inference_is_candidate_not_confirmed(self):
        components = {'C1': make_cap()}
        nets = {
            'P1V8_AON': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_derating(components, nets)[0]
        self.assertEqual('候选判断', row['结论类型'])
        self.assertEqual('网络首 token', row['推断来源类型'])
        self.assertEqual('single_positive_rail_token', row['原因代码'])

    def test_derating_custom_map_is_confirmed(self):
        components = {'C1': make_cap()}
        nets = {
            'VDD_SENSE': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_derating(components, nets, custom_volt_map={'VDD_SENSE': 1.2})[0]
        self.assertEqual('确定结论', row['结论类型'])
        self.assertEqual('自定义映射', row['推断来源类型'])
        self.assertEqual('custom_voltage_map', row['原因代码'])

    def test_derating_passes_high_rated_caps_when_global_max_voltage_is_not_above_12v(self):
        components = {'C1': make_cap(rated='50V')}
        nets = {
            'SIG_ALERT': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
            'P12V_AUX': [{'refdes': 'U1', 'pin': '1', 'pin_name': 'VIN'}],
        }
        row = pstx_analyzer.analyze_derating(components, nets)[0]
        self.assertTrue(row['状态'].startswith('✅'))
        self.assertEqual('12.0', row['推断工作电压(V)'])
        self.assertEqual('P12V_AUX', row['推断来源网络'])
        self.assertEqual('global_max_voltage_under_12v_high_rated_cap', row['原因代码'])

    def test_derating_does_not_apply_50v_override_when_global_max_voltage_exceeds_12v(self):
        components = {'C1': make_cap(rated='50V')}
        nets = {
            'SIG_ALERT': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
            'P20V_SYS': [{'refdes': 'U1', 'pin': '1', 'pin_name': 'VIN'}],
        }
        row = pstx_analyzer.analyze_derating(components, nets)[0]
        self.assertIn('无法推断', row['状态'])
        self.assertEqual('no_positive_voltage_evidence', row['原因代码'])

    def test_derating_marks_mirrored_diff_caps_as_ac_coupling_candidate(self):
        components = {
            'C1': make_cap(refdes='C1'),
            'C2': make_cap(refdes='C2'),
        }
        nets = {
            'PCIE_TXA_P': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'PCIE_TXB_P': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
            'PCIE_TXA_N': [{'refdes': 'C2', 'pin': '1', 'pin_name': '1'}],
            'PCIE_TXB_N': [{'refdes': 'C2', 'pin': '2', 'pin_name': '2'}],
        }
        rows = {row['位号']: row for row in pstx_analyzer.analyze_derating(components, nets)}
        self.assertIn('疑似 AC 耦合', rows['C1']['状态'])
        self.assertEqual('无法判断', rows['C1']['结论类型'])
        self.assertEqual('AC 耦合候选', rows['C1']['推断来源类型'])
        self.assertEqual('ac_coupling_candidate', rows['C1']['原因代码'])
        self.assertIn('疑似 AC 耦合', rows['C2']['状态'])

    def test_derating_does_not_treat_lone_negative_suffix_as_ac_coupling(self):
        components = {'C1': make_cap()}
        nets = {
            'PCIE_TXA_N': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'PCIE_TXB_N': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_derating(components, nets)[0]
        self.assertEqual('no_ground_reference', row['原因代码'])
        self.assertNotIn('AC 耦合', row['状态'])

    def test_analyze_derating_prefers_real_page_for_page_column(self):
        components = {'C1': make_cap(page='PAGE242', page_real='PAGE518')}
        nets = {
            'P1V8_AON': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_derating(components, nets)[0]
        self.assertEqual('PAGE518', row['页面'])

    def test_analyze_derating_does_not_fallback_to_logical_page_when_real_page_missing(self):
        components = {'C1': make_cap(page='PAGE242', page_real='')}
        nets = {
            'P1V8_AON': [{'refdes': 'C1', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'C1', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_derating(components, nets)[0]
        self.assertEqual('', row['页面'])

    def test_analyze_resistors_uses_conservative_power_detection(self):
        components = {
            'R1': {
                'refdes': 'R1',
                'part_name': 'RES_0402',
                'hq_code': '',
                'value': '10k',
                'package': '0402',
                'material': '',
                'tolerance': '',
                'voltage': '',
                'current': '',
                'power': '',
                'bom_option': '',
                'bom_cost': '',
                'room': '',
                'drawing': 'SCH_PAGE1',
                'page': 'PAGE1',
                'comp_type': 'RES',
                'nets': {'1': 'DDR_VDDQ_EN', '2': 'ALERT_N'},
            },
        }
        nets = {
            'DDR_VDDQ_EN': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}],
            'ALERT_N': [{'refdes': 'R1', 'pin': '2', 'pin_name': '2'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets)
        self.assertEqual({}, result['pullups'])

    def test_analyze_resistors_duplicate_pullup_is_candidate(self):
        components = {
            'R1': {'refdes': 'R1', 'part_name': 'RES_0402', 'hq_code': '', 'value': '10k', 'package': '0402',
                   'material': '', 'tolerance': '', 'voltage': '', 'current': '', 'power': '',
                   'bom_option': '', 'bom_cost': '', 'room': '', 'drawing': '', 'page': 'PAGE1',
                   'comp_type': 'RES', 'nets': {'1': 'P3V3', '2': 'ALERT_N'}},
            'R2': {'refdes': 'R2', 'part_name': 'RES_0402', 'hq_code': '', 'value': '4.7k', 'package': '0402',
                   'material': '', 'tolerance': '', 'voltage': '', 'current': '', 'power': '',
                   'bom_option': '', 'bom_cost': '', 'room': '', 'drawing': '', 'page': 'PAGE2',
                   'comp_type': 'RES', 'nets': {'1': 'P3V3', '2': 'ALERT_N'}},
        }
        nets = {
            'P3V3': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}, {'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
            'ALERT_N': [{'refdes': 'R1', 'pin': '2', 'pin_name': '2'}, {'refdes': 'R2', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_resistors(components, nets)['dup_pullups'][0]
        self.assertEqual('候选判断', row['结论类型'])
        self.assertEqual('multiple_pullup_paths', row['原因代码'])
        self.assertNotIn('单板场景', row)

    def test_analyze_resistors_duplicate_pullup_keeps_bom_options_without_single_board_inference(self):
        components = {
            'R1': make_res('R1', 'P3V3', 'ALERT_N', bom_option='MAIN'),
            'R2': make_res('R2', 'P3V3', 'ALERT_N', bom_option='ALT'),
        }
        nets = {
            'P3V3': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}, {'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
            'ALERT_N': [{'refdes': 'R1', 'pin': '2', 'pin_name': '2'}, {'refdes': 'R2', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_resistors(components, nets)['dup_pullups'][0]
        self.assertIn('MAIN', row['BOM_OPTION'])
        self.assertIn('ALT', row['BOM_OPTION'])
        self.assertEqual(2, row['上拉数量'])

    def test_analyze_resistors_finds_series_bias_on_both_sides(self):
        components = {
            'R1': make_res('R1', 'NET_A', 'NET_B', value='100R'),
            'R2': make_res('R2', 'P3V3', 'NET_A', value='1k'),
            'R3': make_res('R3', 'NET_B', 'GND', value='2k'),
        }
        nets = {
            'NET_A': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}, {'refdes': 'R2', 'pin': '2', 'pin_name': '2'}],
            'NET_B': [{'refdes': 'R1', 'pin': '2', 'pin_name': '2'}, {'refdes': 'R3', 'pin': '1', 'pin_name': '1'}],
            'P3V3': [{'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'R3', 'pin': '2', 'pin_name': '2'}],
        }
        rows = pstx_analyzer.analyze_resistors(components, nets)['divider_risks']
        self.assertEqual(2, len(rows))
        row_map = {(row['偏置类型'], row['受影响网络']): row for row in rows}
        self.assertIn(('上拉', 'NET_B'), row_map)
        self.assertIn(('下拉', 'NET_A'), row_map)
        self.assertEqual('NET_A', row_map[('上拉', 'NET_B')]['偏置所在网络'])
        self.assertEqual('GND', row_map[('下拉', 'NET_A')]['偏置参考网络'])

    def test_analyze_resistors_reports_chip_pin_status_for_u_pu_xu(self):
        components = {
            'U1': make_ic('U1', {'1': 'GPIO1'}, page='PAGE242', page_real='PAGE518'),
            'PU2A1': make_ic('PU2A1', {'A1': 'GPIO1'}, page='PAGE242', page_real='PAGE518'),
            'XU3': make_ic('XU3', {'B2': 'NET_SER'}, page='PAGE300', page_real='PAGE612'),
            'R1': make_res('R1', 'GPIO1', 'NET_SER', value='22R', page='PAGE242', page_real='PAGE518'),
            'R2': make_res('R2', 'P3V3', 'GPIO1', value='10k', page='PAGE242', page_real='PAGE518'),
            'R3': make_res('R3', 'GPIO1', 'GND', value='100k', page='PAGE242', page_real='PAGE518'),
        }
        nets = {
            'GPIO1': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'GPIO1'},
                {
                    'refdes': 'PU2A1',
                    'pin': 'A1',
                    'pin_name': (
                        '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'
                        '@GPU_2SW_BOARD_LIB.HQPWR_EFUSE_TPS259260_12VIN_4A(SCH_1):PAGE1_I14'
                        '@HQ_IC.TPS259261DRCR_11P_HDL(CHIPS)'
                    ),
                },
                {'refdes': 'R1', 'pin': '1', 'pin_name': '1'},
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R3', 'pin': '1', 'pin_name': '1'},
            ],
            'NET_SER': [
                {'refdes': 'XU3', 'pin': 'B2', 'pin_name': 'DATA'},
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [{'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'R3', 'pin': '2', 'pin_name': '2'}],
        }
        rows = pstx_analyzer.analyze_resistors(components, nets)['chip_pin_rows']
        row_map = {(row['芯片位号'], row['引脚']): row for row in rows}
        self.assertEqual('是', row_map[('U1', '1')]['有串阻'])
        self.assertEqual('是', row_map[('U1', '1')]['有上拉'])
        self.assertEqual('是', row_map[('U1', '1')]['有下拉'])
        self.assertEqual('A1', row_map[('PU2A1', 'A1')]['后缀组'])
        self.assertEqual('HQPWR_EFUSE_TPS259260_12VIN_4A', row_map[('PU2A1', 'A1')]['子模块'])
        self.assertEqual(
            'GPU_2SW_BOARD / HQPWR_EFUSE_TPS259260_12VIN_4A',
            row_map[('PU2A1', 'A1')]['子模块路径'],
        )
        self.assertEqual('PAGE518', row_map[('PU2A1', 'A1')]['页面'])
        self.assertNotIn('逻辑页', row_map[('PU2A1', 'A1')])
        self.assertEqual('是', row_map[('XU3', 'B2')]['有串阻'])
        self.assertEqual('否', row_map[('XU3', 'B2')]['有上拉'])
        self.assertEqual(1, row_map[('XU3', 'B2')]['隔串阻上拉数量'])
        self.assertEqual('R2', row_map[('XU3', 'B2')]['隔串阻上拉位号'])
        self.assertEqual('P3V3', row_map[('XU3', 'B2')]['隔串阻上拉电源'])
        self.assertEqual('R1', row_map[('XU3', 'B2')]['隔串阻上拉串阻链'])
        self.assertEqual(1, row_map[('XU3', 'B2')]['隔串阻下拉数量'])
        self.assertEqual('R1', row_map[('XU3', 'B2')]['隔串阻下拉串阻链'])

    def test_ground_detection_treats_analog_ground_variants_as_ground(self):
        for net_name in ['AGND', 'AGND1', 'AGND_ADC', 'GNDA', 'VSSA', 'AVSS', '0V', '0']:
            self.assertTrue(pstx_analyzer._net_is_gnd(net_name), net_name)

    def test_analyze_resistors_treats_agnd_variants_as_ground_not_series_path(self):
        components = {
            'U1': make_ic('U1', {'1': 'SIG'}),
            'R1': make_res('R1', 'SIG', 'AGND1', value='22R'),
            'R2': make_res('R2', 'P3V3', 'AGND1', value='10k'),
        }
        nets = {
            'SIG': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'GPIO'},
                {'refdes': 'R1', 'pin': '1', 'pin_name': '1'},
            ],
            'AGND1': [
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [{'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
        }

        result = pstx_analyzer.analyze_resistors(components, nets)
        row = {
            (item['芯片位号'], item['引脚']): item
            for item in result['chip_pin_rows']
        }[('U1', '1')]

        self.assertEqual('否', row['有串阻'])
        self.assertEqual('是', row['有下拉'])
        self.assertEqual('否', row['有上拉'])
        self.assertEqual([], result['indirect_pullups'].get('SIG', []))
        self.assertEqual([], result['divider_risks'])

    def test_analyze_resistors_finds_bias_across_multiple_series_resistors(self):
        components = {
            'XU1': make_ic('XU1', {'B2': 'NET_IN'}, page='PAGE10', page_real='PAGE510'),
            'R1': make_res('R1', 'NET_IN', 'NET_MID', value='22R', page='PAGE10', page_real='PAGE510'),
            'R2': make_res('R2', 'NET_MID', 'NET_BIAS', value='33R', page='PAGE11', page_real='PAGE511'),
            'R3': make_res('R3', 'P3V3', 'NET_BIAS', value='10k', page='PAGE12', page_real='PAGE512'),
            'R4': make_res('R4', 'NET_BIAS', 'GND', value='100k', page='PAGE12', page_real='PAGE512'),
        }
        nets = {
            'NET_IN': [
                {'refdes': 'XU1', 'pin': 'B2', 'pin_name': 'GPIO_CHAIN'},
                {'refdes': 'R1', 'pin': '1', 'pin_name': '1'},
            ],
            'NET_MID': [
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R2', 'pin': '1', 'pin_name': '1'},
            ],
            'NET_BIAS': [
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R3', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R4', 'pin': '1', 'pin_name': '1'},
            ],
            'P3V3': [{'refdes': 'R3', 'pin': '1', 'pin_name': '1'}],
            'GND': [{'refdes': 'R4', 'pin': '2', 'pin_name': '2'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets)
        row = {
            (item['芯片位号'], item['引脚']): item
            for item in result['chip_pin_rows']
        }[('XU1', 'B2')]
        self.assertEqual(1, row['隔串阻上拉数量'])
        self.assertEqual('R3', row['隔串阻上拉位号'])
        self.assertEqual('NET_BIAS', row['隔串阻上拉来源网络'])
        self.assertEqual('P3V3', row['隔串阻上拉电源'])
        self.assertEqual('R1 -> R2', row['隔串阻上拉串阻链'])
        self.assertEqual(1, row['隔串阻下拉数量'])
        self.assertEqual('R4', row['隔串阻下拉位号'])
        self.assertEqual('R1 -> R2', row['隔串阻下拉串阻链'])

        net_in_rows = [item for item in result['divider_risks'] if item['受影响网络'] == 'NET_IN']
        self.assertEqual(2, len(net_in_rows))
        risk_map = {(item['偏置类型'], item['偏置位号']): item for item in net_in_rows}
        self.assertEqual('R1 -> R2', risk_map[('上拉', 'R3')]['串阻位号'])
        self.assertEqual('NET_IN -> NET_MID -> NET_BIAS', risk_map[('上拉', 'R3')]['串阻经过网络'])
        self.assertEqual(2, risk_map[('上拉', 'R3')]['串阻跳数'])
        self.assertEqual('P3V3', risk_map[('上拉', 'R3')]['偏置参考网络'])
        self.assertEqual('R1 -> R2', risk_map[('下拉', 'R4')]['串阻位号'])

    def test_analyze_resistors_keeps_parallel_series_paths_to_same_remote_pullup(self):
        components = {
            'U1': make_ic('U1', {'1': 'SIG'}),
            'R1': make_res('R1', 'SIG', 'MID_A', value='22R'),
            'R2': make_res('R2', 'MID_A', 'REMOTE', value='22R'),
            'R3': make_res('R3', 'SIG', 'MID_B', value='33R'),
            'R4': make_res('R4', 'MID_B', 'REMOTE', value='33R'),
            'R5': make_res('R5', 'P3V3', 'REMOTE', value='4.7k'),
        }
        nets = {
            'SIG': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'GPIO'},
                {'refdes': 'R1', 'pin': '1', 'pin_name': '1'},
                {'refdes': 'R3', 'pin': '1', 'pin_name': '1'},
            ],
            'MID_A': [
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R2', 'pin': '1', 'pin_name': '1'},
            ],
            'MID_B': [
                {'refdes': 'R3', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R4', 'pin': '1', 'pin_name': '1'},
            ],
            'REMOTE': [
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R4', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R5', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [{'refdes': 'R5', 'pin': '1', 'pin_name': '1'}],
        }

        result = pstx_analyzer.analyze_resistors(components, nets)

        sig_pullups = result['indirect_pullups']['SIG']
        self.assertEqual(
            {'R1 -> R2', 'R3 -> R4'},
            {item['via_refdes_chain'] for item in sig_pullups},
        )
        sig_dividers = [
            row for row in result['divider_risks']
            if row['受影响网络'] == 'SIG' and row['偏置位号'] == 'R5'
        ]
        self.assertEqual(2, len(sig_dividers))

    def test_analyze_resistors_od_oc_skips_missing_when_pullup_is_across_multiple_series_resistors(self):
        components = {
            'U1': make_ic('U1', {'1': 'SMBALERT_N'}),
            'R1': make_res('R1', 'SMBALERT_N', 'SMBALERT_BUF1', value='22R'),
            'R2': make_res('R2', 'SMBALERT_BUF1', 'SMBALERT_BUF2', value='22R'),
            'R3': make_res('R3', 'P3V3', 'SMBALERT_BUF2', value='4.7k'),
        }
        nets = {
            'SMBALERT_N': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'SMBALERT_N'},
                {'refdes': 'R1', 'pin': '1', 'pin_name': '1'},
            ],
            'SMBALERT_BUF1': [
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R2', 'pin': '1', 'pin_name': '1'},
            ],
            'SMBALERT_BUF2': [
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R3', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [{'refdes': 'R3', 'pin': '1', 'pin_name': '1'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets)
        self.assertEqual([], result['od_missing'])

    def test_analyze_resistors_prefers_real_page_for_duplicate_and_divider_rows(self):
        components = {
            'R1': make_res('R1', 'P3V3', 'ALERT_N', page='PAGE1', page_real='PAGE501'),
            'R2': make_res('R2', 'P3V3', 'ALERT_N', value='4.7k', page='PAGE2', page_real='PAGE502'),
            'R3': make_res('R3', 'ALERT_N', 'GPIO_BUF', value='22R', page='PAGE3', page_real='PAGE503'),
        }
        nets = {
            'P3V3': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}, {'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
            'ALERT_N': [
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R3', 'pin': '1', 'pin_name': '1'},
            ],
            'GPIO_BUF': [{'refdes': 'R3', 'pin': '2', 'pin_name': '2'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets)
        self.assertEqual('PAGE501, PAGE502', result['dup_pullups'][0]['页面'])
        divider_pages = {row['页面'] for row in result['divider_risks']}
        self.assertEqual({'PAGE503, PAGE501', 'PAGE503, PAGE502'}, divider_pages)

    def test_analyze_resistors_does_not_fallback_to_logical_page_when_real_page_missing(self):
        components = {
            'R1': make_res('R1', 'P3V3', 'ALERT_N', page='PAGE1', page_real=''),
            'R2': make_res('R2', 'P3V3', 'ALERT_N', value='4.7k', page='PAGE2', page_real=''),
        }
        nets = {
            'P3V3': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}, {'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
            'ALERT_N': [{'refdes': 'R1', 'pin': '2', 'pin_name': '2'}, {'refdes': 'R2', 'pin': '2', 'pin_name': '2'}],
        }
        row = pstx_analyzer.analyze_resistors(components, nets)['dup_pullups'][0]
        self.assertEqual('', row['页面'])

    def test_analyze_resistors_od_oc_skips_missing_when_pullup_is_across_series_resistor(self):
        components = {
            'U1': make_ic('U1', {'1': 'SMBALERT_N'}),
            'R1': make_res('R1', 'SMBALERT_N', 'SMBALERT_BUF', value='22R'),
            'R2': make_res('R2', 'P3V3', 'SMBALERT_BUF', value='4.7k'),
        }
        nets = {
            'SMBALERT_N': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'SMBALERT_N'},
                {'refdes': 'R1', 'pin': '1', 'pin_name': '1'},
            ],
            'SMBALERT_BUF': [
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
                {'refdes': 'R2', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [{'refdes': 'R2', 'pin': '1', 'pin_name': '1'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets)
        self.assertEqual([], result['od_missing'])

    def test_analyze_resistors_od_oc_requires_stronger_evidence_than_single_weak_token(self):
        components = {
            'U1': make_ic('U1', {'1': 'FAULT_MISC'}),
        }
        nets = {
            'FAULT_MISC': [{'refdes': 'U1', 'pin': '1', 'pin_name': 'GPIO1'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets)
        self.assertEqual([], result['od_missing'])

    def test_analyze_resistors_od_oc_reports_strong_candidate_without_pullup(self):
        components = {
            'U1': make_ic('U1', {'1': 'SMBALERT_N'}),
        }
        nets = {
            'SMBALERT_N': [{'refdes': 'U1', 'pin': '1', 'pin_name': 'SMBALERT_N'}],
        }
        row = pstx_analyzer.analyze_resistors(components, nets)['od_missing'][0]
        self.assertEqual('SMBALERT_N', row['网络名'])
        self.assertIn('SMBALERT', row['判定依据'])
        self.assertEqual('od_oc_strong_name_without_pullup', row['原因代码'])

    def test_analyze_resistors_ignores_depop_bias_resistors(self):
        components = {
            'U1': make_ic('U1', {'1': 'SMBALERT_N'}),
            'R1': make_res('R1', 'P3V3', 'SMBALERT_N', value='4.7k', bom_option='DEPOP'),
        }
        nets = {
            'SMBALERT_N': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'SMBALERT_N'},
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets)
        self.assertEqual({}, result['pullups'])
        self.assertEqual(1, len(result['od_missing']))

    def test_analyze_resistors_can_include_depop_when_requested(self):
        components = {
            'U1': make_ic('U1', {'1': 'SMBALERT_N'}),
            'R1': make_res('R1', 'P3V3', 'SMBALERT_N', value='4.7k', bom_option='DEPOP'),
        }
        nets = {
            'SMBALERT_N': [
                {'refdes': 'U1', 'pin': '1', 'pin_name': 'SMBALERT_N'},
                {'refdes': 'R1', 'pin': '2', 'pin_name': '2'},
            ],
            'P3V3': [{'refdes': 'R1', 'pin': '1', 'pin_name': '1'}],
        }
        result = pstx_analyzer.analyze_resistors(components, nets, exclude_depop=False)
        self.assertIn('SMBALERT_N', result['pullups'])
        self.assertEqual([], result['od_missing'])

    def test_analyze_networks_detects_diff_pairs_case_insensitively(self):
        nets = {
            'pcie_tx_p': [{'refdes': 'U1', 'pin': '1', 'pin_name': '1'}],
            'pcie_tx_n': [{'refdes': 'U1', 'pin': '2', 'pin_name': '2'}],
        }
        components = {'U1': {'page': 'PAGE1'}}
        result = pstx_analyzer.analyze_networks(nets, components)
        self.assertIn('pcie_tx', result['diff_pairs'])
        self.assertEqual('pcie_tx_p', result['diff_pairs']['pcie_tx']['P'])
        self.assertEqual('pcie_tx_n', result['diff_pairs']['pcie_tx']['N'])
        self.assertEqual('候选判断', result['diff_pair_rows'][0]['结论类型'])

    def test_analyze_networks_does_not_treat_lone_n_suffix_as_diff_pair(self):
        nets = {
            'pcie_tx_n': [{'refdes': 'U1', 'pin': '1', 'pin_name': '1'}],
        }
        components = {'U1': {'page': 'PAGE1'}}
        result = pstx_analyzer.analyze_networks(nets, components)
        self.assertEqual({}, result['diff_pairs'])

    def test_analyze_networks_uses_unknown_for_blank_page(self):
        nets = {}
        components = {
            'U1': {'page': ''},
            'U2': {'page': 'PAGE1'},
        }
        result = pstx_analyzer.analyze_networks(nets, components)
        self.assertEqual(1, result['page_counter']['UNKNOWN'])
        self.assertEqual(1, result['page_counter']['PAGE1'])

    def test_analyze_networks_prefers_real_page_for_page_rows(self):
        nets = {}
        components = {
            'U1': {'page': 'PAGE242', 'page_real': 'PAGE518'},
            'U2': {'page': 'PAGE242', 'page_real': 'PAGE518'},
            'U3': {'page': 'PAGE300', 'page_real': ''},
        }
        result = pstx_analyzer.analyze_networks(nets, components)
        self.assertEqual(2, result['page_counter']['PAGE518'])
        self.assertEqual(1, result['page_counter']['PAGE300'])
        self.assertEqual(['PAGE300', 'PAGE518'], [row['页面'] for row in result['page_rows']])

    def test_analyze_networks_normalizes_and_naturally_sorts_page_rows(self):
        nets = {}
        components = {
            'U1': {'page': 'PAGE01'},
            'U2': {'page': 'PAGE1'},
            'U3': {'page': 'PAGE10'},
            'U4': {'page': 'PAGE2'},
            'U5': {'page': ''},
        }
        result = pstx_analyzer.analyze_networks(nets, components)
        self.assertEqual(2, result['page_counter']['PAGE1'])
        self.assertEqual(['PAGE1', 'PAGE2', 'PAGE10', 'UNKNOWN'],
                         [row['页面'] for row in result['page_rows']])

    def test_analyze_networks_prefers_real_pages_over_hierarchical_logic_pages(self):
        nets = {}
        components = {
            'U1': {'page': 'PAGE1', 'page_real': 'PAGE1'},
            'U2': {'page': 'PAGE242 / PAGE1', 'page_real': 'PAGE601'},
            'U3': {'page': 'PAGE242 / PAGE2', 'page_real': 'PAGE602'},
        }
        result = pstx_analyzer.analyze_networks(nets, components)
        self.assertEqual(1, result['page_counter']['PAGE1'])
        self.assertEqual(1, result['page_counter']['PAGE601'])
        self.assertEqual(1, result['page_counter']['PAGE602'])

    def test_check_drc_does_not_report_testpoint_or_unnamed_net_as_single_pin_issue(self):
        components = {
            'TP1': {
                'refdes': 'TP1', 'part_name': 'testpoint', 'hq_code': '', 'value': '',
                'package': 'TP', 'material': '', 'tolerance': '', 'voltage': '',
                'current': '', 'power': '', 'bom_option': '', 'bom_cost': '',
                'room': '', 'drawing': 'SCH_PAGE1', 'page': 'PAGE1',
                'comp_type': 'TESTPOINT', 'nets': {},
            },
            'U1': {
                'refdes': 'U1', 'part_name': 'IC_CPU', 'hq_code': 'PN1', 'value': 'CPU',
                'package': 'BGA', 'material': '', 'tolerance': '', 'voltage': '',
                'current': '', 'power': '', 'bom_option': '', 'bom_cost': '',
                'room': '', 'drawing': 'SCH_PAGE1', 'page': 'PAGE1',
                'comp_type': 'IC', 'nets': {},
            },
        }
        nets = {
            'NET_ALONE': [{'refdes': 'TP1', 'pin': '1', 'pin_name': 'TP'}],
            'UNNAMED_1': [{'refdes': 'U1', 'pin': 'A1', 'pin_name': 'GPIO'}],
        }
        result = pstx_analyzer.check_drc(components, nets)
        self.assertEqual([], result['single_pin_nets'])
        self.assertEqual('候选判断', result['unnamed_nets'][0]['结论类型'])
        self.assertEqual('unnamed_net', result['unnamed_nets'][0]['原因代码'])

    def test_check_drc_bom_option_suggests_main(self):
        components = {
            'U1': {
                'refdes': 'U1', 'part_name': 'IC_CPU', 'hq_code': 'PN1', 'value': 'CPU',
                'package': 'BGA', 'material': '', 'tolerance': '', 'voltage': '',
                'current': '', 'power': '', 'bom_option': 'MIAN', 'bom_cost': '',
                'room': '', 'drawing': 'SCH_PAGE1', 'page': 'PAGE1',
                'comp_type': 'IC', 'nets': {},
            },
        }
        result = pstx_analyzer.check_drc(components, {})
        self.assertEqual('MAIN', result['bom_option_typos'][0]['疑似应为'])

    def test_check_drc_missing_value_is_confirmed_high_confidence(self):
        components = {
            'U1': {
                'refdes': 'U1', 'part_name': 'IC_CPU', 'hq_code': 'PN1', 'value': '',
                'package': 'BGA', 'material': '', 'tolerance': '', 'voltage': '',
                'current': '', 'power': '', 'bom_option': '', 'bom_cost': '',
                'room': '', 'drawing': 'SCH_PAGE1', 'page': 'PAGE1',
                'comp_type': 'IC', 'nets': {},
            },
        }
        row = pstx_analyzer.check_drc(components, {})['missing_value'][0]
        self.assertEqual('确定结论', row['结论类型'])
        self.assertEqual('高', row['严重级别'])
        self.assertEqual('高', row['置信度'])
        self.assertEqual('missing_value', row['原因代码'])

    def test_check_drc_lists_all_bom_option_components_with_depop_flag(self):
        components = {
            'U1': make_ic('U1', bom_option='DEPOP', page='PAGE1', page_real='PAGE518'),
            'R1': make_res('R1', 'P3V3', 'GPIO1', bom_option='ALT'),
            'C1': make_cap(),
        }
        rows = pstx_analyzer.check_drc(components, {})['bom_option_components']
        row_map = {row['位号']: row for row in rows}
        self.assertEqual({'U1', 'R1'}, set(row_map))
        self.assertEqual('是', row_map['U1']['是否DEPOP'])
        self.assertEqual('否', row_map['R1']['是否DEPOP'])
        self.assertEqual('PAGE518', row_map['U1']['页面'])

    def test_check_drc_prefers_real_page_for_issue_rows(self):
        components = {'U1': make_ic('U1', page='PAGE242', page_real='PAGE518')}
        components['U1']['hq_code'] = ''
        nets = {
            'GPIO_ALONE': [{'refdes': 'U1', 'pin': 'A1', 'pin_name': 'GPIO'}],
        }
        result = pstx_analyzer.check_drc(components, nets)
        self.assertEqual('PAGE518', result['missing_hq_code'][0]['页面'])
        self.assertEqual('PAGE518', result['single_pin_nets'][0]['页面'])

    def test_check_drc_does_not_fallback_to_logical_page_when_real_page_missing(self):
        components = {'U1': make_ic('U1', page='PAGE242', page_real='')}
        components['U1']['hq_code'] = ''
        nets = {
            'GPIO_ALONE': [{'refdes': 'U1', 'pin': 'A1', 'pin_name': 'GPIO'}],
        }
        result = pstx_analyzer.check_drc(components, nets)
        self.assertEqual('', result['missing_hq_code'][0]['页面'])
        self.assertEqual('', result['single_pin_nets'][0]['页面'])


class ExportTests(unittest.TestCase):
    def test_export_to_excel_includes_resistor_analysis_sheet(self):
        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        os.unlink(path)
        try:
            out = pstx_analyzer.export_to_excel({
                'project_name': 'demo',
                'bom_normal_detail': [],
                'bom_depop_detail': [],
                'bom_normal_merged': [],
                'bom_depop_merged': [],
                'net_analysis': {},
                'drc': {},
                'derating': [],
                'resistor_analysis': {
                    'divider_risks': [{'状态': '❌ 高风险'}],
                    'dup_pullups': [],
                    'dup_pulldowns': [],
                    'od_missing': [],
                    'chip_pin_rows': [],
                },
                'csa_geometry': {
                    'page_count': 1,
                    'cross_count': 1,
                    'circle_count': 0,
                    'summary_rows': [{'页面': 'PAGE3', 'DOT四向十字数': 1}],
                    'dot_cross_rows': [{'页面': 'PAGE3', '坐标': '(450,0)'}],
                    'circle_rows': [],
                },
            }, path)
            wb = load_workbook(out)
            try:
                self.assertIn('电阻检查', wb.sheetnames)
                self.assertIn('芯片引脚电阻', wb.sheetnames)
                self.assertIn('规范检查', wb.sheetnames)
            finally:
                wb.close()
        finally:
            if os.path.exists(path):
                os.unlink(path)
            if 'out' in locals() and os.path.exists(out):
                os.unlink(out)

    def test_xl_write_rows_aligns_values_by_header_name(self):
        wb = Workbook()
        ws = wb.active
        rows = [
            {'A': 'a1', 'B': 'b1', 'C': 'c1'},
            {'C': 'c2', 'A': 'a2', 'B': 'b2'},
        ]
        pstx_analyzer._xl_write_rows(ws, rows, pstx_analyzer._BL)
        self.assertEqual(['A', 'B', 'C'], [ws.cell(1, i).value for i in range(1, 4)])
        self.assertEqual(['a2', 'b2', 'c2'], [ws.cell(3, i).value for i in range(1, 4)])
        wb.close()


if __name__ == '__main__':
    unittest.main()
