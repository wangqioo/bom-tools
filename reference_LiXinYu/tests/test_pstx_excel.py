# -*- coding: utf-8 -*-
"""Tests for PSTX Excel export formatting."""

import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from pstx_exports import excel as pstx_excel


class ExcelExportTests(unittest.TestCase):

    def test_export_to_excel_includes_resistor_analysis_sheet(self):
        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        os.unlink(path)
        try:
            out = pstx_excel.export_to_excel({
                'project_name': 'demo',
                'bom_normal_detail': [],
                'bom_depop_detail': [],
                'bom_normal_merged': [],
                'bom_depop_merged': [],
                'net_analysis': {},
                'drc': {},
                'derating': [],
                'resistor_analysis': {
                    'divider_risks': [{'状态': '❌ 高风险'}],
                    'dup_pullups': [],
                    'dup_pulldowns': [],
                    'chip_pin_rows': [],
                },
                'csa_geometry': {
                    'page_count': 1,
                    'cross_count': 1,
                    'circle_count': 0,
                    'summary_rows': [{'页面': 'PAGE3', 'DOT四向十字数': 1}],
                    'dot_cross_rows': [{'页面': 'PAGE3', '坐标': '(450,0)'}],
                    'circle_rows': [],
                },
            }, path)
            wb = load_workbook(out)
            try:
                self.assertIn('电阻检查', wb.sheetnames)
                self.assertIn('芯片引脚电阻', wb.sheetnames)
                self.assertIn('规范检查', wb.sheetnames)
            finally:
                wb.close()
        finally:
            if os.path.exists(path):
                os.unlink(path)
            if 'out' in locals() and os.path.exists(out):
                os.unlink(out)

    def test_xl_write_rows_aligns_values_by_header_name(self):
        wb = Workbook()
        ws = wb.active
        rows = [
            {'A': 'a1', 'B': 'b1', 'C': 'c1'},
            {'C': 'c2', 'A': 'a2', 'B': 'b2'},
        ]
        pstx_excel._xl_write_rows(ws, rows, pstx_excel._BL)
        self.assertEqual(['A', 'B', 'C'], [ws.cell(1, i).value for i in range(1, 4)])
        self.assertEqual(['a2', 'b2', 'c2'], [ws.cell(3, i).value for i in range(1, 4)])
        wb.close()

    def test_export_bom_to_excel_supports_depop_modes(self):
        data = {
            'project_name': 'bom-demo',
            'bom_normal_merged': [
                {'序号': 1, '料号': 'PN-SAME', '位号列表': 'R1', '数量': 1, '描述': 'RES', '值': '10k', '封装': '0402', '类型': '电阻'},
            ],
            'bom_depop_merged': [
                {'序号': 1, '料号': 'PN-SAME', '位号列表': 'R2', '数量': 1, '描述': 'RES', '值': '10k', '封装': '0402', '类型': '电阻'},
            ],
            'bom_total_merged': [
                {
                    '序号': 1,
                    '料号': 'PN-SAME',
                    '位号列表': 'R1, R2',
                    '数量': 2,
                    '贴装数量': 1,
                    'DEPOP数量': 1,
                    'BOM状态': '贴装 / DEPOP',
                    '描述': 'RES',
                    '值': '10k',
                    '封装': '0402',
                    '类型': '电阻',
                },
            ],
        }
        rows_all = pstx_excel.build_bom_export_rows(data, 'all')
        rows_mounted = pstx_excel.build_bom_export_rows(data, 'mounted')
        rows_split = pstx_excel.build_bom_export_rows(data, 'split')

        self.assertEqual(1, len(rows_all))
        self.assertEqual(2, rows_all[0]['数量'])
        self.assertEqual(1, len(rows_mounted))
        self.assertEqual('R1', rows_mounted[0]['位号列表'])
        self.assertEqual(['贴装', 'DEPOP'], [row['BOM状态'] for row in rows_split])
        self.assertEqual(['PN-SAME', 'PN-SAME'], [row['料号'] for row in rows_split])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        os.unlink(path)
        try:
            out = pstx_excel.export_bom_to_excel(data, path, mode='split')
            wb = load_workbook(out)
            try:
                self.assertIn('说明', wb.sheetnames)
                self.assertIn('BOM_分条', wb.sheetnames)
                ws = wb['BOM_分条']
                self.assertEqual('BOM状态', ws['B1'].value)
                self.assertEqual('贴装', ws['B2'].value)
                self.assertEqual('DEPOP', ws['B3'].value)
            finally:
                wb.close()
        finally:
            if os.path.exists(path):
                os.unlink(path)
            if 'out' in locals() and os.path.exists(out):
                os.unlink(out)


if __name__ == '__main__':
    unittest.main()
