import os
import json
import socket
import sqlite3
import tarfile
import tempfile
import time
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from pstx_integrations.aster.service import clear_aster_runtime_config
from pstx_harness.model import HarnessModelResponse
from pstx_webapp import app_factory as webapp_factory
from pstx_webapp import agent_context as webapp_agent_context
from pstx_webapp import compare_view as webapp_compare_view
from pstx_webapp import form_parsing as webapp_form_parsing
from pstx_webapp import json_utils as webapp_json_utils
from pstx_webapp import pages as webapp_pages
from pstx_webapp import project_io as webapp_project_io
from pstx_webapp.report_tables import build_report_table, build_review_plan, build_table_display_policy
from pstx_webapp import run_store as webapp_run_store
from pstx_webapp import server as webapp_server
from pstx_webapp import state as webapp_state
from pstx_webapp.compare_payload import build_compare_payload


PRT_SAMPLE = (
    "PART_NAME\n"
    "U1 'IC_CPU'\n"
    "HQ_CODE='PN_U1'\n"
    "VALUE='CPU'\n"
    "PACKAGE='BGA'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'\n"
    "PART_NAME\n"
    "R1 'RES_0402'\n"
    "VALUE='4.7k'\n"
    "PACKAGE='0402'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I2'\n"
)

NET_SAMPLE = (
    "NET_NAME\n"
    "'SMBALERT_N'\n"
    "NODE_NAME U1 1\n"
    "'SMBALERT_N':\n"
    "NODE_NAME R1 2\n"
    "'2':\n"
    "NET_NAME\n"
    "'P3V3'\n"
    "NODE_NAME R1 1\n"
    "'1':\n"
)

PRT_SAMPLE_TOPOLOGY = (
    "PART_NAME\n"
    "U46 'LCMXO3LF_9400C_HDL-HQ11112042009,LCMXO3LF-9400C-5BG484C'\n"
    "HQ_CODE='HQ11112042009'\n"
    "VALUE='LCMXO3LF-9400C-5BG484C'\n"
    "PACKAGE='BGA484'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE131_I1'\n"
    "PART_NAME\n"
    "U12 'TXS0108_LEVEL_TRANSLATOR'\n"
    "HQ_CODE='HQ-LS'\n"
    "VALUE='TXS0108'\n"
    "PACKAGE='QFN'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE88_I2'\n"
    "PART_NAME\n"
    "J8 'CONN_PCIE'\n"
    "VALUE='PCIE_CONN'\n"
    "PACKAGE='CONN'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE155_I3'\n"
)

NET_SAMPLE_TOPOLOGY = (
    "NET_NAME\n"
    "'I2C_SCL'\n"
    "NODE_NAME U46 A1\n"
    "'A1':\n"
    "NODE_NAME U12 A1\n"
    "'A1':\n"
    "NET_NAME\n"
    "'PCE_TX0_P'\n"
    "NODE_NAME U46 B1\n"
    "'B1':\n"
    "NODE_NAME J8 1\n"
    "'1':\n"
    "NET_NAME\n"
    "'PCE_TX0_N'\n"
    "NODE_NAME U46 B2\n"
    "'B2':\n"
    "NODE_NAME J8 2\n"
    "'2':\n"
)

PRT_SAMPLE_COMPARE = (
    "PART_NAME\n"
    "U1 'IC_CPU'\n"
    "HQ_CODE='PN_U1B'\n"
    "VALUE='CPU_B'\n"
    "PACKAGE='BGA'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'\n"
    "PART_NAME\n"
    "R1 'RES_0402'\n"
    "VALUE='4.7k'\n"
    "PACKAGE='0402'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I2'\n"
    "PART_NAME\n"
    "C2 'CAP_0402'\n"
    "VALUE='1uF'\n"
    "PACKAGE='0402'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I3'\n"
)

NET_SAMPLE_COMPARE = (
    "NET_NAME\n"
    "'SMBALERT_ALT_N'\n"
    "NODE_NAME U1 1\n"
    "'SMBALERT_ALT_N':\n"
    "NODE_NAME R1 2\n"
    "'2':\n"
    "NET_NAME\n"
    "'P3V3'\n"
    "NODE_NAME R1 1\n"
    "'1':\n"
    "NODE_NAME C2 1\n"
    "'1':\n"
    "NET_NAME\n"
    "'GND'\n"
    "NODE_NAME C2 2\n"
    "'2':\n"
)

CSA_DOT_CROSS_SAMPLE = (
    "FILE_TYPE = MACRO_DRAWING;\n"
    "SET PAGE_NUMBER P3;\n"
    "WIRE 16 -1 (400 0)(500 0);\n"
    "FORCEPROP 2 LAST SIG_NAME CROSS_DOT_H\n"
    "WIRE 16 -1 (450 -50)(450 50);\n"
    "FORCEPROP 2 LAST SIG_NAME CROSS_DOT_V\n"
    "DOT 1 (450 0);\n"
    "CIRCLE 16 -1 (1000 1000)(1100 1000);\n"
)

PRT_SAMPLE_DEPOP = (
    "PART_NAME\n"
    "U1 'IC_CPU'\n"
    "HQ_CODE='PN_U1'\n"
    "VALUE='CPU'\n"
    "PACKAGE='BGA'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'\n"
    "PART_NAME\n"
    "R1 'RES_0402'\n"
    "VALUE='4.7k'\n"
    "PACKAGE='0402'\n"
    "BOM_OPTION='DEPOP'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I2'\n"
)

PRT_SAMPLE_DEPOP_WITH_XY = (
    "PART_NAME\n"
    "U1 'IC_CPU'\n"
    "HQ_CODE='PN_U1'\n"
    "VALUE='CPU'\n"
    "PACKAGE='BGA'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'\n"
    "PART_NAME\n"
    "R1 'RES_0402'\n"
    "VALUE='4.7k'\n"
    "PACKAGE='0402'\n"
    "BOM_OPTION='DEPOP'\n"
    "XY='(1000,1000)'\n"
    "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I2'\n"
)

NET_SAMPLE_DEPOP = (
    "NET_NAME\n"
    "'SMBALERT_N'\n"
    "NODE_NAME U1 1\n"
    "'SMBALERT_N':\n"
    "NODE_NAME R1 2\n"
    "'2':\n"
    "NET_NAME\n"
    "'P3V3'\n"
    "NODE_NAME R1 1\n"
    "'1':\n"
)

PRT_SAMPLE_PAGE_V2 = (
    "PART_NAME\n"
    "C1A104 'CAP_HDL-HQ17101005HS0,100NF,10%,0402,X7R,50V':\n"
    "SECTION_NUMBER 1\n"
    " '@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE144_I70"
    "@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE1_I17"
    "@HQ_CAP.CAP_HDL(CHIPS)':\n"
    " C_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page144_i70"
    "@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page1_i17"
    "@hq_cap.cap_hdl(chips)',\n"
    " P_PATH='@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page114_i70"
    "@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1):page1_i17"
    "@hq_cap.cap_hdl(chips)',\n"
    " DRAWING='@GPU_2SW_BOARD_LIB.I2C_REPEATER_9617_CBB_V3(SCH_1):PAGE1'\n"
    " BOM_OPTION='DEPOP'\n"
    " HQ_CODE='HQ17101005HS0'\n"
    " VALUE='100NF'\n"
    " PACKAGE='0402'\n"
)

NET_SAMPLE_PAGE_V2 = (
    "NET_NAME\n"
    "'P1V8_AON'\n"
    "NODE_NAME C1A104 1\n"
    "'1':\n"
    "NET_NAME\n"
    "'GND'\n"
    "NODE_NAME C1A104 2\n"
    "'2':\n"
)


def build_project_root() -> Path:
    return build_project_root_with_samples(PRT_SAMPLE, NET_SAMPLE)


def build_project_root_with_samples(prt_text: str, net_text: str) -> Path:
    root = Path(tempfile.mkdtemp())
    packaged = root / 'packaged'
    packaged.mkdir(parents=True)
    (packaged / 'pstxprt.dat').write_text(prt_text, encoding='utf-8')
    (packaged / 'pstxnet.dat').write_text(net_text, encoding='utf-8')
    (packaged / 'pstxref.dat').write_text('xref placeholder', encoding='utf-8')

    sch_dir = root / 'sch_1'
    sch_dir.mkdir(parents=True)
    (sch_dir / 'page518.csv').write_text('NAME,PAGE_NUMBER\nTOP,242\n', encoding='utf-8')
    return root


def write_minimal_pdf(path: Path, *, page_count: int = 1, width: int = 600, height: int = 800) -> None:
    objects = [
        "1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        "2 0 obj << /Type /Pages /Kids [" + " ".join(f"{idx + 3} 0 R" for idx in range(page_count)) + f"] /Count {page_count} >> endobj\n",
    ]
    for idx in range(page_count):
        obj_id = idx + 3
        objects.append(
            f"{obj_id} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] /Resources << >> >> endobj\n"
        )
    path.write_bytes(("%PDF-1.4\n" + "".join(objects) + "trailer << /Root 1 0 R >>\n%%EOF\n").encode("ascii"))


def build_project_root_for_page_v2() -> Path:
    root = Path(tempfile.mkdtemp())
    packaged = root / 'packaged'
    packaged.mkdir(parents=True)
    (packaged / 'pstxprt.dat').write_text(PRT_SAMPLE_PAGE_V2, encoding='utf-8')
    (packaged / 'pstxnet.dat').write_text(NET_SAMPLE_PAGE_V2, encoding='utf-8')
    (packaged / 'pstxref.dat').write_text('xref placeholder', encoding='utf-8')

    sch_dir = root / 'sch_1'
    sch_dir.mkdir(parents=True)
    (sch_dir / 'page114.csv').write_text('"PAGE_NUMBER" = 144;\n', encoding='utf-8')
    (sch_dir / 'page.map').write_text('144 114 TOP\n', encoding='utf-8')
    (root / 'module_order').write_text(
        'Version 15.0\n'
        'START_MODULEORDER\n'
        '@gpu_2sw_board_lib.gpu_2sw_board(sch_1):page114_i70'
        '@gpu_2sw_board_lib.i2c_repeater_9617_cbb_v3(sch_1) 0 1 177 34 0\n'
        'END_MODULEORDER\n',
        encoding='utf-8',
    )
    return root


class WebUiTests(unittest.TestCase):
    def test_webapp_app_factory_exports_only_app_factory(self):
        self.assertEqual(["create_app"], webapp_factory.__all__)
        self.assertFalse(hasattr(webapp_factory, "_resolve_port"))

    def test_webapp_state_clear_resets_session_caches(self):
        webapp_state.RUN_CACHE["run-test"] = {"ok": True}
        webapp_state.AGENT_CONTEXT_CACHE["run-test"] = {"answers": []}
        webapp_state.AGENT_RUN_CACHE.remember({"agent_run_id": "agent-test"}, agent_run_id="agent-test")
        webapp_state.clear_web_session_state()
        self.assertFalse(webapp_state.RUN_CACHE)
        self.assertFalse(webapp_state.AGENT_CONTEXT_CACHE)
        self.assertIsNone(webapp_state.AGENT_RUN_CACHE.get("agent-test"))

    def test_webapp_run_store_remember_and_project_summary(self):
        payload = {
            "bundle": {
                "project_name": "demo",
                "project_root": "/tmp/demo",
                "components": {"U1": {}, "R1": {}},
                "nets": {"N1": {}},
                "drc": {"missing_hq_code": [{"refdes": "U1"}]},
            },
            "report": {
                "metrics": [{"label": "贴装总数", "value": 2}],
                "include_depop": True,
            },
        }
        webapp_run_store.remember_run("run-demo", payload)

        self.assertIs(payload, webapp_run_store.get_run("run-demo"))
        summary = webapp_run_store.build_project_summary(
            "run-demo",
            payload,
            drc_issue_keys=("missing_hq_code",),
        )
        self.assertEqual("demo", summary["project_name"])
        self.assertEqual(2, summary["component_count"])
        self.assertEqual(1, summary["net_count"])
        self.assertEqual(1, summary["drc_count"])
        self.assertEqual(2, summary["metric_map"]["贴装总数"])
        self.assertEqual(["run-demo"], [item["run_id"] for item in webapp_run_store.list_project_summaries()])

    def test_webapp_compare_view_helpers_are_canonical(self):
        self.assertEqual(500, webapp_compare_view.coerce_compare_detail_limit(""))
        self.assertEqual(5000, webapp_compare_view.coerce_compare_detail_limit("99999"))
        with self.assertRaisesRegex(ValueError, "必须大于 0"):
            webapp_compare_view.coerce_compare_detail_limit("0")

        rows = webapp_compare_view.build_compare_scalar_metrics(
            {"metric_map": {"贴装总数": 5}, "component_count": 10, "net_count": 3, "drc_count": 1},
            {"metric_map": {"贴装总数": 7}, "component_count": 8, "net_count": 3, "drc_count": 2},
        )
        row_map = {item["指标"]: item for item in rows}
        self.assertEqual("+2", row_map["贴装总数"]["变化"])
        self.assertEqual("-2", row_map["元件数"]["变化"])
        self.assertNotIn("网络数", row_map)

    def test_webapp_json_utils_are_canonical(self):
        self.assertEqual('{"a": 1}', webapp_json_utils.json_fingerprint({"a": 1}))
        self.assertEqual("ab…", webapp_json_utils.compact_value("abcd", limit=3))

    def test_webapp_agent_context_helpers_are_canonical(self):
        context = webapp_agent_context.get_agent_context("run-context")
        webapp_agent_context.append_agent_context_answers(
            context,
            ({"question_id": "q1", "answer": "补充规格", "applies_to": {"refdes": "U1"}},),
            source_agent_run_id="agent-a",
        )
        public = webapp_agent_context.agent_context_public("run-context", context)
        self.assertEqual(1, public["answer_count"])
        self.assertEqual("补充规格", public["answers"][0]["answer"])
        self.assertEqual("memory", public["storage"])

    def test_webapp_server_and_form_helpers_are_canonical(self):
        mapping, warnings = webapp_form_parsing.parse_voltage_map_text("P3V3=3.3\nbad-line")
        self.assertEqual({"P3V3": 3.3}, mapping)
        self.assertTrue(warnings)
        self.assertTrue(webapp_form_parsing.parse_checkbox_flag("on"))
        self.assertFalse(webapp_server.port_is_available(70000))

    def test_webapp_pages_render_helpers_keep_template_contract(self):
        calls = []

        def fake_render(template_name, **context):
            calls.append((template_name, context))
            return f"rendered:{template_name}"

        self.assertEqual("rendered:index.html", webapp_pages.render_home_page(
            fake_render,
            request_host="127.0.0.1:45555",
            default_host="127.0.0.1",
            default_port=44441,
        ))
        self.assertEqual(("index.html", {"listen_host": "127.0.0.1", "listen_port": "45555"}), calls[-1])
        self.assertEqual("rendered:compare.html", webapp_pages.render_named_page(fake_render, "compare"))
        self.assertEqual("rendered:topology.html", webapp_pages.render_named_page(fake_render, "topology"))
        self.assertEqual("rendered:report.html", webapp_pages.render_report_page(
            fake_render,
            run_id="run-a",
            report={"ok": True},
        ))
        self.assertEqual((
            "report.html",
            {"run_id": "run-a", "report": {"ok": True}, "debug_ui": False, "debug_fixture": False},
        ), calls[-1])

    def setUp(self):
        clear_aster_runtime_config()
        webapp_state.clear_web_session_state()
        self._old_diagnostics_log_file = os.environ.get('PSTX_DIAGNOSTICS_LOG_FILE')
        self._old_dfmea_data_dir = os.environ.get('PSTX_DFMEA_DATA_DIR')
        self._diagnostics_tmp = tempfile.TemporaryDirectory()
        self._dfmea_tmp = tempfile.TemporaryDirectory()
        os.environ['PSTX_DIAGNOSTICS_LOG_FILE'] = str(Path(self._diagnostics_tmp.name) / 'pstx_diagnostics.log')
        os.environ['PSTX_DFMEA_DATA_DIR'] = self._dfmea_tmp.name
        self.app = webapp_factory.create_app()
        self.app.testing = True
        self.client = self.app.test_client()
        self.temp_roots = []

    def tearDown(self):
        clear_aster_runtime_config()
        if self._old_diagnostics_log_file is None:
            os.environ.pop('PSTX_DIAGNOSTICS_LOG_FILE', None)
        else:
            os.environ['PSTX_DIAGNOSTICS_LOG_FILE'] = self._old_diagnostics_log_file
        if self._old_dfmea_data_dir is None:
            os.environ.pop('PSTX_DFMEA_DATA_DIR', None)
        else:
            os.environ['PSTX_DFMEA_DATA_DIR'] = self._old_dfmea_data_dir
        self._diagnostics_tmp.cleanup()
        self._dfmea_tmp.cleanup()
        for root in self.temp_roots:
            for path in sorted(root.rglob('*'), reverse=True):
                if path.is_file():
                    path.unlink()
                elif path.is_dir():
                    path.rmdir()
            if root.exists():
                root.rmdir()

    def make_root(self) -> Path:
        root = build_project_root()
        self.temp_roots.append(root)
        return root

    def make_root_with_samples(self, prt_text: str, net_text: str) -> Path:
        root = build_project_root_with_samples(prt_text, net_text)
        self.temp_roots.append(root)
        return root

    def make_page_v2_root(self) -> Path:
        root = build_project_root_for_page_v2()
        self.temp_roots.append(root)
        return root

    def with_env(self, updates: dict):
        old_values = {key: os.environ.get(key) for key in updates}
        for key, value in updates.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value

        def restore():
            for key, value in old_values.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value

        self.addCleanup(restore)

    def make_fake_feishu_data(self) -> Path:
        root = Path(tempfile.mkdtemp())
        self.temp_roots.append(root)
        (root / 'feishu_libraries.json').write_text(
            json.dumps({
                'base_url': 'https://mcenter.example.local',
                'origin': 'cli_demo',
                'user_id': '100001',
                'libraries': [{'id': 'lib1', 'name': '优选库'}],
            }, ensure_ascii=False),
            encoding='utf-8',
        )
        conn = sqlite3.connect(root / 'feishu_cache.db')
        conn.execute(
            """
            CREATE TABLE materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lib_id TEXT,
                lib_name TEXT,
                sheet_name TEXT,
                key_value TEXT COLLATE NOCASE,
                hq_no TEXT,
                brand TEXT,
                spec TEXT,
                description TEXT,
                raw_data TEXT,
                synced_at TEXT
            )
            """
        )
        conn.execute(
            "INSERT INTO materials(lib_id,lib_name,sheet_name,key_value,hq_no,brand,spec,description,raw_data,synced_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)",
            ('lib1', '优选库', 'Sheet1', 'RES_0402', 'HQ-R', 'ACME', 'RES_0402', 'resistor', '{}', '2026-04-26'),
        )
        conn.commit()
        conn.close()
        self.with_env({'PSTX_FEISHU_DATA_DIR': str(root)})
        return root

    def add_fake_feishu_material(self,
                                  *,
                                  key_value: str,
                                  hq_no: str,
                                  spec: str = '',
                                  pi: str = '',
                                  selection_order: str = '') -> None:
        data_dir = Path(os.environ['PSTX_FEISHU_DATA_DIR'])
        conn = sqlite3.connect(data_dir / 'feishu_cache.db')
        existing = {
            row[1]
            for row in conn.execute("PRAGMA table_info(materials)").fetchall()
        }
        for column_name in ['pi', 'selection_order', 'extra_fields']:
            if column_name not in existing:
                conn.execute(f"ALTER TABLE materials ADD COLUMN {column_name} TEXT")
        conn.execute(
            "INSERT INTO materials(lib_id,lib_name,sheet_name,key_value,hq_no,brand,spec,description,pi,selection_order,extra_fields,raw_data,synced_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                'lib1',
                '优选库',
                'Sheet1',
                key_value,
                hq_no,
                'ACME',
                spec,
                'from feishu',
                pi,
                selection_order,
                '{}',
                '{}',
                '2026-04-26',
            ),
        )
        conn.commit()
        conn.close()

    def test_parse_voltage_map_text_reports_invalid_lines(self):
        mapping, warnings = webapp_form_parsing.parse_voltage_map_text("P1V8=1.8\nINVALID\nBAD=abc")
        self.assertEqual({'P1V8': 1.8}, mapping)
        self.assertEqual(2, len(warnings))

    def test_discover_project_files_uses_packaged_under_project_root(self):
        root = self.make_root()
        project_root, prt_path, net_path, ref_path = webapp_project_io.discover_project_files(str(root))
        self.assertEqual(root, project_root)
        self.assertEqual(root / 'packaged' / 'pstxprt.dat', prt_path)
        self.assertEqual(root / 'packaged' / 'pstxnet.dat', net_path)
        self.assertEqual(root / 'packaged' / 'pstxref.dat', ref_path)

    def test_discover_project_files_resolves_cpm_container_worklib_project(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            container = Path(temp_dir) / 'A' / 'B'
            project_root = container / 'worklib' / 'MAIN_MOD'
            packaged = project_root / 'packaged'
            packaged.mkdir(parents=True)
            (container / 'MAIN_MOD.cpm').write_text('placeholder', encoding='utf-8')
            (packaged / 'pstxprt.dat').write_text(PRT_SAMPLE, encoding='utf-8')
            (packaged / 'pstxnet.dat').write_text(NET_SAMPLE, encoding='utf-8')
            (packaged / 'pstxref.dat').write_text('xref', encoding='utf-8')

            resolved, prt_path, net_path, ref_path, snapshot = webapp_project_io.discover_project_files_with_snapshot(str(container))

        self.assertEqual(project_root, resolved)
        self.assertEqual(project_root / 'packaged' / 'pstxprt.dat', prt_path)
        self.assertEqual(project_root / 'packaged' / 'pstxnet.dat', net_path)
        self.assertEqual(project_root / 'packaged' / 'pstxref.dat', ref_path)
        self.assertFalse(snapshot['enabled'])
        self.assertEqual('MAIN_MOD', snapshot['module_name'])

    def test_discover_project_files_copies_archive_to_local_snapshot_before_analysis(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            self.with_env({'PSTX_PROJECT_SNAPSHOT_DIR': str(Path(temp_dir) / 'snapshots')})
            container = Path(temp_dir) / 'smb' / 'board'
            project_root = container / 'worklib' / 'MAIN_MOD'
            packaged = project_root / 'packaged'
            packaged.mkdir(parents=True)
            (container / 'MAIN_MOD.cpm').write_text('placeholder', encoding='utf-8')
            (packaged / 'pstxprt.dat').write_text(PRT_SAMPLE, encoding='utf-8')
            (packaged / 'pstxnet.dat').write_text(NET_SAMPLE, encoding='utf-8')
            (packaged / 'pstxref.dat').write_text('xref', encoding='utf-8')
            archive_path = container / 'MAIN_MOD_project.zip'
            with zipfile.ZipFile(archive_path, 'w') as archive:
                for path in project_root.rglob('*'):
                    if path.is_file():
                        archive.write(path, path.relative_to(container))

            resolved, prt_path, net_path, _ref_path, snapshot = webapp_project_io.discover_project_files_with_snapshot(str(container))

            self.assertTrue(snapshot['enabled'])
            self.assertEqual(str(archive_path), snapshot['source_archive'])
            self.assertTrue(Path(snapshot['local_archive']).is_file())
            self.assertTrue(Path(snapshot['snapshot_root']).is_dir())
            self.assertIn('snapshots', str(resolved))
            self.assertNotEqual(project_root, resolved)
            self.assertEqual(resolved / 'packaged' / 'pstxprt.dat', prt_path)
            self.assertEqual(resolved / 'packaged' / 'pstxnet.dat', net_path)

    def test_archive_snapshot_prefers_outer_cpm_module_when_archive_has_extra_cpm(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            self.with_env({'PSTX_PROJECT_SNAPSHOT_DIR': str(Path(temp_dir) / 'snapshots')})
            container = Path(temp_dir) / 'smb' / 'board'
            project_root = container / 'worklib' / 'MAIN_MOD'
            packaged = project_root / 'packaged'
            packaged.mkdir(parents=True)
            (container / 'MAIN_MOD.cpm').write_text('placeholder', encoding='utf-8')
            (packaged / 'pstxprt.dat').write_text(PRT_SAMPLE, encoding='utf-8')
            (packaged / 'pstxnet.dat').write_text(NET_SAMPLE, encoding='utf-8')
            archive_path = container / 'MAIN_MOD_project.zip'
            with zipfile.ZipFile(archive_path, 'w') as archive:
                archive.write(container / 'MAIN_MOD.cpm', 'MAIN_MOD.cpm')
                archive.writestr('OTHER_MOD.cpm', 'placeholder')
                for path in project_root.rglob('*'):
                    if path.is_file():
                        archive.write(path, path.relative_to(container))

            resolved, _prt_path, _net_path, _ref_path, snapshot = webapp_project_io.discover_project_files_with_snapshot(str(container))

            self.assertTrue(snapshot['enabled'])
            self.assertEqual('MAIN_MOD', snapshot['module_name'])
            self.assertEqual('MAIN_MOD', resolved.name)
            self.assertIn('snapshots', str(resolved))

    def test_old_worklib_project_input_uses_sibling_archive_snapshot(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            self.with_env({'PSTX_PROJECT_SNAPSHOT_DIR': str(Path(temp_dir) / 'snapshots')})
            container = Path(temp_dir) / 'smb' / 'board'
            project_root = container / 'worklib' / 'MAIN_MOD'
            packaged = project_root / 'packaged'
            packaged.mkdir(parents=True)
            (container / 'MAIN_MOD.cpm').write_text('placeholder', encoding='utf-8')
            (packaged / 'pstxprt.dat').write_text(PRT_SAMPLE, encoding='utf-8')
            (packaged / 'pstxnet.dat').write_text(NET_SAMPLE, encoding='utf-8')
            archive_path = container.parent / 'MAIN_MOD_board_snapshot.zip'
            with zipfile.ZipFile(archive_path, 'w') as archive:
                for path in project_root.rglob('*'):
                    if path.is_file():
                        archive.write(path, path.relative_to(container))

            resolved, _prt_path, _net_path, _ref_path, snapshot = webapp_project_io.discover_project_files_with_snapshot(str(project_root))

            self.assertTrue(snapshot['enabled'])
            self.assertEqual(str(archive_path), snapshot['source_archive'])
            self.assertIn('snapshots', str(resolved))
            self.assertNotEqual(project_root, resolved)

    def test_archive_snapshot_rejects_tar_special_members(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            self.with_env({'PSTX_PROJECT_SNAPSHOT_DIR': str(Path(temp_dir) / 'snapshots')})
            archive_path = Path(temp_dir) / 'bad_project.tar'
            with tarfile.open(archive_path, 'w') as archive:
                info = tarfile.TarInfo('badfifo')
                info.type = tarfile.FIFOTYPE
                archive.addfile(info)

            with self.assertRaisesRegex(ValueError, '非普通文件'):
                webapp_project_io.discover_project_files_with_snapshot(str(archive_path))

    def test_analyze_uses_local_snapshot_for_cpm_container_archive(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            self.with_env({'PSTX_PROJECT_SNAPSHOT_DIR': str(Path(temp_dir) / 'snapshots')})
            container = Path(temp_dir) / 'smb' / 'board'
            project_root = container / 'worklib' / 'MAIN_MOD'
            packaged = project_root / 'packaged'
            packaged.mkdir(parents=True)
            (container / 'MAIN_MOD.cpm').write_text('placeholder', encoding='utf-8')
            (packaged / 'pstxprt.dat').write_text(PRT_SAMPLE, encoding='utf-8')
            (packaged / 'pstxnet.dat').write_text(NET_SAMPLE, encoding='utf-8')
            (packaged / 'pstxref.dat').write_text('xref', encoding='utf-8')
            archive_path = container / 'MAIN_MOD_project.zip'
            with zipfile.ZipFile(archive_path, 'w') as archive:
                for path in project_root.rglob('*'):
                    if path.is_file():
                        archive.write(path, path.relative_to(container))

            response = self.client.post('/api/analyze', data={'project_root': str(container), 'project_name': 'snapshot-demo'})

            self.assertEqual(200, response.status_code)
            run_id = response.get_json()['run_id']
            cached = webapp_run_store.get_run(run_id)
            snapshot = cached['bundle']['project_input_snapshot']
            self.assertTrue(snapshot['enabled'])
            self.assertEqual(str(archive_path), snapshot['source_archive'])
            self.assertTrue(Path(snapshot['local_archive']).is_file())
            self.assertIn('snapshots', cached['bundle']['project_root'])
            self.assertIn('snapshots', cached['report']['input_files'][0]['filename'])
            projects_payload = self.client.get('/api/projects').get_json()
            project_summary = next(item for item in projects_payload['projects'] if item['run_id'] == run_id)
            self.assertTrue(project_summary['project_input_snapshot']['enabled'])
            self.assertEqual(str(archive_path), project_summary['project_input_snapshot']['source_archive'])

    def test_feishu_bom_status_endpoint_reads_bridge_status(self):
        self.make_fake_feishu_data()
        response = self.client.get('/api/feishu-bom/status')
        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['available'])
        self.assertTrue(payload['configured'])
        self.assertEqual(1, payload['library_count'])
        self.assertEqual(1, payload['cache_count'])
        self.assertEqual('优选库', payload['cache_stats'][0]['lib_name'])
        self.assertTrue(payload['online_debug_log_file'].endswith('feishu_bom_debug.log'))

    def test_feishu_bom_sheets_endpoint_uses_rewritten_client(self):
        calls = []
        old_fetch = webapp_factory.fetch_feishu_sheet_list

        def fake_fetch(**kwargs):
            calls.append(kwargs)
            return {
                'ok': True,
                'spreadsheet_token': 'token123',
                'sheet_count': 1,
                'sheets': [{'sheet_id': 'sh1', 'title': 'Sheet1', 'row_count': 3}],
            }

        webapp_factory.fetch_feishu_sheet_list = fake_fetch
        self.addCleanup(lambda: setattr(webapp_factory, 'fetch_feishu_sheet_list', old_fetch))

        response = self.client.post(
            '/api/feishu-bom/sheets',
            json={
                'base_url': 'https://mcenter.example.local',
                'origin': 'cli_demo',
                'user_id': '100001',
                'spreadsheet_token_or_url': 'https://example.feishu.cn/sheets/token123',
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual('sh1', payload['sheets'][0]['sheet_id'])
        self.assertEqual('https://example.feishu.cn/sheets/token123', calls[0]['spreadsheet_token_or_url'])
        self.assertEqual('100001', calls[0]['user_id'])

    def test_feishu_bom_sync_endpoint_updates_cache(self):
        calls = []
        old_sync = webapp_factory.sync_feishu_library

        def fake_sync(**kwargs):
            calls.append(kwargs)
            return {
                'ok': True,
                'library_id': kwargs['library_id'] or 'lib1',
                'library_name': kwargs['library_name'],
                'synced_rows': 2,
                'skipped_sheets': 0,
                'per_sheet': [{'sheet_id': 'sh1', 'status': 'synced', 'row_count': 2}],
            }

        webapp_factory.sync_feishu_library = fake_sync
        self.addCleanup(lambda: setattr(webapp_factory, 'sync_feishu_library', old_sync))

        response = self.client.post(
            '/api/feishu-bom/sync',
            json={
                'base_url': 'https://mcenter.example.local',
                'origin': 'cli_demo',
                'user_id': '100001',
                'library_id': 'lib1',
                'library_name': '优选库',
                'spreadsheet_token_or_url': 'token123',
                'sheets': [{
                    'sheet_id': 'sh1',
                    'title': 'Sheet1',
                    'spec_model_col': '厂家型号',
                    'hq_code_col': 'HQ料号',
                    'pi_col': 'PI',
                }],
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual(2, payload['synced_rows'])
        self.assertEqual('优选库', calls[0]['library_name'])
        self.assertEqual('厂家型号', calls[0]['sheets'][0]['spec_model_col'])
        self.assertEqual('HQ料号', calls[0]['sheets'][0]['hq_code_col'])

    def test_feishu_sync_page_is_available_before_project_analysis(self):
        response = self.client.get('/feishu-sync')
        self.assertEqual(200, response.status_code)
        body = response.get_data(as_text=True)
        self.assertIn('飞书优选库同步页', body)
        self.assertIn('Agent 辅助表头识别', body)

    def test_feishu_db_page_is_available_and_lists_cache(self):
        self.make_fake_feishu_data()
        page = self.client.get('/feishu-db')
        self.assertEqual(200, page.status_code)
        page_text = page.get_data(as_text=True)
        self.assertIn('飞书缓存数据库', page_text)
        self.assertIn('新增、编辑或剔除单行本地缓存', page_text)
        self.assertIn('feishu-db-row-editor', page_text)

        overview = self.client.get('/api/feishu-bom/database')
        self.assertEqual(200, overview.status_code)
        payload = overview.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual(1, payload['cache_count'])
        self.assertEqual('lib1', payload['libraries'][0]['lib_id'])

        rows = self.client.get('/api/feishu-bom/database/rows?lib_id=lib1&query=RES')
        self.assertEqual(200, rows.status_code)
        row_payload = rows.get_json()
        self.assertTrue(row_payload['ok'])
        self.assertEqual(1, row_payload['total'])
        self.assertEqual('RES_0402', row_payload['rows'][0]['key_value'])

    def test_feishu_bom_database_delete_row_endpoint(self):
        self.make_fake_feishu_data()
        rows = self.client.get('/api/feishu-bom/database/rows?lib_id=lib1').get_json()
        row_id = rows['rows'][0]['id']

        response = self.client.delete(f'/api/feishu-bom/database/rows/{row_id}')
        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual(1, payload['deleted_rows'])
        self.assertEqual('RES_0402', payload['deleted_row']['key_value'])

        after_rows = self.client.get('/api/feishu-bom/database/rows?lib_id=lib1').get_json()
        self.assertEqual(0, after_rows['total'])

        missing = self.client.delete(f'/api/feishu-bom/database/rows/{row_id}')
        self.assertEqual(404, missing.status_code)
        self.assertFalse(missing.get_json()['ok'])

    def test_feishu_bom_database_create_and_update_row_endpoints(self):
        self.make_fake_feishu_data()

        create = self.client.post('/api/feishu-bom/database/rows', json={
            'lib_id': 'lib1',
            'lib_name': '优选库',
            'sheet_name': '手工维护',
            'key_value': 'CAP-100N',
            'hq_no': 'HQ17101005',
            'pi': 'LiXinYu',
            'selection_order': '1',
            'extra_fields': {'封装': '0402'},
        })
        self.assertEqual(201, create.status_code)
        created_payload = create.get_json()
        self.assertTrue(created_payload['ok'])
        self.assertEqual('CAP-100N', created_payload['row']['key_value'])

        update = self.client.patch(
            f"/api/feishu-bom/database/rows/{created_payload['row_id']}",
            json={
                'key_value': 'CAP-220N',
                'hq_no': 'HQ17101006',
                'selection_order': '2',
                'extra_fields': '{"封装":"0201"}',
            },
        )
        self.assertEqual(200, update.status_code)
        updated_payload = update.get_json()
        self.assertTrue(updated_payload['ok'])
        self.assertEqual('CAP-220N', updated_payload['row']['key_value'])
        self.assertEqual('2', updated_payload['row']['selection_order'])

    def test_feishu_bom_preview_sheet_endpoint_returns_rows(self):
        old_preview = webapp_factory.preview_feishu_sheet

        def fake_preview(**kwargs):
            return {
                'ok': True,
                'sheet_id': kwargs['sheet_id'],
                'header_row': kwargs['header_row'],
                'rows': [['厂家型号', 'HQ料号'], ['ABC-123', 'HQ001']],
                'headers': ['厂家型号', 'HQ料号'],
                'mapping_suggestion': {
                    'header_row': 1,
                    'mapping': {'spec_model_col': '厂家型号', 'hq_code_col': 'HQ料号'},
                },
            }

        webapp_factory.preview_feishu_sheet = fake_preview
        self.addCleanup(lambda: setattr(webapp_factory, 'preview_feishu_sheet', old_preview))

        response = self.client.post(
            '/api/feishu-bom/preview-sheet',
            json={
                'base_url': 'https://mcenter.example.local',
                'origin': 'cli_demo',
                'user_id': '100001',
                'spreadsheet_token_or_url': 'token123',
                'sheet_id': 'sh1',
                'header_row': 2,
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual('sh1', payload['sheet_id'])
        self.assertEqual(2, payload['header_row'])

    def test_feishu_bom_suggest_mapping_endpoint_uses_local_heuristic_by_default(self):
        response = self.client.post(
            '/api/feishu-bom/suggest-mapping',
            json={
                'sheet_title': 'Sheet1',
                'rows': [
                    ['备注', '', ''],
                    ['厂家型号', 'HQ料号', '制造商'],
                    ['ABC-123', 'HQ001', 'ACME'],
                ],
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertFalse(payload['agent']['used'])
        self.assertEqual('header-title-and-optional-field-suggestion', payload['agent']['role'])
        self.assertEqual(2, payload['suggestion']['header_row'])
        self.assertEqual('厂家型号', payload['suggestion']['mapping']['spec_model_col'])
        self.assertEqual('HQ料号', payload['suggestion']['mapping']['hq_code_col'])

    def test_feishu_bom_agent_suggest_mapping_uses_isolated_conversation_per_sheet(self):
        calls = []
        old_ask = webapp_factory.ask_aster_model

        def fake_ask(prompt, *, inputs=None, environ=None):
            calls.append({
                'prompt': prompt,
                'inputs': dict(inputs or {}),
                'environ': dict(environ or {}),
            })
            return {
                'ok': True,
                'mode': 'live',
                'provider': 'fake-aster',
                'answer': json.dumps({
                    'header_row': 2,
                    'headers': ['厂家型号', 'HQ料号', 'PI'],
                    'optional_titles': ['PI'],
                    'confidence': 'high',
                    'notes': ['仅基于当前 Sheet 判断。'],
                }, ensure_ascii=False),
                'metadata': {'conversation_id': 'fake-response-conv'},
            }

        webapp_factory.ask_aster_model = fake_ask
        self.addCleanup(lambda: setattr(webapp_factory, 'ask_aster_model', old_ask))
        self.with_env({
            'PSTX_ASTER_MODE': 'live',
            'PSTX_ASTER_BACKEND': 'chat-flow',
            'ASTER_API_KEY': 'fake-key',
            'ASTER_EMP_NO': '100001',
            'PSTX_ASTER_CONVERSATION_ID': 'shared-context-should-not-pass',
        })

        for title in ['Sheet-A', 'Sheet-B']:
            response = self.client.post(
                '/api/feishu-bom/suggest-mapping',
                json={
                    'sheet_title': title,
                    'use_agent': True,
                    'rows': [
                        ['说明', '', ''],
                        ['厂家型号', 'HQ料号', 'PI'],
                        [f'{title}-PN', 'HQ001', 'PI-A'],
                    ],
                },
            )
            self.assertEqual(200, response.status_code)
            payload = response.get_json()
            self.assertTrue(payload['agent']['used'])
            self.assertTrue(payload['agent']['metadata']['isolated_conversation'])

        self.assertEqual(2, len(calls))
        request_ids = [call['inputs']['sheet_agent_request_id'] for call in calls]
        self.assertNotEqual(request_ids[0], request_ids[1])
        for call in calls:
            self.assertEqual('', call['environ'].get('PSTX_ASTER_CONVERSATION_ID'))
            self.assertEqual('false', call['environ'].get('PSTX_ASTER_AUTO_GENERATE_NAME'))
            self.assertNotIn('local_suggestion', call['inputs'])
            self.assertIn('required_targets', call['inputs'])
            self.assertIn('HQ料号', call['inputs']['required_targets'])
            self.assertEqual('current_sheet_only', call['inputs']['isolation']['scope'])
            self.assertIn('必须仅基于本次 preview_rows 独立判断', call['prompt'])
            self.assertIn('不要硬猜', call['prompt'])
            self.assertIn('必须保留表格中的原始 title 文本', call['prompt'])

    def test_feishu_bom_database_delete_library_endpoint(self):
        self.make_fake_feishu_data()
        response = self.client.delete('/api/feishu-bom/database/libraries/lib1')
        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual(1, payload['deleted_rows'])

        overview = self.client.get('/api/feishu-bom/database').get_json()
        self.assertEqual(0, overview['cache_count'])

    def test_feishu_bom_preview_matches_report_bom_rows(self):
        self.make_fake_feishu_data()
        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'feishu-demo',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']
        preview = self.client.post(
            f'/api/report/{run_id}/feishu-bom/preview',
            json={'source': 'bom_normal_detail', 'key_field': '描述', 'limit': 10},
        )
        self.assertEqual(200, preview.status_code)
        payload = preview.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual(1, payload['matched_count'])
        self.assertEqual(1, payload['unmatched_count'])
        matched_rows = [row for row in payload['rows'] if row['匹配状态'] == '已匹配']
        self.assertEqual('R1', matched_rows[0]['位号'])
        self.assertEqual('HQ-R', matched_rows[0]['HQ料号'])
        total_preview = self.client.post(
            f'/api/report/{run_id}/feishu-bom/preview',
            json={'source': 'bom_total_detail', 'key_field': '描述', 'limit': 10},
        )
        self.assertEqual(200, total_preview.status_code)
        self.assertEqual('总 BOM 明细', total_preview.get_json()['source_label'])

    def test_report_adds_feishu_hq_code_review_without_changing_original_bom(self):
        self.make_fake_feishu_data()
        self.add_fake_feishu_material(
            key_value='IC_CPU_SPEC',
            hq_no='PN_U1',
            spec='IC_CPU_SPEC',
            pi='LiXinYu',
            selection_order='A1',
        )
        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'feishu-link-demo',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']
        report = self.client.get(f'/api/report/{run_id}').get_json()

        bom_section = next(section for section in report['sections'] if section['id'] == 'bom')
        original_bom = next(table for table in bom_section['tables'] if table['id'] == 'bom_normal_merged')
        self.assertFalse(any(table['id'] == 'feishu_hq_links' for table in bom_section['tables']))
        self.assertIn('飞书规格型号', original_bom['columns'])
        self.assertIn('PI', original_bom['columns'])
        self.assertIn('选型顺序', original_bom['columns'])
        matched = next(row for row in original_bom['rows'] if row['料号'] == 'PN_U1')
        self.assertEqual('PN_U1', matched['飞书HQ料号'])
        self.assertEqual('IC_CPU_SPEC', matched['飞书规格型号'])
        self.assertEqual('LiXinYu', matched['PI'])
        self.assertEqual('A1', matched['选型顺序'])
        self.assertIn('通过', matched['飞书校对结论'])
        skipped = next(row for row in original_bom['rows'] if row['料号'] == '')
        self.assertIn('Cadence 料号为空', skipped['飞书校对结论'])

        resistor_section = next(section for section in report['sections'] if section['id'] == 'resistor')
        chip_table = next(table for table in resistor_section['tables'] if table['id'] == 'chip_pin_rows')
        self.assertIn('芯片飞书规格型号', chip_table['columns'])
        self.assertIn('芯片PI', chip_table['columns'])
        chip_row = next(row for row in chip_table['rows'] if row['芯片位号'] == 'U1')
        self.assertEqual('IC_CPU_SPEC', chip_row['芯片飞书规格型号'])
        self.assertEqual('LiXinYu', chip_row['芯片PI'])
        self.assertEqual('A1', chip_row['芯片选型顺序'])

    def test_harness_status_endpoint_lists_readonly_tools(self):
        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        response = self.client.get('/api/harness/status')

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual('local-harness', payload['mode'])
        self.assertGreaterEqual(payload['tool_count'], 6)
        self.assertTrue(all(tool['readonly'] for tool in payload['tools']))
        self.assertEqual('mock', payload['model_provider']['mode'])

    def test_harness_tools_endpoint_returns_agent_tool_schema(self):
        response = self.client.get('/api/harness/tools')

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        tool_map = {tool['name']: tool for tool in payload['tools']}
        self.assertIn('get_table_rows', tool_map)
        self.assertIn('read_project_text', tool_map)
        self.assertIn('search_feishu_cache_rows', tool_map)
        self.assertIn('batch_search_feishu_cache_rows', tool_map)
        self.assertIn('get_feishu_cache_row', tool_map)
        self.assertIn('summarize_dfmea_readiness', tool_map)
        self.assertIn('search_component_identity_cards', tool_map)
        self.assertIn('batch_get_component_identity_cards', tool_map)
        self.assertIn('search_datasheets', tool_map)
        self.assertIn('get_datasheet_excerpt', tool_map)
        self.assertIn('list_datasheet_documents', tool_map)
        self.assertIn('list_datasheet_review_templates', tool_map)
        self.assertIn('get_datasheet_review_template', tool_map)
        self.assertIn('search_datasheet_chunks', tool_map)
        self.assertIn('search_datasheet_parameters', tool_map)
        self.assertIn('get_datasheet_parameter', tool_map)
        self.assertIn('batch_search_datasheet_chunks', tool_map)
        self.assertIn('get_datasheet_chunk', tool_map)
        self.assertIn('get_datasheet_page_excerpt', tool_map)
        self.assertIn('list_agent_ref_sources', tool_map)
        self.assertIn('search_agent_ref_pdfs', tool_map)
        self.assertIn('get_agent_ref_pdf_excerpt', tool_map)
        self.assertIn('list_review_checklist_sources', tool_map)
        self.assertIn('search_review_checklists', tool_map)
        self.assertIn('get_review_checklist_excerpt', tool_map)
        self.assertIn('summarize_chip_topology', tool_map)
        self.assertIn('list_business_dictionary', tool_map)
        self.assertIn('summarize_llm_topology_netlist', tool_map)
        self.assertIn('summarize_topology_review_tasks', tool_map)
        self.assertIn('query_chip_topology', tool_map)
        self.assertIn('query_llm_topology_netlist', tool_map)
        self.assertIn('batch_query_chip_topology', tool_map)
        self.assertIn('batch_query_llm_topology_netlist', tool_map)
        self.assertIn('get_chip_topology_edge', tool_map)
        self.assertIn('get_llm_topology_node', tool_map)
        self.assertIn('get_llm_topology_edge', tool_map)
        self.assertIn('get_topology_review_task', tool_map)
        self.assertIn('batch_expand_topology_review_tasks', tool_map)
        self.assertIn('list_document_search_sources', tool_map)
        self.assertIn('search_documents', tool_map)
        self.assertIn('batch_search_documents', tool_map)
        self.assertIn('get_document_excerpt', tool_map)
        self.assertIn('batch_query_report_entities', tool_map)
        self.assertIn('summarize_table_column_values', tool_map)
        self.assertIn('list_harness_skills', tool_map)
        self.assertIn('select_harness_skills', tool_map)
        self.assertIn('get_harness_skill', tool_map)
        self.assertTrue(tool_map['read_project_text']['file_access'])
        self.assertFalse(tool_map['read_project_text']['mutating'])
        self.assertEqual('read_project_file', tool_map['read_project_text']['approval_scope'])
        self.assertEqual('harness_skill', tool_map['get_harness_skill']['evidence_kind'])
        self.assertIn('evidence_kind', tool_map['get_table_rows'])
        self.assertIn('supports_parallel', tool_map['get_table_rows'])
        self.assertIn('table_id', tool_map['get_table_rows']['input_schema']['required'])
        self.assertIn('query', tool_map['search_feishu_cache_rows']['input_schema']['required'])
        self.assertIn('query', tool_map['search_component_identity_cards']['input_schema']['required'])
        self.assertIn('query', tool_map['search_datasheets']['input_schema']['required'])
        self.assertIn('query', tool_map['search_datasheet_chunks']['input_schema']['required'])
        self.assertIn('template_id', tool_map['get_datasheet_review_template']['input_schema']['required'])
        self.assertIn('parameter_id', tool_map['get_datasheet_parameter']['input_schema']['required'])
        self.assertIn('doc_id', tool_map['get_datasheet_chunk']['input_schema']['required'])
        self.assertIn('chunk_id', tool_map['get_datasheet_chunk']['input_schema']['required'])
        self.assertIn('query', tool_map['search_agent_ref_pdfs']['input_schema']['required'])
        self.assertIn('query', tool_map['search_review_checklists']['input_schema']['required'])
        self.assertIn('query', tool_map['query_chip_topology']['input_schema']['required'])
        self.assertIn('query', tool_map['query_llm_topology_netlist']['input_schema']['required'])
        self.assertIn('edge_id', tool_map['get_chip_topology_edge']['input_schema']['required'])
        self.assertIn('edge_id', tool_map['get_llm_topology_edge']['input_schema']['required'])
        self.assertIn('task_id', tool_map['get_topology_review_task']['input_schema']['required'])
        self.assertIn('task_ids', tool_map['batch_expand_topology_review_tasks']['input_schema']['required'])
        self.assertIn('refdes', tool_map['get_llm_topology_node']['input_schema']['required'])
        self.assertIn('query', tool_map['search_documents']['input_schema']['required'])
        self.assertIn('doc_id', tool_map['get_document_excerpt']['input_schema']['required'])

    def test_harness_profiles_endpoint_returns_profiles(self):
        response = self.client.get('/api/harness/profiles')

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        profiles = {item['id']: item for item in payload['profiles']}
        self.assertEqual('quick_scan', payload['default_profile'])
        self.assertIn('auto', profiles)
        self.assertIn('bom_depop', profiles)
        self.assertIn('feishu_bom_qa', profiles)
        self.assertIn('datasheet_qa', profiles)
        self.assertIn('dfmea_prep', profiles)
        self.assertIn('chip_topology', profiles)
        self.assertIn('document_search', profiles)
        self.assertIn('agent_ref_qa', profiles)
        self.assertIn('review_checklist_qa', profiles)
        self.assertIn('full_review', profiles)
        self.assertIn('get_evidence_pack', profiles['bom_depop']['tools'])
        self.assertIn('summarize_table_column_values', profiles['page_mapping']['tools'])
        self.assertIn('get_harness_skill', profiles['page_mapping']['tools'])
        self.assertIn('search_feishu_cache_rows', profiles['feishu_bom_qa']['tools'])
        self.assertIn('batch_search_feishu_cache_rows', profiles['feishu_bom_qa']['tools'])
        self.assertIn('search_datasheet_chunks', profiles['datasheet_qa']['tools'])
        self.assertIn('select_harness_skills', profiles['datasheet_qa']['tools'])
        self.assertIn('list_datasheet_review_templates', profiles['datasheet_qa']['tools'])
        self.assertIn('search_datasheet_parameters', profiles['datasheet_qa']['tools'])
        self.assertIn('get_datasheet_chunk', profiles['datasheet_qa']['tools'])
        self.assertIn('get_datasheet_parameter', profiles['datasheet_qa']['tools'])
        self.assertIn('batch_search_datasheet_chunks', profiles['datasheet_qa']['tools'])
        self.assertIn('summarize_dfmea_readiness', profiles['dfmea_prep']['tools'])
        self.assertIn('batch_get_component_identity_cards', profiles['dfmea_prep']['tools'])
        self.assertIn('search_datasheet_chunks', profiles['dfmea_prep']['tools'])
        self.assertIn('get_datasheet_review_template', profiles['dfmea_prep']['tools'])
        self.assertIn('search_datasheet_parameters', profiles['dfmea_prep']['tools'])
        self.assertIn('get_datasheet_chunk', profiles['dfmea_prep']['tools'])
        self.assertIn('search_datasheets', profiles['dfmea_prep']['tools'])
        self.assertIn('batch_match_component_datasheets', profiles['dfmea_prep']['tools'])
        self.assertIn('summarize_dfmea_datasheet_coverage', profiles['dfmea_prep']['tools'])
        self.assertIn('list_business_dictionary', profiles['chip_topology']['tools'])
        self.assertIn('summarize_llm_topology_netlist', profiles['chip_topology']['tools'])
        self.assertIn('summarize_topology_review_tasks', profiles['chip_topology']['tools'])
        self.assertIn('batch_query_llm_topology_netlist', profiles['chip_topology']['tools'])
        self.assertIn('get_llm_topology_node', profiles['chip_topology']['tools'])
        self.assertIn('get_llm_topology_edge', profiles['chip_topology']['tools'])
        self.assertIn('get_topology_review_task', profiles['chip_topology']['tools'])
        self.assertIn('batch_expand_topology_review_tasks', profiles['chip_topology']['tools'])
        self.assertNotIn('summarize_chip_topology', profiles['chip_topology']['tools'])
        self.assertNotIn('batch_query_chip_topology', profiles['chip_topology']['tools'])
        self.assertNotIn('get_chip_topology_edge', profiles['chip_topology']['tools'])
        self.assertIn('search_documents', profiles['document_search']['tools'])
        self.assertIn('get_document_excerpt', profiles['document_search']['tools'])
        self.assertIn('search_agent_ref_pdfs', profiles['agent_ref_qa']['tools'])
        self.assertIn('get_agent_ref_pdf_excerpt', profiles['agent_ref_qa']['tools'])
        self.assertIn('search_review_checklists', profiles['review_checklist_qa']['tools'])
        self.assertIn('get_review_checklist_excerpt', profiles['review_checklist_qa']['tools'])

    def test_datasheet_status_and_reindex_api(self):
        root = Path(tempfile.mkdtemp())
        self.temp_roots.append(root)
        source = root / 'datasheets'
        source.mkdir()
        (source / 'HQ100_GPU_CORE_TEST_IC.pdf').write_bytes(b'%PDF fake')
        self.with_env({
            'PSTX_DATASHEET_DIR': str(source),
            'PSTX_DATASHEET_DATA_DIR': str(root / 'datasheet_data'),
            'PSTX_PDF_EXTRACTOR': 'auto',
            'PSTX_MINERU_BIN': None,
            'PSTX_MINERU_DEVICE': 'auto',
            'PSTX_MINERU_METHOD': 'auto',
            'PSTX_MINERU_MODEL_SOURCE': 'auto',
        })

        status = self.client.get('/api/datasheets/status')
        self.assertEqual(200, status.status_code)
        status_payload = status.get_json()
        self.assertTrue(status_payload['ok'])
        self.assertTrue(status_payload['configured'])
        self.assertEqual('auto', status_payload['extractor']['mode'])
        self.assertIn('mineru', status_payload['extractor'])
        self.assertEqual('auto', status_payload['extractor']['mineru']['device'])
        self.assertEqual('auto', status_payload['extractor']['mineru']['method'])
        self.assertEqual('auto', status_payload['extractor']['mineru']['model_source'])

        with mock.patch(
            'pstx_knowledge.datasheets._extract_pdf_pages',
            return_value=('indexed', ['HQ100 GPU_CORE_TEST_IC datasheet limits'], 'fake', ''),
        ):
            reindex = self.client.post('/api/datasheets/reindex', json={'force': True, 'max_files': 10})
        self.assertEqual(200, reindex.status_code)
        reindex_payload = reindex.get_json()
        self.assertTrue(reindex_payload['ok'])
        self.assertEqual(1, reindex_payload['indexed_count'])
        status_after = self.client.get('/api/datasheets/status').get_json()
        self.assertEqual(1, status_after['chunk_count'])

        bad = self.client.post('/api/datasheets/reindex', json={'max_files': 'bad'})
        self.assertEqual(400, bad.status_code)

    def test_agent_lab_status_reindex_and_ask_api(self):
        root = Path(tempfile.mkdtemp())
        self.temp_roots.append(root)
        ref_dir = root / 'ref'
        ref_dir.mkdir()
        checklist_dir = root / 'ref_checklist'
        checklist_dir.mkdir()
        (ref_dir / 'agent_lab_manual.pdf').write_bytes(b'%PDF fake')
        (checklist_dir / 'review_cases.md').write_text('真实 review 问题：U46 多 symbol HQ_CODE 检查。', encoding='utf-8')
        self.with_env({
            'PSTX_AGENT_REF_DIR': str(ref_dir),
            'PSTX_AGENT_REF_DATA_DIR': str(root / 'agent_ref_data'),
            'PSTX_AGENT_CHECKLIST_REF_DIR': str(checklist_dir),
            'PSTX_AGENT_CHECKLIST_DATA_DIR': str(root / 'agent_checklist_data'),
            'PSTX_ASTER_MODE': 'mock',
        })

        status = self.client.get('/api/agent-lab/status')
        self.assertEqual(200, status.status_code)
        status_payload = status.get_json()
        self.assertTrue(status_payload['ok'])
        self.assertEqual('review_checklist_qa', status_payload['default_profile'])
        self.assertEqual(1, status_payload['ref']['pdf_count'])
        self.assertEqual(1, status_payload['checklist']['file_count'])

        with mock.patch(
            'pstx_knowledge.reference_library._extract_pdf_pages',
            return_value=('indexed', ['Agent Lab PDF evidence boundary test'], 'fake', ''),
        ):
            reindex = self.client.post('/api/agent-lab/ref/reindex', json={'force': True, 'max_files': 10})
        self.assertEqual(200, reindex.status_code)
        self.assertEqual(1, reindex.get_json()['indexed_count'])
        checklist_reindex = self.client.post('/api/agent-lab/checklist/reindex', json={'force': True, 'max_files': 10})
        self.assertEqual(200, checklist_reindex.status_code)
        self.assertEqual(1, checklist_reindex.get_json()['indexed_count'])

        ask = self.client.post('/api/agent-lab/ask', json={
            'profile': 'agent_ref_qa',
            'question': '请检索 Agent Lab PDF',
            'max_steps': 2,
            'max_tool_calls': 2,
        })
        self.assertEqual(200, ask.status_code)
        ask_payload = ask.get_json()
        self.assertIn('agent_run_id', ask_payload)
        self.assertIn('lab', ask_payload)

    def test_harness_review_endpoint_reviews_existing_report(self):
        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'harness-demo',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']

        response = self.client.post(
            f'/api/report/{run_id}/harness/review',
            json={'task': 'full_review', 'question': '请重点看 DRC', 'max_rows_per_table': 2, 'include_model': False},
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual('local-harness', payload['mode'])
        self.assertEqual('full_review', payload['task'])
        self.assertEqual('请重点看 DRC', payload['question'])
        self.assertTrue(any(pack['id'] == 'drc' for pack in payload['evidence_packs']))
        self.assertFalse(payload['model_metadata']['included'])

    def test_harness_review_endpoint_rejects_invalid_request_and_missing_run(self):
        invalid = self.client.post(
            '/api/report/notfound/harness/review',
            json={'task': 'full_review'},
        )
        self.assertEqual(404, invalid.status_code)
        self.assertIn('未找到报告', invalid.get_json()['error'])

        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'harness-invalid',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        run_id = analyze.get_json()['run_id']
        bad_limit = self.client.post(
            f'/api/report/{run_id}/harness/review',
            json={'max_rows_per_table': 0},
        )
        self.assertEqual(400, bad_limit.status_code)
        self.assertIn('max_rows_per_table', bad_limit.get_json()['error'])

    def test_harness_agent_endpoint_runs_mock_agent_loop(self):
        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'agent-demo',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']

        response = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={
                'question': '请先看 DRC',
                'max_steps': 12,
                'max_tool_calls': 24,
                'enable_subagents': True,
                'subagent_profiles': ['bom_depop', 'derating'],
                'max_subagents': 2,
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual('local-agent-harness', payload['mode'])
        self.assertEqual('quick_scan', payload['profile'])
        self.assertTrue(payload['agent_run_id'])
        self.assertTrue(payload['trace_summary'])
        self.assertTrue(payload['citations'])
        self.assertEqual('final_answer', payload['model_metadata']['stopped_reason'])
        self.assertEqual('get_table_rows', payload['tool_calls'][0]['tool'])
        self.assertTrue(payload['observations'])
        self.assertEqual(2, payload['subagent_summary']['planned_count'])
        self.assertEqual(2, len(payload['subagents']))
        self.assertEqual(2, payload['trace_summary']['subagent_count'])
        self.assertTrue(payload['request']['enable_subagents'])

        replay = self.client.get(f"/api/harness/agent-runs/{payload['agent_run_id']}")
        self.assertEqual(200, replay.status_code)
        replay_payload = replay.get_json()
        self.assertTrue(replay_payload['ok'])
        self.assertEqual(payload['agent_run_id'], replay_payload['agent_run']['agent_run_id'])
        self.assertEqual(payload['agent_run_id'], replay_payload['trace']['agent_run_id'])
        self.assertEqual(payload['profile'], replay_payload['trace']['profile'])
        self.assertGreaterEqual(replay_payload['trace']['tool_call_count'], 1)
        self.assertTrue(replay_payload['trace']['execution_journal'])
        self.assertEqual('agent-run-journal/v1', replay_payload['trace']['journal_summary']['version'])
        self.assertEqual('agent-continuation-pack/v1', replay_payload['trace']['continuation_pack']['version'])
        context_payload = self.client.get(f"/api/report/{run_id}/harness/context").get_json()
        self.assertEqual('agent-continuation-pack/v1', context_payload['context']['latest_continuation_pack']['version'])

    def test_harness_agent_async_run_persists_status_and_artifacts(self):
        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'agent-async-demo',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']

        response = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={
                'question': '请后台快速审查 DRC',
                'max_steps': 3,
                'max_tool_calls': 4,
                'async': True,
            },
        )
        self.assertEqual(202, response.status_code)
        queued = response.get_json()
        self.assertTrue(queued['async'])
        self.assertEqual('queued', queued['status'])
        self.assertTrue(queued['agent_run_id'].startswith('report_'))

        agent_run_id = queued['agent_run_id']
        status_payload = {}
        for _ in range(80):
            status = self.client.get(f'/api/harness/agent-runs/{agent_run_id}')
            self.assertEqual(200, status.status_code)
            status_payload = status.get_json()
            if status_payload['status'] in {'completed', 'waiting_for_user', 'failed'}:
                break
            time.sleep(0.05)

        self.assertEqual('completed', status_payload['status'])
        self.assertTrue(status_payload['result_available'])
        self.assertEqual(agent_run_id, status_payload['agent_run']['agent_run_id'])
        self.assertIn('workspace', status_payload)
        self.assertIn('current_phase', status_payload)
        self.assertIn('progress', status_payload)
        self.assertIn('partial_trace', status_payload)
        self.assertFalse(status_payload['can_cancel'])

        artifacts = self.client.get(f'/api/harness/agent-runs/{agent_run_id}/artifacts')
        self.assertEqual(200, artifacts.status_code)
        artifact_payload = artifacts.get_json()
        self.assertTrue(artifact_payload['ok'])
        artifact_names = {item['name'] for item in artifact_payload['artifacts']}
        self.assertIn('result.json', artifact_names)
        self.assertIn('answer.md', artifact_names)
        self.assertIn('trace.json', artifact_names)
        self.assertIn('evidence_cards.json', artifact_names)

    def test_harness_agent_async_dispatch_creates_child_runs(self):
        class DispatchProvider:
            provider = 'dispatch-mock'
            mode = 'mock'

            def generate_agent_step(self, prompt, *, inputs=None):
                return HarnessModelResponse(
                    answer=json.dumps({
                        'dispatch_tasks': [{
                            'task_id': 'ds-u1',
                            'title': 'U1 datasheet',
                            'profile': 'datasheet_qa',
                            'question': '读取 U1 datasheet 关键参数。',
                        }],
                        'reason': 'datasheet 分支适合后台执行。',
                    }, ensure_ascii=False),
                    provider=self.provider,
                    mode=self.mode,
                )

        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        webapp_state.clear_web_session_state()
        with mock.patch.object(webapp_factory, 'MockHarnessModelProvider', DispatchProvider):
            app = webapp_factory.create_app()
            app.testing = True
            client = app.test_client()
            root = self.make_root()
            analyze = client.post(
                '/api/analyze',
                data={
                    'project_root': str(root),
                    'project_name': 'agent-dispatch-demo',
                    'ratio_limit': '70',
                    'custom_volt_map': '',
                },
            )
            self.assertEqual(200, analyze.status_code)
            run_id = analyze.get_json()['run_id']

            response = client.post(
                f'/api/report/{run_id}/harness/agent',
                json={
                    'question': '请后台拆分规格书复核',
                    'max_steps': 1,
                    'max_tool_calls': 1,
                    'async': True,
                },
            )
            self.assertEqual(202, response.status_code)
            parent_run_id = response.get_json()['agent_run_id']
            status_payload = {}
            for _ in range(80):
                status = client.get(f'/api/harness/agent-runs/{parent_run_id}')
                self.assertEqual(200, status.status_code)
                status_payload = status.get_json()
                if status_payload['status'] in {'completed', 'waiting_for_user', 'failed'}:
                    break
                time.sleep(0.05)

            self.assertEqual('completed', status_payload['status'])
            self.assertEqual(['ds-u1'], [item['task_id'] for item in status_payload['dispatch_tasks']])
            self.assertEqual(1, len(status_payload['child_agent_run_ids']))
            child_run_id = status_payload['child_agent_run_ids'][0]
            child_status = client.get(f'/api/harness/agent-runs/{child_run_id}').get_json()
            self.assertTrue(child_status['ok'])
            self.assertEqual(parent_run_id, child_status['parent_agent_run_id'])
            self.assertEqual('ds-u1', child_status['dispatch_task']['task_id'])

    def test_harness_agent_endpoint_rejects_bad_request_and_path_escape(self):
        missing = self.client.post('/api/report/notfound/harness/agent', json={'question': 'x'})
        self.assertEqual(404, missing.status_code)
        self.assertIn('未找到报告', missing.get_json()['error'])

        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'agent-invalid',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        run_id = analyze.get_json()['run_id']
        bad_limit = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={'max_steps': 0},
        )
        self.assertEqual(400, bad_limit.status_code)
        self.assertIn('max_steps', bad_limit.get_json()['error'])

        bad_profile = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={'profile': 'unsafe'},
        )
        self.assertEqual(400, bad_profile.status_code)
        self.assertIn('未知 agent profile', bad_profile.get_json()['error'])

        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        escape = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={'profile': 'full_review', 'question': '请测试越权读取', 'debug': True},
        )
        self.assertEqual(400, escape.status_code)
        payload = escape.get_json()
        self.assertFalse(payload['ok'])
        self.assertIn('项目根目录之外', payload['answer'])

        replay_missing = self.client.get('/api/harness/agent-runs/notfound')
        self.assertEqual(404, replay_missing.status_code)
        self.assertIn('未找到 agent_run_id', replay_missing.get_json()['error'])

    def test_harness_agent_project_context_clarification_loop(self):
        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'dfmea-context-demo',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']

        first = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={'profile': 'dfmea_prep', 'max_steps': 2, 'max_tool_calls': 2},
        )
        self.assertEqual(200, first.status_code)
        first_payload = first.get_json()
        self.assertEqual('waiting_for_user', first_payload['status'])
        self.assertTrue(first_payload['needs_user_input']['questions'])

        context_response = self.client.get(f'/api/report/{run_id}/harness/context')
        self.assertEqual(200, context_response.status_code)
        context_payload = context_response.get_json()['context']
        self.assertEqual(0, context_payload['answer_count'])
        self.assertTrue(context_payload['pending_questions'])
        self.assertEqual('agent-project-session-memory/v1', context_payload['session_memory_summary']['version'])
        self.assertIn(first_payload['agent_run_id'], context_payload['session_memory_summary']['source_agent_run_ids'])
        self.assertIn('evidence_memory_cards', context_payload)
        first_evidence_ids = {item.get('id') for item in first_payload.get('final_evidence') or []}
        memory_evidence_ids = {item.get('id') for item in context_payload.get('evidence_memory_cards') or []}
        self.assertTrue(first_evidence_ids.intersection(memory_evidence_ids))

        question = first_payload['needs_user_input']['questions'][0]
        second = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={
                'profile': 'dfmea_prep',
                'max_steps': 2,
                'max_tool_calls': 2,
                'continue_agent_run_id': first_payload['agent_run_id'],
                'context_answers': [{
                    'question_id': question['question_id'],
                    'answer': 'U1/PU 类器件规格由硬件 owner 人工待查，当前先按准备度报告处理。',
                    'applies_to': question.get('applies_to', {}),
                }],
            },
        )
        self.assertEqual(200, second.status_code)
        second_payload = second.get_json()
        self.assertIn(second_payload['status'], {'completed', 'limited'})
        self.assertEqual(1, second_payload['project_context_summary']['answer_count'])
        self.assertFalse(second_payload['project_context_summary']['pending_questions'])
        active_pack = second_payload['project_context_summary']['active_continuation_pack']
        self.assertEqual(first_payload['agent_run_id'], active_pack['agent_run_id'])
        self.assertEqual('agent-continuation-pack/v1', active_pack['version'])
        self.assertEqual(
            'agent-project-session-memory/v1',
            second_payload['project_context_summary']['session_memory_summary']['version'],
        )

        bad_continue = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={'profile': 'dfmea_prep', 'continue_agent_run_id': 'missing-run'},
        )
        self.assertEqual(400, bad_continue.status_code)
        self.assertIn('continue_agent_run_id', bad_continue.get_json()['error'])

        bad_answer = self.client.post(
            f'/api/report/{run_id}/harness/agent',
            json={'profile': 'dfmea_prep', 'context_answers': [{'question_id': 'q1', 'answer': ''}]},
        )
        self.assertEqual(400, bad_answer.status_code)
        self.assertIn('answer', bad_answer.get_json()['error'])

        cleared = self.client.post(f'/api/report/{run_id}/harness/context/clear')
        self.assertEqual(200, cleared.status_code)
        self.assertEqual(0, cleared.get_json()['context']['answer_count'])

    def test_agent_eval_status_and_run_api(self):
        status = self.client.get('/api/agent-eval/status')
        self.assertEqual(200, status.status_code)
        status_payload = status.get_json()
        self.assertTrue(status_payload['ok'])
        self.assertGreaterEqual(status_payload['case_count'], 6)
        self.assertIn('mock_quick_scan', [item['case_id'] for item in status_payload['cases']])

        run_all = self.client.post('/api/agent-eval/run', json={})
        self.assertEqual(200, run_all.status_code)
        run_payload = run_all.get_json()
        self.assertTrue(run_payload['ok'])
        self.assertEqual(100, run_payload['score'])
        self.assertEqual(0, run_payload['failed_count'])

        run_one = self.client.post('/api/agent-eval/run', json={'case_ids': ['invalid_citation_flagged']})
        self.assertEqual(200, run_one.status_code)
        one_payload = run_one.get_json()
        self.assertEqual(1, one_payload['case_count'])
        self.assertGreaterEqual(one_payload['cases'][0]['metrics']['invalid_citation_count'], 1)

        missing = self.client.post('/api/agent-eval/run', json={'case_ids': ['missing_case']})
        self.assertEqual(400, missing.status_code)
        self.assertIn('未知 eval case', missing.get_json()['error'])

    def test_read_local_text_file_decodes_gb18030_without_replacement(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / 'pstxprt.dat'
            path.write_bytes(
                "PART_NAME\nC1 '电容'\nVALUE='10微法'\n".encode('gb18030')
            )

            text, meta = webapp_project_io.read_local_text_file(path, 'pstxprt.dat', True)

        self.assertIn('电容', text)
        self.assertIn('10微法', text)
        self.assertEqual('gb18030', meta['encoding'])

    def test_webapp_project_io_discovers_packaged_project_files(self):
        root = build_project_root()
        self.temp_roots.append(root)

        project_root, prt_path, net_path, ref_path = webapp_project_io.discover_project_files(str(root / 'packaged'))

        self.assertEqual(root, project_root)
        self.assertEqual(root / 'packaged' / 'pstxprt.dat', prt_path)
        self.assertEqual(root / 'packaged' / 'pstxnet.dat', net_path)
        self.assertEqual(root / 'packaged' / 'pstxref.dat', ref_path)

    def test_resolve_port_falls_back_when_preferred_port_is_busy(self):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.bind((webapp_server.DEFAULT_HOST, 0))
            sock.listen(1)
            busy_port = sock.getsockname()[1]
            resolved = webapp_server.resolve_port(busy_port, max_attempts=3)
        self.assertNotEqual(busy_port, resolved)
        self.assertGreaterEqual(resolved, busy_port + 1)

    def test_default_port_uses_reserved_localhost_port(self):
        self.assertEqual(44441, webapp_server.DEFAULT_PORT)

    def test_home_page_uses_product_title_and_runtime_port(self):
        response = self.client.get('/', headers={'Host': '127.0.0.1:8766'})
        self.assertEqual(200, response.status_code)
        text = response.get_data(as_text=True)
        self.assertIn('PSTX 原理图审查平台', text)
        self.assertIn('8766', text)
        self.assertIn('项目根路径', text)
        self.assertIn('使用说明', text)
        self.assertIn('href="/guide"', text)
        self.assertIn('报告定位统一使用“页码”', text)
        self.assertIn('class="help-tip"', text)
        self.assertIn('DEPOP 参与排查', text)
        self.assertIn('总 BOM 视图', text)
        self.assertIn('id="project-manager"', text)
        self.assertIn('历史分析项目', text)
        self.assertIn('project-manager-body', text)
        self.assertNotIn('class="home-side"', text)
        self.assertNotIn('自定义电压映射', text)
        self.assertNotIn('name="custom_volt_map"', text)
        self.assertNotIn('综合动效', text)
        self.assertNotIn('单项目打开', text)

    def test_guide_page_explains_usage_and_rule_logic(self):
        response = self.client.get('/guide')
        self.assertEqual(200, response.status_code)
        text = response.get_data(as_text=True)
        self.assertIn('使用说明与判定逻辑', text)
        self.assertIn('快速使用流程', text)
        self.assertIn('全模块判定细则', text)
        self.assertIn('上下拉判定', text)
        self.assertIn('串阻判定逻辑', text)
        self.assertIn('隔串阻上下拉', text)
        self.assertIn('分压风险', text)
        self.assertNotIn('OD/OC 候选', text)
        self.assertIn('芯片 Pin 电阻状态表', text)
        self.assertIn('页码 = 用户看到的总体页', text)
        self.assertIn('原理图总页数', text)
        self.assertIn('module_order.dat', text)
        self.assertIn('未命名网络页码', text)
        self.assertIn('Split symbol 大芯片', text)
        self.assertIn('电容降额判定', text)
        self.assertIn('BOM_OPTION 打圈覆盖', text)
        self.assertIn('飞书在线同步与飞书 HQ 校对', text)
        self.assertIn('页级 Cadence 语义比对', text)
        self.assertIn('DFMEA 准备度', text)
        self.assertIn('ref_checklist', text)
        self.assertIn('证据压缩不替代原始证据', text)
        self.assertIn('Agent 与 Harness 如何工作', text)
        self.assertIn('class="help-tip"', text)

    def test_diagnostics_status_tail_and_export_api(self):
        self.client.get('/')
        status = self.client.get('/api/diagnostics/status')
        self.assertEqual(200, status.status_code)
        status_payload = status.get_json()
        self.assertTrue(status_payload['ok'])
        self.assertTrue(status_payload['log_file']['path'].endswith('pstx_diagnostics.log'))
        self.assertTrue(status_payload['feishu_log_file']['path'].endswith('feishu_bom_debug.log'))

        tail = self.client.get('/api/diagnostics/tail?limit=20')
        self.assertEqual(200, tail.status_code)
        tail_payload = tail.get_json()
        self.assertTrue(tail_payload['ok'])
        tail_text = json.dumps(tail_payload, ensure_ascii=False)
        self.assertIn('web.request.start', tail_text)
        self.assertIn('web.request.finish', tail_text)

        bad_tail = self.client.get('/api/diagnostics/tail?limit=bad')
        self.assertEqual(400, bad_tail.status_code)

        export = self.client.get('/api/diagnostics/export')
        self.assertEqual(200, export.status_code)
        with tempfile.TemporaryDirectory() as temp_dir:
            zip_path = Path(temp_dir) / 'diagnostics.zip'
            zip_path.write_bytes(export.get_data())
            with zipfile.ZipFile(zip_path) as archive:
                self.assertIn('diagnostics_status.json', archive.namelist())
                self.assertIn('logs/pstx_diagnostics.log', archive.namelist())

    def test_debug_report_page_serves_fixture_preview(self):
        response = self.client.get('/debug/report')
        self.assertEqual(200, response.status_code)
        text = response.get_data(as_text=True)
        self.assertIn('Debug UI 假项目', text)
        self.assertIn('data-debug-ui="true"', text)
        self.assertIn('data-debug-fixture="true"', text)
        self.assertIn('window.PSTX_DEBUG_REPORT', text)
        self.assertIn('Debug fixture', text)
        self.assertIn('report-layer-guide', text)
        self.assertIn('pages/report.js', text)

        app_js = self.client.get('/static/app.js')
        try:
            self.assertEqual(200, app_js.status_code)
            app_js_text = app_js.get_data(as_text=True)
            self.assertIn('REPORT_TABLE_LEVELS', app_js_text)
            self.assertIn('report-level-group', app_js_text)
            self.assertIn('table-level-badge', app_js_text)
            self.assertIn('table-trust-badge', app_js_text)
            self.assertIn('quiet-table-group', app_js_text)
            self.assertIn('section-scan-meta', app_js_text)
            self.assertIn('report-decision-strip', app_js_text)
            self.assertIn('PSTX_DEBUG_REPORT', app_js_text)
            self.assertIn('table-empty-state', app_js_text)
        finally:
            app_js.close()

    def test_debug_dfmea_page_serves_fixture_style_preview(self):
        response = self.client.get('/debug/dfmea')
        self.assertEqual(200, response.status_code)
        text = response.get_data(as_text=True)
        self.assertIn('DFMEA 工作台', text)
        self.assertIn('data-debug-ui="true"', text)
        self.assertIn('data-debug-fixture="true"', text)
        self.assertIn('debug-dfmea', text)
        self.assertIn('pages/dfmea.js', text)
        self.assertIn('name="refdes_text" class="dfmea-refdes-display"', text)
        self.assertNotIn('name="refdes_text" readonly', text)

        app_js = self.client.get('/static/app.js')
        try:
            self.assertEqual(200, app_js.status_code)
            app_js_text = app_js.get_data(as_text=True)
            self.assertIn('bootUiDebugMode', app_js_text)
            self.assertIn('ui-debug-panel', app_js_text)
            self.assertIn('scheduleAutoRenderMoreRows', app_js_text)
            self.assertNotIn('table-render-more', app_js_text)
            self.assertNotIn('继续渲染更多', app_js_text)
        finally:
            app_js.close()

        dfmea_js = self.client.get('/static/pages/dfmea.js')
        try:
            self.assertEqual(200, dfmea_js.status_code)
            dfmea_js_text = dfmea_js.get_data(as_text=True)
            self.assertIn('DEBUG_PENDING', dfmea_js_text)
            self.assertIn('Debug fixture', dfmea_js_text)
            self.assertIn('dfmea-pending-table', dfmea_js_text)
            self.assertIn('RENDER_BATCH_SIZE', dfmea_js_text)
            self.assertIn('ensurePendingTableShell', dfmea_js_text)
            self.assertIn('onPendingHostClick', dfmea_js_text)
            self.assertIn('rowMatchesGlobalQuery', dfmea_js_text)
            self.assertIn("params.delete('q')", dfmea_js_text)
            self.assertIn('dfmea-load-more', dfmea_js_text)
            self.assertIn('dfmea-empty-row', dfmea_js_text)
            self.assertIn('is-collapsed', dfmea_js_text)
            self.assertIn('function setRefdesDisplay', dfmea_js_text)
            self.assertIn('return Array.from(state.selected);', dfmea_js_text)
            self.assertIn("card.querySelectorAll('[data-action=\"edit\"]')", dfmea_js_text)
            self.assertIn("card.querySelectorAll('[data-action=\"delete\"]')", dfmea_js_text)
            self.assertNotIn('parseRefdesText', dfmea_js_text)
        finally:
            dfmea_js.close()

        app_css = self.client.get('/static/app.css')
        try:
            self.assertEqual(200, app_css.status_code)
            app_css_text = app_css.get_data(as_text=True)
            self.assertIn('.ui-debug-mode', app_css_text)
            self.assertIn('.ui-debug-panel', app_css_text)
            self.assertIn('.dfmea-refdes-display', app_css_text)
            self.assertIn('.dfmea-table-more', app_css_text)
            self.assertIn('body[data-page="dfmea"]::before', app_css_text)
            self.assertIn('contain: layout paint style', app_css_text)
            self.assertNotIn('.table-render-more', app_css_text)
        finally:
            app_css.close()

    def test_report_review_plan_layers_tables_by_review_intent(self):
        sections = [
            {
                'id': 'drc',
                'title': '设计检查',
                'lead': '设计检查项',
                'tables': [
                    build_report_table('missing_hq_code', '缺少料号', [{'位号': 'U1', '页码': 'PAGE1'}, {'位号': 'U2', '页码': 'PAGE2'}]),
                    build_report_table('tbd_attrs', 'TBD 待确认属性', [{'位号': 'R2', '网络': 'I2C_SCL'}]),
                ],
            },
            {
                'id': 'network',
                'title': '网络分析',
                'lead': '网络信息',
                'tables': [
                    build_report_table('page_rows', '页码元件分布', [{'页码': 'PAGE1', '元件数': 2}]),
                    build_report_table('page_mapping_rows', '主模块页/页码映射检查', [{'主模块页': 'PAGE10', '页码': 'PAGE20'}]),
                ],
            },
        ]

        plan = build_review_plan(sections)
        policy = build_table_display_policy(sections)

        self.assertEqual(['missing_hq_code'], [item['table_id'] for item in plan['focus_items']])
        self.assertEqual(['page_rows'], [item['table_id'] for item in plan['info_items']])
        self.assertEqual(['page_mapping_rows'], [item['table_id'] for item in plan['debug_items']])
        self.assertEqual('tbd_attrs', plan['review_groups'][0]['items'][0]['table_id'])
        self.assertEqual('明确异常', plan['focus_items'][0]['trust_label'])
        self.assertEqual('待确认项', plan['review_groups'][0]['items'][0]['trust_label'])
        self.assertEqual('证据明细', plan['debug_items'][0]['trust_label'])
        self.assertIn('trust_counts', plan['summary'])
        self.assertEqual(2, plan['summary']['trust_counts']['明确异常'])
        self.assertEqual(['U1', 'U2'], plan['focus_items'][0]['related_refdes'])
        self.assertEqual(['PAGE1', 'PAGE2'], plan['focus_items'][0]['related_pages'])
        self.assertIn('page_mapping_rows', plan['hidden_table_ids'])
        self.assertEqual('info', next(item for item in policy if item['table_id'] == 'page_rows')['level'])
        self.assertEqual('信息统计', next(item for item in policy if item['table_id'] == 'page_rows')['trust_label'])
        self.assertTrue(next(item for item in policy if item['table_id'] == 'page_mapping_rows')['default_collapsed'])

    def test_compare_net_view_aggregates_before_display_truncation(self):
        def component(refdes: str, net: str) -> dict:
            return {
                'CDS_PART_NAME': 'IC',
                'value': 'IC',
                'package': 'QFN',
                'nets': {'1': net},
            }

        left_components = {
            'U1': component('U1', 'NET_A'),
            'U2': component('U2', 'NET_B'),
            'U3': component('U3', 'NET_C'),
        }
        right_components = {
            'U1': component('U1', 'NET_A2'),
            'U2': component('U2', 'NET_B2'),
            'U3': component('U3', 'NET_C2'),
        }
        left_nets = {
            name: [{'refdes': refdes, 'pin': '1', 'pin_name': 'IO'}]
            for refdes, name in [('U1', 'NET_A'), ('U2', 'NET_B'), ('U3', 'NET_C')]
        }
        right_nets = {
            name: [{'refdes': refdes, 'pin': '1', 'pin_name': 'IO'}]
            for refdes, name in [('U1', 'NET_A2'), ('U2', 'NET_B2'), ('U3', 'NET_C2')]
        }
        payloads = {
            'left': {'bundle': {'components': left_components, 'nets': left_nets}, 'report': {}},
            'right': {'bundle': {'components': right_components, 'nets': right_nets}, 'report': {}},
        }

        compare_payload = build_compare_payload(
            'left',
            'right',
            get_run_payload=lambda run_id: payloads[run_id],
            detail_limit=1,
        )

        self.assertEqual(1, len(compare_payload['key_pin_net_diff']['rows']))
        self.assertEqual(1, len(compare_payload['net_view_diff']['rows']))
        self.assertEqual(3, compare_payload['net_view_diff']['total_rows'])
        self.assertEqual(3, compare_payload['diff_totals']['net_view'])
        self.assertNotIn('_all_rows', compare_payload['key_pin_net_diff'])
        self.assertNotIn('_all_rows', compare_payload['net_diff'])

    def test_localhost_web_flow_uses_real_page_for_page_summary_and_keeps_query_pages_split(self):
        root = self.make_root()
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'demo',
                'ratio_limit': '70',
                'custom_volt_map': 'P3V3=3.3',
                'include_total_bom': 'on',
            },
        )
        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        run_id = payload['run_id']

        report_json = self.client.get(f'/api/report/{run_id}')
        self.assertEqual(200, report_json.status_code)
        report_payload = report_json.get_json()
        self.assertEqual('demo', report_payload['project_name'])
        self.assertTrue(report_payload['include_total_bom'])
        self.assertTrue(any(metric['label'] == '总BOM 总数' for metric in report_payload['metrics']))
        self.assertTrue(any('总 BOM：开启' in line for line in report_payload['summary_lines']))
        self.assertTrue(any(section['id'] == 'resistor' for section in report_payload['sections']))
        self.assertTrue(report_payload['top_insights'])
        self.assertTrue(any(card['id'] == 'bom' for card in report_payload['section_cards']))
        self.assertIn('review_plan', report_payload)
        self.assertIn('table_display_policy', report_payload)
        self.assertIn('focus_items', report_payload['review_plan'])
        self.assertIn('review_groups', report_payload['review_plan'])
        self.assertIn('trust_counts', report_payload['review_plan']['summary'])
        self.assertEqual(
            'info',
            next(item for item in report_payload['table_display_policy'] if item['table_id'] == 'page_rows')['level'],
        )
        self.assertEqual(
            'debug',
            next(item for item in report_payload['table_display_policy'] if item['table_id'] == 'page_mapping_rows')['level'],
        )
        self.assertEqual(
            '证据明细',
            next(item for item in report_payload['table_display_policy'] if item['table_id'] == 'page_mapping_rows')['trust_label'],
        )
        bom_section = next(section for section in report_payload['sections'] if section['id'] == 'bom')
        total_bom_table = next(table for table in bom_section['tables'] if table['id'] == 'bom_total_merged')
        self.assertEqual('总 BOM', total_bom_table['title'])
        self.assertEqual('信息统计', total_bom_table['trust_label'])
        self.assertIn('贴装数量', total_bom_table['columns'])
        self.assertIn('DEPOP数量', total_bom_table['columns'])

        network_section = next(section for section in report_payload['sections'] if section['id'] == 'network')
        page_table = next(table for table in network_section['tables'] if table['id'] == 'page_rows')
        self.assertEqual(['PAGE518'], [row['页码'] for row in page_table['rows']])
        mapping_table = next(table for table in network_section['tables'] if table['id'] == 'page_mapping_rows')
        self.assertEqual('PAGE242', mapping_table['rows'][0]['主模块页'])
        self.assertEqual('PAGE518', mapping_table['rows'][0]['页码'])
        self.assertEqual('是', mapping_table['rows'][0]['是否一一对应'])

        resistor_section = next(section for section in report_payload['sections'] if section['id'] == 'resistor')
        chip_table = next(table for table in resistor_section['tables'] if table['id'] == 'chip_pin_rows')
        self.assertIn('页码', chip_table['columns'])
        self.assertNotIn('页面', chip_table['columns'])
        self.assertNotIn('真实页', chip_table['columns'])
        self.assertEqual('PAGE518', chip_table['rows'][0]['页码'])

        report_page = self.client.get(f'/report/{run_id}')
        self.assertEqual(200, report_page.status_code)
        report_html = report_page.get_data(as_text=True)
        self.assertIn('PSTX 原理图审查平台', report_html)
        self.assertIn('report-topbar', report_html)
        self.assertIn('report-inspector', report_html)
        self.assertIn('inspector-toggle', report_html)
        self.assertIn('收起右栏', report_html)
        self.assertIn('inspector-content', report_html)
        self.assertIn('id="report-decision-strip"', report_html)
        self.assertIn('id="review-plan"', report_html)
        self.assertIn('审查任务分层', report_html)
        self.assertIn('完整指标', report_html)
        self.assertIn('完整分区索引', report_html)
        self.assertIn('明确异常', report_html)
        self.assertIn('规则候选', report_html)
        self.assertIn('nav-brand', report_html)
        self.assertIn('status-pill', report_html)
        self.assertIn('summary-details', report_html)
        self.assertIn('sidebar-toggle', report_html)
        self.assertIn('query-results', report_html)
        self.assertIn('project-manager', report_html)
        self.assertIn(f'href="/topology?run_id={run_id}"', report_html)
        self.assertIn(f'href="/compare?left_run_id={run_id}"', report_html)
        self.assertIn(f'href="/dfmea?run_id={run_id}"', report_html)
        self.assertIn('aster-summary-button', report_html)
        self.assertIn('aster-float-launcher', report_html)
        self.assertIn('aria-label="打开 Aster AI 客服审查助手"', report_html)
        self.assertIn('aria-expanded="false"', report_html)
        self.assertIn('aster-bubble-avatar', report_html)
        self.assertIn('aster-bubble-status', report_html)
        self.assertIn('aster-panel-minimize', report_html)
        self.assertIn('aster-panel-reset-position', report_html)
        self.assertIn('data-drag-handle="aster"', report_html)
        self.assertIn('AI 浮窗助手', report_html)
        self.assertIn('输入问题即可连续追问', report_html)
        self.assertIn('</div>\n  <button id="aster-float-launcher"', report_html)
        self.assertIn('aster-auth-status', report_html)
        self.assertIn('AI 配置已移至独立页面', report_html)
        self.assertIn('href="/ai-settings"', report_html)
        self.assertIn('href="/guide"', report_html)
        self.assertNotIn('aster-credential-form', report_html)
        self.assertNotIn('Room Validate Origin', report_html)
        self.assertIn('harness-agent-form', report_html)
        self.assertIn('harness-agent-enable-subagents', report_html)
        self.assertIn('harness-agent-max-tool-calls', report_html)
        self.assertIn('agent-trace-drawer', report_html)
        self.assertIn('发送问题', report_html)
        self.assertNotIn('Agent 审查工作台', report_html)
        self.assertNotIn('name="base_url"', report_html)
        self.assertNotIn('Casebook', report_html)
        self.assertNotIn('Campaign', report_html)
        self.assertNotIn('Quality Gate', report_html)
        self.assertNotIn('Quality History', report_html)

        agent_eval_page = self.client.get('/agent-eval')
        self.assertEqual(200, agent_eval_page.status_code)
        agent_eval_html = agent_eval_page.get_data(as_text=True)
        self.assertIn('Agent Eval Center', agent_eval_html)
        self.assertIn('agent-eval-case-list', agent_eval_html)
        self.assertIn('agent-eval-run-all', agent_eval_html)

        agent_lab_page = self.client.get('/agent-lab')
        self.assertEqual(200, agent_lab_page.status_code)
        agent_lab_html = agent_lab_page.get_data(as_text=True)
        self.assertIn('Agent 能力实验室', agent_lab_html)
        self.assertIn('agent-lab-form', agent_lab_html)
        self.assertIn('agent-lab-reindex', agent_lab_html)
        self.assertIn('agent-lab-checklist-reindex', agent_lab_html)
        self.assertIn('id="agent-trace-close"', agent_lab_html)
        self.assertIn('data-agent-trace-close', agent_lab_html)
        self.assertNotIn('agent-trace-panel', agent_lab_html)
        self.assertIn('ref/', agent_lab_html)
        self.assertIn('ref_checklist/', agent_lab_html)

        compare_page = self.client.get('/compare')
        self.assertEqual(200, compare_page.status_code)
        compare_html = compare_page.get_data(as_text=True)
        self.assertIn('项目对比工作台', compare_html)
        self.assertIn('默认先按 Net 视角查看网络迁移', compare_html)
        self.assertIn('compare-page-form', compare_html)
        self.assertIn('compare-result-host', compare_html)
        self.assertIn('href="/guide"', compare_html)
        self.assertIn('value="300"', compare_html)
        self.assertIn('aster-float-launcher', compare_html)
        self.assertIn('AI 对比助手', compare_html)
        self.assertIn('harness-agent-form', compare_html)
        self.assertIn('agent-trace-drawer', compare_html)
        self.assertIn('AI 配置已移至独立页面', compare_html)
        self.assertNotIn('aster-credential-form', compare_html)

        topology_page = self.client.get(f'/topology?run_id={run_id}')
        self.assertEqual(200, topology_page.status_code)
        topology_html = topology_page.get_data(as_text=True)
        self.assertIn('芯片 / 连接器拓扑视图', topology_html)
        self.assertIn('topology-controls', topology_html)
        self.assertIn('topology-graph', topology_html)
        self.assertIn('topology-redraw', topology_html)
        self.assertIn('topology-view', topology_html)
        self.assertIn('topology-supply-mode', topology_html)
        self.assertIn('topology-edge-label-mode', topology_html)
        self.assertIn('id="topology-include-connectors" name="include_connectors" type="checkbox">', topology_html)
        self.assertIn('static/pages/topology.js', topology_html)

        debug_topology_page = self.client.get('/debug/topology')
        self.assertEqual(200, debug_topology_page.status_code)
        self.assertIn('Debug Topology Fixture', debug_topology_page.get_data(as_text=True))

        ai_settings_page = self.client.get('/ai-settings')
        self.assertEqual(200, ai_settings_page.status_code)
        ai_settings_html = ai_settings_page.get_data(as_text=True)
        self.assertIn('AI 配置中心', ai_settings_html)
        self.assertIn('aster-credential-form', ai_settings_html)
        self.assertIn('服务地址固定为', ai_settings_html)
        self.assertIn('https://aigc.huaqin.com', ai_settings_html)
        self.assertIn('Room Validate Origin', ai_settings_html)

        for removed_path in ['/casebook', '/campaign', '/quality-gate', '/quality-history']:
            self.assertEqual(404, self.client.get(removed_path).status_code)

        aster_response = self.client.get(f'/api/report/{run_id}/aster-summary')
        self.assertEqual(200, aster_response.status_code)
        aster_payload = aster_response.get_json()
        self.assertTrue(aster_payload['ok'])
        self.assertEqual('mock', aster_payload['mode'])
        self.assertEqual('local-aster-mock', aster_payload['provider'])
        self.assertTrue(aster_payload['section_focus'])
        self.assertTrue(aster_payload['review_checklist'])
        self.assertTrue(aster_payload['manual_review'])
        self.assertTrue(any('不访问真实 Aster' in item for item in aster_payload['safeguards']))

        query_json = self.client.post(
            f'/api/report/{run_id}/query',
            json={'mode': '位号', 'keyword': 'U1'},
        )
        self.assertEqual(200, query_json.status_code)
        query_payload = query_json.get_json()
        self.assertIn('U1', '\n'.join(query_payload['lines']))
        self.assertEqual('component', query_payload['view'])
        meta_map = {item['label']: item['value'] for item in query_payload['summary']['meta']}
        self.assertEqual('PAGE518', meta_map['页码'])
        self.assertEqual('是', meta_map['主模块页映射一一对应'])

        network_query_json = self.client.post(
            f'/api/report/{run_id}/query',
            json={'mode': '网络名', 'keyword': 'SMBALERT_N'},
        )
        self.assertEqual(200, network_query_json.status_code)
        network_query_payload = network_query_json.get_json()
        node_row = network_query_payload['cards'][0]['items'][0]
        self.assertEqual('PAGE518', node_row['页码'])
        self.assertEqual('是', node_row['主模块页映射一一对应'])

        export_response = self.client.get(f'/api/report/{run_id}/export')
        self.assertEqual(200, export_response.status_code)
        self.assertTrue(export_response.data.startswith(b'PK'))

        app_js = self.client.get('/static/app.js')
        try:
            self.assertEqual(200, app_js.status_code)
            app_js_text = app_js.get_data(as_text=True)
            self.assertIn('table-sort-column', app_js_text)
            self.assertIn('reportDecisionItemNode', app_js_text)
            self.assertIn('renderReportDecisionStrip', app_js_text)
            self.assertIn('table-trust-badge', app_js_text)
            self.assertIn('review-plan-trust', app_js_text)
            self.assertIn('column-resize-handle', app_js_text)
            self.assertIn('toolbar-density', app_js_text)
            self.assertIn('metricIconForLabel', app_js_text)
            self.assertIn('navIconForSection', app_js_text)
            self.assertIn('TABLE_INITIAL_RENDER_LIMIT', app_js_text)
            self.assertIn('scheduleTableMount', app_js_text)
            self.assertIn('scheduleScrollShadowUpdate', app_js_text)
            self.assertIn('sidebar-toggle', app_js_text)
            self.assertIn('columnFilterMatches', app_js_text)
            self.assertIn('column-filter-builder', app_js_text)
            self.assertIn('column-filter-add', app_js_text)
            self.assertIn('renderProjectManager', app_js_text)
            self.assertIn('bootComparePage', app_js_text)
            self.assertIn('bootCompareHarnessAgent', app_js_text)
            self.assertIn('/api/compare', app_js_text)
            self.assertIn('/api/compare/harness-agent', app_js_text)
            self.assertIn('/api/compare/harness/profiles', app_js_text)
            self.assertIn('/api/feishu-bom/database/rows/', app_js_text)
            self.assertIn('确认从本地缓存剔除', app_js_text)
            self.assertIn('feishu-row-action-head', app_js_text)
            self.assertIn('关键器件 Pin/Net 连接差异', app_js_text)
            self.assertIn('R/C/L Pin/Net 连接差异', app_js_text)
            self.assertIn('Net 视角变化', app_js_text)
            self.assertIn('compare-net-focus', app_js_text)
            self.assertIn('compare-perspective-controls', app_js_text)
            self.assertIn('compare-diff-preview', app_js_text)
            self.assertIn('bootReveals(result)', app_js_text)
            self.assertIn('default_density', app_js_text)
            self.assertIn('staggerChildren', app_js_text)
            self.assertIn('restartMotion', app_js_text)
            self.assertIn('bootPageMotion', app_js_text)
            self.assertIn('applyGlobalStaggers', app_js_text)
            self.assertIn('showLoadingMask', app_js_text)
            self.assertIn('bootRuntimeHints', app_js_text)
            self.assertIn('runWhenBrowserIsIdle', app_js_text)
            self.assertIn('MAX_STAGGERED_NODES_PER_SELECTOR', app_js_text)
            self.assertIn('bootAsterSummary', app_js_text)
            self.assertIn('bootAsterFloatingPanel', app_js_text)
            self.assertIn('bootInspectorToggle', app_js_text)
            self.assertIn('pstx-report-inspector-collapsed', app_js_text)
            self.assertIn('is-inspector-collapsed', app_js_text)
            self.assertIn('bootAsterChatAgent', app_js_text)
            self.assertIn('pstx_aster_panel_position', app_js_text)
            self.assertIn('setPointerCapture', app_js_text)
            self.assertIn("saved = 'closed'", app_js_text)
            self.assertIn("aria-expanded", app_js_text)
            self.assertIn("aria-hidden", app_js_text)
            self.assertIn('bootAsterStatus', app_js_text)
            self.assertIn('bootAsterCredentialForm', app_js_text)
            self.assertIn('asterChecklistStatusLabel', app_js_text)
            self.assertIn('review_checklist', app_js_text)
            self.assertIn('manual_review', app_js_text)
            self.assertIn('renderAsterError', app_js_text)
            self.assertIn('diagnostic_hints', app_js_text)
            self.assertIn('Aster 调用失败', app_js_text)
            self.assertIn('/aster-summary', app_js_text)
            self.assertIn('/api/aster/status', app_js_text)
            self.assertIn('/api/aster/runtime-config', app_js_text)
            self.assertNotIn('bootHarnessAgentWorkbench', app_js_text)
            self.assertIn('/api/harness/profiles', app_js_text)
            self.assertIn('/harness/agent', app_js_text)
            self.assertIn('/api/harness/agent-runs', app_js_text)
            self.assertIn('renderNeedsUserInputForm', app_js_text)
            self.assertIn('ensureAgentChatThread', app_js_text)
            self.assertIn('appendAgentChatMessage', app_js_text)
            self.assertIn('inferReportAgentProfile', app_js_text)
            self.assertIn('inferCompareAgentProfile', app_js_text)
            self.assertIn("profileSelect.value || 'auto'", app_js_text)
            self.assertIn('capability_plan', app_js_text)
            self.assertIn('清空本地对话', app_js_text)
            self.assertIn('对比页浮窗已进入连续对话模式', app_js_text)
            self.assertIn('AGENT_STATUS_STAGES', app_js_text)
            self.assertIn('attachAgentStageController', app_js_text)
            self.assertIn('data-agent-stage-status', app_js_text)
            self.assertIn('规划取证路线', app_js_text)
            self.assertIn('校验证据引用', app_js_text)
            self.assertIn('收集报告摘要', app_js_text)
            self.assertIn('/harness/context/clear', app_js_text)
            self.assertIn('context_answers', app_js_text)
            self.assertIn('continue_agent_run_id', app_js_text)
            self.assertIn('enable_subagents', app_js_text)
            self.assertIn('并行 Subagents', app_js_text)
            self.assertIn('bootAgentEvalPage', app_js_text)
            self.assertIn('/api/agent-eval/status', app_js_text)
            self.assertIn('/api/agent-eval/run', app_js_text)
            self.assertIn('bootAgentLabPage', app_js_text)
            self.assertIn('/api/agent-lab/status', app_js_text)
            self.assertIn('/api/agent-lab/ref/reindex', app_js_text)
            self.assertIn('/api/agent-lab/checklist/reindex', app_js_text)
            self.assertIn('/api/agent-lab/ask', app_js_text)
            self.assertIn('agent_ref_qa', app_js_text)
            self.assertIn('review_checklist_qa', app_js_text)
            self.assertIn('[data-agent-trace-close]', app_js_text)
            self.assertIn('agent-result-dismiss', app_js_text)
            self.assertIn('is-body-collapsed', app_js_text)
            self.assertNotIn('bootCasebookPage', app_js_text)
            self.assertNotIn('/api/casebook/cases', app_js_text)
            self.assertNotIn('/casebook/from-agent-run', app_js_text)
            self.assertNotIn('bootCampaignPage', app_js_text)
            self.assertNotIn('/api/campaign/status', app_js_text)
            self.assertNotIn('/api/campaign/run', app_js_text)
            self.assertNotIn('bootQualityGatePage', app_js_text)
            self.assertNotIn('/api/quality-gate/status', app_js_text)
            self.assertNotIn('/api/quality-gate/evaluate', app_js_text)
            self.assertNotIn('bootQualityHistoryPage', app_js_text)
            self.assertNotIn('/api/quality-history/status', app_js_text)
        finally:
            app_js.close()

        app_css = self.client.get('/static/app.css')
        try:
            self.assertEqual(200, app_css.status_code)
            app_css_text = app_css.get_data(as_text=True)
            self.assertIn('.column-resize-handle', app_css_text)
            self.assertIn('table-layout: fixed', app_css_text)
            self.assertIn('.report-topbar', app_css_text)
            self.assertIn('.report-decision-strip', app_css_text)
            self.assertIn('.report-secondary-panel', app_css_text)
            self.assertIn('.table-trust-badge', app_css_text)
            self.assertIn('.review-plan-trust', app_css_text)
            self.assertIn('.report-inspector', app_css_text)
            self.assertIn('.report-layout.is-inspector-collapsed', app_css_text)
            self.assertIn('.report-inspector.is-collapsed', app_css_text)
            self.assertIn('writing-mode: vertical-rl', app_css_text)
            self.assertIn('.nav-logo-mark', app_css_text)
            self.assertIn('.metric-icon', app_css_text)
            self.assertIn('content-visibility: auto', app_css_text)
            self.assertIn('.table-render-footer', app_css_text)
            self.assertIn('.column-filter-panel', app_css_text)
            self.assertIn('.column-filter-row', app_css_text)
            self.assertIn('.project-manager', app_css_text)
            self.assertIn('.compare-shell', app_css_text)
            self.assertIn('.compare-domain-section', app_css_text)
            self.assertIn('.compare-diff-preview', app_css_text)
            self.assertIn('.compare-stat-grid', app_css_text)
            self.assertIn('@keyframes soft-rise', app_css_text)
            self.assertIn('.table-open-pulse', app_css_text)
            self.assertIn('.compare-result-enter', app_css_text)
            self.assertIn('@keyframes page-panel-in', app_css_text)
            self.assertIn('.query-result-enter', app_css_text)
            self.assertIn('.loading-mask.is-visible', app_css_text)
            self.assertIn('Microsoft YaHei UI', app_css_text)
            self.assertIn('scrollbar-gutter: stable', app_css_text)
            self.assertIn('html.is-windows', app_css_text)
            self.assertIn('.aster-assist-panel', app_css_text)
            self.assertIn('.aster-float-launcher', app_css_text)
            self.assertIn('.aster-chat-form', app_css_text)
            self.assertIn('.aster-advanced-settings', app_css_text)
            self.assertIn('.aster-assist-panel.is-dragging', app_css_text)
            self.assertIn('.aster-bubble-avatar', app_css_text)
            self.assertIn('.aster-bubble-status', app_css_text)
            self.assertIn('@keyframes aster-bubble-soft-pulse', app_css_text)
            self.assertIn('@keyframes aster-bubble-float', app_css_text)
            self.assertIn('@keyframes ambient-soft-breathe', app_css_text)
            self.assertIn('@keyframes agent-loading-dot', app_css_text)
            self.assertNotIn('@keyframes ambient-drift', app_css_text)
            self.assertNotIn('backdrop-filter: blur', app_css_text)
            self.assertIn('.aster-window-head', app_css_text)
            self.assertIn('.aster-checklist', app_css_text)
            self.assertIn('.aster-manual-review', app_css_text)
            self.assertIn('.aster-focus-grid', app_css_text)
            self.assertIn('.aster-auth-status', app_css_text)
            self.assertIn('.aster-auth-grid', app_css_text)
            self.assertIn('.aster-credential-form', app_css_text)
            self.assertIn('.feishu-row-action-head', app_css_text)
            self.assertIn('.feishu-row-actions', app_css_text)
            self.assertIn('position: sticky;', app_css_text)
            self.assertIn('right: 0;', app_css_text)
            self.assertIn('.danger-ghost-btn', app_css_text)
            self.assertNotIn('.agent-workbench', app_css_text)
            self.assertIn('.agent-result-shell', app_css_text)
            self.assertIn('.agent-result-shell.is-compact', app_css_text)
            self.assertIn('.agent-result-body', app_css_text)
            self.assertIn('.agent-result-title-actions', app_css_text)
            self.assertIn('.agent-result-shell.is-body-collapsed .agent-result-body', app_css_text)
            self.assertIn('.agent-result-compact-note', app_css_text)
            self.assertIn('.agent-chat-thread', app_css_text)
            self.assertIn('.agent-chat-message', app_css_text)
            self.assertIn('.agent-chat-loading', app_css_text)
            self.assertIn('.agent-trace-drawer', app_css_text)
            self.assertIn('.agent-subagent', app_css_text)
            self.assertIn('.agent-eval-workspace', app_css_text)
            self.assertIn('.agent-eval-case-item', app_css_text)
            self.assertIn('.agent-eval-result-shell', app_css_text)
            self.assertIn('.agent-lab-workspace', app_css_text)
            self.assertIn('.agent-lab-doc-item', app_css_text)
            self.assertIn('.agent-lab-form', app_css_text)
            self.assertIn('.topology-workspace', app_css_text)
            self.assertIn('.topology-graph', app_css_text)
            self.assertIn('.topology-node-orb', app_css_text)
            self.assertIn('.bom-export-menu', app_css_text)
            self.assertNotIn('.casebook-workspace', app_css_text)
            self.assertNotIn('.campaign-workspace', app_css_text)
            self.assertNotIn('.quality-gate-workspace', app_css_text)
            self.assertNotIn('.quality-history-workspace', app_css_text)
        finally:
            app_css.close()

    def test_report_page_exports_bom_by_depop_mode(self):
        prt_text = (
            "PART_NAME\n"
            "R1 'RES_0402'\n"
            "HQ_CODE='PN-SAME'\n"
            "VALUE='10k'\n"
            "PACKAGE='0402'\n"
            "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I1'\n"
            "PART_NAME\n"
            "R2 'RES_0402'\n"
            "HQ_CODE='PN-SAME'\n"
            "VALUE='10k'\n"
            "PACKAGE='0402'\n"
            "BOM_OPTION='DEPOP'\n"
            "DRAWING='@GPU_2SW_BOARD_LIB.GPU_2SW_BOARD(SCH_1):PAGE242_I2'\n"
        )
        net_text = (
            "NET_NAME\n"
            "'P3V3'\n"
            "NODE_NAME R1 1\n"
            "'1':\n"
            "NODE_NAME R2 1\n"
            "'1':\n"
        )
        root = self.make_root_with_samples(prt_text, net_text)
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'bom-export-demo',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        report_page = self.client.get(f'/report/{run_id}')
        self.assertEqual(200, report_page.status_code)
        report_html = report_page.get_data(as_text=True)
        self.assertIn('导出 BOM', report_html)
        self.assertIn(f'/api/report/{run_id}/bom/export?mode=all', report_html)
        self.assertIn(f'/api/report/{run_id}/bom/export?mode=mounted', report_html)
        self.assertIn(f'/api/report/{run_id}/bom/export?mode=split', report_html)

        all_export = self.client.get(f'/api/report/{run_id}/bom/export?mode=all')
        mounted_export = self.client.get(f'/api/report/{run_id}/bom/export?mode=mounted')
        split_export = self.client.get(f'/api/report/{run_id}/bom/export?mode=split')
        try:
            self.assertEqual(200, all_export.status_code)
            self.assertEqual(200, mounted_export.status_code)
            self.assertEqual(200, split_export.status_code)
            self.assertIn('bom-export-demo_bom_split.xlsx', split_export.headers.get('Content-Disposition', ''))

            with tempfile.TemporaryDirectory() as temp_dir:
                split_path = Path(temp_dir) / 'split.xlsx'
                split_path.write_bytes(split_export.data)
                workbook = load_workbook(split_path)
                try:
                    ws = workbook['BOM_分条']
                    headers = [cell.value for cell in ws[1]]
                    status_col = headers.index('BOM状态') + 1
                    pn_col = headers.index('料号') + 1
                    qty_col = headers.index('数量') + 1
                    self.assertEqual(['贴装', 'DEPOP'], [ws.cell(row, status_col).value for row in (2, 3)])
                    self.assertEqual(['PN-SAME', 'PN-SAME'], [ws.cell(row, pn_col).value for row in (2, 3)])
                    self.assertEqual([1, 1], [ws.cell(row, qty_col).value for row in (2, 3)])
                finally:
                    workbook.close()

                all_path = Path(temp_dir) / 'all.xlsx'
                all_path.write_bytes(all_export.data)
                workbook = load_workbook(all_path)
                try:
                    ws = workbook['BOM_含DEPOP']
                    headers = [cell.value for cell in ws[1]]
                    self.assertEqual(2, ws.cell(2, headers.index('数量') + 1).value)
                finally:
                    workbook.close()
        finally:
            all_export.close()
            mounted_export.close()
            split_export.close()

        invalid = self.client.get(f'/api/report/{run_id}/bom/export?mode=bad')
        self.assertEqual(400, invalid.status_code)
        self.assertFalse(invalid.get_json()['ok'])

    def test_report_schematic_pdf_annotation_api_returns_overlay(self):
        root = self.make_root()
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'pdf-annotation-demo',
                'include_depop': 'on',
                'ratio_limit': '70',
                'custom_volt_map': '',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']
        pdf_dir = Path(tempfile.mkdtemp())
        self.temp_roots.append(pdf_dir)
        pdf_path = pdf_dir / 'schematic.pdf'
        write_minimal_pdf(pdf_path, page_count=1)

        response = self.client.post(
            f'/api/report/{run_id}/schematic-pdf/annotations',
            json={
                'pdf_path': str(pdf_path),
                'targets': [{
                    'kind': 'coordinate',
                    'page': 'PAGE1',
                    'label': '降额提醒',
                    'severity': 'warning',
                    'pdf_page_number': 1,
                    'pdf_bbox': [10, 20, 50, 70],
                }],
                'stdout': 'full',
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        annotation = payload['schematic_pdf_annotation']
        self.assertEqual('pstx-schematic-pdf-annotation.v1', annotation['schema_version'])
        self.assertEqual(1, annotation['summary']['matched_count'])
        self.assertEqual('explicit_pdf_bbox', annotation['annotations'][0]['confidence'])
        self.assertEqual('rect', annotation['page_overlays'][0]['overlays'][0]['shape'])

    def test_topology_page_api_returns_chip_connector_graph(self):
        root = self.make_root_with_samples(PRT_SAMPLE_TOPOLOGY, NET_SAMPLE_TOPOLOGY)
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'topology-demo',
                'ratio_limit': '70',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        default_response = self.client.get(f'/api/report/{run_id}/topology?limit=50')
        self.assertEqual(200, default_response.status_code)
        default_topology = default_response.get_json()['topology']
        self.assertEqual('summary', default_topology['view'])
        self.assertEqual('grouped', default_topology['supply_mode'])
        self.assertFalse(default_topology['include_connectors'])
        self.assertIn('topology_cache_status', default_topology)
        self.assertFalse(any(node['refdes'] == 'J8' for node in default_topology['nodes']))

        topology_response = self.client.get(f'/api/report/{run_id}/topology?include_connectors=1&limit=50')
        self.assertEqual(200, topology_response.status_code)
        payload = topology_response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual('topology-demo', payload['project_name'])
        self.assertIn('topology_cache_status', payload)
        self.assertIn('topology_timing', payload)
        topology = payload['topology']
        self.assertEqual('llm-topology.v1', topology['schema_version'])
        self.assertGreaterEqual(topology['node_count'], 2)
        self.assertGreaterEqual(topology['edge_count'], 1)
        self.assertTrue(any(node['refdes'] == 'J8' for node in topology['nodes']))
        self.assertTrue(any(edge['source_refdes'] == 'U46' or edge['target_refdes'] == 'U46' for edge in topology['edges']))

        full_response = self.client.get(f'/api/report/{run_id}/topology?include_connectors=1&view=full&supply_mode=details&limit=50')
        self.assertEqual(200, full_response.status_code)
        full_topology = full_response.get_json()['topology']
        self.assertEqual('full', full_topology['view'])
        self.assertEqual('details', full_topology['supply_mode'])

        bad_limit = self.client.get(f'/api/report/{run_id}/topology?limit=bad')
        self.assertEqual(400, bad_limit.status_code)

    def test_report_cadence_page_api_returns_connectivity_semantics(self):
        root = self.make_root()
        (root / 'sch_1' / 'page114.csa').write_text(
            '\n'.join([
                'WIRE 16 -1 (0 0)(100 0);',
                'FORCEPROP 2 LAST SIG_NAME I2C_SCL;',
                'NET_LABEL 1 (50 0) I2C_SCL;',
                'PORT 1 (100 0) I2C_SCL OUTPUT;',
            ]),
            encoding='utf-8',
        )
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'cadence-page-demo',
                'ratio_limit': '70',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        page_response = self.client.get(f'/api/report/{run_id}/cadence-page?page=114&stdout=objects')
        self.assertEqual(200, page_response.status_code)
        payload = page_response.get_json()
        self.assertTrue(payload['ok'])
        page = payload['cadence_page']
        self.assertEqual('pstx-cadence-page.v1', page['schema_version'])
        self.assertEqual(1, page['connectivity_summary']['semantic_counts']['NET_LABEL'])
        label_id = next(item['object_id'] for item in page['objects'] if item['type'] == 'NET_LABEL')

        detail_response = self.client.get(f'/api/report/{run_id}/cadence-page?page=114&object_id={label_id}')
        self.assertEqual(200, detail_response.status_code)
        self.assertEqual(label_id, detail_response.get_json()['cadence_page']['object']['object_id'])

        bad_page = self.client.get(f'/api/report/{run_id}/cadence-page?page=bad')
        self.assertEqual(400, bad_page.status_code)

        index_response = self.client.get(
            f'/api/report/{run_id}/cadence-index?stdout=full&query=I2C&kind=all&limit=20'
        )
        self.assertEqual(200, index_response.status_code)
        index_payload = index_response.get_json()
        self.assertTrue(index_payload['ok'])
        cadence_index = index_payload['cadence_index']
        self.assertEqual('pstx-cadence-index.v1', cadence_index['schema_version'])
        self.assertEqual(1, cadence_index['digest']['net_count'])
        self.assertEqual('I2C_SCL', cadence_index['net_rows'][0]['name'])
        self.assertEqual('I2C_SCL', cadence_index['port_rows'][0]['name'])

        bad_index_page = self.client.get(f'/api/report/{run_id}/cadence-index?page=bad')
        self.assertEqual(400, bad_index_page.status_code)

    def test_report_payload_exposes_module_scope_section(self):
        root = build_project_root_for_page_v2()
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'demo',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        report_payload = self.client.get(f'/api/report/{run_id}').get_json()

        self.assertTrue(any(metric['label'] == '子模块数' and metric['value'] == 1 for metric in report_payload['metrics']))
        self.assertTrue(any('模块视角：识别到 1 个子模块实例' in line for line in report_payload['summary_lines']))
        module_section = next(section for section in report_payload['sections'] if section['id'] == 'module')
        summary_table = next(table for table in module_section['tables'] if table['id'] == 'module_scope_rows')
        component_table = next(table for table in module_section['tables'] if table['id'] == 'module_component_rows')
        submodule_row = next(row for row in summary_table['rows'] if row['模块类型'] == '子模块')
        self.assertEqual('i2c_repeater_9617_cbb_v3', submodule_row['模块名'])
        self.assertEqual('PAGE114', submodule_row['父级Symbol页码'])
        self.assertEqual('I70', submodule_row['父级Symbol实例'])
        component_row = next(row for row in component_table['rows'] if row['位号'] == 'C1A104')
        self.assertEqual('子模块', component_row['模块类型'])
        self.assertEqual('PAGE177', component_row['页码'])

    def test_aster_live_mode_missing_config_reports_displayable_error(self):
        self.with_env({
            'PSTX_ASTER_MODE': 'live',
            'ASTER_BASE_URL': None,
            'ASTER_API_KEY': None,
            'ASTER_EMP_NO': None,
        })
        root = self.make_root()
        analyze_response = self.client.post('/api/analyze', data={'project_root': str(root)})
        run_id = analyze_response.get_json()['run_id']

        response = self.client.get(f'/api/report/{run_id}/aster-summary')
        self.assertEqual(400, response.status_code)
        payload = response.get_json()
        self.assertFalse(payload['ok'])
        self.assertEqual('config', payload['error_type'])
        self.assertNotIn('ASTER_BASE_URL', payload['error'])
        self.assertIn('ASTER_EMP_NO', payload['error'])
        self.assertIn('log_file', payload)
        self.assertNotIn('accessToken', payload['error'])

    def test_aster_status_endpoint_redacts_credentials(self):
        self.with_env({
            'PSTX_ASTER_MODE': 'live',
            'PSTX_ASTER_BACKEND': 'chat-flow',
            'ASTER_BASE_URL': 'https://aster.example.local/api?secret=query-secret',
            'ASTER_EMP_NO': '100019100',
            'ASTER_API_KEY': 'super-secret-key',
        })

        response = self.client.get('/api/aster/status')
        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual('ready', payload['status'])
        self.assertTrue(payload['live_ready'])
        payload_text = str(payload)
        self.assertNotIn('super-secret-key', payload_text)
        self.assertNotIn('query-secret', payload_text)
        item_map = {item['name']: item for item in payload['items']}
        self.assertEqual('https://aigc.huaqin.com', item_map['ASTER_FIXED_BASE_URL']['value'])
        self.assertEqual('fixed', item_map['ASTER_FIXED_BASE_URL']['source'])
        self.assertTrue(item_map['ASTER_API_KEY']['configured'])
        self.assertNotIn('value', item_map['ASTER_API_KEY'])

    def test_aster_runtime_config_can_set_and_clear_without_echoing_secret(self):
        response = self.client.post('/api/aster/runtime-config', json={
            'mode': 'live',
            'backend': 'chat-flow',
            'base_url': 'https://runtime-aster.example.local/api',
            'emp_no': '100019100',
            'api_key': 'runtime-secret-key',
            'origin': 'runtime-origin.example.local',
        })
        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertEqual('ready', payload['status'])
        self.assertTrue(payload['runtime_override_active'])
        self.assertNotIn('runtime-secret-key', str(payload))
        item_map = {item['name']: item for item in payload['items']}
        self.assertNotIn('ASTER_BASE_URL', item_map)
        self.assertEqual('https://aigc.huaqin.com', item_map['ASTER_FIXED_BASE_URL']['value'])
        self.assertEqual('runtime', item_map['ASTER_API_KEY']['source'])
        self.assertNotIn('value', item_map['ASTER_API_KEY'])
        self.assertEqual('runtime-origin.example.local', item_map['ASTER_ORIGIN']['value'])

        clear_response = self.client.delete('/api/aster/runtime-config')
        self.assertEqual(200, clear_response.status_code)
        clear_payload = clear_response.get_json()
        self.assertFalse(clear_payload['runtime_override_active'])

    def test_dfmea_workbench_page_api_and_excel_export(self):
        root = self.make_root_with_samples(PRT_SAMPLE_DEPOP, NET_SAMPLE_DEPOP)
        analyze = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'dfmea-demo',
                'ratio_limit': '70',
            },
        )
        self.assertEqual(200, analyze.status_code)
        run_id = analyze.get_json()['run_id']

        page = self.client.get(f'/dfmea?run_id={run_id}')
        self.assertEqual(200, page.status_code)
        page_html = page.get_data(as_text=True)
        self.assertIn('DFMEA 工作台', page_html)
        self.assertIn('dfmea-pending-list', page_html)
        self.assertIn('pages/dfmea.js', page_html)
        self.assertIn('name="exclude_rc"', page_html)

        default_payload = self.client.get(f'/api/report/{run_id}/dfmea/workbench').get_json()
        self.assertTrue(default_payload['ok'])
        self.assertEqual(['U1'], [row['refdes'] for row in default_payload['pending_components']])
        self.assertEqual('page', default_payload['sort'])

        with_depop = self.client.get(f'/api/report/{run_id}/dfmea/workbench?include_depop=1').get_json()
        self.assertEqual(['U1', 'R1'], [row['refdes'] for row in with_depop['pending_components']])
        without_rc = self.client.get(f'/api/report/{run_id}/dfmea/workbench?include_depop=1&exclude_rc=1').get_json()
        self.assertTrue(without_rc['exclude_rc'])
        self.assertEqual(['U1'], [row['refdes'] for row in without_rc['pending_components']])

        created = self.client.post(
            f'/api/report/{run_id}/dfmea/groups',
            json={
                'refdes': ['U1', 'R1'],
                'function_requirement': '核心控制与偏置',
                'failure_mode': '开路',
                'failure_effect': '系统功能异常',
                'failure_cause': '焊接异常',
                'prevention_detection': 'ICT/FCT',
            },
        )
        self.assertEqual(200, created.status_code)
        group_id = created.get_json()['group_id']

        after_group = self.client.get(f'/api/report/{run_id}/dfmea/workbench?include_depop=1').get_json()
        self.assertEqual([], [row['refdes'] for row in after_group['pending_components']])
        self.assertEqual('U1, R1', after_group['groups'][0]['refdes_text'])
        self.assertEqual('PAGE518', after_group['groups'][0]['pages_text'])
        self.assertEqual(['U1', 'R1'], [row['refdes'] for row in after_group['groups'][0]['components']])

        updated = self.client.patch(
            f'/api/report/{run_id}/dfmea/groups/{group_id}',
            json={
                'refdes': ['U1'],
                'function_requirement': '核心控制',
                'failure_mode': '短路',
                'failure_effect': '无法启动',
                'failure_cause': '过压',
                'prevention_detection': '降额检查',
            },
        )
        self.assertEqual(200, updated.status_code)
        updated_payload = self.client.get(f'/api/report/{run_id}/dfmea/workbench?include_depop=1').get_json()
        self.assertEqual(['R1'], [row['refdes'] for row in updated_payload['pending_components']])
        self.assertEqual('U1', updated_payload['groups'][0]['refdes_text'])

        export_response = self.client.get(f'/api/report/{run_id}/dfmea/export')
        try:
            self.assertEqual(200, export_response.status_code)
            workbook_path = Path(self._dfmea_tmp.name) / 'export.xlsx'
            workbook_path.write_bytes(export_response.data)
            workbook = load_workbook(workbook_path)
            sheet = workbook.active
            self.assertEqual(
                ['组ID', '位号', '页码', '功能/需求', '潜在失效模式', '潜在失效后果', '潜在失效原因/机理', '现有预防/探测方案', '更新时间'],
                [cell.value for cell in sheet[1]],
            )
            self.assertEqual('U1', sheet['B2'].value)
            self.assertEqual('PAGE518', sheet['C2'].value)
        finally:
            export_response.close()

        deleted = self.client.delete(f'/api/report/{run_id}/dfmea/groups/{group_id}')
        self.assertEqual(200, deleted.status_code)
        restored = self.client.get(f'/api/report/{run_id}/dfmea/workbench?include_depop=1').get_json()
        self.assertEqual(['U1', 'R1'], [row['refdes'] for row in restored['pending_components']])

    def test_dfmea_api_missing_run_returns_json_error(self):
        response = self.client.get('/api/report/not-found/dfmea/workbench')
        self.assertEqual(404, response.status_code)
        payload = response.get_json()
        self.assertFalse(payload['ok'])
        self.assertIn('未找到 DFMEA 对应报告', payload['error'])

    def test_compare_refdes_category_treats_power_passive_prefix_as_passive(self):
        self.assertEqual('passive', webapp_compare_view.refdes_category('PC16A10'))
        self.assertEqual('passive', webapp_compare_view.refdes_category('PR10A1'))
        self.assertEqual('passive', webapp_compare_view.refdes_category('PL2A5'))
        self.assertEqual('connector', webapp_compare_view.refdes_category('P1'))
        self.assertTrue(webapp_compare_view.is_passive_refdes('PC16A10'))

    def test_project_list_and_compare_api_tracks_multiple_runs(self):
        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        self.make_fake_feishu_data()
        self.add_fake_feishu_material(
            key_value='IC_CPU_SPEC_A',
            hq_no='PN_U1',
            spec='IC_CPU_SPEC_A',
            pi='PI-A',
            selection_order='A1',
        )
        self.add_fake_feishu_material(
            key_value='IC_CPU_SPEC_B',
            hq_no='PN_U1B',
            spec='IC_CPU_SPEC_B',
            pi='PI-B',
            selection_order='B1',
        )
        left_root = self.make_root()
        right_root = self.make_root_with_samples(PRT_SAMPLE_COMPARE, NET_SAMPLE_COMPARE)

        left_response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(left_root),
                'project_name': 'alpha',
                'ratio_limit': '70',
                'custom_volt_map': 'P3V3=3.3',
            },
        )
        self.assertEqual(200, left_response.status_code)
        left_run_id = left_response.get_json()['run_id']

        right_response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(right_root),
                'project_name': 'beta',
                'ratio_limit': '70',
                'custom_volt_map': 'P3V3=3.3',
            },
        )
        self.assertEqual(200, right_response.status_code)
        right_run_id = right_response.get_json()['run_id']

        projects_response = self.client.get('/api/projects')
        self.assertEqual(200, projects_response.status_code)
        projects_payload = projects_response.get_json()
        self.assertEqual(2, projects_payload['count'])
        self.assertEqual(['beta', 'alpha'], [item['project_name'] for item in projects_payload['projects']])

        compare_response = self.client.post(
            '/api/compare',
            json={'left_run_id': left_run_id, 'right_run_id': right_run_id},
        )
        self.assertEqual(200, compare_response.status_code)
        compare_payload = compare_response.get_json()
        self.assertTrue(compare_payload['ok'])
        self.assertEqual('alpha', compare_payload['left']['project_name'])
        self.assertEqual('beta', compare_payload['right']['project_name'])
        self.assertGreaterEqual(compare_payload['component_diff']['added_count'], 1)
        self.assertGreaterEqual(compare_payload['component_diff']['changed_count'], 1)
        self.assertTrue(any(row['位号'] == 'C2' and row['类型'] == '新增' for row in compare_payload['component_diff']['rows']))
        self.assertTrue(any(row['位号'] == 'U1' and row['类型'] == '变化' for row in compare_payload['component_diff']['rows']))
        self.assertEqual(0, compare_payload['key_component_diff']['removed_count'])
        self.assertTrue(any(row['位号'] == 'U1' and row['器件类别'] == '芯片' and row['左侧网络'] == 'SMBALERT_N' and row['右侧网络'] == 'SMBALERT_ALT_N' for row in compare_payload['key_pin_net_diff']['rows']))
        self.assertTrue(any(row['位号'] == 'U1' and row['左侧PI'] == 'PI-A' and row['右侧PI'] == 'PI-B' for row in compare_payload['key_pin_net_diff']['rows']))
        self.assertTrue(any(row['位号'] == 'R1' and row['器件类别'] == 'R/C/L' and row['左侧网络'] == 'SMBALERT_N' and row['右侧网络'] == 'SMBALERT_ALT_N' for row in compare_payload['passive_pin_net_diff']['rows']))
        self.assertIn('net_view_diff', compare_payload)
        self.assertTrue(any(
            row['类型'] == '网络迁移'
            and row['左侧网络'] == 'SMBALERT_N'
            and row['右侧网络'] == 'SMBALERT_ALT_N'
            and row['关键器件数'] >= 1
            and row['R/C/L数'] >= 1
            for row in compare_payload['net_view_diff']['rows']
        ))
        self.assertNotIn('feishu_hq_diff', compare_payload)
        component_u1 = next(row for row in compare_payload['component_diff']['rows'] if row['位号'] == 'U1')
        self.assertIn('PI', component_u1['变化字段'])
        self.assertIn('PI-A', component_u1['左侧'])
        self.assertIn('PI-B', component_u1['右侧'])
        self.assertGreaterEqual(compare_payload['net_diff']['added_count'], 1)
        self.assertGreater(compare_payload['diff_totals']['key_pin_nets'], 0)
        self.assertGreater(compare_payload['diff_totals']['net_view'], 0)
        self.assertGreater(compare_payload['diff_totals']['passive_pin_nets'], 0)
        self.assertNotIn('feishu_hq', compare_payload['diff_totals'])
        self.assertGreater(compare_payload['diff_totals']['components'], 0)
        self.assertEqual(500, compare_payload['detail_limit'])
        self.assertIn('compare_sections', compare_payload)
        section_ids = [section['id'] for section in compare_payload['compare_sections']]
        self.assertIn('net_view', section_ids)
        self.assertIn('key_components', section_ids)
        self.assertIn('key_pin_nets', section_ids)
        self.assertIn('passive_pin_nets', section_ids)
        self.assertIn('components', section_ids)
        self.assertIn('nets', section_ids)
        self.assertTrue(all('table' in section for section in compare_payload['compare_sections']))
        self.assertEqual('net', next(section for section in compare_payload['compare_sections'] if section['id'] == 'net_view')['group'])
        self.assertTrue(all(section['table'].get('default_density') == 'comfortable' for section in compare_payload['compare_sections']))

        limited_compare = self.client.post(
            '/api/compare',
            json={'left_run_id': left_run_id, 'right_run_id': right_run_id, 'detail_limit': 1},
        )
        self.assertEqual(200, limited_compare.status_code)
        limited_payload = limited_compare.get_json()
        self.assertEqual(1, limited_payload['detail_limit'])
        self.assertLessEqual(len(limited_payload['component_diff']['rows']), 1)

        invalid_limit = self.client.post(
            '/api/compare',
            json={'left_run_id': left_run_id, 'right_run_id': right_run_id, 'detail_limit': 'bad'},
        )
        self.assertEqual(400, invalid_limit.status_code)

        compare_profiles = self.client.get('/api/compare/harness/profiles')
        self.assertEqual(200, compare_profiles.status_code)
        compare_profile_payload = compare_profiles.get_json()
        self.assertTrue(compare_profile_payload['ok'])
        compare_profile_ids = [item['id'] for item in compare_profile_payload['profiles']]
        self.assertIn('auto', compare_profile_ids)
        self.assertIn('compare_full_review', compare_profile_ids)
        self.assertIn('compare_datasheet_qa', compare_profile_ids)
        compare_profile_map = {item['id']: item for item in compare_profile_payload['profiles']}
        self.assertIn('batch_query_compare_diff', compare_profile_map['compare_quick_scan']['tools'])
        self.assertIn('get_harness_skill', compare_profile_map['compare_quick_scan']['tools'])
        self.assertIn('batch_get_cadence_page_objects', compare_profile_map['compare_cadence_pages']['tools'])
        self.assertIn('list_datasheet_review_templates', compare_profile_map['compare_datasheet_qa']['tools'])
        self.assertIn('select_harness_skills', compare_profile_map['compare_datasheet_qa']['tools'])
        self.assertIn('batch_search_datasheet_chunks', compare_profile_map['compare_datasheet_qa']['tools'])
        self.assertIn('search_datasheet_parameters', compare_profile_map['compare_datasheet_qa']['tools'])
        self.assertIn('get_datasheet_chunk', compare_profile_map['compare_datasheet_qa']['tools'])
        self.assertIn('search_datasheet_chunks', compare_profile_map['compare_bom_feishu']['tools'])

        compare_harness_agent = self.client.post(
            '/api/compare/harness-agent',
            json={
                'left_run_id': left_run_id,
                'right_run_id': right_run_id,
                'question': '请重点看 U1 和 PI 变化',
                'max_steps': 8,
                'max_tool_calls': 14,
            },
        )
        self.assertEqual(200, compare_harness_agent.status_code)
        compare_harness_payload = compare_harness_agent.get_json()
        self.assertTrue(compare_harness_payload['ok'])
        self.assertEqual('local-compare-agent-harness', compare_harness_payload['mode'])
        self.assertTrue(compare_harness_payload['agent_run_id'])
        self.assertTrue(compare_harness_payload['trace_summary'])
        self.assertTrue(compare_harness_payload['citations'])
        self.assertIn(compare_harness_payload['agent_run_id'], webapp_state.AGENT_RUN_CACHE)

        async_compare_agent = self.client.post(
            '/api/compare/harness-agent',
            json={
                'left_run_id': left_run_id,
                'right_run_id': right_run_id,
                'question': '后台看 U1 和 PI 变化',
                'max_steps': 3,
                'max_tool_calls': 4,
                'async': True,
            },
        )
        self.assertEqual(202, async_compare_agent.status_code)
        async_payload = async_compare_agent.get_json()
        self.assertTrue(async_payload['agent_run_id'].startswith('compare_'))
        status_payload = {}
        for _ in range(80):
            status = self.client.get(f"/api/harness/agent-runs/{async_payload['agent_run_id']}")
            self.assertEqual(200, status.status_code)
            status_payload = status.get_json()
            if status_payload['status'] in {'completed', 'waiting_for_user', 'failed'}:
                break
            time.sleep(0.05)
        self.assertEqual('completed', status_payload['status'])
        self.assertEqual(async_payload['agent_run_id'], status_payload['agent_run']['agent_run_id'])
        self.assertIn('current_phase', status_payload)
        self.assertIn('progress', status_payload)

        webapp_state.AGENT_DURABLE_STORE.update_record(
            async_payload['agent_run_id'],
            status='incomplete',
            current_phase='incomplete',
            result={},
            error='test resume compare',
        )
        compare_continue = self.client.post(
            f"/api/harness/agent-runs/{async_payload['agent_run_id']}/continue",
            json={'question': '继续检查 U1 和 PI 变化'},
        )
        self.assertEqual(202, compare_continue.status_code)
        continued_payload = {}
        for _ in range(80):
            status = self.client.get(f"/api/harness/agent-runs/{async_payload['agent_run_id']}")
            self.assertEqual(200, status.status_code)
            continued_payload = status.get_json()
            if continued_payload['status'] in {'completed', 'waiting_for_user', 'failed'}:
                break
            time.sleep(0.05)
        self.assertEqual('completed', continued_payload['status'])
        self.assertEqual(async_payload['agent_run_id'], continued_payload['agent_run']['agent_run_id'])

        bad_compare = self.client.post(
            '/api/compare',
            json={'left_run_id': left_run_id, 'right_run_id': left_run_id},
        )
        self.assertEqual(400, bad_compare.status_code)

        missing_compare = self.client.post(
            '/api/compare',
            json={'left_run_id': left_run_id, 'right_run_id': 'missing-run'},
        )
        self.assertEqual(404, missing_compare.status_code)
        self.assertFalse(missing_compare.get_json()['ok'])

        bad_compare_agent = self.client.post(
            '/api/compare/harness-agent',
            json={'left_run_id': left_run_id, 'right_run_id': right_run_id, 'profile': 'unsafe'},
        )
        self.assertEqual(400, bad_compare_agent.status_code)
        self.assertIn('未知 compare agent profile', bad_compare_agent.get_json()['error'])

    def test_compare_agent_async_dispatch_creates_child_runs(self):
        class CompareDispatchProvider:
            provider = 'compare-dispatch-mock'
            mode = 'mock'

            def generate_agent_step(self, prompt, *, inputs=None):
                return HarnessModelResponse(
                    answer=json.dumps({
                        'dispatch_tasks': [{
                            'task_id': 'cmp-u1',
                            'title': 'U1 datasheet compare',
                            'profile': 'compare_datasheet_qa',
                            'question': '对比 U1 datasheet 关键参数。',
                        }],
                        'reason': 'datasheet 对比分支适合后台执行。',
                    }, ensure_ascii=False),
                    provider=self.provider,
                    mode=self.mode,
                )

        self.with_env({'PSTX_ASTER_MODE': 'mock'})
        webapp_state.clear_web_session_state()
        with mock.patch.object(webapp_factory, 'CompareMockModelProvider', CompareDispatchProvider):
            app = webapp_factory.create_app()
            app.testing = True
            client = app.test_client()
            left_root = self.make_root_with_samples(PRT_SAMPLE, NET_SAMPLE)
            right_root = self.make_root_with_samples(PRT_SAMPLE_COMPARE, NET_SAMPLE_COMPARE)
            left_run_id = client.post(
                '/api/analyze',
                data={'project_root': str(left_root), 'project_name': 'left-dispatch', 'ratio_limit': '70', 'custom_volt_map': ''},
            ).get_json()['run_id']
            right_run_id = client.post(
                '/api/analyze',
                data={'project_root': str(right_root), 'project_name': 'right-dispatch', 'ratio_limit': '70', 'custom_volt_map': ''},
            ).get_json()['run_id']

            response = client.post(
                '/api/compare/harness-agent',
                json={
                    'left_run_id': left_run_id,
                    'right_run_id': right_run_id,
                    'question': '请后台拆分 datasheet 对比',
                    'max_steps': 1,
                    'max_tool_calls': 1,
                    'async': True,
                },
            )
            self.assertEqual(202, response.status_code)
            parent_run_id = response.get_json()['agent_run_id']
            status_payload = {}
            for _ in range(80):
                status = client.get(f'/api/harness/agent-runs/{parent_run_id}')
                self.assertEqual(200, status.status_code)
                status_payload = status.get_json()
                if status_payload['status'] in {'completed', 'waiting_for_user', 'failed'}:
                    break
                time.sleep(0.05)

            self.assertEqual('completed', status_payload['status'])
            self.assertEqual(['cmp-u1'], [item['task_id'] for item in status_payload['dispatch_tasks']])
            self.assertEqual(1, len(status_payload['child_agent_run_ids']))
            child_run_id = status_payload['child_agent_run_ids'][0]
            child_status = client.get(f'/api/harness/agent-runs/{child_run_id}').get_json()
            self.assertTrue(child_status['ok'])
            self.assertEqual(parent_run_id, child_status['parent_agent_run_id'])
            self.assertEqual('cmp-u1', child_status['dispatch_task']['task_id'])

    def test_web_report_includes_csa_geometry_review_section(self):
        root = self.make_root()
        (root / 'sch_1' / 'page3.csa').write_text(CSA_DOT_CROSS_SAMPLE, encoding='utf-8')
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'csa-demo',
                'ratio_limit': '70',
                'custom_volt_map': 'P3V3=3.3',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']
        report_payload = self.client.get(f'/api/report/{run_id}').get_json()
        self.assertEqual('pstx-analysis-timings.v1', report_payload['analysis_timings']['schema_version'])
        self.assertTrue(any(row['stage'] == 'report_payload' for row in report_payload['analysis_timings']['stages']))

        csa_section = next(section for section in report_payload['sections'] if section['id'] == 'csa')
        self.assertEqual('规范检查', csa_section['title'])
        cross_table = next(table for table in csa_section['tables'] if table['id'] == 'csa_dot_cross_rows')
        connectivity_table = next(table for table in csa_section['tables'] if table['id'] == 'cadence_connectivity_rows')
        self.assertFalse(any(table['id'] == 'csa_circle_rows' for table in csa_section['tables']))
        self.assertEqual(1, cross_table['count'])
        self.assertEqual('(450,0)', cross_table['rows'][0]['坐标'])
        self.assertEqual(2, connectivity_table['count'])
        page3_connectivity_row = next(row for row in connectivity_table['rows'] if row['页码'] == 'PAGE3')
        self.assertEqual('PAGE3', page3_connectivity_row['页码'])
        self.assertIn('网络标签', connectivity_table['columns'])
        self.assertTrue(any(metric['label'] == '规范候选' and metric['value'] == 1 for metric in report_payload['metrics']))

        csa_response = self.client.get(f'/api/report/{run_id}/csa-geometry?stdout=full&limit=1')
        self.assertEqual(200, csa_response.status_code)
        csa_payload = csa_response.get_json()
        self.assertTrue(csa_payload['ok'])
        self.assertEqual('pstx-csa-geometry.v1', csa_payload['csa_geometry']['schema_version'])
        self.assertEqual(1, csa_payload['csa_geometry']['digest']['cross_count'])
        self.assertEqual(1, len(csa_payload['csa_geometry']['dot_cross_rows']))

        overlay_response = self.client.get(
            f'/api/report/{run_id}/csa-geometry?stdout=full&include_connectivity=1&page=3'
        )
        self.assertEqual(200, overlay_response.status_code)
        overlay_payload = overlay_response.get_json()['csa_geometry']
        self.assertEqual(3, overlay_payload['digest']['page_filter'])
        semantic_overlay = overlay_payload['semantic_overlay']
        self.assertEqual('pstx-csa-connectivity-overlay.v1', semantic_overlay['schema_version'])
        self.assertEqual(1, semantic_overlay['digest']['dot_cross_matched_count'])
        self.assertIn('CROSS_DOT_H', semantic_overlay['dot_cross_overlay_rows'][0]['signal_names'])

        bad_csa = self.client.get(f'/api/report/{run_id}/csa-geometry?limit=bad')
        self.assertEqual(400, bad_csa.status_code)
        bad_page = self.client.get(f'/api/report/{run_id}/csa-geometry?page=bad')
        self.assertEqual(400, bad_page.status_code)

    def test_web_report_links_bom_option_components_to_csa_circle_marks(self):
        root = self.make_root_with_samples(PRT_SAMPLE_DEPOP_WITH_XY, NET_SAMPLE_DEPOP)
        (root / 'sch_1' / 'page518.csa').write_text(
            "FILE_TYPE = MACRO_DRAWING;\n"
            "CIRCLE 16 -1 (1000 1000)(1100 1000);\n",
            encoding='utf-8',
        )
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'bom-circle-demo',
                'ratio_limit': '70',
                'custom_volt_map': 'P3V3=3.3',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']
        report_payload = self.client.get(f'/api/report/{run_id}').get_json()

        bom_section = next(section for section in report_payload['sections'] if section['id'] == 'bom')
        drc_section = next(section for section in report_payload['sections'] if section['id'] == 'drc')
        coverage_table = next(table for table in bom_section['tables'] if table['id'] == 'bom_option_circle_coverage')
        issue_table = next(table for table in drc_section['tables'] if table['id'] == 'bom_option_circle_issues')
        self.assertEqual(1, coverage_table['count'])
        self.assertEqual('R1', coverage_table['rows'][0]['位号'])
        self.assertEqual('已打圈', coverage_table['rows'][0]['覆盖状态'])
        self.assertEqual('100%', coverage_table['rows'][0]['中心重合度'])
        self.assertEqual(0, issue_table['count'])
        self.assertTrue(any(metric['label'] == 'BOM圈问题' and metric['value'] == 0 for metric in report_payload['metrics']))

    def test_web_analysis_excludes_depop_by_default_but_keeps_detection_list(self):
        root = self.make_root_with_samples(PRT_SAMPLE_DEPOP, NET_SAMPLE_DEPOP)
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'demo-depop',
                'ratio_limit': '70',
                'custom_volt_map': 'P3V3=3.3',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        report_payload = self.client.get(f'/api/report/{run_id}').get_json()
        self.assertFalse(report_payload['include_depop'])
        self.assertEqual(1, report_payload['excluded_depop_count'])
        self.assertTrue(any('DEPOP 排查：关闭' in line for line in report_payload['summary_lines']))

        bom_section = next(section for section in report_payload['sections'] if section['id'] == 'bom')
        drc_section = next(section for section in report_payload['sections'] if section['id'] == 'drc')
        self.assertFalse(any(table['id'] == 'bom_option_components' for table in drc_section['tables']))
        bom_option_table = next(table for table in bom_section['tables'] if table['id'] == 'bom_option_components')
        self.assertEqual('R1', bom_option_table['rows'][0]['位号'])
        self.assertEqual('是', bom_option_table['rows'][0]['是否DEPOP'])

        query_payload = self.client.post(
            f'/api/report/{run_id}/query',
            json={'mode': '位号', 'keyword': 'R1'},
        ).get_json()
        self.assertEqual('missing', query_payload['match_type'])

    def test_web_analysis_can_include_depop_when_switch_is_on(self):
        root = self.make_root_with_samples(PRT_SAMPLE_DEPOP, NET_SAMPLE_DEPOP)
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'demo-depop-on',
                'ratio_limit': '70',
                'custom_volt_map': 'P3V3=3.3',
                'include_depop': 'on',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        report_payload = self.client.get(f'/api/report/{run_id}').get_json()
        self.assertTrue(report_payload['include_depop'])
        self.assertEqual(1, report_payload['depop_count'])
        self.assertTrue(any('DEPOP 排查：开启' in line for line in report_payload['summary_lines']))

        query_payload = self.client.post(
            f'/api/report/{run_id}/query',
            json={'mode': '位号', 'keyword': 'R1'},
        ).get_json()
        self.assertEqual('exact', query_payload['match_type'])


    def _archived_test_web_report_shows_real_and_user_visible_pages_from_p_path(self):
        root = self.make_page_v2_root()
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'page-v2',
                'ratio_limit': '70',
                'include_depop': 'on',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        report_payload = self.client.get(f'/api/report/{run_id}').get_json()
        bom_section = next(section for section in report_payload['sections'] if section['id'] == 'bom')
        bom_option_table = next(table for table in bom_section['tables'] if table['id'] == 'bom_option_components')
        self.assertEqual('PAGE114', bom_option_table['rows'][0]['页面'])
        self.assertEqual('PAGE177', bom_option_table['rows'][0]['页码'])

        network_section = next(section for section in report_payload['sections'] if section['id'] == 'network')
        mapping_table = next(table for table in network_section['tables'] if table['id'] == 'page_mapping_rows')
        self.assertEqual('PAGE144', mapping_table['rows'][0]['主模块页'])
        self.assertEqual('PAGE114', mapping_table['rows'][0]['页码'])

        query_payload = self.client.post(
            f'/api/report/{run_id}/query',
            json={'mode': '浣嶅彿', 'keyword': 'C1A104'},
        ).get_json()
        meta_map = {item['label']: item['value'] for item in query_payload['summary']['meta']}
        self.assertEqual('PAGE177', meta_map['页码'])

    def _archived_test_web_report_page_v2_query_meta_contains_real_and_user_visible_pages(self):
        root = self.make_page_v2_root()
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'page-v2-query',
                'ratio_limit': '70',
                'include_depop': 'on',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        query_payload = self.client.post(
            f'/api/report/{run_id}/query',
            json={'mode': '浣嶅彿', 'keyword': 'C1A104'},
        ).get_json()
        self.assertEqual('exact', query_payload['match_type'])
        meta_values = [item['value'] for item in query_payload['summary']['meta']]
        self.assertIn('PAGE114', meta_values)
        self.assertIn('PAGE177', meta_values)

    def test_web_report_page_v2_tables_show_user_visible_page(self):
        root = self.make_page_v2_root()
        response = self.client.post(
            '/api/analyze',
            data={
                'project_root': str(root),
                'project_name': 'page-v2-report',
                'ratio_limit': '70',
            },
        )
        self.assertEqual(200, response.status_code)
        run_id = response.get_json()['run_id']

        report_payload = self.client.get(f'/api/report/{run_id}').get_json()
        bom_section = next(section for section in report_payload['sections'] if section['id'] == 'bom')
        bom_option_table = next(table for table in bom_section['tables'] if table['id'] == 'bom_option_components')
        self.assertEqual('PAGE177', bom_option_table['rows'][0]['页面'])
        self.assertEqual('PAGE177', bom_option_table['rows'][0]['页码'])

        network_section = next(section for section in report_payload['sections'] if section['id'] == 'network')
        mapping_table = next(table for table in network_section['tables'] if table['id'] == 'page_mapping_rows')
        self.assertEqual('PAGE144', mapping_table['rows'][0]['主模块页'])
        self.assertEqual('PAGE114', mapping_table['rows'][0]['页码'])

if __name__ == '__main__':
    unittest.main()
