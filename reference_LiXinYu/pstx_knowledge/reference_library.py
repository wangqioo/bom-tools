# -*- coding: utf-8 -*-
"""Local reference indexes for the Agent capability lab."""

from __future__ import annotations

import csv
import os
import sqlite3
import time
import zipfile
from pathlib import Path
from typing import Iterable, List
from xml.etree import ElementTree as ET

from pstx_knowledge.datasheets import _extract_pdf_pages, _safe_text, _snippet, _terms


AGENT_REF_DIR_ENV = "PSTX_AGENT_REF_DIR"
AGENT_REF_DATA_DIR_ENV = "PSTX_AGENT_REF_DATA_DIR"
AGENT_CHECKLIST_REF_DIR_ENV = "PSTX_AGENT_CHECKLIST_REF_DIR"
AGENT_CHECKLIST_DATA_DIR_ENV = "PSTX_AGENT_CHECKLIST_DATA_DIR"
DEFAULT_AGENT_REF_DIR = "ref"
DEFAULT_AGENT_REF_DATA_DIR = "agent_ref_data"
DEFAULT_AGENT_CHECKLIST_REF_DIR = "ref_checklist"
DEFAULT_AGENT_CHECKLIST_DATA_DIR = "agent_checklist_data"
DB_NAME = "agent_ref_index.db"
CHECKLIST_DB_NAME = "review_checklist_index.db"
CHECKLIST_SUFFIXES = {".pdf", ".txt", ".md", ".markdown", ".csv", ".tsv", ".xlsx", ".xlsm", ".docx", ".xls"}
TEXT_SUFFIXES = {".txt", ".md", ".markdown"}
CSV_SUFFIXES = {".csv", ".tsv"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


def _now() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())


def agent_ref_dir() -> Path:
    raw = os.environ.get(AGENT_REF_DIR_ENV) or DEFAULT_AGENT_REF_DIR
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    return path


def agent_checklist_ref_dir() -> Path:
    raw = os.environ.get(AGENT_CHECKLIST_REF_DIR_ENV) or DEFAULT_AGENT_CHECKLIST_REF_DIR
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    return path


def _data_dir() -> Path:
    raw = os.environ.get(AGENT_REF_DATA_DIR_ENV) or DEFAULT_AGENT_REF_DATA_DIR
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    path.mkdir(parents=True, exist_ok=True)
    return path


def _checklist_data_dir() -> Path:
    raw = os.environ.get(AGENT_CHECKLIST_DATA_DIR_ENV) or DEFAULT_AGENT_CHECKLIST_DATA_DIR
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    path.mkdir(parents=True, exist_ok=True)
    return path


def agent_ref_db_path() -> Path:
    return _data_dir() / DB_NAME


def agent_checklist_db_path() -> Path:
    return _checklist_data_dir() / CHECKLIST_DB_NAME


def _connect_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    _init_schema(conn)
    return conn


def _connect() -> sqlite3.Connection:
    return _connect_db(agent_ref_db_path())


def _connect_checklist() -> sqlite3.Connection:
    return _connect_db(agent_checklist_db_path())


def _init_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            path TEXT NOT NULL UNIQUE,
            rel_path TEXT NOT NULL DEFAULT '',
            title TEXT NOT NULL,
            size INTEGER NOT NULL DEFAULT 0,
            mtime REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'pending',
            page_count INTEGER NOT NULL DEFAULT 0,
            extractor TEXT NOT NULL DEFAULT '',
            error TEXT NOT NULL DEFAULT '',
            indexed_at TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS pages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id INTEGER NOT NULL,
            page INTEGER NOT NULL,
            text TEXT NOT NULL DEFAULT '',
            FOREIGN KEY(doc_id) REFERENCES documents(id) ON DELETE CASCADE,
            UNIQUE(doc_id, page)
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_agent_ref_documents_status ON documents(status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_agent_ref_pages_doc_page ON pages(doc_id, page)")
    conn.commit()


def _pdf_files(root: Path) -> List[Path]:
    if not root.is_dir():
        return []
    files = [
        file
        for file in root.rglob("*")
        if file.is_file() and not file.is_symlink() and file.suffix.lower() == ".pdf"
    ]
    return sorted(set(files), key=lambda item: item.as_posix().lower())


def _checklist_files(root: Path) -> List[Path]:
    if not root.is_dir():
        return []
    files = [
        file
        for file in root.rglob("*")
        if file.is_file() and not file.is_symlink() and file.suffix.lower() in CHECKLIST_SUFFIXES
    ]
    return sorted(set(files), key=lambda item: item.as_posix().lower())


def _rel_path(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.name


def _path_in_root(path: object, root: Path) -> bool:
    try:
        candidate = Path(str(path or "")).expanduser().resolve(strict=False)
        resolved_root = root.expanduser().resolve(strict=False)
        candidate.relative_to(resolved_root)
        return True
    except (OSError, ValueError):
        return False


def _decode_text_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _normalize_reference_text(text: str) -> str:
    return "\n".join(line.rstrip() for line in str(text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")).strip()


def _chunk_text(text: str, *, max_chars: int = 4500) -> List[str]:
    normalized = _normalize_reference_text(text)
    if not normalized:
        return []
    chunks: List[str] = []
    lines = normalized.splitlines()
    current: List[str] = []
    current_chars = 0
    for line in lines:
        line_chars = len(line) + 1
        if current and current_chars + line_chars > max_chars:
            chunks.append("\n".join(current).strip())
            current = []
            current_chars = 0
        if line_chars > max_chars:
            if current:
                chunks.append("\n".join(current).strip())
                current = []
                current_chars = 0
            for index in range(0, len(line), max_chars):
                chunks.append(line[index:index + max_chars])
            continue
        current.append(line)
        current_chars += line_chars
    if current:
        chunks.append("\n".join(current).strip())
    return [chunk for chunk in chunks if chunk]


def _extract_text_file_pages(path: Path) -> tuple[str, List[str], str, str]:
    text = _decode_text_bytes(path.read_bytes())
    chunks = _chunk_text(text)
    if not chunks:
        return "needs_manual_review", [], "text", "文本文件为空或未抽取到可检索内容。"
    return "indexed", chunks, "text", ""


def _extract_csv_pages(path: Path) -> tuple[str, List[str], str, str]:
    text = _decode_text_bytes(path.read_bytes())
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    lines = []
    for row_index, row in enumerate(rows, start=1):
        values = [_safe_text(cell, 240) for cell in row]
        lines.append(f"row {row_index}: " + " | ".join(values))
    chunks = _chunk_text("\n".join(lines))
    if not chunks:
        return "needs_manual_review", [], "csv", "CSV/TSV 文件为空或未抽取到可检索内容。"
    return "indexed", chunks, "csv", ""


def _extract_excel_pages(path: Path) -> tuple[str, List[str], str, str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        return "needs_manual_review", [], "openpyxl", f"未安装 openpyxl，无法读取 Excel：{exc}"
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return "needs_manual_review", [], "openpyxl", f"Excel 读取失败：{exc}"
    pages: List[str] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            lines = [f"Sheet: {sheet_name}"]
            chunk_index = 1
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [_safe_text(cell, 240) for cell in row]
                if not any(values):
                    continue
                lines.append(f"row {row_index}: " + " | ".join(values))
                if len("\n".join(lines)) >= 4200:
                    pages.append(f"{sheet_name} chunk {chunk_index}\n" + "\n".join(lines))
                    chunk_index += 1
                    lines = [f"Sheet: {sheet_name}"]
            if len(lines) > 1:
                pages.append(f"{sheet_name} chunk {chunk_index}\n" + "\n".join(lines))
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    if not pages:
        return "needs_manual_review", [], "openpyxl", "Excel 未抽取到非空单元格。"
    return "indexed", pages, "openpyxl", ""


def _extract_docx_pages(path: Path) -> tuple[str, List[str], str, str]:
    try:
        with zipfile.ZipFile(path) as archive:
            names = [
                name for name in archive.namelist()
                if name.startswith("word/") and name.endswith(".xml") and ("document" in name or "footnotes" in name or "endnotes" in name)
            ]
            texts: List[str] = []
            for name in names:
                root = ET.fromstring(archive.read(name))
                for node in root.iter():
                    if node.tag.endswith("}t") and node.text:
                        texts.append(node.text)
                    elif node.tag.endswith("}tab"):
                        texts.append("\t")
                    elif node.tag.endswith("}br") or node.tag.endswith("}p"):
                        texts.append("\n")
    except Exception as exc:
        return "needs_manual_review", [], "docx", f"DOCX 读取失败：{exc}"
    chunks = _chunk_text("".join(texts))
    if not chunks:
        return "needs_manual_review", [], "docx", "DOCX 未抽取到正文文本。"
    return "indexed", chunks, "docx", ""


def _extract_checklist_pages(path: Path) -> tuple[str, List[str], str, str]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf_pages(path)
    if suffix in TEXT_SUFFIXES:
        return _extract_text_file_pages(path)
    if suffix in CSV_SUFFIXES:
        return _extract_csv_pages(path)
    if suffix in EXCEL_SUFFIXES:
        return _extract_excel_pages(path)
    if suffix == ".docx":
        return _extract_docx_pages(path)
    if suffix == ".xls":
        return "needs_manual_review", [], "unsupported", "旧版 .xls 暂不解析，请转换为 .xlsx 后放入 ref_checklist。"
    return "needs_manual_review", [], "unsupported", f"暂不支持的 checklist 文件类型：{suffix}"


def reindex_agent_ref(*, force: bool = False, max_files: int = 1000) -> dict:
    root = agent_ref_dir()
    all_files = _pdf_files(root)
    files = all_files[: max(1, int(max_files or 1000))]
    indexed_count = 0
    skipped_count = 0
    failed_count = 0
    with _connect() as conn:
        current_paths = {str(path) for path in all_files}
        stale_rows = conn.execute("SELECT id,path FROM documents").fetchall()
        for row in stale_rows:
            if row["path"] in current_paths:
                continue
            conn.execute("DELETE FROM pages WHERE doc_id=?", (int(row["id"]),))
            conn.execute("DELETE FROM documents WHERE id=?", (int(row["id"]),))
        for path in files:
            stat = path.stat()
            existing = conn.execute("SELECT * FROM documents WHERE path=?", (str(path),)).fetchone()
            if existing and not force and int(existing["size"]) == stat.st_size and float(existing["mtime"]) == stat.st_mtime:
                skipped_count += 1
                continue
            status, pages, extractor, error = _extract_pdf_pages(path)
            if status != "indexed":
                failed_count += 1
            conn.execute(
                """
                INSERT INTO documents(path,rel_path,title,size,mtime,status,page_count,extractor,error,indexed_at)
                VALUES(?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(path) DO UPDATE SET
                    rel_path=excluded.rel_path,
                    title=excluded.title,
                    size=excluded.size,
                    mtime=excluded.mtime,
                    status=excluded.status,
                    page_count=excluded.page_count,
                    extractor=excluded.extractor,
                    error=excluded.error,
                    indexed_at=excluded.indexed_at
                """,
                (
                    str(path),
                    _rel_path(path, root),
                    path.name,
                    stat.st_size,
                    stat.st_mtime,
                    status,
                    len(pages),
                    extractor,
                    _safe_text(error, 2000),
                    _now(),
                ),
            )
            doc_id = int(conn.execute("SELECT id FROM documents WHERE path=?", (str(path),)).fetchone()["id"])
            conn.execute("DELETE FROM pages WHERE doc_id=?", (doc_id,))
            for page_index, text in enumerate(pages, start=1):
                conn.execute(
                    "INSERT OR REPLACE INTO pages(doc_id,page,text) VALUES(?,?,?)",
                    (doc_id, page_index, text or ""),
                )
            indexed_count += 1
        conn.commit()
    return {
        "ok": True,
        "ref_dir": str(root),
        "db_path": str(agent_ref_db_path()),
        "pdf_count": len(all_files),
        "indexed_count": indexed_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
        "summary": f"扫描 ref PDF {len(all_files)} 个，更新 {indexed_count} 个，跳过 {skipped_count} 个，失败/需人工 {failed_count} 个。",
    }


def reindex_review_checklists(*, force: bool = False, max_files: int = 1000) -> dict:
    root = agent_checklist_ref_dir()
    all_files = _checklist_files(root)
    files = all_files[: max(1, int(max_files or 1000))]
    indexed_count = 0
    skipped_count = 0
    failed_count = 0
    with _connect_checklist() as conn:
        current_paths = {str(path) for path in all_files}
        stale_rows = conn.execute("SELECT id,path FROM documents").fetchall()
        for row in stale_rows:
            if row["path"] in current_paths:
                continue
            conn.execute("DELETE FROM pages WHERE doc_id=?", (int(row["id"]),))
            conn.execute("DELETE FROM documents WHERE id=?", (int(row["id"]),))
        for path in files:
            stat = path.stat()
            existing = conn.execute("SELECT * FROM documents WHERE path=?", (str(path),)).fetchone()
            if existing and not force and int(existing["size"]) == stat.st_size and float(existing["mtime"]) == stat.st_mtime:
                skipped_count += 1
                continue
            status, pages, extractor, error = _extract_checklist_pages(path)
            if status != "indexed":
                failed_count += 1
            conn.execute(
                """
                INSERT INTO documents(path,rel_path,title,size,mtime,status,page_count,extractor,error,indexed_at)
                VALUES(?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(path) DO UPDATE SET
                    rel_path=excluded.rel_path,
                    title=excluded.title,
                    size=excluded.size,
                    mtime=excluded.mtime,
                    status=excluded.status,
                    page_count=excluded.page_count,
                    extractor=excluded.extractor,
                    error=excluded.error,
                    indexed_at=excluded.indexed_at
                """,
                (
                    str(path),
                    _rel_path(path, root),
                    path.name,
                    stat.st_size,
                    stat.st_mtime,
                    status,
                    len(pages),
                    extractor,
                    _safe_text(error, 2000),
                    _now(),
                ),
            )
            doc_id = int(conn.execute("SELECT id FROM documents WHERE path=?", (str(path),)).fetchone()["id"])
            conn.execute("DELETE FROM pages WHERE doc_id=?", (doc_id,))
            for page_index, text in enumerate(pages, start=1):
                conn.execute(
                    "INSERT OR REPLACE INTO pages(doc_id,page,text) VALUES(?,?,?)",
                    (doc_id, page_index, text or ""),
                )
            indexed_count += 1
        conn.commit()
    return {
        "ok": True,
        "ref_dir": str(root),
        "db_path": str(agent_checklist_db_path()),
        "file_count": len(all_files),
        "indexed_count": indexed_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
        "supported_suffixes": sorted(CHECKLIST_SUFFIXES),
        "summary": f"扫描 ref_checklist 文件 {len(all_files)} 个，更新 {indexed_count} 个，跳过 {skipped_count} 个，失败/需人工 {failed_count} 个。",
    }


def build_agent_ref_status() -> dict:
    root = agent_ref_dir()
    pdf_count = len(_pdf_files(root))
    with _connect() as conn:
        doc_count = int(conn.execute("SELECT COUNT(*) FROM documents").fetchone()[0])
        page_count = int(conn.execute("SELECT COUNT(*) FROM pages").fetchone()[0])
        indexed_count = int(conn.execute("SELECT COUNT(*) FROM documents WHERE status='indexed'").fetchone()[0])
        failed_count = int(conn.execute("SELECT COUNT(*) FROM documents WHERE status!='indexed'").fetchone()[0])
        last_row = conn.execute("SELECT indexed_at FROM documents ORDER BY indexed_at DESC LIMIT 1").fetchone()
        docs = [
            {
                "doc_id": int(row["id"]),
                "title": row["title"],
                "rel_path": row["rel_path"],
                "status": row["status"],
                "page_count": int(row["page_count"] or 0),
                "error": row["error"],
            }
            for row in conn.execute(
                "SELECT id,title,rel_path,status,page_count,error FROM documents ORDER BY title LIMIT 40"
            )
        ]
    return {
        "ok": True,
        "configured": True,
        "ref_dir": str(root),
        "db_path": str(agent_ref_db_path()),
        "pdf_count": pdf_count,
        "document_count": doc_count,
        "indexed_count": indexed_count,
        "page_count": page_count,
        "failed_count": failed_count,
        "last_indexed_at": str(last_row["indexed_at"]) if last_row else "",
        "documents": docs,
        "summary": f"ref 目录包含 {pdf_count} 个 PDF，索引中有 {indexed_count} 个可检索文档。",
    }


def build_review_checklist_status() -> dict:
    root = agent_checklist_ref_dir()
    file_count = len(_checklist_files(root))
    with _connect_checklist() as conn:
        doc_count = int(conn.execute("SELECT COUNT(*) FROM documents").fetchone()[0])
        page_count = int(conn.execute("SELECT COUNT(*) FROM pages").fetchone()[0])
        indexed_count = int(conn.execute("SELECT COUNT(*) FROM documents WHERE status='indexed'").fetchone()[0])
        failed_count = int(conn.execute("SELECT COUNT(*) FROM documents WHERE status!='indexed'").fetchone()[0])
        last_row = conn.execute("SELECT indexed_at FROM documents ORDER BY indexed_at DESC LIMIT 1").fetchone()
        docs = [
            {
                "doc_id": int(row["id"]),
                "title": row["title"],
                "rel_path": row["rel_path"],
                "status": row["status"],
                "page_count": int(row["page_count"] or 0),
                "extractor": row["extractor"],
                "error": row["error"],
            }
            for row in conn.execute(
                "SELECT id,title,rel_path,status,page_count,extractor,error FROM documents ORDER BY title LIMIT 40"
            )
        ]
    return {
        "ok": True,
        "configured": True,
        "ref_dir": str(root),
        "db_path": str(agent_checklist_db_path()),
        "file_count": file_count,
        "document_count": doc_count,
        "indexed_count": indexed_count,
        "page_count": page_count,
        "failed_count": failed_count,
        "last_indexed_at": str(last_row["indexed_at"]) if last_row else "",
        "supported_suffixes": sorted(CHECKLIST_SUFFIXES),
        "documents": docs,
        "summary": f"ref_checklist 目录包含 {file_count} 个文件，索引中有 {indexed_count} 个可检索文档。",
    }


def search_agent_ref(query: str, *, limit: int = 20, offset: int = 0) -> dict:
    terms = _terms(query)
    if not terms:
        return {"ok": False, "error": "search_agent_ref 需要 query。", "matches": []}
    with _connect() as conn:
        rows = conn.execute(
            """
            SELECT d.id AS doc_id,d.title,d.rel_path,d.path,d.status,d.error,p.page,p.text
            FROM documents d
            LEFT JOIN pages p ON p.doc_id=d.id
            WHERE d.status='indexed'
            ORDER BY d.title,p.page
            """
        ).fetchall()
    root = agent_ref_dir()
    matches = []
    for row in rows:
        if not _path_in_root(row["path"], root):
            continue
        haystack = f"{row['title']} {row['rel_path']} {row['text'] or ''}".lower()
        matched_terms = [term for term in terms if term.lower() in haystack]
        if not matched_terms:
            continue
        score = len(matched_terms)
        title_text = f"{row['title']} {row['rel_path']}".lower()
        if any(term.lower() in title_text for term in terms):
            score += 2
        matches.append({
            "doc_id": int(row["doc_id"]),
            "title": row["title"],
            "rel_path": row["rel_path"],
            "path": row["path"],
            "status": row["status"],
            "page": int(row["page"] or 1),
            "score": score,
            "matched_terms": matched_terms,
            "snippet": _snippet(row["text"] or row["title"], matched_terms),
        })
    matches.sort(key=lambda item: (-int(item["score"]), item["title"].lower(), int(item["page"])))
    start = max(0, int(offset or 0))
    limit = max(1, min(int(limit or 20), 100))
    return {
        "ok": True,
        "query": query,
        "terms": terms,
        "total_matches": len(matches),
        "limit": limit,
        "offset": start,
        "matches": matches[start:start + limit],
        "summary": f"ref PDF 检索 `{query}` 命中 {len(matches)} 个页级片段。",
    }


def _search_index(conn: sqlite3.Connection,
                  query: str,
                  *,
                  limit: int,
                  offset: int,
                  label: str,
                  root: Path) -> dict:
    terms = _terms(query)
    if not terms:
        return {"ok": False, "error": f"{label} 检索需要 query。", "matches": []}
    rows = conn.execute(
        """
        SELECT d.id AS doc_id,d.title,d.rel_path,d.path,d.status,d.error,p.page,p.text
        FROM documents d
        LEFT JOIN pages p ON p.doc_id=d.id
        WHERE d.status='indexed'
        ORDER BY d.title,p.page
        """
    ).fetchall()
    matches = []
    for row in rows:
        if not _path_in_root(row["path"], root):
            continue
        haystack = f"{row['title']} {row['rel_path']} {row['text'] or ''}".lower()
        matched_terms = [term for term in terms if term.lower() in haystack]
        if not matched_terms:
            continue
        score = len(matched_terms)
        title_text = f"{row['title']} {row['rel_path']}".lower()
        if any(term.lower() in title_text for term in terms):
            score += 2
        matches.append({
            "doc_id": int(row["doc_id"]),
            "title": row["title"],
            "rel_path": row["rel_path"],
            "path": row["path"],
            "status": row["status"],
            "page": int(row["page"] or 1),
            "score": score,
            "matched_terms": matched_terms,
            "snippet": _snippet(row["text"] or row["title"], matched_terms),
        })
    matches.sort(key=lambda item: (-int(item["score"]), item["title"].lower(), int(item["page"])))
    start = max(0, int(offset or 0))
    safe_limit = max(1, min(int(limit or 20), 100))
    return {
        "ok": True,
        "query": query,
        "terms": terms,
        "total_matches": len(matches),
        "limit": safe_limit,
        "offset": start,
        "matches": matches[start:start + safe_limit],
        "summary": f"{label} 检索 `{query}` 命中 {len(matches)} 个片段。",
    }


def search_review_checklists(query: str, *, limit: int = 20, offset: int = 0) -> dict:
    with _connect_checklist() as conn:
        return _search_index(
            conn,
            query,
            limit=limit,
            offset=offset,
            label="review checklist",
            root=agent_checklist_ref_dir(),
        )


def get_agent_ref_excerpt(doc_id: int, page: int, *, max_chars: int = 2400) -> dict:
    with _connect() as conn:
        row = conn.execute(
            """
            SELECT d.id AS doc_id,d.title,d.rel_path,d.path,d.status,d.error,p.page,p.text
            FROM documents d
            LEFT JOIN pages p ON p.doc_id=d.id AND p.page=?
            WHERE d.id=?
            """,
            (int(page), int(doc_id)),
        ).fetchone()
    if not row:
        return {"ok": False, "error": f"未找到 ref PDF doc_id={doc_id} page={page}。"}
    if not _path_in_root(row["path"], agent_ref_dir()):
        return {"ok": False, "error": f"ref PDF doc_id={doc_id} 不属于当前 ref 目录。"}
    text = str(row["text"] or "")
    max_chars = max(1, min(int(max_chars or 2400), 12000))
    return {
        "ok": True,
        "doc_id": int(row["doc_id"]),
        "title": row["title"],
        "rel_path": row["rel_path"],
        "path": row["path"],
        "status": row["status"],
        "page": int(row["page"] or page),
        "chars": len(text),
        "truncated": len(text) > max_chars,
        "content": text[:max_chars],
        "summary": f"读取 ref PDF {row['title']} 第 {int(row['page'] or page)} 页，返回 {min(len(text), max_chars)} 字符。",
    }


def get_review_checklist_excerpt(doc_id: int, page: int, *, max_chars: int = 2400) -> dict:
    with _connect_checklist() as conn:
        row = conn.execute(
            """
            SELECT d.id AS doc_id,d.title,d.rel_path,d.path,d.status,d.error,p.page,p.text
            FROM documents d
            LEFT JOIN pages p ON p.doc_id=d.id AND p.page=?
            WHERE d.id=?
            """,
            (int(page), int(doc_id)),
        ).fetchone()
    if not row:
        return {"ok": False, "error": f"未找到 review checklist doc_id={doc_id} page={page}。"}
    if not _path_in_root(row["path"], agent_checklist_ref_dir()):
        return {"ok": False, "error": f"review checklist doc_id={doc_id} 不属于当前 ref_checklist 目录。"}
    text = str(row["text"] or "")
    max_chars = max(1, min(int(max_chars or 2400), 12000))
    return {
        "ok": True,
        "doc_id": int(row["doc_id"]),
        "title": row["title"],
        "rel_path": row["rel_path"],
        "path": row["path"],
        "status": row["status"],
        "page": int(row["page"] or page),
        "chars": len(text),
        "truncated": len(text) > max_chars,
        "content": text[:max_chars],
        "summary": f"读取 review checklist {row['title']} 片段 {int(row['page'] or page)}，返回 {min(len(text), max_chars)} 字符。",
    }
