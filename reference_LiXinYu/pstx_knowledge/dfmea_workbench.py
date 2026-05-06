# -*- coding: utf-8 -*-
"""Local DFMEA grouping workbench.

This module owns the small SQLite store used by the manual DFMEA page.  It
stores only tool-local grouping data and component snapshots derived from the
current report run; it never writes back to PSTX/Cadence/Feishu sources.
"""

from __future__ import annotations

import io
import json
import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from pstx_knowledge.component_identity import classify_refdes
from pstx_core.pages import _natural_sort_key


DFMEA_DATA_DIR_ENV = "PSTX_DFMEA_DATA_DIR"
DEFAULT_DFMEA_DATA_DIR = "dfmea_data"
DFMEA_DB_NAME = "dfmea_workbench.db"

DFMEA_TEXT_FIELDS = (
    "function_requirement",
    "failure_mode",
    "failure_effect",
    "failure_cause",
    "prevention_detection",
)

DFMEA_FIELD_LABELS = {
    "function_requirement": "功能/需求",
    "failure_mode": "潜在失效模式",
    "failure_effect": "潜在失效后果",
    "failure_cause": "潜在失效原因/机理",
    "prevention_detection": "现有预防/探测方案",
}

EXPORT_HEADERS = [
    "组ID",
    "位号",
    "页码",
    "功能/需求",
    "潜在失效模式",
    "潜在失效后果",
    "潜在失效原因/机理",
    "现有预防/探测方案",
    "更新时间",
]

TYPE_ORDER = {
    "U": 10,
    "PU": 11,
    "XU": 12,
    "J": 20,
    "P": 21,
    "CN": 22,
    "R": 30,
    "C": 31,
    "L": 32,
}


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _clean_text(value, limit: int = 500) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return text if len(text) <= limit else text[:limit - 1] + "…"


def _json_dumps(value) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def dfmea_data_dir() -> Path:
    configured = os.environ.get(DFMEA_DATA_DIR_ENV)
    return Path(configured).expanduser() if configured else Path(DEFAULT_DFMEA_DATA_DIR)


def dfmea_db_path() -> Path:
    return dfmea_data_dir() / DFMEA_DB_NAME


def _connect() -> sqlite3.Connection:
    path = dfmea_db_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    _init_db(conn)
    return conn


def _init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS dfmea_projects (
            run_id TEXT PRIMARY KEY,
            project_name TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS dfmea_components (
            run_id TEXT NOT NULL,
            refdes TEXT NOT NULL,
            page TEXT NOT NULL DEFAULT '',
            sort_page_num INTEGER NOT NULL DEFAULT 999999,
            refdes_type TEXT NOT NULL DEFAULT '',
            category TEXT NOT NULL DEFAULT '',
            hq_no TEXT NOT NULL DEFAULT '',
            value TEXT NOT NULL DEFAULT '',
            package TEXT NOT NULL DEFAULT '',
            bom_option TEXT NOT NULL DEFAULT '',
            is_depop INTEGER NOT NULL DEFAULT 0,
            summary_json TEXT NOT NULL DEFAULT '{}',
            updated_at TEXT NOT NULL,
            PRIMARY KEY (run_id, refdes)
        );

        CREATE TABLE IF NOT EXISTS dfmea_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT NOT NULL,
            function_requirement TEXT NOT NULL DEFAULT '',
            failure_mode TEXT NOT NULL DEFAULT '',
            failure_effect TEXT NOT NULL DEFAULT '',
            failure_cause TEXT NOT NULL DEFAULT '',
            prevention_detection TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS dfmea_group_components (
            group_id INTEGER NOT NULL,
            run_id TEXT NOT NULL,
            refdes TEXT NOT NULL,
            page TEXT NOT NULL DEFAULT '',
            refdes_type TEXT NOT NULL DEFAULT '',
            sort_index INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (group_id, refdes),
            UNIQUE (run_id, refdes),
            FOREIGN KEY (group_id) REFERENCES dfmea_groups(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_dfmea_components_run ON dfmea_components(run_id);
        CREATE INDEX IF NOT EXISTS idx_dfmea_group_components_run ON dfmea_group_components(run_id);
        """
    )
    conn.commit()


def _page_value(comp: Mapping[str, object]) -> str:
    for key in ("page_user_visible_pages", "page_submodule_mapped", "user_visible_page", "页码", "page_real", "真实页", "page"):
        value = _clean_text(comp.get(key), 120)
        if value:
            return value
    return ""


def _main_page_value(comp: Mapping[str, object]) -> str:
    for key in ("page_logical_pages", "page_logical", "主模块页", "page_original"):
        value = _clean_text(comp.get(key), 120)
        if value:
            return value
    return ""


def _first_value(comp: Mapping[str, object], keys: Sequence[str], limit: int = 240) -> str:
    for key in keys:
        value = _clean_text(comp.get(key), limit)
        if value:
            return value
    return ""


def _is_depop(bom_option: str) -> bool:
    value = str(bom_option or "").upper()
    return "DEPOP" in value or "DNP" in value or "NO MOUNT" in value or "NOSTUFF" in value


def _refdes_type(refdes: str) -> str:
    text = str(refdes or "").strip().upper()
    power_passive = re.match(r"^P([RCL])\d", text)
    if power_passive:
        return power_passive.group(1)
    match = re.match(r"^[A-Z]+", text)
    return match.group(0) if match else ""


def _sort_page_num(page: str) -> int:
    match = re.search(r"(\d+)", str(page or ""))
    return int(match.group(1)) if match else 999999


def _type_sort_value(refdes_type: str) -> int:
    return TYPE_ORDER.get(str(refdes_type or "").upper(), 500)


def _component_snapshot(refdes: str, comp: Mapping[str, object]) -> dict:
    refdes = str(refdes or comp.get("refdes") or "").strip()
    category, candidate, confidence = classify_refdes(refdes, dict(comp))
    bom_option = _first_value(comp, ("bom_option", "BOM_OPTION"), 160)
    page = _page_value(comp)
    summary = {
        "refdes": refdes,
        "candidate": candidate,
        "confidence": confidence,
        "main_module_page": _main_page_value(comp),
        "part_type": _first_value(comp, ("part_type", "CDS_PART_NAME", "cds_part_name"), 240),
        "nets": comp.get("nets") if isinstance(comp.get("nets"), dict) else {},
    }
    return {
        "refdes": refdes,
        "page": page,
        "sort_page_num": _sort_page_num(page),
        "refdes_type": _refdes_type(refdes),
        "category": category,
        "hq_no": _first_value(comp, ("hq_code", "HQ_CODE", "料号", "part_number", "PART_NUMBER"), 160),
        "value": _first_value(comp, ("value", "VALUE", "值"), 240),
        "package": _first_value(comp, ("package", "PACKAGE", "封装"), 160),
        "bom_option": bom_option,
        "is_depop": 1 if _is_depop(bom_option) else 0,
        "summary_json": _json_dumps(summary),
    }


def _iter_source_components(bundle: Mapping[str, object]) -> Iterable[tuple[str, Mapping[str, object]]]:
    components = bundle.get("all_components")
    if not isinstance(components, Mapping):
        components = bundle.get("components")
    if not isinstance(components, Mapping):
        return []
    return [
        (str(refdes), comp)
        for refdes, comp in components.items()
        if isinstance(comp, Mapping) and str(refdes or comp.get("refdes") or "").strip()
    ]


def sync_dfmea_project(run_id: str, report: Mapping[str, object], bundle: Mapping[str, object]) -> dict:
    run_id = str(run_id or "").strip()
    if not run_id:
        raise ValueError("run_id 不能为空")
    project_name = _clean_text(
        report.get("project_name") if isinstance(report, Mapping) else "",
        240,
    ) or _clean_text(bundle.get("project_name") if isinstance(bundle, Mapping) else "", 240)
    now = _now()
    rows = [_component_snapshot(refdes, comp) for refdes, comp in _iter_source_components(bundle)]
    with _connect() as conn:
        conn.execute(
            """
            INSERT INTO dfmea_projects(run_id, project_name, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(run_id) DO UPDATE SET
                project_name=excluded.project_name,
                updated_at=excluded.updated_at
            """,
            (run_id, project_name, now, now),
        )
        for row in rows:
            conn.execute(
                """
                INSERT INTO dfmea_components(
                    run_id, refdes, page, sort_page_num, refdes_type, category,
                    hq_no, value, package, bom_option, is_depop, summary_json, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(run_id, refdes) DO UPDATE SET
                    page=excluded.page,
                    sort_page_num=excluded.sort_page_num,
                    refdes_type=excluded.refdes_type,
                    category=excluded.category,
                    hq_no=excluded.hq_no,
                    value=excluded.value,
                    package=excluded.package,
                    bom_option=excluded.bom_option,
                    is_depop=excluded.is_depop,
                    summary_json=excluded.summary_json,
                    updated_at=excluded.updated_at
                """,
                (
                    run_id,
                    row["refdes"],
                    row["page"],
                    row["sort_page_num"],
                    row["refdes_type"],
                    row["category"],
                    row["hq_no"],
                    row["value"],
                    row["package"],
                    row["bom_option"],
                    row["is_depop"],
                    row["summary_json"],
                    now,
                ),
            )
        conn.commit()
    return {"ok": True, "component_count": len(rows), "db_path": str(dfmea_db_path())}


def _row_to_component(row: sqlite3.Row) -> dict:
    return {
        "refdes": row["refdes"],
        "页码": row["page"],
        "page": row["page"],
        "refdes_type": row["refdes_type"],
        "category": row["category"],
        "hq_no": row["hq_no"],
        "value": row["value"],
        "package": row["package"],
        "bom_option": row["bom_option"],
        "is_depop": bool(row["is_depop"]),
        "summary": json.loads(row["summary_json"] or "{}"),
        "updated_at": row["updated_at"],
    }


def _component_sort_key(row: dict, sort: str):
    sort = str(sort or "page")
    refdes = row.get("refdes", "")
    if sort == "refdes":
        return (_natural_sort_key(refdes),)
    if sort == "type":
        return (_type_sort_value(row.get("refdes_type")), _natural_sort_key(refdes))
    if sort == "hq":
        return (str(row.get("hq_no") or "ZZZ").upper(), _natural_sort_key(refdes))
    return (
        _sort_page_num(row.get("page")),
        _type_sort_value(row.get("refdes_type")),
        _natural_sort_key(refdes),
    )


def _matches_query(row: dict, query: str) -> bool:
    query = str(query or "").strip().lower()
    if not query:
        return True
    haystack = " ".join(
        str(row.get(key) or "")
        for key in ("refdes", "page", "refdes_type", "category", "hq_no", "value", "package", "bom_option")
    ).lower()
    return query in haystack


def _fetch_groups(conn: sqlite3.Connection, run_id: str) -> list[dict]:
    groups = []
    group_rows = conn.execute(
        "SELECT * FROM dfmea_groups WHERE run_id = ? ORDER BY id",
        (run_id,),
    ).fetchall()
    for group in group_rows:
        comp_rows = conn.execute(
            """
            SELECT c.*
            FROM dfmea_group_components gc
            JOIN dfmea_components c
              ON c.run_id = gc.run_id
             AND UPPER(c.refdes) = UPPER(gc.refdes)
            WHERE gc.group_id = ?
            ORDER BY gc.sort_index, gc.refdes
            """,
            (group["id"],),
        ).fetchall()
        refs = [row["refdes"] for row in comp_rows]
        pages = _unique_pages(row["page"] for row in comp_rows)
        components = [_row_to_component(row) for row in comp_rows]
        groups.append({
            "id": group["id"],
            "run_id": group["run_id"],
            "refdes": refs,
            "refdes_text": ", ".join(refs),
            "pages": pages,
            "pages_text": ", ".join(pages),
            "components": components,
            **{field: group[field] for field in DFMEA_TEXT_FIELDS},
            "created_at": group["created_at"],
            "updated_at": group["updated_at"],
        })
    return groups


def _unique_pages(pages: Iterable[str]) -> list[str]:
    seen = set()
    result = []
    for page in sorted([str(page or "").strip() for page in pages if str(page or "").strip()], key=_natural_sort_key):
        if page not in seen:
            seen.add(page)
            result.append(page)
    return result


def get_dfmea_workbench(run_id: str,
                        report: Mapping[str, object],
                        bundle: Mapping[str, object],
                        *,
                        include_depop: bool = False,
                        exclude_rc: bool = False,
                        sort: str = "page",
                        query: str = "") -> dict:
    sync = sync_dfmea_project(run_id, report, bundle)
    with _connect() as conn:
        grouped_refdes = {
            row["refdes"]
            for row in conn.execute(
                "SELECT refdes FROM dfmea_group_components WHERE run_id = ?",
                (run_id,),
            ).fetchall()
        }
        rows = [
            _row_to_component(row)
            for row in conn.execute(
                "SELECT * FROM dfmea_components WHERE run_id = ?",
                (run_id,),
            ).fetchall()
        ]
        pending = [
            row
            for row in rows
            if row["refdes"] not in grouped_refdes
            and (include_depop or not row["is_depop"])
            and (not exclude_rc or str(row.get("refdes_type") or "").upper() not in {"R", "C"})
            and _matches_query(row, query)
        ]
        pending.sort(key=lambda row: _component_sort_key(row, sort))
        groups = _fetch_groups(conn, run_id)
    return {
        "ok": True,
        "run_id": run_id,
        "db_path": sync["db_path"],
        "sort": sort or "page",
        "include_depop": bool(include_depop),
        "exclude_rc": bool(exclude_rc),
        "query": query or "",
        "pending_count": len(pending),
        "group_count": len(groups),
        "total_component_count": sync["component_count"],
        "pending_components": pending,
        "groups": groups,
        "fields": dict(DFMEA_FIELD_LABELS),
        "export_headers": list(EXPORT_HEADERS),
    }


def _normalize_refdes_list(values) -> list[str]:
    if isinstance(values, str):
        raw = re.split(r"[,\s，、]+", values)
    elif isinstance(values, Sequence):
        raw = [str(item) for item in values]
    else:
        raw = []
    result = []
    seen = set()
    for value in raw:
        refdes = str(value or "").strip()
        if not refdes:
            continue
        key = refdes.upper()
        if key in seen:
            continue
        seen.add(key)
        result.append(refdes)
    return result


def _clean_group_payload(payload: Mapping[str, object], existing: Mapping[str, object] | None = None) -> dict:
    fields = {}
    for field in DFMEA_TEXT_FIELDS:
        if field in payload:
            fields[field] = _clean_text(payload.get(field), 4000)
        elif existing is not None:
            fields[field] = _clean_text(existing.get(field), 4000)
        else:
            fields[field] = ""
    return fields


def _component_rows_for_refdes(conn: sqlite3.Connection, run_id: str, refdes_list: Sequence[str]) -> list[sqlite3.Row]:
    rows = []
    for refdes in refdes_list:
        row = conn.execute(
            "SELECT * FROM dfmea_components WHERE run_id = ? AND UPPER(refdes) = UPPER(?)",
            (run_id, refdes),
        ).fetchone()
        if row is None:
            raise ValueError(f"未找到元器件位号：{refdes}")
        rows.append(row)
    return rows


def _ensure_refdes_available(conn: sqlite3.Connection,
                             run_id: str,
                             refdes_list: Sequence[str],
                             *,
                             exclude_group_id: int | None = None) -> None:
    for refdes in refdes_list:
        params: list[object] = [run_id, refdes]
        sql = "SELECT group_id FROM dfmea_group_components WHERE run_id = ? AND UPPER(refdes) = UPPER(?)"
        if exclude_group_id is not None:
            sql += " AND group_id <> ?"
            params.append(exclude_group_id)
        existing = conn.execute(sql, tuple(params)).fetchone()
        if existing is not None:
            raise ValueError(f"位号 {refdes} 已经保存到组 {existing['group_id']}，请先编辑或删除原分组。")


def create_dfmea_group(run_id: str, payload: Mapping[str, object]) -> dict:
    refdes_list = _normalize_refdes_list(payload.get("refdes"))
    if not refdes_list:
        raise ValueError("请选择至少一个元器件位号。")
    now = _now()
    fields = _clean_group_payload(payload)
    with _connect() as conn:
        rows = _component_rows_for_refdes(conn, run_id, refdes_list)
        _ensure_refdes_available(conn, run_id, refdes_list)
        cursor = conn.execute(
            """
            INSERT INTO dfmea_groups(
                run_id, function_requirement, failure_mode, failure_effect,
                failure_cause, prevention_detection, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                fields["function_requirement"],
                fields["failure_mode"],
                fields["failure_effect"],
                fields["failure_cause"],
                fields["prevention_detection"],
                now,
                now,
            ),
        )
        group_id = int(cursor.lastrowid)
        _replace_group_components(conn, run_id, group_id, rows)
        conn.commit()
    return {"ok": True, "group_id": group_id}


def _replace_group_components(conn: sqlite3.Connection,
                              run_id: str,
                              group_id: int,
                              rows: Sequence[sqlite3.Row]) -> None:
    conn.execute("DELETE FROM dfmea_group_components WHERE group_id = ?", (group_id,))
    sorted_rows = sorted(
        rows,
        key=lambda row: (_sort_page_num(row["page"]), _type_sort_value(row["refdes_type"]), _natural_sort_key(row["refdes"])),
    )
    for index, row in enumerate(sorted_rows):
        conn.execute(
            """
            INSERT INTO dfmea_group_components(group_id, run_id, refdes, page, refdes_type, sort_index)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (group_id, run_id, row["refdes"], row["page"], row["refdes_type"], index),
        )


def update_dfmea_group(run_id: str, group_id: int, payload: Mapping[str, object]) -> dict:
    group_id = int(group_id)
    now = _now()
    refdes_list = _normalize_refdes_list(payload.get("refdes")) if "refdes" in payload else None
    with _connect() as conn:
        group = conn.execute(
            "SELECT * FROM dfmea_groups WHERE run_id = ? AND id = ?",
            (run_id, group_id),
        ).fetchone()
        if group is None:
            raise ValueError(f"未找到 DFMEA 分组：{group_id}")
        fields = _clean_group_payload(payload, group)
        conn.execute(
            """
            UPDATE dfmea_groups SET
                function_requirement = ?,
                failure_mode = ?,
                failure_effect = ?,
                failure_cause = ?,
                prevention_detection = ?,
                updated_at = ?
            WHERE run_id = ? AND id = ?
            """,
            (
                fields["function_requirement"],
                fields["failure_mode"],
                fields["failure_effect"],
                fields["failure_cause"],
                fields["prevention_detection"],
                now,
                run_id,
                group_id,
            ),
        )
        if refdes_list is not None:
            if not refdes_list:
                raise ValueError("分组至少需要保留一个位号。")
            rows = _component_rows_for_refdes(conn, run_id, refdes_list)
            _ensure_refdes_available(conn, run_id, refdes_list, exclude_group_id=group_id)
            _replace_group_components(conn, run_id, group_id, rows)
        conn.commit()
    return {"ok": True, "group_id": group_id}


def delete_dfmea_group(run_id: str, group_id: int) -> dict:
    with _connect() as conn:
        cursor = conn.execute(
            "DELETE FROM dfmea_groups WHERE run_id = ? AND id = ?",
            (run_id, int(group_id)),
        )
        conn.commit()
    if cursor.rowcount <= 0:
        raise ValueError(f"未找到 DFMEA 分组：{group_id}")
    return {"ok": True, "group_id": int(group_id)}


def export_dfmea_workbook(run_id: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with _connect() as conn:
        groups = _fetch_groups(conn, run_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "DFMEA"
    ws.append(EXPORT_HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for group in groups:
        ws.append([
            group["id"],
            group["refdes_text"],
            group["pages_text"],
            group["function_requirement"],
            group["failure_mode"],
            group["failure_effect"],
            group["failure_cause"],
            group["prevention_detection"],
            group["updated_at"],
        ])
    widths = [12, 38, 24, 32, 32, 34, 36, 38, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


__all__ = [
    "DFMEA_DATA_DIR_ENV",
    "DEFAULT_DFMEA_DATA_DIR",
    "DFMEA_DB_NAME",
    "DFMEA_FIELD_LABELS",
    "EXPORT_HEADERS",
    "create_dfmea_group",
    "delete_dfmea_group",
    "dfmea_data_dir",
    "dfmea_db_path",
    "export_dfmea_workbook",
    "get_dfmea_workbench",
    "sync_dfmea_project",
    "update_dfmea_group",
]
