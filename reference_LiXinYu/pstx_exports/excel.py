# -*- coding: utf-8 -*-
"""Excel export for PSTX analysis bundles."""

from __future__ import annotations

import os
import subprocess
import sys
from typing import Dict, List

try:
    import openpyxl
except ImportError:
    print("未检测到 openpyxl，正在自动安装...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from pstx_rules.result_meta import DRC_ISSUE_KEYS, count_result_kinds, iter_list_rows


_BL = PatternFill("solid", fgColor="1F4E79")
_OR = PatternFill("solid", fgColor="C55A11")
_GR = PatternFill("solid", fgColor="375623")
_GY = PatternFill("solid", fgColor="595959")
_RF = PatternFill("solid", fgColor="FFCCCC")
_WF = Font(color="FFFFFF", bold=True, size=10)
_BF = Font(bold=True, size=10)
_NF = Font(size=10)
_CA = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LA = Alignment(horizontal="left", vertical="center", wrap_text=True)
_TH = Side(style="thin")
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


BOM_EXPORT_MODES: Dict[str, dict] = {
    "all": {
        "label": "包含 DEPOP，同料号合并",
        "sheet": "BOM_含DEPOP",
        "fill": _GY,
    },
    "mounted": {
        "label": "不包含 DEPOP，仅贴装",
        "sheet": "BOM_不含DEPOP",
        "fill": _BL,
    },
    "split": {
        "label": "贴装与 DEPOP 分条",
        "sheet": "BOM_分条",
        "fill": _OR,
    },
}


_BOM_EXPORT_COLUMN_ORDER = [
    "序号",
    "BOM状态",
    "料号",
    "位号列表",
    "数量",
    "贴装数量",
    "DEPOP数量",
    "描述",
    "值",
    "封装",
    "耐压",
    "耐压/额定电压",
    "额定功率",
    "精度",
    "材质",
    "类型",
]


def _xl_hdr(ws, row_idx, fill):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = _WF
            cell.alignment = _CA
            cell.border = _BD


def _xl_autowidth(ws, mx=50):
    for col in ws.columns:
        vals = [str(c.value or "") for c in col]
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(v) for v in vals), default=8) + 2,
            mx,
        )


def _xl_write_rows(ws, rows: List[dict], fill, hl_col=None, freeze=True):
    if not rows:
        ws.append(["（无数据）"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    _xl_hdr(ws, ws.max_row, fill)
    highlight_index = headers.index(hl_col) if hl_col in headers else None
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
        row_index = ws.max_row
        red = highlight_index is not None and "❌" in str(ws.cell(row_index, highlight_index + 1).value or "")
        for cell in ws[row_index]:
            cell.border = _BD
            cell.alignment = _LA
            cell.font = _NF
            if red:
                cell.fill = _RF
    _xl_autowidth(ws)
    if freeze:
        ws.freeze_panes = "A2"


def _xl_section(ws, title, fill):
    ws.append([title])
    for cell in ws[ws.max_row]:
        cell.fill = fill
        cell.font = _WF
        cell.border = _BD
    ws.append([])


def _normalize_bom_export_mode(mode: str) -> str:
    normalized = str(mode or "all").strip().lower().replace("-", "_")
    aliases = {
        "include_depop": "all",
        "including_depop": "all",
        "all_including_depop": "all",
        "total": "all",
        "normal": "mounted",
        "mounted_only": "mounted",
        "exclude_depop": "mounted",
        "without_depop": "mounted",
        "split_depop": "split",
        "separate": "split",
    }
    normalized = aliases.get(normalized, normalized)
    if normalized not in BOM_EXPORT_MODES:
        allowed = ", ".join(BOM_EXPORT_MODES)
        raise ValueError(f"不支持的 BOM 导出模式：{mode}，可用模式：{allowed}")
    return normalized


def _ordered_bom_export_row(row: dict, *, index: int, status: str | None = None) -> dict:
    source = {key: value for key, value in dict(row or {}).items() if key != "_ctype"}
    ordered = {"序号": index}
    if status is not None:
        ordered["BOM状态"] = status
    elif "BOM状态" in source:
        ordered["BOM状态"] = source.get("BOM状态", "")
    for column in _BOM_EXPORT_COLUMN_ORDER:
        if column in {"序号", "BOM状态"}:
            continue
        if column in source:
            ordered[column] = source.get(column, "")
    for key, value in source.items():
        if key in ordered or key in {"序号", "BOM状态"}:
            continue
        ordered[key] = value
    return ordered


def build_bom_export_rows(data: dict, mode: str = "all") -> List[dict]:
    """Build material-number grouped BOM rows for report-local BOM downloads."""
    mode = _normalize_bom_export_mode(mode)
    if mode == "all":
        rows = list(data.get("bom_total_merged", []) or [])
        return [_ordered_bom_export_row(row, index=index) for index, row in enumerate(rows, 1)]
    if mode == "mounted":
        rows = list(data.get("bom_normal_merged", []) or [])
        return [_ordered_bom_export_row(row, index=index) for index, row in enumerate(rows, 1)]

    split_rows = []
    for row in data.get("bom_normal_merged", []) or []:
        split_rows.append((row, "贴装"))
    for row in data.get("bom_depop_merged", []) or []:
        split_rows.append((row, "DEPOP"))
    return [
        _ordered_bom_export_row(row, index=index, status=status)
        for index, (row, status) in enumerate(split_rows, 1)
    ]


def export_bom_to_excel(data: dict, out_path: str, *, mode: str = "all") -> str:
    """Export one BOM view as a small Excel workbook."""
    mode = _normalize_bom_export_mode(mode)
    base, ext = os.path.splitext(out_path)
    n, path = 1, out_path
    while os.path.exists(path):
        path = f"{base}({n}){ext}"
        n += 1

    meta = BOM_EXPORT_MODES[mode]
    rows = build_bom_export_rows(data, mode)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("说明")
    summary_rows = [
        ("项目名称", data.get("project_name", "")),
        ("导出口径", meta["label"]),
        ("物料条目数", len(rows)),
        ("物料总数", sum(int(row.get("数量", 0) or 0) for row in rows)),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _BD
            cell.font = _BF if cell.column == 1 else _NF
            cell.alignment = _LA
    _xl_autowidth(ws)

    ws = wb.create_sheet(str(meta["sheet"]))
    _xl_write_rows(ws, rows, meta["fill"])

    wb.save(path)
    return path


def export_to_excel(data: dict, out_path: str) -> str:
    base, ext = os.path.splitext(out_path)
    n, path = 1, out_path
    while os.path.exists(path):
        path = f"{base}({n}){ext}"
        n += 1

    wb = Workbook()
    wb.remove(wb.active)
    project = data.get("project_name", "")
    net_analysis = data.get("net_analysis", {})
    drc = data.get("drc", {})
    derating = data.get("derating", [])
    resistor = data.get("resistor_analysis", {})
    csa = data.get("csa_geometry", {})
    bom_normal = data.get("bom_normal_merged", [])
    bom_depop = data.get("bom_depop_merged", [])
    bom_total = data.get("bom_total_merged", [])
    network_rows = iter_list_rows(net_analysis, ["power_net_rows", "gnd_net_rows", "diff_pair_rows", "single_node_rows"])
    drc_rows = iter_list_rows(drc, DRC_ISSUE_KEYS)
    resistor_rows = iter_list_rows(resistor, ["divider_risks", "dup_pullups", "dup_pulldowns"])
    net_kind_counts = count_result_kinds(network_rows)
    drc_kind_counts = count_result_kinds(drc_rows)
    derating_kind_counts = count_result_kinds(derating)
    resistor_kind_counts = count_result_kinds(resistor_rows)

    ws = wb.create_sheet("概览")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    drc_total = sum(len(drc.get(key, [])) for key in DRC_ISSUE_KEYS)
    derating_fail = sum(1 for row in derating if row.get("状态", "").startswith("❌"))
    for label, val in [
        ("项目名称", project),
        ("贴装元件种类数", len(bom_normal)),
        ("贴装元件总数", sum(row.get("数量", 0) for row in bom_normal)),
        ("DEPOP 元件种类数", len(bom_depop)),
        ("DEPOP 元件总数", sum(row.get("数量", 0) for row in bom_depop)),
        ("总BOM 元件种类数", len(bom_total)),
        ("总BOM 元件总数", sum(row.get("数量", 0) for row in bom_total)),
        ("BOM_OPTION 打圈覆盖问题数", len(drc.get("bom_option_circle_issues", []))),
        ("网络总数", net_analysis.get("total", "")),
        ("候选单节点网络数", len(net_analysis.get("single_node", {}))),
        ("候选电源网络数", len(net_analysis.get("power_nets", {}))),
        ("候选差分对数", len(net_analysis.get("diff_pairs", {}))),
        ("DRC 问题总数", drc_total),
        ("电容降额不合格数", derating_fail),
        ("网络候选判断数", net_kind_counts.get("候选判断", 0)),
        ("DRC 确定结论数", drc_kind_counts.get("确定结论", 0)),
        ("DRC 候选判断数", drc_kind_counts.get("候选判断", 0)),
        ("降额确定结论数", derating_kind_counts.get("确定结论", 0)),
        ("降额候选判断数", derating_kind_counts.get("候选判断", 0)),
        ("降额无法判断数", derating_kind_counts.get("无法判断", 0)),
        ("电阻候选判断数", resistor_kind_counts.get("候选判断", 0)),
        ("电阻无法判断数", resistor_kind_counts.get("无法判断", 0)),
        ("CSA 扫描页数", csa.get("page_count", 0)),
        ("CSA DOT四向十字数", csa.get("cross_count", 0)),
    ]:
        ws.append([label, val])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _BD
            cell.font = _BF if cell.column == 1 else _NF
            cell.alignment = _LA

    ws = wb.create_sheet("BOM_贴装")
    _xl_write_rows(ws, bom_normal, _BL)
    ws = wb.create_sheet("BOM_DEPOP")
    _xl_write_rows(ws, bom_depop, _OR)
    ws = wb.create_sheet("BOM_总")
    _xl_write_rows(ws, bom_total, _GY)
    ws = wb.create_sheet("BOM_明细")
    all_detail = (
        [{"DEPOP": "", **row} for row in data.get("bom_normal_detail", [])]
        + [{"DEPOP": "Y", **row} for row in data.get("bom_depop_detail", [])]
    )
    _xl_write_rows(ws, all_detail, _GY)

    ws = wb.create_sheet("网络分析")
    ws.freeze_panes = None
    _xl_section(ws, "候选电源网络", _BL)
    _xl_write_rows(ws, net_analysis.get("power_net_rows", []), _BL, freeze=False)
    ws.append([])
    _xl_section(ws, "候选 GND 网络", _GR)
    _xl_write_rows(ws, net_analysis.get("gnd_net_rows", []), _GR, freeze=False)
    ws.append([])
    _xl_section(ws, "候选差分对", _OR)
    _xl_write_rows(ws, net_analysis.get("diff_pair_rows", []), _OR, freeze=False)
    ws.append([])
    _xl_section(ws, "单节点网络概览", _GY)
    _xl_write_rows(ws, net_analysis.get("single_node_rows", []), _GY, freeze=False)
    ws.append([])
    _xl_section(ws, "页码元件分布", _BL)
    _xl_write_rows(ws, net_analysis.get("page_rows", []), _BL, freeze=False)
    ws.append([])
    _xl_section(ws, "主模块页/页码映射检查", _GY)
    _xl_write_rows(ws, data.get("page_mapping_rows", []), _GY, freeze=False)
    _xl_autowidth(ws)

    ws = wb.create_sheet("设计检查")
    ws.freeze_panes = None
    for title, key, fill in [
        ("TBD 待确认属性", "tbd_attrs", _OR),
        ("缺少料号", "missing_hq_code", _RF),
        ("缺少 VALUE", "missing_value", _RF),
        ("缺少封装", "missing_package", _RF),
        ("单端候选网络", "single_pin_nets", _GY),
        ("未命名网络", "unnamed_nets", _GY),
        ("BOM_OPTION 候选拼写", "bom_option_typos", _OR),
        ("BOM_OPTION 打圈覆盖问题", "bom_option_circle_issues", _OR),
        ("BOM_OPTION 元件", "bom_option_components", _BL),
        ("BOM_OPTION 打圈覆盖明细", "bom_option_circle_coverage", _GY),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, drc.get(key, []), fill, freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    ws = wb.create_sheet("电阻检查")
    ws.freeze_panes = None
    for title, key, fill in [
        ("串阻分压候选风险", "divider_risks", _OR),
        ("重复上拉候选", "dup_pullups", _RF),
        ("重复下拉候选", "dup_pulldowns", _RF),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, resistor.get(key, []), fill, hl_col="状态", freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    ws = wb.create_sheet("芯片引脚电阻")
    _xl_write_rows(ws, resistor.get("chip_pin_rows", []), _BL)

    ws = wb.create_sheet("规范检查")
    ws.freeze_panes = None
    _xl_section(ws, "CSA 页级汇总", _BL)
    _xl_write_rows(ws, csa.get("summary_rows", []), _BL, freeze=False)
    ws.append([])
    _xl_section(ws, "CSA DOT四向十字交叉", _OR)
    _xl_write_rows(ws, csa.get("dot_cross_rows", []), _OR, freeze=False)
    _xl_autowidth(ws)

    ws = wb.create_sheet("降额分析")
    _xl_write_rows(ws, derating, _BL, hl_col="状态")

    wb.save(path)
    return path
