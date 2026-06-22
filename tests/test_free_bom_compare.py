# -*- coding: utf-8 -*-
import json
import sys
import unittest
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
WEB_APP = ROOT / "web_app2"
TESTS = ROOT / "tests"
if str(TESTS) not in sys.path:
    sys.path.insert(0, str(TESTS))
from test_env import configure_test_environment  # noqa: E402

configure_test_environment()
if str(WEB_APP) not in sys.path:
    sys.path.insert(0, str(WEB_APP))

from app import app  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402


def xlsx_bytes(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


class FreeBomCompareTests(unittest.TestCase):
    def test_free_bom_compare_accepts_two_arbitrary_excels(self):
        client = app.test_client()
        sheets_resp = client.post(
            "/api/bom_compare/free_sheets",
            data={
                "left_file": (xlsx_bytes(["Item", "MPN", "Qty", "Remark"], [["A1", "R-10K", 2, "old"]]), "left.xlsx"),
                "right_file": (xlsx_bytes(["Code", "Model", "Quantity", "Remark"], [["A1", "R-10K", 3, "new"]]), "right.xlsx"),
                "left_header_row": "1",
                "right_header_row": "1",
            },
            content_type="multipart/form-data",
        )
        sheets_payload = sheets_resp.get_json()
        self.assertTrue(sheets_payload["success"], sheets_payload)
        self.assertEqual(sheets_payload["left_format"], "generic")
        self.assertEqual(sheets_payload["right_format"], "generic")

        compare_resp = client.post(
            "/api/bom_compare/free",
            data={
                "left_file": (xlsx_bytes(["Item", "MPN", "Qty", "Remark"], [["A1", "R-10K", 2, "old"], ["A2", "C-1U", 1, "same"]]), "left.xlsx"),
                "right_file": (xlsx_bytes(["Code", "Model", "Quantity", "Remark"], [["A1", "R-10K", 3, "new"], ["A3", "L-2U2", 1, "added"]]), "right.xlsx"),
                "config": json.dumps({
                    "left_header_row": 1,
                    "right_header_row": 1,
                    "left_key_col": "Item",
                    "right_key_col": "Code",
                    "field_pairs": [
                        {"left": "MPN", "right": "Model"},
                        {"left": "Qty", "right": "Quantity"},
                        {"left": "Remark", "right": "Remark"},
                    ],
                }),
            },
            content_type="multipart/form-data",
        )
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"], payload)
        self.assertEqual(payload["left_only"], 1)
        self.assertEqual(payload["right_only"], 1)
        self.assertEqual(payload["changed"], 1)
        self.assertEqual(payload["same"], 0)
        self.assertTrue(payload["download"].endswith(".xlsx"))

    def test_free_bom_report_exports_review_friendly_sheets(self):
        client = app.test_client()
        compare_resp = client.post(
            "/api/bom_compare/free",
            data={
                "left_file": (xlsx_bytes(["Item", "MPN", "Qty", "Remark"], [["A1", "R-10K", 2, "old remark with enough detail for width"], ["A2", "C-1U", 1, "removed"]]), "left.xlsx"),
                "right_file": (xlsx_bytes(["Code", "Model", "Quantity", "Remark"], [["A1", "R-22K", 3, "new remark with enough detail for width"], ["A3", "L-2U2", 1, "added"]]), "right.xlsx"),
                "config": json.dumps({
                    "left_header_row": 1,
                    "right_header_row": 1,
                    "left_key_col": "Item",
                    "right_key_col": "Code",
                    "field_pairs": [
                        {"left": "MPN", "right": "Model"},
                        {"left": "Qty", "right": "Quantity"},
                        {"left": "Remark", "right": "Remark"},
                    ],
                }),
            },
            content_type="multipart/form-data",
        )
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"], payload)

        report_name = payload["download"].rsplit("/", 1)[-1]
        report_path = WEB_APP / "outputs" / report_name
        wb = load_workbook(report_path, data_only=False)
        try:
            self.assertEqual(wb.sheetnames, ["\u5dee\u5f02\u603b\u89c8", "\u65b0\u589e\u7269\u6599", "\u5220\u9664\u7269\u6599", "\u53d8\u66f4\u7269\u6599", "\u91cd\u590d\u548c\u7a7a\u952e"])
            self.assertNotIn("\u5dee\u5f02\u660e\u7ec6", wb.sheetnames)
            summary = wb["\u5dee\u5f02\u603b\u89c8"]
            summary_values = {summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value for row in range(1, summary.max_row + 1)}
            summary_rows = {summary.cell(row=row, column=1).value: row for row in range(1, summary.max_row + 1)}
            self.assertEqual(summary["A1"].value, "BOM Tools 导出报告")
            self.assertEqual(summary_values["导出来源"], "BOM Tools 平台 v2.2.1")
            self.assertEqual(summary_values["平台版本"], "v2.2.1")
            self.assertEqual(summary_values["工具版本"], "v1.1.3")
            self.assertIsNone(summary.freeze_panes)
            self.assertEqual(summary_values["工具名称"], "通用 BOM 对比")
            self.assertEqual(summary_values["基准 BOM 文件"], "left.xlsx")
            self.assertEqual(summary_values["对比 BOM 文件"], "right.xlsx")
            self.assertEqual(summary_values["基准 Sheet / 表头行"], "Sheet / 1")
            self.assertEqual(summary_values["对比 Sheet / 表头行"], "Sheet / 1")
            self.assertIn("MPN <-> Model", summary_values["比对字段"])
            self.assertEqual(summary_values["\u65b0\u589e\u7269\u6599"], 1)
            self.assertEqual(summary_values["\u5220\u9664\u7269\u6599"], 1)
            self.assertEqual(summary_values["\u53d8\u66f4\u7269\u6599"], 1)
            self.assertEqual(summary.cell(row=summary_rows["\u65b0\u589e\u7269\u6599"], column=2).hyperlink.target, "#'\u65b0\u589e\u7269\u6599'!A1")
            self.assertEqual(summary.cell(row=summary_rows["\u5220\u9664\u7269\u6599"], column=2).hyperlink.target, "#'\u5220\u9664\u7269\u6599'!A1")
            self.assertEqual(summary.cell(row=summary_rows["\u53d8\u66f4\u7269\u6599"], column=2).hyperlink.target, "#'\u53d8\u66f4\u7269\u6599'!A1")

            added = wb["\u65b0\u589e\u7269\u6599"]
            self.assertEqual([cell.value for cell in added[1]], ["\u5dee\u5f02\u7c7b\u578b", "\u5339\u914d\u952e", "\u5bf9\u6bd4BOM\u884c\u53f7", "Code", "Model", "Quantity", "Remark"])
            self.assertEqual([cell.value for cell in added[2]], ["\u65b0\u589e\u7269\u6599", "A3", 3, "A3", "L-2U2", "1", "added"])

            removed = wb["\u5220\u9664\u7269\u6599"]
            self.assertEqual([cell.value for cell in removed[1]], ["\u5dee\u5f02\u7c7b\u578b", "\u5339\u914d\u952e", "\u57fa\u51c6BOM\u884c\u53f7", "Item", "MPN", "Qty", "Remark"])
            self.assertEqual([cell.value for cell in removed[2]], ["\u5220\u9664\u7269\u6599", "A2", 3, "A2", "C-1U", "1", "removed"])

            changed = wb["\u53d8\u66f4\u7269\u6599"]
            self.assertEqual([cell.value for cell in changed[1]], ["\u5339\u914d\u952e", "\u57fa\u51c6BOM\u884c\u53f7", "\u5bf9\u6bd4BOM\u884c\u53f7", "\u53d8\u66f4\u5b57\u6bb5\u6570", "MPN <-> Model", "Qty <-> Quantity", "Remark"])
            self.assertEqual([cell.value for cell in changed[2]], [
                "A1",
                2,
                2,
                3,
                "R-10K -> R-22K",
                "2 -> 3",
                "old remark with enough detail for width -> new remark with enough detail for width",
            ])
            self.assertGreater(changed.column_dimensions["G"].width, changed.column_dimensions["B"].width)
            self.assertGreaterEqual(changed.column_dimensions["G"].width, 60)
        finally:
            wb.close()

    def test_free_bom_refdes_diff_reports_only_delta_refs(self):
        client = app.test_client()
        compare_resp = client.post(
            "/api/bom_compare/free",
            data={
                "left_file": (xlsx_bytes(["Item", "位号", "Qty"], [["A1", "R1,R2,R3", 3], ["A2", "C1,C2", 2]]), "left.xlsx"),
                "right_file": (xlsx_bytes(["Code", "RefDes", "Quantity"], [["A1", "R3,R1,R4", 3], ["A2", "C2,C1", 2]]), "right.xlsx"),
                "config": json.dumps({
                    "left_header_row": 1,
                    "right_header_row": 1,
                    "left_key_col": "Item",
                    "right_key_col": "Code",
                    "field_pairs": [
                        {"left": "位号", "right": "RefDes"},
                        {"left": "Qty", "right": "Quantity"},
                    ],
                }),
            },
            content_type="multipart/form-data",
        )
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"], payload)
        self.assertEqual(payload["changed"], 1)
        self.assertEqual(payload["same"], 1)

        report_name = payload["download"].rsplit("/", 1)[-1]
        wb = load_workbook(WEB_APP / "outputs" / report_name, data_only=True)
        try:
            changed = wb["变更物料"]
            headers = [cell.value for cell in changed[1]]
            removed_col = headers.index("位号 <-> RefDes 删除位号") + 1
            added_col = headers.index("位号 <-> RefDes 新增位号") + 1
            self.assertEqual(changed.cell(2, 1).value, "A1")
            self.assertEqual(changed.cell(2, 4).value, 1)
            self.assertEqual(changed.cell(2, removed_col).value, "R2")
            self.assertEqual(changed.cell(2, added_col).value, "R4")
        finally:
            wb.close()


    def test_free_bom_preview_returns_rows_before_header_selection(self):
        client = app.test_client()
        resp = client.post(
            "/api/bom_compare/free_preview",
            data={
                "left_file": (xlsx_bytes(["note", "not header"], [["Item", "Qty"], ["A1", 2]]), "left.xlsx"),
            },
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload)
        self.assertIn("left", payload)
        self.assertEqual(payload["left"]["current_sheet"], "Sheet")
        rows = payload["left"]["rows"]
        self.assertEqual(rows[0]["row_number"], 1)
        self.assertEqual(rows[1]["values"][:2], ["Item", "Qty"])
        self.assertEqual(rows[2]["values"][:2], ["A1", "2"])

    def test_bom_compare_template_defaults_to_free_bom_first(self):
        html = (WEB_APP / "templates" / "partials" / "tools" / "bom-compare.html").read_text(encoding="utf-8")
        first_button = html.index('class="bomcmp-tab-btn active"')
        self.assertIn('data-bomcmp-tab="free-bom"', html[first_button:first_button + 140])
        self.assertLess(html.index('id="bomcmp-tab-free-bom"'), html.index('id="bomcmp-tab-customer-hq"'))
        self.assertIn('id="freeLeftPreviewRows"', html)
        self.assertIn('id="freeRightPreviewRows"', html)
        self.assertIn('第二步：确认表头行', html)
        self.assertIn('请在上方预览表格中点击表头所在行', html)
        self.assertIn('第三步：选择匹配键和字段', html)
        self.assertIn('3.1 选择匹配键', html)
        self.assertIn('3.2 选择需要比对的字段', html)
        self.assertIn('未勾选的字段不会进入差异判断', html)
        self.assertIn('添加不同列名字段', html)
        self.assertIn('Qty 对 Quantity', html)
        self.assertIn('第四步：导出报告', html)
        self.assertIn('data-tool-version="free-bom-compare"', html)
        self.assertIn("tool_versions['free-bom-compare']", html)
        self.assertIn('data-tool-version="customer-hq-compare"', html)
        self.assertIn("tool_versions['customer-hq-compare']", html)
        self.assertIn('data-tool-version="hq-version-compare"', html)
        self.assertIn("tool_versions['hq-version-compare']", html)
        self.assertIn('data-tool-version="machine-hq-version-compare"', html)
        self.assertIn("tool_versions['machine-hq-version-compare']", html)
        self.assertIn('data-tool-version="cadence-hq-compare"', html)
        self.assertIn("tool_versions['cadence-hq-compare']", html)
        self.assertIn('Cadence BOM 预览', html)
        self.assertIn('第二步：确认 Sheet 和表头行', html)
        self.assertIn('第三步：选择匹配键和字段', html)
        self.assertIn('PART_NUMBER 对 料号', html)
        self.assertIn('第四步：导出报告', html)
        self.assertIn('第二步：确认匹配规则和字段映射', html)
        self.assertIn('2.1 选择匹配规则', html)
        self.assertIn('2.2 映射客户 BOM 字段', html)
        self.assertIn('数量会参与差异判断', html)
        self.assertIn('第三步：预览或导出', html)

    def test_bom_compare_field_selector_uses_chips_and_default_whitelist(self):
        js = (WEB_APP / "static" / "js" / "app.js").read_text(encoding="utf-8")
        css = (WEB_APP / "static" / "css" / "app.css").read_text(encoding="utf-8")
        for field in [
            "\u578b\u53f7",
            "\u7269\u6599\u63cf\u8ff0",
            "\u5355\u8017",
            "\u66ff\u4ee3\u5173\u7cfb",
            "\u4f4d\u53f7",
            "\u751f\u4ea7\u5382\u5bb6",
        ]:
            self.assertIn(field, js)
        self.assertIn("bomShouldDefaultCompareField", js)
        self.assertIn("bomCompareFieldChip", js)
        self.assertIn("disabled?' disabled':'", js)
        self.assertIn(":checked:not(:disabled)", js)
        self.assertIn(".chk-list .field-chip", css)
        self.assertIn(".chk-list .field-chip.disabled", css)
        self.assertIn("input:checked", css)


if __name__ == "__main__":
    unittest.main()
