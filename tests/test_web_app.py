import io
import json
import sys
import unittest
import uuid
from pathlib import Path
from unittest.mock import patch
from urllib.parse import unquote

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WEB_APP = ROOT / "web_app2"
TESTS = ROOT / "tests"
if str(TESTS) not in sys.path:
    sys.path.insert(0, str(TESTS))
from test_env import configure_test_environment  # noqa: E402
configure_test_environment()
if str(WEB_APP) not in sys.path:
    sys.path.insert(0, str(WEB_APP))

from app import app  # noqa: E402
from bom_checklist import _run_checks as _run_bom_checklist_checks  # noqa: E402
from bom_compare import _save_uploaded_hq_excel  # noqa: E402
from feishu import _hq_read_sheet, _is_preferred_level, _write_cache  # noqa: E402
from manufacturer_alias import normalize_manufacturer_name  # noqa: E402
from shared import _save_uploaded_excel  # noqa: E402


def _xlsx_bytes(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _hq_export_bytes(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(["料号", "PROJECT", "描述", "DESC", "项目配置名", "CFG"])
    ws.append(["版本", "I.1", "替代项", "", "BOM名称", "BOM"])
    ws.append(["序号", "料号", "型号", "物料描述", "单耗", "替代关系", "位号", "生产厂家"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _plm_full_bom_bytes(sheet_rows):
    bom = 'BOM'
    dbg = 'DBG\u4e1a\u52a1BOM'
    ctrl = 'DBGBOM\u5236\u63a7\u4fe1\u606f'
    base_headers = [
        '\u5e8f\u53f7', '\u6599\u53f7', '\u4e0a\u9636BOM\u540d\u79f0', 'BOM\u5c42\u7ea7', '\u578b\u53f7', '\u7269\u6599\u63cf\u8ff0', '\u5355\u8017', '\u66ff\u4ee3\u5173\u7cfb', '\u4f4d\u53f7', '\u751f\u4ea7\u5382\u5bb6',
        '\u662f\u5426\u73af\u4fdd', '\u6e7f\u654f\u5c5e\u6027', '\u7269\u6599\u66f4\u65b0\u65f6\u95f4', 'MOQ', '\u63d0\u524d\u671f', 'ECO_NO', 'ECN\u751f\u6548\u65f6\u95f4', '\u8d28\u91cf\u6807\u51c6',
        '\u4e34\u65f6\u6280\u672f\u5c01\u6837\u65e5\u671f', '\u5b9e\u9645\u6280\u672f\u5c01\u6837\u65e5\u671f', '\u5c01\u6837\u5c5e\u6027', '\u5c01\u6837\u5907\u6ce8', '\u5de5\u7a0b\u5e08\u5907\u6ce8', '\u73af\u4fdd\u8d44\u6599', '\u73af\u4fdd\u9f50\u5957',
        '\u5ba2\u6237\u96f6\u4ef6\u7f16\u7801', '\u5ba2\u6237\u96f6\u4ef6\u540d\u79f0', '\u5ba2\u6237\u751f\u4ea7\u5de5\u5382\u7269\u6599\u6599\u53f7', '\u539f\u4ea7\u56fd', '\u96f6\u4ef6\u5236\u9020\u5546', '\u96f6\u4ef6\u5236\u9020\u5546\u5730\u5740', '\u5355\u4f4d\u51c0\u91cd',
        '\u5ba2\u6237\u6599\u53f7', '\u7814\u53d1\u5c5e\u6027', '\u6700\u5c0f\u5305\u88c5(\u5173\u8054\u4e0b\u5355)', '\u6700\u5c0f\u5305\u88c5(\u4e0d\u5173\u8054\u4e0b\u5355)', '\u6700\u5c0f\u8bf7\u8d2d\u91cf(\u5173\u8054\u4e0b\u5355)', '\u6700\u5c0f\u8bf7\u8d2d\u91cf(\u4e0d\u5173\u8054\u4e0b\u5355)',
        '\u751f\u4ea7\u5382\u5bb6\uff08\u82f1\u6587\uff09', '\u751f\u4ea7\u5382\u5bb6\uff08\u4e2d\u6587\uff09', '\u8fdb\u53e3\u5546', '\u6e7f\u654f\u7b49\u7ea7', '\u5ba2\u6237\u96f6\u4ef6\u540d\u79f0(\u4e2d\u6587)', '\u7269\u6599\u63cf\u8ff0-\u82f1\u6587', '\u5236\u9020\u5546-\u82f1\u6587',
        '\u6e7f\u654f\u5c5e\u6027-\u82f1\u6587', '\u4e0b\u5355\u6599\u53f7', '\u8f6f\u4ef6\u7248\u672c', '\u786c\u4ef6\u7248\u672c', 'MODEL', '\u8d27\u8fd0\u7ec4\u7ec7', '\u4e3b\u8f85BOM\u5b9a\u4e49', '\u6781\u6027&1\u811a\u4f4d\u7f6e',
        '\u5668\u4ef6\u95f4\u8ddd(P1)', '\u8f7d\u5e26\u5bbd\u5ea6(W)', '\u6761\u7801\u89c4\u52191', '\u6761\u7801\u89c4\u52192', '\u6761\u7801\u89c4\u52193', 'HW\u5ba2\u6237\u7269\u6599\u5c5e\u6027', '\u82f1\u6587\u63cf\u8ff0', '\u4f9b\u5e94\u5546',
        '\u6570\u636e\u6cbb\u7406\u6807\u8bb0', 'MBG\u4f18\u9009\u5c5e\u6027', 'CBG\u4f18\u9009\u5c5e\u6027', 'DBG\u4f18\u9009\u5c5e\u6027', '\u91c7\u8d2d\u6a21\u5f0f', '\u662f\u5426\u53ef\u91cf\u4ea7\u4e0b\u5355', 'ABG\u4f18\u9009\u5c5e\u6027', 'IFM_PART', 'PCD_PART', 'ECCN',
    ]
    extras = {
        bom: [],
        dbg: ['\u5355\u4f4d\u603b\u7528\u91cf', '\u9996\u5236\u7a0b', '\u6b21\u5236\u7a0b', '\u6b21\u5236\u7a0b\u5355\u8017', '\u6b21\u5236\u7a0b\u4f4d\u53f7'],
        ctrl: ['\u5355\u4f4d\u603b\u7528\u91cf', '\u635f\u8017\u7387', '\u56fa\u5b9a\u5907\u635f\u503c', '\u9996\u5236\u7a0b', '\u6b21\u5236\u7a0b', '\u6b21\u5236\u7a0b\u5355\u8017', '\u6b21\u5236\u7a0b\u4f4d\u53f7'],
    }
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in [bom, dbg, ctrl]:
        ws = wb.create_sheet(sheet_name)
        ws.append(['BOM\u5386\u53f2\u4fee\u6539\u8bb0\u5f55\u62a5\u8868'])
        ws.append([])
        ws.append(['\u63d0\u4ea4\u65f6\u95f4', '2026-05-13 14:50:41'])
        ws.append(['\u91cf\u4ea7/\u8bd5\u4ea7', '\u8bd5\u4ea7'])
        ws.append([None] * 12)
        ws.append(['\u6599\u53f7', 'HQ31200063SB0', '\u63cf\u8ff0', 'Demo', '\u9879\u76ee\u914d\u7f6e\u540d', 'NS8551AAA', '\u5de5\u7a0b\u5e08', 'Tester'])
        ws.append(['\u7248\u672c', 'I.4', '\u66ff\u4ee3\u9879', '', 'BOM\u540d\u79f0', 'Demo full BOM', '\u5f52\u6863\u90e8\u95e8', 'SYSTEMADMIN'])
        headers = base_headers + extras[sheet_name]
        ws.append(headers)
        for row in sheet_rows.get(sheet_name, sheet_rows[bom]):
            data = {'\u5e8f\u53f7': row[0], '\u6599\u53f7': row[1], '\u4e0a\u9636BOM\u540d\u79f0': 'HQ31200063SB0', 'BOM\u5c42\u7ea7': '1', '\u578b\u53f7': row[2], '\u7269\u6599\u63cf\u8ff0': row[3], '\u5355\u8017': row[4], '\u66ff\u4ee3\u5173\u7cfb': row[5], '\u4f4d\u53f7': row[6], '\u751f\u4ea7\u5382\u5bb6': row[7]}
            data.update(row[8] if len(row) > 8 else {})
            ws.append([data.get(h, '') for h in headers])
        ws.append([])
        ws.append(['BOM\u7248\u672c', '\u65e5\u671f', '\u4fee\u8ba2\u88c5\u914d\u4ef6', '\u4fee\u8ba2\u7ec4\u4ef6', '\u7ec4\u4ef6\u578b\u53f7', '\u7ec4\u4ef6\u63cf\u8ff0', '\u52a8\u4f5c\uff08\u65b0\u589e\uff0c\u66f4\u6539\uff0c\u7981\u7528\uff09', '\u4fee\u8ba2\u66ff\u4ee3\u4ef6', '\u66ff\u4ee3\u4ef6\u578b\u53f7', '\u66ff\u4ee3\u4ef6\u63cf\u8ff0', '\u52a8\u4f5c\uff08\u65b0\u589e\uff0c\u7981\u7528\uff09', '\u65e7\u6570\u91cf', '\u65b0\u6570\u91cf', '\u65e7\u4f4d\u53f7', '\u65b0\u4f4d\u53f7'])
        ws.append(['I.4', '2026-05-12', 'HQ31200063SB0', 'HISTORY', 'HistoryModel', 'HistoryDesc', '\u65b0\u589e\u7269\u6599', '', '', '', '', '', '1', '', 'H1'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf



class WebAppTests(unittest.TestCase):
    def test_index_renders_main_shell(self):
        resp = app.test_client().get("/")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn("BOM Tools", html)
        self.assertIn("\u786c\u4ef6\u8bbe\u8ba1\u8f85\u52a9\u5e73\u53f0 v2.2.6", html)
        self.assertIn("css/app.css?v=2.2.6", html)
        self.assertIn("js/app.js?v=2.2.6", html)
        self.assertIn("version: \"2.2.6\"", html)
        self.assertIn("toolVersions", html)
        self.assertIn("currentUser", html)
        self.assertIn("free-bom-compare", html)
        self.assertIn("1.1.4", html)
        self.assertIn("飞书优选库+关系库匹配", html)
        self.assertIn("BOM比对工具合集", html)
        self.assertIn("\u5c0f\u5de5\u5177\u5408\u96c6", html)
        self.assertIn("\u8ba1\u7b97\u54c8\u5e0c\u503c", html)
        self.assertIn("单板HQ BOM版本对比", html)
        self.assertIn("整机HQ BOM版本对比", html)
        self.assertIn("Cadence导出BOM对比HQ BOM", html)
        self.assertIn("\u5ba2\u6237BOM\u8f6c\u6362\u6210HQ\u683c\u5f0f\u5355\u677fBOM", html)


    def test_plm_auto_defaults_account_from_logged_in_employee_id(self):
        html = (WEB_APP / "templates" / "partials" / "tools" / "plm-auto.html").read_text(encoding="utf-8")
        js = (WEB_APP / "static" / "js" / "app.js").read_text(encoding="utf-8")
        self.assertIn('id="paUser"', html)
        self.assertIn('id="paAttUser"', html)
        self.assertIn('id="paSingleValue"', html)
        self.assertIn('id="paProgressPanel"', html)
        self.assertIn('id="paRunSingle"', html)
        self.assertIn('查询方式一：规格型号 / HQ 料号直接查询', html)
        self.assertIn('多个规格型号或 HQ 料号可用半角逗号或全角逗号分隔', html)
        self.assertIn('中间空格会保留', html)
        self.assertIn('查询方式二：Excel 批量查询', html)
        self.assertIn('执行进度与结果', html)
        self.assertNotIn('id="paRun"', html)
        self.assertNotIn('value="100448405"', html)
        self.assertIn("function currentEmployeeId()", js)
        self.assertIn("BOM_TOOLS_BOOTSTRAP.currentUser", js)
        self.assertIn("$('paUser').value = currentEmployeeId();", js)
        self.assertIn("$('paAttUser').value = currentEmployeeId();", js)
        self.assertIn("function plmAutoSetProgress", js)
        self.assertIn("single_value", js)
    def test_about_project_nav_is_last_and_hides_direct_contact_sentence(self):
        resp = app.test_client().get("/")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        nav_start = html.index("<nav>")
        nav_end = html.index("</nav>", nav_start)
        nav_html = html[nav_start:nav_end]
        self.assertGreater(nav_html.index('data-tool="manual"'), nav_html.index('data-tool="admin-users"'))
        self.assertGreater(nav_html.index('data-tool="about-project"'), nav_html.index('data-tool="manual"'))
        self.assertNotIn("\u8054\u7cfb\u65b9\u5f0f\uff1a\u98de\u4e66\u641c\u7d22\u5de5\u53f7 100448405", html)
        self.assertNotIn("\u98de\u4e66\u641c\u7d22\u5de5\u53f7 100448405 \u6216\u59d3\u540d", html)

    def test_about_project_contains_polished_easter_egg(self):
        resp = app.test_client().get("/")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn('id="aboutEggTrigger"', html)
        self.assertIn('id="aboutEggPanel"', html)
        self.assertIn('class="about-egg show"', html)
        self.assertIn('class="about-egg-trigger active"', html)
        self.assertIn('aria-hidden="false"', html)
        self.assertIn("BOM TOOLS / INTERNAL BUILD", html)
        self.assertIn("Keep the BOM clean.", html)
        css = (WEB_APP / "static" / "css" / "app.css").read_text(encoding="utf-8")
        self.assertIn(".about-egg-trigger", css)
        self.assertIn("@keyframes aboutPulse", css)
        js = (WEB_APP / "static" / "js" / "app.js").read_text(encoding="utf-8")
        self.assertIn("function initAboutProject", js)
        self.assertIn("tool==='about-project'", js)

    def test_bom_checklist_nav_and_api_detect_basic_issues(self):
        resp = app.test_client().get("/")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn('data-tool="bom-checklist"', html)
        self.assertIn("BOM Checklist", html)
        self.assertIn("当前检查项说明", html)

        data = {
            "file": (_xlsx_bytes(["料号", "型号", "型号"], [["A", "M1", "M1"], [None, None, None], ["B", "M2", "M2"]]), "check.xlsx"),
            "header_row": "1",
        }
        payload = app.test_client().post(
            "/api/bom_checklist/run",
            data=data,
            content_type="multipart/form-data",
        ).get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["duplicate_headers"]["status"], "fail")
        self.assertEqual(checks["blank_rows"]["status"], "warn")
        self.assertEqual(checks["required_headers"]["status"], "warn")
        self.assertEqual(payload["summary"]["data_rows"], 2)

    def test_bom_checklist_web_documents_every_backend_check(self):
        html = app.test_client().get("/").get_data(as_text=True)
        headers = ["序号", "料号", "上阶BOM名称", "BOM层级", "型号", "物料描述", "单耗", "替代关系", "位号", "生产厂家", "是否环保", "湿敏属性"]
        data_rows = [
            (2, ["1", "HQ-PCB", "ROOT", "1", "PCB", "PCB main", "1", "主料", "PCB", "HQ", "I", "非湿敏器件"]),
        ]
        _, checks = _run_bom_checklist_checks(headers, data_rows, [])

        for check in checks:
            self.assertIn(f'data-check-id="{check["id"]}"', html)

    def test_bom_checklist_pcba_rules_detect_structural_issues(self):
        headers = ["序号", "料号", "上阶BOM名称", "BOM层级", "型号", "物料描述", "单耗", "替代关系", "位号", "生产厂家", "是否环保", "湿敏属性"]
        rows = [
            ["1", "HQ-PCBA", "ROOT", 1, "PCBA", "PCB main", 1, "主料", "", "HQ", "I", "非湿敏器件"],
            ["2", "HQ-DEPOP", "ROOT", 1, "R1", "DEPOP resistor", "", "主料", "R1", "Maker", "I", "非湿敏器件"],
            ["3", "HQ-ALT", "ROOT", 1, "C1", "capacitor", 1, "替代料", "C1", "Maker", "I", "非湿敏器件"],
            ["1", "HQ-CHILD", "MISSING-PARENT", 2, "U1", "child", 1, "主料", "U1", "Maker", "I", "非湿敏器件"],
        ]
        data = {
            "file": (_xlsx_bytes(headers, rows), "pcba-bad.xlsx"),
            "header_row": "1",
        }
        payload = app.test_client().post("/api/bom_checklist/run", data=data, content_type="multipart/form-data").get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["pcba_standard_headers"]["status"], "pass")
        self.assertEqual(checks["pcba_depop_removed"]["status"], "fail")
        self.assertEqual(checks["pcba_pcb_location"]["status"], "fail")
        self.assertEqual(checks["pcba_qty_by_substitute_type"]["status"], "fail")
        self.assertEqual(checks["pcba_parent_bom_reference"]["status"], "fail")


    def test_bom_checklist_chip_labels_detect_missing_and_stage(self):
        headers = ["序号", "料号", "上阶BOM名称", "BOM层级", "型号", "物料描述", "单耗", "替代关系", "位号", "生产厂家", "是否环保", "湿敏属性"]
        chip_rows = [
            ["1", "HQ-BMC", "PCBA-ROOT", 1, "AST2600", "BMC AST2600 controller", 1, "主料", "U1", "Aspeed", "I", "非湿敏器件"],
            ["2", "HQ-MAC", "PCBA-ROOT", 1, "WGI210AT", "网卡芯片_MAC_千兆", 1, "主料", "U2", "Intel", "I", "非湿敏器件"],
        ]
        payload = app.test_client().post("/api/bom_checklist/run", data={"file": (_xlsx_bytes(headers, chip_rows), "missing-label.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["pcba_bmc_bios_auth_labels"]["status"], "fail")
        self.assertIn("HQ11111801009", checks["pcba_bmc_bios_auth_labels"]["message"])
        self.assertIn("MAC 标签", checks["pcba_bmc_bios_auth_labels"]["message"])
        self.assertIn("SN 标签", checks["pcba_bmc_bios_auth_labels"]["message"])

        loose_rows = chip_rows + [
            ["3", "HQ60410125009", "ROOT", 2, "BIOS LABEL", "AMI BIOS label", "", "主料", "", "AMI", "I", "非湿敏器件"],
            ["4", "HQ11111801009", "ROOT", 2, "BMC LABEL", "AST2600 label", "", "主料", "", "AMI", "I", "非湿敏器件"],
            ["5", "HQ-MAC-LABEL", "ROOT", 2, "MAC LABEL", "MAC address label", "", "主料", "", "", "I", "非湿敏器件"],
            ["6", "HQ-SN-LABEL", "ROOT", 2, "SN LABEL", "Serial Number label", "", "主料", "", "", "I", "非湿敏器件"],
        ]
        payload = app.test_client().post("/api/bom_checklist/run", data={"file": (_xlsx_bytes(headers, loose_rows), "loose-label.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["pcba_bmc_bios_auth_labels"]["status"], "warn")
        self.assertIn("组装虚拟阶", checks["pcba_bmc_bios_auth_labels"]["message"])

        assembly_rows = chip_rows + [
            ["3", "HQ60410125009", "PCBA-ROOT", 1, "PCBA ASSY BIOS LABEL", "AMI BIOS label", "", "主料", "BIOS_LABEL", "AMI", "I", "非湿敏器件"],
            ["4", "HQ11111801009", "PCBA-ROOT", 1, "PCBA ASSY BMC LABEL", "AST2600 label", "", "主料", "BMC_LABEL", "AMI", "I", "非湿敏器件"],
            ["5", "HQ-MAC-LABEL", "PCBA-ROOT", 1, "PCBA ASSY MAC LABEL", "MAC address label", "", "主料", "MAC_LABEL", "", "I", "非湿敏器件"],
            ["6", "HQ-SN-LABEL", "PCBA-ROOT", 1, "PCBA ASSY SN LABEL", "Serial Number label", "", "主料", "SN_LABEL", "", "I", "非湿敏器件"],
        ]
        payload = app.test_client().post("/api/bom_checklist/run", data={"file": (_xlsx_bytes(headers, assembly_rows), "assembly-label.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["pcba_bmc_bios_auth_labels"]["status"], "pass")


    def test_bom_checklist_chip_labels_use_standard_pcba_sample(self):
        sample_path = ROOT / "samples" / "pcba_bom_standard_sample.xlsx"
        with sample_path.open("rb") as f:
            payload = app.test_client().post(
                "/api/bom_checklist/run",
                data={"file": (io.BytesIO(f.read()), "pcba_bom_standard_sample.xlsx"), "header_row": "8"},
                content_type="multipart/form-data",
            ).get_json()
        checks = {item["id"]: item for item in payload["checks"]}
        label_check = checks["pcba_bmc_bios_auth_labels"]
        self.assertEqual(label_check["status"], "fail")
        self.assertIn("HQ11111788009", label_check["message"])
        self.assertIn("MAC 标签", label_check["message"])
        self.assertIn("SN 标签", label_check["message"])
        self.assertEqual(label_check["rows"], [11, 224])


    def test_bom_checklist_flash_socket_and_smt_flash_are_mutually_exclusive(self):
        headers = ["序号", "料号", "上阶BOM名称", "BOM层级", "型号", "物料描述", "单耗", "替代关系", "位号", "生产厂家", "是否环保", "湿敏属性"]
        socket_row = ["1", "HQ-FLASH-SOCKET", "PCBA-ROOT", 1, "FLASH SOCKET", "SPI Flash Socket", 1, "主料", "U100", "Maker", "I", "非湿敏器件"]
        flash_row = ["2", "HQ-SPI-FLASH", "PCBA-ROOT", 1, "W25Q128JVSIQ", "SPI Flash_128Mbit_SOIC-8", 1, "主料", "U101", "Winbond", "I", "非湿敏器件"]

        payload = app.test_client().post("/api/bom_checklist/run", data={"file": (_xlsx_bytes(headers, [socket_row, flash_row]), "flash-conflict.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["pcba_flash_socket_smt_conflict"]["status"], "fail")
        self.assertEqual(checks["pcba_flash_socket_smt_conflict"]["rows"], [2, 3])

        payload = app.test_client().post("/api/bom_checklist/run", data={"file": (_xlsx_bytes(headers, [socket_row]), "flash-socket.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["pcba_flash_socket_smt_conflict"]["status"], "warn")
        self.assertIn("colay", checks["pcba_flash_socket_smt_conflict"]["message"])

        payload = app.test_client().post("/api/bom_checklist/run", data={"file": (_xlsx_bytes(headers, [flash_row]), "flash-smt.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        checks = {item["id"]: item for item in payload["checks"]}
        self.assertEqual(checks["pcba_flash_socket_smt_conflict"]["status"], "pass")


    def test_bom_detect_returns_fifty_preview_rows(self):
        rows = [["PN-%02d" % i, "Model-%02d" % i] for i in range(1, 56)]
        data = {
            "file": (_xlsx_bytes(["PN", "Model"], rows), "preview.xlsx"),
            "header_row": "1",
        }
        resp = app.test_client().post(
            "/api/bom/detect",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(len(payload["preview"]), 50)
        self.assertEqual(payload["preview"][0], ["PN-01", "Model-01"])
        self.assertEqual(payload["preview"][-1], ["PN-50", "Model-50"])

    def test_download_rejects_missing_and_traversal_paths(self):
        client = app.test_client()
        self.assertEqual(client.get("/download/not-found.xlsx").status_code, 404)
        self.assertEqual(client.get("/download/../app.py").status_code, 404)

    def test_preferred_level_does_not_treat_non_preferred_as_preferred(self):
        self.assertTrue(_is_preferred_level("优选"))
        self.assertTrue(_is_preferred_level("7"))
        self.assertTrue(_is_preferred_level("Preferred"))
        self.assertFalse(_is_preferred_level("非优选"))
        self.assertFalse(_is_preferred_level("不优选"))
        self.assertFalse(_is_preferred_level("6"))
        self.assertFalse(_is_preferred_level(""))

    def test_pref_rate_rejects_invalid_header_row_before_excel_processing(self):
        data = {
            "file": (_xlsx_bytes(["HQ料号"], [["HQ-1"]]), "bom.xlsx"),
            "config": json.dumps({"header_row": "bad", "local_key_col": "HQ料号"}),
        }
        resp = app.test_client().post(
            "/api/feishu/pref_rate",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(payload["success"])
        self.assertIn("表头行", payload["error"])


    def test_feishu_local_sheets_uses_selected_header_row(self):
        data = {
            "file": (_xlsx_bytes(["not", "headers"], [["PN", "Maker"], ["A", "M"]]), "local.xlsx"),
            "header_row": "2",
        }
        resp = app.test_client().post(
            "/api/feishu/local_sheets",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["headers"], ["PN", "Maker"])


    def test_feishu_local_sheets_reuses_uploaded_file_by_uid(self):
        client = app.test_client()
        source = _xlsx_bytes(["not", "headers"], [["PN", "Maker"], ["A", "M"]])
        first = client.post(
            "/api/feishu/local_sheets",
            data={"file": (source, "local.xlsx"), "header_row": "1"},
            content_type="multipart/form-data",
        ).get_json()
        self.assertTrue(first["success"], first.get("error"))
        self.assertEqual(first["headers"], ["not", "headers"])

        second = client.post(
            "/api/feishu/local_sheets",
            data={"uid": first["uid"], "header_row": "2", "sheet_name": first["current_sheet"]},
            content_type="multipart/form-data",
        ).get_json()
        self.assertTrue(second["success"], second.get("error"))
        self.assertEqual(second["uid"], first["uid"])
        self.assertEqual(second["headers"], ["PN", "Maker"])


    def test_bom_detect_reuses_uploaded_file_by_uid(self):
        client = app.test_client()
        source = _xlsx_bytes(["not", "headers"], [["Part", "Qty"], ["A", 1]])
        first = client.post("/api/bom/detect", data={"file": (source, "bom.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        self.assertTrue(first["success"], first.get("error"))
        second = client.post("/api/bom/detect", data={"uid": first["uid"], "header_row": "2", "sheet_name": first["current_sheet"]}, content_type="multipart/form-data").get_json()
        self.assertTrue(second["success"], second.get("error"))
        self.assertEqual(second["uid"], first["uid"])
        self.assertIn("Part", second["headers"])

    def test_plm_detect_reuses_uploaded_file_by_uid(self):
        client = app.test_client()
        source = _xlsx_bytes(["not", "headers"], [["SEQ", "HQ PN", "QTY"], ["1", "A", 1]])
        first = client.post("/api/plm/detect", data={"file": (source, "plm.xlsx"), "header_row": "1"}, content_type="multipart/form-data").get_json()
        self.assertTrue(first["success"], first.get("error"))
        second = client.post("/api/plm/detect", data={"uid": first["uid"], "header_row": "2", "sheet_name": first["current_sheet"]}, content_type="multipart/form-data").get_json()
        self.assertTrue(second["success"], second.get("error"))
        self.assertEqual(second["uid"], first["uid"])
        self.assertTrue(any("HQ PN" in header for header in second["headers"]))


    def test_plm_spec_extract_deduplicates_repeated_values(self):
        client = app.test_client()
        resp = client.post(
            "/api/plm/spec_extract",
            data={
                "file": (_xlsx_bytes(
                    ["料号", "HQ料号"],
                    [["A 001", ""], ["A001", ""], ["B002", ""], ["C003", "已存在"]],
                ), "spec.xlsx"),
                "config": json.dumps({
                    "sheet_name": "Sheet",
                    "header_row": 1,
                    "col_name": "料号",
                    "exclude_col_name": "HQ料号",
                }),
            },
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["count"], 2)
        self.assertEqual(payload["skipped_duplicates"], 1)
        self.assertEqual(payload["skipped_excluded"], 1)

        download = client.get(payload["download"])
        self.assertEqual(download.status_code, 200)
        download_buf = io.BytesIO(download.data)
        wb = openpyxl.load_workbook(download_buf, data_only=True)
        values = [wb.active.cell(row=i, column=1).value for i in range(2, wb.active.max_row + 1)]
        wb.close()
        download_buf.close()
        self.assertEqual(values, ["A001", "B002"])

    def test_customer_hq_detect_reuses_uploaded_file_by_uid(self):
        client = app.test_client()
        source = _xlsx_bytes(["not", "headers", "here"], [["Seq", "Brand", "Qty"], ["1", "Maker", 2]])
        first = client.post(
            "/api/plm/customer_hq_detect",
            data={"file": (source, "customer.xlsx"), "header_row": "1"},
            content_type="multipart/form-data",
        ).get_json()
        self.assertTrue(first["success"], first.get("error"))
        second = client.post(
            "/api/plm/customer_hq_detect",
            data={"uid": first["uid"], "header_row": "2", "sheet_name": first["current_sheet"]},
            content_type="multipart/form-data",
        ).get_json()
        self.assertTrue(second["success"], second.get("error"))
        self.assertEqual(second["uid"], first["uid"])
        self.assertTrue(any("Brand" in header for header in second["headers"]))



    def test_feishu_sheet_read_uses_header_width_and_pads_sparse_rows(self):
        calls = []

        class FakeResponse:
            def __init__(self, values):
                self.values = values

            def raise_for_status(self):
                return None

            def json(self):
                return {"code": 0, "data": {"valueRange": {"values": self.values}}}

        def fake_get(url, params=None, timeout=None):
            range_name = (params or {}).get("range", "")
            calls.append(range_name)
            if range_name == "sheet!A1:Z1":
                return FakeResponse([["PN", "Desc", "HQ PN", ""]])
            if range_name == "sheet!A1:C3":
                return FakeResponse([["PN", "Desc", "HQ PN"], ["A", "D1"], ["B", "D2", "HQ-B"]])
            return FakeResponse([])

        with patch("feishu._requests.get", side_effect=fake_get):
            rows = _hq_read_sheet("https://example.test", "origin", "user", "token", "sheet", row_count=3, col_count=26)

        self.assertEqual(rows, [["PN", "Desc", "HQ PN"], ["A", "D1", ""], ["B", "D2", "HQ-B"]])
        self.assertNotIn("sheet!A1:Z3", calls)


    def test_feishu_sheet_read_falls_back_to_column_ranges_when_block_read_fails(self):
        class FakeResponse:
            def __init__(self, values):
                self.values = values

            def raise_for_status(self):
                return None

            def json(self):
                return {"code": 0, "data": {"valueRange": {"values": self.values}}}

        def fake_get(url, params=None, timeout=None):
            range_name = (params or {}).get("range", "")
            if range_name == "sheet!A1:Z1":
                return FakeResponse([["PN", "Desc"]])
            if range_name == "sheet!A1:B2":
                raise RuntimeError("range failed")
            if range_name == "sheet!A1:A2":
                return FakeResponse([["PN"], ["A"]])
            if range_name == "sheet!B1:B2":
                return FakeResponse([["Desc"], ["fallback"]])
            return FakeResponse([])

        with patch("feishu._requests.get", side_effect=fake_get):
            rows = _hq_read_sheet("https://example.test", "origin", "user", "token", "sheet", row_count=2, col_count=26)

        self.assertEqual(rows, [["PN", "Desc"], ["A", "fallback"]])


    def test_feishu_clear_cache_deletes_server_cache_file(self):
        suffix = uuid.uuid4().hex[:8]
        token = f"token-clear-{suffix}"
        sheet_id = f"sid-clear-{suffix}"
        key, _, _ = _write_cache(token, sheet_id, [["PN"], ["A"]])
        cache_path = WEB_APP / "cache" / f"feishu_{key}.json"
        self.assertTrue(cache_path.exists())

        payload = app.test_client().post("/api/feishu/cache/clear", json={
            "token": token,
            "sheet_id": sheet_id,
        }).get_json()

        self.assertTrue(payload["success"], payload.get("error"))
        self.assertTrue(payload["deleted"])
        self.assertFalse(cache_path.exists())

        invalid = app.test_client().post("/api/feishu/cache/clear", json={"token": None})
        self.assertEqual(invalid.status_code, 200)
        self.assertFalse(invalid.get_json()["success"])
        self.assertEqual(invalid.get_json()["error"], "\u8bf7\u63d0\u4f9b Token \u548c Sheet ID")



    def test_feishu_match_reports_missing_model_local_mapping(self):
        data = {
            "file": (_xlsx_bytes(["Model", "Maker"], [["A", "M"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [{
                    "name": "MLCC",
                    "token": "token-missing-model",
                    "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-missing-model",
                        "sheet_name": "Preferred",
                        "local_key_names": ["", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                    }],
                }],
            }),
        }
        payload = app.test_client().post("/api/feishu/match", data=data, content_type="multipart/form-data").get_json()
        self.assertFalse(payload["success"])
        self.assertIn("\u672a\u9009\u62e9\u578b\u53f7", payload["error"])
        self.assertNotIn("\u6ca1\u6709\u53ef\u7528", payload["error"])


    def test_feishu_match_reports_missing_maker_local_mapping(self):
        data = {
            "file": (_xlsx_bytes(["Model", "Maker"], [["A", "M"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [{
                    "name": "MLCC",
                    "token": "token-missing-maker",
                    "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-missing-maker",
                        "sheet_name": "Preferred",
                        "local_key_names": ["Model", ""],
                        "feishu_key_names": ["PN", "Maker"],
                    }],
                }],
            }),
        }
        payload = app.test_client().post("/api/feishu/match", data=data, content_type="multipart/form-data").get_json()
        self.assertFalse(payload["success"])
        self.assertIn("\u672a\u9009\u62e9\u5382\u5546", payload["error"])
        self.assertNotIn("\u6ca1\u6709\u53ef\u7528", payload["error"])


    def test_feishu_match_deduplicates_identical_rows_and_merges_sources(self):
        key_a, _, _ = _write_cache("token-a", "sid-a", [
            ["PN", "Maker", "HQ PN", "Model"],
            ["A", "MakerA", "HQ-1", "M1"],
            ["A", "MakerA", "HQ-1", "M1"],
            ["A", "MakerA", "HQ-2", "M2"],
        ])
        key_b, _, _ = _write_cache("token-b", "sid-b", [
            ["PN", "Maker", "HQ PN", "Model"],
            ["A", "MakerA", "HQ-1", "M1"],
        ])
        fetch_map = [
            {"output": "HQ PN", "alias": "HQ PN"},
            {"output": "Model", "alias": "Model"},
        ]
        data = {
            "file": (_xlsx_bytes(["PN", "Maker"], [["A", "MakerA"], ["B", "MakerB"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [
                    {"name": "LibA", "token": "token-a", "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-a",
                        "sheet_name": "SheetA",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                        "fetch_col_map": fetch_map,
                        "cache_key": key_a,
                    }]},
                    {"name": "LibB", "token": "token-b", "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-b",
                        "sheet_name": "SheetB",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                        "fetch_col_map": fetch_map,
                        "cache_key": key_b,
                    }]},
                ],
            }),
        }
        resp = app.test_client().post(
            "/api/feishu/match",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["total"], 2)
        self.assertEqual(payload["matched"], 1)
        self.assertEqual(payload["unmatched"], 1)

        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb.active
        self.assertEqual([cell.value for cell in ws[1]], ["PN", "Maker", "HQ PN", "Model", "\u6765\u6e90\u8868\u683c"])
        self.assertEqual([ws["A2"].value, ws["B2"].value, ws["C2"].value, ws["D2"].value], ["A", "MakerA", "HQ-1", "M1"])
        self.assertEqual(ws["E2"].value, "LibA - SheetA\uff1bLibB - SheetB")
        self.assertEqual([ws["A3"].value, ws["B3"].value, ws["C3"].value, ws["D3"].value, ws["E3"].value], [None, None, "HQ-2", "M2", "LibA - SheetA"])
        self.assertEqual([ws["A4"].value, ws["B4"].value, ws["C4"].value, ws["D4"].value, ws["E4"].value], ["B", "MakerB", None, None, "\u672a\u5339\u914d"])
        wb.close()


    def test_feishu_match_prefers_relation_table_over_preferred_library(self):
        key_pref, _, _ = _write_cache("token-pref", "sid-pref", [
            ["PN", "Maker", "HQ PN", "Model"],
            ["A", "MakerA", "HQ-PREF", "M-PREF"],
        ])
        key_rel, _, _ = _write_cache("token-rel", "sid-rel", [
            ["PN", "Maker", "HQ PN", "Model"],
            ["A", "MakerA", "HQ-REL", "M-REL"],
        ])
        fetch_map = [
            {"output": "HQ PN", "alias": "HQ PN"},
            {"output": "Model", "alias": "Model"},
        ]
        data = {
            "file": (_xlsx_bytes(["PN", "Maker"], [["A", "MakerA"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [
                    {"name": "MLCC Preferred Library", "token": "token-pref", "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-pref",
                        "sheet_name": "Preferred",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                        "fetch_col_map": fetch_map,
                        "cache_key": key_pref,
                    }]},
                    {"name": "\u5ba2\u6237\u7269\u6599\u578b\u53f7\u4e0eHQ\u6599\u53f7\u5bf9\u5e94\u5173\u7cfb", "token": "token-rel", "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-rel",
                        "sheet_name": "\u5b57\u8282",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                        "fetch_col_map": fetch_map,
                        "cache_key": key_rel,
                    }]},
                ],
            }),
        }
        resp = app.test_client().post(
            "/api/feishu/match",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["matched"], 1)
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb.active
        self.assertEqual(ws.max_row, 2)
        self.assertEqual([ws["A2"].value, ws["B2"].value, ws["C2"].value, ws["D2"].value], ["A", "MakerA", "HQ-REL", "M-REL"])
        self.assertEqual(ws["E2"].value, "\u5ba2\u6237\u7269\u6599\u578b\u53f7\u4e0eHQ\u6599\u53f7\u5bf9\u5e94\u5173\u7cfb - \u5b57\u8282")
        wb.close()


    def test_feishu_match_can_include_preferred_hits_with_relation_hits(self):
        key_pref, _, _ = _write_cache("token-pref-with-rel", "sid-pref-with-rel", [
            ["PN", "Maker", "HQ PN"],
            ["A", "MakerA", "HQ-PREF"],
        ])
        key_rel, _, _ = _write_cache("token-rel-with-pref", "sid-rel-with-pref", [
            ["PN", "Maker", "HQ PN"],
            ["A", "MakerA", "HQ-REL"],
        ])
        fetch_map = [{"output": "HQ PN", "alias": "HQ PN"}]
        data = {
            "file": (_xlsx_bytes(["PN", "Maker"], [["A", "MakerA"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "include_preferred_with_relation": True,
                "tables": [
                    {"name": "MLCC", "token": "token-pref-with-rel", "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-pref-with-rel",
                        "sheet_name": "Preferred",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                        "fetch_col_map": fetch_map,
                        "enable_recommendations": False,
                        "cache_key": key_pref,
                    }]},
                    {"name": "\u5ba2\u6237\u7269\u6599\u578b\u53f7\u4e0eHQ\u6599\u53f7\u5bf9\u5e94\u5173\u7cfb", "token": "token-rel-with-pref", "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-rel-with-pref",
                        "sheet_name": "\u5173\u7cfb",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                        "fetch_col_map": fetch_map,
                        "cache_key": key_rel,
                    }]},
                ],
            }),
        }
        resp = app.test_client().post("/api/feishu/match", data=data, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb.active
        self.assertEqual(ws.max_row, 3)
        self.assertEqual([ws["C2"].value, ws["C3"].value], ["HQ-PREF", "HQ-REL"])
        wb.close()


    def test_feishu_match_expands_recommendations_by_hq_description_when_enabled(self):
        key, _, _ = _write_cache("token-mlcc-rec", "sid-mlcc-rec", [
            ["Model", "Maker", "Desc", "HQ PN", "PI"],
            ["CAP-A", "MakerA", "D1", "HQ-STRICT", "1"],
            ["CAP-B", "MakerB", "D1", "HQ-REC-HIGH", "9"],
            ["CAP-C", "MakerC", "D1", "HQ-REC-MID", "5"],
            ["CAP-D", "MakerD", "D2", "HQ-OTHER", "9"],
        ])
        fetch_map = [
            {"output": "HQ PN", "alias": "HQ PN"},
            {"output": "HQ\u63cf\u8ff0", "alias": "Desc"},
            {"output": "\u4f18\u9009\u7b49\u7ea7", "alias": "PI"},
        ]
        data = {
            "file": (_xlsx_bytes(["Model", "Maker"], [["CAP-A", "MakerA"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [{
                    "name": "MLCC",
                    "token": "token-mlcc-rec",
                    "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-mlcc-rec",
                        "sheet_name": "Preferred",
                        "local_key_names": ["Model", "Maker"],
                        "feishu_key_names": ["Model", "Maker"],
                        "fetch_col_map": fetch_map,
                        "enable_recommendations": True,
                        "cache_key": key,
                    }],
                }],
            }),
        }
        resp = app.test_client().post("/api/feishu/match", data=data, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb.active
        self.assertEqual(ws.max_row, 4)
        self.assertEqual([ws["C2"].value, ws["C3"].value, ws["C4"].value], ["HQ-STRICT", "HQ-REC-HIGH", "HQ-REC-MID"])
        self.assertIn("\u4f18\u9009\u53ef\u66ff\u4ee3\u63a8\u8350", ws["F3"].value)
        wb.close()


    def test_feishu_match_can_disable_recommendations_per_sheet(self):
        key, _, _ = _write_cache("token-mlcc-rec-off", "sid-mlcc-rec-off", [
            ["Model", "Maker", "Desc", "HQ PN", "PI"],
            ["CAP-A", "MakerA", "D1", "HQ-STRICT", "1"],
            ["CAP-B", "MakerB", "D1", "HQ-REC", "9"],
        ])
        fetch_map = [
            {"output": "HQ PN", "alias": "HQ PN"},
            {"output": "HQ\u63cf\u8ff0", "alias": "Desc"},
            {"output": "\u4f18\u9009\u7b49\u7ea7", "alias": "PI"},
        ]
        data = {
            "file": (_xlsx_bytes(["Model", "Maker"], [["CAP-A", "MakerA"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [{
                    "name": "MLCC",
                    "token": "token-mlcc-rec-off",
                    "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-mlcc-rec-off",
                        "sheet_name": "Preferred",
                        "local_key_names": ["Model", "Maker"],
                        "feishu_key_names": ["Model", "Maker"],
                        "fetch_col_map": fetch_map,
                        "enable_recommendations": False,
                        "cache_key": key,
                    }],
                }],
            }),
        }
        resp = app.test_client().post("/api/feishu/match", data=data, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        self.assertEqual(wb.active.max_row, 2)
        wb.close()


    def test_feishu_match_requires_all_configured_keys_to_be_non_empty(self):
        key, _, _ = _write_cache("token-two-key", "sid-two-key", [
            ["PN", "Maker", "HQ PN"],
            ["A", "", "HQ-EMPTY-MAKER"],
            ["A", "MakerA", "HQ-A"],
        ])
        data = {
            "file": (_xlsx_bytes(["PN", "Maker"], [["A", ""], ["A", "MakerA"]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [{
                    "name": "LibTwoKey",
                    "token": "token-two-key",
                    "sheets": [{
                        "enabled": True,
                        "sheet_id": "sid-two-key",
                        "sheet_name": "SheetTwoKey",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "Maker"],
                        "fetch_col_map": [{"output": "HQ PN", "alias": "HQ PN"}],
                        "cache_key": key,
                    }],
                }],
            }),
        }
        resp = app.test_client().post(
            "/api/feishu/match",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["total"], 2)
        self.assertEqual(payload["matched"], 1)
        self.assertEqual(payload["unmatched"], 1)

        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb.active
        self.assertEqual([ws["A2"].value, ws["B2"].value, ws["C2"].value, ws["D2"].value], ["A", None, None, "\u672a\u5339\u914d"])
        self.assertEqual([ws["A3"].value, ws["B3"].value, ws["C3"].value], ["A", "MakerA", "HQ-A"])
        wb.close()


    def test_feishu_match_maps_local_manufacturer_alias_key(self):
        suffix = uuid.uuid4().hex[:8]
        canonical = f"HQ Maker {suffix}"
        alias = f"Customer Maker {suffix}"
        create_resp = app.test_client().post("/api/manufacturer_aliases", data={
            "canonical_name": canonical,
            "alias": alias,
            "source": "unit-test",
        })
        self.assertTrue(create_resp.get_json()["success"])

        key, _, _ = _write_cache("token-mfg", f"sid-{suffix}", [
            ["PN", "HQ Maker", "HQ PN"],
            ["A", canonical, "HQ-1"],
        ])
        data = {
            "file": (_xlsx_bytes(["PN", "Maker"], [["A", alias]]), "local.xlsx"),
            "config": json.dumps({
                "sheet_name": "Sheet",
                "header_row": 1,
                "tables": [{
                    "name": "LibMfg",
                    "token": "token-mfg",
                    "sheets": [{
                        "enabled": True,
                        "sheet_id": f"sid-{suffix}",
                        "sheet_name": "SheetMfg",
                        "local_key_names": ["PN", "Maker"],
                        "feishu_key_names": ["PN", "HQ Maker"],
                        "local_key_transforms": ["", "manufacturer_alias"],
                        "fetch_col_map": [{"output": "HQ PN", "alias": "HQ PN"}],
                        "cache_key": key,
                    }],
                }],
            }),
        }
        resp = app.test_client().post(
            "/api/feishu/match",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["matched"], 1)
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb.active
        self.assertEqual([ws["A2"].value, ws["B2"].value, ws["C2"].value], ["A", alias, "HQ-1"])
        wb.close()


    def test_manufacturer_alias_create_lookup_duplicate_and_delete(self):
        suffix = uuid.uuid4().hex[:8]
        canonical = f"HQ Maker {suffix}"
        alias = f"Maker-{suffix}"
        client = app.test_client()

        create_resp = client.post("/api/manufacturer_aliases", data={
            "canonical_name": canonical,
            "alias": alias,
            "source": "unit-test",
            "note": "case and punctuation normalization",
        })
        created_payload = create_resp.get_json()
        self.assertTrue(created_payload["success"])
        created = created_payload["alias"]
        self.assertEqual(created["canonical_name"], canonical)
        self.assertEqual(created["normalized_alias"], normalize_manufacturer_name(alias))

        lookup_resp = client.get(f"/api/manufacturer_aliases/lookup?name=maker_{suffix.upper()}")
        lookup_payload = lookup_resp.get_json()
        self.assertTrue(lookup_payload["success"])
        self.assertEqual(lookup_payload["match"]["canonical_name"], canonical)

        duplicate_resp = client.post("/api/manufacturer_aliases", data={
            "canonical_name": "Other HQ Name",
            "alias": f" maker.{suffix} ",
        })
        duplicate_payload = duplicate_resp.get_json()
        self.assertFalse(duplicate_payload["success"])
        self.assertEqual(duplicate_payload["existing"]["id"], created["id"])

        list_resp = client.get(f"/api/manufacturer_aliases?q={alias}")
        list_payload = list_resp.get_json()
        self.assertTrue(list_payload["success"])
        self.assertEqual(list_payload["match"]["id"], created["id"])
        self.assertTrue(any(item["id"] == created["id"] for item in list_payload["aliases"]))

        delete_resp = client.delete(f"/api/manufacturer_aliases/{created['id']}")
        self.assertTrue(delete_resp.get_json()["success"])
        missing_lookup = client.get(f"/api/manufacturer_aliases/lookup?name={alias}").get_json()
        self.assertIsNone(missing_lookup["match"])

    def test_plm_customer_hq_converts_customer_bom_to_single_board_hq_format(self):
        source = _xlsx_bytes(
            ["Seq", "HQ PN", "Model", "Description", "Qty", "Refdes", "Maker", "\u662f\u5426\u73af\u4fdd", "\u6e7f\u654f\u5c5e\u6027", "\u004d\u0042\u0047\u4f18\u9009\u5c5e\u6027", "\u0043\u0042\u0047\u4f18\u9009\u5c5e\u6027", "\u0044\u0042\u0047\u4f18\u9009\u5c5e\u6027", "\u4e3b\u5236\u63a7", "\u5b50\u5236\u63a7", "\u5b50\u5236\u63a7\u6570\u91cf", "\u0041\u0042\u0047\u4f18\u9009\u5c5e\u6027"],
            [
                ["6", "HQ1", "M-A", "Cap 1uF", 2, "C1,C2", "MakerA", "\u2160\u7ea7(\u65e0\u5364\u6b27\u76df\u73af\u4fdd)", "\u6e7f\u654f\u5668\u4ef6", "\u65e0", "\u65e0", "\u9650\u9009", "SMT", "", "", "\u65e0"],
                ["6", "HQ2", "M-B", "Cap 1uF", 2, "C1,C2", "MakerB", "\u2160\u7ea7(\u65e0\u5364\u6b27\u76df\u73af\u4fdd)", "\u975e\u6e7f\u654f\u5668\u4ef6", "\u53ef\u9009", "\u9650\u9009", "\u4f18\u9009", "", "", "", "\u9650\u9009"],
                ["7", "HQ3", "R-A", "Res 10K", 4, "R1", "MakerC", "\u2160\u7ea7(\u65e0\u5364\u6b27\u76df\u73af\u4fdd)", "\u6e7f\u654f\u5668\u4ef6", "\u65e0", "\u65e0", "\u9650\u9009", "DIP", "", "", "\u65e0"],
            ],
        )
        resp = app.test_client().post(
            "/api/plm/customer_hq_convert",
            data={
                "file": (source, "customer.xlsx"),
                "sheet": "Sheet",
                "header_row": "1",
                "col_seq": "A",
                "col_hqpn": "B",
                "col_model": "C",
                "col_name": "D",
                "col_qty": "E",
                "col_refdes": "F",
                "col_brand": "G",
                "part_no": "HQ31200063SB0",
                "description": "Demo Board PCBA",
                "config_name": "DEMO",
                "engineer": "Tester",
                "version": "I.1",
                "bom_name": "Demo Board PCBA",
                "archive_dept": "HW",
            },
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["total"], 3)
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb.active
        self.assertEqual(ws.title, "BOM")
        self.assertEqual(ws.max_column, 19)
        self.assertEqual(ws["A1"].value, "\u6599\u53f7")
        self.assertEqual(ws["B1"].value, "HQ31200063SB0")
        self.assertEqual(ws["D1"].value, "Demo Board PCBA")
        self.assertEqual(ws["F1"].value, "DEMO")
        self.assertEqual(ws["H1"].value, "Tester")
        self.assertEqual([ws.cell(row=3, column=i).value for i in range(1, 20)], [
            "\u5e8f\u53f7", "\u6599\u53f7", "\u578b\u53f7", "\u7269\u6599\u63cf\u8ff0", "\u5355\u8017", "\u66ff\u4ee3\u5173\u7cfb", "\u4f4d\u53f7", "\u751f\u4ea7\u5382\u5bb6", "\u662f\u5426\u73af\u4fdd", "\u6e7f\u654f\u5c5e\u6027", "\u5907\u6ce8", "\u4e3b\u8f85BOM\u6807\u8bb0", "\u004d\u0042\u0047\u4f18\u9009\u5c5e\u6027", "\u0043\u0042\u0047\u4f18\u9009\u5c5e\u6027", "\u0044\u0042\u0047\u4f18\u9009\u5c5e\u6027", "\u4e3b\u5236\u63a7", "\u5b50\u5236\u63a7", "\u5b50\u5236\u63a7\u6570\u91cf", "\u0041\u0042\u0047\u4f18\u9009\u5c5e\u6027"
        ])
        self.assertEqual([ws["A4"].value, ws["B4"].value, ws["C4"].value, ws["D4"].value, ws["E4"].value, ws["G4"].value, ws["H4"].value], ["6", "HQ1", "M-A", "Cap 1uF", 2, "C1,C2", "MakerA"])
        self.assertEqual([ws["I4"].value, ws["J4"].value, ws["M4"].value, ws["O4"].value, ws["P4"].value, ws["S4"].value], ["\u2160\u7ea7(\u65e0\u5364\u6b27\u76df\u73af\u4fdd)", "\u6e7f\u654f\u5668\u4ef6", "\u65e0", "\u9650\u9009", "SMT", "\u65e0"])
        self.assertEqual([ws["A5"].value, ws["B5"].value, ws["C5"].value, ws["D5"].value, ws["E5"].value, ws["G5"].value, ws["H5"].value], ["6", "HQ2", "M-B", "Cap 1uF", None, None, "MakerB"])
        self.assertEqual([ws["M5"].value, ws["N5"].value, ws["O5"].value, ws["S5"].value], ["\u53ef\u9009", "\u9650\u9009", "\u4f18\u9009", "\u9650\u9009"])
        self.assertEqual([ws["A6"].value, ws["B6"].value, ws["E6"].value, ws["G6"].value, ws["H6"].value], ["7", "HQ3", 4, "R1", "MakerC"])
        wb.close()

    def test_hq_bom_version_compare_reports_added_removed_and_changed(self):
        base_bom = _hq_export_bytes(
            [["1", "A", "R1", "DESC-A", 1, "", "R1", "厂商A"],
             ["2", "B", "C1", "DESC-B", 2, "", "C1", "厂商B"],
             ["3", "C", "L1", "DESC-C", 3, "", "L1", "厂商C"]]
        )
        compare_bom = _hq_export_bytes(
            [["1", "A", "R1", "DESC-A", 1, "", "R1", "厂商A"],
             ["2", "B", "C2", "DESC-B", 2, "", "C1", "厂商B"],
             ["4", "D", "L2", "DESC-D", 4, "", "L2", "厂商D"]]
        )
        data = {
            "old_file": (base_bom, "base.xlsx"),
            "new_file": (compare_bom, "compare.xlsx"),
            "config": json.dumps({
                "header_row": 3,
                "key_col": "料号",
                "compare_cols": ["型号", "单耗"],
            }),
        }
        resp = app.test_client().post(
            "/api/bom_compare/hq_version",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(payload["success"])
        self.assertEqual(payload["added"], 1)
        self.assertEqual(payload["removed"], 1)
        self.assertEqual(payload["changed"], 1)
        self.assertEqual(payload["unchanged"], 1)
        self.assertTrue(payload["download"].startswith("/download/"))
        filename = unquote(payload["download"].split("/download/", 1)[1])
        report_path = WEB_APP / "outputs" / filename
        wb = openpyxl.load_workbook(report_path, data_only=True)
        self.assertEqual(
            wb.sheetnames,
            ["差异总览", "新增物料", "删除物料", "变更物料", "重复料号"],
        )
        self.assertNotIn("差异明细", wb.sheetnames)
        detail = wb["变更物料"]
        self.assertEqual(detail["A1"].value, "差异类型")
        self.assertEqual(detail["B1"].value, "料号")
        self.assertEqual(detail["C1"].value, "基准版本行号")
        self.assertEqual(detail["D1"].value, "对比版本行号")
        self.assertEqual(detail["E1"].value, "变更字段")
        self.assertEqual(detail["F1"].value, "基准值")
        self.assertEqual(detail["G1"].value, "对比值")
        added_sheet = wb["新增物料"]
        added_headers = [cell.value for cell in added_sheet[1]]
        self.assertIn("序号", added_headers)
        self.assertIn("物料描述", added_headers)
        self.assertIn("生产厂家", added_headers)
        self.assertIn("位号", added_headers)
        added_row = {header: added_sheet.cell(row=2, column=idx + 1).value for idx, header in enumerate(added_headers)}
        self.assertEqual(added_row["料号"], "D")
        self.assertEqual(added_row["序号"], "4")
        self.assertEqual(added_row["物料描述"], "DESC-D")
        self.assertEqual(added_row["生产厂家"], "\u5382\u5546D")
        self.assertEqual(added_row["位号"], "L2")

        removed_sheet = wb["删除物料"]
        removed_headers = [cell.value for cell in removed_sheet[1]]
        self.assertIn("序号", removed_headers)
        self.assertIn("物料描述", removed_headers)
        self.assertIn("生产厂家", removed_headers)
        self.assertIn("位号", removed_headers)
        removed_row = {header: removed_sheet.cell(row=2, column=idx + 1).value for idx, header in enumerate(removed_headers)}
        self.assertEqual(removed_row["料号"], "C")
        self.assertEqual(removed_row["序号"], "3")
        self.assertEqual(removed_row["物料描述"], "DESC-C")
        self.assertEqual(removed_row["生产厂家"], "\u5382\u5546C")
        self.assertEqual(removed_row["位号"], "L1")

        wb.close()

    def test_xls_pair_conversion_uses_distinct_paths_for_same_request(self):
        class DummyUpload:
            filename = "demo.xls"

            def save(self, path):
                Path(path).write_bytes(b"xls")

        converted = []

        def fake_convert(src_path, uid, prefix="bomcmp_converted"):
            out_path = WEB_APP / "uploads" / f"{prefix}_converted_{uid}.xlsx"
            converted.append((Path(src_path).name, out_path.name))
            return str(out_path)

        with patch("shared._convert_xls_with_xlrd", side_effect=RuntimeError("not a real xls")):
            with patch("shared._convert_xls_with_excel", side_effect=fake_convert):
                old_path = _save_uploaded_hq_excel(DummyUpload(), "bomcmp_old", "sameuid")
                new_path = _save_uploaded_hq_excel(DummyUpload(), "bomcmp_new", "sameuid")

        self.assertNotEqual(old_path, new_path)
        self.assertEqual(
            [name for _, name in converted],
            ["bomcmp_old_converted_sameuid.xlsx", "bomcmp_new_converted_sameuid.xlsx"],
        )


    def test_plain_excel_upload_converts_xls_instead_of_rejecting(self):
        class DummyUpload:
            filename = "customer.xls"

            def save(self, path):
                Path(path).write_bytes(b"xls")

        converted = []

        def fake_convert(src_path, uid, prefix="converted"):
            out_path = WEB_APP / "uploads" / f"{prefix}_converted_{uid}.xlsx"
            converted.append((Path(src_path).name, out_path.name))
            return str(out_path)

        with patch("shared._convert_xls_with_excel", side_effect=fake_convert):
            left_path = _save_uploaded_excel(DummyUpload(), "bomcmp_customer_left", "sameuid")
            right_path = _save_uploaded_excel(DummyUpload(), "bomcmp_free_right", "sameuid")

        self.assertNotEqual(left_path, right_path)
        self.assertEqual(
            [name for _, name in converted],
            ["bomcmp_customer_left_converted_sameuid.xlsx", "bomcmp_free_right_converted_sameuid.xlsx"],
        )

    def test_plain_excel_upload_uses_headless_xls_converter_first(self):
        class DummyUpload:
            filename = "customer.xls"

            def save(self, path):
                Path(path).write_bytes(b"xls")

        def fake_convert(src_path, uid, prefix="converted"):
            out_path = WEB_APP / "uploads" / f"{prefix}_converted_{uid}.xlsx"
            out_path.write_bytes(b"xlsx")
            return str(out_path)

        with patch("shared._convert_xls_with_xlrd", side_effect=fake_convert) as native_convert:
            with patch("shared._convert_xls_with_excel", side_effect=AssertionError("Excel fallback should not run")):
                path = _save_uploaded_excel(DummyUpload(), "bom_native", "sameuid")

        self.assertEqual(Path(path).name, "bom_native_converted_sameuid.xlsx")
        native_convert.assert_called_once()


    def test_hq_bom_version_exports_single_detail_sheet_with_full_attrs(self):
        headers = ["序号", "料号", "型号", "物料描述", "单耗", "替代关系", "位号", "生产厂家", "制程", "是否量产标识"]

        def build(rows):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM"
            ws.append(["料号", "PROJECT", "描述", "DESC", "项目配置名", "CFG"])
            ws.append(["版本", "I.1", "替代项", "", "BOM名称", "BOM"])
            ws.append(headers)
            for row in rows:
                ws.append(row)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        base_bom = build([
            ["1", "A", "M1", "DESC-A", 1, "", "R1", "厂商A", "SMT", "是"],
            ["2", "B", "M2", "DESC-B", 2, "", "C1", "厂商B", "DIP", "否"],
            ["4", "D", "M4", "DESC-D", 5, "", "U1", "厂商D", "DIP", "否"],
        ])
        compare_bom = build([
            ["1", "A", "M1X", "DESC-A", 3, "", "R1,R2", "厂商A", "SMT2", "否"],
            ["2", "B", "M2", "DESC-B", 2, "", "C1", "厂商B", "DIP", "否"],
            ["3", "C", "M3", "DESC-C", 4, "", "L1", "厂商C", "SMT", "是"],
        ])
        resp = app.test_client().post(
            "/api/bom_compare/hq_version",
            data={
                "old_file": (base_bom, "base.xlsx"),
                "new_file": (compare_bom, "compare.xlsx"),
                "config": json.dumps({
                    "key_col": "料号",
                    "compare_cols": ["型号", "单耗", "位号", "制程", "是否量产标识"],
                }),
            },
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload)
        report_path = WEB_APP / "outputs" / payload["download"].split("/")[-1]
        wb = openpyxl.load_workbook(report_path, data_only=True)
        self.assertEqual(wb.sheetnames, ["\u5dee\u5f02\u603b\u89c8", "\u65b0\u589e\u7269\u6599", "\u5220\u9664\u7269\u6599", "\u53d8\u66f4\u7269\u6599", "\u91cd\u590d\u6599\u53f7"])
        changed = wb["\u53d8\u66f4\u7269\u6599"]
        changed_rows = list(changed.iter_rows(min_row=2, values_only=True))
        by_key = {row[1]: row for row in changed_rows}
        self.assertNotIn("B", by_key)
        self.assertEqual(by_key["A"][0], "\u53d8\u66f4")
        self.assertIn(by_key["A"][4], {"\u578b\u53f7", "\u5355\u8017", "\u4f4d\u53f7", "\u5236\u7a0b", "\u662f\u5426\u91cf\u4ea7\u6807\u8bc6"})

        row_fills = [row[0].fill.fgColor.rgb for row in changed.iter_rows(min_row=2) if row[1].value == "A"]
        self.assertGreater(len(row_fills), 1)
        self.assertEqual(len(set(row_fills)), 1)

        removed = wb["\u5220\u9664\u7269\u6599"]
        removed_headers = [cell.value for cell in removed[1]]
        removed_row = {header: removed.cell(row=2, column=idx + 1).value for idx, header in enumerate(removed_headers)}
        self.assertEqual(removed_row["\u6599\u53f7"], "D")
        self.assertEqual(removed_row["\u5355\u8017"], "5")
        self.assertEqual(removed_row["\u4f4d\u53f7"], "U1")
        self.assertEqual(removed_row["\u5236\u7a0b"], "DIP")
        self.assertEqual(removed_row["\u662f\u5426\u91cf\u4ea7\u6807\u8bc6"], "\u5426")

        added = wb["\u65b0\u589e\u7269\u6599"]
        added_headers = [cell.value for cell in added[1]]
        added_row = {header: added.cell(row=2, column=idx + 1).value for idx, header in enumerate(added_headers)}
        self.assertEqual(added_row["\u6599\u53f7"], "C")
        self.assertEqual(added_row["\u5355\u8017"], "4")
        self.assertEqual(added_row["\u4f4d\u53f7"], "L1")
        self.assertEqual(added_row["\u5236\u7a0b"], "SMT")
        self.assertEqual(added_row["\u662f\u5426\u91cf\u4ea7\u6807\u8bc6"], "\u662f")
        wb.close()


    def test_hq_bom_version_summary_links_to_detail_sheets(self):
        base_bom = _hq_export_bytes([
            ["1", "A", "M1", "DESC-A", 1, "", "R1", "\u5382\u5546A"],
            ["2", "B", "M2", "DESC-B", 2, "", "C1", "\u5382\u5546B"],
        ])
        compare_bom = _hq_export_bytes([
            ["1", "A", "M1X", "DESC-A", 1, "", "R1", "\u5382\u5546A"],
            ["3", "C", "M3", "DESC-C", 3, "", "L1", "\u5382\u5546C"],
        ])
        resp = app.test_client().post(
            "/api/bom_compare/hq_version",
            data={
                "old_file": (base_bom, "base.xlsx"),
                "new_file": (compare_bom, "compare.xlsx"),
                "config": json.dumps({
                    "key_col": "\u6599\u53f7",
                    "compare_cols": ["\u578b\u53f7"],
                }),
            },
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload)
        report_path = WEB_APP / "outputs" / payload["download"].split("/")[-1]
        wb = openpyxl.load_workbook(report_path, data_only=False)
        try:
            summary = wb["\u5dee\u5f02\u603b\u89c8"]
            link_by_name = {
                summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).hyperlink.target
                for row in range(1, summary.max_row + 1)
                if summary.cell(row=row, column=2).hyperlink
            }
            self.assertEqual(link_by_name["\u65b0\u589e"], "#'\u65b0\u589e\u7269\u6599'!A1")
            self.assertEqual(link_by_name["\u5220\u9664"], "#'\u5220\u9664\u7269\u6599'!A1")
            self.assertEqual(link_by_name["\u53d8\u66f4"], "#'\u53d8\u66f4\u7269\u6599'!A1")
            self.assertEqual(link_by_name["\u57fa\u51c6\u7248\u672c\u91cd\u590d\u952e"], "#'\u91cd\u590d\u6599\u53f7'!A1")
            self.assertEqual(link_by_name["\u5bf9\u6bd4\u7248\u672c\u91cd\u590d\u952e"], "#'\u91cd\u590d\u6599\u53f7'!A1")
        finally:
            wb.close()

    def test_hq_bom_version_compare_reports_refdes_delta_as_part_change(self):
        base_bom = _hq_export_bytes(
            [["1", "A", "M1", "DESC-A", 3, "", "R1,R2,R3", "厂商A"]]
        )
        compare_bom = _hq_export_bytes(
            [["1", "A", "M1", "DESC-A", 3, "", "R1,R3,R4", "厂商A"]]
        )
        data = {
            "old_file": (base_bom, "base.xlsx"),
            "new_file": (compare_bom, "compare.xlsx"),
            "config": json.dumps({
                "header_row": 3,
                "key_col": "料号",
                "compare_cols": ["型号", "单耗", "位号"],
            }),
        }
        resp = app.test_client().post(
            "/api/bom_compare/hq_version",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(payload["success"])
        self.assertEqual(payload["added"], 0)
        self.assertEqual(payload["removed"], 0)
        self.assertEqual(payload["changed"], 1)
        self.assertEqual(payload["unchanged"], 0)

        filename = unquote(payload["download"].split("/download/", 1)[1])
        report_path = WEB_APP / "outputs" / filename
        wb = openpyxl.load_workbook(report_path, data_only=True)
        self.assertNotIn("差异明细", wb.sheetnames)
        detail = wb["变更物料"]
        self.assertEqual(detail["C1"].value, "基准版本行号")
        self.assertEqual(detail["D1"].value, "对比版本行号")
        changed_rows = [row for row in detail.iter_rows(min_row=2, values_only=True) if row[0] == "变更"]
        self.assertEqual(len(changed_rows), 1)
        self.assertEqual(changed_rows[0][1], "A")
        self.assertEqual(changed_rows[0][4], "位号")
        self.assertEqual(changed_rows[0][5], "R2")
        self.assertEqual(changed_rows[0][6], "R4")
        wb.close()

    def test_plm_full_bom_version_compare_merges_bom_and_dbg_sheets(self):
        bom = "BOM"
        dbg = "DBG\u4e1a\u52a1BOM"
        ctrl = "DBGBOM\u5236\u63a7\u4fe1\u606f"
        main = "\u4e3b\u6599"
        maker_a = "\u5382\u5546A"
        maker_b = "\u5382\u5546B"
        maker_c = "\u5382\u5546C"
        total_qty = "\u5355\u4f4d\u603b\u7528\u91cf"
        first_process = "\u9996\u5236\u7a0b"
        loss = "\u635f\u8017\u7387"
        fixed_loss = "\u56fa\u5b9a\u5907\u635f\u503c"
        model = "\u578b\u53f7"
        old_rows = {
            bom: [["1", "A", "M1", "DESC-A", 1, main, "R1", maker_a], ["2", "B", "M2", "DESC-B", 2, main, "R2", maker_b]],
            dbg: [["1", "A", "M1-DBG", "DESC-A", 1, main, "R1", maker_a, {total_qty: 1, first_process: "SMT"}], ["2", "B", "M2-DBG", "DESC-B", 2, main, "R2", maker_b, {total_qty: 2, first_process: "SMT"}]],
            ctrl: [["1", "A", "M1-CTRL", "DESC-A", 1, main, "R1", maker_a, {total_qty: 1, loss: "0.005", fixed_loss: "50", first_process: "SMT"}], ["2", "B", "M2-CTRL", "DESC-B", 2, main, "R2", maker_b, {total_qty: 2, loss: "0.005", fixed_loss: "50", first_process: "SMT"}]],
        }
        new_rows = {
            bom: [["1", "A", "M1-NEW", "DESC-A", 1, main, "R1", maker_a], ["3", "C", "M3", "DESC-C", 3, main, "R3", maker_c]],
            dbg: [["1", "A", "M1-NEW-DBG", "DESC-A", 1, main, "R1", maker_a, {total_qty: 1, first_process: "SMT"}], ["3", "C", "M3-DBG", "DESC-C", 3, main, "R3", maker_c, {total_qty: 3, first_process: "SMT"}]],
            ctrl: [["1", "A", "M1-NEW-CTRL", "DESC-A", 1, main, "R1", maker_a, {total_qty: 1, loss: "0.01", fixed_loss: "50", first_process: "SMT"}], ["3", "C", "M3-CTRL", "DESC-C", 3, main, "R3", maker_c, {total_qty: 3, loss: "0.005", fixed_loss: "50", first_process: "SMT"}], ["4", "D", "M4-CTRL", "DESC-D", 4, main, "R4", maker_c, {loss: "0.99"}]],
        }
        data = {
            "old_file": (_plm_full_bom_bytes(old_rows), "old_plm.xlsx"),
            "new_file": (_plm_full_bom_bytes(new_rows), "new_plm.xlsx"),
            "config": json.dumps({"key_col": "\u6599\u53f7", "compare_cols": [model, first_process, loss]}),
        }
        resp = app.test_client().post("/api/bom_compare/hq_version", data=data, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["format"], "plm_full")
        self.assertEqual(payload["sheets"], [bom, dbg])
        self.assertEqual(payload["old_total"], 2)
        self.assertEqual(payload["new_total"], 2)
        self.assertEqual(payload["added"], 1)
        self.assertEqual(payload["removed"], 1)
        self.assertEqual(payload["changed"], 1)
        self.assertEqual(payload["unchanged"], 0)

        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        self.assertIn("\u65b0\u589e\u7269\u6599", wb.sheetnames)
        self.assertIn("\u5220\u9664\u7269\u6599", wb.sheetnames)
        self.assertIn("\u53d8\u66f4\u7269\u6599", wb.sheetnames)
        self.assertNotIn("\u5168\u90e8\u65b0\u589e\u7269\u6599", wb.sheetnames)
        self.assertNotIn("\u5168\u90e8\u5220\u9664\u7269\u6599", wb.sheetnames)
        self.assertNotIn("\u5168\u90e8\u53d8\u66f4\u7269\u6599", wb.sheetnames)

        added_sheet = wb["\u65b0\u589e\u7269\u6599"]
        added_headers = [cell.value for cell in added_sheet[1]]
        self.assertNotIn("Sheet", added_headers)
        self.assertIn("\u5e8f\u53f7", added_headers)
        self.assertIn("\u7269\u6599\u63cf\u8ff0", added_headers)
        self.assertIn(first_process, added_headers)
        added_rows = [dict(zip(added_headers, row)) for row in added_sheet.iter_rows(min_row=2, values_only=True)]
        added = next(row for row in added_rows if row["\u6599\u53f7"] == "C")
        self.assertEqual(added["\u5e8f\u53f7"], "3")
        self.assertEqual(added["\u7269\u6599\u63cf\u8ff0"], "DESC-C")
        self.assertEqual(added[model], "M3")
        self.assertEqual(added[first_process], "SMT")

        removed_sheet = wb["\u5220\u9664\u7269\u6599"]
        removed_headers = [cell.value for cell in removed_sheet[1]]
        self.assertNotIn("Sheet", removed_headers)
        self.assertIn(first_process, removed_headers)
        removed_rows = [dict(zip(removed_headers, row)) for row in removed_sheet.iter_rows(min_row=2, values_only=True)]
        removed = next(row for row in removed_rows if row["\u6599\u53f7"] == "B")
        self.assertEqual(removed[model], "M2")
        self.assertEqual(removed[first_process], "SMT")

        changed_sheet = wb["\u53d8\u66f4\u7269\u6599"]
        changed_rows = list(changed_sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual([row[1] for row in changed_rows], ["A"])
        self.assertEqual(changed_rows[0][4], model)
        self.assertEqual(changed_rows[0][5], "M1")
        self.assertEqual(changed_rows[0][6], "M1-NEW")
        wb.close()


    def test_hq_bom_export_format_defaults_to_header_row_three(self):
        data = {
            "file": (_hq_export_bytes([["1", "HQ1", "M1", "DESC1", 1, "", "R1", "厂商"]]), "hq.xlsx"),
            "header_row": "3",
        }
        resp = app.test_client().post(
            "/api/bom_compare/local_sheets",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["current_sheet"], "BOM")
        self.assertEqual(payload["detected_key"], "料号")
        self.assertIn("单耗", payload["headers"])


    def test_hq_version_sheets_hide_empty_compare_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["料号", "PROJECT", "描述", "DESC", "项目配置名", "CFG"])
        ws.append(["版本", "I.1", "替代项", "", "BOM名称", "BOM"])
        ws.append(["序号", "料号", "型号", "物料描述", "单耗", "替代关系", "位号", "生产厂家", "空白扩展列"])
        ws.append(["1", "A", "M1", "DESC", 1, "主料", "R1", "厂商A", ""])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = app.test_client().post(
            "/api/bom_compare/local_sheets",
            data={"file": (buf, "hq.xlsx")},
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertIn("生产厂家", payload["headers"])
        self.assertNotIn("空白扩展列", payload["headers"])

    def test_hq_bom_compare_rejects_non_standard_format(self):
        data = {
            "file": (_xlsx_bytes(["HQ料号", "规格型号", "单耗"], [["A", "R1", 1]]), "not-standard.xlsx"),
            "header_row": "1",
        }
        resp = app.test_client().post(
            "/api/bom_compare/local_sheets",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertFalse(payload["success"])
        self.assertIn("不支持当前文件格式", payload["error"])



    def test_customer_hq_preview_standardizes_customer_and_hq_rows(self):
        suffix = uuid.uuid4().hex[:8]
        client = app.test_client()
        client.post("/api/manufacturer_aliases", data={
            "canonical_name": f"HQ Maker {suffix}",
            "alias": f"Customer Maker {suffix}",
            "source": "test",
        })
        customer = _xlsx_bytes(
            ["\u4f9b\u5e94\u5546", "\u89c4\u683c\u578b\u53f7", "\u6570\u91cf"],
            [[f"Customer Maker {suffix}", "M1", 2]],
        )
        hq = _hq_export_bytes([
            ["1", "P1", "M1", "", 2, "\u4e3b\u6599", "R1", f"HQ Maker {suffix}"],
        ])
        resp = client.post("/api/bom_compare/customer_hq_preview", data={
            "left_file": (customer, "customer.xlsx"),
            "right_file": (hq, "hq.xlsx"),
            "config": json.dumps({
                "left_header_row": 1,
                "match_mode": "identity",
                "mapping": {
                    "manufacturer": "\u4f9b\u5e94\u5546",
                    "model": "\u89c4\u683c\u578b\u53f7",
                    "quantity": "\u6570\u91cf",
                },
            }),
        }, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["match_mode"], "identity")
        self.assertEqual(payload["customer_total"], 1)
        self.assertEqual(payload["hq_total"], 1)
        self.assertEqual(payload["customer_invalid"], 0)
        row = payload["customer_preview"][0]
        self.assertIn("\u578b\u53f7:M1", row["match_key"])
        self.assertEqual(row["manufacturer_mapped"], f"HQ Maker {suffix}")
        self.assertEqual(row["quantity"], "2")
        self.assertEqual(payload["hq_preview"][0]["part_no"], "P1")

    def test_machine_hq_version_treats_equivalent_numbers_as_same(self):
        old_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", "8.0", "主料", "R1", "厂商A"],
        ]).getvalue()
        new_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", "8", "主料", "R1", "厂商A"],
        ]).getvalue()

        resp = app.test_client().post("/api/bom_compare/machine_hq_version", data={
            "old_file": (io.BytesIO(old_bytes), "old.xlsx"),
            "new_file": (io.BytesIO(new_bytes), "new.xlsx"),
            "config": json.dumps({"key_col": "料号", "compare_cols": ["单耗"]}),
        }, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["changed"], 0)
        self.assertEqual(payload["unchanged"], 1)

    def test_generic_cadence_standard_export_auto_header_and_refdes_expand(self):
        refdes = "位号"
        part_no = "料号"
        model = "型号"
        main = "主料"
        maker = "厂商A"
        left_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 2, main, "R1,R2", maker],
            ["2", "B", "M2", "", 1, main, "R3", maker],
        ]).getvalue()
        right_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 1, main, "R1", maker],
            ["2", "A", "M1", "", 1, main, "R2", maker],
            ["3", "C", "M3", "", 1, main, "R3", maker],
        ]).getvalue()

        sheets_resp = app.test_client().post("/api/bom_compare/generic_sheets", data={
            "left_file": (io.BytesIO(left_bytes), "cadence.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "hq.xlsx"),
            "compare_type": "cadence_hq",
            "left_header_row": "1",
            "right_header_row": "3",
        }, content_type="multipart/form-data")
        sheets_payload = sheets_resp.get_json()
        self.assertTrue(sheets_payload["success"])
        self.assertEqual(sheets_payload["left_format"], "cadence_standard")
        self.assertEqual(sheets_payload["left_header_row"], 3)
        self.assertEqual(sheets_payload["detected_left_key"], part_no)
        self.assertEqual(sheets_payload["detected_right_key"], part_no)

        compare_resp = app.test_client().post("/api/bom_compare/generic", data={
            "left_file": (io.BytesIO(left_bytes), "cadence.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "hq.xlsx"),
            "config": json.dumps({
                "compare_type": "cadence_hq",
                "left_header_row": 1,
                "right_header_row": 3,
                "left_key_col": refdes,
                "right_key_col": refdes,
                "field_pairs": [{"left": part_no, "right": part_no}, {"left": model, "right": model}],
            }),
        }, content_type="multipart/form-data")
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"])
        self.assertTrue(payload["expanded_refdes"])
        self.assertEqual(payload["left_header_row"], 3)
        self.assertEqual(payload["left_total"], 3)
        self.assertEqual(payload["right_total"], 3)
        self.assertEqual(payload["left_only"], 0)
        self.assertEqual(payload["right_only"], 0)
        self.assertEqual(payload["changed"], 1)
        self.assertEqual(payload["same"], 2)

    def test_generic_cadence_preview_returns_rows_before_header_selection(self):
        cadence = _xlsx_bytes(["REFDES", "PART_NUMBER", "QTY"], [["R1", "A", 1], ["R2", "A", 1]])
        hq = _hq_export_bytes([
            ["1", "A", "M1", "", 1, "\u4e3b\u6599", "R1", "\u5382\u5546A"],
        ])

        resp = app.test_client().post("/api/bom_compare/generic_preview", data={
            "left_file": (cadence, "cadence.xlsx"),
            "right_file": (hq, "hq.xlsx"),
            "compare_type": "cadence_hq",
        }, content_type="multipart/form-data")
        payload = resp.get_json()

        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["left"]["current_sheet"], "Sheet")
        self.assertEqual(payload["left"]["rows"][1]["values"][:3], ["R1", "A", "1"])
        self.assertIn("right", payload)
        self.assertGreaterEqual(len(payload["right"]["rows"]), 3)


    def test_generic_cadence_summary_links_to_detail_sheets(self):
        part_no = "\u6599\u53f7"
        model = "\u578b\u53f7"
        refdes = "\u4f4d\u53f7"
        main = "\u4e3b\u6599"
        maker = "\u5382\u5546A"
        left_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 1, main, "R1", maker],
            ["2", "B", "M2", "", 1, main, "R2", maker],
        ]).getvalue()
        right_bytes = _hq_export_bytes([
            ["1", "A", "M1X", "", 1, main, "R1", maker],
            ["3", "C", "M3", "", 1, main, "R3", maker],
        ]).getvalue()

        compare_resp = app.test_client().post("/api/bom_compare/generic", data={
            "left_file": (io.BytesIO(left_bytes), "cadence.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "hq.xlsx"),
            "config": json.dumps({
                "compare_type": "cadence_hq",
                "left_header_row": 3,
                "right_header_row": 3,
                "left_key_col": part_no,
                "right_key_col": part_no,
                "field_pairs": [{"left": model, "right": model}, {"left": refdes, "right": refdes}],
            }),
        }, content_type="multipart/form-data")
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"], payload)
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=False)
        try:
            summary = wb["\u5dee\u5f02\u603b\u89c8"]
            link_by_name = {
                summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).hyperlink.target
                for row in range(1, summary.max_row + 1)
                if summary.cell(row=row, column=2).hyperlink
            }
            self.assertEqual(link_by_name["\u4ec5 Cadence BOM \u5b58\u5728"], "#'\u4ec5Cadence BOM\u5b58\u5728'!A1")
            self.assertEqual(link_by_name["\u4ec5 HQ BOM \u5b58\u5728"], "#'\u4ec5HQ BOM\u5b58\u5728'!A1")
            self.assertEqual(link_by_name["\u5b57\u6bb5\u53d8\u66f4"], "#'\u5b57\u6bb5\u53d8\u66f4'!A1")
            self.assertEqual(link_by_name["Cadence BOM \u91cd\u590d\u952e"], "#'\u91cd\u590d\u952e'!A1")
            self.assertEqual(link_by_name["HQ BOM \u91cd\u590d\u952e"], "#'\u91cd\u590d\u952e'!A1")
        finally:
            wb.close()


    def test_generic_cadence_treats_refdes_order_as_same(self):
        part_no = "料号"
        refdes = "位号"
        main = "主料"
        maker = "厂商A"
        left_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 3, main, "R1,R2,R3", maker],
        ]).getvalue()
        right_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 3, main, "R3,R1,R2", maker],
        ]).getvalue()
        compare_resp = app.test_client().post("/api/bom_compare/generic", data={
            "left_file": (io.BytesIO(left_bytes), "cadence.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "hq.xlsx"),
            "config": json.dumps({
                "compare_type": "cadence_hq",
                "left_header_row": 1,
                "right_header_row": 3,
                "left_key_col": part_no,
                "right_key_col": part_no,
                "field_pairs": [{"left": refdes, "right": refdes}],
            }),
        }, content_type="multipart/form-data")
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["changed"], 0)
        self.assertEqual(payload["same"], 1)
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        self.assertNotIn("差异明细", wb.sheetnames)
        changed_sheet = wb["字段变更"]
        self.assertEqual(changed_sheet.max_row, 1)
        wb.close()


    def test_generic_cadence_refdes_field_diff_reports_only_delta_refs(self):
        part_no = "料号"
        refdes = "位号"
        main = "主料"
        maker = "厂商A"
        left_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 3, main, "R1,R2,R3", maker],
            ["2", "B", "M2", "", 3, main, "C1,C2,C3", maker],
        ]).getvalue()
        right_bytes = _hq_export_bytes([
            ["1", "A", "M1X", "", 3, main, "R1,R3,R4", maker],
            ["2", "B", "M2X", "", 3, main, "C1,C3,C4", maker],
        ]).getvalue()
        compare_resp = app.test_client().post("/api/bom_compare/generic", data={
            "left_file": (io.BytesIO(left_bytes), "cadence.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "hq.xlsx"),
            "config": json.dumps({
                "compare_type": "cadence_hq",
                "left_header_row": 3,
                "right_header_row": 3,
                "left_key_col": part_no,
                "right_key_col": part_no,
                "field_pairs": [{"left": refdes, "right": refdes}, {"left": "型号", "right": "型号"}],
            }),
        }, content_type="multipart/form-data")
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["changed"], 2)
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename)
        rows = list(wb["字段变更"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 4)
        row_a_refdes = next(row for row in rows if row[1] == "A" and row[4] == refdes)
        self.assertEqual(row_a_refdes[5], "R2")
        self.assertEqual(row_a_refdes[6], "R4")
        fills_by_key = {}
        ws = wb["字段变更"]
        for row_idx in range(2, ws.max_row + 1):
            key = ws.cell(row_idx, 2).value
            fills_by_key.setdefault(key, set()).add(ws.cell(row_idx, 1).fill.fgColor.rgb)
        self.assertEqual(len(fills_by_key), 2)
        self.assertTrue(all(len(fills) == 1 for fills in fills_by_key.values()))
        self.assertEqual(len({next(iter(fills)) for fills in fills_by_key.values()}), 2)
        wb.close()
    def test_generic_cadence_ignores_empty_columns_by_default(self):
        part_no = "料号"
        model = "型号"
        empty_col = "物料描述"
        main = "主料"
        maker = "厂商A"
        left_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 1, main, "R1", maker],
        ]).getvalue()
        right_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 1, main, "R1", maker],
        ]).getvalue()

        sheets_resp = app.test_client().post("/api/bom_compare/generic_sheets", data={
            "left_file": (io.BytesIO(left_bytes), "cadence.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "hq.xlsx"),
            "compare_type": "cadence_hq",
            "left_header_row": "1",
            "right_header_row": "3",
        }, content_type="multipart/form-data")
        sheets_payload = sheets_resp.get_json()
        self.assertTrue(sheets_payload["success"])
        self.assertNotIn(empty_col, sheets_payload["left_headers"])
        self.assertIn(empty_col, sheets_payload["left_ignored_headers"])

        compare_resp = app.test_client().post("/api/bom_compare/generic", data={
            "left_file": (io.BytesIO(left_bytes), "cadence.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "hq.xlsx"),
            "config": json.dumps({
                "compare_type": "cadence_hq",
                "left_header_row": 1,
                "right_header_row": 3,
                "left_key_col": part_no,
                "right_key_col": part_no,
                "field_pairs": [{"left": empty_col, "right": model}, {"left": model, "right": model}],
            }),
        }, content_type="multipart/form-data")
        payload = compare_resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["changed"], 0)
        self.assertEqual(payload["same"], 1)
        self.assertIn(empty_col, payload["left_ignored_headers"])



    def test_generic_compare_allows_later_valid_field_pair_when_first_is_empty(self):
        part_no = "\u6599\u53f7"
        model = "\u578b\u53f7"
        left_bytes = _hq_export_bytes([
            ["1", "A", "M1", "", 1, "\u4e3b\u6599", "R1", "\u5382\u5546A"],
        ]).getvalue()
        right_bytes = _hq_export_bytes([
            ["1", "A", "M2", "", 1, "\u4e3b\u6599", "R1", "\u5382\u5546A"],
        ]).getvalue()
        resp = app.test_client().post("/api/bom_compare/generic", data={
            "left_file": (io.BytesIO(left_bytes), "left.xlsx"),
            "right_file": (io.BytesIO(right_bytes), "right.xlsx"),
            "config": json.dumps({
                "compare_type": "cadence_hq",
                "left_header_row": 3,
                "right_header_row": 3,
                "left_key_col": part_no,
                "right_key_col": part_no,
                "field_pairs": [
                    {"left": "", "right": ""},
                    {"left": model, "right": model},
                ],
            }),
        }, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["changed"], 1)

    def test_generic_cadence_compare_accepts_plm_full_hq_bom_on_right(self):
        main = "\u4e3b\u6599"
        maker_a = "\u5382\u5546A"
        maker_b = "\u5382\u5546B"
        refdes = "\u4f4d\u53f7"
        cadence = _xlsx_bytes(["REFDES", "PART_NUMBER"], [["R1", "A"], ["R9", "Z"]])
        hq = _plm_full_bom_bytes({"BOM": [["1", "A", "M1", "DESC-A", 1, main, "R1", maker_a], ["2", "B", "M2", "DESC-B", 2, main, "R2", maker_b]]})
        data = {
            "left_file": (cadence, "cadence.xlsx"),
            "right_file": (hq, "hq_plm.xlsx"),
            "compare_type": "cadence_hq",
            "left_header_row": "1",
            "right_header_row": "3",
        }
        resp = app.test_client().post("/api/bom_compare/generic_sheets", data=data, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["right_format"], "plm_full")
        self.assertEqual(payload["right_header_row"], 8)
        self.assertEqual(payload["detected_left_key"], "PART_NUMBER")
        self.assertEqual(payload["detected_right_key"], "料号")


    def test_generic_sheets_detects_left_and_right_keys(self):
        cadence = _xlsx_bytes(["REFDES", "PART_NUMBER", "QTY"], [["R1", "A", 1]])
        hq = _hq_export_bytes([["1", "A", "M1", "", 1, "", "R1", ""]])
        data = {
            "left_file": (cadence, "cadence.xlsx"),
            "right_file": (hq, "hq.xlsx"),
            "left_header_row": "1",
            "right_header_row": "3",
        }
        resp = app.test_client().post(
            "/api/bom_compare/generic_sheets",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["detected_left_key"], "PART_NUMBER")
        self.assertEqual(payload["detected_right_key"], "料号")

    def test_bug_report_submit_and_list(self):
        data = {
            "reporter": "张三",
            "employee_id": "100001",
            "module": "BOM比对工具合集",
            "severity": "一般",
            "title": "测试问题",
            "description": "这里是问题描述",
            "steps": "1. 打开页面",
            "expected": "可以正常使用",
        }
        resp = app.test_client().post("/api/bug_reports", data=data)
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["report"]["reporter"], "张三")

        list_resp = app.test_client().get("/api/bug_reports")
        list_payload = list_resp.get_json()
        self.assertTrue(list_payload["success"])
        self.assertTrue(any(item["id"] == payload["report"]["id"] for item in list_payload["reports"]))
        self.assertTrue(any(item["id"] == "seed-bug-bom-header-detect" for item in list_payload["reports"]))

    def test_bug_report_defaults_and_validation_messages_are_readable(self):
        client = app.test_client()
        missing_resp = client.post("/api/bug_reports", data={"reporter": "张三"})
        missing_payload = missing_resp.get_json()
        self.assertFalse(missing_payload["success"])
        self.assertIn("问题标题", missing_payload["error"])

        create_resp = client.post("/api/bug_reports", data={
            "reporter": "默认值测试",
            "employee_id": "100013",
            "title": "默认字段检查",
            "description": "未填写模块和严重程度时应使用可读默认值",
        })
        payload = create_resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["report"]["module"], "未指定")
        self.assertEqual(payload["report"]["severity"], "一般")
        self.assertEqual(payload["report"]["status"], "待处理")

    def test_bug_report_accepts_excel_attachment(self):
        data = {
            "reporter": "李四",
            "employee_id": "100002",
            "title": "附件测试",
            "description": "上传 Excel 附件",
            "images": (_xlsx_bytes(["A"], [["B"]]), "case.xlsx"),
        }
        resp = app.test_client().post(
            "/api/bug_reports",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["report"]["attachments"][0]["name"], "case.xlsx")

    def test_bug_report_rejects_unsupported_attachment(self):
        data = {
            "reporter": "王五",
            "employee_id": "100003",
            "title": "附件测试",
            "description": "上传不支持附件",
            "images": (io.BytesIO(b"binary"), "tool.exe"),
        }
        resp = app.test_client().post(
            "/api/bug_reports",
            data=data,
            content_type="multipart/form-data",
        )
        payload = resp.get_json()
        self.assertFalse(payload["success"])
        self.assertIn("仅支持上传", payload["error"])


    def test_bug_report_status_can_be_updated(self):
        data = {
            "reporter": "孙八",
            "employee_id": "100006",
            "title": "状态测试",
            "description": "用于验证状态修改",
        }
        create_resp = app.test_client().post("/api/bug_reports", data=data)
        created = create_resp.get_json()["report"]

        update_resp = app.test_client().post(
            f"/api/bug_reports/{created['id']}/status",
            json={"status": "处理中"},
        )
        payload = update_resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["report"]["status"], "处理中")

        invalid_resp = app.test_client().post(
            f"/api/bug_reports/{created['id']}/status",
            json={"status": "随意状态"},
        )
        self.assertFalse(invalid_resp.get_json()["success"])


    def test_feature_request_submit_and_list(self):
        data = {
            "requester": "赵六",
            "employee_id": "100004",
            "module": "BOM比对工具合集",
            "priority": "较高",
            "request_type": "新功能",
            "title": "增加导出汇总",
            "background": "需要快速查看差异总数",
            "requirement": "生成差异汇总页",
            "value": "减少手工统计",
            "acceptance": "导出文件包含汇总页",
        }
        resp = app.test_client().post("/api/feature_requests", data=data)
        payload = resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["request"]["requester"], "赵六")
        self.assertEqual(payload["request"]["status"], "待评估")
        self.assertEqual(payload["request"]["likes"], 0)

        list_resp = app.test_client().get("/api/feature_requests")
        list_payload = list_resp.get_json()
        self.assertTrue(list_payload["success"])
        self.assertTrue(any(item["id"] == payload["request"]["id"] for item in list_payload["requests"]))
        self.assertTrue(any(item["id"] == "seed-bom-compare-summary" for item in list_payload["requests"]))

    def test_feature_request_rejects_missing_required_fields(self):
        resp = app.test_client().post("/api/feature_requests", data={"requester": "赵六"})
        payload = resp.get_json()
        self.assertFalse(payload["success"])
        self.assertIn("需求标题", payload["error"])


    def test_feature_request_like_increments_count(self):
        data = {
            "requester": "钱七",
            "employee_id": "100005",
            "title": "点赞测试需求",
            "requirement": "需要可以点赞",
        }
        create_resp = app.test_client().post("/api/feature_requests", data=data)
        created = create_resp.get_json()["request"]

        like_resp = app.test_client().post(f"/api/feature_requests/{created['id']}/like")
        payload = like_resp.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["request"]["likes"], 1)

        second_resp = app.test_client().post(f"/api/feature_requests/{created['id']}/like")
        self.assertEqual(second_resp.get_json()["request"]["likes"], 2)


if __name__ == "__main__":
    unittest.main()
