import io
import json
import sys
import unittest
import uuid
from pathlib import Path
from urllib.parse import unquote

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WEB_APP = ROOT / "web_app2"
if str(WEB_APP) not in sys.path:
    sys.path.insert(0, str(WEB_APP))

from app import app  # noqa: E402


def _xlsx_bytes(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _hq_export_bytes(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(["\u6599\u53f7", "PROJECT", "\u63cf\u8ff0", "DESC", "\u9879\u76ee\u914d\u7f6e\u540d", "CFG"])
    ws.append(["\u7248\u672c", "I.1", "\u66ff\u4ee3\u9879", "", "BOM\u540d\u79f0", "BOM"])
    ws.append(["\u5e8f\u53f7", "\u6599\u53f7", "\u578b\u53f7", "\u7269\u6599\u63cf\u8ff0", "\u5355\u8017", "\u66ff\u4ee3\u5173\u7cfb", "\u4f4d\u53f7", "\u751f\u4ea7\u5382\u5bb6"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class CustomerHqExportReportTests(unittest.TestCase):
    def test_summary_links_to_detail_sheets(self):
        suffix = uuid.uuid4().hex[:8]
        client = app.test_client()
        client.post("/api/manufacturer_aliases", data={
            "canonical_name": f"HQ Maker {suffix}",
            "alias": f"Customer Maker {suffix}",
            "source": "test",
        })
        customer = _xlsx_bytes(
            ["\u4f9b\u5e94\u5546", "\u89c4\u683c\u578b\u53f7", "\u6570\u91cf"],
            [[f"Customer Maker {suffix}", "M1", 3], ["Only Customer", "M2", 1]],
        )
        hq = _hq_export_bytes([
            ["1", "P1", "M1", "", 2, "\u4e3b\u6599", "R1", f"HQ Maker {suffix}"],
            ["2", "P2", "M3", "", 1, "\u4e3b\u6599", "R2", "Only HQ"],
        ])
        resp = client.post("/api/bom_compare/customer_hq_export", data={
            "left_file": (customer, "customer.xlsx"),
            "right_file": (hq, "hq.xlsx"),
            "config": json.dumps({
                "left_header_row": 1,
                "match_mode": "identity",
                "mapping": {
                    "manufacturer": "\u4f9b\u5e94\u5546",
                    "model": "\u89c4\u683c\u578b\u53f7",
                    "quantity": "\u6570\u91cf",
                },
            }),
        }, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["changed"], 1)
        self.assertEqual(payload["customer_only"], 1)
        self.assertEqual(payload["hq_only"], 1)

        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        ws = wb["\u5dee\u5f02\u603b\u89c8"]
        summary_values = {
            ws.cell(row, 1).value: ws.cell(row, 2).value
            for row in range(1, ws.max_row + 1)
        }
        self.assertEqual(ws["A1"].value, "BOM Tools 导出报告")
        self.assertEqual(summary_values["导出来源"], "BOM Tools 平台 v2.2.0")
        self.assertEqual(summary_values["平台版本"], "v2.2.0")
        self.assertEqual(summary_values["工具名称"], "客户 BOM 对比 HQ BOM")
        self.assertEqual(summary_values["工具版本"], "v1.1.0")
        self.assertEqual(summary_values["客户 BOM 文件"], "customer.xlsx")
        self.assertEqual(summary_values["HQ BOM 文件"], "hq.xlsx")
        self.assertEqual(summary_values["客户 Sheet / 表头行"], "自动选择 / 1")
        self.assertEqual(summary_values["HQ Sheet"], "自动识别")
        self.assertEqual(summary_values["匹配模式"], "按型号+制造商匹配")
        self.assertIn("规格型号: 规格型号", summary_values["字段映射"])
        self.assertIn("制造商: 供应商", summary_values["字段映射"])
        links = {
            ws.cell(row, 1).value: ws.cell(row, 1).hyperlink.target
            for row in range(3, ws.max_row + 1)
            if ws.cell(row, 1).hyperlink
        }
        self.assertEqual(links["\u5b57\u6bb5\u5dee\u5f02"], "#'\u5b57\u6bb5\u5dee\u5f02'!A1")
        self.assertEqual(links["\u4ec5\u5ba2\u6237\u5b58\u5728"], "#'\u4ec5\u5ba2\u6237\u5b58\u5728'!A1")
        self.assertEqual(links["\u4ec5 HQ \u5b58\u5728"], "#'\u4ec5HQ\u5b58\u5728'!A1")
        wb.close()



    def test_export_auto_fits_column_widths_and_row_heights(self):
        long_model = "MODEL-WITH-LONG-COMMERCIAL-NAME-AND-PACKAGE-DETAILS-1234567890"
        long_maker = "Very Long Customer Manufacturer Name With Multiple Business Units And Suffixes"
        long_refdes = ",".join(f"R{i:03d}" for i in range(1, 41))
        customer = _xlsx_bytes(
            ["供应商", "规格型号", "数量", "位号"],
            [[long_maker, long_model, 123, long_refdes]],
        )
        hq = _hq_export_bytes([
            ["1", "P1", long_model, "", 456, "主料", "R001,R002,R003", long_maker],
        ])
        resp = app.test_client().post("/api/bom_compare/customer_hq_export", data={
            "left_file": (customer, "customer-long.xlsx"),
            "right_file": (hq, "hq-long.xlsx"),
            "config": json.dumps({
                "left_header_row": 1,
                "match_mode": "identity",
                "mapping": {
                    "manufacturer": "供应商",
                    "model": "规格型号",
                    "quantity": "数量",
                    "refdes": "位号",
                },
            }),
        }, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename)
        try:
            summary = wb["差异总览"]
            detail = wb["匹配明细"]
            field_diff = wb["字段差异"]
            self.assertGreaterEqual(summary.column_dimensions["B"].width, 60)
            self.assertGreater(summary.row_dimensions[14].height, 18)
            self.assertGreaterEqual(detail.column_dimensions["H"].width, 60)
            self.assertGreaterEqual(detail.column_dimensions["J"].width, 60)
            self.assertTrue(any(
                (field_diff.row_dimensions[row].height or 0) > 18
                for row in range(2, field_diff.max_row + 1)
            ))
        finally:
            wb.close()

    def test_refdes_field_diff_reports_only_delta_refs(self):
        customer = _xlsx_bytes(
            ["供应商", "规格型号", "数量", "位号"],
            [["MakerA", "M1", 3, "R1,R2,R3"]],
        )
        hq = _hq_export_bytes([
            ["1", "P1", "M1", "", 3, "主料", "R1,R3,R4", "MakerA"],
        ])
        resp = app.test_client().post("/api/bom_compare/customer_hq_export", data={
            "left_file": (customer, "customer.xlsx"),
            "right_file": (hq, "hq.xlsx"),
            "config": json.dumps({
                "left_header_row": 1,
                "match_mode": "identity",
                "mapping": {
                    "manufacturer": "供应商",
                    "model": "规格型号",
                    "quantity": "数量",
                    "refdes": "位号",
                },
            }),
        }, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        self.assertEqual(payload["changed"], 1)

        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        rows = list(wb["字段差异"].iter_rows(min_row=2, values_only=True))
        refdes_rows = [row for row in rows if row[1] == "位号差异"]
        self.assertEqual(len(refdes_rows), 1)
        self.assertEqual(refdes_rows[0][3], "R2")
        self.assertEqual(refdes_rows[0][5], "R4")
        wb.close()

    def test_field_diff_groups_same_match_key_with_same_fill(self):
        customer = _xlsx_bytes(
            ["供应商", "规格型号", "数量", "位号"],
            [["MakerA", "M1", 3, "R1,R2"], ["MakerB", "M2", 5, "C1,C2"]],
        )
        hq = _hq_export_bytes([
            ["1", "P1", "M1", "", 4, "主料", "R1,R3", "MakerA"],
            ["2", "P2", "M2", "", 6, "主料", "C1,C3", "MakerB"],
        ])
        resp = app.test_client().post("/api/bom_compare/customer_hq_export", data={
            "left_file": (customer, "customer.xlsx"),
            "right_file": (hq, "hq.xlsx"),
            "config": json.dumps({
                "left_header_row": 1,
                "match_mode": "identity",
                "mapping": {"manufacturer": "供应商", "model": "规格型号", "quantity": "数量", "refdes": "位号"},
            }),
        }, content_type="multipart/form-data")
        payload = resp.get_json()
        self.assertTrue(payload["success"], payload.get("error"))
        filename = unquote(payload["download"].split("/download/", 1)[1])
        wb = openpyxl.load_workbook(WEB_APP / "outputs" / filename)
        ws = wb["字段差异"]
        fills_by_key = {}
        for row_idx in range(2, ws.max_row + 1):
            key = ws.cell(row_idx, 1).value
            fills_by_key.setdefault(key, set()).add(ws.cell(row_idx, 1).fill.fgColor.rgb)
        self.assertEqual(len(fills_by_key), 2)
        self.assertTrue(all(len(fills) == 1 for fills in fills_by_key.values()))
        self.assertEqual(len({next(iter(fills)) for fills in fills_by_key.values()}), 2)
        wb.close()
if __name__ == "__main__":
    unittest.main()
