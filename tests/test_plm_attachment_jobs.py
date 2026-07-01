# -*- coding: utf-8 -*-
import sys
import time
import unittest
from pathlib import Path
from unittest.mock import patch
from io import BytesIO
from zipfile import ZipFile


ROOT = Path(__file__).resolve().parents[1]
WEB_APP = ROOT / "web_app2"
TESTS = ROOT / "tests"
if str(TESTS) not in sys.path:
    sys.path.insert(0, str(TESTS))
from test_env import configure_test_environment  # noqa: E402

configure_test_environment()
if str(WEB_APP) not in sys.path:
    sys.path.insert(0, str(WEB_APP))

from app import app  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402


class PlmAttachmentJobTests(unittest.TestCase):
    def test_plm_spec_reverse_single_value_creates_background_job(self):
        from plm import _PLM_SPEC_REVERSE_JOBS

        run_calls = []

        class FakePlaywrightContext:
            def __enter__(self):
                return object()

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_run_plm_feature(playwright, *, username, password, feature, upload_file, output_dir, headless=False, log=None):
            wb_in = load_workbook(upload_file, data_only=True)
            ws_in = wb_in.active
            single_value = ws_in.cell(row=2, column=1).value
            all_values = [ws_in.cell(row=row, column=1).value for row in range(2, ws_in.max_row + 1)]
            wb_in.close()
            run_calls.append((username, upload_file.name, single_value, all_values))
            if log:
                log("启动浏览器")
                log("上传文件：" + upload_file.name)
                log("导出完成：mock.xlsx")
            out_path = Path(output_dir) / "mock_spec_reverse.xlsx"
            wb = Workbook()
            wb.active.append(["HQ料号"])
            wb.active.append(["HQTEST001"])
            wb.save(out_path)
            return out_path

        def wait_done(client, job_id):
            for _ in range(40):
                status = client.get(f"/api/plm/auto_spec_reverse/status/{job_id}").get_json()
                if status.get("status") == "done":
                    return status
                if status.get("status") == "error":
                    self.fail(status.get("error"))
                time.sleep(0.05)
            self.fail("job did not finish")

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            with patch("playwright.sync_api.sync_playwright", return_value=FakePlaywrightContext()), patch(
                "plm.automation.run_plm_feature", fake_run_plm_feature
            ):
                resp = client.post(
                    "/api/plm/auto_spec_reverse",
                    data={"username": "100448405", "password": "pw", "single_value": " HQ TEST 001 "},
                )
                payload = resp.get_json()
                self.assertTrue(payload["success"], payload)
                self.assertIn("/api/plm/auto_spec_reverse/status/", payload["status_url"])
                status = wait_done(client, payload["job_id"])
                self.assertEqual(status["download"], "/download/mock_spec_reverse.xlsx")
                self.assertEqual(status["progress"], 100)
                self.assertIn("上传文件", status["log"])
                self.assertEqual(run_calls[0][0], "100448405")
                self.assertEqual(run_calls[0][2], "HQ TEST 001")
                self.assertEqual(run_calls[0][3], ["HQ TEST 001"])

            _PLM_SPEC_REVERSE_JOBS.pop(payload["job_id"], None)

    def test_plm_spec_reverse_single_value_accepts_comma_separated_values(self):
        from plm import _PLM_SPEC_REVERSE_JOBS

        run_calls = []

        class FakePlaywrightContext:
            def __enter__(self):
                return object()

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_run_plm_feature(playwright, *, username, password, feature, upload_file, output_dir, headless=False, log=None):
            wb_in = load_workbook(upload_file, data_only=True)
            values = [wb_in.active.cell(row=row, column=1).value for row in range(2, wb_in.active.max_row + 1)]
            wb_in.close()
            run_calls.append(values)
            out_path = Path(output_dir) / "mock_spec_reverse_multi.xlsx"
            wb = Workbook()
            wb.active.append(["HQ料号"])
            wb.active.append(["HQTEST001"])
            wb.save(out_path)
            return out_path

        def wait_done(client, job_id):
            for _ in range(40):
                status = client.get(f"/api/plm/auto_spec_reverse/status/{job_id}").get_json()
                if status.get("status") == "done":
                    return status
                if status.get("status") == "error":
                    self.fail(status.get("error"))
                time.sleep(0.05)
            self.fail("job did not finish")

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            with patch("playwright.sync_api.sync_playwright", return_value=FakePlaywrightContext()), patch(
                "plm.automation.run_plm_feature", fake_run_plm_feature
            ):
                resp = client.post(
                    "/api/plm/auto_spec_reverse",
                    data={"username": "100448405", "password": "pw", "single_value": " HQ TEST 001, A B 002 ，HQ003 "},
                )
                payload = resp.get_json()
                self.assertTrue(payload["success"], payload)
                wait_done(client, payload["job_id"])
                self.assertEqual(run_calls[0], ["HQ TEST 001", "A B 002", "HQ003"])

            _PLM_SPEC_REVERSE_JOBS.pop(payload["job_id"], None)

    def test_plm_attachment_batch_excel_enqueues_unique_hqpns(self):
        from plm import _PLM_ATTACHMENT_JOBS

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["Index", "HQ料号", "Remark"])
        ws.append([1, "HQ17001007NC0", "single"])
        ws.append([2, "HQ162020006H0", "multi"])
        ws.append([3, "HQ17001007NC0", "duplicate"])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        queued = []

        def fake_enqueue(job_id, username, password, hqpn, batch_id=""):
            queued.append((job_id, username, password, hqpn, batch_id))

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            with patch("plm._enqueue_attachment_job", fake_enqueue):
                resp = client.post(
                    "/api/plm/auto_hq_attachments/batch",
                    data={
                        "username": "100448405",
                        "password": "pw",
                        "header_row": "1",
                        "col_hqpn": "B",
                        "file": (bio, "bom.xlsx"),
                    },
                    content_type="multipart/form-data",
                )
                payload = resp.get_json()
                self.assertTrue(payload["success"], payload)
                self.assertEqual(payload["count"], 2)
                self.assertIn("batch_id", payload)
                self.assertIn("/api/plm/auto_hq_attachments/batch/status/", payload["status_url"])
                self.assertEqual([j["hqpn"] for j in payload["jobs"]], ["HQ17001007NC0", "HQ162020006H0"])
                self.assertEqual([item[3] for item in queued], ["HQ17001007NC0", "HQ162020006H0"])
                self.assertTrue(all(item[4] == payload["batch_id"] for item in queued))

            for job in payload["jobs"]:
                _PLM_ATTACHMENT_JOBS.pop(job["job_id"], None)

    def test_plm_attachment_batch_status_builds_total_zip(self):
        from plm import _PLM_ATTACHMENT_BATCHES, _PLM_ATTACHMENT_JOBS, _new_attachment_job, _update_attachment_job

        out1 = WEB_APP / "outputs" / "HQTEST1_附件.zip"
        out2 = WEB_APP / "outputs" / "HQTEST2_附件.zip"
        out1.parent.mkdir(parents=True, exist_ok=True)
        out1.write_bytes(b"PK\x05\x06" + b"\x00" * 18)
        out2.write_bytes(b"PK\x05\x06" + b"\x00" * 18)

        job1 = _new_attachment_job("HQTEST1")
        job2 = _new_attachment_job("HQTEST2")
        _update_attachment_job(job1, status="done", progress=100, stage="下载完成", download="/download/HQTEST1_附件.zip", filename=out1.name, source_path=str(out1))
        _update_attachment_job(job2, status="done", progress=100, stage="下载完成", download="/download/HQTEST2_附件.zip", filename=out2.name, source_path=str(out2))
        batch_id = "batchziptest1234567890"
        _PLM_ATTACHMENT_BATCHES[batch_id] = {
            "id": batch_id,
            "job_ids": [job1, job2],
            "download": "",
            "filename": "",
            "source_path": "",
            "created_at": time.time(),
            "updated_at": time.time(),
        }

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            resp = client.get(f"/api/plm/auto_hq_attachments/batch/status/{batch_id}")
            payload = resp.get_json()
            self.assertTrue(payload["success"], payload)
            self.assertEqual(payload["status"], "done")
            self.assertEqual(payload["progress"], 100)
            self.assertEqual(payload["done"], 2)
            self.assertTrue(payload["download"].endswith(".zip"))
            zip_path = WEB_APP / "outputs" / payload["filename"]
            with ZipFile(zip_path) as zf:
                self.assertEqual(sorted(zf.namelist()), sorted([out1.name, out2.name]))

        _PLM_ATTACHMENT_BATCHES.pop(batch_id, None)
        _PLM_ATTACHMENT_JOBS.pop(job1, None)
        _PLM_ATTACHMENT_JOBS.pop(job2, None)

    def test_plm_attachment_batch_cancel_marks_queued_jobs(self):
        from plm import _PLM_ATTACHMENT_BATCHES, _PLM_ATTACHMENT_JOBS, _new_attachment_job

        job1 = _new_attachment_job("HQTEST1")
        job2 = _new_attachment_job("HQTEST2")
        batch_id = "batchcanceltest123456"
        _PLM_ATTACHMENT_BATCHES[batch_id] = {
            "id": batch_id,
            "job_ids": [job1, job2],
            "download": "",
            "filename": "",
            "source_path": "",
            "created_at": time.time(),
            "updated_at": time.time(),
        }

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            resp = client.post(f"/api/plm/auto_hq_attachments/batch/cancel/{batch_id}")
            payload = resp.get_json()
            self.assertTrue(payload["success"], payload)
            self.assertEqual(payload["cancelled"], 2)
            self.assertEqual(payload["finished"], 2)
            self.assertEqual({job["status"] for job in payload["jobs"]}, {"cancelled"})

        _PLM_ATTACHMENT_BATCHES.pop(batch_id, None)
        _PLM_ATTACHMENT_JOBS.pop(job1, None)
        _PLM_ATTACHMENT_JOBS.pop(job2, None)

    def test_plm_attachment_batch_excel_detects_hqpn_column(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["序号", "HQ PN", "说明"])
        ws.append([1, "HQ17001007NC0", "single"])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            resp = client.post(
                "/api/plm/auto_hq_attachments/excel_detect",
                data={"header_row": "1", "file": (bio, "bom.xlsx")},
                content_type="multipart/form-data",
            )
            payload = resp.get_json()
            self.assertTrue(payload["success"], payload)
            self.assertEqual(payload["detected"].get("hqpn"), "B")
            self.assertEqual(payload["headers"], ["序号", "HQ PN", "说明"])

            second = client.post(
                "/api/plm/auto_hq_attachments/excel_detect",
                data={"header_row": "2", "uid": payload["uid"], "sheet_name": payload["current_sheet"]},
                content_type="multipart/form-data",
            ).get_json()
            self.assertTrue(second["success"], second)
            self.assertEqual(second["uid"], payload["uid"])
            self.assertEqual(second["headers"], ["1", "HQ17001007NC0", "single"])


    def test_plm_attachment_single_jobs_start_new_session_after_success(self):
        from plm import _PLM_ATTACHMENT_JOBS

        init_calls = []
        run_calls = []

        class FakePage:
            def goto(self, *args, **kwargs):
                return None

            def locator(self, *args, **kwargs):
                return self

            def filter(self, *args, **kwargs):
                return self

            def is_closed(self):
                return False

        class FakeContext:
            def new_page(self):
                return FakePage()

            def close(self):
                pass

        class FakeBrowser:
            def new_context(self, accept_downloads=True):
                return FakeContext()

            def close(self):
                pass

        class FakeChromium:
            def launch(self, headless=False):
                return FakeBrowser()

        class FakePlaywright:
            chromium = FakeChromium()

        class FakePlaywrightContext:
            def __enter__(self):
                return FakePlaywright()

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_wait_for_eip_ready(page, username, password):
            init_calls.append(("eip", username))

        def fake_click_opening_page(page, locator, timeout=30000):
            return FakePage()

        def fake_open_search(context, page, username, password, log=None):
            init_calls.append(("search", username))
            if log:
                log("Open PLM search page")
            return FakePage()

        def fake_download(context, search_page, *, hqpn, output_dir, username="", password="", log=None):
            run_calls.append(hqpn)
            out_path = Path(output_dir) / f"{hqpn}_attachment.zip"
            out_path.write_bytes(b"PK\x05\x06" + b"\x00" * 18)
            if log:
                log("Downloaded selected attachments: document_contents.zip")
            return out_path, search_page

        def wait_done(client, job_id):
            for _ in range(40):
                status = client.get(f"/api/plm/auto_hq_attachments/status/{job_id}").get_json()
                if status.get("status") == "done":
                    return status
                if status.get("status") == "error":
                    self.fail(status.get("error"))
                time.sleep(0.05)
            self.fail("job did not finish")

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            with patch("playwright.sync_api.sync_playwright", return_value=FakePlaywrightContext()), patch(
                "plm.automation.wait_for_eip_ready", fake_wait_for_eip_ready
            ), patch("plm.automation.click_opening_page", fake_click_opening_page), patch(
                "plm.automation.login_if_present", lambda *args, **kwargs: False
            ), patch("plm.automation._wait_for_plm_home", lambda context, page, username, password: page), patch(
                "plm.automation._open_plm_search_page", fake_open_search
            ), patch("plm.automation.download_hq_attachment_from_search_page", fake_download):
                job_ids = []
                for hqpn in ("HQTEST1", "HQTEST2"):
                    resp = client.post(
                        "/api/plm/auto_hq_attachments",
                        data={"username": "100448405", "password": "pw", "hqpn": hqpn},
                    )
                    payload = resp.get_json()
                    self.assertTrue(payload["success"])
                    job_ids.append(payload["job_id"])

                first = wait_done(client, job_ids[0])
                second = wait_done(client, job_ids[1])

                self.assertEqual(first["download"], "/download/HQTEST1_attachment.zip")
                self.assertEqual(second["download"], "/download/HQTEST2_attachment.zip")
                self.assertEqual(run_calls, ["HQTEST1", "HQTEST2"])
                self.assertEqual(init_calls.count(("search", "100448405")), 2)

            for job_id in job_ids:
                _PLM_ATTACHMENT_JOBS.pop(job_id, None)
    def test_plm_attachment_job_retries_once_with_existing_session(self):
        from plm import _PLM_ATTACHMENT_JOBS

        init_calls = []
        run_calls = []

        class FakePage:
            def goto(self, *args, **kwargs):
                return None

            def locator(self, *args, **kwargs):
                return self

            def filter(self, *args, **kwargs):
                return self

            def is_closed(self):
                return False

        class FakeContext:
            def new_page(self):
                return FakePage()

            def close(self):
                pass

        class FakeBrowser:
            def new_context(self, accept_downloads=True):
                return FakeContext()

            def close(self):
                pass

        class FakeChromium:
            def launch(self, headless=False):
                return FakeBrowser()

        class FakePlaywright:
            chromium = FakeChromium()

        class FakePlaywrightContext:
            def __enter__(self):
                return FakePlaywright()

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_wait_for_eip_ready(page, username, password):
            init_calls.append(("eip", username))

        def fake_click_opening_page(page, locator, timeout=30000):
            return FakePage()

        def fake_open_search(context, page, username, password, log=None):
            init_calls.append(("search", username))
            if log:
                log("Open PLM search page")
            return FakePage()

        def fake_download(context, search_page, *, hqpn, output_dir, username="", password="", log=None):
            run_calls.append(hqpn)
            if len(run_calls) == 1:
                raise RuntimeError("search page not ready")
            out_path = Path(output_dir) / f"{hqpn}_retry_attachment.zip"
            out_path.write_bytes(b"PK\x05\x06" + b"\x00" * 18)
            if log:
                log("Downloaded selected attachments: document_contents.zip")
            return out_path, search_page

        def wait_done(client, job_id):
            for _ in range(40):
                status = client.get(f"/api/plm/auto_hq_attachments/status/{job_id}").get_json()
                if status.get("status") == "done":
                    return status
                if status.get("status") == "error":
                    self.fail(status.get("error"))
                time.sleep(0.05)
            self.fail("job did not finish")

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            with patch("playwright.sync_api.sync_playwright", return_value=FakePlaywrightContext()), patch(
                "plm.automation.wait_for_eip_ready", fake_wait_for_eip_ready
            ), patch("plm.automation.click_opening_page", fake_click_opening_page), patch(
                "plm.automation.login_if_present", lambda *args, **kwargs: False
            ), patch("plm.automation._wait_for_plm_home", lambda context, page, username, password: page), patch(
                "plm.automation._open_plm_search_page", fake_open_search
            ), patch("plm.automation.download_hq_attachment_from_search_page", fake_download):
                resp = client.post(
                    "/api/plm/auto_hq_attachments",
                    data={"username": "100448406", "password": "pw", "hqpn": "HQRETRY"},
                )
                payload = resp.get_json()
                self.assertTrue(payload["success"], payload)
                status = wait_done(client, payload["job_id"])

                self.assertEqual(status["download"], "/download/HQRETRY_retry_attachment.zip")
                self.assertEqual(run_calls, ["HQRETRY", "HQRETRY"])
                self.assertEqual(init_calls.count(("eip", "100448406")), 1)
                self.assertGreaterEqual(init_calls.count(("search", "100448406")), 2)
                self.assertIn("首次下载失败", status["log"])

            _PLM_ATTACHMENT_JOBS.pop(payload["job_id"], None)

    def test_plm_attachment_job_reports_progress_and_download(self):
        from plm import _PLM_ATTACHMENT_JOBS

        out_path = WEB_APP / "outputs" / "mock_plm_attachment.zip"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(b"PK\x05\x06" + b"\x00" * 18)

        def fake_download(context, search_page, *, hqpn, output_dir, username="", password="", log=None):
            log("Open PLM search page")
            log("Downloaded selected attachments: document_contents.zip")
            return out_path, search_page

        with app.test_client() as client:
            client.post("/api/login", json={"employee_id": "ADMIN"})
            with patch("plm._enqueue_attachment_job") as enqueue:
                enqueue.side_effect = lambda job_id, username, password, hqpn, batch_id="": (
                    fake_download(None, None, hqpn=hqpn, output_dir=WEB_APP / "outputs", log=lambda message: __import__("plm")._append_attachment_log(job_id, message)),
                    __import__("plm")._update_attachment_job(
                        job_id,
                        status="done",
                        stage="下载完成",
                        progress=100,
                        download="/download/mock_plm_attachment.zip",
                        filename="mock_plm_attachment.zip",
                        source_path=str(out_path),
                    ),
                )
                resp = client.post(
                    "/api/plm/auto_hq_attachments",
                    data={"username": "100448405", "password": "pw", "hqpn": "HQTEST"},
                )
                payload = resp.get_json()
                self.assertTrue(payload["success"])
                job_id = payload["job_id"]

                status = client.get(f"/api/plm/auto_hq_attachments/status/{job_id}").get_json()
                self.assertEqual(status["status"], "done")
                self.assertEqual(status["progress"], 100)
                self.assertEqual(status["download"], "/download/mock_plm_attachment.zip")
                self.assertIn("Downloaded selected attachments", status["log"])

            _PLM_ATTACHMENT_JOBS.pop(job_id, None)


if __name__ == "__main__":
    unittest.main()
