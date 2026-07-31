import io
import sys
import unittest
from pathlib import Path
from urllib.parse import unquote

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WEB_APP = ROOT / "web_app2"
if str(WEB_APP) not in sys.path:
    sys.path.insert(0, str(WEB_APP))

from app import app  # noqa: E402
from plm import PLM_HEADERS  # noqa: E402


def _xlsx_bytes(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


class CustomerHqStandardExportTests(unittest.TestCase):
    def test_customer_hq_export_uses_hq_plm_headers_and_maps_legacy_process_fields(self):
        headers = [
            "\u5e8f\u53f7", "HQ PN", "\u751f\u4ea7\u5382\u5bb6", "\u578b\u53f7", "\u5355\u8017", "\u7269\u6599\u63cf\u8ff0", "\u4f4d\u53f7",
            "\u4e3b\u5236\u63a7", "\u5b50\u5236\u63a7", "\u5b50\u5236\u63a7\u6570\u91cf", "\u662f\u5426\u53ef\u91cf\u4ea7\u4e0b\u5355", "\u6b21\u5236\u7a0b\u4f4d\u53f7",
            "\u6e7f\u654f\u5c5e\u6027", "\u4e3b\u8f85BOM\u6807\u8bb0", "IFM_PART", "PCD_PART", "\u662f\u5426\u53d7EAR\u7ba1\u63a7", "ECCN",
        ]
        rows = [[
            "10", "HQ-001", "Maker", "Model-A", 3, "Description", "R1,R2",
            "SMT", "DIP", 2, "\u662f", "R3", "MSL-3", "\u4e8c\u4f9b", "IFM-1", "PCD-1", "\u5426", "5A992",
        ]]
        response = app.test_client().post(
            "/api/plm/customer_hq_convert",
            data={
                "file": (_xlsx_bytes(headers, rows), "customer.xlsx"),
                "sheet": "Input",
                "header_row": "1",
                "col_seq": "A",
                "col_hqpn": "B",
                "col_brand": "C",
                "col_model": "D",
                "col_qty": "E",
                "col_name": "F",
                "col_refdes": "G",
            },
            content_type="multipart/form-data",
        )
        payload = response.get_json()
        self.assertTrue(payload["success"], payload.get("error"))

        filename = unquote(payload["download"].split("/download/", 1)[1])
        workbook = openpyxl.load_workbook(WEB_APP / "outputs" / filename, data_only=True)
        try:
            sheet = workbook["BOM"]
            self.assertEqual([sheet.cell(3, column).value for column in range(1, 26)], PLM_HEADERS)
            self.assertEqual(sheet.max_column, len(PLM_HEADERS))
            self.assertEqual(sheet.cell(3, 16).value, "\u9996\u5236\u7a0b")
            self.assertEqual(sheet.cell(3, 17).value, "\u6b21\u5236\u7a0b")
            self.assertEqual(sheet.cell(3, 18).value, "\u6b21\u5236\u7a0b\u5355\u8017")
            self.assertEqual(sheet.cell(4, 16).value, "SMT")
            self.assertEqual(sheet.cell(4, 17).value, "DIP")
            self.assertEqual(sheet.cell(4, 18).value, "2")
            self.assertEqual(sheet.cell(4, 19).value, "\u662f")
            self.assertEqual(sheet.cell(4, 20).value, "R3")
            self.assertEqual(sheet.cell(4, 10).value, "MSL-3")
            self.assertEqual(sheet.cell(4, 12).value, "\u4e8c\u4f9b")
            self.assertEqual(sheet.cell(4, 22).value, "IFM-1")
            self.assertEqual(sheet.cell(4, 23).value, "PCD-1")
            self.assertEqual(sheet.cell(4, 24).value, "\u5426")
            self.assertEqual(sheet.cell(4, 25).value, "5A992")
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
