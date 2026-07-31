# -*- coding: utf-8 -*-
"""Customer BOM to HQ single-board BOM conversion for the PLM tool."""

import os
import uuid

from activity import track_tool_activity
from openpyxl.styles import Color
from bom import _detect_columns as _detect_customer_bom_columns, _safe_qty
from shared import (
    OUTPUT_DIR,
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    Workbook,
    get_column_letter,
    jsonify,
    request,
    _cell_str,
    _col_int,
    _open_workbook,
    _request_int,
    _save_or_reuse_uploaded_excel,
    _save_uploaded_excel,
)

from . import PLM_HEADERS, plm_bp


HQ_SINGLE_BOARD_HEADERS = PLM_HEADERS
HQ_SINGLE_BOARD_FIELDS = [
    "seq", "hqpn", "model", "description", "qty", "alternate", "refdes", "manufacturer",
    "environmental", "thermal_sensitive", "note", "main_aux", "mbg_preferred", "cbg_preferred",
    "dbg_preferred", "first_process", "second_process", "second_process_qty", "mass_orderable",
    "second_process_refdes", "abg_preferred", "ifm_part", "pcd_part", "ear_control", "eccn",
]

# Accept both historical customer-BOM labels and current HQ/PLM standard labels.
OPTIONAL_HEADER_TO_FIELD = {
    "\u66ff\u4ee3\u5173\u7cfb": "alternate",
    "\u662f\u5426\u73af\u4fdd": "environmental",
    "\u6e7f\u654f\u5c5e\u6027": "thermal_sensitive",
    "\u6e29\u654f\u5c5e\u6027": "thermal_sensitive",
    "\u5907\u6ce8": "note",
    "\u4e3b\u8f85BOM\u6807\u8bb0": "main_aux",
    "MBG\u4f18\u9009\u5c5e\u6027": "mbg_preferred",
    "CBG\u4f18\u9009\u5c5e\u6027": "cbg_preferred",
    "DBG\u4f18\u9009\u5c5e\u6027": "dbg_preferred",
    "\u4e3b\u5236\u63a7": "first_process",
    "\u4e3b\u5236\u7a0b": "first_process",
    "\u9996\u5236\u7a0b": "first_process",
    "\u5b50\u5236\u63a7": "second_process",
    "\u6b21\u5236\u7a0b": "second_process",
    "\u5b50\u5236\u63a7\u6570\u91cf": "second_process_qty",
    "\u6b21\u5236\u7a0b\u5355\u8017": "second_process_qty",
    "\u662f\u5426\u53ef\u91cf\u4ea7\u4e0b\u5355": "mass_orderable",
    "\u6b21\u5236\u7a0b\u4f4d\u53f7": "second_process_refdes",
    "ABG\u4f18\u9009\u5c5e\u6027": "abg_preferred",
    "IFM_PART": "ifm_part",
    "PCD_PART": "pcd_part",
    "\u662f\u5426\u53d7EAR\u7ba1\u63a7": "ear_control",
    "ECCN": "eccn",
}
OPTIONAL_HEADER_TO_FIELD = {
    "".join(key.replace("\uFF08", "(").replace("\uFF09", ")").split()).lower(): value
    for key, value in OPTIONAL_HEADER_TO_FIELD.items()
}

SAMPLE_COLUMN_WIDTHS = [
    5.88671875, 21.44140625, 31.21875, 43.0, 13.6640625, 13.0, 31.21875, 11.6640625, 21.44140625,
    11.6640625, 13.0, 19.5546875, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0,
]


def _detect_seq_column(ws, header_row):
    for ci in range(1, ws.max_column + 1):
        header = _cell_str(ws.cell(row=header_row, column=ci).value).lower()
        if header in ("序号", "seq", "no.", "no"):
            return ci
    return None


def _detect_exact_header_columns(ws, header_row):
    result = {}
    for ci in range(1, ws.max_column + 1):
        header = _cell_str(ws.cell(row=header_row, column=ci).value)
        normalized_header = "".join(header.replace("\uFF08", "(").replace("\uFF09", ")").split()).lower()
        field = OPTIONAL_HEADER_TO_FIELD.get(normalized_header)
        if not field and normalized_header.startswith("\u4e3b\u8f85bom\u6807\u8bb0"):
            field = "main_aux"
        if field:
            result[field] = ci
    return result


def _customer_hq_detect_payload(path, sheet_name, header_row):
    wb = _open_workbook(path, data_only=True)
    try:
        if not sheet_name or sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0] if wb.sheetnames else ""
        ws = wb[sheet_name]
        all_cols, best = _detect_customer_bom_columns(ws, header_row)
        seq_col = _detect_seq_column(ws, header_row)
        headers = [ws.cell(row=header_row, column=ci).value for ci in range(1, ws.max_column + 1)]
        preview = []
        for ri in range(header_row + 1, min(header_row + 51, ws.max_row + 1)):
            row = [ws.cell(row=ri, column=ci).value for ci in range(1, ws.max_column + 1)]
            if any(_cell_str(v) for v in row):
                preview.append([_cell_str(v) for v in row])

        detected = {}
        if seq_col:
            detected["seq"] = get_column_letter(seq_col)
        role_map = {
            "brand_combined": "brand",
            "brand_split": "brand",
            "brand_code": "brand",
            "model_split": "model",
            "model_code": "model",
            "qty": "qty",
            "name": "name",
        }
        for role, info in best.items():
            mapped = role_map.get(role, role)
            if mapped not in detected:
                detected[mapped] = info["letter"]

        return {
            "success": True,
            "sheets": wb.sheetnames,
            "current_sheet": sheet_name,
            "headers": [f"{info['letter']}:{info['header']}" for _, info in sorted(all_cols.items())],
            "preview_headers": [_cell_str(h) for h in headers],
            "preview": preview,
            "detected": detected,
        }
    finally:
        wb.close()


def _form_text(name, default=""):
    return _cell_str(request.form.get(name)) or default


def _write_hq_single_board_bom(rows, out_path, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    thin = Side(style="thin", color=Color(auto=1))
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor=Color(indexed=23))
    label_font = Font(name="Arial", size=12, bold=True)
    body_font = Font(name="宋体", size=10)

    meta_rows = [
        ["料号", meta["part_no"], "描述", meta["description"], "项目配置名", meta["config_name"], "工程师", meta["engineer"]],
        ["版本", meta["version"], "替代项", meta["alternate_item"], "BOM名称", meta["bom_name"], "归档部门", meta["archive_dept"]],
    ]
    for ri, values in enumerate(meta_rows, 1):
        for ci, value in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = body_font
            cell.number_format = "@"
            if ci % 2 == 1:
                cell.font = label_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, header in enumerate(HQ_SINGLE_BOARD_HEADERS, 1):
        cell = ws.cell(row=3, column=ci, value=header)
        cell.font = label_font
        cell.border = border
        cell.fill = header_fill
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, row in enumerate(rows, 4):
        for ci, field in enumerate(HQ_SINGLE_BOARD_FIELDS, 1):
            value = row.get(field, "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if ci != 5:
                cell.number_format = "@"

    for ci, width in enumerate(SAMPLE_COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width
    wb.save(out_path)
    wb.close()


@plm_bp.route("/api/plm/customer_hq_detect", methods=["POST"])
def api_customer_hq_detect():
    file = request.files.get("file")
    try:
        uid, path = _save_or_reuse_uploaded_excel(file, "plm_customer_hq_pre", request.form.get("uid", ""))
        wb = _open_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)})

    sheet_name = request.form.get("sheet_name", "")
    if not sheet_name or sheet_name not in sheets:
        sheet_name = sheets[0] if sheets else ""
    header_row = _request_int("header_row", 1)
    if header_row is None:
        return jsonify({"success": False, "error": "表头行必须是大于等于 1 的数字"})

    try:
        payload = _customer_hq_detect_payload(path, sheet_name, header_row)
        payload["uid"] = uid
        return jsonify(payload)
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)})


@plm_bp.route("/api/plm/customer_hq_convert", methods=["POST"])
@track_tool_activity("客户BOM转换成HQ格式单板BOM")
def api_customer_hq_convert():
    file = request.files.get("file")
    if not file:
        return jsonify({"success": False, "error": "请上传客户 BOM 文件"})

    uid = str(uuid.uuid4())[:8]
    try:
        path = _save_uploaded_excel(file, "plm_customer_hq_in", uid)
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)})

    sheet_name = request.form.get("sheet", "")
    header_row = _request_int("header_row", 1)
    if header_row is None:
        return jsonify({"success": False, "error": "表头行必须是大于等于 1 的数字"})

    col_seq_str = request.form.get("col_seq", "")
    col_hqpn_str = request.form.get("col_hqpn", "")
    col_brand_str = request.form.get("col_brand", "")
    col_model_str = request.form.get("col_model", "")
    col_qty_str = request.form.get("col_qty", "")
    col_name_str = request.form.get("col_name", "")
    col_refdes_str = request.form.get("col_refdes", "")

    wb_ro = _open_workbook(path, read_only=True, data_only=True)
    try:
        sheets = wb_ro.sheetnames
    finally:
        wb_ro.close()
    if not sheet_name or sheet_name not in sheets:
        sheet_name = sheets[0]

    if not all([col_seq_str, col_brand_str, col_qty_str, col_name_str]):
        detected = _customer_hq_detect_payload(path, sheet_name, header_row).get("detected", {})
        col_seq_str = col_seq_str or detected.get("seq", "")
        col_brand_str = col_brand_str or detected.get("brand", "")
        col_model_str = col_model_str or detected.get("model", "")
        col_qty_str = col_qty_str or detected.get("qty", "")
        col_name_str = col_name_str or detected.get("name", "")

    col_seq = _col_int(col_seq_str)
    col_hqpn = _col_int(col_hqpn_str) if str(col_hqpn_str).strip() else None
    col_brand = _col_int(col_brand_str)
    col_model = _col_int(col_model_str) if str(col_model_str).strip() else None
    col_qty = _col_int(col_qty_str)
    col_name = _col_int(col_name_str)
    col_refdes = _col_int(col_refdes_str) if str(col_refdes_str).strip() else None
    if not col_seq:
        return jsonify({"success": False, "error": "请指定序号列"})
    if not col_brand:
        return jsonify({"success": False, "error": "请指定生产厂家列"})
    if not col_model:
        return jsonify({"success": False, "error": "请指定型号列"})
    if not col_qty:
        return jsonify({"success": False, "error": "请指定单耗/用量列"})
    if not col_name:
        return jsonify({"success": False, "error": "请指定物料描述列"})

    meta = {
        "part_no": _form_text("part_no"),
        "description": _form_text("description", _form_text("project_name", "客户BOM")),
        "config_name": _form_text("config_name", _form_text("project_name", "客户BOM")),
        "engineer": _form_text("engineer"),
        "version": _form_text("version"),
        "alternate_item": _form_text("alternate_item"),
        "bom_name": _form_text("bom_name", _form_text("description", _form_text("project_name", "客户BOM"))),
        "archive_dept": _form_text("archive_dept"),
    }

    wb = _open_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        optional_cols = _detect_exact_header_columns(ws, header_row)
        rows = []
        skipped = 0
        invalid_seq_rows = []
        seen_seq = set()
        for ri in range(header_row + 1, ws.max_row + 1):
            row_vals = {ci: ws.cell(row=ri, column=ci).value for ci in range(1, ws.max_column + 1)}
            seq = _cell_str(row_vals.get(col_seq))
            manufacturer = _cell_str(row_vals.get(col_brand))
            model = _cell_str(row_vals.get(col_model))
            description = _cell_str(row_vals.get(col_name))
            if not seq and not any([manufacturer, model, description]):
                skipped += 1
                continue
            if not seq:
                invalid_seq_rows.append(ri)
                continue
            is_main = seq not in seen_seq
            seen_seq.add(seq)
            optional_values = {field: _cell_str(row_vals.get(col_idx)) for field, col_idx in optional_cols.items()}
            rows.append({
                "seq": seq,
                "hqpn": _cell_str(row_vals.get(col_hqpn)) if col_hqpn else "",
                "model": model,
                "description": description,
                "qty": _safe_qty(row_vals.get(col_qty)) if is_main else "",
                "alternate": optional_values.get("alternate", ""),
                "refdes": _cell_str(row_vals.get(col_refdes)) if (is_main and col_refdes) else "",
                "manufacturer": manufacturer,
                "environmental": optional_values.get("environmental", ""),
                "thermal_sensitive": optional_values.get("thermal_sensitive", ""),
                "note": optional_values.get("note", ""),
                "main_aux": optional_values.get("main_aux", ""),
                "mbg_preferred": optional_values.get("mbg_preferred", ""),
                "cbg_preferred": optional_values.get("cbg_preferred", ""),
                "dbg_preferred": optional_values.get("dbg_preferred", ""),
                "first_process": optional_values.get("first_process", ""),
                "second_process": optional_values.get("second_process", ""),
                "second_process_qty": optional_values.get("second_process_qty", ""),
                "abg_preferred": optional_values.get("abg_preferred", ""),
                "mass_orderable": optional_values.get("mass_orderable", ""),
                "second_process_refdes": optional_values.get("second_process_refdes", ""),
                "ifm_part": optional_values.get("ifm_part", ""),
                "pcd_part": optional_values.get("pcd_part", ""),
                "ear_control": optional_values.get("ear_control", ""),
                "eccn": optional_values.get("eccn", ""),
            })
    finally:
        wb.close()

    if invalid_seq_rows:
        preview_rows = ', '.join(str(row) for row in invalid_seq_rows[:10])
        suffix = ' ...' if len(invalid_seq_rows) > 10 else ''
        return jsonify({
            "success": False,
            "error": f"序号列存在空值（Excel 行 {preview_rows}{suffix}）。请补齐序号后再转换，避免错误识别主料和替代料。",
        })

    out_name = f"客户BOM_HQ单板BOM_{uid}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_name)
    _write_hq_single_board_bom(rows, out_path, meta)

    return jsonify({
        "success": True,
        "download": f"/download/{out_name}",
        "total": len(rows),
        "skipped": skipped,
        "columns": {
            "seq": get_column_letter(col_seq),
            "hqpn": get_column_letter(col_hqpn) if col_hqpn else "",
            "brand": get_column_letter(col_brand),
            "model": get_column_letter(col_model),
            "qty": get_column_letter(col_qty),
            "name": get_column_letter(col_name),
            "refdes": get_column_letter(col_refdes) if col_refdes else "",
        },
    })
