# -*- coding: utf-8 -*-
"""飞书多表格匹配 — Blueprint"""

import os, uuid, json, hashlib, time
from flask import Blueprint
from activity import track_tool_activity
from shared import (
    requests as _requests,
    openpyxl, Workbook, Font, PatternFill, Alignment, Border, Side,
    get_column_letter,
    request, jsonify,
    UPLOAD_DIR, OUTPUT_DIR, CACHE_DIR, _cell_str,
    _open_workbook, _resolve_feishu_base_url, _save_uploaded_excel, _to_int,
)
from manufacturer_alias import lookup_manufacturer

os.makedirs(CACHE_DIR, exist_ok=True)

# ── 服务端数据缓存（以 token + sheet_id 为粒度）─────────────────

def _mk_cache_key(token, sheet_id):
    raw = f"{token}:{sheet_id}"
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def _cache_path(key):
    return os.path.join(CACHE_DIR, f"feishu_{key}.json")


def _write_cache(token, sheet_id, rows, row_count_at_cache=0):
    """缓存单个 sheet 的全部行数据"""
    key = _mk_cache_key(token, sheet_id)
    payload = {
        "token": token,
        "sheet_id": sheet_id,
        "fetched_at": time.time(),
        "row_count_at_cache": row_count_at_cache,
        "rows": rows,
    }
    with open(_cache_path(key), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    headers = [_cell_str(v) for v in (rows[0] if rows else [])]
    return key, max(0, len(rows) - 1), headers


def _read_cache(key):
    path = _cache_path(key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _is_preferred_level(value):
    """判断优选等级是否为优选料。"""
    text = str(value).strip() if value is not None else ''
    if not text:
        return False
    lowered = text.lower()
    negative_markers = ('非优选', '不优选', '非推薦', '非推荐', 'not preferred')
    if any(marker in lowered for marker in negative_markers):
        return False
    if '优选' in text or 'preferred' in lowered:
        return True
    try:
        return float(text) >= 7
    except ValueError:
        return False


feishu_bp = Blueprint('feishu', __name__)


# ── 飞书 API ────────────────────────────────────────────────

def _hq_get_sheets(base_url, origin, user_id, token):
    url = f"{base_url.rstrip('/')}/fs/sheet/v1/spreadsheetsMetainfo"
    r = _requests.get(url, params={
        "origin": origin, "userId": user_id,
        "spreadsheetToken": token,
    }, timeout=15)
    r.raise_for_status()
    d = r.json()
    if d.get("code") not in (0, 200):
        raise RuntimeError(f"获取 Sheet 列表失败：{d.get('msg')}（code={d.get('code')}）")
    return [s for s in d["data"]["sheets"] if s.get("title")]


def _hq_read_sheet(base_url, origin, user_id, token,
                   sheet_id, row_count=200000, col_count=100,
                   batch_size=3000, progress_cb=None):
    end_col = get_column_letter(max(col_count, 26))
    all_rows, start = [], 1
    while start <= max(row_count, 1):
        end = min(start + batch_size - 1, row_count)
        params = {
            "origin": origin, "userId": user_id,
            "spreadsheetToken": token,
            "range": f"{sheet_id}!A{start}:{end_col}{end}",
        }
        r = _requests.get(
            f"{base_url.rstrip('/')}/fs/sheet/v1/getSheetsValue",
            params=params, timeout=60,
        )
        r.raise_for_status()
        d = r.json()
        if d.get("code") not in (0, 200):
            raise RuntimeError(f"读取失败：{d.get('msg')}")
        batch = d["data"]["valueRange"].get("values") or []
        if all_rows and batch:
            batch = batch[1:]
        if not batch:
            break
        all_rows.extend(batch)
        if progress_cb:
            progress_cb(len(all_rows))
        expected = end - start + 1
        skip = 1 if start > 1 else 0
        if len(batch) < expected - skip:
            break
        start = end + 1
    while all_rows and not any(_cell_str(v) for v in all_rows[-1]):
        all_rows.pop()
    return all_rows


def _map_local_key_value(value, transform=''):
    text = _cell_str(value)
    if transform == 'manufacturer_alias' and text:
        match = lookup_manufacturer(text)
        if match:
            return _cell_str(match.get('canonical_name'))
    return text


def _match_source_priority(name):
    text = str(name or '')
    return 1 if ('对应关系' in text or '关系库' in text) else 0


def _do_match_multi(local_ws, local_header_row, prepared_tables, all_fetch_cols, out_file):
    max_local_col = local_ws.max_column
    local_header = [local_ws.cell(row=local_header_row, column=ci).value
                    for ci in range(1, max_local_col + 1)]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "匹配结果"
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")
    hq_fill = PatternFill("solid", fgColor="FFFF00")
    src_fill = PatternFill("solid", fgColor="BDD7EE")

    out_hdrs = list(local_header) + all_fetch_cols + ["来源表格"]
    for ci, h in enumerate(out_hdrs, 1):
        c = ws_out.cell(row=1, column=ci, value=h or "")
        c.font = Font(bold=True)
        c.fill = (PatternFill("solid", fgColor="FFC000") if ci > max_local_col else hdr_fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    ws_out.row_dimensions[1].height = 22
    for ci in range(1, max_local_col + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 18
    for ci in range(max_local_col + 1, len(out_hdrs) + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 24

    dr = 2
    total = matched = unmatched = 0

    for ri in range(local_header_row + 1, local_ws.max_row + 1):
        row_vals = [local_ws.cell(row=ri, column=ci).value
                    for ci in range(1, max_local_col + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        total += 1

        candidate_matches = []
        for pt in prepared_tables:
            transforms = pt.get("local_key_transforms", [])
            key = tuple(
                _map_local_key_value(
                    row_vals[lc - 1],
                    transforms[i] if i < len(transforms) else '',
                )
                for i, lc in enumerate(pt["local_key_cols"])
            )
            if not any(k for k in key):
                continue
            matches = pt["lookup"].get(key, [])
            if not matches:
                continue
            for mdict in matches:
                fetch_values = []
                for col_name in all_fetch_cols:
                    lookup_name = pt.get("col_lookup", {}).get(col_name, col_name)
                    fetch_values.append(mdict.get(lookup_name, ""))
                candidate_matches.append({
                    "values": fetch_values,
                    "source": pt["name"],
                    "priority": pt.get("source_priority", 0),
                })

        grouped_matches = {}
        if candidate_matches:
            max_priority = max(item["priority"] for item in candidate_matches)
            for item in candidate_matches:
                if item["priority"] != max_priority:
                    continue
                group_key = tuple(item["values"])
                grouped = grouped_matches.setdefault(group_key, {
                    "values": item["values"],
                    "sources": [],
                })
                if item["source"] not in grouped["sources"]:
                    grouped["sources"].append(item["source"])

        if grouped_matches:
            first = True
            for grouped in grouped_matches.values():
                for ci, val in enumerate(row_vals, 1):
                    c = ws_out.cell(row=dr, column=ci, value=val if first else None)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = bdr
                for j, value in enumerate(grouped["values"]):
                    c = ws_out.cell(row=dr, column=max_local_col + j + 1, value=value)
                    c.fill = hq_fill
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = bdr
                src_col = max_local_col + len(all_fetch_cols) + 1
                c = ws_out.cell(row=dr, column=src_col, value="；".join(grouped["sources"]))
                c.fill = src_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = bdr
                first = False
                dr += 1
            matched += 1
        else:
            for ci, val in enumerate(row_vals, 1):
                c = ws_out.cell(row=dr, column=ci, value=val)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = bdr
            for j in range(len(all_fetch_cols)):
                ws_out.cell(row=dr, column=max_local_col + j + 1).border = bdr
            c = ws_out.cell(row=dr, column=max_local_col + len(all_fetch_cols) + 1, value="未匹配")
            c.border = bdr
            unmatched += 1
            dr += 1

    wb_out.save(out_file)
    return total, matched, unmatched


# ── 路由 ─────────────────────────────────────────────────────

@feishu_bp.route('/api/feishu/load', methods=['POST'])
@track_tool_activity('飞书Sheet缓存')
def api_feishu_load():
    """拉取单个 Sheet 全部数据并缓存到服务端"""
    data = request.get_json(silent=True) or {}
    try:
        base_url = _resolve_feishu_base_url(data.get('base_url'))
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    origin   = data.get('origin', '')
    user_id  = data.get('user_id', '')
    token    = data.get('token', '').strip()
    sheet_id = data.get('sheet_id', '').strip()
    if not token or not sheet_id:
        return jsonify({'success': False, 'error': '请填写 Token 和 Sheet ID'})
    try:
        sheets_meta = _hq_get_sheets(base_url, origin, user_id, token)
        target = next((s for s in sheets_meta if s['sheetId'] == sheet_id), None)
        if not target:
            return jsonify({'success': False, 'error': 'Sheet 不存在或已被删除'})
        row_count = target.get('rowCount', 200000) or 200000
        col_count = target.get('columnCount', 100) or 100
        rows = _hq_read_sheet(base_url, origin, user_id, token,
                              sheet_id, row_count=row_count, col_count=col_count)
        if not rows:
            return jsonify({'success': False, 'error': '读取到 0 行数据'})
        row_count_at_cache = target.get('rowCount', 0)
        key, data_rows, headers = _write_cache(token, sheet_id, rows, row_count_at_cache)
        return jsonify({
            'success': True,
            'cache_key': key,
            'row_count': data_rows,
            'headers': headers,
            'fetched_at': time.time(),
            'row_count_at_cache': row_count_at_cache,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@feishu_bp.route('/api/feishu/sheets', methods=['POST'])
def api_feishu_sheets():
    """获取飞书表格的 Sheet 列表，同时读取每个 Sheet 的第一行表头"""
    data = request.get_json(silent=True) or {}
    try:
        base_url = _resolve_feishu_base_url(data.get('base_url'))
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    origin = data.get('origin', '')
    user_id = data.get('user_id', '')
    token = data.get('token', '').strip()
    if not token:
        return jsonify({'success': False, 'error': '请填写 Token'})
    try:
        sheets = _hq_get_sheets(base_url, origin, user_id, token)
        sheet_headers = {}
        for s in sheets:
            sid = s['sheetId']
            col_count = s.get('columnCount', 100) or 100
            try:
                rows = _hq_read_sheet(base_url, origin, user_id, token,
                                      sid, row_count=2, col_count=col_count, batch_size=20)
                sheet_headers[sid] = [_cell_str(v) for v in (rows[0] if rows else [])]
            except Exception:
                sheet_headers[sid] = []
        return jsonify({'success': True, 'sheets': sheets, 'sheet_headers': sheet_headers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@feishu_bp.route('/api/feishu/match', methods=['POST'])
@track_tool_activity('飞书优选库匹配')
def api_feishu_match():
    """执行飞书多表格匹配（per-sheet 配置格式）"""
    local_file = request.files.get('file')
    if not local_file:
        return jsonify({'success': False, 'error': '请上传本地 Excel 文件'})

    config_str = request.form.get('config', '{}')
    try:
        config = json.loads(config_str)
    except Exception:
        return jsonify({'success': False, 'error': 'config 参数格式错误'})

    try:
        base_url = _resolve_feishu_base_url(config.get('base_url'))
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    origin     = config.get('origin', '')
    user_id    = config.get('user_id', '')
    sheet_name = config.get('sheet_name', '')
    header_row = _to_int(config.get('header_row', 1), 1)
    if header_row is None:
        return jsonify({'success': False, 'error': '表头行必须是大于等于 1 的数字'})
    tables_cfg = config.get('tables', [])

    uid = str(uuid.uuid4())[:8]
    try:
        local_path = _save_uploaded_excel(local_file, "fs_local", uid)
        wb_local = _open_workbook(local_path, data_only=True)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    sheets = wb_local.sheetnames
    if not sheet_name or sheet_name not in sheets:
        sheet_name = sheets[0]
    ws_local = wb_local[sheet_name]
    local_header = [ws_local.cell(row=header_row, column=ci).value
                    for ci in range(1, ws_local.max_column + 1)]
    local_header_strs = [_cell_str(h) for h in local_header]

    logs = []
    prepared_tables = []
    all_fetch_cols_ordered = []
    seen_fetch_cols = set()

    for tcfg in tables_cfg:
        table_name = tcfg.get('name', '')
        token = tcfg.get('token', '').strip()
        if not token:
            continue

        for scfg in tcfg.get('sheets', []):
            if not scfg.get('enabled', False):
                continue
            sid        = scfg.get('sheet_id', '').strip()
            sname      = scfg.get('sheet_name', sid)
            full_name  = f"{table_name} - {sname}"
            local_keys = [k for k in scfg.get('local_key_names', []) if k]
            feishu_keys= [k for k in scfg.get('feishu_key_names', []) if k]
            local_key_transforms = list(scfg.get('local_key_transforms', []))
            fetch_cols = [c for c in scfg.get('fetch_col_names', []) if c]
            if len(local_key_transforms) < len(local_keys):
                local_key_transforms += [''] * (len(local_keys) - len(local_key_transforms))
            elif len(local_key_transforms) > len(local_keys):
                local_key_transforms = local_key_transforms[:len(local_keys)]

            if not sid:
                logs.append(f"[{full_name}] 跳过：Sheet ID 为空")
                continue
            if not local_keys or not feishu_keys:
                logs.append(f"[{full_name}] 跳过：匹配键未配置")
                continue
            if len(local_keys) != len(feishu_keys):
                logs.append(f"[{full_name}] 跳过：本地键与飞书键数量不匹配")
                continue

            # Map local key names → column indices
            local_key_cols = []
            for kn in local_keys:
                try:
                    local_key_cols.append(local_header_strs.index(kn) + 1)
                except ValueError:
                    logs.append(f"[{full_name}] 跳过：本地列「{kn}」不存在")
                    local_key_cols = []
                    break
            if not local_key_cols:
                continue

            # Prefer server-side cache
            cache_key = scfg.get('cache_key', '')
            rows = None
            if cache_key:
                cached = _read_cache(cache_key)
                if cached:
                    rows = cached['rows']
                    logs.append(f"[{full_name}] 使用缓存（{len(rows)-1} 行，"
                                f"{time.strftime('%Y-%m-%d %H:%M', time.localtime(cached['fetched_at']))}）")
                else:
                    logs.append(f"[{full_name}] 缓存已失效，实时拉取...")
            if rows is None:
                try:
                    logs.append(f"[{full_name}] 正在读取 Sheet 数据...")
                    sheets_meta = _hq_get_sheets(base_url, origin, user_id, token)
                    target = next((s for s in sheets_meta if s['sheetId'] == sid), None)
                    if not target:
                        logs.append(f"[{full_name}] 跳过：Sheet 已失效，请重新配置")
                        continue
                    rows = _hq_read_sheet(base_url, origin, user_id, token, sid,
                                         row_count=target.get('rowCount', 200000) or 200000,
                                         col_count=target.get('columnCount', 100) or 100)
                    if not rows:
                        logs.append(f"[{full_name}] 读取到 0 行，跳过")
                        continue
                    logs.append(f"[{full_name}] 读取 {len(rows)-1} 行")
                except Exception as e:
                    logs.append(f"[{full_name}] 读取失败：{e}")
                    continue

            fs_headers = [_cell_str(v) for v in rows[0]]
            logs.append(f"[{full_name}] 共 {len(fs_headers)} 列")

            bad_keys = [k for k in feishu_keys if k not in fs_headers]
            if bad_keys:
                logs.append(f"[{full_name}] 跳过：飞书列 {bad_keys} 不存在")
                continue

            feishu_key_indices = [fs_headers.index(k) for k in feishu_keys]
            lookup = {}
            for row in rows[1:]:
                padded = list(row) + [""] * max(0, len(fs_headers) - len(row))
                key = tuple(_cell_str(padded[i]) if i < len(padded) else "" for i in feishu_key_indices)
                row_dict = {fs_headers[i]: _cell_str(padded[i]) if i < len(padded) else ""
                            for i in range(len(fs_headers))}
                lookup.setdefault(key, []).append(row_dict)

            # Build col_lookup: output_name → actual feishu column name
            # Priority: fetch_col_map (global mapped with aliases) > fetch_col_names (direct)
            col_lookup = {}
            fetch_col_map_cfg = scfg.get('fetch_col_map', [])  # [{output, alias}]
            for fm in fetch_col_map_cfg:
                out_name = (fm.get('output') or '').strip()
                alias = (fm.get('alias') or '').strip() or out_name
                if out_name:
                    col_lookup[out_name] = alias
                    if out_name not in seen_fetch_cols:
                        all_fetch_cols_ordered.append(out_name)
                        seen_fetch_cols.add(out_name)
            # Old per-sheet direct columns (identity mapping, output == lookup)
            for cn in fetch_cols:
                if cn and cn not in col_lookup:
                    col_lookup[cn] = cn
                    if cn not in seen_fetch_cols:
                        all_fetch_cols_ordered.append(cn)
                        seen_fetch_cols.add(cn)

            prepared_tables.append({
                "name": full_name,
                "local_key_cols": local_key_cols,
                "local_key_transforms": local_key_transforms,
                "source_priority": _match_source_priority(full_name),
                "lookup": lookup,
                "col_lookup": col_lookup,
            })
            logs.append(f"[{full_name}] 就绪，{len(lookup)} 个唯一键")

    if not prepared_tables:
        wb_local.close()
        return jsonify({'success': False, 'error': '没有可用的 Sheet，请检查配置', 'logs': logs})

    out_name = f"飞书匹配结果_{uid}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_name)

    try:
        total, matched, unmatched = _do_match_multi(
            ws_local, header_row, prepared_tables, all_fetch_cols_ordered, out_path)
        wb_local.close()
        logs.append(f"完成：共 {total} 行，匹配 {matched}，未匹配 {unmatched}")
        return jsonify({
            'success': True,
            'download': f'/download/{out_name}',
            'total': total,
            'matched': matched,
            'unmatched': unmatched,
            'logs': logs,
        })
    except Exception as e:
        wb_local.close()
        logs.append(f"匹配出错：{e}")
        return jsonify({'success': False, 'error': str(e), 'logs': logs})


@feishu_bp.route('/api/feishu/local_sheets', methods=['POST'])
def api_feishu_local_sheets():
    """获取本地 Excel 的 Sheet 列表和列标题"""
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'error': '请上传文件'})
    uid = str(uuid.uuid4())[:8]
    try:
        path = _save_uploaded_excel(file, "fs_pre", uid)
        wb = _open_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

    sheet_name = request.form.get('sheet_name', '')
    if not sheet_name or sheet_name not in sheets:
        sheet_name = sheets[0] if sheets else ''
    header_row = _to_int(request.form.get('header_row', 1), 1)
    if header_row is None:
        return jsonify({'success': False, 'error': '表头行必须是大于等于 1 的数字'})

    try:
        wb2 = _open_workbook(path, data_only=True)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    ws = wb2[sheet_name] if sheet_name else wb2[wb2.sheetnames[0]]
    headers = [_cell_str(ws.cell(row=header_row, column=ci).value)
               for ci in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]
    wb2.close()

    return jsonify({'success': True, 'sheets': sheets, 'current_sheet': sheet_name,
                    'headers': headers, 'uid': uid})

@feishu_bp.route('/api/feishu/pref_rate', methods=['POST'])
@track_tool_activity('查询BOM优选率')
def api_pref_rate():
    """查询BOM优选率：按 HQ料号 在所有优选库缓存中查找优选等级"""
    local_file = request.files.get('file')
    if not local_file:
        return jsonify({'success': False, 'error': '未上传文件'})

    config_str = request.form.get('config', '{}')
    try:
        config = json.loads(config_str)
    except Exception:
        return jsonify({'success': False, 'error': 'config 参数格式错误'})

    header_row   = _to_int(config.get('header_row', 1), 1)
    if header_row is None:
        return jsonify({'success': False, 'error': '表头行必须是大于等于 1 的数字'})
    sheet_name   = config.get('sheet_name', '')
    local_key_col = config.get('local_key_col', '')
    tables_cfg   = config.get('tables', [])   # [{name, sheets:[{sid,name,cache_key,fetch_col_aliases}]}]

    uid = str(uuid.uuid4())[:8]
    try:
        local_path = _save_uploaded_excel(local_file, "pref", uid)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})

    try:
        try:
            wb = _open_workbook(local_path, data_only=True)
        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)})
        sheets = wb.sheetnames
        if not sheet_name or sheet_name not in sheets:
            sheet_name = sheets[0]
        ws = wb[sheet_name]

        headers = [_cell_str(ws.cell(row=header_row, column=ci).value)
                   for ci in range(1, ws.max_column + 1)]
        if local_key_col not in headers:
            return jsonify({'success': False,
                            'error': f'列 "{local_key_col}" 不存在，请检查表头行设置'})
        key_col_idx = headers.index(local_key_col)

        # Build combined lookup: hq_no → {pref, source}
        combined = {}
        for tcfg in tables_cfg:
            tname = tcfg.get('name', '')
            for scfg in tcfg.get('sheets', []):
                cache_key = scfg.get('cache_key', '')
                aliases   = scfg.get('fetch_col_aliases', {})
                if not cache_key:
                    continue
                cached = _read_cache(cache_key)
                if not cached or not cached.get('rows'):
                    continue
                rows = cached['rows']
                fs_hdrs = [_cell_str(v) for v in rows[0]]
                hq_col   = aliases.get('HQ料号', 'HQ料号')
                pref_col = aliases.get('优选等级', '优选等级')
                if hq_col not in fs_hdrs or pref_col not in fs_hdrs:
                    continue
                hi = fs_hdrs.index(hq_col)
                pi = fs_hdrs.index(pref_col)
                for row in rows[1:]:
                    padded = list(row) + [''] * max(0, max(hi, pi) + 1 - len(row))
                    hv = _cell_str(padded[hi]) if hi < len(padded) else ''
                    pv = _cell_str(padded[pi]) if pi < len(padded) else ''
                    if hv and hv not in combined:
                        combined[hv] = {'pref': pv, 'source': tname}

        # Write output
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        pref_fill    = PatternFill('solid', fgColor='E8F5E9')   # 绿：优选
        nonpref_fill = PatternFill('solid', fgColor='FFF9C4')   # 黄：匹配到但非优选
        nomatch_fill = PatternFill('solid', fgColor='F5F5F5')   # 灰：未匹配
        hdr_font = Font(bold=True)
        bdr = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'))
        center = Alignment(horizontal='center', vertical='center')
        left   = Alignment(horizontal='left',   vertical='center')

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = '优选率查询'

        out_headers = headers + ['优选等级', '来源']
        for ci, h in enumerate(out_headers, 1):
            c = ws_out.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.border = bdr; c.alignment = center

        total = matched = preferred = dr = 0
        for ri in range(header_row + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=ri, column=ci).value
                        for ci in range(1, ws.max_column + 1)]
            if not any(v is not None and str(v).strip() for v in row_vals):
                continue
            total += 1
            dr += 1
            kv = _cell_str(row_vals[key_col_idx]) if key_col_idx < len(row_vals) else ''
            m  = combined.get(kv) if kv else None
            if m:
                matched += 1
                is_pref = _is_preferred_level(m['pref'])
                if is_pref:
                    preferred += 1
                fill = pref_fill if is_pref else nonpref_fill
            else:
                fill = nomatch_fill
            out_row = list(row_vals) + [m['pref'] if m else '', m['source'] if m else '']
            for ci, v in enumerate(out_row, 1):
                c = ws_out.cell(row=dr + 1, column=ci, value=v)
                c.fill = fill; c.border = bdr
                c.alignment = center if ci > len(headers) else left

        wb.close()
        out_name = f'pref_rate_{uid}.xlsx'
        wb_out.save(os.path.join(OUTPUT_DIR, out_name))

        return jsonify({
            'success': True,
            'download': f'/download/{out_name}',
            'total': total,
            'matched': matched,
            'unmatched': total - matched,
            'preferred': preferred,
            'non_preferred': matched - preferred,
            # 优选率 = 优选料 / 已匹配料（未匹配不参与）
            'rate': f'{preferred / matched * 100:.1f}%' if matched else 'N/A',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})



