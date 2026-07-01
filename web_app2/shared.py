# -*- coding: utf-8 -*-
"""BOM Tools Web v2 — 公共模块"""

import os, sys, time, re, subprocess, uuid
from zipfile import BadZipFile

_parent = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _parent not in sys.path:
    sys.path.insert(0, _parent)

import openpyxl
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

from flask import request, jsonify, send_file

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
CACHE_DIR  = os.path.join(os.path.dirname(__file__), "cache")
PLATFORM_VERSION = "2.2.2"
TOOL_VERSIONS = {
    "bom": "5.10.0",
    "bom-checklist": "0.1.4",
    "feishu": "3.0.4",
    "manufacturer-alias": "1.0.0",
    "pref-rate": "1.0.0",
    "plm": "1.7.1",
    "plm-auto": "1.2.0",
    "bom-compare": "0.3.2",
    "free-bom-compare": "1.1.4",
    "customer-hq-compare": "1.1.1",
    "hq-version-compare": "1.0.1",
    "machine-hq-version-compare": "1.0.1",
    "cadence-hq-compare": "1.0.1",
    "toolbox": "1.0.0",
    "manual": "1.0.0",
}

BAD_EXCEL_ERROR = "\u65e0\u6cd5\u8bfb\u53d6\u6587\u4ef6\uff0c\u53ef\u80fd\u539f\u56e0\uff1a\u2460 \u6587\u4ef6\u662f .xls \u65e7\u683c\u5f0f\uff08\u8bf7\u53e6\u5b58\u4e3a .xlsx\uff09\uff1b\u2461 \u516c\u53f8\u52a0\u89e3\u5bc6\u8f6f\u4ef6\u672a\u542f\u52a8\u5bfc\u81f4\u6587\u4ef6\u88ab\u52a0\u5bc6\uff0c\u8bf7\u68c0\u67e5\u540e\u91cd\u8bd5"
XLS_CONVERT_ERROR = "\u65e0\u6cd5\u76f4\u63a5\u8bfb\u53d6\u8be5 .xls \u6587\u4ef6\u3002\u8bf7\u786e\u8ba4\u670d\u52a1\u5668\u4e3a Windows \u4e14\u5df2\u5b89\u88c5\u53ef\u89e3\u5bc6\u6b64\u6587\u4ef6\u7684 Excel\uff0c\u6216\u5148\u5728 Excel \u4e2d\u53e6\u5b58\u4e3a .xlsx \u540e\u518d\u4e0a\u4f20\u3002"

FEISHU_BASE_URLS = {
    "huaqin": "https://mcenter.huaqin.com",
}

FEISHU_PRESET_TABLES = [
    {"name": "MLCC",               "token": "shthq7d9W17DSo7cwuFhtIg7KPf", "category": "优选库"},
    {"name": "电阻",                "token": "shthqdJvubPLY8mrO8qkOMXmGiw", "category": "优选库"},
    {"name": "电解电容",             "token": "shthqO56eTG9DyJX60nDaaMGp0e", "category": "优选库"},
    {"name": "网络变压器/电感器",     "token": "shthquBGGQB8twAgmJWSBwFY5he", "category": "优选库"},
    {"name": "磁珠",                "token": "shthq1EJlpgHfBqBGNediRZdt8c", "category": "优选库"},
    {"name": "晶体晶振",             "token": "shthqpy6hg6hVD78VNPElmcSF6d", "category": "优选库"},
    {"name": "保险丝",               "token": "shthqpsZCFkwGjn62CjRYCpKrHg", "category": "优选库"},
    {"name": "纽扣电池",             "token": "shthqngCNYdufotGIcL6Vn6gWWh", "category": "优选库"},
    {"name": "滤波器/共模扼流圈",     "token": "shthqaZbFrblnh3V0A3ahhOj4Og", "category": "优选库"},
    {"name": "Power IC优选库",       "token": "shthqz7lKPJt9UGF4FOIU1uyTUh", "category": "优选库"},
    {"name": "功能IC优选库",          "token": "shthq4b1PTCh1HqyalUal6aTYte", "category": "优选库"},
    {"name": "DBG分立器件优选库",     "token": "shthqEZrwmemvVULwrhmyAmlfzd", "category": "优选库"},
    {"name": "连接器",               "token": "shthqE9sVI2DkIBYkSkLcUdNxvn", "category": "优选库"},
    {"name": "Cable",               "token": "shthqpYELkJAH7b0uPn1HRcEyLg",  "category": "优选库"},
    {"name": "客户物料型号与HQ料号对应关系", "token": "shthq1R9G7zSp5hvTISGNDOWjme", "category": "对应关系库"},
]


def _cell_str(val):
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, list):
        return " ".join(
            item.get("text", "") or item.get("link", "")
            for item in val if isinstance(item, dict)
        ).strip()
    return str(val).strip()


def _cleanup_old_files(directory, minutes=30):
    now = time.time()
    for f in os.listdir(directory):
        fp = os.path.join(directory, f)
        try:
            if os.path.isfile(fp) and now - os.path.getmtime(fp) > minutes * 60:
                os.remove(fp)
        except Exception:
            pass


def _col_int(s):
    if not s:
        return None
    s = str(s).strip().upper()
    try:
        return int(s) if s.isdigit() else column_index_from_string(s)
    except Exception:
        return None

def _to_int(value, default=1, min_value=1):
    try:
        result = int(value if value is not None else default)
    except (TypeError, ValueError):
        return None
    if min_value is not None and result < min_value:
        return None
    return result


def _request_int(name, default=1, min_value=1):
    return _to_int(request.form.get(name, default), default, min_value)


def _ps_single_quote(value):
    return "'" + str(value).replace("'", "''") + "'"


def _convert_xls_with_excel(src_path, uid, prefix="converted"):
    if os.name != "nt":
        raise ValueError(XLS_CONVERT_ERROR)
    out_path = os.path.join(UPLOAD_DIR, f"{prefix}_converted_{uid}.xlsx")
    script = (
        "$ErrorActionPreference='Stop';"
        f"$src={_ps_single_quote(src_path)};"
        f"$dst={_ps_single_quote(out_path)};"
        "$excel=New-Object -ComObject Excel.Application;"
        "$excel.Visible=$false;$excel.DisplayAlerts=$false;"
        "try{$wb=$excel.Workbooks.Open($src);$wb.SaveAs($dst,51);$wb.Close($false)}"
        "finally{$excel.Quit();[System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)|Out-Null}"
    )
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            check=True, capture_output=True, text=True, timeout=90,
        )
    except Exception as exc:
        raise ValueError(XLS_CONVERT_ERROR) from exc
    if not os.path.exists(out_path):
        raise ValueError(XLS_CONVERT_ERROR)
    return out_path


def _save_uploaded_excel(file, prefix, uid):
    if not file:
        raise ValueError("\u8bf7\u4e0a\u4f20\u6587\u4ef6")
    filename = file.filename or ""
    lower = filename.lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        raw_path = os.path.join(UPLOAD_DIR, f"{prefix}_{uid}.xls")
        file.save(raw_path)
        return _convert_xls_with_excel(raw_path, uid, prefix)
    path = os.path.join(UPLOAD_DIR, f"{prefix}_{uid}.xlsx")
    file.save(path)
    return path




def _valid_upload_uid(uid):
    uid = str(uid or '').strip()
    return bool(uid) and len(uid) <= 32 and all(c.isalnum() or c in '_-' for c in uid)


def _uploaded_excel_path(prefix, uid):
    if not _valid_upload_uid(uid):
        return None
    for name in (f"{prefix}_{uid}.xlsx", f"{prefix}_converted_{uid}.xlsx"):
        path = os.path.join(UPLOAD_DIR, name)
        if os.path.exists(path):
            return path
    return None


def _save_or_reuse_uploaded_excel(file, prefix, uid=None):
    if file:
        uid = str(uuid.uuid4())[:8] if 'uuid' in globals() else str(int(time.time() * 1000))[-8:]
        return uid, _save_uploaded_excel(file, prefix, uid)
    path = _uploaded_excel_path(prefix, uid)
    if not path:
        raise ValueError('Local file cache expired; please upload the file again')
    return str(uid).strip(), path

def _open_workbook(path, **kwargs):
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except BadZipFile as exc:
        raise ValueError(BAD_EXCEL_ERROR) from exc


def _resolve_feishu_base_url(value):
    key_or_url = str(value or "").strip()
    if not key_or_url:
        return FEISHU_BASE_URLS["huaqin"]
    if key_or_url in FEISHU_BASE_URLS:
        return FEISHU_BASE_URLS[key_or_url]
    allowed = {url.rstrip("/") for url in FEISHU_BASE_URLS.values()}
    normalized = key_or_url.rstrip("/")
    if normalized in allowed:
        return normalized
    raise ValueError("\u4e0d\u652f\u6301\u7684\u98de\u4e66\u7f51\u5173\u5730\u5740\uff0c\u8bf7\u4f7f\u7528\u7cfb\u7edf\u9884\u8bbe\u5730\u5740")

