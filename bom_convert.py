# -*- coding: utf-8 -*-
"""
BOM转换脚本：整机BOM配置表 → PLM系统上传格式
使用方法：python bom_convert.py
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# ========== 配置区 ==========
INPUT_FILE = "BOM导入模板 .xlsx"       # 源文件名，改成你的实际文件名
INPUT_SHEET = "SW节点整机BOM配置"      # 源文件的Sheet名
HEADER_ROW = 4                         # 表头所在行（第4行）
DATA_START_ROW = 5                     # 数据起始行

# 源文件列号（A=1, B=2, ...）
COL_SEQ = 1        # A列：序号
COL_PN = 5         # E列：HQ PN（料号）
COL_QTY = 11       # K列：用量（根据图1，用量在K列）

OUTPUT_FILE = "PLM上传BOM.xlsx"        # 输出文件名
# ============================


def find_columns(ws, header_row):
    """自动识别列位置，以防列顺序不同"""
    col_map = {}
    for cell in ws[header_row]:
        if cell.value:
            val = str(cell.value).strip()
            if val == "序号":
                col_map["seq"] = cell.column
            elif val in ("HQ PN", "HQPN", "料号"):
                col_map["pn"] = cell.column
            elif val == "用量":
                col_map["qty"] = cell.column
    return col_map


def convert_bom(input_file, input_sheet, output_file):
    print(f"读取文件：{input_file}，Sheet：{input_sheet}")

    try:
        wb_in = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"[错误] 无法打开文件：{e}")
        return

    if input_sheet not in wb_in.sheetnames:
        print(f"[错误] 找不到Sheet '{input_sheet}'，现有Sheet：{wb_in.sheetnames}")
        return

    ws_in = wb_in[input_sheet]

    # 自动识别列
    col_map = find_columns(ws_in, HEADER_ROW)
    seq_col = col_map.get("seq", COL_SEQ)
    pn_col = col_map.get("pn", COL_PN)
    qty_col = col_map.get("qty", COL_QTY)

    print(f"列位置识别：序号={seq_col}, 料号={pn_col}, 用量={qty_col}")

    # 读取并过滤数据
    rows_to_import = []
    for row_idx in range(DATA_START_ROW, ws_in.max_row + 1):
        seq_val = ws_in.cell(row=row_idx, column=seq_col).value
        pn_val = ws_in.cell(row=row_idx, column=pn_col).value
        qty_val = ws_in.cell(row=row_idx, column=qty_col).value

        # 跳过完全空行
        if seq_val is None and pn_val is None:
            continue

        # 用量为空 → 不导入
        if qty_val is None or str(qty_val).strip() == "":
            continue

        # 用量为0 → 导入但单耗留空
        try:
            qty_num = float(qty_val)
        except (ValueError, TypeError):
            qty_num = None

        if qty_num == 0:
            single_consume = None   # 单耗留空，不填0
        else:
            single_consume = qty_val

        rows_to_import.append({
            "seq": seq_val,
            "pn": pn_val,
            "qty": single_consume,
        })

    print(f"共筛选出 {len(rows_to_import)} 条物料（用量为空的已排除）")

    # 创建输出文件
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "BOM"

    # 第1行：料号信息行（留空，用户填写）
    ws_out["A1"] = "料号："
    ws_out["B1"] = ""
    ws_out["C1"] = "描述："
    ws_out["E1"] = "项目配置名："
    ws_out["G1"] = "工程师："

    # 第2行
    ws_out["A2"] = "版本："
    ws_out["C2"] = "替代项："
    ws_out["E2"] = "BOM名称："
    ws_out["G2"] = "归档部门："

    # 第3行：表头
    headers = ["序号", "料号", "型号", "物料描述", "单耗", "替代关系\n(A:完全替代/N:\n独供/X:不完全替代)", "位号"]
    header_style = Font(bold=True, color="FF0000")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_out.cell(row=3, column=col_idx, value=h)
        cell.font = header_style
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_out.row_dimensions[3].height = 50

    # 写入数据（从第4行开始）
    for data_row_idx, item in enumerate(rows_to_import, start=4):
        ws_out.cell(row=data_row_idx, column=1, value=item["seq"])
        ws_out.cell(row=data_row_idx, column=2, value=item["pn"])
        # 型号、物料描述留空
        if item["qty"] is not None:
            ws_out.cell(row=data_row_idx, column=5, value=item["qty"])
        # 替代关系、位号留空

    # 调整列宽
    col_widths = [8, 20, 15, 20, 8, 20, 15]
    for i, w in enumerate(col_widths, start=1):
        ws_out.column_dimensions[ws_out.cell(row=1, column=i).column_letter].width = w

    wb_out.save(output_file)
    print(f"转换完成！输出文件：{output_file}")
    print(f"  - 有用量物料（单耗已填）：{sum(1 for r in rows_to_import if r['qty'] is not None)} 条")
    print(f"  - 用量为0物料（单耗留空）：{sum(1 for r in rows_to_import if r['qty'] is None)} 条")


if __name__ == "__main__":
    convert_bom(INPUT_FILE, INPUT_SHEET, OUTPUT_FILE)
