# -*- coding: utf-8 -*-
"""飞书多表格匹配 — Blueprint"""

import os, uuid, json, hashlib, time, re
from flask import Blueprint
from activity import track_tool_activity
from shared import (
    requests as _requests,
    openpyxl, Workbook, Font, PatternFill, Alignment, Border, Side,
    get_column_letter,
    request, jsonify,
    UPLOAD_DIR, OUTPUT_DIR, CACHE_DIR, _cell_str,
    _open_workbook, _resolve_feishu_base_url, _save_uploaded_excel, _save_or_reuse_uploaded_excel, _to_int,
)
from manufacturer_alias import lookup_manufacturer

os.makedirs(CACHE_DIR, exist_ok=True)

# ── 服务端数据缓存（以 token + sheet_id 为粒度）─────────────────

def _mk_cache_key(token, sheet_id):
    raw = f"{token}:{sheet_id}"
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def _cache_path(key):
    return os.path.join(CACHE_DIR, f"feishu_{key}.json")


def _write_cache(token, sheet_id, rows, row_count_at_cache=0):
    """缓存单个 sheet 的全部行数据"""
    key = _mk_cache_key(token, sheet_id)
    payload = {
        "token": token,
        "sheet_id": sheet_id,
        "fetched_at": time.time(),
        "row_count_at_cache": row_count_at_cache,
        "rows": rows,
    }
    with open(_cache_path(key), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    headers = [_cell_str(v) for v in (rows[0] if rows else [])]
    return key, max(0, len(rows) - 1), headers


def _read_cache(key):
    path = _cache_path(key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None



def _delete_cache(token, sheet_id):
    key = _mk_cache_key(token, sheet_id)
    path = _cache_path(key)
    existed = os.path.exists(path)
    if existed:
        os.remove(path)
    return key, existed


def _is_preferred_level(value):
    """判断优选等级是否为优选料。"""
    text = str(value).strip() if value is not None else ''
    if not text:
        return False
    lowered = text.lower()
    negative_markers = ('非优选', '不优选', '非推薦', '非推荐', 'not preferred')
    if any(marker in lowered for marker in negative_markers):
        return False
    if '优选' in text or 'preferred' in lowered:
        return True
    try:
        return float(text) >= 7
    except ValueError:
        return False


feishu_bp = Blueprint('feishu', __name__)


# ── 飞书 API ────────────────────────────────────────────────

def _hq_get_sheets(base_url, origin, user_id, token):
    url = f"{base_url.rstrip('/')}/fs/sheet/v1/spreadsheetsMetainfo"
    r = _requests.get(url, params={
        "origin": origin, "userId": user_id,
        "spreadsheetToken": token,
    }, timeout=15)
    r.raise_for_status()
    d = r.json()
    if d.get("code") not in (0, 200):
        raise RuntimeError(f"获取 Sheet 列表失败：{d.get('msg')}（code={d.get('code')}）")
    return [s for s in d["data"]["sheets"] if s.get("title")]



def _hq_get_values_range(base_url, origin, user_id, token, range_name, timeout=60):
    params = {
        "origin": origin,
        "userId": user_id,
        "spreadsheetToken": token,
        "range": range_name,
        "valueRenderOption": "FormattedValue",
        "dateTimeRenderOption": "FormattedString",
    }
    r = _requests.get(
        f"{base_url.rstrip('/')}/fs/sheet/v1/getSheetsValue",
        params=params,
        timeout=timeout,
    )
    r.raise_for_status()
    d = r.json()
    if d.get("code") not in (0, 200):
        raise RuntimeError(f"\u8bfb\u53d6\u5931\u8d25\uff1a{d.get('msg')}")
    return d["data"]["valueRange"].get("values") or []


def _normalize_dense_rows(rows, width, header=None):
    dense = []
    for row in rows or []:
        values = list(row) if isinstance(row, list) else [row]
        dense.append(values[:width] + [""] * max(0, width - len(values)))
    if header is not None and dense:
        dense[0] = list(header)[:width] + [""] * max(0, width - len(header))
    while dense and not any(_cell_str(v) for v in dense[-1]):
        dense.pop()
    return dense


def _hq_read_sheet_by_columns(base_url, origin, user_id, token,
                              sheet_id, row_count, last_header_col, header,
                              batch_size=3000, progress_cb=None):
    column_batch_size = max(batch_size, 10000)
    column_values = []
    for col_idx in range(1, last_header_col + 1):
        col_letter = get_column_letter(col_idx)
        values = []
        start = 1
        while start <= row_count:
            end = min(start + column_batch_size - 1, row_count)
            expected = end - start + 1
            batch = _hq_get_values_range(
                base_url, origin, user_id, token,
                f"{sheet_id}!{col_letter}{start}:{col_letter}{end}",
            )
            if not batch:
                break
            for item in batch:
                values.append(item[0] if isinstance(item, list) and item else item)
            if len(batch) < expected:
                break
            start = end + 1
        column_values.append(values)
        if progress_cb:
            progress_cb(col_idx)

    all_rows = []
    for row_idx in range(row_count):
        all_rows.append([
            col[row_idx] if row_idx < len(col) else ""
            for col in column_values
        ])
    return _normalize_dense_rows(all_rows, last_header_col, header)


def _hq_read_sheet(base_url, origin, user_id, token,
                   sheet_id, row_count=200000, col_count=100,
                   batch_size=3000, progress_cb=None):
    """Read a sheet using fast range blocks, with dense rows and column fallback."""
    row_count = max(int(row_count or 1), 1)
    header_end_col = get_column_letter(max(int(col_count or 26), 26))
    header_rows = _hq_get_values_range(
        base_url, origin, user_id, token,
        f"{sheet_id}!A1:{header_end_col}1",
        timeout=30,
    )
    header = list(header_rows[0]) if header_rows else []
    last_header_col = 0
    for idx, value in enumerate(header, 1):
        if _cell_str(value):
            last_header_col = idx
    if not last_header_col:
        return []
    header = header[:last_header_col]
    end_col = get_column_letter(last_header_col)

    try:
        all_rows, start = [], 1
        while start <= row_count:
            end = min(start + batch_size - 1, row_count)
            expected = end - start + 1
            batch = _hq_get_values_range(
                base_url, origin, user_id, token,
                f"{sheet_id}!A{start}:{end_col}{end}",
                timeout=90,
            )
            if all_rows and batch:
                batch = batch[1:]
            if not batch:
                break
            all_rows.extend(batch)
            if progress_cb:
                progress_cb(len(all_rows))
            skip = 1 if start > 1 else 0
            if len(batch) < expected - skip:
                break
            start = end + 1
        if all_rows:
            return _normalize_dense_rows(all_rows, last_header_col, header)
    except Exception:
        pass

    return _hq_read_sheet_by_columns(
        base_url, origin, user_id, token, sheet_id,
        row_count, last_header_col, header,
        batch_size=batch_size, progress_cb=progress_cb,
    )


def _map_local_key_value(value, transform=''):
    text = _cell_str(value)
    if transform == 'manufacturer_alias' and text:
        match = lookup_manufacturer(text)
        if match:
            return _cell_str(match.get('canonical_name'))
    return text



def _match_source_priority(name):
    text = str(name or '')
    return 1 if ('\u5bf9\u5e94\u5173\u7cfb' in text or '\u5173\u7cfb\u5e93' in text) else 0


def _is_passive_recommendation_source(name):
    text = str(name or '').lower()
    return 'mlcc' in text or '\u7535\u963b' in text


_HQ_NUMBER_OUTPUT = "HQ\u6599\u53f7"
_HQ_NUMBER_ALIASES = (
    "HQ\u6599\u53f7",
    "HQ\u7f16\u7801",
    "\u7269\u6599\u7f16\u7801",
    "\u534e\u52e4\u6599\u53f7",
    "\u534e\u52e4\u7269\u6599\u53f7",
    "\u6599\u53f7",
    "HQ PN",
    "HQPN",
)


def _find_hq_number_col(headers, col_lookup):
    candidates = [col_lookup.get(_HQ_NUMBER_OUTPUT), _HQ_NUMBER_OUTPUT, *_HQ_NUMBER_ALIASES]
    for name in candidates:
        name = _cell_str(name)
        if name and name in headers:
            return name
    lowered = {_cell_str(h).lower(): h for h in headers}
    for name in candidates:
        key = _cell_str(name).lower()
        if key and key in lowered:
            return lowered[key]
    return ''


def _extract_hq_number(row_dict, values, all_fetch_cols, col_lookup):
    hq_index = all_fetch_cols.index(_HQ_NUMBER_OUTPUT) if _HQ_NUMBER_OUTPUT in all_fetch_cols else -1
    if 0 <= hq_index < len(values):
        value = _cell_str(values[hq_index])
        if value:
            return value
    for name in [col_lookup.get(_HQ_NUMBER_OUTPUT), *_HQ_NUMBER_ALIASES]:
        value = _cell_str(row_dict.get(_cell_str(name), ''))
        if value:
            return value
    return ''


def _build_hq_lookup(rows, headers, col_lookup):
    hq_col = _find_hq_number_col(headers, col_lookup)
    lookup = {}
    if not hq_col:
        return lookup, ''
    for row in rows:
        hq_no = _cell_str(row.get(hq_col, ''))
        if hq_no:
            lookup.setdefault(hq_no, []).append(row)
    return lookup, hq_col


def _fetch_values_for_row(row_dict, all_fetch_cols, col_lookup):
    return [row_dict.get(col_lookup.get(col_name, col_name), '') for col_name in all_fetch_cols]


def _preferred_sort_score(value):
    text = _cell_str(value)
    if not text:
        return (0, '')
    lowered = text.lower()
    negative_markers = ('\u975e\u4f18\u9009', '\u4e0d\u4f18\u9009', '\u975e\u63a8\u8350', 'not preferred', '\u6dd8\u6c70')
    if any(marker in lowered for marker in negative_markers):
        return (-1, text)
    try:
        return (float(text), text)
    except ValueError:
        pass
    number_match = re.search(r'\d+(?:\.\d+)?', text)
    if number_match:
        try:
            return (float(number_match.group(0)), text)
        except ValueError:
            pass
    if '\u4f18\u9009' in text or 'preferred' in lowered:
        return (9, text)
    if text.upper() in ('PI', 'P'):
        return (8, text)
    if '\u9650\u9009' in text:
        return (2, text)
    return (1, text)


def _source_label(item):
    source = item.get('source', '')
    match_type = item.get('match_type', '')
    if match_type == '\u4f18\u9009\u53ef\u66ff\u4ee3\u63a8\u8350':
        return f"{source}\uff08{match_type}\uff09"
    return source


def _merge_match_groups(items):
    grouped = {}
    for item in items:
        group_key = tuple(item['values'])
        grouped_item = grouped.setdefault(group_key, {
            'values': item['values'],
            'sources': [],
            'rows': [],
            'tables': [],
        })
        source = _source_label(item)
        if source not in grouped_item['sources']:
            grouped_item['sources'].append(source)
        grouped_item['rows'].append(item.get('row') or {})
        grouped_item['tables'].append(item.get('table') or {})
    return list(grouped.values())


def _do_match_multi(local_ws, local_header_row, prepared_tables, all_fetch_cols, out_file):
    max_local_col = local_ws.max_column
    local_header = [local_ws.cell(row=local_header_row, column=ci).value
                    for ci in range(1, max_local_col + 1)]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "\u5339\u914d\u7ed3\u679c"
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")
    hq_fill = PatternFill("solid", fgColor="FFFF00")
    src_fill = PatternFill("solid", fgColor="BDD7EE")

    out_hdrs = list(local_header) + all_fetch_cols + ["\u6765\u6e90\u8868\u683c"]
    for ci, h in enumerate(out_hdrs, 1):
        c = ws_out.cell(row=1, column=ci, value=h or "")
        c.font = Font(bold=True)
        c.fill = (PatternFill("solid", fgColor="FFC000") if ci > max_local_col else hdr_fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    ws_out.row_dimensions[1].height = 22
    for ci in range(1, max_local_col + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 18
    for ci in range(max_local_col + 1, len(out_hdrs) + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 24

    dr = 2
    total = matched = unmatched = 0

    for ri in range(local_header_row + 1, local_ws.max_row + 1):
        row_vals = [local_ws.cell(row=ri, column=ci).value
                    for ci in range(1, max_local_col + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        total += 1

        candidate_matches = []
        for pt in prepared_tables:
            transforms = pt.get("local_key_transforms", [])
            key = tuple(
                _map_local_key_value(
                    row_vals[lc - 1],
                    transforms[i] if i < len(transforms) else '',
                )
                for i, lc in enumerate(pt["local_key_cols"])
            )
            if not key or not all(k for k in key):
                continue
            matches = pt["lookup"].get(key, [])
            if not matches:
                continue
            for mdict in matches:
                candidate_matches.append({
                    "values": _fetch_values_for_row(mdict, all_fetch_cols, pt.get("col_lookup", {})),
                    "source": pt["name"],
                    "priority": pt.get("source_priority", 0),
                    "match_type": "\u4e25\u683c\u5339\u914d",
                    "include_with_relation": pt.get("include_with_relation", False),
                    "table": pt,
                    "row": mdict,
                })

        if candidate_matches:
            include_all_strict_sources = any(item.get("include_with_relation") for item in candidate_matches)
            if include_all_strict_sources:
                seen_direct = {(id(item.get("table")), id(item.get("row"))) for item in candidate_matches}
                relation_hq_numbers = []
                for item in candidate_matches:
                    if item.get("priority", 0) <= 0:
                        continue
                    hq_no = _extract_hq_number(
                        item.get("row") or {},
                        item.get("values") or [],
                        all_fetch_cols,
                        (item.get("table") or {}).get("col_lookup", {}),
                    )
                    if hq_no:
                        relation_hq_numbers.append(hq_no)
                for hq_no in relation_hq_numbers:
                    for pt in prepared_tables:
                        if pt.get("source_priority", 0) > 0:
                            continue
                        for mdict in (pt.get("hq_lookup") or {}).get(hq_no, []):
                            direct_key = (id(pt), id(mdict))
                            if direct_key in seen_direct:
                                continue
                            seen_direct.add(direct_key)
                            candidate_matches.append({
                                "values": _fetch_values_for_row(mdict, all_fetch_cols, pt.get("col_lookup", {})),
                                "source": pt["name"],
                                "priority": pt.get("source_priority", 0),
                                "match_type": "\u4e25\u683c\u5339\u914d",
                                "include_with_relation": True,
                                "table": pt,
                                "row": mdict,
                            })

        output_groups = []
        if candidate_matches:
            max_priority = max(item["priority"] for item in candidate_matches)
            include_all_strict_sources = any(item.get("include_with_relation") for item in candidate_matches)
            selected_matches = candidate_matches if include_all_strict_sources else [
                item for item in candidate_matches if item["priority"] == max_priority
            ]
            strict_groups = _merge_match_groups(selected_matches)
            output_groups.extend(strict_groups)

            seen_values = {tuple(group["values"]) for group in output_groups}
            recommendation_items = []
            for group in strict_groups:
                for pt, strict_row in zip(group.get("tables") or [], group.get("rows") or []):
                    if not pt.get("passive_recommendation"):
                        continue
                    desc_col = pt.get("description_col")
                    pref_col = pt.get("preferred_col")
                    if not desc_col or not pref_col:
                        continue
                    desc_value = _cell_str(strict_row.get(desc_col, ""))
                    if not desc_value:
                        continue
                    for mdict in pt.get("rows", []):
                        if _cell_str(mdict.get(desc_col, "")) != desc_value:
                            continue
                        values = _fetch_values_for_row(mdict, all_fetch_cols, pt.get("col_lookup", {}))
                        if tuple(values) in seen_values:
                            continue
                        seen_values.add(tuple(values))
                        recommendation_items.append({
                            "values": values,
                            "source": pt["name"],
                            "priority": pt.get("source_priority", 0),
                            "match_type": "\u4f18\u9009\u53ef\u66ff\u4ee3\u63a8\u8350",
                            "sort_score": _preferred_sort_score(mdict.get(pref_col, "")),
                            "row": mdict,
                            "table": pt,
                        })
            recommendation_items.sort(
                key=lambda item: (
                    item.get("sort_score", (0, ""))[0],
                    item.get("sort_score", (0, ""))[1],
                ),
                reverse=True,
            )
            output_groups.extend(_merge_match_groups(recommendation_items))

        if output_groups:
            first = True
            for grouped in output_groups:
                for ci, val in enumerate(row_vals, 1):
                    c = ws_out.cell(row=dr, column=ci, value=val if first else None)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = bdr
                for j, value in enumerate(grouped["values"]):
                    c = ws_out.cell(row=dr, column=max_local_col + j + 1, value=value)
                    c.fill = hq_fill
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = bdr
                src_col = max_local_col + len(all_fetch_cols) + 1
                c = ws_out.cell(row=dr, column=src_col, value="\uff1b".join(grouped["sources"]))
                c.fill = src_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = bdr
                first = False
                dr += 1
            matched += 1
        else:
            for ci, val in enumerate(row_vals, 1):
                c = ws_out.cell(row=dr, column=ci, value=val)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = bdr
            for j in range(len(all_fetch_cols)):
                ws_out.cell(row=dr, column=max_local_col + j + 1).border = bdr
            c = ws_out.cell(row=dr, column=max_local_col + len(all_fetch_cols) + 1, value="\u672a\u5339\u914d")
            c.border = bdr
            unmatched += 1
            dr += 1

    wb_out.save(out_file)
    return total, matched, unmatched


# ── 路由 ─────────────────────────────────────────────────────

@feishu_bp.route('/api/feishu/load', methods=['POST'])
@track_tool_activity('飞书Sheet缓存')
def api_feishu_load():
    """拉取单个 Sheet 全部数据并缓存到服务端"""
    data = request.get_json(silent=True) or {}
    try:
        base_url = _resolve_feishu_base_url(data.get('base_url'))
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    origin   = data.get('origin', '')
    user_id  = data.get('user_id', '')
    token    = data.get('token', '').strip()
    sheet_id = data.get('sheet_id', '').strip()
    if not token or not sheet_id:
        return jsonify({'success': False, 'error': '请填写 Token 和 Sheet ID'})
    try:
        sheets_meta = _hq_get_sheets(base_url, origin, user_id, token)
        target = next((s for s in sheets_meta if s['sheetId'] == sheet_id), None)
        if not target:
            return jsonify({'success': False, 'error': 'Sheet 不存在或已被删除'})
        row_count = target.get('rowCount', 200000) or 200000
        col_count = target.get('columnCount', 100) or 100
        rows = _hq_read_sheet(base_url, origin, user_id, token,
                              sheet_id, row_count=row_count, col_count=col_count)
        if not rows:
            return jsonify({'success': False, 'error': '读取到 0 行数据'})
        row_count_at_cache = target.get('rowCount', 0)
        key, data_rows, headers = _write_cache(token, sheet_id, rows, row_count_at_cache)
        return jsonify({
            'success': True,
            'cache_key': key,
            'row_count': data_rows,
            'headers': headers,
            'fetched_at': time.time(),
            'row_count_at_cache': row_count_at_cache,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@feishu_bp.route('/api/feishu/cache/clear', methods=['POST'])
def api_feishu_cache_clear():
    data = request.get_json(silent=True) or {}
    token = str(data.get('token') or '').strip()
    sheet_id = str(data.get('sheet_id') or '').strip()
    if not token or not sheet_id:
        return jsonify({'success': False, 'error': '\u8bf7\u63d0\u4f9b Token \u548c Sheet ID'})
    try:
        key, existed = _delete_cache(token, sheet_id)
        return jsonify({'success': True, 'cache_key': key, 'deleted': existed})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@feishu_bp.route('/api/feishu/sheets', methods=['POST'])
def api_feishu_sheets():
    """获取飞书表格的 Sheet 列表，同时读取每个 Sheet 的第一行表头"""
    data = request.get_json(silent=True) or {}
    try:
        base_url = _resolve_feishu_base_url(data.get('base_url'))
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    origin = data.get('origin', '')
    user_id = data.get('user_id', '')
    token = data.get('token', '').strip()
    if not token:
        return jsonify({'success': False, 'error': '请填写 Token'})
    try:
        sheets = _hq_get_sheets(base_url, origin, user_id, token)
        sheet_headers = {}
        for s in sheets:
            sid = s['sheetId']
            col_count = s.get('columnCount', 100) or 100
            try:
                rows = _hq_read_sheet(base_url, origin, user_id, token,
                                      sid, row_count=2, col_count=col_count, batch_size=20)
                sheet_headers[sid] = [_cell_str(v) for v in (rows[0] if rows else [])]
            except Exception:
                sheet_headers[sid] = []
        return jsonify({'success': True, 'sheets': sheets, 'sheet_headers': sheet_headers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@feishu_bp.route('/api/feishu/match', methods=['POST'])
@track_tool_activity('飞书优选库匹配')
def api_feishu_match():
    """执行飞书多表格匹配（per-sheet 配置格式）"""
    local_file = request.files.get('file')
    if not local_file:
        return jsonify({'success': False, 'error': '请上传本地 Excel 文件'})

    config_str = request.form.get('config', '{}')
    try:
        config = json.loads(config_str)
    except Exception:
        return jsonify({'success': False, 'error': 'config 参数格式错误'})

    try:
        base_url = _resolve_feishu_base_url(config.get('base_url'))
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    origin     = config.get('origin', '')
    user_id    = config.get('user_id', '')
    sheet_name = config.get('sheet_name', '')
    header_row = _to_int(config.get('header_row', 1), 1)
    if header_row is None:
        return jsonify({'success': False, 'error': '表头行必须是大于等于 1 的数字'})
    tables_cfg = config.get('tables', [])
    include_preferred_with_relation = bool(config.get('include_preferred_with_relation', False))

    uid = str(uuid.uuid4())[:8]
    try:
        local_path = _save_uploaded_excel(local_file, "fs_local", uid)
        wb_local = _open_workbook(local_path, data_only=True)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    sheets = wb_local.sheetnames
    if not sheet_name or sheet_name not in sheets:
        sheet_name = sheets[0]
    ws_local = wb_local[sheet_name]
    local_header = [ws_local.cell(row=header_row, column=ci).value
                    for ci in range(1, ws_local.max_column + 1)]
    local_header_strs = [_cell_str(h) for h in local_header]

    for tcfg in tables_cfg:
        for scfg in tcfg.get('sheets', []):
            if not scfg.get('enabled', False):
                continue
            local_key_names = [_cell_str(k) for k in scfg.get('local_key_names', [])]
            if len(local_key_names) < 1 or not local_key_names[0]:
                wb_local.close()
                return jsonify({'success': False, 'error': '\u672a\u9009\u62e9\u578b\u53f7\uff0c\u8bf7\u5148\u9009\u62e9\u672c\u5730 Sheet \u7684\u578b\u53f7\u6620\u5c04'})
            if len(local_key_names) < 2 or not local_key_names[1]:
                wb_local.close()
                return jsonify({'success': False, 'error': '\u672a\u9009\u62e9\u5382\u5546\uff0c\u8bf7\u5148\u9009\u62e9\u672c\u5730 Sheet \u7684\u5382\u5546\u6620\u5c04'})
            if local_key_names[0] not in local_header_strs:
                wb_local.close()
                return jsonify({'success': False, 'error': f'\u672c\u5730\u578b\u53f7\u6620\u5c04\u5217\u300c{local_key_names[0]}\u300d\u4e0d\u5b58\u5728\uff0c\u8bf7\u91cd\u65b0\u9009\u62e9'})
            if local_key_names[1] not in local_header_strs:
                wb_local.close()
                return jsonify({'success': False, 'error': f'\u672c\u5730\u5382\u5546\u6620\u5c04\u5217\u300c{local_key_names[1]}\u300d\u4e0d\u5b58\u5728\uff0c\u8bf7\u91cd\u65b0\u9009\u62e9'})

    logs = []
    prepared_tables = []
    all_fetch_cols_ordered = []
    seen_fetch_cols = set()

    for tcfg in tables_cfg:
        table_name = tcfg.get('name', '')
        token = tcfg.get('token', '').strip()
        if not token:
            continue

        for scfg in tcfg.get('sheets', []):
            if not scfg.get('enabled', False):
                continue
            sid        = scfg.get('sheet_id', '').strip()
            sname      = scfg.get('sheet_name', sid)
            full_name  = f"{table_name} - {sname}"
            local_keys = [k for k in scfg.get('local_key_names', []) if k]
            feishu_keys= [k for k in scfg.get('feishu_key_names', []) if k]
            local_key_transforms = list(scfg.get('local_key_transforms', []))
            fetch_cols = [c for c in scfg.get('fetch_col_names', []) if c]
            if len(local_key_transforms) < len(local_keys):
                local_key_transforms += [''] * (len(local_keys) - len(local_key_transforms))
            elif len(local_key_transforms) > len(local_keys):
                local_key_transforms = local_key_transforms[:len(local_keys)]

            if not sid:
                logs.append(f"[{full_name}] 跳过：Sheet ID 为空")
                continue
            if not local_keys or not feishu_keys:
                logs.append(f"[{full_name}] 跳过：匹配键未配置")
                continue
            if len(local_keys) != len(feishu_keys):
                logs.append(f"[{full_name}] 跳过：本地键与飞书键数量不匹配")
                continue

            # Map local key names → column indices
            local_key_cols = []
            for kn in local_keys:
                try:
                    local_key_cols.append(local_header_strs.index(kn) + 1)
                except ValueError:
                    logs.append(f"[{full_name}] 跳过：本地列「{kn}」不存在")
                    local_key_cols = []
                    break
            if not local_key_cols:
                continue

            # Prefer server-side cache
            cache_key = scfg.get('cache_key', '')
            rows = None
            if cache_key:
                cached = _read_cache(cache_key)
                if cached:
                    rows = cached['rows']
                    logs.append(f"[{full_name}] 使用缓存（{len(rows)-1} 行，"
                                f"{time.strftime('%Y-%m-%d %H:%M', time.localtime(cached['fetched_at']))}）")
                else:
                    logs.append(f"[{full_name}] 缓存已失效，实时拉取...")
            if rows is None:
                try:
                    logs.append(f"[{full_name}] 正在读取 Sheet 数据...")
                    sheets_meta = _hq_get_sheets(base_url, origin, user_id, token)
                    target = next((s for s in sheets_meta if s['sheetId'] == sid), None)
                    if not target:
                        logs.append(f"[{full_name}] 跳过：Sheet 已失效，请重新配置")
                        continue
                    rows = _hq_read_sheet(base_url, origin, user_id, token, sid,
                                         row_count=target.get('rowCount', 200000) or 200000,
                                         col_count=target.get('columnCount', 100) or 100)
                    if not rows:
                        logs.append(f"[{full_name}] 读取到 0 行，跳过")
                        continue
                    logs.append(f"[{full_name}] 读取 {len(rows)-1} 行")
                except Exception as e:
                    logs.append(f"[{full_name}] 读取失败：{e}")
                    continue

            fs_headers = [_cell_str(v) for v in rows[0]]
            logs.append(f"[{full_name}] 共 {len(fs_headers)} 列")

            bad_keys = [k for k in feishu_keys if k not in fs_headers]
            if bad_keys:
                logs.append(f"[{full_name}] 跳过：飞书列 {bad_keys} 不存在")
                continue

            feishu_key_indices = [fs_headers.index(k) for k in feishu_keys]
            lookup = {}
            for row in rows[1:]:
                padded = list(row) + [""] * max(0, len(fs_headers) - len(row))
                key = tuple(_cell_str(padded[i]) if i < len(padded) else "" for i in feishu_key_indices)
                if not key or not all(key):
                    continue
                row_dict = {fs_headers[i]: _cell_str(padded[i]) if i < len(padded) else ""
                            for i in range(len(fs_headers))}
                lookup.setdefault(key, []).append(row_dict)

            # Build col_lookup: output_name → actual feishu column name
            # Priority: fetch_col_map (global mapped with aliases) > fetch_col_names (direct)
            col_lookup = {}
            fetch_col_map_cfg = scfg.get('fetch_col_map', [])  # [{output, alias}]
            for fm in fetch_col_map_cfg:
                out_name = (fm.get('output') or '').strip()
                alias = (fm.get('alias') or '').strip() or out_name
                if out_name:
                    col_lookup[out_name] = alias
                    if out_name not in seen_fetch_cols:
                        all_fetch_cols_ordered.append(out_name)
                        seen_fetch_cols.add(out_name)
            # Old per-sheet direct columns (identity mapping, output == lookup)
            for cn in fetch_cols:
                if cn and cn not in col_lookup:
                    col_lookup[cn] = cn
                    if cn not in seen_fetch_cols:
                        all_fetch_cols_ordered.append(cn)
                        seen_fetch_cols.add(cn)

            table_rows = []
            for row in rows[1:]:
                padded = list(row) + [""] * max(0, len(fs_headers) - len(row))
                table_rows.append({
                    fs_headers[i]: _cell_str(padded[i]) if i < len(padded) else ""
                    for i in range(len(fs_headers))
                })
            hq_lookup, hq_lookup_col = _build_hq_lookup(table_rows, fs_headers, col_lookup)
            enable_recommendations = bool(scfg.get('enable_recommendations', False))

            prepared_tables.append({
                "name": full_name,
                "local_key_cols": local_key_cols,
                "local_key_transforms": local_key_transforms,
                "source_priority": _match_source_priority(full_name),
                "include_with_relation": include_preferred_with_relation,
                "passive_recommendation": enable_recommendations,
                "description_col": col_lookup.get("HQ\u63cf\u8ff0"),
                "preferred_col": col_lookup.get("\u4f18\u9009\u7b49\u7ea7"),
                "lookup": lookup,
                "hq_lookup": hq_lookup,
                "hq_lookup_col": hq_lookup_col,
                "rows": table_rows,
                "col_lookup": col_lookup,
            })
            logs.append(f"[{full_name}] 就绪，{len(lookup)} 个唯一键")

    if not prepared_tables:
        wb_local.close()
        return jsonify({'success': False, 'error': '没有可用的 Sheet，请检查配置', 'logs': logs})

    out_name = f"飞书匹配结果_{uid}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_name)

    try:
        total, matched, unmatched = _do_match_multi(
            ws_local, header_row, prepared_tables, all_fetch_cols_ordered, out_path)
        wb_local.close()
        logs.append(f"完成：共 {total} 行，匹配 {matched}，未匹配 {unmatched}")
        return jsonify({
            'success': True,
            'download': f'/download/{out_name}',
            'total': total,
            'matched': matched,
            'unmatched': unmatched,
            'logs': logs,
        })
    except Exception as e:
        wb_local.close()
        logs.append(f"匹配出错：{e}")
        return jsonify({'success': False, 'error': str(e), 'logs': logs})



@feishu_bp.route('/api/feishu/local_sheets', methods=['POST'])
def api_feishu_local_sheets():
    """Return local Excel sheets and headers; reuse uploaded preview files by uid."""
    file = request.files.get('file')
    try:
        uid, path = _save_or_reuse_uploaded_excel(file, "fs_pre", request.form.get('uid', ''))
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

    header_row = _to_int(request.form.get('header_row', 1), 1)
    if header_row is None:
        return jsonify({'success': False, 'error': 'Header row must be a number greater than or equal to 1'})

    try:
        wb = _open_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        sheet_name = request.form.get('sheet_name', '')
        if not sheet_name or sheet_name not in sheets:
            sheet_name = sheets[0] if sheets else ''
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        row_iter = ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        header_values = next(row_iter, [])
        headers = [_cell_str(v) for v in header_values]
        headers = [h for h in headers if h]
        wb.close()
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

    return jsonify({'success': True, 'sheets': sheets, 'current_sheet': sheet_name,
                    'headers': headers, 'uid': uid})

@feishu_bp.route('/api/feishu/pref_rate', methods=['POST'])
@track_tool_activity('查询BOM优选率')
def api_pref_rate():
    """查询BOM优选率：按 HQ料号 在所有优选库缓存中查找优选等级"""
    local_file = request.files.get('file')
    if not local_file:
        return jsonify({'success': False, 'error': '未上传文件'})

    config_str = request.form.get('config', '{}')
    try:
        config = json.loads(config_str)
    except Exception:
        return jsonify({'success': False, 'error': 'config 参数格式错误'})

    header_row   = _to_int(config.get('header_row', 1), 1)
    if header_row is None:
        return jsonify({'success': False, 'error': '表头行必须是大于等于 1 的数字'})
    sheet_name   = config.get('sheet_name', '')
    local_key_col = config.get('local_key_col', '')
    tables_cfg   = config.get('tables', [])   # [{name, sheets:[{sid,name,cache_key,fetch_col_aliases}]}]

    uid = str(uuid.uuid4())[:8]
    try:
        local_path = _save_uploaded_excel(local_file, "pref", uid)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)})

    try:
        try:
            wb = _open_workbook(local_path, data_only=True)
        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)})
        sheets = wb.sheetnames
        if not sheet_name or sheet_name not in sheets:
            sheet_name = sheets[0]
        ws = wb[sheet_name]

        headers = [_cell_str(ws.cell(row=header_row, column=ci).value)
                   for ci in range(1, ws.max_column + 1)]
        if local_key_col not in headers:
            return jsonify({'success': False,
                            'error': f'列 "{local_key_col}" 不存在，请检查表头行设置'})
        key_col_idx = headers.index(local_key_col)

        # Build combined lookup: hq_no → {pref, source}
        combined = {}
        for tcfg in tables_cfg:
            tname = tcfg.get('name', '')
            for scfg in tcfg.get('sheets', []):
                cache_key = scfg.get('cache_key', '')
                aliases   = scfg.get('fetch_col_aliases', {})
                if not cache_key:
                    continue
                cached = _read_cache(cache_key)
                if not cached or not cached.get('rows'):
                    continue
                rows = cached['rows']
                fs_hdrs = [_cell_str(v) for v in rows[0]]
                hq_col   = aliases.get('HQ料号', 'HQ料号')
                pref_col = aliases.get('优选等级', '优选等级')
                if hq_col not in fs_hdrs or pref_col not in fs_hdrs:
                    continue
                hi = fs_hdrs.index(hq_col)
                pi = fs_hdrs.index(pref_col)
                for row in rows[1:]:
                    padded = list(row) + [''] * max(0, max(hi, pi) + 1 - len(row))
                    hv = _cell_str(padded[hi]) if hi < len(padded) else ''
                    pv = _cell_str(padded[pi]) if pi < len(padded) else ''
                    if hv and hv not in combined:
                        combined[hv] = {'pref': pv, 'source': tname}

        # Write output
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        pref_fill    = PatternFill('solid', fgColor='E8F5E9')   # 绿：优选
        nonpref_fill = PatternFill('solid', fgColor='FFF9C4')   # 黄：匹配到但非优选
        nomatch_fill = PatternFill('solid', fgColor='F5F5F5')   # 灰：未匹配
        hdr_font = Font(bold=True)
        bdr = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'))
        center = Alignment(horizontal='center', vertical='center')
        left   = Alignment(horizontal='left',   vertical='center')

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = '优选率查询'

        out_headers = headers + ['优选等级', '来源']
        for ci, h in enumerate(out_headers, 1):
            c = ws_out.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.border = bdr; c.alignment = center

        total = matched = preferred = dr = 0
        for ri in range(header_row + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=ri, column=ci).value
                        for ci in range(1, ws.max_column + 1)]
            if not any(v is not None and str(v).strip() for v in row_vals):
                continue
            total += 1
            dr += 1
            kv = _cell_str(row_vals[key_col_idx]) if key_col_idx < len(row_vals) else ''
            m  = combined.get(kv) if kv else None
            if m:
                matched += 1
                is_pref = _is_preferred_level(m['pref'])
                if is_pref:
                    preferred += 1
                fill = pref_fill if is_pref else nonpref_fill
            else:
                fill = nomatch_fill
            out_row = list(row_vals) + [m['pref'] if m else '', m['source'] if m else '']
            for ci, v in enumerate(out_row, 1):
                c = ws_out.cell(row=dr + 1, column=ci, value=v)
                c.fill = fill; c.border = bdr
                c.alignment = center if ci > len(headers) else left

        wb.close()
        out_name = f'pref_rate_{uid}.xlsx'
        wb_out.save(os.path.join(OUTPUT_DIR, out_name))

        return jsonify({
            'success': True,
            'download': f'/download/{out_name}',
            'total': total,
            'matched': matched,
            'unmatched': total - matched,
            'preferred': preferred,
            'non_preferred': matched - preferred,
            # 优选率 = 优选料 / 已匹配料（未匹配不参与）
            'rate': f'{preferred / matched * 100:.1f}%' if matched else 'N/A',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})



