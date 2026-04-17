# -*- coding: utf-8 -*-
"""
PSTX 原理图分析工具 v1.1
解析 Cadence Packager-XL 导出的 pstxprt.dat / pstxnet.dat

功能：BOM 管理 / 网络拓扑 / DRC / 电容降额 / 元件查询 / Excel 导出

依赖：pip install openpyxl
运行：python pstx_analyzer.py
"""

import sys
import subprocess

try:
    import openpyxl
except ImportError:
    print("未检测到 openpyxl，正在自动安装...")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

import os
import re
import threading
import tkinter as tk
from collections import Counter, defaultdict
from tkinter import ttk, filedialog, messagebox, scrolledtext
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════
# 一、PST 文件解析
# ══════════════════════════════════════════════════════════

def _join_continuations(text: str) -> str:
    lines = text.split('\n')
    result, buf = [], ''
    for line in lines:
        stripped = line.rstrip()
        if stripped.endswith('~'):
            buf += stripped[:-1]
        else:
            buf += line
            result.append(buf)
            buf = ''
    if buf:
        result.append(buf)
    return '\n'.join(result)


def _extract_attrs(text: str) -> Dict[str, str]:
    attrs = {}
    for m in re.finditer(r"\b([A-Z][A-Z0-9_]*)\s*=\s*'([^']*)'", text):
        key, val = m.group(1), m.group(2)
        if key not in attrs:
            attrs[key] = val
    return attrs


def _extract_page(drawing: str) -> str:
    m = re.search(r'(PAGE\d+)', drawing, re.IGNORECASE)
    return m.group(1).upper() if m else ''


def _get_comp_type(refdes: str, part_name: str) -> str:
    pn = part_name.lower()
    type_rules = [
        (['cap_pol'],                           'CAP_POL'),
        (['cap_hdl', 'cap_'],                   'CAP'),
        (['res_hdl', 'res_'],                   'RES'),
        (['ind_hdl', 'ind_', 'ferrite', 'fer_hdl', 'fb_hdl'], 'IND'),
        (['osc_', 'crystal', 'xtal'],           'XTAL'),
        (['conn_', 'connector'],                'CONN'),
        (['led_'],                              'LED'),
        (['diode', '_d_hdl'],                   'DIODE'),
        (['mosfet', 'mos_', 'nmos', 'pmos', 'nfet', 'pfet'], 'FET'),
        (['bjt', 'transistor', 'npn', 'pnp'],  'BJT'),
        (['fuse'],                              'FUSE'),
        (['sw_hdl', 'switch'],                  'SWITCH'),
        (['testpoint', 'test_point', 'tp_hdl'], 'TESTPOINT'),
        (['transformer', 'xfmr'],              'TRANSFORMER'),
    ]
    for keywords, ctype in type_rules:
        if any(k in pn for k in keywords):
            return ctype
    prefix = (re.match(r'[A-Za-z]+', refdes) or re.match(r'', '')).group(0).upper()
    prefix_map = {
        'C': 'CAP', 'PC': 'CAP', 'R': 'RES', 'L': 'IND', 'FB': 'IND',
        'U': 'IC', 'J': 'CONN', 'P': 'CONN', 'CN': 'CONN', 'Q': 'FET',
        'D': 'DIODE', 'LED': 'LED', 'Y': 'XTAL', 'F': 'FUSE',
        'SW': 'SWITCH', 'TP': 'TESTPOINT', 'T': 'TRANSFORMER',
    }
    return prefix_map.get(prefix, 'IC')


def parse_pstxprt(content: str) -> Dict[str, dict]:
    text = _join_continuations(content)
    components = {}
    for block in re.split(r'\nPART_NAME\n', text)[1:]:
        m = re.match(r"(\S+)\s+'([^']+)'", block.split('\n')[0].strip())
        if not m:
            continue
        refdes, part_name = m.group(1), m.group(2)
        attrs = _extract_attrs(block)
        components[refdes] = {
            'refdes':     refdes,
            'part_name':  part_name,
            'hq_code':    attrs.get('HQ_CODE', ''),
            'value':      attrs.get('VALUE', ''),
            'package':    attrs.get('PACKAGE', ''),
            'material':   attrs.get('MATERIAL', ''),
            'tolerance':  attrs.get('TOLERANCE', ''),
            'voltage':    attrs.get('VOLTAGE', ''),
            'current':    attrs.get('CURRENT', ''),
            'power':      attrs.get('POWER', ''),
            'bom_option': attrs.get('BOM_OPTION', ''),
            'bom_cost':   attrs.get('BOM_COST', ''),
            'room':       attrs.get('ROOM', ''),
            'drawing':    attrs.get('DRAWING', ''),
            'page':       _extract_page(attrs.get('DRAWING', '')),
            'comp_type':  _get_comp_type(refdes, part_name),
        }
    return components


def parse_pstxnet(content: str) -> Dict[str, List[dict]]:
    text = _join_continuations(content)
    nets = {}
    node_re     = re.compile(r'NODE_NAME\s+(\S+)\s+(\S+)')
    pin_name_re = re.compile(r"'([^']+)'\s*:")
    for block in re.split(r'\nNET_NAME\n', text)[1:]:
        m = re.search(r"'([^']+)'", block)
        if not m:
            continue
        net_name = m.group(1)
        nodes = []
        for nm in node_re.finditer(block):
            after    = block[nm.end(): nm.end() + 200]
            pn_match = pin_name_re.search(after)
            nodes.append({
                'refdes':   nm.group(1),
                'pin':      nm.group(2),
                'pin_name': pn_match.group(1) if pn_match else nm.group(2),
            })
        if nodes:
            nets[net_name] = nodes
    return nets


def parse_all(prt_content: str, net_content: str):
    components = parse_pstxprt(prt_content)
    nets       = parse_pstxnet(net_content)
    comp_nets: Dict[str, Dict[str, str]] = {}
    for net_name, nodes in nets.items():
        for node in nodes:
            rd = node['refdes']
            if rd not in comp_nets:
                comp_nets[rd] = {}
            comp_nets[rd][node['pin']] = net_name
    for refdes, comp in components.items():
        comp['nets'] = comp_nets.get(refdes, {})
    return components, nets, comp_nets


# ══════════════════════════════════════════════════════════
# 二、BOM 分析
# ══════════════════════════════════════════════════════════

COMP_TYPE_CN = {
    'CAP': '电容', 'CAP_POL': '电解/钽电容', 'RES': '电阻',
    'IND': '电感/磁珠', 'IC': 'IC 芯片', 'CONN': '连接器',
    'DIODE': '二极管', 'LED': 'LED', 'FET': 'MOS/FET',
    'BJT': '三极管', 'XTAL': '晶振', 'FUSE': '保险丝',
    'SWITCH': '开关', 'TESTPOINT': '测试点', 'TRANSFORMER': '变压器',
}
_TYPE_ORDER = list(COMP_TYPE_CN.keys())


def build_bom(components: Dict):
    detail_normal, detail_depop = [], []
    for comp in components.values():
        ctype = comp.get('comp_type', '')
        row = {
            '位号':          comp['refdes'],
            '料号':          comp.get('hq_code', ''),
            '描述':          comp.get('part_name', ''),
            '值':            comp.get('value', ''),
            '封装':          comp.get('package', ''),
            '耐压/额定电压': comp.get('voltage', ''),
            '额定功率':      comp.get('power', ''),
            '精度':          comp.get('tolerance', ''),
            '材质':          comp.get('material', ''),
            '类型':          COMP_TYPE_CN.get(ctype, ctype),
            '_ctype':        ctype,
            '页面':          comp.get('page', ''),
            'ROOM':          comp.get('room', ''),
        }
        (detail_depop if comp.get('bom_option') == 'DEPOP' else detail_normal).append(row)

    def _merge(detail):
        if not detail:
            return []
        groups = {}
        for row in detail:
            key = row['料号'] or row['描述']
            if key not in groups:
                groups[key] = {
                    '料号': row['料号'], '位号列表': [], '数量': 0,
                    '描述': row['描述'], '值': row['值'], '封装': row['封装'],
                    '耐压': row['耐压/额定电压'], '额定功率': row['额定功率'],
                    '精度': row['精度'], '材质': row['材质'],
                    '类型': row['类型'], '_ctype': row['_ctype'],
                }
            groups[key]['位号列表'].append(row['位号'])
            groups[key]['数量'] += 1
        merged = list(groups.values())
        merged.sort(key=lambda r: (
            _TYPE_ORDER.index(r['_ctype']) if r['_ctype'] in _TYPE_ORDER else 99, r['料号']))
        for i, r in enumerate(merged, 1):
            r['序号'] = i
            r['位号列表'] = ', '.join(sorted(r['位号列表']))
            del r['_ctype']
        return merged

    def _clean(rows):
        return [{k: v for k, v in r.items() if k != '_ctype'} for r in rows]

    return _clean(detail_normal), _clean(detail_depop), _merge(detail_normal), _merge(detail_depop)


# ══════════════════════════════════════════════════════════
# 三、网络分析
# ══════════════════════════════════════════════════════════

def analyze_networks(nets: Dict, components: Dict) -> dict:
    single_node = {k: v for k, v in nets.items() if len(v) == 1}
    gnd_nets    = {k: v for k, v in nets.items()
                   if re.search(r'GND|AGND|SGND|PGND|DGND', k, re.I)}
    power_nets  = {k: v for k, v in nets.items()
                   if re.search(r'^P\d|^[0-9]+V|VCC|VDD|VBAT|VCORE|VCCIO|PVDD|AVDD|DVDD', k, re.I)
                   and k not in gnd_nets}
    diff_pairs: Dict[str, dict] = {}
    for net_name in nets:
        for sp, sn in [('_P','_N'),('_DP','_DN'),('.P','.N'),
                       ('_TXPLUS','_TXMINUS'),('_RXPLUS','_RXMINUS')]:
            if net_name.endswith(sp):
                base, cp = net_name[:-len(sp)], net_name[:-len(sp)] + sn
                if cp in nets:
                    diff_pairs[base] = {'P': net_name, 'N': cp}
            elif net_name.endswith(sn):
                base, cp = net_name[:-len(sn)], net_name[:-len(sn)] + sp
                if cp in nets and base not in diff_pairs:
                    diff_pairs[base] = {'P': cp, 'N': net_name}
    page_counter: Counter = Counter()
    for comp in components.values():
        page_counter[comp.get('page', 'UNKNOWN')] += 1
    return {
        'total': len(nets), 'single_node': single_node,
        'gnd_nets': gnd_nets, 'power_nets': power_nets,
        'diff_pairs': diff_pairs, 'page_counter': page_counter,
    }


# ══════════════════════════════════════════════════════════
# 四、DRC 设计检查
# ══════════════════════════════════════════════════════════

_VALID_BOM_OPTIONS = {'', 'DEPOP', 'OPTION', 'MAIN_PLD', 'MAIN', 'ALT', 'DNP'}
_FUZZY_KEYWORDS    = ['DEPOP', 'OPTION']


def _edit_distance(a: str, b: str) -> int:
    a, b = a.upper(), b.upper()
    if a == b: return 0
    if not a:  return len(b)
    if not b:  return len(a)
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        prev = dp[:]
        dp[0] = i + 1
        for j, cb in enumerate(b):
            dp[j+1] = min(prev[j] + (0 if ca == cb else 1), dp[j]+1, prev[j+1]+1)
    return dp[len(b)]


def check_drc(components: Dict, nets: Dict) -> dict:
    missing_hq, missing_val, missing_pkg, tbd_attrs, single_pin, unnamed = [], [], [], [], [], []
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype == 'TESTPOINT':
            continue
        base = {'位号': refdes, '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': comp.get('page', '')}
        if not comp.get('hq_code'):  missing_hq.append(base.copy())
        if not comp.get('value'):    missing_val.append(base.copy())
        if not comp.get('package'):  missing_pkg.append(base.copy())
        for attr in ('voltage', 'current', 'power'):
            val = comp.get(attr, '')
            if val and 'TBD' in val.upper():
                tbd_attrs.append({'位号': refdes, '属性': attr.upper(), '当前值': val,
                                  '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': comp.get('page','')})
    for net_name, nodes in nets.items():
        if len(nodes) == 1:
            n = nodes[0]
            single_pin.append({'网络名': net_name, '连接元件': n['refdes'],
                                '引脚': n['pin_name'],
                                '页面': components.get(n['refdes'], {}).get('page', '')})
        if re.search(r'^UNNAMED_', net_name, re.I):
            unnamed.append({'网络名': net_name, '节点数': len(nodes)})
    option_map: Dict[str, List[str]] = defaultdict(list)
    for refdes, comp in components.items():
        option_map[(comp.get('bom_option') or '').strip().upper()].append(refdes)
    typos = []
    for val, refs in sorted(option_map.items()):
        if val in _VALID_BOM_OPTIONS:
            continue
        min_d   = min(_edit_distance(val, kw) for kw in _FUZZY_KEYWORDS)
        nearest = min(_FUZZY_KEYWORDS, key=lambda kw: _edit_distance(val, kw))
        typos.append({'实际填写值': val, '疑似应为': nearest if min_d <= 2 else '未知',
                      '编辑距离': min_d, '使用该值的位号': ', '.join(sorted(refs)),
                      '数量': len(refs), '风险': '❌ 疑似拼错' if min_d <= 2 else '⚠ 未知值'})
    return {
        'missing_hq_code': missing_hq, 'missing_value': missing_val,
        'missing_package': missing_pkg, 'tbd_attrs': tbd_attrs,
        'single_pin_nets': single_pin, 'unnamed_nets': unnamed,
        'bom_option_typos': typos,
    }


# ══════════════════════════════════════════════════════════
# 五、电容降额分析
# ══════════════════════════════════════════════════════════

_VOLT_RULES: List[Tuple[str, float]] = [
    (r'P48V', 48.0), (r'P24V', 24.0), (r'P19V', 19.0), (r'P15V', 15.0),
    (r'P12V', 12.0), (r'\b12V', 12.0), (r'P9V',  9.0),  (r'P7V',  7.4),
    (r'P5V(?!\d)', 5.0), (r'\b5V', 5.0), (r'P3V3', 3.3), (r'\b3V3', 3.3),
    (r'P3V', 3.3),  (r'P2V5', 2.5), (r'2V5', 2.5),  (r'P1V8', 1.8),
    (r'1V8', 1.8),  (r'P1V5', 1.5), (r'1V5', 1.5),  (r'P1V2', 1.2),
    (r'1V2', 1.2),  (r'P1V05', 1.05), (r'1V05', 1.05), (r'P1V(?!\d)', 1.0),
    (r'1V0', 1.0),  (r'P0V9', 0.9), (r'0V9', 0.9),  (r'P0V8', 0.8),
    (r'GND', 0.0),  (r'AGND|PGND|DGND|SGND', 0.0),
]


# OD/PG 信号模式：这类网络的电压由外部上拉决定，无法从网络名直接推断
_OD_SKIP_PATTERNS = re.compile(
    r'\bPG\b|PGOOD|_PG_|_PGD\b|PG_N|PWRGD|POWER_GOOD'
    r'|\bFAULT\b|_FAULT|VR_FAULT'
    r'|\bALERT\b|_ALERT|SMBALERT'
    r'|\bSDA\b|\bSCL\b'
    r'|\bOC_N\b|_OC\b'
    r'|\bPRSNT\b|\bPRESENT\b'
    r'|\bINT_N\b|\bIRQ_N\b',
    re.IGNORECASE,
)


def _infer_voltage(net_name: str) -> Optional[float]:
    if _OD_SKIP_PATTERNS.search(net_name):
        return None          # OD/PG 信号跳过，避免错误推断
    for pattern, volt in _VOLT_RULES:
        if re.search(pattern, net_name, re.IGNORECASE):
            return volt
    return None


def _is_od_net(net_name: str) -> bool:
    return bool(_OD_SKIP_PATTERNS.search(net_name))


def analyze_derating(components: Dict, nets: Dict,
                     pct: float = 70.0,
                     custom_volt_map: Optional[Dict[str, float]] = None) -> List[dict]:
    """pct: 工作电压上限占额定电压的百分比，默认 70（即工作电压 ≤ 额定 × 70% 视为合格）"""
    comp_nets: Dict[str, List[str]] = defaultdict(list)
    for net_name, nodes in nets.items():
        for node in nodes:
            comp_nets[node['refdes']].append(net_name)

    rows = []
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype not in ('CAP', 'CAP_POL'):
            continue
        rated_str = comp.get('voltage', '')
        if not rated_str:
            status, derating, max_v, from_net = '⚪ 无额定电压', None, None, ''
        else:
            m = re.match(r'([\d.]+)\s*V', rated_str.strip(), re.I)
            rated_v = float(m.group(1)) if m else None
            if rated_v is None:
                status, derating, max_v, from_net = '⚪ 无法解析额定电压', None, None, ''
            else:
                max_v, from_net = None, ''
                for net_name in comp_nets.get(refdes, []):
                    v = None
                    if custom_volt_map:
                        for key, vv in custom_volt_map.items():
                            if key.upper() in net_name.upper():
                                v = vv; break
                    if v is None:
                        v = _infer_voltage(net_name)
                    if v is not None and v > 0 and (max_v is None or v > max_v):
                        max_v, from_net = v, net_name
                if max_v is None:
                    # 判断是否因为连接了 OD/PG 网络导致无法推断
                    od_nets = [n for n in comp_nets.get(refdes, []) if _is_od_net(n)]
                    if od_nets:
                        status = f'⚪ OD/PG信号（{od_nets[0]}），工作电压由上拉决定，请手动确认'
                    else:
                        status = '⚪ 无法推断工作电压'
                    derating = None
                else:
                    usage_pct = max_v / rated_v * 100        # 工作电压占额定的 %
                    derating  = rated_v / max_v              # 仍保留降额比供参考
                    if usage_pct <= pct:
                        status = f'✅ 合格 ({usage_pct:.0f}% ≤ {pct:.0f}%)'
                    else:
                        status = f'❌ 不合格 ({usage_pct:.0f}% > {pct:.0f}%)'
        rows.append({
            '位号':            refdes,
            '值':              comp.get('value', ''),
            '封装':            comp.get('package', ''),
            '类型':            COMP_TYPE_CN.get(ctype, ctype),
            '额定电压':        rated_str,
            '推断工作电压(V)': str(max_v) if max_v is not None else '',
            '推断来源网络':    from_net,
            '所有连接网络':    ', '.join(comp_nets.get(refdes, [])),
            '降额比':          f'{derating:.2f}' if derating is not None else '',
            '状态':            status,
            '页面':            comp.get('page', ''),
            'DEPOP':           'Y' if comp.get('bom_option') == 'DEPOP' else '',
        })
    rows.sort(key=lambda r: 0 if r['状态'].startswith('❌') else 1)
    return rows


# ══════════════════════════════════════════════════════════
# 六、电阻检查（上拉 / 下拉 / 串阻）
# ══════════════════════════════════════════════════════════

def _parse_ohms(value_str: str) -> Optional[float]:
    """解析电阻值字符串为欧姆数，支持 k/M/R/Ω 后缀，如 10k→10000, 4.7k→4700, 100R→100"""
    if not value_str:
        return None
    s = re.sub(r'\s', '', value_str.upper())
    s = s.replace('Ω', 'R').replace('OHM', 'R').replace('OHMS', 'R')
    m = re.match(r'^([\d.]+)([KMGR]?)$', s)
    if not m:
        return None
    val = float(m.group(1))
    return val * {'K': 1e3, 'M': 1e6, 'G': 1e9, 'R': 1, '': 1}.get(m.group(2), 1)


def _net_is_power(net: str) -> bool:
    return bool(re.search(r'^P\d|^[0-9]+V|VCC|VDD|VBAT|VCORE|VCCIO|PVDD|AVDD|DVDD', net, re.I))


def _net_is_gnd(net: str) -> bool:
    return bool(re.search(r'GND|AGND|SGND|PGND|DGND', net, re.I))


# 常见 OD/OC 信号名特征（不区分大小写）
_OD_OC_PATTERNS = [
    r'\bSDA\b', r'\bSCL\b', r'SMBA', r'SMBALERT',
    r'\bFAULT\b', r'_FAULT', r'VR_FAULT', r'VRD_FAULT',
    r'\bALERT\b', r'_ALERT',
    r'\bPGD\b', r'_PGD', r'VRD_PGD', r'VR_PGD',
    r'OC_N', r'_OC\b',
    r'\bPRESENT\b', r'\bPRSNT\b', r'_PRSNT',
    r'INT_N', r'IRQ_N',
]


def analyze_resistors(components: Dict, nets: Dict) -> dict:
    """检测上拉/下拉/串阻相关设计问题"""
    pullups:   Dict[str, list] = defaultdict(list)
    pulldowns: Dict[str, list] = defaultdict(list)
    series_list: list = []

    for refdes, comp in components.items():
        if comp.get('comp_type') != 'RES':
            continue
        pin_nets = list(comp.get('nets', {}).values())
        if len(pin_nets) < 2:
            continue
        net_a, net_b = pin_nets[0], pin_nets[1]
        ohms    = _parse_ohms(comp.get('value', ''))
        val_str = comp.get('value', '')
        page    = comp.get('page', '')

        a_pwr, b_pwr = _net_is_power(net_a), _net_is_power(net_b)
        a_gnd, b_gnd = _net_is_gnd(net_a),   _net_is_gnd(net_b)

        if a_pwr and not b_pwr and not b_gnd:
            pullups[net_b].append({'refdes': refdes, 'ohms': ohms, 'value': val_str,
                                   'power_net': net_a, 'page': page})
        elif b_pwr and not a_pwr and not a_gnd:
            pullups[net_a].append({'refdes': refdes, 'ohms': ohms, 'value': val_str,
                                   'power_net': net_b, 'page': page})
        elif a_gnd and not b_gnd and not b_pwr:
            pulldowns[net_b].append({'refdes': refdes, 'ohms': ohms,
                                     'value': val_str, 'page': page})
        elif b_gnd and not a_gnd and not a_pwr:
            pulldowns[net_a].append({'refdes': refdes, 'ohms': ohms,
                                     'value': val_str, 'page': page})
        elif not a_pwr and not b_pwr and not a_gnd and not b_gnd:
            series_list.append({'refdes': refdes, 'net_a': net_a, 'net_b': net_b,
                                 'ohms': ohms, 'value': val_str, 'page': page})

    # ── 检查1：重复上拉 ───────────────────────────────────
    dup_pullups = []
    for sig_net, pu_list in sorted(pullups.items()):
        if len(pu_list) > 1:
            dup_pullups.append({
                '信号网络':  sig_net,
                '上拉数量':  len(pu_list),
                '位号':      ', '.join(r['refdes'] for r in pu_list),
                '阻值':      ', '.join(r['value']  for r in pu_list),
                '上拉电源':  ', '.join(dict.fromkeys(r['power_net'] for r in pu_list)),
                '页面':      ', '.join(dict.fromkeys(r['page']      for r in pu_list)),
            })

    # ── 检查2：重复下拉 ───────────────────────────────────
    dup_pulldowns = []
    for sig_net, pd_list in sorted(pulldowns.items()):
        if len(pd_list) > 1:
            dup_pulldowns.append({
                '信号网络': sig_net,
                '下拉数量': len(pd_list),
                '位号':     ', '.join(r['refdes'] for r in pd_list),
                '阻值':     ', '.join(r['value']  for r in pd_list),
                '页面':     ', '.join(dict.fromkeys(r['page'] for r in pd_list)),
            })

    # ── 检查3：串阻 + 上拉分压风险 ───────────────────────
    divider_risks = []
    seen_pairs: set = set()
    for sr in series_list:
        for pu_net in (sr['net_a'], sr['net_b']):
            if pu_net not in pullups:
                continue
            for pu in pullups[pu_net]:
                pair_key = (sr['refdes'], pu['refdes'])
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                sr_ohms, pu_ohms = sr['ohms'], pu['ohms']
                if sr_ohms is not None and pu_ohms is not None and pu_ohms > 0:
                    ratio = sr_ohms / pu_ohms
                    if pu_ohms < 1000:
                        status = '❌ 高风险' if ratio > 0.1 else '✅ 正常'
                    else:
                        status = '⚠️ 关注'   if ratio > 0.1 else '✅ 正常'
                else:
                    ratio, status = None, '⚪ 阻值未填，无法计算'
                divider_risks.append({
                    '串阻位号':   sr['refdes'],
                    '串阻值':     sr['value'],
                    '上拉位号':   pu['refdes'],
                    '上拉值':     pu['value'],
                    '上拉电源':   pu['power_net'],
                    '公共节点':   pu_net,
                    '串/拉比':    f'{ratio:.3f}' if ratio is not None else '',
                    '上拉 < 1k':  '是' if (pu_ohms or 0) < 1000 else '否',
                    '状态':       status,
                    '页面':       sr['page'],
                })
            break
    divider_risks.sort(key=lambda r: 0 if r['状态'].startswith('❌')
                       else 1 if r['状态'].startswith('⚠') else 2)

    # ── 检查4：OD/OC 信号缺上拉 ──────────────────────────
    od_missing = []
    for net_name in sorted(nets.keys()):
        if any(re.search(p, net_name, re.I) for p in _OD_OC_PATTERNS):
            if net_name not in pullups:
                nodes = nets[net_name]
                od_missing.append({
                    '网络名':   net_name,
                    '节点数':   len(nodes),
                    '连接元件': ', '.join(n['refdes'] for n in nodes[:6]),
                    '说明':     '疑似 OD/OC 信号，未找到上拉电阻',
                })

    return {
        'dup_pullups':    dup_pullups,
        'dup_pulldowns':  dup_pulldowns,
        'divider_risks':  divider_risks,
        'od_missing':     od_missing,
        'pullups':        dict(pullups),
        'pulldowns':      dict(pulldowns),
    }


# ══════════════════════════════════════════════════════════
# 八、Excel 导出
# ══════════════════════════════════════════════════════════

_BL = PatternFill("solid", fgColor="1F4E79")
_OR = PatternFill("solid", fgColor="C55A11")
_GR = PatternFill("solid", fgColor="375623")
_GY = PatternFill("solid", fgColor="595959")
_RF = PatternFill("solid", fgColor="FFCCCC")
_WF = Font(color="FFFFFF", bold=True, size=10)
_BF = Font(bold=True, size=10)
_NF = Font(size=10)
_CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LA = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_TH = Side(style='thin')
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


def _xl_hdr(ws, row_idx, fill):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill; cell.font = _WF; cell.alignment = _CA; cell.border = _BD


def _xl_autowidth(ws, mx=50):
    for col in ws.columns:
        vals = [str(c.value or '') for c in col]
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(v) for v in vals), default=8) + 2, mx)


def _xl_write_rows(ws, rows: List[dict], fill, hl_col=None, freeze=True):
    if not rows:
        ws.append(['（无数据）']); return
    hdrs = list(rows[0].keys())
    ws.append(hdrs); _xl_hdr(ws, ws.max_row, fill)
    hl_idx = hdrs.index(hl_col) if hl_col in hdrs else None
    for row in rows:
        ws.append(list(row.values()))
        ri  = ws.max_row
        red = hl_idx is not None and '❌' in str(ws.cell(ri, hl_idx+1).value or '')
        for cell in ws[ri]:
            cell.border = _BD; cell.alignment = _LA; cell.font = _NF
            if red: cell.fill = _RF
    _xl_autowidth(ws)
    if freeze: ws.freeze_panes = 'A2'


def _xl_section(ws, title, fill):
    ws.append([title])
    for cell in ws[ws.max_row]:
        cell.fill = fill; cell.font = _WF; cell.border = _BD
    ws.append([])


def export_to_excel(data: dict, out_path: str) -> str:
    base, ext = os.path.splitext(out_path)
    n, path = 1, out_path
    while os.path.exists(path):
        path = f'{base}({n}){ext}'; n += 1

    wb = Workbook(); wb.remove(wb.active)
    project = data.get('project_name', '')
    na  = data.get('net_analysis', {})
    drc = data.get('drc', {})
    drt = data.get('derating', [])
    mn  = data.get('bom_normal_merged', [])
    md  = data.get('bom_depop_merged', [])

    # 概览
    ws = wb.create_sheet('概览')
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 16
    drc_total = sum(len(v) for v in drc.values() if isinstance(v, list))
    fail = sum(1 for r in drt if r.get('状态', '').startswith('❌'))
    for label, val in [
        ('项目名称', project),
        ('贴装元件种类数', len(mn)),
        ('贴装元件总数',  sum(r.get('数量', 0) for r in mn)),
        ('DEPOP 元件种类数', len(md)),
        ('DEPOP 元件总数',  sum(r.get('数量', 0) for r in md)),
        ('网络总数', na.get('total', '')),
        ('单端网络数（疑似漏连）', len(na.get('single_node', {}))),
        ('电源网络数', len(na.get('power_nets', {}))),
        ('差分对数', len(na.get('diff_pairs', {}))),
        ('DRC 问题总数', drc_total),
        ('电容降额不合格数', fail),
    ]:
        ws.append([label, val])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _BD
            cell.font = _BF if cell.column == 1 else _NF
            cell.alignment = _LA

    # BOM
    ws = wb.create_sheet('BOM_贴装'); _xl_write_rows(ws, mn, _BL)
    ws = wb.create_sheet('BOM_DEPOP'); _xl_write_rows(ws, md, _OR)
    ws = wb.create_sheet('BOM_明细')
    all_d = [{'DEPOP': '', **r} for r in data.get('bom_normal_detail', [])] + \
            [{'DEPOP': 'Y', **r} for r in data.get('bom_depop_detail', [])]
    _xl_write_rows(ws, all_d, _GY)

    # 网络分析
    ws = wb.create_sheet('网络分析'); ws.freeze_panes = None
    _xl_section(ws, '电源网络', _BL)
    _xl_write_rows(ws, [{'网络名': k, '节点数': len(v)}
                        for k, v in sorted(na.get('power_nets', {}).items(), key=lambda x: -len(x[1]))],
                   _BL, freeze=False)
    ws.append([])
    _xl_section(ws, 'GND 网络', _GR)
    _xl_write_rows(ws, [{'网络名': k, '节点数': len(v)}
                        for k, v in sorted(na.get('gnd_nets', {}).items(), key=lambda x: -len(x[1]))],
                   _GR, freeze=False)
    ws.append([])
    _xl_section(ws, '差分对', _OR)
    _xl_write_rows(ws, [{'基础名': b, 'P端网络': pr['P'], 'N端网络': pr['N']}
                        for b, pr in sorted(na.get('diff_pairs', {}).items())],
                   _OR, freeze=False)
    ws.append([])
    _xl_section(ws, '单端网络（疑似漏连）', _GY)
    _xl_write_rows(ws, [{'网络名': k, '连接元件': v[0]['refdes'], '引脚': v[0]['pin_name']}
                        for k, v in sorted(na.get('single_node', {}).items())],
                   _GY, freeze=False)
    ws.append([])
    _xl_section(ws, '各页面元件数', _BL)
    _xl_write_rows(ws, [{'页面': p, '元件数': c}
                        for p, c in sorted(na.get('page_counter', {}).items())],
                   _BL, freeze=False)
    _xl_autowidth(ws)

    # 设计检查
    ws = wb.create_sheet('设计检查'); ws.freeze_panes = None
    for title, key, fill in [
        ('TBD 待确认属性', 'tbd_attrs',       _OR),
        ('缺少料号',       'missing_hq_code',  _RF),
        ('缺少 VALUE',     'missing_value',     _RF),
        ('缺少封装',       'missing_package',   _RF),
        ('单端网络',       'single_pin_nets',   _GY),
        ('未命名网络',     'unnamed_nets',      _GY),
        ('BOM_OPTION 拼写', 'bom_option_typos', _OR),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, drc.get(key, []), fill, freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    # 降额
    ws = wb.create_sheet('降额分析')
    _xl_write_rows(ws, drt, _BL, hl_col='状态')

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════
# 九、GUI
# ══════════════════════════════════════════════════════════

def _make_tree(parent, columns, height=12):
    """创建带双向滚动条的 Treeview，返回 (outer_frame, tree)"""
    outer = tk.Frame(parent)
    tree  = ttk.Treeview(outer, columns=columns, show='headings', height=height)
    vsb   = ttk.Scrollbar(outer, orient='vertical',   command=tree.yview)
    hsb   = ttk.Scrollbar(outer, orient='horizontal',  command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    tree.grid(row=0, column=0, sticky='nsew')
    vsb.grid(row=0, column=1, sticky='ns')
    hsb.grid(row=1, column=0, sticky='ew')
    outer.grid_rowconfigure(0, weight=1)
    outer.grid_columnconfigure(0, weight=1)
    return outer, tree


def _natural_key(s: str):
    """自然排序 key：将字符串中的数字段按数值比较，如 C2 < C10 < C100"""
    return [int(p) if p.isdigit() else p.lower() for p in re.split(r'(\d+)', s)]


def _sort_tree(tree, col, reverse: bool):
    """点击表头时对 Treeview 排序，自动区分纯数值 / 混合字符串，循环切换升降序。"""
    items = [(tree.set(iid, col), iid) for iid in tree.get_children('')]
    try:
        items.sort(key=lambda t: (float(t[0]) if t[0] else float('-inf')), reverse=reverse)
    except ValueError:
        items.sort(key=lambda t: _natural_key(t[0]), reverse=reverse)
    for idx, (_, iid) in enumerate(items):
        tree.move(iid, '', idx)
    # 更新所有列标题：当前排序列显示箭头，其他列恢复原名
    arrow = ' ▲' if not reverse else ' ▼'
    for c in tree['columns']:
        base = tree.heading(c, 'text').rstrip(' ▲▼')
        if c == col:
            tree.heading(c, text=base + arrow,
                         command=lambda _c=c: _sort_tree(tree, _c, not reverse))
        else:
            tree.heading(c, text=base,
                         command=lambda _c=c: _sort_tree(tree, _c, False))


def _fill_tree(tree, rows: list, columns: list = None):
    tree.delete(*tree.get_children())
    if not rows:
        return
    cols = columns or list(rows[0].keys())
    tree['columns'] = cols
    for c in cols:
        tree.heading(c, text=c, anchor='w',
                     command=lambda _c=c: _sort_tree(tree, _c, False))
        tree.column(c, width=min(max(len(c)*9, 80), 220), anchor='w', stretch=True)
    for row in rows:
        tree.insert('', 'end', values=[str(row.get(c, '')) for c in cols])


class PstxApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title('PSTX 原理图分析工具 v1.1')
        self.geometry('1040x720')
        self.minsize(880, 580)
        self.resizable(True, True)

        self._components = {}; self._nets = {}
        self._dn = []; self._dd = []; self._mn = []; self._md = []
        self._na = {}; self._drc = {}; self._drt = []; self._res = {}

        self.prt_path    = tk.StringVar()
        self.net_path    = tk.StringVar()
        self.ref_path    = tk.StringVar()
        self.project_var = tk.StringVar()
        self.bom_filter  = tk.StringVar(value='贴装')
        self.bom_search  = tk.StringVar()
        self.query_mode  = tk.StringVar(value='位号')
        self.query_text  = tk.StringVar()
        self.ratio_var   = tk.DoubleVar(value=70.0)

        self.bom_filter.trace_add('write', lambda *_: self._refresh_bom())
        self.bom_search.trace_add('write', lambda *_: self._refresh_bom())

        self._build_ui()

    # ──────── UI 搭建 ─────────────────────────────────────

    def _section(self, parent, title):
        f = ttk.LabelFrame(parent, text=title, padding=8)
        f.pack(fill='x', padx=10, pady=4)
        return f

    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=6)
        self.nb = nb
        for text, builder in [
            ('  文件加载  ', self._build_load),
            ('  BOM 管理  ', self._build_bom),
            ('  网络分析  ', self._build_net),
            ('  设计检查  ', self._build_drc),
            ('  电阻检查  ', self._build_res),
            ('  电容降额  ', self._build_derating),
            ('  元件查询  ', self._build_query),
            ('  日志      ', self._build_log),
        ]:
            f = ttk.Frame(nb); nb.add(f, text=text); builder(f)

    # ── Tab：文件加载 ──────────────────────────────────────

    def _build_load(self, p):
        # ── 主操作区：自动识别（突出显示）──────────────────────
        fa = tk.Frame(p, bg='#e8f0fe', relief='groove', bd=1)
        fa.pack(fill='x', padx=10, pady=(10, 4))
        inner = tk.Frame(fa, bg='#e8f0fe')
        inner.pack(padx=14, pady=10)
        tk.Label(inner, text='快速加载', font=('Arial', 11, 'bold'),
                 bg='#e8f0fe', fg='#1a3a8f').pack(side='left', padx=(0, 10))
        tk.Button(inner, text='  选择文件夹…  ', font=('Arial', 10, 'bold'),
                  bg='#2d6cdf', fg='white', relief='flat', padx=8, pady=4,
                  command=self._auto_detect).pack(side='left')
        self.auto_detect_lbl = tk.Label(
            inner,
            text='选择 worklib 中该项目的文件夹，自动识别并填入下方路径',
            bg='#e8f0fe', fg='#444')
        self.auto_detect_lbl.pack(side='left', padx=12)

        # ── 分隔线 + 说明 ───────────────────────────────────────
        sep_row = tk.Frame(p)
        sep_row.pack(fill='x', padx=10, pady=(6, 0))
        ttk.Separator(sep_row, orient='horizontal').pack(side='left', fill='x', expand=True, pady=6)
        tk.Label(sep_row, text='  或手动选择各文件  ', fg='#888').pack(side='left')
        ttk.Separator(sep_row, orient='horizontal').pack(side='left', fill='x', expand=True, pady=6)

        # ── 次级区：单文件手动选择 ─────────────────────────────
        for label, var in [
            ('pstxprt.dat  【必须】元件属性 — 位号、料号、封装、电气参数等', self.prt_path),
            ('pstxnet.dat  【必须】网络连接 — 引脚与网络的映射关系',          self.net_path),
            ('pstxref.dat  【可选】交叉参考 — 元件与原理图符号的交叉索引',    self.ref_path),
        ]:
            f = self._section(p, label)
            tk.Label(f, text='文件路径：').grid(row=0, column=0, sticky='w')
            ttk.Entry(f, textvariable=var, width=58).grid(row=0, column=1, padx=6)
            ttk.Button(f, text='浏览…',
                       command=lambda v=var: self._browse_dat(v)).grid(row=0, column=2)

        fp = self._section(p, '项目名称（导出报告用）')
        tk.Label(fp, text='项目名称：').grid(row=0, column=0, sticky='w')
        ttk.Entry(fp, textvariable=self.project_var, width=40).grid(
            row=0, column=1, padx=6, sticky='w')

        br = tk.Frame(p); br.pack(pady=14)
        self.parse_btn = tk.Button(
            br, text='开始解析', font=('Arial', 13, 'bold'),
            bg='#2d6cdf', fg='white', relief='flat',
            padx=24, pady=10, command=self._run_parse)
        self.parse_btn.pack(side='left', padx=8)
        self.load_status = tk.Label(br, text='', font=('Arial', 11))
        self.load_status.pack(side='left', padx=8)

        fo = self._section(p, '解析概览')
        self.overview_text = tk.Text(fo, height=7, font=('Consolas', 9),
                                      state='disabled', bg='#f5f5f5', relief='flat')
        self.overview_text.pack(fill='x')

    # ── Tab：BOM 管理 ──────────────────────────────────────

    def _build_bom(self, p):
        ctrl = self._section(p, '筛选 / 搜索')
        for val, txt in [('贴装', '贴装元件'), ('DEPOP', 'DEPOP'), ('全部', '全部')]:
            ttk.Radiobutton(ctrl, text=txt, variable=self.bom_filter,
                            value=val).pack(side='left', padx=10)
        ttk.Separator(ctrl, orient='vertical').pack(side='left', fill='y', padx=6)
        tk.Label(ctrl, text='搜索：').pack(side='left')
        ttk.Entry(ctrl, textvariable=self.bom_search, width=26).pack(side='left', padx=4)

        outer, self.bom_tree = _make_tree(p, ['位号', '料号', '值', '封装', '类型', '页面'], height=16)
        outer.pack(fill='both', expand=True, padx=10, pady=4)

        bot = tk.Frame(p); bot.pack(fill='x', padx=10, pady=4)
        self.bom_count_lbl = tk.Label(bot, text='', fg='#444')
        self.bom_count_lbl.pack(side='left')
        ttk.Button(bot, text='导出 Excel', command=self._export_excel).pack(side='right')

    # ── Tab：网络分析 ──────────────────────────────────────

    def _build_net(self, p):
        fs = self._section(p, '汇总')
        self.net_stat = tk.Text(fs, height=2, font=('Consolas', 9),
                                 state='disabled', bg='#f5f5f5', relief='flat')
        self.net_stat.pack(fill='x')

        sub = ttk.Notebook(p); sub.pack(fill='both', expand=True, padx=10, pady=4)
        for title, cols, attr in [
            ('电源网络',     ['网络名', '节点数'],           '_tree_power'),
            ('GND 网络',     ['网络名', '节点数'],           '_tree_gnd'),
            ('差分对',       ['基础名', 'P端网络', 'N端网络'],'_tree_diff'),
            ('单端网络',     ['网络名', '连接元件', '引脚'],  '_tree_single'),
            ('各页面元件数', ['页面', '元件数'],             '_tree_pages'),
        ]:
            f = ttk.Frame(sub); sub.add(f, text=f'  {title}  ')
            outer, tree = _make_tree(f, cols, height=14)
            outer.pack(fill='both', expand=True)
            setattr(self, attr, tree)

    # ── Tab：设计检查 ──────────────────────────────────────

    def _build_drc(self, p):
        sub = ttk.Notebook(p); sub.pack(fill='both', expand=True, padx=10, pady=8)
        for title, cols, attr in [
            ('缺料号',     ['位号', '类型', '页面'],                                       '_tree_drc_hq'),
            ('缺 VALUE',   ['位号', '类型', '页面'],                                       '_tree_drc_val'),
            ('缺封装',     ['位号', '类型', '页面'],                                       '_tree_drc_pkg'),
            ('TBD 属性',   ['位号', '属性', '当前值', '类型', '页面'],                     '_tree_drc_tbd'),
            ('单端网络',   ['网络名', '连接元件', '引脚', '页面'],                         '_tree_drc_single'),
            ('BOM_OPTION', ['实际填写值', '疑似应为', '编辑距离', '使用该值的位号', '风险'], '_tree_drc_opt'),
        ]:
            f = ttk.Frame(sub); sub.add(f, text=f'  {title}  ')
            outer, tree = _make_tree(f, cols, height=15)
            outer.pack(fill='both', expand=True)
            setattr(self, attr, tree)

    # ── Tab：电阻检查 ──────────────────────────────────────

    def _build_res(self, p):
        sub = ttk.Notebook(p); sub.pack(fill='both', expand=True, padx=10, pady=8)
        for title, cols, attr in [
            ('串阻分压风险',
             ['串阻位号', '串阻值', '上拉位号', '上拉值', '上拉电源',
              '公共节点', '串/拉比', '上拉 < 1k', '状态', '页面'],
             '_tree_res_div'),
            ('重复上拉',
             ['信号网络', '上拉数量', '位号', '阻值', '上拉电源', '页面'],
             '_tree_res_dup_pu'),
            ('重复下拉',
             ['信号网络', '下拉数量', '位号', '阻值', '页面'],
             '_tree_res_dup_pd'),
            ('OD/OC 缺上拉',
             ['网络名', '节点数', '连接元件', '说明'],
             '_tree_res_od'),
        ]:
            f = ttk.Frame(sub); sub.add(f, text=f'  {title}  ')
            outer, tree = _make_tree(f, cols, height=15)
            outer.pack(fill='both', expand=True)
            setattr(self, attr, tree)

    # ── Tab：电容降额 ──────────────────────────────────────

    def _build_derating(self, p):
        fc = self._section(p, '参数设置')
        tk.Label(fc, text='工作电压上限（额定电压的 %）：').grid(row=0, column=0, sticky='w')
        ttk.Scale(fc, from_=50, to=100, orient='horizontal',
                  variable=self.ratio_var, length=200).grid(row=0, column=1, padx=8)
        self.ratio_lbl = tk.Label(fc, text='70%', width=6, font=('Arial', 11, 'bold'))
        self.ratio_lbl.grid(row=0, column=2)
        self.ratio_var.trace_add('write', lambda *_: self.ratio_lbl.configure(
            text=f'{self.ratio_var.get():.0f}%'))

        tk.Label(fc, text='自定义电压映射\n（每行 NET前缀=电压V）：',
                 justify='left').grid(row=1, column=0, sticky='nw', pady=6)
        self.volt_entry = tk.Text(fc, height=3, width=38, font=('Consolas', 9))
        self.volt_entry.grid(row=1, column=1, columnspan=2, padx=8, sticky='w')
        self.volt_entry.insert('1.0', '# 示例：VBUS=5.0\n# P12V_AUX=12.0')
        btn_row = tk.Frame(fc); btn_row.grid(row=2, column=1, sticky='w', pady=4)
        ttk.Button(btn_row, text='重新计算',
                   command=self._recalc_derating).pack(side='left')
        self._rules_visible = False
        self._rules_btn = ttk.Button(btn_row, text='查看内置电压匹配规则 ▾',
                                     command=self._toggle_rules)
        self._rules_btn.pack(side='left', padx=12)

        # 可折叠的规则说明区
        self._rules_frame = tk.Frame(fc, relief='sunken', bd=1, bg='#f8f8f8')
        self._rules_frame.grid(row=3, column=0, columnspan=3, sticky='ew', padx=0, pady=(0, 4))
        self._rules_frame.grid_remove()          # 默认隐藏

        rules_txt = scrolledtext.ScrolledText(
            self._rules_frame, font=('Consolas', 9), height=10,
            bg='#f8f8f8', relief='flat', state='normal')
        rules_txt.pack(fill='both', expand=True, padx=6, pady=4)

        algo_text = (
            "【工作电压推断算法】\n"
            "  1. 读取该电容连接的所有网络名\n"
            "  2. 对每个网络名逐条匹配下方规则，取最大非零值作为工作电压\n"
            "  3. 若用户填写了自定义映射（NET前缀=电压），优先匹配\n"
            "  4. 合格条件：工作电压 ≤ 额定电压 × 上限百分比（默认 70%）\n\n"
            "【内置电压匹配规则（按匹配优先级）】\n"
            f"  {'网络名关键字':<20} {'推断电压 (V)'}\n"
            f"  {'-'*36}\n"
        )
        for pattern, volt in _VOLT_RULES:
            # 去掉正则符号，展示可读关键字
            readable = re.sub(r'[\\b\\(\\)\\?!^]', '', pattern).replace('\\', '')
            algo_text += f"  {readable:<20} {volt}\n"
        algo_text += (
            "\n【⚪ 无法判断的常见原因】\n"
            "  · 原理图未填写 VOLTAGE 属性\n"
            "  · 连接网络全用自定义命名（如 NET_A、SIGNAL_1），无法匹配规则\n"
            "    → 解决：在上方自定义映射框里手动添加 NET前缀=电压\n"
        )
        rules_txt.insert('1.0', algo_text)
        rules_txt.configure(state='disabled')

        cols = ['位号', '值', '封装', '类型', '额定电压', '推断工作电压(V)',
                '推断来源网络', '降额比', '状态', '页面', 'DEPOP']
        outer, self.drt_tree = _make_tree(p, cols, height=14)
        outer.pack(fill='both', expand=True, padx=10, pady=4)
        self.drt_stat = tk.Label(p, text='', fg='#555')
        self.drt_stat.pack(anchor='w', padx=10, pady=2)

    # ── Tab：元件查询 ──────────────────────────────────────

    def _build_query(self, p):
        fc = self._section(p, '查询')
        for val, txt in [('位号', '按位号 (refdes)'), ('网络名', '按网络名')]:
            ttk.Radiobutton(fc, text=txt, variable=self.query_mode,
                            value=val).pack(side='left', padx=10)
        ttk.Separator(fc, orient='vertical').pack(side='left', fill='y', padx=6)
        ttk.Entry(fc, textvariable=self.query_text, width=30).pack(side='left', padx=4)
        ttk.Button(fc, text='查询', command=self._do_query).pack(side='left', padx=4)

        fr = self._section(p, '查询结果')
        self.query_result = scrolledtext.ScrolledText(
            fr, font=('Consolas', 10), state='disabled',
            bg='#1e1e1e', fg='#d4d4d4', relief='flat', height=22)
        self.query_result.pack(fill='both', expand=True)

    # ── Tab：日志 ──────────────────────────────────────────

    def _build_log(self, p):
        self.log = scrolledtext.ScrolledText(
            p, font=('Consolas', 9), state='disabled',
            bg='#1e1e1e', fg='#d4d4d4', relief='flat')
        self.log.pack(fill='both', expand=True, padx=8, pady=8)
        ttk.Button(p, text='清空日志',
                   command=self._clear_log).pack(anchor='e', padx=8, pady=4)

    # ──────── 事件 ────────────────────────────────────────

    def _browse_dat(self, var):
        path = filedialog.askopenfilename(
            title='选择 .dat 文件',
            filetypes=[('DAT 文件', '*.dat'), ('所有文件', '*.*')])
        if path:
            var.set(path); self._log(f'选择文件：{path}')

    def _auto_detect(self):
        folder = filedialog.askdirectory(title='选择包含 PST 文件的文件夹')
        if not folder:
            return
        # 在当前目录 + 一级子目录中搜索
        candidates = [folder] + [
            os.path.join(folder, d) for d in os.listdir(folder)
            if os.path.isdir(os.path.join(folder, d))
        ]
        targets = {'pstxprt.dat': self.prt_path,
                   'pstxnet.dat': self.net_path,
                   'pstxref.dat': self.ref_path}
        found, missing = {}, []
        for name, var in targets.items():
            hit = None
            for d in candidates:
                p = os.path.join(d, name)
                if os.path.isfile(p):
                    hit = p; break
            if hit:
                var.set(hit)
                found[name] = hit
            else:
                missing.append(name)
        # 更新提示标签
        if found:
            msg = f'找到 {len(found)} 个：{", ".join(found.keys())}'
            if missing:
                msg += f'    未找到：{", ".join(missing)}'
            color = '#2a8a2a' if 'pstxprt.dat' in found and 'pstxnet.dat' in found else '#b06000'
        else:
            msg = f'未在该目录下找到任何 PST 文件'
            color = 'red'
        self.auto_detect_lbl.configure(text=msg, fg=color)
        self._log(f'\n自动识别文件夹：{folder}')
        for name, path in found.items():
            self._log(f'  ✅ {name} → {path}')
        for name in missing:
            self._log(f'  ⚪ {name} 未找到')

    # ──────── 解析流程 ────────────────────────────────────

    def _run_parse(self):
        if not self.prt_path.get().strip() or not self.net_path.get().strip():
            messagebox.showerror('错误', '请先选择 pstxprt.dat 和 pstxnet.dat（pstxref.dat 可选）')
            return
        self.parse_btn.configure(state='disabled')
        self._start_spinner('解析中')
        threading.Thread(target=self._do_parse, daemon=True).start()

    def _do_parse(self):
        try:
            self._log('\n── 开始解析 ──────────────────')
            with open(self.prt_path.get(), encoding='utf-8', errors='replace') as f:
                prt = f.read()
            with open(self.net_path.get(), encoding='utf-8', errors='replace') as f:
                net = f.read()
            ref_p = self.ref_path.get().strip()
            if ref_p and os.path.isfile(ref_p):
                with open(ref_p, encoding='utf-8', errors='replace') as f:
                    _ref = f.read()
                self._log(f'  pstxprt：{len(prt):,} 字节    pstxnet：{len(net):,} 字节    pstxref：{len(_ref):,} 字节')
            else:
                self._log(f'  pstxprt：{len(prt):,} 字节    pstxnet：{len(net):,} 字节    pstxref：未加载')

            comps, nets, _ = parse_all(prt, net)
            self._log(f'  元件：{len(comps)} 个    网络：{len(nets)} 个')

            dn, dd, mn, md = build_bom(comps)
            na  = analyze_networks(nets, comps)
            drc = check_drc(comps, nets)
            drt = analyze_derating(comps, nets, self.ratio_var.get(), self._volt_map())
            res = analyze_resistors(comps, nets)

            self._components = comps; self._nets = nets
            self._dn = dn; self._dd = dd; self._mn = mn; self._md = md
            self._na = na; self._drc = drc; self._drt = drt; self._res = res

            drc_total = sum(len(v) for v in drc.values() if isinstance(v, list))
            self._log(f'  贴装 {len(mn)} 种 / {sum(r.get("数量",0) for r in mn)} 个')
            self._log(f'  DRC 问题：{drc_total} 项    '
                      f'降额不合格：{sum(1 for r in drt if r["状态"].startswith("❌"))} 个')
            self.after(0, self._on_parse_done)
        except Exception as e:
            import traceback
            self._log(f'❌ {e}\n{traceback.format_exc()}')
            self.after(0, lambda: self._stop_spinner('❌ 解析失败'))
            self.after(0, lambda: messagebox.showerror('错误', str(e)))
        finally:
            self.after(0, lambda: self.parse_btn.configure(state='normal'))

    def _on_parse_done(self):
        self._stop_spinner('✅ 解析完成')
        self._update_overview()
        self._refresh_bom()
        self._refresh_net()
        self._refresh_drc()
        self._refresh_res()
        self._refresh_derating()
        self.nb.select(1)

    def _update_overview(self):
        na = self._na; drc = self._drc; mn = self._mn; md = self._md; drt = self._drt
        drc_total = sum(len(v) for v in drc.values() if isinstance(v, list))
        fail = sum(1 for r in drt if r.get('状态', '').startswith('❌'))
        lines = [
            f'贴装元件：{len(mn)} 种 / {sum(r.get("数量",0) for r in mn)} 个',
            f'DEPOP：   {len(md)} 种 / {sum(r.get("数量",0) for r in md)} 个',
            f'网络：{na.get("total",0)} 个   电源：{len(na.get("power_nets",{}))}   '
            f'GND：{len(na.get("gnd_nets",{}))}   差分对：{len(na.get("diff_pairs",{}))}',
            f'单端网络（疑似漏连）：{len(na.get("single_node",{}))}',
            f'DRC 问题：{drc_total}   电容降额不合格：{fail}',
        ]
        self.overview_text.configure(state='normal')
        self.overview_text.delete('1.0', 'end')
        self.overview_text.insert('end', '\n'.join(lines))
        self.overview_text.configure(state='disabled')

    # ──────── BOM ─────────────────────────────────────────

    def _refresh_bom(self):
        mode = self.bom_filter.get()
        kw   = self.bom_search.get().strip().lower()
        src  = (self._dn if mode == '贴装' else
                self._dd if mode == 'DEPOP' else
                self._dn + self._dd)
        if kw:
            src = [r for r in src if any(kw in str(v).lower() for v in r.values())]
        _fill_tree(self.bom_tree, src, ['位号', '料号', '值', '封装', '类型', '页面'])
        self.bom_count_lbl.configure(text=f'共 {len(src)} 行')

    # ──────── 网络 ─────────────────────────────────────────

    def _refresh_net(self):
        na = self._na
        stat = (f'网络总数：{na.get("total",0)}    '
                f'电源：{len(na.get("power_nets",{}))}    '
                f'GND：{len(na.get("gnd_nets",{}))}    '
                f'差分对：{len(na.get("diff_pairs",{}))}    '
                f'单端（疑似漏连）：{len(na.get("single_node",{}))}')
        self.net_stat.configure(state='normal')
        self.net_stat.delete('1.0', 'end')
        self.net_stat.insert('end', stat)
        self.net_stat.configure(state='disabled')

        _fill_tree(self._tree_power,
                   [{'网络名': k, '节点数': len(v)}
                    for k, v in sorted(na.get('power_nets', {}).items(), key=lambda x: -len(x[1]))],
                   ['网络名', '节点数'])
        _fill_tree(self._tree_gnd,
                   [{'网络名': k, '节点数': len(v)}
                    for k, v in sorted(na.get('gnd_nets', {}).items(), key=lambda x: -len(x[1]))],
                   ['网络名', '节点数'])
        _fill_tree(self._tree_diff,
                   [{'基础名': b, 'P端网络': pr['P'], 'N端网络': pr['N']}
                    for b, pr in sorted(na.get('diff_pairs', {}).items())],
                   ['基础名', 'P端网络', 'N端网络'])
        _fill_tree(self._tree_single,
                   [{'网络名': k, '连接元件': v[0]['refdes'], '引脚': v[0]['pin_name']}
                    for k, v in sorted(na.get('single_node', {}).items())],
                   ['网络名', '连接元件', '引脚'])
        _fill_tree(self._tree_pages,
                   [{'页面': pg, '元件数': cnt}
                    for pg, cnt in sorted(na.get('page_counter', {}).items())],
                   ['页面', '元件数'])

    # ──────── DRC ─────────────────────────────────────────

    def _refresh_drc(self):
        drc = self._drc
        _fill_tree(self._tree_drc_hq,     drc.get('missing_hq_code', []),  ['位号', '类型', '页面'])
        _fill_tree(self._tree_drc_val,    drc.get('missing_value', []),    ['位号', '类型', '页面'])
        _fill_tree(self._tree_drc_pkg,    drc.get('missing_package', []),  ['位号', '类型', '页面'])
        _fill_tree(self._tree_drc_tbd,    drc.get('tbd_attrs', []),
                   ['位号', '属性', '当前值', '类型', '页面'])
        _fill_tree(self._tree_drc_single, drc.get('single_pin_nets', []),
                   ['网络名', '连接元件', '引脚', '页面'])
        _fill_tree(self._tree_drc_opt,    drc.get('bom_option_typos', []),
                   ['实际填写值', '疑似应为', '编辑距离', '使用该值的位号', '风险'])

    # ──────── 电阻检查 ─────────────────────────────────────

    def _refresh_res(self):
        ra = self._res
        _fill_tree(self._tree_res_div,
                   ra.get('divider_risks', []),
                   ['串阻位号', '串阻值', '上拉位号', '上拉值', '上拉电源',
                    '公共节点', '串/拉比', '上拉 < 1k', '状态', '页面'])
        _fill_tree(self._tree_res_dup_pu,
                   ra.get('dup_pullups', []),
                   ['信号网络', '上拉数量', '位号', '阻值', '上拉电源', '页面'])
        _fill_tree(self._tree_res_dup_pd,
                   ra.get('dup_pulldowns', []),
                   ['信号网络', '下拉数量', '位号', '阻值', '页面'])
        _fill_tree(self._tree_res_od,
                   ra.get('od_missing', []),
                   ['网络名', '节点数', '连接元件', '说明'])

    # ──────── 降额 ─────────────────────────────────────────

    def _refresh_derating(self):
        cols = ['位号', '值', '封装', '类型', '额定电压', '推断工作电压(V)',
                '推断来源网络', '降额比', '状态', '页面', 'DEPOP']
        _fill_tree(self.drt_tree, self._drt, cols)
        total = len(self._drt)
        fail  = sum(1 for r in self._drt if r.get('状态', '').startswith('❌'))
        ok    = sum(1 for r in self._drt if r.get('状态', '').startswith('✅'))
        self.drt_stat.configure(
            text=f'共 {total} 个电容  |  ✅ 合格 {ok}  |  ❌ 不合格 {fail}  |  ⚪ 无法判断 {total-fail-ok}')

    def _recalc_derating(self):
        if not self._components:
            messagebox.showwarning('提示', '请先解析文件'); return
        self._drt = analyze_derating(
            self._components, self._nets, self.ratio_var.get(), self._volt_map())
        self._refresh_derating()
        self._log(f'降额重新计算完成（上限={self.ratio_var.get():.0f}%）')

    def _toggle_rules(self):
        self._rules_visible = not self._rules_visible
        if self._rules_visible:
            self._rules_frame.grid()
            self._rules_btn.configure(text='收起内置电压匹配规则 ▴')
        else:
            self._rules_frame.grid_remove()
            self._rules_btn.configure(text='查看内置电压匹配规则 ▾')

    def _volt_map(self):
        result = {}
        for line in self.volt_entry.get('1.0', 'end').splitlines():
            line = line.strip()
            if not line or line.startswith('#'): continue
            if '=' in line:
                k, _, v = line.partition('=')
                try: result[k.strip()] = float(v.strip())
                except ValueError: pass
        return result or None

    # ──────── 查询 ─────────────────────────────────────────

    def _do_query(self):
        kw   = self.query_text.get().strip()
        mode = self.query_mode.get()
        if not kw: return
        if not self._components:
            messagebox.showwarning('提示', '请先解析文件'); return

        lines = []
        if mode == '位号':
            comp = self._components.get(kw)
            if comp:
                lines.append(f'═══ 元件：{kw} ═══')
                for k, v in comp.items():
                    if k == 'nets': continue
                    lines.append(f'  {k:<16} {v}')
                lines += ['', '  引脚 → 网络：']
                for pin, net in sorted(comp.get('nets', {}).items()):
                    lines.append(f'    pin {pin:<6} → {net}')
            else:
                matched = sorted(r for r in self._components if kw.upper() in r.upper())
                lines.append('未找到精确匹配，模糊结果：' if matched else f'未找到位号：{kw}')
                lines.extend(f'  {r}' for r in matched[:50])
        else:
            nodes = self._nets.get(kw)
            if nodes:
                lines.append(f'═══ 网络：{kw}（{len(nodes)} 个节点）═══')
                for n in nodes:
                    comp = self._components.get(n['refdes'], {})
                    desc = comp.get('value', '') or comp.get('part_name', '')
                    lines.append(f'  {n["refdes"]:<10} pin {n["pin"]:<6} ({n["pin_name"]:<12}) {desc}')
            else:
                matched = sorted(k for k in self._nets if kw.upper() in k.upper())
                lines.append('未找到精确匹配，模糊结果：' if matched else f'未找到网络：{kw}')
                lines.extend(f'  {nm}  ({len(self._nets[nm])} nodes)' for nm in matched[:50])

        self.query_result.configure(state='normal')
        self.query_result.delete('1.0', 'end')
        self.query_result.insert('end', '\n'.join(lines))
        self.query_result.configure(state='disabled')

    # ──────── 导出 ─────────────────────────────────────────

    def _export_excel(self):
        if not self._components:
            messagebox.showwarning('提示', '请先解析文件'); return
        out = filedialog.asksaveasfilename(
            title='保存分析报告', initialfile='pstx_分析报告.xlsx',
            defaultextension='.xlsx', filetypes=[('Excel 文件', '*.xlsx')])
        if not out: return
        self._log('\n导出 Excel 中…')
        threading.Thread(target=self._do_export, args=(out,), daemon=True).start()

    def _do_export(self, path):
        try:
            actual = export_to_excel({
                'project_name':      self.project_var.get().strip() or '未命名项目',
                'bom_normal_detail': self._dn, 'bom_depop_detail': self._dd,
                'bom_normal_merged': self._mn, 'bom_depop_merged': self._md,
                'net_analysis': self._na, 'drc': self._drc, 'derating': self._drt,
            }, path)
            self._log(f'✅ 导出完成：{actual}')
            self.after(0, lambda: messagebox.showinfo('完成', f'导出成功！\n{actual}'))
            folder = os.path.dirname(os.path.abspath(actual))
            try:
                if sys.platform == 'win32':    os.startfile(folder)
                elif sys.platform == 'darwin': subprocess.Popen(['open', folder])
                else:                          subprocess.Popen(['xdg-open', folder])
            except Exception:
                pass
        except Exception as e:
            import traceback
            self._log(f'❌ 导出失败：{e}\n{traceback.format_exc()}')
            self.after(0, lambda: messagebox.showerror('错误', str(e)))

    # ──────── Spinner + 日志 ──────────────────────────────

    def _start_spinner(self, label='处理中'):
        self._spinning = True; self._spin_step = 0; self._spin_label = label; self._spin()

    def _spin(self):
        if not self._spinning: return
        f = ['◐', '◓', '◑', '◒'][self._spin_step % 4]
        self.load_status.configure(text=f'{f} {self._spin_label}，请稍候…', fg='#2d6cdf')
        self._spin_step += 1
        self._spin_job = self.after(200, self._spin)

    def _stop_spinner(self, msg=''):
        self._spinning = False
        if hasattr(self, '_spin_job'): self.after_cancel(self._spin_job)
        color = '#2a8a2a' if msg.startswith('✅') else ('red' if '❌' in msg else '#333')
        self.load_status.configure(text=msg, fg=color)

    def _log(self, msg):
        def _w():
            self.log.configure(state='normal')
            self.log.insert('end', msg + '\n'); self.log.see('end')
            self.log.configure(state='disabled')
        self.after(0, _w)

    def _clear_log(self):
        self.log.configure(state='normal'); self.log.delete('1.0', 'end')
        self.log.configure(state='disabled')


if __name__ == '__main__':
    app = PstxApp()
    app.mainloop()
